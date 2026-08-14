import csv
from pathlib import Path

from openpyxl import load_workbook

from pami_scraper import exportar_resultados, procesar_lote


ENCABEZADOS_PROBABLES = {
    "beneficio",
    "beneficios",
    "nro_beneficio",
    "numero_beneficio",
    "numero beneficio",
    "nro beneficio",
    "beneficio_pami",
}


def main() -> None:
    print("==========================================")
    print(" Consulta de padrón PAMI por beneficios")
    print("==========================================")
    print("")

    beneficios = pedir_beneficios()
    if not beneficios:
        print("No se cargaron beneficios. El programa finaliza.")
        return

    headless = pedir_modo_navegador()

    print("")
    print(f"Se van a procesar {len(beneficios)} beneficio(s).")
    print("El navegador puede tardar unos segundos en abrir.")
    print("")

    resultados = procesar_lote(beneficios, headless=headless)

    if not resultados:
        print("No hubo resultados para exportar.")
        return

    ruta_excel = exportar_resultados(resultados)

    print("")
    print("Proceso finalizado.")
    print(f"Excel generado: {ruta_excel.resolve()}")
    print("")


def pedir_beneficios() -> list[str]:
    while True:
        print("Elegí cómo querés cargar los beneficios:")
        print("1. TXT (un beneficio por línea)")
        print("2. CSV")
        print("3. Excel (.xlsx)")
        print("4. Pegar lista manualmente")
        opcion = input("Opción: ").strip()
        print("")

        try:
            if opcion == "1":
                ruta = pedir_ruta_archivo("Ingresá la ruta del archivo TXT: ")
                return leer_txt(ruta)
            if opcion == "2":
                ruta = pedir_ruta_archivo("Ingresá la ruta del archivo CSV: ")
                return leer_csv(ruta)
            if opcion == "3":
                ruta = pedir_ruta_archivo("Ingresá la ruta del archivo Excel (.xlsx): ")
                return leer_excel(ruta)
            if opcion == "4":
                return leer_lista_manual()
        except Exception as exc:
            print(f"No se pudo leer la entrada: {exc}")
            print("")
            continue

        print("Opción no válida. Probá de nuevo.")
        print("")


def pedir_modo_navegador() -> bool:
    print("Modo de depuración visual: por defecto el navegador se verá abierto.")
    respuesta = input("¿Querés ejecutar sin mostrar el navegador? [s/N]: ").strip().lower()
    return respuesta in {"s", "si", "sí"}


def pedir_ruta_archivo(mensaje: str) -> Path:
    ruta = Path(input(mensaje).strip().strip('"'))
    if not ruta.exists():
        raise FileNotFoundError("El archivo no existe.")
    return ruta


def leer_txt(ruta: Path) -> list[str]:
    return [linea.strip() for linea in ruta.read_text(encoding="utf-8").splitlines() if linea.strip()]


def leer_csv(ruta: Path) -> list[str]:
    with ruta.open("r", encoding="utf-8-sig", newline="") as archivo:
        reader = csv.DictReader(archivo)
        if not reader.fieldnames:
            raise ValueError("El CSV no tiene encabezados.")

        columna = detectar_columna(reader.fieldnames)
        return [fila.get(columna, "").strip() for fila in reader if fila.get(columna, "").strip()]


def leer_excel(ruta: Path) -> list[str]:
    libro = load_workbook(ruta, read_only=True, data_only=True)
    hoja = libro.active
    filas = list(hoja.iter_rows(values_only=True))
    if not filas:
        raise ValueError("El archivo Excel está vacío.")

    encabezados = [str(valor or "").strip() for valor in filas[0]]
    columna = detectar_columna(encabezados)
    indice = encabezados.index(columna)

    beneficios = []
    for fila in filas[1:]:
        valor = fila[indice] if indice < len(fila) else None
        if valor is not None and str(valor).strip():
            beneficios.append(str(valor).strip())

    return beneficios


def leer_lista_manual() -> list[str]:
    print("Pegá la lista de beneficios, uno por línea.")
    print("Cuando termines, presioná Enter en una línea vacía.")

    beneficios = []
    while True:
        linea = input().strip()
        if not linea:
            break
        beneficios.append(linea)

    print("")
    return beneficios


def detectar_columna(encabezados: list[str]) -> str:
    encabezados_limpios = [str(item or "").strip() for item in encabezados if str(item or "").strip()]
    if not encabezados_limpios:
        raise ValueError("No se encontraron encabezados válidos.")

    for encabezado in encabezados_limpios:
        if encabezado.lower() in ENCABEZADOS_PROBABLES:
            return encabezado

    if len(encabezados_limpios) == 1:
        return encabezados_limpios[0]

    raise ValueError(
        "No pude identificar la columna de beneficios. Usá un encabezado como 'beneficio' o dejá una sola columna."
    )


if __name__ == "__main__":
    main()
