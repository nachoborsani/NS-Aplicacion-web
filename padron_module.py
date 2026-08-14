import os
import queue
import re
import threading
from pathlib import Path
from threading import Event
from tkinter import filedialog, messagebox, ttk

import customtkinter as ctk
from openpyxl import load_workbook

from app_paths import get_log_file, get_output_dir
from app_settings import get_medico_default, load_medicos_config
from excel_models import guardar_modelo_padron
from pami_scraper import (
    FUENTE_CARTILLA,
    FUENTE_PADRON,
    MODO_BENEFICIO,
    MODO_DNI,
    exportar_resultados,
    procesar_lote,
)


PADRON_HEADERS = {
    MODO_BENEFICIO: {
        "beneficio",
        "beneficios",
        "nro_beneficio",
        "numero_beneficio",
        "numero beneficio",
        "nro beneficio",
        "beneficio_pami",
    },
    MODO_DNI: {
        "dni",
        "documento",
        "nro_documento",
        "numero_documento",
        "numero documento",
        "nro documento",
        "doc",
    },
}

VISIBLE_COLUMNS_BENEFICIO = [
    ("numero_original", "BENEFICIARIO"),
    ("nombre_afiliado", "NOMBRE AFILIADO"),
    ("medico_cabecera", "MEDICO CABECERA"),
    ("clasificacion", "CLASIFICACION"),
]

VISIBLE_COLUMNS_DNI = [
    ("numero_original", "DNI BUSCADO"),
    ("beneficio_encontrado", "BENEFICIO"),
    ("nombre_afiliado", "NOMBRE AFILIADO"),
    ("medico_cabecera", "MEDICO CABECERA"),
    ("clasificacion", "CLASIFICACION"),
]

COLUMNAS_CARTILLA = {
    "beneficio": {
        "beneficio",
        "beneficios",
        "nro_beneficio",
        "numero_beneficio",
        "numero beneficio",
        "nro beneficio",
        "beneficio_pami",
        "afiliacion",
        "numero_afiliacion",
        "nro_afiliacion",
    },
    "dni": PADRON_HEADERS[MODO_DNI],
}


class PadronModuleFrame(ctk.CTkFrame):
    def __init__(self, master, on_back) -> None:
        super().__init__(master, fg_color="transparent")
        self.on_back = on_back
        self.event_queue: queue.Queue = queue.Queue()
        self.worker_thread: threading.Thread | None = None
        self.processing = False
        self.cancel_requested = Event()
        self.current_results: list[dict] = []
        self.last_output_excel: Path | None = None
        self.visible_columns = list(VISIBLE_COLUMNS_BENEFICIO)
        self.medicos_disponibles = load_medicos_config()
        self.active_cell_value = ""
        self.active_column_id = ""
        self.active_column_title = ""

        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(3, weight=1)
        self._build_ui()
        self.after(150, self._process_ui_queue)

    def _build_ui(self) -> None:
        top = ctk.CTkFrame(self, corner_radius=18, fg_color="#f3f7fb")
        top.grid(row=0, column=0, padx=18, pady=(18, 10), sticky="ew")
        top.grid_columnconfigure(2, weight=1)

        ctk.CTkButton(top, text="Volver", width=100, command=self._go_home, fg_color="#9aafc3", hover_color="#7f95aa").grid(
            row=0, column=0, padx=(18, 10), pady=18, sticky="w"
        )
        ctk.CTkLabel(top, text="Consultar padron PAMI", font=ctk.CTkFont(size=24, weight="bold"), text_color="#16324f").grid(
            row=0, column=1, padx=(0, 10), pady=18, sticky="w"
        )
        ctk.CTkLabel(
            top,
            text="Buscar por Cartilla medica o Padron prestadores, clasificar por medico objetivo y guardar el Excel solo cuando lo necesites.",
            font=ctk.CTkFont(size=13),
            text_color="#51657a",
        ).grid(row=1, column=1, columnspan=2, padx=(0, 18), pady=(0, 16), sticky="w")

        content = ctk.CTkFrame(self, corner_radius=18, fg_color="#ffffff")
        content.grid(row=1, column=0, padx=18, pady=(0, 18), sticky="nsew")
        content.grid_columnconfigure(0, weight=1)
        content.grid_rowconfigure(4, weight=1)

        search_frame = ctk.CTkFrame(content, corner_radius=12, fg_color="#eef3f8")
        search_frame.grid(row=0, column=0, padx=18, pady=(18, 10), sticky="ew")
        search_frame.grid_columnconfigure(6, weight=1)

        ctk.CTkLabel(search_frame, text="Fuente:", font=ctk.CTkFont(size=14, weight="bold"), text_color="#16324f").grid(
            row=0, column=0, padx=(14, 10), pady=12, sticky="w"
        )
        self.source_selector = ctk.CTkSegmentedButton(
            search_frame,
            values=["Cartilla médica", "Padrón prestadores"],
            command=self.on_source_change,
        )
        self.source_selector.grid(row=0, column=1, padx=10, pady=12, sticky="w")
        self.source_selector.set("Cartilla médica")

        ctk.CTkLabel(search_frame, text="Buscar por:", font=ctk.CTkFont(size=14, weight="bold"), text_color="#16324f").grid(
            row=0, column=2, padx=(18, 10), pady=12, sticky="w"
        )
        self.search_mode_selector = ctk.CTkSegmentedButton(search_frame, values=["Beneficio", "DNI"], command=self.on_search_mode_change)
        self.search_mode_selector.grid(row=0, column=3, padx=10, pady=12, sticky="w")
        self.search_mode_selector.set("Beneficio")

        ctk.CTkLabel(search_frame, text="Medico objetivo:", font=ctk.CTkFont(size=14, weight="bold"), text_color="#16324f").grid(
            row=0, column=4, padx=(18, 10), pady=12, sticky="w"
        )
        self.target_doctor_var = ctk.StringVar(value=get_medico_default())
        self.target_doctor_selector = ctk.CTkComboBox(
            search_frame, values=self.medicos_disponibles, variable=self.target_doctor_var, width=320, state="readonly"
        )
        self.target_doctor_selector.grid(row=0, column=5, padx=10, pady=12, sticky="w")

        self.input_container = ctk.CTkFrame(content, corner_radius=12, fg_color="#ffffff")
        self.input_container.grid(row=1, column=0, padx=18, pady=(0, 12), sticky="ew")
        self.input_container.grid_columnconfigure(0, weight=1)

        self.input_text = ctk.CTkTextbox(self.input_container, height=120, font=ctk.CTkFont(size=14))
        self.input_text.grid(row=0, column=0, sticky="ew")

        self.cartilla_inputs_frame = ctk.CTkFrame(self.input_container, corner_radius=12, fg_color="#ffffff")
        self.cartilla_inputs_frame.grid(row=0, column=0, sticky="ew")
        self.cartilla_inputs_frame.grid_columnconfigure((0, 1), weight=1)

        ctk.CTkLabel(
            self.cartilla_inputs_frame,
            text="BENEF",
            font=ctk.CTkFont(size=13, weight="bold"),
            text_color="#16324f",
        ).grid(row=0, column=0, padx=(0, 10), pady=(0, 6), sticky="w")
        ctk.CTkLabel(
            self.cartilla_inputs_frame,
            text="DNI",
            font=ctk.CTkFont(size=13, weight="bold"),
            text_color="#16324f",
        ).grid(row=0, column=1, padx=(10, 0), pady=(0, 6), sticky="w")

        self.cartilla_benef_text = ctk.CTkTextbox(self.cartilla_inputs_frame, height=120, font=ctk.CTkFont(size=14))
        self.cartilla_benef_text.grid(row=1, column=0, padx=(0, 10), sticky="ew")
        self.cartilla_dni_text = ctk.CTkTextbox(self.cartilla_inputs_frame, height=120, font=ctk.CTkFont(size=14))
        self.cartilla_dni_text.grid(row=1, column=1, padx=(10, 0), sticky="ew")

        controls = ctk.CTkFrame(content, corner_radius=12, fg_color="#eef3f8")
        controls.grid(row=2, column=0, padx=18, pady=(0, 12), sticky="ew")
        controls.grid_columnconfigure(5, weight=1)

        self.load_button = ctk.CTkButton(controls, text="Cargar TXT / Excel", width=180, command=self.load_file)
        self.load_button.grid(row=0, column=0, padx=(14, 10), pady=14, sticky="w")
        self.template_button = ctk.CTkButton(controls, text="Descargar modelo Excel", width=200, command=self.download_template)
        self.template_button.grid(row=0, column=1, padx=10, pady=14, sticky="w")
        self.process_button = ctk.CTkButton(
            controls, text="Procesar", width=150, height=40, font=ctk.CTkFont(size=16, weight="bold"), command=self.start_processing
        )
        self.process_button.grid(row=0, column=2, padx=10, pady=14, sticky="w")
        self.stop_button = ctk.CTkButton(
            controls, text="Detener", width=120, height=40, fg_color="#c56c6c", hover_color="#a95555", state="disabled", command=self.request_stop
        )
        self.stop_button.grid(row=0, column=3, padx=10, pady=14, sticky="w")
        self.show_browser_var = ctk.BooleanVar(value=False)
        self.show_browser_checkbox = ctk.CTkCheckBox(controls, text="Mostrar navegador", variable=self.show_browser_var)
        self.show_browser_checkbox.grid(row=0, column=4, padx=10, pady=14, sticky="w")
        self.clear_button = ctk.CTkButton(controls, text="Limpiar", width=110, fg_color="#9aafc3", hover_color="#7f95aa", command=self.clear_panel)
        self.clear_button.grid(row=0, column=6, padx=(10, 10), pady=14, sticky="e")
        self.save_button = ctk.CTkButton(controls, text="Guardar Excel", width=140, fg_color="#4f6378", hover_color="#3d4d61", command=self.save_excel)
        self.save_button.grid(row=0, column=7, padx=(0, 14), pady=14, sticky="e")

        status = ctk.CTkFrame(content, corner_radius=12, fg_color="#f8fafc")
        status.grid(row=4, column=0, padx=18, pady=(0, 18), sticky="nsew")
        status.grid_columnconfigure(0, weight=1)
        status.grid_rowconfigure(3, weight=1)

        self.status_label = ctk.CTkLabel(status, text="Preparado para comenzar.", font=ctk.CTkFont(size=15, weight="bold"), text_color="#16324f")
        self.status_label.grid(row=0, column=0, padx=16, pady=(14, 8), sticky="w")
        self.progress = ctk.CTkProgressBar(status, height=18)
        self.progress.grid(row=1, column=0, padx=16, pady=(0, 8), sticky="ew")
        self.progress.set(0)
        self.summary_label = ctk.CTkLabel(status, text="Sin resultados todavia.", font=ctk.CTkFont(size=13), text_color="#4f6378")
        self.summary_label.grid(row=2, column=0, padx=16, pady=(0, 10), sticky="w")

        table_container = ctk.CTkFrame(status, corner_radius=10, fg_color="#ffffff")
        table_container.grid(row=3, column=0, padx=16, pady=(0, 12), sticky="nsew")
        table_container.grid_columnconfigure(0, weight=1)
        table_container.grid_rowconfigure(0, weight=1)

        self.tree = ttk.Treeview(
            table_container,
            columns=[column_id for column_id, _ in self.visible_columns],
            show="headings",
            height=8,
            style="Report.Treeview",
        )
        self.tree.grid(row=0, column=0, sticky="nsew")
        self._configure_tree_columns()
        self.tree.bind("<ButtonRelease-1>", self.on_tree_click)
        self.tree.bind("<Double-1>", self.on_tree_double_click)
        self.tree.bind("<Control-c>", self.on_tree_copy_shortcut)
        self.tree.bind("<MouseWheel>", self.on_tree_mousewheel)
        scrollbar_y = ttk.Scrollbar(table_container, orient="vertical", command=self.tree.yview, style="Report.Vertical.TScrollbar")
        scrollbar_y.grid(row=0, column=1, sticky="ns")
        self.tree.configure(yscrollcommand=scrollbar_y.set)

        actions = ctk.CTkFrame(status, corner_radius=10, fg_color="#eef3f8")
        actions.grid(row=4, column=0, padx=16, pady=(0, 16), sticky="ew")
        actions.grid_columnconfigure(4, weight=1)

        self.copy_row_button = ctk.CTkButton(actions, text="Copiar fila seleccionada", command=self.copy_selected_row)
        self.copy_row_button.grid(row=0, column=0, padx=(12, 10), pady=12, sticky="w")
        self.column_copy_var = ctk.StringVar(value=self.visible_columns[0][1])
        self.column_selector = ctk.CTkComboBox(
            actions, values=[title for _, title in self.visible_columns], variable=self.column_copy_var, width=220, state="readonly"
        )
        self.column_selector.grid(row=0, column=1, padx=10, pady=12, sticky="w")
        self.copy_column_button = ctk.CTkButton(actions, text="Copiar columna", command=self.copy_selected_column)
        self.copy_column_button.grid(row=0, column=2, padx=10, pady=12, sticky="w")
        self.open_folder_button = ctk.CTkButton(actions, text="Abrir carpeta", fg_color="#4f6378", hover_color="#3d4d61", command=self.open_output_folder)
        self.open_folder_button.grid(row=0, column=5, padx=(10, 12), pady=12, sticky="e")
        self.on_source_change(self.source_selector.get())

    def _go_home(self) -> None:
        if self.processing:
            messagebox.showwarning("Atencion", "Espera a que termine o detene el proceso antes de volver.")
            return
        self.on_back()

    def _current_search_mode(self) -> str:
        return MODO_DNI if self.search_mode_selector.get() == "DNI" else MODO_BENEFICIO

    def _current_source(self) -> str:
        return FUENTE_CARTILLA if self.source_selector.get().startswith("Cartilla") else FUENTE_PADRON

    def _get_inputs_from_text(self) -> list[str]:
        if self._current_source() == FUENTE_CARTILLA:
            beneficios = [line.strip() for line in self.cartilla_benef_text.get("1.0", "end").splitlines() if line.strip()]
            dnis = [line.strip() for line in self.cartilla_dni_text.get("1.0", "end").splitlines() if line.strip()]
            if not beneficios and not dnis:
                return []
            if len(beneficios) != len(dnis):
                raise ValueError("En Cartilla medica la cantidad de BENEF y DNI debe coincidir.")
            return [f"{beneficio},{dni}" for beneficio, dni in zip(beneficios, dnis)]

        raw = self.input_text.get("1.0", "end").strip()
        return [line.strip() for line in raw.splitlines() if line.strip()] if raw else []

    def _detect_padron_column(self, headers: list[str], modo: str) -> str:
        clean_headers = [str(item or "").strip() for item in headers if str(item or "").strip()]
        if not clean_headers:
            raise ValueError("No se encontraron encabezados validos.")
        for header in clean_headers:
            if header.lower() in PADRON_HEADERS[modo]:
                return header
        if len(clean_headers) == 1:
            return clean_headers[0]
        if modo == MODO_DNI:
            raise ValueError("No pude identificar la columna de DNI. Usa un encabezado como 'dni' o deja una sola columna.")
        raise ValueError("No pude identificar la columna de beneficios. Usa un encabezado como 'beneficio' o deja una sola columna.")

    def _detect_cartilla_columns(self, headers: list[str]) -> tuple[str, str]:
        clean_headers = [str(item or "").strip() for item in headers if str(item or "").strip()]
        if not clean_headers:
            raise ValueError("No se encontraron encabezados validos.")

        beneficio_col = ""
        dni_col = ""
        for header in clean_headers:
            lower = header.lower()
            if lower in COLUMNAS_CARTILLA["beneficio"] and not beneficio_col:
                beneficio_col = header
            if lower in COLUMNAS_CARTILLA["dni"] and not dni_col:
                dni_col = header

        if not beneficio_col or not dni_col:
            raise ValueError("Para Cartilla médica el Excel debe incluir columnas 'beneficio' y 'dni'.")
        return beneficio_col, dni_col

    def _read_inputs_file(self, path: Path) -> list[str]:
        if path.suffix.lower() == ".txt":
            return [line.strip() for line in path.read_text(encoding="utf-8").splitlines() if line.strip()]
        if path.suffix.lower() == ".xlsx":
            workbook = load_workbook(path, read_only=True, data_only=True)
            rows = list(workbook.active.iter_rows(values_only=True))
            if not rows:
                raise ValueError("El archivo Excel esta vacio.")
            headers = [str(value or "").strip() for value in rows[0]]
            if self._current_source() == FUENTE_CARTILLA:
                beneficio_col, dni_col = self._detect_cartilla_columns(headers)
                idx_beneficio = headers.index(beneficio_col)
                idx_dni = headers.index(dni_col)
                return [
                    f"{str(row[idx_beneficio]).strip()},{str(row[idx_dni]).strip()}"
                    for row in rows[1:]
                    if idx_beneficio < len(row)
                    and idx_dni < len(row)
                    and row[idx_beneficio] is not None
                    and row[idx_dni] is not None
                    and str(row[idx_beneficio]).strip()
                    and str(row[idx_dni]).strip()
                ]
            column = self._detect_padron_column(headers, self._current_search_mode())
            index = headers.index(column)
            return [str(row[index]).strip() for row in rows[1:] if index < len(row) and row[index] is not None and str(row[index]).strip()]
        raise ValueError("Formato no soportado. Usa TXT o Excel.")

    def load_file(self) -> None:
        file_path = filedialog.askopenfilename(
            title="Seleccionar archivo",
            filetypes=[("Archivos compatibles", "*.txt *.xlsx"), ("TXT", "*.txt"), ("Excel", "*.xlsx")],
        )
        if not file_path:
            return
        try:
            items = self._read_inputs_file(Path(file_path))
        except Exception as exc:
            messagebox.showerror("Error", f"No se pudo leer el archivo.\n\n{exc}")
            return
        if self._current_source() == FUENTE_CARTILLA:
            beneficios = []
            dnis = []
            for item in items:
                partes = [parte.strip() for parte in str(item).split(",", 1)]
                if len(partes) == 2:
                    beneficios.append(partes[0])
                    dnis.append(partes[1])
            self.cartilla_benef_text.delete("1.0", "end")
            self.cartilla_dni_text.delete("1.0", "end")
            self.cartilla_benef_text.insert("1.0", "\n".join(beneficios))
            self.cartilla_dni_text.insert("1.0", "\n".join(dnis))
        else:
            self.input_text.delete("1.0", "end")
            self.input_text.insert("1.0", "\n".join(items))
        self.status_label.configure(text=f"Archivo cargado: {Path(file_path).name}")
        if self._current_source() == FUENTE_CARTILLA:
            etiqueta = "registro(s) beneficio,dni"
        else:
            etiqueta = "DNI(s)" if self._current_search_mode() == MODO_DNI else "beneficio(s)"
        self.summary_label.configure(text=f"Se cargaron {len(items)} {etiqueta}.")

    def download_template(self) -> None:
        modo = self._current_search_mode()
        ruta = filedialog.asksaveasfilename(
            title="Guardar modelo Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile=(
                "modelo_padron_cartilla.xlsx"
                if self._current_source() == FUENTE_CARTILLA
                else "modelo_padron_dni.xlsx" if modo == MODO_DNI else "modelo_padron_beneficio.xlsx"
            ),
        )
        if not ruta:
            return
        try:
            guardar_modelo_padron(ruta, modo, fuente_consulta=self._current_source())
            messagebox.showinfo("Modelo guardado", f"Modelo guardado en:\n\n{ruta}")
        except Exception as exc:
            messagebox.showerror("Error", f"No se pudo guardar el modelo.\n\n{exc}")

    def start_processing(self) -> None:
        if self.processing:
            return
        try:
            items = self._get_inputs_from_text()
        except Exception as exc:
            messagebox.showwarning("Atencion", str(exc))
            return
        if not items:
            texto = (
                "Pega beneficio,dni o carga un archivo antes de procesar."
                if self._current_source() == FUENTE_CARTILLA
                else "Pega DNI o carga un archivo antes de procesar." if self._current_search_mode() == MODO_DNI else "Pega beneficios o carga un archivo antes de procesar."
            )
            messagebox.showwarning("Atencion", texto)
            return

        self.processing = True
        self.cancel_requested.clear()
        self.current_results = []
        self.last_output_excel = None
        self.progress.set(0)
        self._clear_table()
        self._set_controls_enabled(False)
        self.status_label.configure(text="Preparando...")
        self.summary_label.configure(text=f"Se procesaran {len(items)} registro(s). El Excel se guarda solo si lo pedis.")

        self.worker_thread = threading.Thread(
            target=self._run_worker,
            args=(items, not self.show_browser_var.get(), self._current_search_mode(), self.target_doctor_var.get().strip(), self._current_source()),
            daemon=True,
        )
        self.worker_thread.start()

    def _run_worker(self, items: list[str], headless: bool, modo: str, medico: str, fuente: str) -> None:
        try:
            resultados = procesar_lote(
                items,
                headless=headless,
                modo_busqueda=modo,
                medico_objetivo=medico,
                fuente_consulta=fuente,
                progress_callback=self._progress_callback,
                status_callback=self._status_callback,
                should_cancel=self.cancel_requested.is_set,
            )
            self.event_queue.put(("finished", {"resultados": resultados, "cancelled": self.cancel_requested.is_set()}))
        except Exception as exc:
            self.event_queue.put(("fatal_error", str(exc)))

    def _progress_callback(self, current: int, total: int, resultado: dict) -> None:
        self.event_queue.put(("progress", {"current": current, "total": total, "resultado": resultado}))

    def _status_callback(self, text: str) -> None:
        self.event_queue.put(("status", text))

    def _process_ui_queue(self) -> None:
        had_events = False
        try:
            while True:
                event, payload = self.event_queue.get_nowait()
                had_events = True
                if event == "status":
                    self.status_label.configure(text=payload)
                elif event == "progress":
                    current, total, resultado = payload["current"], payload["total"], payload["resultado"]
                    self.progress.set(current / total if total else 0)
                    self.status_label.configure(text=f"Procesando {current} de {total}")
                    self.summary_label.configure(
                        text=f"Ultimo resultado: {resultado['clasificacion']}"
                        + (f" | {resultado['nombre_afiliado']}" if resultado["nombre_afiliado"] else "")
                    )
                    self.current_results.append(resultado)
                    self._append_result_row(resultado)
                elif event == "finished":
                    self.processing = False
                    self._set_controls_enabled(True)
                    if not payload["cancelled"] and self.current_results:
                        self.progress.set(1)
                    self.status_label.configure(text="Proceso detenido" if payload["cancelled"] else "Finalizado")
                    self.summary_label.configure(
                        text=f"{'Proceso detenido' if payload['cancelled'] else 'Proceso finalizado'}. Resultados en pantalla: {len(self.current_results)}."
                    )
                    if self.current_results:
                        detail = "Lo procesado quedo cargado en pantalla." if payload["cancelled"] else "Los resultados quedaron cargados en pantalla."
                        messagebox.showinfo(self.status_label.cget("text"), f"{detail}\n\nLog: {get_log_file()}")
                        if not payload["cancelled"]:
                            self._notify_registered_doctor_mismatch()
                    else:
                        messagebox.showwarning("Sin resultados", "No hubo resultados para mostrar.")
                elif event == "fatal_error":
                    self.processing = False
                    self._set_controls_enabled(True)
                    self.status_label.configure(text="Error")
                    self.summary_label.configure(text="El proceso se detuvo por un error general.")
                    messagebox.showerror("Error", f"Ocurrio un error general.\n\n{payload}\n\nLog: {get_log_file()}")
        except queue.Empty:
            pass
        finally:
            delay = 120 if (self.processing or had_events) else 350
            self.after(delay, self._process_ui_queue)

    def save_excel(self) -> None:
        if not self.current_results:
            messagebox.showwarning("Atencion", "Todavia no hay resultados para guardar.")
            return
        ruta = filedialog.asksaveasfilename(
            title="Guardar resultados en Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile="resultados_padron.xlsx",
        )
        if not ruta:
            return
        try:
            self.last_output_excel = exportar_resultados(self.current_results, ruta_excel=ruta)
            messagebox.showinfo("Excel guardado", f"Archivo guardado en:\n\n{self.last_output_excel}")
        except Exception as exc:
            messagebox.showerror("Error", f"No se pudo guardar el Excel.\n\n{exc}")

    def open_output_folder(self) -> None:
        target = self.last_output_excel.parent if self.last_output_excel and self.last_output_excel.exists() else get_output_dir()
        os.startfile(str(target.resolve()))

    def clear_panel(self) -> None:
        if self.processing:
            return
        self.input_text.delete("1.0", "end")
        self.cartilla_benef_text.delete("1.0", "end")
        self.cartilla_dni_text.delete("1.0", "end")
        self.current_results = []
        self.last_output_excel = None
        self.progress.set(0)
        self.status_label.configure(text="Preparado para comenzar.")
        self.summary_label.configure(text="Panel limpio. Podes pegar nuevos datos.")
        self._clear_table()

    def request_stop(self) -> None:
        if not self.processing:
            return
        self.cancel_requested.set()
        self.stop_button.configure(state="disabled")
        self.status_label.configure(text="Deteniendo...")
        self.summary_label.configure(text="Se va a detener al terminar el afiliado que esta en curso.")

    def copy_selected_row(self) -> None:
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning("Atencion", "Selecciona una fila para copiar.")
            return
        row_text = "\t".join(str(value) for value in self.tree.item(selected[0], "values"))
        self.clipboard_clear()
        self.clipboard_append(row_text)
        self.update()
        messagebox.showinfo("Copiado", "La fila seleccionada fue copiada al portapapeles.")

    def copy_selected_column(self) -> None:
        if not self.current_results:
            messagebox.showwarning("Atencion", "Todavia no hay resultados para copiar.")
            return
        selected_title = self.column_copy_var.get()
        selected_column_id = next((column_id for column_id, title in self.visible_columns if title == selected_title), None)
        if not selected_column_id:
            messagebox.showwarning("Atencion", "No se pudo identificar la columna seleccionada.")
            return
        column_text = "\n".join(str(result.get(selected_column_id, "")) for result in self.current_results)
        self.clipboard_clear()
        self.clipboard_append(column_text)
        self.update()
        messagebox.showinfo("Copiado", f"La columna '{selected_title}' fue copiada al portapapeles.")

    def on_tree_click(self, event) -> None:
        region = self.tree.identify_region(event.x, event.y)
        if region != "cell":
            return

        row_id = self.tree.identify_row(event.y)
        column_token = self.tree.identify_column(event.x)
        if not row_id or not column_token:
            return

        column_index = int(column_token.replace("#", "")) - 1
        if column_index < 0 or column_index >= len(self.visible_columns):
            return

        values = self.tree.item(row_id, "values")
        self.tree.selection_set(row_id)
        self.active_column_id, self.active_column_title = self.visible_columns[column_index]
        self.active_cell_value = str(values[column_index]) if column_index < len(values) else ""
        self.summary_label.configure(
            text=f"Celda seleccionada: {self.active_column_title} = {self.active_cell_value or '(vacio)'}"
        )

    def on_tree_double_click(self, event) -> None:
        region = self.tree.identify_region(event.x, event.y)
        if region == "cell":
            self.on_tree_click(event)
            self._copy_active_cell()
            return

        if region == "heading":
            column_token = self.tree.identify_column(event.x)
            if not column_token:
                return
            column_index = int(column_token.replace("#", "")) - 1
            if column_index < 0 or column_index >= len(self.visible_columns):
                return
            _, title = self.visible_columns[column_index]
            self.column_copy_var.set(title)
            self.copy_selected_column()

    def on_tree_copy_shortcut(self, _event=None):
        if self.active_cell_value:
            self._copy_active_cell()
            return "break"
        return None

    def on_tree_mousewheel(self, event):
        self.tree.yview_scroll(int(-1 * (event.delta / 120)), "units")
        return "break"

    def _copy_active_cell(self) -> None:
        if not self.active_column_title:
            return
        self.clipboard_clear()
        self.clipboard_append(self.active_cell_value)
        self.update()
        self.summary_label.configure(
            text=f"Copiado: {self.active_column_title} = {self.active_cell_value or '(vacio)'}"
        )

    def on_search_mode_change(self, selected_value: str) -> None:
        if self._current_source() == FUENTE_CARTILLA:
            self.visible_columns = list(VISIBLE_COLUMNS_BENEFICIO)
            self._configure_tree_columns()
            self.column_selector.configure(values=[title for _, title in self.visible_columns])
            self.column_copy_var.set(self.visible_columns[0][1])
            self._clear_table()
            self.current_results = []
            self.summary_label.configure(text="Modo Cartilla médica seleccionado. Usa beneficio,dni por línea o un Excel con ambas columnas.")
            return

        modo = MODO_DNI if selected_value == "DNI" else MODO_BENEFICIO
        self.visible_columns = list(VISIBLE_COLUMNS_DNI if modo == MODO_DNI else VISIBLE_COLUMNS_BENEFICIO)
        self._configure_tree_columns()
        self.column_selector.configure(values=[title for _, title in self.visible_columns])
        self.column_copy_var.set(self.visible_columns[0][1])
        self._clear_table()
        self.current_results = []
        self.summary_label.configure(
            text="Modo DNI seleccionado. Pega DNI o carga TXT/Excel."
            if modo == MODO_DNI
            else "Modo beneficio seleccionado. Pega beneficios o carga TXT/Excel."
        )

    def on_source_change(self, selected_value: str) -> None:
        es_cartilla = selected_value.startswith("Cartilla")
        self.search_mode_selector.configure(state=("disabled" if es_cartilla else "normal"))
        if es_cartilla:
            self.input_text.grid_remove()
            self.cartilla_inputs_frame.grid()
        else:
            self.cartilla_inputs_frame.grid_remove()
            self.input_text.grid()
        self.visible_columns = list(VISIBLE_COLUMNS_BENEFICIO if es_cartilla or self._current_search_mode() == MODO_BENEFICIO else VISIBLE_COLUMNS_DNI)
        self._configure_tree_columns()
        self.column_selector.configure(values=[title for _, title in self.visible_columns])
        self.column_copy_var.set(self.visible_columns[0][1])
        self.current_results = []
        self._clear_table()
        self.input_text.delete("1.0", "end")
        self.cartilla_benef_text.delete("1.0", "end")
        self.cartilla_dni_text.delete("1.0", "end")
        self.progress.set(0)
        if es_cartilla:
            self.status_label.configure(text="Modo Cartilla medica")
            self.summary_label.configure(text="Carga una fila por paciente en dos columnas: BENEF y DNI, o usa un Excel con ambas columnas.")
        else:
            self.status_label.configure(text="Modo Padron prestadores")
            self.on_search_mode_change(self.search_mode_selector.get())

    def _append_result_row(self, resultado: dict) -> None:
        self.tree.insert("", "end", values=[resultado.get(column_id, "") for column_id, _ in self.visible_columns])

    def _configure_tree_columns(self) -> None:
        visible_ids = [column_id for column_id, _ in self.visible_columns]
        self.tree.configure(columns=visible_ids)
        for column_id in self.tree["columns"]:
            self.tree.heading(column_id, text="")
            self.tree.column(column_id, width=0, stretch=False)
        for column_id, title in self.visible_columns:
            self.tree.heading(column_id, text=title)
            width = 260 if column_id == "medico_cabecera" else 170 if column_id in {"numero_original", "beneficio_encontrado"} else 190
            self.tree.column(column_id, width=width, anchor="w", stretch=True)

    def _clear_table(self) -> None:
        for item in self.tree.get_children():
            self.tree.delete(item)

    def _set_controls_enabled(self, enabled: bool) -> None:
        state = "normal" if enabled else "disabled"
        self.load_button.configure(state=state)
        self.template_button.configure(state=state)
        self.process_button.configure(state=state)
        self.show_browser_checkbox.configure(state=state)
        self.clear_button.configure(state=state)
        self.search_mode_selector.configure(state=("disabled" if self._current_source() == FUENTE_CARTILLA else state))
        self.source_selector.configure(state=state)
        self.target_doctor_selector.configure(state=state)
        self.stop_button.configure(state=("normal" if not enabled else "disabled"))

    def _notify_registered_doctor_mismatch(self) -> None:
        medico_seleccionado = self.target_doctor_var.get().strip()
        medico_seleccionado_norm = _normalizar_texto_comparable(medico_seleccionado)
        medicos_registrados = {_normalizar_texto_comparable(medico): medico for medico in self.medicos_disponibles}

        coincidencias: dict[str, int] = {}
        for resultado in self.current_results:
            medico_cabecera = str(resultado.get("medico_cabecera", "")).strip()
            if not medico_cabecera:
                continue

            medico_encontrado_norm = _normalizar_texto_comparable(medico_cabecera)
            if not medico_encontrado_norm or medico_encontrado_norm == medico_seleccionado_norm:
                continue

            medico_registrado = medicos_registrados.get(medico_encontrado_norm)
            if medico_registrado:
                coincidencias[medico_registrado] = coincidencias.get(medico_registrado, 0) + 1

        if not coincidencias:
            return

        detalle = "\n".join(
            f"- {medico}: {cantidad} resultado(s)"
            for medico, cantidad in sorted(coincidencias.items(), key=lambda item: item[0])
        )
        messagebox.showwarning(
            "Revisar medico seleccionado",
            "Atencion: en los resultados aparecio al menos otro medico que tambien esta registrado en la app.\n\n"
            f"Medico seleccionado: {medico_seleccionado}\n\n"
            "Se detecto:\n"
            f"{detalle}\n\n"
            "Puede que hayas procesado con un medico objetivo distinto al que correspondia.",
        )

    def refresh_medicos(self, medicos: list[str] | None = None) -> None:
        self.medicos_disponibles = medicos or load_medicos_config()
        medico_actual = self.target_doctor_var.get().strip()
        if medico_actual not in self.medicos_disponibles:
            medico_actual = self.medicos_disponibles[0]
        self.target_doctor_selector.configure(values=self.medicos_disponibles)
        self.target_doctor_var.set(medico_actual)


def _normalizar_texto_comparable(texto: str) -> str:
    reemplazos = str.maketrans(
        {
            "Á": "A",
            "É": "E",
            "Í": "I",
            "Ó": "O",
            "Ú": "U",
            "á": "A",
            "é": "E",
            "í": "I",
            "ó": "O",
            "ú": "U",
        }
    )
    return " ".join(re.sub(r"\s+", " ", str(texto or "").translate(reemplazos).upper()).split())
