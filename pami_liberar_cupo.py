import threading
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Callable, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from playwright.sync_api import Browser, BrowserContext, Page, Playwright, sync_playwright

from app_logging import log_message
from pami_scraper import configurar_playwright


CUP_LOGIN_URL = "https://cup.pami.org.ar/controllers/loginController.php?redirect=https://pe.pami.org.ar"
PAMI_TRANSMISION_URL = "https://pe.pami.org.ar/controllers/transmision.php?registros_por_pagina=50"
PAMI_PANEL_ACEPTACION_URL = "https://pe.pami.org.ar/controllers/efector.php?registros_por_pagina=50"
_NO_EVALUATE_ARG = object()
_NAVIGATION_TRANSIENT_ERRORS = (
    "execution context was destroyed",
    "most likely because of a navigation",
    "cannot find context with specified id",
    "target page, context or browser has been closed",
)


@dataclass
class SesionLiberarCupo:
    playwright: Playwright
    browser: Browser
    context: BrowserContext
    page: Page


@dataclass
class ResultadoLiberacion:
    n_orden: str
    estado: str
    mensaje: str = ""
    nombre: str = ""
    beneficio: str = ""
    practica: str = ""
    turno: str = ""
    f_vencimiento: str = ""


@dataclass
class ResumenLiberacion:
    ok: int = 0
    errores: int = 0
    omitidos: int = 0
    detalle: list[ResultadoLiberacion] = field(default_factory=list)


def exportar_reporte_no_validadas(rows: list[dict], destino: str | Path, filtros: dict | None = None) -> Path:
    destino = Path(destino)
    destino.parent.mkdir(parents=True, exist_ok=True)
    filtros = filtros or {}
    exportado = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

    headers = [
        "Nro. OME",
        "Turno",
        "Beneficio/GP",
        "Paciente",
        "Practica",
        "Estado",
        "Vencimiento aceptacion",
        "Filtro turno desde",
        "Filtro turno hasta",
        "Exportado",
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "No validadas"
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="2F6FA3")
    for cell in ws[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill

    for row in rows:
        ws.append(
            [
                row.get("n_orden", ""),
                row.get("turno", ""),
                row.get("beneficio", ""),
                row.get("nombre", ""),
                row.get("practica", ""),
                row.get("estado", "") or "No validada / pendiente de liberar",
                row.get("f_vencimiento", ""),
                filtros.get("fecha_desde", ""),
                filtros.get("fecha_hasta", ""),
                exportado,
            ]
        )

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    widths = [18, 22, 18, 34, 72, 34, 22, 20, 20, 22]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    wb.save(destino)
    return destino


class PamiLiberarCupoController:
    def __init__(
        self,
        log_callback: Optional[Callable[[str], None]] = None,
        status_callback: Optional[Callable[[str], None]] = None,
    ) -> None:
        self.log_callback = log_callback or (lambda _: None)
        self.status_callback = status_callback or (lambda _: None)
        self._lock = threading.RLock()
        self._sesion: SesionLiberarCupo | None = None
        self._stop_requested = False

    def abrir_pami(self, usuario: str | None = None, clave: str | None = None, headless: bool = False) -> None:
        with self._lock:
            self._cleanup_dead_session()
            if self._sesion is not None:
                self._dispose_session("Sesion anterior cerrada.", "Sesion anterior descartada antes de abrir Liberar Cupo.")

            configurar_playwright()
            playwright = sync_playwright().start()
            launch_args = {"headless": headless}
            if not headless:
                launch_args["args"] = ["--window-size=1440,920"]
            browser = playwright.chromium.launch(**launch_args)
            context = browser.new_context(ignore_https_errors=True, viewport={"width": 1440, "height": 920})
            page = context.new_page()
            self._configure_page(page)

            page.goto(CUP_LOGIN_URL if (usuario and clave) else PAMI_TRANSMISION_URL, wait_until="domcontentloaded")
            self._sesion = SesionLiberarCupo(playwright=playwright, browser=browser, context=context, page=page)

            if usuario and clave:
                self._auto_login_cup(page, usuario, clave)
                page.goto(PAMI_TRANSMISION_URL, wait_until="domcontentloaded")
                page.wait_for_timeout(1200)

            self._status("PAMI abierto para Liberar Cupo.")
            self._log("Navegador listo para revisar cupos no validados.")

    def cerrar_navegador(self) -> None:
        with self._lock:
            if self._sesion is None:
                return
            self._dispose_session("Navegador cerrado.", "Sesion de Liberar Cupo cerrada.")

    def sesion_activa(self) -> bool:
        self._cleanup_dead_session()
        return self._sesion is not None

    def solicitar_detencion(self) -> None:
        self._stop_requested = True
        self._status("Detencion solicitada. Se detiene al terminar la OME actual.")
        self._log("Detencion solicitada por el usuario.")

    def detectar_candidatas(
        self,
        fecha_desde: str = "",
        fecha_hasta: str = "",
        max_paginas: int = 10,
    ) -> list[dict]:
        page = self._get_page()
        self._stop_requested = False
        self._ensure_transmision_page(page)
        self._ensure_transmision_registros_por_pagina(page, "50")
        self._apply_transmision_filters(page, fecha_desde, fecha_hasta)
        self._ensure_transmision_registros_por_pagina(page, "50")

        candidatas: list[dict] = []
        vistos: set[str] = set()
        pagina = 0
        while pagina < max(1, int(max_paginas or 1)):
            if self._stop_requested:
                break
            page.wait_for_timeout(700)
            filas = self._evaluate_with_navigation_retry(page, _EXTRAER_CANDIDATAS_TRANSMISION_SCRIPT) or []
            for item in filas:
                if not isinstance(item, dict):
                    continue
                orden = str(item.get("n_orden", "") or "").strip()
                if not orden or orden in vistos:
                    continue
                vistos.add(orden)
                candidatas.append(item)
            pagina += 1
            self._log(f"Pagina de Transmision revisada: {pagina} | candidatas acumuladas: {len(candidatas)}")
            if not self._avanzar_pagina_transmision(page):
                break

        self._log(f"Candidatas no validadas detectadas: {len(candidatas)}")
        self._status(f"Detectadas {len(candidatas)} OME(s) para revisar.")
        return candidatas

    def liberar_omes(self, candidatas: list[dict]) -> ResumenLiberacion:
        page = self._get_page()
        self._stop_requested = False
        resumen = ResumenLiberacion()
        total = len(candidatas)
        for idx, item in enumerate(candidatas, start=1):
            if self._stop_requested:
                break
            orden = str(item.get("n_orden", "") or "").strip()
            if not orden:
                resumen.omitidos += 1
                continue
            self._status(f"[{idx}/{total}] Cancelando aceptacion OME {orden}")
            try:
                resultado = self._cancelar_aceptacion(page, item)
                if resultado.estado == "LIBERADA":
                    resumen.ok += 1
                elif resultado.estado == "OMITIDA":
                    resumen.omitidos += 1
                else:
                    resumen.errores += 1
                resumen.detalle.append(resultado)
            except Exception as exc:
                mensaje = str(exc)
                self._log(f"OME {orden}: ERROR - {mensaje}")
                resumen.errores += 1
                resumen.detalle.append(
                    ResultadoLiberacion(
                        n_orden=orden,
                        estado="ERROR",
                        mensaje=mensaje,
                        nombre=str(item.get("nombre", "")),
                        beneficio=str(item.get("beneficio", "")),
                        practica=str(item.get("practica", "")),
                        turno=str(item.get("turno", "")),
                        f_vencimiento=str(item.get("f_vencimiento", "")),
                    )
                )
                try:
                    page = self._goto_with_recovery(page, PAMI_PANEL_ACEPTACION_URL)
                    page.wait_for_timeout(800)
                except Exception:
                    pass

        self._status(f"Liberacion terminada: {resumen.ok} OK, {resumen.errores} error(es), {resumen.omitidos} omitida(s).")
        self._log(f"Liberacion terminada: OK={resumen.ok} errores={resumen.errores} omitidos={resumen.omitidos}")
        return resumen

    def cerrar(self) -> None:
        self.cerrar_navegador()

    def _cancelar_aceptacion(self, page: Page, item: dict) -> ResultadoLiberacion:
        orden = str(item.get("n_orden", "") or "").strip()
        if not orden:
            raise RuntimeError("Registro sin Nro. de Orden.")

        self._buscar_orden_en_aceptacion(page, orden)
        detalle = self._leer_fila_aceptacion(page, orden)
        estado = str(detalle.get("estado", "") or "").strip().upper()
        if estado and "ACEPTADA" not in estado:
            mensaje = f"Estado actual no cancelable automaticamente: {estado}"
            self._log(f"OME {orden}: omitida. {mensaje}")
            return ResultadoLiberacion(
                n_orden=orden,
                estado="OMITIDA",
                mensaje=mensaje,
                nombre=str(item.get("nombre", "")),
                beneficio=str(item.get("beneficio", "")),
                practica=str(item.get("practica", "")),
                turno=str(item.get("turno", "")),
                f_vencimiento=str(detalle.get("f_vencimiento", "")),
            )

        clicked = self._evaluate_with_navigation_retry(page, _CLICK_CANCELAR_ACEPTACION_SCRIPT, orden)
        if clicked != "OK":
            raise RuntimeError(f"No se encontro boton rojo de cancelar aceptacion: {clicked}")
        page.wait_for_timeout(900)

        self._confirmar_cancelacion_si_corresponde(page)
        page.wait_for_timeout(1200)

        if self._orden_sigue_cancelable(page, orden):
            raise RuntimeError("PAMI aun muestra la OME como cancelable despues de confirmar.")

        vencimiento = str(detalle.get("f_vencimiento", "") or "").strip()
        mensaje = f"Cupo liberado OK{f' | vencimiento {vencimiento}' if vencimiento else ''}."
        self._log(f"OME {orden}: {mensaje}")
        return ResultadoLiberacion(
            n_orden=orden,
            estado="LIBERADA",
            mensaje=mensaje,
            nombre=str(item.get("nombre") or detalle.get("nombre") or ""),
            beneficio=str(item.get("beneficio") or detalle.get("beneficio") or ""),
            practica=str(item.get("practica") or detalle.get("practica") or ""),
            turno=str(item.get("turno") or detalle.get("turno") or ""),
            f_vencimiento=vencimiento,
        )

    def _buscar_orden_en_aceptacion(self, page: Page, n_orden: str) -> None:
        page = self._goto_with_recovery(page, PAMI_PANEL_ACEPTACION_URL)
        page.wait_for_timeout(900)
        if "cup.pami.org.ar" in (page.url or ""):
            raise RuntimeError("La sesion de PAMI expiro. Inicia sesion nuevamente.")
        result = self._evaluate_with_navigation_retry(page, _BUSCAR_ORDEN_ACEPTACION_SCRIPT, n_orden, tolerate_navigation=True)
        if result not in {"OK", "NAVIGATED_AFTER_ACTION"}:
            raise RuntimeError(f"No se pudo buscar la OME en Panel de Aceptacion: {result}")
        page.wait_for_timeout(1400)

    def _leer_fila_aceptacion(self, page: Page, n_orden: str) -> dict:
        data = self._evaluate_with_navigation_retry(page, _LEER_FILA_ACEPTACION_SCRIPT, n_orden)
        if not isinstance(data, dict) or not data.get("encontrada"):
            raise RuntimeError("La OME no aparecio en Panel de Aceptacion.")
        return data

    def _confirmar_cancelacion_si_corresponde(self, page: Page) -> None:
        try:
            page.locator(".modal.show, .modal.in, .modal[style*='display: block'], .bootbox, [role='dialog']").first.wait_for(
                state="visible",
                timeout=3000,
            )
        except Exception:
            return

        result = self._evaluate_with_navigation_retry(page, _CONFIRMAR_MODAL_CANCELACION_SCRIPT, tolerate_navigation=True)
        if result not in {"OK", "NAVIGATED_AFTER_ACTION"}:
            raise RuntimeError(f"No se pudo confirmar la cancelacion: {result}")
        try:
            page.locator(".modal.show, .modal.in, .modal[style*='display: block'], .bootbox, [role='dialog']").first.wait_for(
                state="hidden",
                timeout=8000,
            )
        except Exception:
            pass

    def _orden_sigue_cancelable(self, page: Page, n_orden: str) -> bool:
        try:
            if "efector.php" not in (page.url or ""):
                page = self._goto_with_recovery(page, PAMI_PANEL_ACEPTACION_URL)
                page.wait_for_timeout(900)
                self._evaluate_with_navigation_retry(page, _BUSCAR_ORDEN_ACEPTACION_SCRIPT, n_orden, tolerate_navigation=True)
                page.wait_for_timeout(900)
            return bool(self._evaluate_with_navigation_retry(page, _ORDEN_SIGUE_CANCELABLE_SCRIPT, n_orden))
        except Exception:
            return False

    def _ensure_transmision_page(self, page: Page) -> None:
        if "transmision.php" not in (page.url or ""):
            page = self._goto_with_recovery(page, PAMI_TRANSMISION_URL)
            page.wait_for_timeout(1200)
        if "cup.pami.org.ar" in (page.url or ""):
            raise RuntimeError("La sesion de PAMI expiro. Inicia sesion nuevamente.")

    def _apply_transmision_filters(self, page: Page, fecha_desde: str, fecha_hasta: str) -> None:
        result = self._evaluate_with_navigation_retry(
            page,
            """
            ({desde, hasta}) => {
              const visible = (el) => !!el && (el.offsetParent !== null || el.getClientRects().length > 0);
              const setValue = (selector, value) => {
                const el = document.querySelector(selector);
                if (!el) return false;
                el.value = value || '';
                el.dispatchEvent(new Event('input', { bubbles: true }));
                el.dispatchEvent(new Event('change', { bubbles: true }));
                return true;
              };
              setValue('input[name="f_turno_desde"], #f_turno_desde', desde || '');
              setValue('input[name="f_turno_hasta"], #f_turno_hasta', hasta || '');
              setValue('select[name="c_validada"], #c_validada', '');
              setValue('select[name="transmitida"], #transmitida', '');
              const buscar = Array.from(document.querySelectorAll('button, input[type="submit"], input[type="button"], a'))
                .find((el) => visible(el) && (el.textContent || el.value || '').trim().toLowerCase() === 'buscar');
              if (!buscar) return 'NO_BUSCAR';
              buscar.click();
              return 'OK';
            }
            """,
            {"desde": fecha_desde, "hasta": fecha_hasta},
            tolerate_navigation=True,
        )
        if result not in {"OK", "NAVIGATED_AFTER_ACTION"}:
            self._log(f"No se pudo confirmar click en Buscar de Transmision: {result}")
        self._wait_page_settled(page, timeout=5000)
        page.wait_for_timeout(900)
        self._log(f"Filtros aplicados en Transmision: desde={fecha_desde or '-'} hasta={fecha_hasta or '-'} validada=- transmitida=-")

    def _ensure_transmision_registros_por_pagina(self, page: Page, valor: str = "50") -> None:
        actualizado = self._evaluate_with_navigation_retry(
            page,
            """
            (valor) => {
              const visible = (el) => !!el && (el.offsetParent !== null || el.getClientRects().length > 0);
              const candidatos = [
                document.querySelector('select[name="registros_por_pagina"]'),
                document.querySelector('select[name="rpp"]'),
                document.querySelector('select[name="per_page"]'),
              ].filter(Boolean);

              let sel = candidatos[0] || null;
              if (!sel) {
                sel = Array.from(document.querySelectorAll('select')).find((item) =>
                  visible(item) && Array.from(item.options || []).some((opt) =>
                    (opt.value || '').trim() === valor || (opt.textContent || '').includes(valor)
                  )
                ) || null;
              }

              if (!sel) return 'missing';
              if ((sel.value || '').trim() === valor) return 'ok';

              sel.value = valor;
              sel.dispatchEvent(new Event('input', { bubbles: true }));
              sel.dispatchEvent(new Event('change', { bubbles: true }));
              return 'changed';
            }
            """,
            valor,
            tolerate_navigation=True,
        )
        if actualizado in {"changed", "NAVIGATED_AFTER_ACTION"}:
            self._wait_page_settled(page, timeout=5000)
            page.wait_for_timeout(700)
            self._log(f"Registros por pagina ajustado a {valor}.")
        elif actualizado == "missing":
            self._log("No se encontro selector de registros por pagina; se continua con la paginacion disponible.")

    def _avanzar_pagina_transmision(self, page: Page) -> bool:
        try:
            advanced = self._evaluate_with_navigation_retry(
                page,
                """
                () => {
                  const visible = (el) => !!el && (el.offsetParent !== null || el.getClientRects().length > 0);
                  const norm = (value) => String(value || '')
                    .normalize('NFD')
                    .replace(/[\\u0300-\\u036f]/g, '')
                    .replace(/\\s+/g, ' ')
                    .trim()
                    .toLowerCase();
                  const active = Array.from(document.querySelectorAll('a, button, li, span'))
                    .find((el) => /^\\d+$/.test((el.textContent || '').trim()) && /active|current|selected|btn-primary/.test(el.className || ''));
                  const current = active ? Number((active.textContent || '').trim()) : 1;
                  const mas = Array.from(document.querySelectorAll('a, button'))
                    .find((el) => visible(el) && norm(el.textContent || el.value).includes('mas resultados'));
                  if (mas) {
                    mas.click();
                    return true;
                  }
                  const next = Array.from(document.querySelectorAll('a, button'))
                    .find((el) => visible(el) && !el.disabled && /^\\d+$/.test((el.textContent || '').trim()) && Number((el.textContent || '').trim()) === current + 1);
                  if (!next) return false;
                  next.click();
                  return true;
                }
                """,
                tolerate_navigation=True,
            )
            if advanced:
                self._wait_page_settled(page, timeout=5000)
                page.wait_for_timeout(900)
            return bool(advanced)
        except Exception:
            return False

    def _auto_login_cup(self, page: Page, usuario: str, clave: str) -> None:
        usuario = (usuario or "").strip()
        clave = clave or ""
        if not usuario or not clave:
            return
        page = self._goto_with_recovery(page, CUP_LOGIN_URL)
        page.wait_for_timeout(900)
        user_input = page.locator('input[name="usuario"], input[type="text"], #usuario').first
        pass_input = page.locator('input[name="password"], input[type="password"], #password').first
        user_input.wait_for(state="visible")
        user_input.fill("")
        user_input.type(usuario, delay=75)
        pass_input.wait_for(state="visible")
        pass_input.fill("")
        pass_input.type(clave, delay=75)
        submit = page.locator('button[type="submit"], input[type="submit"], button:has-text("Ingresar"), button:has-text("Entrar")').first
        submit.click()
        page.wait_for_load_state("domcontentloaded")
        page.wait_for_timeout(1800)
        if "cup.pami.org.ar" in (page.url or ""):
            raise RuntimeError("No se pudo iniciar sesion en CUP PAMI.")
        self._log("Sesion iniciada automaticamente en CUP PAMI.")

    def _get_page(self) -> Page:
        self._cleanup_dead_session()
        if self._sesion is None:
            raise RuntimeError("No hay una sesion activa de PAMI. Ejecuta Detectar no validadas para abrir PAMI automaticamente.")
        return self._sesion.page

    def _configure_page(self, page: Page) -> None:
        page.set_default_timeout(25000)
        page.on("console", self._on_console)
        page.on("pageerror", self._on_page_error)

    def _recover_active_page(self, preferred: Page | None = None) -> Page | None:
        sesion = self._sesion
        if sesion is None:
            return None
        try:
            if not sesion.browser.is_connected():
                self._dispose_session("El navegador fue cerrado.", "Sesion descartada porque Chromium se desconecto.")
                return None
            pages = [item for item in list(sesion.context.pages) if not item.is_closed()]
            if preferred is not None and any(item == preferred for item in pages):
                sesion.page = preferred
                return preferred
            if pages:
                sesion.page = pages[-1]
                self._log("Se recupero una pestaña activa de PAMI.")
                return sesion.page
            sesion.page = sesion.context.new_page()
            self._configure_page(sesion.page)
            self._log("Se abrio una nueva pestaña en la sesion activa de PAMI.")
            return sesion.page
        except Exception:
            self._dispose_session("La sesion no esta disponible.", "Sesion descartada por error de Playwright.")
            return None

    def _goto_with_recovery(self, page: Page, url: str, *, wait_until: str = "domcontentloaded") -> Page:
        page = self._recover_active_page(page) or page
        try:
            page.goto(url, wait_until=wait_until)
            return page
        except Exception as exc:
            if not self._is_navigation_transient_error(exc):
                raise
            self._log("La pestaña de PAMI quedo invalida durante la navegacion; se intenta recuperar.")
            recovered = self._recover_active_page(None)
            if recovered is None:
                raise RuntimeError("El navegador de PAMI se cerro durante el proceso. Volve a ejecutar la accion para reabrir la sesion.") from exc
            recovered.goto(url, wait_until=wait_until)
            return recovered

    def _evaluate_with_navigation_retry(
        self,
        page: Page,
        script: str,
        arg=_NO_EVALUATE_ARG,
        *,
        retries: int = 3,
        tolerate_navigation: bool = False,
    ):
        last_exc: Exception | None = None
        for attempt in range(1, max(1, retries) + 1):
            try:
                page = self._recover_active_page(page) or page
                self._wait_page_settled(page, timeout=3500)
                if arg is _NO_EVALUATE_ARG:
                    return page.evaluate(script)
                return page.evaluate(script, arg)
            except Exception as exc:
                last_exc = exc
                if not self._is_navigation_transient_error(exc):
                    raise
                self._log(f"PAMI recargo la pagina durante la lectura. Reintento {attempt}/{retries}.")
                page = self._recover_active_page(page) or page
                self._wait_page_settled(page, timeout=7000)
                try:
                    page.wait_for_timeout(700)
                except Exception:
                    pass

        if tolerate_navigation and last_exc is not None and self._is_navigation_transient_error(last_exc):
            self._log("PAMI navego despues de la accion; se continua con la pagina recargada.")
            return "NAVIGATED_AFTER_ACTION"
        if last_exc is not None:
            raise last_exc
        return None

    def _is_navigation_transient_error(self, exc: Exception) -> bool:
        message = str(exc or "").lower()
        return any(fragment in message for fragment in _NAVIGATION_TRANSIENT_ERRORS)

    def _wait_page_settled(self, page: Page, timeout: int = 5000) -> None:
        try:
            page.wait_for_load_state("domcontentloaded", timeout=timeout)
        except Exception:
            pass

    def _cleanup_dead_session(self) -> None:
        if self._sesion is None:
            return
        try:
            recovered = self._recover_active_page(self._sesion.page)
            if recovered is None:
                return
        except Exception:
            self._dispose_session("La sesion no esta disponible.", "Sesion descartada por error de Playwright.")
            return
        try:
            _ = list(self._sesion.context.pages)
        except Exception:
            self._dispose_session("El contexto del navegador no esta disponible.", "Contexto cerrado.")

    def _dispose_session(self, status_message: str, log_message_text: str) -> None:
        sesion = self._sesion
        self._sesion = None
        if sesion is not None:
            for fn in (sesion.context.close, sesion.browser.close, sesion.playwright.stop):
                try:
                    fn()
                except Exception:
                    pass
        self._status(status_message)
        self._log(log_message_text)

    def _log(self, message: str) -> None:
        self.log_callback(message)
        log_message(f"[LIBERAR CUPO] {message}")

    def _status(self, message: str) -> None:
        self.status_callback(message)

    def _on_console(self, msg) -> None:
        try:
            text = msg.text
        except Exception:
            text = str(msg)
        text = str(text or "").strip()
        if not text:
            return
        lowered = text.lower()
        noisy_fragments = (
            "mismo efector undefined",
            "inicio_fin_prestador_modulo_contrato",
            "inicio_fin_prestador_cred_contrato",
        )
        if any(fragment in lowered for fragment in noisy_fragments):
            return
        self._log(f"[PAMI] {text}")

    def _on_page_error(self, exc) -> None:
        self._log(f"Error de pagina: {exc}")


_EXTRAER_CANDIDATAS_TRANSMISION_SCRIPT = r"""
() => {
  const visible = (el) => !!el && (el.offsetParent !== null || el.getClientRects().length > 0);
  const text = (el) => (el?.textContent || el?.value || '').replace(/\s+/g, ' ').trim();
  const isGreen = (el) => {
    const target = el?.closest('button,a,span') || el;
    const cls = `${target?.className || ''} ${el?.className || ''}`.toLowerCase();
    return cls.includes('btn-success') || cls.includes('green');
  };
  const isBlue = (el) => {
    const target = el?.closest('button,a,span') || el;
    const cls = `${target?.className || ''} ${el?.className || ''}`.toLowerCase();
    return cls.includes('btn-primary') || cls.includes('btn-info') || cls.includes('blue');
  };
  const headers = Array.from(document.querySelectorAll('table thead th')).map((th) => text(th).toLowerCase());
  const idx = (names, fallback) => {
    for (const name of names) {
      const found = headers.findIndex((h) => h.includes(name));
      if (found >= 0) return found;
    }
    return fallback;
  };
  const iOrden = idx(['orden'], 0);
  const iEmision = idx(['emision'], 1);
  const iBenef = idx(['beneficio', 'afiliado'], 2);
  const iNombre = idx(['apellido', 'nombre'], 3);
  const iPractica = idx(['practica'], 4);
  const iTurno = idx(['turno'], 5);
  const iTransmitida = idx(['trasmitida', 'transmitida'], 6);

  return Array.from(document.querySelectorAll('table tbody tr')).map((tr) => {
    const tds = Array.from(tr.querySelectorAll('td'));
    if (tds.length < 6) return null;
    const celdas = tds.map(text);
    const acciones = tds[tds.length - 1] || tr;
    const check = acciones.querySelector('.fa-check');
    const transmitidaTexto = celdas[iTransmitida] || '';
    const transmitida = /^si\b/i.test(transmitidaTexto);
    const validada = check ? isBlue(check) : transmitida;
    const validarPendiente = check ? isGreen(check) : !validada;
    const hasCancelTarget = acciones.querySelector('.fa-arrow-right, i.transmitir, .transmitir, .fa-share, .fa-arrow-circle-right');
    if (!validarPendiente || transmitida || !hasCancelTarget) return null;
    return {
      n_orden: celdas[iOrden] || '',
      f_emision: celdas[iEmision] || '',
      beneficio: celdas[iBenef] || '',
      nombre: celdas[iNombre] || '',
      practica: celdas[iPractica] || '',
      turno: celdas[iTurno] || '',
      transmitida_texto: transmitidaTexto,
      validada: false,
      transmitida: false
    };
  }).filter((item) => item && item.n_orden);
}
"""


_BUSCAR_ORDEN_ACEPTACION_SCRIPT = r"""
(nroOrden) => {
  window.alert = () => true;
  window.confirm = () => true;
  const visible = (el) => !!el && (el.offsetParent !== null || el.getClientRects().length > 0);
  const setValue = (el, value) => {
    if (!el) return false;
    el.value = value || '';
    el.dispatchEvent(new Event('input', { bubbles: true }));
    el.dispatchEvent(new Event('change', { bubbles: true }));
    el.dispatchEvent(new KeyboardEvent('keyup', { bubbles: true }));
    return true;
  };
  const ordenInput =
    document.querySelector('input[name="n_orden"]') ||
    document.querySelector('#n_orden') ||
    Array.from(document.querySelectorAll('input[type="text"], input:not([type])'))
      .find((el) => visible(el) && /orden/i.test(`${el.name || ''} ${el.id || ''} ${el.placeholder || ''}`));
  if (!ordenInput) return 'NO_ORDEN_INPUT';
  setValue(ordenInput, nroOrden);

  const fechaDesde = document.querySelector('#f_emision_desde, input[name="f_emision_desde"]');
  setValue(fechaDesde, '01/01/2025');
  const fechaHasta = document.querySelector('#f_emision_hasta, input[name="f_emision_hasta"]');
  setValue(fechaHasta, '');

  const buscar = Array.from(document.querySelectorAll('button, input[type="submit"], input[type="button"], a'))
    .find((el) => visible(el) && (el.textContent || el.value || '').trim().toLowerCase() === 'buscar');
  if (!buscar) return 'NO_BUSCAR';
  buscar.click();
  return 'OK';
}
"""


_LEER_FILA_ACEPTACION_SCRIPT = r"""
(nroOrden) => {
  const norm = (value) => String(value || '').replace(/\s+/g, ' ').trim();
  for (const tr of document.querySelectorAll('table tbody tr')) {
    const tds = Array.from(tr.querySelectorAll('td'));
    const celdas = tds.map((td) => norm(td.textContent));
    if (!celdas.length || celdas[0] !== nroOrden) continue;
    return {
      encontrada: true,
      n_orden: celdas[0] || '',
      f_emision: celdas[1] || '',
      f_vencimiento: celdas[2] || '',
      beneficio: celdas[3] || '',
      nombre: celdas[4] || '',
      estado: celdas[5] || '',
      practica: celdas[6] || '',
      tiene_cancelar: !!tr.querySelector('i.boton-historial[data-estado="cancelar"], .boton-historial[data-estado="cancelar"], i.boton-historial[data-estado="anular"], .boton-historial[data-estado="anular"], .fa-ban, .fa-times, .fa-remove')
    };
  }
  return { encontrada: false };
}
"""


_CLICK_CANCELAR_ACEPTACION_SCRIPT = r"""
(nroOrden) => {
  window.alert = () => true;
  window.confirm = () => true;
  const visible = (el) => !!el && (el.offsetParent !== null || el.getClientRects().length > 0);
  for (const tr of document.querySelectorAll('table tbody tr')) {
    const tds = Array.from(tr.querySelectorAll('td'));
    if ((tds[0]?.textContent || '').replace(/\s+/g, ' ').trim() !== nroOrden) continue;
    const candidates = Array.from(tr.querySelectorAll(
      'i.boton-historial[data-estado="cancelar"], .boton-historial[data-estado="cancelar"], ' +
      'i.boton-historial[data-estado="anular"], .boton-historial[data-estado="anular"], ' +
      '.fa-ban, .fa-times, .fa-remove, button, a, span'
    )).filter(visible);
    const boton = candidates.find((el) => {
      const target = el.closest('button,a,span') || el;
      const cls = `${target.className || ''} ${el.className || ''}`.toLowerCase();
      const title = `${target.title || ''} ${el.title || ''} ${target.getAttribute('data-original-title') || ''}`.toLowerCase();
      return cls.includes('fa-ban') ||
        cls.includes('fa-times') ||
        cls.includes('fa-remove') ||
        cls.includes('btn-danger') ||
        title.includes('cancelar acept');
    });
    const target = boton?.closest('button,a,span') || boton;
    if (!target) return 'NO_CANCEL_BUTTON';
    target.click();
    return 'OK';
  }
  return 'NO_ROW';
}
"""


_CONFIRMAR_MODAL_CANCELACION_SCRIPT = r"""
() => {
  window.alert = () => true;
  window.confirm = () => true;
  const visible = (el) => !!el && (el.offsetParent !== null || el.getClientRects().length > 0);
  const modal =
    document.querySelector('.modal.show, .modal.in, .modal[style*="display: block"], .bootbox, [role="dialog"]') ||
    Array.from(document.querySelectorAll('.modal, .bootbox, [role="dialog"]')).find(visible);
  if (!modal) return 'NO_MODAL';

  const selects = Array.from(modal.querySelectorAll('select')).filter(visible);
  for (const select of selects) {
    if (select.value) continue;
    const opt = Array.from(select.options || []).find((item) => String(item.value || '').trim() && !item.disabled);
    if (opt) {
      select.value = opt.value;
      select.dispatchEvent(new Event('change', { bubbles: true }));
    }
  }

  const buttons = Array.from(modal.querySelectorAll('button, input[type="button"], input[type="submit"], a')).filter(visible);
  const negative = /(cerrar|volver|limpiar|no\b)/i;
  const button = buttons.find((el) => {
    const txt = (el.textContent || el.value || el.title || '').replace(/\s+/g, ' ').trim();
    const cls = `${el.className || ''}`.toLowerCase();
    if (negative.test(txt)) return false;
    return /cancelar acept|confirmar|aceptar|guardar|si\b|sí\b/i.test(txt) || cls.includes('btn-danger') || cls.includes('btn-success') || cls.includes('cambiar-estado');
  });
  if (!button) return 'NO_CONFIRM_BUTTON';
  button.click();
  return 'OK';
}
"""


_ORDEN_SIGUE_CANCELABLE_SCRIPT = r"""
(nroOrden) => {
  for (const tr of document.querySelectorAll('table tbody tr')) {
    const tds = Array.from(tr.querySelectorAll('td'));
    if ((tds[0]?.textContent || '').replace(/\s+/g, ' ').trim() !== nroOrden) continue;
    return !!tr.querySelector('i.boton-historial[data-estado="cancelar"], .boton-historial[data-estado="cancelar"], i.boton-historial[data-estado="anular"], .boton-historial[data-estado="anular"], .fa-ban, .fa-times, .fa-remove');
  }
  return false;
}
"""
