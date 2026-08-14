import json
import os
import queue
import re
import shutil
import subprocess
import sys
import threading
import traceback
import zipfile
from calendar import monthrange
from datetime import datetime, timedelta
from pathlib import Path
from tkinter import filedialog, messagebox, ttk

import customtkinter as ctk
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from app_credentials import sync_profile_records, upsert_shared_credentials_from_records
from app_paths import get_data_dir, get_log_file
from gmail_informes import (
    descargar_informes,
    get_connected_email,
    get_gmail_credentials_path,
    get_gmail_token_path,
)
from pami_documentacion import (
    IMAGE_EXTENSIONS,
    PamiDocumentacionController,
    _codigos_equivalentes_documentacion,
    _codigo_practica,
    _es_archivo_temporal_informe,
    _extraer_fecha_archivo,
    _limpiar_nombre_paciente_detectado,
    _normalizar_nombre_paciente_detectado_final,
    _practica_keywords_pdf,
    preparar_lote_documentacion,
    preparar_lote_documentacion_desde_cima,
    preparar_lote_documentacion_para_pami,
)


class DocumentacionOmeFrame(ctk.CTkFrame):
    def __init__(self, master, on_back=None) -> None:
        super().__init__(master, fg_color="transparent")
        self.on_back = on_back
        self.data_dir = get_data_dir()
        self.profiles_file = Path(self.data_dir) / "usuarios_documentacion.json"
        self.config_file = Path(self.data_dir) / "documentacion_config.json"
        self.config = self._load_config()
        self.saved_profiles = self._load_saved_profiles()
        self.event_queue: queue.Queue = queue.Queue()
        self.action_running = False
        self.stop_requested = False
        self.stop_event = threading.Event()
        self.password_visible = False
        self.ruta_excel: Path | None = None
        self.pdf_paths: list[Path] = []
        self.zip_extract_dir: Path | None = None
        self.lote: list[dict] = []
        self._resultado_por_item: dict[str, dict] = {}
        self.lote_verificado_en_pami = False
        self.controller: PamiDocumentacionController | None = None
        self.worker_profile: dict | None = None
        self.worker_hide_browser = False
        today = datetime.now()
        first_day = today.replace(day=1)
        last_day = today.replace(day=monthrange(today.year, today.month)[1])
        default_desde = first_day.strftime("%d/%m/%Y")
        default_hasta = last_day.strftime("%d/%m/%Y")
        self.gmail_desde_var = ctk.StringVar(value=str(self.config.get("gmail_desde") or default_desde).strip() or default_desde)
        self.gmail_hasta_var = ctk.StringVar(value=str(self.config.get("gmail_hasta") or default_hasta).strip() or default_hasta)
        default_gmail_destino = str(Path.home() / "Downloads" / "INFORMESNS")
        self.gmail_destino_var = ctk.StringVar(value=str(self.config.get("gmail_destino") or default_gmail_destino))
        self.gmail_status_var = ctk.StringVar(value="Gmail no conectado")
        self.gmail_connected = False
        self.show_general_progress = False

        self._build_ui()
        self.after(50, self._ensure_gmail_date_defaults)
        self._load_saved_excel_path()
        self._load_initial_profile_into_form()
        self.after(400, self._start_gmail_status_check)
        self.after(150, self._process_ui_queue)

    def _start_gmail_status_check(self) -> None:
        token_path = get_gmail_token_path()
        if not token_path.exists():
            return

        def runner() -> None:
            try:
                email = get_connected_email(interactive=False)
                if email:
                    self.event_queue.put(("gmail_connected", email))
            except Exception:
                self.event_queue.put(("gmail_status", "Token Gmail encontrado, pero hay que reconectar."))

        threading.Thread(target=runner, daemon=True).start()

    def _build_ui(self) -> None:
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=1)

        header = ctk.CTkFrame(self, corner_radius=8, fg_color="#f7fafc", border_width=1, border_color="#d8e2ec")
        header.grid(row=0, column=0, padx=8, pady=(6, 4), sticky="ew")
        header.grid_columnconfigure(1, weight=1)
        header.grid_columnconfigure(2, weight=0)

        if self.on_back:
            ctk.CTkButton(
                header,
                text="Volver",
                width=78,
                command=self._go_home,
                fg_color="#9aafc3",
                hover_color="#7f95aa",
            ).grid(row=0, column=0, rowspan=2, padx=(8, 10), pady=6, sticky="w")

        ctk.CTkLabel(
            header,
            text="Subir documentacion a OMEs",
            font=ctk.CTkFont(size=20, weight="bold"),
            text_color="#16324f",
        ).grid(row=0, column=1, padx=10, pady=(6, 1), sticky="w")
        ctk.CTkLabel(
            header,
            text="Cruza informes contra la bandeja de Transmision guardada como XLSX y carga documentacion en PAMI.",
            font=ctk.CTkFont(size=11),
            text_color="#51657a",
        ).grid(row=1, column=1, padx=10, pady=(0, 6), sticky="w")
        self.restart_button = ctk.CTkButton(
            header,
            text="Reiniciar app",
            command=self._restart_app,
            width=132,
            height=28,
            fg_color="#4f6378",
            hover_color="#3d4d61",
            font=ctk.CTkFont(size=11, weight="bold"),
        )
        self.restart_button.grid(row=0, column=2, rowspan=2, padx=(8, 10), pady=6, sticky="e")

        content = ctk.CTkFrame(self, corner_radius=8, fg_color="#ffffff")
        content.grid(row=1, column=0, padx=8, pady=(0, 8), sticky="nsew")
        content.grid_columnconfigure(0, weight=1)
        content.grid_rowconfigure(4, weight=1)

        files_frame = ctk.CTkFrame(content, corner_radius=8, fg_color="#f4f8fb", border_width=1, border_color="#d8e2ec")
        files_frame.grid(row=0, column=0, padx=8, pady=(8, 5), sticky="ew")
        files_frame.grid_columnconfigure(1, weight=1)

        self.instructions_label = ctk.CTkLabel(
            files_frame,
            text="1) En PAMI: Panel de prestaciones > Exportar/descargar bandeja de Transmision. "
            "2) Seleccionar aca la bandeja (.xls o .xlsx, no hace falta convertirla) y los informes. "
            "3) Tocar Previsualizar bandeja, revisar que este OK, y despues Subir documentos. "
            "Si el informe ya trae beneficio/DNI, se puede seleccionar solo informes y tocar Verificar en PAMI. "
            "Para turnos CIMA, seleccionar el Excel CIMA y usar Previsualizar turnos CIMA > Verificar en PAMI.",
            text_color="#51657a",
            wraplength=1300,
            justify="left",
        )
        self.instructions_label.grid(row=0, column=0, columnspan=3, padx=12, pady=(12, 4), sticky="w")

        self.select_excel_button = ctk.CTkButton(files_frame, text="Seleccionar bandeja", command=self._select_excel, width=170)
        self.select_excel_button.grid(
            row=1, column=0, padx=12, pady=(6, 6), sticky="w"
        )
        self.excel_label = ctk.CTkLabel(files_frame, text="Sin XLSX seleccionado", text_color="#51657a", wraplength=900)
        self.excel_label.grid(row=1, column=1, padx=8, pady=(6, 6), sticky="w")
        self.clear_excel_button = ctk.CTkButton(
            files_frame,
            text="Quitar XLSX",
            command=self._clear_excel,
            width=105,
            fg_color="#9aafc3",
            hover_color="#7f95aa",
        )
        self.clear_excel_button.grid(row=1, column=2, padx=(8, 12), pady=(6, 6), sticky="e")

        self.select_pdfs_button = ctk.CTkButton(files_frame, text="Seleccionar informes", command=self._select_pdfs, width=170)
        self.select_pdfs_button.grid(
            row=2, column=0, padx=12, pady=(0, 12), sticky="w"
        )
        self.pdf_label = ctk.CTkLabel(files_frame, text="Sin PDFs seleccionados", text_color="#51657a", wraplength=900)
        self.pdf_label.grid(row=2, column=1, columnspan=2, padx=8, pady=(0, 12), sticky="w")

        gmail_frame = ctk.CTkFrame(content, corner_radius=12, fg_color="#eef3f8")
        gmail_frame.grid(row=1, column=0, padx=14, pady=(0, 8), sticky="ew")
        gmail_frame.grid_columnconfigure(10, weight=1)

        ctk.CTkLabel(
            gmail_frame,
            text="Descargar informes desde Gmail",
            font=ctk.CTkFont(size=14, weight="bold"),
            text_color="#16324f",
        ).grid(row=0, column=0, columnspan=2, padx=12, pady=(10, 4), sticky="w")
        ctk.CTkLabel(gmail_frame, text="Desde", text_color="#16324f").grid(row=1, column=0, padx=(12, 6), pady=(4, 10), sticky="w")
        desde_frame = ctk.CTkFrame(gmail_frame, fg_color="transparent")
        desde_frame.grid(row=1, column=1, padx=(0, 12), pady=(4, 10), sticky="w")
        self.gmail_desde_entry = ctk.CTkEntry(desde_frame, textvariable=self.gmail_desde_var, width=100)
        self.gmail_desde_entry.grid(row=0, column=0, sticky="w")
        self.gmail_desde_calendar_button = ctk.CTkButton(
            desde_frame,
            text="...",
            width=30,
            command=lambda: self._open_date_picker(self.gmail_desde_var, "Fecha desde"),
        )
        self.gmail_desde_calendar_button.grid(row=0, column=1, padx=(4, 0), sticky="w")
        ctk.CTkLabel(gmail_frame, text="Hasta", text_color="#16324f").grid(row=1, column=2, padx=(0, 6), pady=(4, 10), sticky="w")
        hasta_frame = ctk.CTkFrame(gmail_frame, fg_color="transparent")
        hasta_frame.grid(row=1, column=3, padx=(0, 12), pady=(4, 10), sticky="w")
        self.gmail_hasta_entry = ctk.CTkEntry(hasta_frame, textvariable=self.gmail_hasta_var, width=100)
        self.gmail_hasta_entry.grid(row=0, column=0, sticky="w")
        self.gmail_hasta_calendar_button = ctk.CTkButton(
            hasta_frame,
            text="...",
            width=30,
            command=lambda: self._open_date_picker(self.gmail_hasta_var, "Fecha hasta"),
        )
        self.gmail_hasta_calendar_button.grid(row=0, column=1, padx=(4, 0), sticky="w")
        ctk.CTkLabel(gmail_frame, text="Carpeta", text_color="#16324f").grid(row=1, column=4, padx=(0, 6), pady=(4, 10), sticky="w")
        self.gmail_destino_entry = ctk.CTkEntry(gmail_frame, textvariable=self.gmail_destino_var, width=360)
        self.gmail_destino_entry.grid(row=1, column=5, padx=(0, 8), pady=(4, 10), sticky="w")
        self.gmail_browse_button = ctk.CTkButton(gmail_frame, text="Examinar", command=self._browse_gmail_destino, width=90)
        self.gmail_browse_button.grid(row=1, column=6, padx=(0, 8), pady=(4, 10), sticky="w")
        self.gmail_connect_button = ctk.CTkButton(gmail_frame, text="Conectar Gmail", command=self._connect_gmail_clicked, width=130)
        self.gmail_connect_button.grid(row=1, column=7, padx=(0, 8), pady=(4, 10), sticky="w")
        self.gmail_download_button = ctk.CTkButton(
            gmail_frame,
            text="Descargar informes",
            command=self._download_gmail_clicked,
            width=140,
            state="disabled",
        )
        self.gmail_download_button.grid(row=1, column=8, padx=(0, 8), pady=(4, 10), sticky="w")
        self.gmail_progress = ctk.CTkProgressBar(gmail_frame, mode="indeterminate", width=140)
        self.gmail_progress.grid(row=1, column=9, padx=(0, 10), pady=(4, 10), sticky="w")
        self.gmail_progress.set(0)
        self.gmail_status_label = ctk.CTkLabel(gmail_frame, textvariable=self.gmail_status_var, text_color="#51657a")
        self.gmail_status_label.grid(row=2, column=0, columnspan=10, padx=12, pady=(0, 10), sticky="w")

        profiles_frame = ctk.CTkFrame(content, corner_radius=12, fg_color="#eef3f8")
        profiles_frame.grid(row=2, column=0, padx=14, pady=(0, 8), sticky="ew")
        profiles_frame.grid_columnconfigure(1, weight=0)
        profiles_frame.grid_columnconfigure(4, weight=1)

        profile_options = self._profile_options()
        self.profile_var = ctk.StringVar(value=profile_options[0] if profile_options else "")
        self.profile_name_var = ctk.StringVar(value="")
        self.profile_user_var = ctk.StringVar(value="")
        self.profile_password_var = ctk.StringVar(value="")
        self.hide_browser_var = ctk.BooleanVar(value=False)

        ctk.CTkLabel(profiles_frame, text="Perfil", font=ctk.CTkFont(size=13, weight="bold"), text_color="#16324f").grid(
            row=0, column=0, padx=(12, 8), pady=10, sticky="w"
        )
        self.profile_combo = ctk.CTkComboBox(
            profiles_frame,
            values=profile_options or [""],
            variable=self.profile_var,
            command=self._on_profile_selected,
            width=240,
        )
        self.profile_combo.grid(row=0, column=1, padx=8, pady=8, sticky="w")
        ctk.CTkButton(profiles_frame, text="Nuevo", command=self._new_profile, width=90).grid(row=0, column=2, padx=8, pady=8)
        ctk.CTkButton(profiles_frame, text="Guardar", command=self._save_current_profile, width=100).grid(row=0, column=3, padx=(8, 12), pady=8)

        ctk.CTkLabel(profiles_frame, text="Cliente", font=ctk.CTkFont(size=13, weight="bold"), text_color="#16324f").grid(
            row=1, column=0, padx=(12, 8), pady=(0, 10), sticky="w"
        )
        self.client_entry = ctk.CTkEntry(profiles_frame, textvariable=self.profile_name_var, placeholder_text="Nombre del cliente")
        self.client_entry.grid(row=1, column=1, padx=8, pady=(0, 10), sticky="w")

        creds = ctk.CTkFrame(profiles_frame, fg_color="transparent")
        creds.grid(row=2, column=1, columnspan=3, padx=8, pady=(0, 12), sticky="w")
        ctk.CTkLabel(creds, text="Usuario", font=ctk.CTkFont(size=13, weight="bold"), text_color="#16324f").grid(
            row=0, column=0, padx=(0, 8), sticky="w"
        )
        self.user_entry = ctk.CTkEntry(creds, textvariable=self.profile_user_var, placeholder_text="Usuario PAMI", width=190)
        self.user_entry.grid(row=0, column=1, padx=(0, 10), sticky="w")
        ctk.CTkLabel(creds, text="Clave", font=ctk.CTkFont(size=13, weight="bold"), text_color="#16324f").grid(
            row=0, column=2, padx=(0, 8), sticky="w"
        )
        self.password_entry = ctk.CTkEntry(creds, textvariable=self.profile_password_var, placeholder_text="Clave PAMI", show="*", width=190)
        self.password_entry.grid(row=0, column=3, padx=(0, 10), sticky="w")
        ctk.CTkButton(creds, text="Ver", command=self._toggle_password, width=70).grid(row=0, column=4, padx=(0, 10), sticky="w")
        ctk.CTkCheckBox(creds, text="No ver navegador", variable=self.hide_browser_var, text_color="#16324f").grid(row=0, column=5, sticky="w")

        actions_frame = ctk.CTkFrame(content, corner_radius=12, fg_color="#eef3f8")
        actions_frame.grid(row=3, column=0, padx=14, pady=(0, 8), sticky="ew")
        actions_frame.grid_columnconfigure(7, weight=0)
        actions_frame.grid_columnconfigure(8, weight=1)
        self.analyze_button = ctk.CTkButton(
            actions_frame,
            text="Analizar informes",
            command=lambda: self._run_background(self._analyze_reports_only),
            width=170,
            fg_color="#3d84c6",
        )
        self.analyze_button.grid(row=0, column=0, padx=12, pady=12, sticky="w")
        self.preview_button = ctk.CTkButton(
            actions_frame,
            text="Previsualizar bandeja",
            command=lambda: self._run_background(self._build_preview),
            width=190,
            fg_color="#3d84c6",
        )
        self.preview_button.grid(row=0, column=1, padx=8, pady=12, sticky="w")
        self.preview_cima_button = ctk.CTkButton(
            actions_frame,
            text="Previsualizar turnos CIMA",
            command=lambda: self._run_background(self._build_preview_cima),
            width=210,
            fg_color="#4f6378",
            hover_color="#3d4d61",
        )
        self.preview_cima_button.grid(row=0, column=2, padx=8, pady=12, sticky="w")
        self.clear_button = ctk.CTkButton(
            actions_frame,
            text="Limpiar panel",
            command=self._clear_panel,
            width=130,
            fg_color="#9aafc3",
            hover_color="#7f95aa",
        )
        self.clear_button.grid(row=0, column=2, padx=8, pady=12, sticky="w")
        self.verify_pami_button = ctk.CTkButton(
            actions_frame,
            text="Verificar en PAMI",
            command=lambda: self._run_background(self._verify_lote_en_pami),
            width=160,
            state="disabled",
        )
        self.verify_pami_button.grid(row=0, column=3, padx=8, pady=12, sticky="w")
        self.upload_button = ctk.CTkButton(
            actions_frame,
            text="Subir documentos",
            command=lambda: self._run_background(self._upload_lote),
            width=150,
            fg_color="#2f9e44",
            hover_color="#25833a",
            state="disabled",
        )
        self.upload_button.grid(row=0, column=4, padx=8, pady=12, sticky="w")
        self.close_button = ctk.CTkButton(actions_frame, text="Cerrar navegador", command=self._close_browser, width=140)
        self.close_button.grid(row=0, column=5, padx=8, pady=12, sticky="w")
        self.stop_button = ctk.CTkButton(
            actions_frame,
            text="Detener",
            command=self._request_stop,
            width=100,
            fg_color="#c92a2a",
            hover_color="#a61e1e",
            state="disabled",
        )
        self.stop_button.grid(row=0, column=6, padx=8, pady=12, sticky="w")
        self.status_label = ctk.CTkLabel(actions_frame, text=f"Log en: {get_log_file()}", text_color="#51657a")
        self.status_label.grid(row=0, column=7, padx=(8, 12), pady=12, sticky="w")
        self.action_progress_label = ctk.CTkLabel(actions_frame, text="", text_color="#51657a")
        self.action_progress_label.grid(row=1, column=0, columnspan=2, padx=12, pady=(0, 10), sticky="w")
        self.action_progress = ctk.CTkProgressBar(actions_frame, mode="indeterminate")
        self.action_progress.grid(row=1, column=2, columnspan=7, padx=(8, 12), pady=(0, 10), sticky="ew")
        self.action_progress.set(0)
        self.profile_var.trace_add("write", lambda *_args: self._apply_cliente_mode())
        self.profile_name_var.trace_add("write", lambda *_args: self._apply_cliente_mode())

        results_frame = ctk.CTkFrame(content, corner_radius=12, fg_color="#eef3f8")
        results_frame.grid(row=4, column=0, padx=14, pady=(0, 8), sticky="nsew")
        results_frame.grid_columnconfigure(0, weight=1)
        results_frame.grid_rowconfigure(0, weight=1)

        columns = ("archivo", "pdf", "dni", "fecha", "ome", "paciente", "beneficio", "practica", "estado")
        self.results_columns = columns
        style = ttk.Style()
        style.configure("Report.Treeview", rowheight=27, font=("Segoe UI", 9))
        style.configure("Report.Treeview.Heading", font=("Segoe UI", 9, "bold"))
        self.results_table = ttk.Treeview(results_frame, columns=columns, show="headings", style="Report.Treeview", height=10)
        headings = {
            "archivo": "PDF",
            "pdf": "Paciente detectado",
            "dni": "DNI",
            "fecha": "Fecha informe",
            "ome": "OME",
            "paciente": "Paciente XLS",
            "beneficio": "Beneficio",
            "practica": "Practica",
            "estado": "Estado",
        }
        widths = {
            "archivo": 120,
            "pdf": 125,
            "dni": 78,
            "fecha": 95,
            "ome": 105,
            "paciente": 150,
            "beneficio": 118,
            "practica": 405,
            "estado": 560,
        }
        for col in columns:
            self.results_table.heading(col, text=headings[col])
            self.results_table.column(col, width=widths[col], minwidth=widths[col], anchor="w", stretch=col in {"practica", "estado"})
        self.results_table.grid(row=0, column=0, padx=(12, 0), pady=12, sticky="nsew")
        self.results_table.bind("<Double-1>", self._on_results_table_double_click)
        scrollbar = ttk.Scrollbar(results_frame, orient="vertical", command=self.results_table.yview, style="Report.Vertical.TScrollbar")
        scrollbar.grid(row=0, column=1, padx=(0, 12), pady=12, sticky="ns")
        horizontal_scrollbar = ttk.Scrollbar(results_frame, orient="horizontal", command=self.results_table.xview)
        horizontal_scrollbar.grid(row=1, column=0, padx=(12, 0), pady=(0, 8), sticky="ew")
        self.results_table.configure(yscrollcommand=scrollbar.set, xscrollcommand=horizontal_scrollbar.set)
        self.copy_beneficio_button = ctk.CTkButton(
            results_frame,
            text="Copiar beneficio",
            command=self._copy_selected_beneficio,
            width=130,
            fg_color="#8aa4bd",
            hover_color="#748ea8",
        )
        self.copy_beneficio_button.grid(row=2, column=0, padx=12, pady=(0, 10), sticky="w")
        self.open_informe_button = ctk.CTkButton(
            results_frame,
            text="Abrir informe",
            command=self._open_selected_informe,
            width=130,
            fg_color="#3d84c6",
        )
        self.open_informe_button.grid(row=2, column=0, padx=(158, 12), pady=(0, 10), sticky="w")
        self.export_report_button = ctk.CTkButton(
            results_frame,
            text="Descargar reporte",
            command=self._export_reporte,
            width=150,
            fg_color="#3d84c6",
        )
        self.export_report_button.grid(row=2, column=0, padx=(304, 12), pady=(0, 10), sticky="w")

        log_frame = ctk.CTkFrame(content, corner_radius=12, fg_color="#eef3f8")
        log_frame.grid(row=5, column=0, padx=14, pady=(0, 14), sticky="ew")
        log_frame.grid_columnconfigure(0, weight=1)
        self.log_text = ctk.CTkTextbox(log_frame, height=120, wrap="word")
        self.log_text.grid(row=0, column=0, padx=12, pady=12, sticky="ew")
        self.log_text.configure(state="disabled")

    def _browse_gmail_destino(self) -> None:
        selected = filedialog.askdirectory(
            title="Seleccionar carpeta destino para informes Gmail",
            initialdir=self.gmail_destino_var.get() or str(Path.home() / "Downloads"),
        )
        if not selected:
            return
        self.gmail_destino_var.set(selected)
        self._save_config({"gmail_destino": selected})

    def _open_date_picker(self, target_var: ctk.StringVar, title: str) -> None:
        try:
            current = self._parse_gmail_date(target_var.get())
        except RuntimeError:
            current = datetime.now()
        selected_month = {"year": current.year, "month": current.month}

        popup = ctk.CTkToplevel(self)
        popup.title(title)
        popup.resizable(False, False)
        popup.transient(self.winfo_toplevel())
        popup.grab_set()

        header = ctk.CTkFrame(popup, fg_color="transparent")
        header.grid(row=0, column=0, padx=10, pady=(10, 4), sticky="ew")
        header.grid_columnconfigure(1, weight=1)
        month_label = ctk.CTkLabel(header, text="", font=ctk.CTkFont(size=14, weight="bold"))

        days_frame = ctk.CTkFrame(popup, fg_color="transparent")
        days_frame.grid(row=1, column=0, padx=10, pady=(4, 10))

        def pick_day(day: int) -> None:
            picked = datetime(selected_month["year"], selected_month["month"], day)
            value = picked.strftime("%d/%m/%Y")
            target_var.set(value)
            key = "gmail_desde" if target_var is self.gmail_desde_var else "gmail_hasta"
            self._save_config({key: value})
            popup.destroy()

        def render_days() -> None:
            for child in days_frame.winfo_children():
                child.destroy()
            year = selected_month["year"]
            month = selected_month["month"]
            month_label.configure(text=f"{month:02d}/{year}")
            for col, day_name in enumerate(("L", "M", "X", "J", "V", "S", "D")):
                ctk.CTkLabel(days_frame, text=day_name, width=32, text_color="#51657a").grid(row=0, column=col, padx=2, pady=2)
            first_weekday = datetime(year, month, 1).weekday()
            last_day = monthrange(year, month)[1]
            row = 1
            col = first_weekday
            for day in range(1, last_day + 1):
                ctk.CTkButton(
                    days_frame,
                    text=str(day),
                    width=32,
                    height=28,
                    command=lambda selected_day=day: pick_day(selected_day),
                ).grid(row=row, column=col, padx=2, pady=2)
                col += 1
                if col > 6:
                    col = 0
                    row += 1

        def change_month(delta: int) -> None:
            month = selected_month["month"] + delta
            year = selected_month["year"]
            if month < 1:
                month = 12
                year -= 1
            elif month > 12:
                month = 1
                year += 1
            selected_month["year"] = year
            selected_month["month"] = month
            render_days()

        ctk.CTkButton(header, text="<", width=34, command=lambda: change_month(-1)).grid(row=0, column=0, padx=(0, 6))
        month_label.grid(row=0, column=1, padx=6)
        ctk.CTkButton(header, text=">", width=34, command=lambda: change_month(1)).grid(row=0, column=2, padx=(6, 0))
        render_days()
        popup.focus_force()

    def _connect_gmail_clicked(self) -> None:
        if self.gmail_connected:
            if not messagebox.askyesno(
                "Cambiar cuenta Gmail",
                "Se borrara la conexion Gmail actual y se abrira el navegador para elegir otra cuenta. Continuar?",
            ):
                return
            token_path = get_gmail_token_path()
            try:
                if token_path.exists():
                    token_path.unlink()
            except Exception as exc:
                messagebox.showerror("Gmail", f"No se pudo borrar el token Gmail:\n{exc}")
                return
            self.gmail_connected = False
            self.event_queue.put(("gmail_status", "Gmail desconectado. Conecta la nueva cuenta."))
            self._update_gmail_connection_ui()
        self.gmail_progress.start()
        self._run_background(self._connect_gmail, show_general_progress=False)

    def _download_gmail_clicked(self) -> None:
        if not self.gmail_connected:
            messagebox.showwarning("Gmail", "Primero conecta Gmail.")
            return
        self.gmail_progress.start()
        self._run_background(self._download_gmail, show_general_progress=False)

    def _connect_gmail(self) -> None:
        credentials_path = get_gmail_credentials_path()
        token_path = get_gmail_token_path()
        self._push_log(f"Conectando Gmail con credenciales: {credentials_path}")
        email = get_connected_email(credentials_path=credentials_path, token_path=token_path, interactive=True)
        self.event_queue.put(("gmail_connected", email or "cuenta Gmail"))
        self._push_log(f"Gmail conectado: {email or 'cuenta Gmail'}")

    def _download_gmail(self) -> None:
        fecha_desde, fecha_hasta = self._gmail_query_dates()
        destino = Path(self.gmail_destino_var.get().strip() or Path.home() / "Downloads" / "INFORMESNS")
        self._save_config({"gmail_destino": str(destino)})
        self._push_log(f"Descargando informes Gmail desde {fecha_desde} hasta {fecha_hasta} en {destino}...")
        count = descargar_informes(
            credentials_path=get_gmail_credentials_path(),
            token_path=get_gmail_token_path(),
            destino=destino,
            fecha_desde=fecha_desde,
            fecha_hasta=fecha_hasta,
            callback_progreso=lambda filename: self.event_queue.put(("gmail_file", filename)),
            should_stop=lambda: self.stop_event.is_set(),
        )
        if self.stop_event.is_set():
            mensaje = f"Descarga Gmail detenida: {count} archivos descargados en {destino}."
        else:
            mensaje = f"{count} archivos descargados en {destino}."
        self.event_queue.put(("gmail_done", mensaje))

    def _gmail_query_dates(self) -> tuple[str, str]:
        self._ensure_gmail_date_defaults()
        desde = self._parse_gmail_date(self.gmail_desde_var.get())
        hasta = self._parse_gmail_date(self.gmail_hasta_var.get())
        if hasta < desde:
            raise RuntimeError("La fecha hasta no puede ser menor que la fecha desde.")
        self._save_config({"gmail_desde": desde.strftime("%d/%m/%Y"), "gmail_hasta": hasta.strftime("%d/%m/%Y")})
        return desde.strftime("%Y/%m/%d"), (hasta + timedelta(days=1)).strftime("%Y/%m/%d")

    def _pami_turno_date_range(self) -> tuple[str, str]:
        self._ensure_gmail_date_defaults()
        desde = self._parse_gmail_date(self.gmail_desde_var.get())
        hasta = self._parse_gmail_date(self.gmail_hasta_var.get())
        if hasta < desde:
            raise RuntimeError("La fecha hasta no puede ser menor que la fecha desde.")
        self._save_config({"gmail_desde": desde.strftime("%d/%m/%Y"), "gmail_hasta": hasta.strftime("%d/%m/%Y")})
        return desde.strftime("%d/%m/%Y"), hasta.strftime("%d/%m/%Y")

    def _ensure_gmail_date_defaults(self) -> None:
        today = datetime.now()
        first_day = today.replace(day=1)
        last_day = today.replace(day=monthrange(today.year, today.month)[1])
        if not str(self.gmail_desde_var.get() or "").strip():
            self.gmail_desde_var.set(first_day.strftime("%d/%m/%Y"))
        if not str(self.gmail_hasta_var.get() or "").strip():
            self.gmail_hasta_var.set(last_day.strftime("%d/%m/%Y"))

    def _parse_gmail_date(self, value: str) -> datetime:
        value = str(value or "").strip()
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%Y/%m/%d"):
            try:
                return datetime.strptime(value, fmt)
            except ValueError:
                continue
        raise RuntimeError("Ingresa las fechas de Gmail en formato dd/mm/aaaa.")

    def _update_gmail_connection_ui(self) -> None:
        if not hasattr(self, "gmail_connect_button"):
            return
        if self.gmail_connected:
            self.gmail_connect_button.configure(text="Cambiar cuenta")
            self.gmail_download_button.configure(state="normal" if not self.action_running else "disabled")
        else:
            self.gmail_connect_button.configure(text="Conectar Gmail")
            self.gmail_download_button.configure(state="disabled")

    def _cliente_actual(self) -> str:
        texto = f"{self.profile_var.get()} {self.profile_name_var.get()}".lower()
        if "cima" in texto:
            return "cima"
        return "caballito"

    def _apply_cliente_mode(self) -> None:
        if not hasattr(self, "preview_button"):
            return
        enabled = not self.action_running
        cliente = self._cliente_actual()
        if cliente == "cima":
            self.instructions_label.configure(
                text=(
                    "CIMA: si tenes la bandeja de Transmision PAMI, selecciona ese XLSX y usa "
                    "Analizar informes para revisar la deteccion antes de cruzar. Despues usa "
                    "Previsualizar bandeja. Si no tenes XLSX y los informes traen beneficio/DNI, "
                    "usa Verificar en PAMI. Para validar contra turnos internos CIMA, "
                    "selecciona el Excel CIMA y usa Previsualizar turnos CIMA."
                )
            )
            self.select_excel_button.configure(text="Seleccionar bandeja")
            self.analyze_button.configure(
                text="Analizar informes",
                state="normal" if enabled else "disabled",
                width=170,
            )
            self.analyze_button.grid(row=0, column=0, padx=12, pady=12, sticky="w")
            self.preview_button.configure(
                text="Previsualizar bandeja",
                state="normal" if enabled else "disabled",
                width=190,
            )
            self.preview_button.grid(row=0, column=1, padx=8, pady=12, sticky="w")
            self.preview_cima_button.configure(
                text="Previsualizar turnos CIMA",
                state="normal" if enabled else "disabled",
                width=205,
            )
            self.preview_cima_button.grid(row=0, column=2, padx=8, pady=12, sticky="w")
            self.verify_pami_button.configure(text="Verificar en PAMI", width=160)
        else:
            self.instructions_label.configure(
                text=(
                    "Caballito: descarga la bandeja desde PAMI (.xls o .xlsx, no hace falta convertirla), "
                    "selecciona esa bandeja y los informes. Si queres revisar primero que entiende la app, "
                    "toca Analizar informes. Despues usa Previsualizar bandeja, revisa el resultado y "
                    "Subir documentos. Si el informe trae beneficio/DNI, tambien podes usar Verificar en PAMI."
                )
            )
            self.select_excel_button.configure(text="Seleccionar bandeja")
            self.analyze_button.configure(text="Analizar informes", state="normal" if enabled else "disabled", width=170)
            self.analyze_button.grid(row=0, column=0, padx=12, pady=12, sticky="w")
            self.preview_button.configure(text="Previsualizar bandeja", state="normal" if enabled else "disabled", width=190)
            self.preview_button.grid(row=0, column=1, padx=8, pady=12, sticky="w")
            self.preview_cima_button.grid_remove()
            self.verify_pami_button.configure(text="Verificar en PAMI", width=160)
        self._layout_common_action_buttons()

    def _layout_common_action_buttons(self) -> None:
        offset = 3 if self._cliente_actual() == "cima" else 2
        self.clear_button.grid(row=0, column=offset, padx=8, pady=12, sticky="w")
        self.verify_pami_button.grid(row=0, column=offset + 1, padx=8, pady=12, sticky="w")
        self.upload_button.grid(row=0, column=offset + 2, padx=8, pady=12, sticky="w")
        self.close_button.grid(row=0, column=offset + 3, padx=8, pady=12, sticky="w")
        self.stop_button.grid(row=0, column=offset + 4, padx=8, pady=12, sticky="w")
        self.status_label.grid(row=0, column=offset + 5, padx=(8, 12), pady=12, sticky="e")
        self.action_progress_label.grid(row=1, column=0, columnspan=2, padx=12, pady=(0, 10), sticky="w")
        self.action_progress.grid(row=1, column=2, columnspan=offset + 4, padx=(8, 12), pady=(0, 10), sticky="ew")

    def _select_excel(self) -> None:
        ruta = filedialog.askopenfilename(
            title="Seleccionar bandeja de Transmision (.xls o .xlsx)",
            filetypes=[("Excel actualizado", "*.xlsx *.xlsm"), ("XLS original de PAMI", "*.xls"), ("Todos los archivos", "*.*")],
        )
        if not ruta:
            return
        self.ruta_excel = Path(ruta)
        self.excel_label.configure(text=str(self.ruta_excel))
        self._save_config({"ruta_excel": str(self.ruta_excel)})

    def _clear_excel(self) -> None:
        if self.action_running:
            messagebox.showwarning("Atencion", "Espera a que termine la accion actual antes de quitar el XLSX.")
            return
        self.ruta_excel = None
        self.excel_label.configure(text="Sin XLSX seleccionado")
        self._save_config({"ruta_excel": ""})
        self.status_label.configure(text="XLSX quitado. Si los informes tienen DNI o beneficio, podes usar Verificar en PAMI.")

    def _select_pdfs(self) -> None:
        rutas = filedialog.askopenfilenames(
            title="Seleccionar informes",
            filetypes=[
                ("Informes", "*.pdf *.doc *.docx *.jpg *.jpeg *.png *.zip"),
                ("PDF", "*.pdf"),
                ("Word", "*.doc *.docx"),
                ("Imagenes", "*.jpg *.jpeg *.png"),
                ("ZIP", "*.zip"),
                ("Todos los archivos", "*.*"),
            ],
        )
        if not rutas:
            return
        self._clear_zip_extract_dir()
        self.pdf_paths = self._expand_selected_informes([Path(ruta) for ruta in rutas])
        self._reset_lote_view(clear_log=True)
        self.pdf_label.configure(text=f"{len(self.pdf_paths)} informes seleccionados")
        self.verify_pami_button.configure(state="normal" if self.pdf_paths else "disabled")
        navegador = "Se reutilizara el navegador PAMI abierto." if self.controller else "La proxima accion abrira o preparara navegador PAMI."
        self.status_label.configure(text=f"Nuevo lote seleccionado. {navegador}")

    def _expand_selected_informes(self, rutas: list[Path]) -> list[Path]:
        informes: list[Path] = []
        zip_files = [ruta for ruta in rutas if ruta.suffix.lower() == ".zip"]
        informe_extensiones = {".pdf", ".doc", ".docx"} | IMAGE_EXTENSIONS
        informes.extend(
            ruta
            for ruta in rutas
            if ruta.suffix.lower() in informe_extensiones and not _es_archivo_temporal_informe(ruta)
        )
        if not zip_files:
            return informes

        extract_root = Path(self.data_dir) / "documentacion_zip_extraidos"
        extract_root.mkdir(parents=True, exist_ok=True)
        self.zip_extract_dir = extract_root
        for zip_path in zip_files:
            destino = extract_root / zip_path.stem
            if destino.exists():
                shutil.rmtree(destino, ignore_errors=True)
            destino.mkdir(parents=True, exist_ok=True)
            try:
                with zipfile.ZipFile(zip_path) as zf:
                    for member in zf.infolist():
                        member_path = Path(member.filename)
                        if (
                            member.is_dir()
                            or member_path.suffix.lower() not in informe_extensiones
                            or _es_archivo_temporal_informe(member_path)
                        ):
                            continue
                        target = destino / member_path.name
                        with zf.open(member) as source, target.open("wb") as output:
                            shutil.copyfileobj(source, output)
                        informes.append(target)
                self._push_log(f"ZIP extraido: {zip_path.name}")
            except zipfile.BadZipFile:
                self._push_log(f"ZIP invalido o corrupto: {zip_path}")
            except Exception as exc:
                self._push_log(f"No se pudo extraer ZIP {zip_path}: {exc}")
        return informes

    def _clear_zip_extract_dir(self) -> None:
        if self.zip_extract_dir and self.zip_extract_dir.exists():
            shutil.rmtree(self.zip_extract_dir, ignore_errors=True)
        self.zip_extract_dir = None

    def _analyze_reports_only(self) -> None:
        if not self.pdf_paths:
            raise RuntimeError("Selecciona uno o mas informes para analizar.")
        self._push_log("Analizando informes sin cruzar contra bandeja ni PAMI...")
        self.lote_verificado_en_pami = False
        self.lote = preparar_lote_documentacion_para_pami(
            self.pdf_paths,
            progress_callback=self._preview_progress_callback,
            cancel_event=self.stop_event,
        )
        self.lote = [self._marcar_fuente_informe(item) for item in self.lote]
        self.event_queue.put(("preview", self.lote))

        total_filas = len(self.lote)
        con_practica = sum(1 for item in self.lote if self._codigos_practica_texto((item.get("prestacion") or {}).get("practica", "")))
        sin_practica = sum(1 for item in self.lote if item.get("estado") == "sin_practica_informe")
        sin_datos = sum(1 for item in self.lote if item.get("estado") == "sin_datos_busqueda")
        para_pami = sum(1 for item in self.lote if item.get("estado") == "para_verificar_pami")
        mensaje = (
            f"Analisis local listo: {len(self.pdf_paths)} informe(s), {total_filas} fila(s); "
            f"{con_practica} con practica detectada"
        )
        if para_pami:
            mensaje += f" | {para_pami} para verificar en PAMI"
        if sin_practica:
            mensaje += f" | {sin_practica} sin practica detectada"
        if sin_datos:
            mensaje += f" | {sin_datos} sin DNI/beneficio"
        if self.stop_event.is_set():
            mensaje += " | detenido por usuario"
        mensaje += "."
        self._push_status(mensaje)
        self.event_queue.put(("done_message", mensaje))

    def _build_preview(self) -> None:
        if not self.ruta_excel or not self.ruta_excel.exists():
            raise RuntimeError("Selecciona el XLSX de Transmision actualizado.")
        if not self.pdf_paths:
            raise RuntimeError("Selecciona uno o mas PDFs.")
        self._push_log("Analizando informes y cruzando contra XLSX de Transmision...")
        self.lote_verificado_en_pami = False
        self.lote = preparar_lote_documentacion(
            self.ruta_excel,
            self.pdf_paths,
            progress_callback=self._preview_progress_callback,
            cancel_event=self.stop_event,
        )
        self.lote = [self._marcar_fuente_bandeja(item) for item in self.lote]
        self.event_queue.put(("preview", self.lote))
        listos = sum(1 for item in self.lote if item.get("estado") == "listo")
        if self.stop_event.is_set():
            self._push_status(f"Previsualizacion detenida: {len(self.lote)} filas generadas.")
        else:
            self._push_status(f"Previsualizacion lista: {listos}/{len(self.lote)} vinculados.")

    def _build_preview_cima(self) -> None:
        if not self.ruta_excel or not self.ruta_excel.exists():
            raise RuntimeError("Selecciona el Excel de turnos CIMA.")
        if not self.pdf_paths:
            raise RuntimeError("Selecciona uno o mas informes.")
        self._push_log("Analizando informes y cruzando contra Excel de turnos CIMA...")
        self.lote_verificado_en_pami = False
        self.lote = preparar_lote_documentacion_desde_cima(
            self.ruta_excel,
            self.pdf_paths,
            progress_callback=self._preview_progress_callback,
            cancel_event=self.stop_event,
        )
        self.lote = [self._marcar_fuente_bandeja(item) for item in self.lote]
        self.event_queue.put(("preview", self.lote))
        verificables = sum(1 for item in self.lote if item.get("estado") == "para_verificar_pami")
        revisar = sum(1 for item in self.lote if item.get("estado") in {"sin_coincidencia", "sin_practica_informe"})
        if self.stop_event.is_set():
            self._push_status(f"Previsualizacion CIMA detenida: {len(self.lote)} filas generadas.")
        else:
            mensaje = f"Previsualizacion CIMA lista: {verificables} para cruzar con PAMI"
            if revisar:
                mensaje += f" | {revisar} para revisar"
            self._push_status(mensaje + ".")

    def _preview_progress_callback(self, current: int, total: int, path: Path, partial_lote: list[dict]) -> None:
        self.event_queue.put(("action_progress", (current, total, f"Analizando {current}/{total}: {path.name}")))
        self.event_queue.put(("preview_partial", partial_lote))
        self._push_log(f"Analizado {current}/{total}: {path.name}")

    def _open_pami(self) -> None:
        self._ensure_controller()
        assert self.controller is not None
        self.controller.abrir_pami()

    def _upload_lote(self) -> None:
        self._ensure_controller()
        assert self.controller is not None

        if self.lote:
            self._push_log(f"Estados antes de subir: {self._resumen_estados(self.lote)}")
        self._push_log("Subir documentos: se usa el estado actual del lote; no se revalida en PAMI.")
        if self.stop_event.is_set():
            mensaje = "Operacion detenida por el usuario."
            self._push_status(mensaje)
            self.event_queue.put(("done_message", mensaje))
            return

        listos = self._listos_sin_duplicar()
        self._push_log(f"Subir documentos: {len(listos)} informe(s) listos para cargar.")
        if not listos:
            sin_dni = sum(
                1
                for item in self.lote
                if item.get("estado") in {"faltante_ome", "sin_coincidencia"}
                and not (str(item.get("dni_pdf", "")).strip() or str(item.get("beneficio_pdf", "")).strip())
            )
            mensaje = "No hay informes vinculados a OMEs pendientes para subir."
            if sin_dni:
                mensaje += f" {sin_dni} quedaron sin verificar porque no tienen DNI ni beneficio en el informe."
            pendientes_verificar = sum(1 for item in self._items_para_revalidar_pami(solo_listos=False))
            if pendientes_verificar:
                mensaje += f" Hay {pendientes_verificar} informe(s) pendientes de Verificar en PAMI."
            self._push_status(mensaje)
            self.event_queue.put(("done_message", mensaje))
            return
        resultados = self.controller.cargar_lote(listos)
        resultados = [self._marcar_resultado_carga(item) for item in resultados]
        self.lote = self._merge_lote_resultados(self.lote, resultados)
        self.event_queue.put(("preview", self.lote))
        transmitidos = sum(1 for item in resultados if item.get("estado") in {"transmitido", "ya_transmitido"})
        errores = sum(1 for item in resultados if item.get("estado") == "error")
        pendientes_no_incluidas = sum(1 for item in self.lote if item.get("estado") == "pendiente_no_incluida")
        mensaje = f"Carga finalizada: {transmitidos}/{len(resultados)} transmitidos"
        if errores:
            mensaje += f" | {errores} con error"
        if pendientes_no_incluidas:
            mensaje += f" | {pendientes_no_incluidas} pendientes no incluidas"
        if self.stop_event.is_set():
            mensaje += " | detenido por usuario"
        mensaje += "."
        self._push_status(mensaje)
        self.event_queue.put(("done_message", mensaje))

    def _merge_revalidacion_para_subida(self, lote_actual: list[dict], resultados: list[dict]) -> list[dict]:
        originales_listos = {
            self._lote_key(item): item
            for item in lote_actual
            if item.get("estado") == "listo" and str((item.get("prestacion") or {}).get("n_orden", "")).strip()
        }
        merged = self._merge_lote_resultados(lote_actual, resultados)
        estados_transitorios = {
            "sin_coincidencia_pami",
            "sin_datos_busqueda",
            "sin_documentacion_pendiente",
        }
        ajustado: list[dict] = []
        preservados = 0
        for item in merged:
            original = originales_listos.get(self._lote_key(item))
            if original and item.get("estado") in estados_transitorios:
                preservado = dict(original)
                preservado["motivo"] = (
                    "La revalidacion PAMI no confirmo un estado accionable; se mantiene la OME "
                    "detectada en bandeja para intentar cargar por numero de orden."
                )
                ajustado.append(preservado)
                preservados += 1
            else:
                ajustado.append(item)
        if preservados:
            self._push_log(
                f"Revalidacion PAMI no confirmo {preservados} OME(s); se intentara cargar por numero de orden original."
            )
        return ajustado

    def _resumen_estados(self, lote: list[dict]) -> str:
        counts: dict[str, int] = {}
        for item in lote:
            estado = str(item.get("estado") or "sin_estado")
            counts[estado] = counts.get(estado, 0) + 1
        return " | ".join(f"{estado}: {count}" for estado, count in sorted(counts.items()))

    def _items_para_revalidar_pami(self, solo_listos: bool = False) -> list[dict]:
        if solo_listos:
            return [
                item
                for item in self.lote
                if item.get("estado") == "listo"
                and not item.get("verificado_pami")
                and not item.get("resultado_carga")
                and self._tiene_datos_para_buscar_en_pami(item)
            ]
        estados_revalidables = {
            "faltante_ome",
            "sin_coincidencia",
            "para_verificar_pami",
            "fecha_no_coincide",
            "sin_coincidencia_pami",
            "no_validada",
        }
        estados_no_resueltos = {
            "faltante_ome",
            "sin_coincidencia",
            "para_verificar_pami",
            "fecha_no_coincide",
            "sin_coincidencia_pami",
            "no_validada",
        }
        items: list[dict] = []
        for item in self.lote:
            estado = item.get("estado")
            if item.get("resultado_carga"):
                continue
            if estado not in estados_revalidables:
                continue
            if item.get("verificado_pami") and estado not in estados_no_resueltos:
                continue
            if self._tiene_datos_para_buscar_en_pami(item):
                items.append(item)
        return items

    def _tiene_datos_para_buscar_en_pami(self, item: dict) -> bool:
        prestacion = item.get("prestacion") or {}
        valores = [
            prestacion.get("beneficio", ""),
            prestacion.get("n_orden", ""),
            item.get("beneficio_pdf", ""),
            item.get("dni_pdf", ""),
        ]
        return any(re.sub(r"\D+", "", str(valor or "")) for valor in valores)

    def _listos_sin_duplicar(self) -> list[dict]:
        listos: list[dict] = []
        vistos: dict[tuple[str, str], dict] = {}
        lote_actualizado: list[dict] = []
        for item in self.lote:
            item_final = dict(item)
            if item_final.get("estado") == "listo":
                clave = self._clave_intento_documentacion(item_final)
                if all(clave) and clave in vistos:
                    item_final["estado"] = "duplicado"
                    item_final["motivo"] = (
                        "Duplicado: ya hay otro informe para este paciente y codigo de practica. "
                        f"Se usara {Path(str(vistos[clave].get('archivo', ''))).name or 'el primero'}."
                    )
                else:
                    if all(clave):
                        vistos[clave] = item_final
                    listos.append(item_final)
            lote_actualizado.append(item_final)
        if lote_actualizado != self.lote:
            self.lote = lote_actualizado
            self.event_queue.put(("preview", self.lote))
        duplicados = sum(1 for item in self.lote if item.get("estado") == "duplicado")
        if duplicados:
            self._push_log(f"{duplicados} informe(s) duplicado(s) omitidos por paciente/codigo.")
        return listos

    def _clave_intento_documentacion(self, item: dict) -> tuple[str, str]:
        prestacion = item.get("prestacion") or {}
        practica = str(prestacion.get("practica", ""))
        match = re.search(r"\b(\d{6})\b", practica)
        codigo = match.group(1) if match else ""
        beneficio = re.sub(r"\D+", "", str(prestacion.get("beneficio", "") or item.get("beneficio_pdf", "")))
        dni = re.sub(r"\D+", "", str(item.get("dni_pdf", "")))
        nombre = self._normalizar_clave(str(prestacion.get("nombre", "") or item.get("pdf_paciente", "")))
        return beneficio or dni or nombre, codigo

    def _normalizar_clave(self, value: str) -> str:
        value = str(value or "").lower()
        replacements = {
            "á": "a",
            "é": "e",
            "í": "i",
            "ó": "o",
            "ú": "u",
            "ñ": "n",
        }
        for src, dst in replacements.items():
            value = value.replace(src, dst)
        return " ".join(value.split())

    def _verify_lote_en_pami(self) -> None:
        if not self.lote:
            if not self.pdf_paths:
                raise RuntimeError("Selecciona uno o mas informes para cruzar con PAMI.")
            self._push_log("Preparando informes para cruzar directo con PAMI sin XLSX...")
            self.lote = preparar_lote_documentacion_para_pami(
                self.pdf_paths,
                progress_callback=self._preview_progress_callback,
                cancel_event=self.stop_event,
            )
            self.lote = [self._marcar_fuente_informe(item) for item in self.lote]
            self.event_queue.put(("preview", self.lote))
            self.lote_verificado_en_pami = False
            verificables = [
                item
                for item in self.lote
                if item.get("estado") == "para_verificar_pami"
            ]
            if not verificables:
                mensaje = "No hay informes con DNI/beneficio y practica detectable para cruzar con PAMI."
                self._push_status(mensaje)
                self.event_queue.put(("done_message", mensaje))
                return
        self._ensure_controller()
        assert self.controller is not None
        verificables = self._items_para_revalidar_pami(solo_listos=False)
        omitidos = len(self.lote) - len(verificables)
        if not verificables:
            mensaje = "No hay informes pendientes de verificar en PAMI; se conserva el resultado actual."
            self._push_status(mensaje)
            self._push_log(mensaje)
            self.event_queue.put(("done_message", mensaje))
            return
        self._push_log(
            f"Verificando en PAMI {len(verificables)} informe(s) pendientes; {omitidos} ya tenian estado resuelto."
        )
        resultados = self.controller.verificar_lote_en_pami(
            verificables,
            progress_callback=self._pami_progress_callback,
        )
        resultados = [self._marcar_verificado_pami(item) for item in resultados]
        self.lote = self._merge_lote_resultados(self.lote, resultados)
        self.lote_verificado_en_pami = all(
            item.get("verificado_pami") or item.get("resultado_carga") or item.get("estado") not in {
                "listo",
                "faltante_ome",
                "sin_coincidencia",
                "para_verificar_pami",
                "fecha_no_coincide",
                "sin_coincidencia_pami",
            }
            for item in self.lote
        )
        self.event_queue.put(("preview", self.lote))
        transmitidos = sum(1 for item in self.lote if item.get("estado") == "ya_transmitido")
        listos = sum(1 for item in self.lote if item.get("estado") == "listo")
        falta_activar = sum(1 for item in self.lote if item.get("estado") == "falta_activar_ome")
        sin_pami = sum(1 for item in self.lote if item.get("estado") == "sin_coincidencia_pami")
        pendientes_extra = sum(1 for item in self.lote if item.get("estado") == "pendiente_no_incluida")
        mensaje = (
            f"Verificacion PAMI lista: {transmitidos} transmitidos | "
            f"{listos} pendientes para subir | {falta_activar} faltan activar"
        )
        if pendientes_extra:
            mensaje += f" | {pendientes_extra} pendientes no incluidas"
        if sin_pami:
            mensaje += f" | {sin_pami} sin coincidencia en PAMI"
        if self.stop_event.is_set():
            mensaje += " | detenido por usuario"
        mensaje += "."
        self._push_status(mensaje)
        self.event_queue.put(("done_message", mensaje))

    def _pami_progress_callback(self, current: int, total: int, item: dict) -> None:
        restantes = max(int(total or 0) - int(current or 0), 0)
        prestacion = item.get("prestacion") or {}
        nombre = (
            str(prestacion.get("nombre", "") or item.get("pdf_paciente", "") or Path(str(item.get("archivo", ""))).name)
            .strip()
        )
        nombre = nombre[:42] + "..." if len(nombre) > 45 else nombre
        message = f"Verificando en PAMI {current}/{total} - faltan {restantes}: {nombre}"
        self.event_queue.put(("action_progress", (current, total, message)))

    def _request_stop(self) -> None:
        if not self.action_running:
            return
        self.stop_requested = True
        self.stop_event.set()
        self.stop_button.configure(state="disabled")
        self._push_status("Deteniendo... se cortara al finalizar el paso actual.")
        self._push_log("Detencion solicitada por el usuario.")

    def _merge_lote_resultados(self, lote_actual: list[dict], resultados: list[dict]) -> list[dict]:
        merged = list(lote_actual)
        used: set[int] = set()
        for resultado in resultados:
            idx = self._find_merge_index(merged, resultado, used)
            if idx is None:
                merged.append(resultado)
                used.add(len(merged) - 1)
            else:
                merged[idx] = self._merge_item_resultado(merged[idx], resultado)
                used.add(idx)
        return merged

    def _merge_item_resultado(self, actual: dict, resultado: dict) -> dict:
        merged = dict(actual)
        force_keys = {"estado", "motivo", "fuente_estado"}
        for key, value in resultado.items():
            if key == "prestacion":
                continue
            if key in force_keys or self._valor_util(value) or key not in merged:
                merged[key] = value

        prestacion_actual = dict(actual.get("prestacion") or {})
        prestacion_resultado = resultado.get("prestacion") or {}
        for key, value in prestacion_resultado.items():
            if self._valor_util(value) or key not in prestacion_actual:
                prestacion_actual[key] = value
        if prestacion_actual:
            merged["prestacion"] = prestacion_actual

        for flag in ("verificado_pami", "resultado_carga"):
            if actual.get(flag) or resultado.get(flag):
                merged[flag] = True
        return merged

    def _valor_util(self, value) -> bool:
        return value not in (None, "", [], {})

    def _marcar_verificado_pami(self, item: dict) -> dict:
        return {**item, "verificado_pami": True, "fuente_estado": "pami"}

    def _marcar_resultado_carga(self, item: dict) -> dict:
        return {
            **item,
            "verificado_pami": True,
            "resultado_carga": True,
            "fuente_estado": "carga",
        }

    def _marcar_fuente_bandeja(self, item: dict) -> dict:
        return {**item, "fuente_estado": item.get("fuente_estado") or "bandeja"}

    def _marcar_fuente_informe(self, item: dict) -> dict:
        return {**item, "fuente_estado": item.get("fuente_estado") or "informe"}

    def _find_merge_index(self, lote: list[dict], resultado: dict, used: set[int]) -> int | None:
        key = self._lote_key(resultado)
        if key[1]:
            for idx, item in enumerate(lote):
                if idx not in used and self._lote_key(item) == key:
                    return idx
        archivo = str(resultado.get("archivo", ""))
        dni = str(resultado.get("dni_pdf", ""))
        beneficio = str(resultado.get("beneficio_pdf", ""))
        practica = str((resultado.get("prestacion") or {}).get("practica", ""))
        for idx, item in enumerate(lote):
            if idx in used or str(item.get("archivo", "")) != archivo:
                continue
            item_prestacion = item.get("prestacion") or {}
            if str(item_prestacion.get("n_orden", "")):
                continue
            if dni and str(item.get("dni_pdf", "")) != dni:
                continue
            if beneficio and str(item.get("beneficio_pdf", "")) and str(item.get("beneficio_pdf", "")) != beneficio:
                continue
            if (
                practica
                and str(item_prestacion.get("practica", ""))
                and not self._practicas_equivalentes(item_prestacion.get("practica", ""), practica)
            ):
                continue
            return idx
        return None

    def _practicas_equivalentes(self, izquierda: str, derecha: str) -> bool:
        izq_codigos = self._codigos_practica_texto(izquierda)
        der_codigos = self._codigos_practica_texto(derecha)
        if izq_codigos and der_codigos:
            return bool(izq_codigos & der_codigos)
        izq_key = self._normalizar_clave(str(izquierda or ""))
        der_key = self._normalizar_clave(str(derecha or ""))
        return bool(izq_key and der_key and (izq_key == der_key or izq_key in der_key or der_key in izq_key))

    def _codigos_practica_texto(self, value: str) -> set[str]:
        codigos: set[str] = set()
        directo = _codigo_practica(str(value or ""))
        if directo:
            codigos.add(directo)
        for token in _practica_keywords_pdf(str(value or "")):
            token = str(token or "")
            if re.fullmatch(r"\d{6}", token):
                codigos.add(token)
        expandidos = set(codigos)
        for codigo in codigos:
            expandidos.update(_codigos_equivalentes_documentacion(codigo))
        return expandidos

    def _clear_panel(self) -> None:
        if self.action_running:
            messagebox.showwarning("Atencion", "Espera a que termine la accion actual antes de limpiar.")
            return
        self.pdf_paths = []
        self._clear_zip_extract_dir()
        self._reset_lote_view(clear_log=True)
        self.pdf_label.configure(text="Sin informes seleccionados")
        navegador = "Se mantiene el navegador PAMI abierto para la proxima accion." if self.controller else "Navegador cerrado."
        self.status_label.configure(text=f"Panel limpio. XLSX activo: {self.ruta_excel or 'sin XLSX'}. {navegador}")

    def _reset_lote_view(self, clear_log: bool = False) -> None:
        self.lote = []
        self.lote_verificado_en_pami = False
        self._render_lote([])
        self.upload_button.configure(state="disabled")
        self.verify_pami_button.configure(state="disabled")
        self.copy_beneficio_button.configure(state="disabled")
        if clear_log:
            self.log_text.configure(state="normal")
            self.log_text.delete("1.0", "end")
            self.log_text.configure(state="disabled")

    def _ensure_controller(self) -> None:
        profile = self.worker_profile or self._current_profile_from_form()
        hide_browser = self.worker_hide_browser if self.worker_profile is not None else self.hide_browser_var.get()
        if hide_browser and (not profile["usuario"] or not profile["clave"]):
            raise RuntimeError("Para ocultar el navegador, completa usuario y clave.")
        if self.controller is None:
            self.controller = PamiDocumentacionController(
                usuario=profile["usuario"],
                clave=profile["clave"],
                log_callback=self._push_log,
                status_callback=self._push_status,
                headless=hide_browser,
            )
        self.controller.set_cancel_event(self.stop_event)
        desde, hasta = self._pami_turno_date_range()
        self.controller.set_rango_turno_fallback(desde, hasta)

    def _run_background(self, target, show_general_progress: bool = True) -> None:
        if self.action_running:
            return
        self.stop_requested = False
        self.stop_event.clear()
        self.worker_profile = self._current_profile_from_form()
        self.worker_hide_browser = self.hide_browser_var.get()
        self.action_running = True
        self.show_general_progress = show_general_progress
        self._set_controls_enabled(False)
        self.stop_button.configure(state="normal")
        if show_general_progress:
            self._start_action_progress(target)

        def runner() -> None:
            try:
                target()
                self.event_queue.put(("action_finished", None))
            except Exception as exc:
                self._push_log(traceback.format_exc())
                self.event_queue.put(("action_error", str(exc)))
            finally:
                self.event_queue.put(("worker_context_done", None))

        threading.Thread(target=runner, daemon=True).start()

    def _process_ui_queue(self) -> None:
        try:
            while True:
                event, payload = self.event_queue.get_nowait()
                if event == "log":
                    self._append_log(payload)
                elif event == "status":
                    self.status_label.configure(text=payload)
                elif event == "gmail_status":
                    self.gmail_status_var.set(payload)
                    self.gmail_status_label.configure(text_color="#51657a")
                elif event == "gmail_connected":
                    self.gmail_connected = True
                    self.gmail_status_var.set(f"Gmail conectado: {payload}")
                    self.gmail_status_label.configure(text_color="#2f9e44")
                    self._update_gmail_connection_ui()
                elif event == "gmail_file":
                    self._append_log(f"Gmail: descargado {payload}")
                elif event == "gmail_done":
                    self.gmail_progress.stop()
                    self.gmail_progress.set(0)
                    self._push_status(payload)
                    messagebox.showinfo("Gmail", payload)
                elif event == "action_progress":
                    current, total, message = payload
                    self._update_action_progress(current, total, message)
                elif event == "preview_partial":
                    self.lote = payload
                    self._render_lote(payload)
                elif event == "preview":
                    self._render_lote(payload)
                    if not self.action_running:
                        self.verify_pami_button.configure(state="normal" if (self.lote or self.pdf_paths) else "disabled")
                        self.upload_button.configure(state="normal" if any(item.get("estado") == "listo" for item in self.lote) else "disabled")
                elif event == "action_finished":
                    self.action_running = False
                    self._stop_action_progress()
                    if hasattr(self, "gmail_progress"):
                        self.gmail_progress.stop()
                        self.gmail_progress.set(0)
                    self._set_controls_enabled(True)
                    self.stop_button.configure(state="disabled")
                elif event == "done_message":
                    messagebox.showinfo("Documentacion OME", payload)
                elif event == "worker_context_done":
                    self.worker_profile = None
                    self.worker_hide_browser = False
                    self.show_general_progress = False
                elif event == "action_error":
                    self.action_running = False
                    self._stop_action_progress()
                    if hasattr(self, "gmail_progress"):
                        self.gmail_progress.stop()
                        self.gmail_progress.set(0)
                    self._set_controls_enabled(True)
                    self.stop_button.configure(state="disabled")
                    self._push_log(f"ERROR: {payload}")
                    messagebox.showerror("Documentacion OME", payload)
        except queue.Empty:
            pass
        self.after(150 if self.action_running else 350, self._process_ui_queue)

    def _render_lote(self, lote: list[dict]) -> None:
        self._resultado_por_item.clear()
        for item_id in self.results_table.get_children():
            self.results_table.delete(item_id)
        for item in lote:
            prestacion = item.get("prestacion") or {}
            estado = item.get("estado", "")
            motivo = item.get("motivo", "")
            estado_texto = self._estado_display(estado, motivo)
            item_id = self.results_table.insert(
                "",
                "end",
                values=(
                    self._short_filename(Path(str(item.get("archivo", "")).strip()).name),
                    self._paciente_detectado_display(item),
                    item.get("dni_pdf", ""),
                    item.get("fecha_informe", ""),
                    prestacion.get("n_orden", ""),
                    prestacion.get("nombre", ""),
                    prestacion.get("beneficio", ""),
                    prestacion.get("practica", ""),
                    estado_texto,
                ),
            )
            self._resultado_por_item[item_id] = item

    def _short_filename(self, filename: str, max_len: int = 34) -> str:
        if len(filename) <= max_len:
            return filename
        path = Path(filename)
        suffix = path.suffix
        stem = path.stem
        keep = max(max_len - len(suffix) - 3, 12)
        return f"{stem[:keep]}...{suffix}"

    def _resultado_seleccionado(self) -> dict | None:
        selected = self.results_table.selection()
        if not selected:
            return None
        item_id = selected[0]
        resultado = self._resultado_por_item.get(item_id)
        if resultado is not None:
            return resultado
        try:
            index = list(self.results_table.get_children()).index(item_id)
        except ValueError:
            return None
        if 0 <= index < len(self.lote):
            return self.lote[index]
        return None

    def _ruta_informe_de_resultado(self, resultado: dict | None) -> Path | None:
        if not isinstance(resultado, dict):
            return None
        ruta_archivo = str(resultado.get("archivo", "") or "").strip()
        if ruta_archivo:
            path = Path(ruta_archivo)
            if path.exists():
                return path

        nombre = Path(ruta_archivo).name if ruta_archivo else ""
        if not nombre:
            selected = self.results_table.selection()
            values = self.results_table.item(selected[0], "values") if selected else ()
            nombre = str(values[0] if values else "").strip()
        if not nombre:
            return None

        nombre_path = Path(nombre)
        for candidate in self.pdf_paths:
            path = Path(candidate)
            if path.exists() and (
                path.name == nombre
                or path.stem == nombre_path.stem
                or self._short_filename(path.name) == nombre
            ):
                return path
        return None

    def _open_selected_informe(self) -> None:
        resultado = self._resultado_seleccionado()
        if resultado is None:
            messagebox.showinfo("Informe", "Selecciona una fila para abrir el informe.")
            return
        path = self._ruta_informe_de_resultado(resultado)
        if path is None:
            messagebox.showwarning("Informe", "No encontre el archivo original del informe. Puede haberse movido o borrado.")
            return
        try:
            os.startfile(str(path))
        except OSError as exc:
            messagebox.showerror("Informe", f"No se pudo abrir el informe:\n{exc}")
            return
        self._push_status(f"Informe abierto: {path.name}")

    def _on_results_table_double_click(self, event) -> None:
        item_id = self.results_table.identify_row(event.y)
        if not item_id:
            return
        self.results_table.selection_set(item_id)
        self.results_table.focus(item_id)
        self._open_selected_informe()

    def _copy_selected_beneficio(self) -> None:
        selected = self.results_table.selection()
        if not selected:
            self._push_status("Selecciona una fila para copiar el beneficio.")
            return
        values = self.results_table.item(selected[0], "values")
        beneficio_idx = getattr(self, "results_columns", ()).index("beneficio")
        beneficio = str(values[beneficio_idx] if len(values) > beneficio_idx else "").strip()
        if not beneficio:
            self._push_status("La fila seleccionada no tiene beneficio para copiar.")
            return
        self.clipboard_clear()
        self.clipboard_append(beneficio)
        self._push_status(f"Beneficio copiado: {beneficio}")

    def _export_reporte(self) -> None:
        if not self.lote:
            messagebox.showwarning("Reporte", "No hay filas para exportar. Primero previsualiza o cruza con PAMI.")
            return
        default_name = f"reporte_documentacion_omes_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        destino = filedialog.asksaveasfilename(
            title="Guardar reporte de documentacion",
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel", "*.xlsx")],
        )
        if not destino:
            return
        try:
            self._guardar_reporte_xlsx(Path(destino))
            self._push_status(f"Reporte guardado: {destino}")
            if messagebox.askyesno("Reporte", f"Reporte guardado:\n{destino}\n\n¿Abrirlo ahora?"):
                try:
                    os.startfile(destino)  # type: ignore[attr-defined]
                except Exception as open_exc:
                    messagebox.showerror("Reporte", f"No se pudo abrir el reporte:\n{open_exc}")
        except Exception as exc:
            messagebox.showerror("Reporte", f"No se pudo guardar el reporte:\n{exc}")

    def _guardar_reporte_xlsx(self, destino: Path) -> None:
        headers = [
            "Archivo",
            "Informe",
            "Fecha informe",
            "Paciente detectado",
            "DNI",
            "OME",
            "Paciente XLS/PAMI",
            "Beneficio",
            "Practica",
            "Estado",
            "Motivo",
            "Fecha turno",
        ]
        wb = Workbook()
        ws = wb.active
        ws.title = "Documentacion OMEs"
        ws.append(headers)
        header_fill = PatternFill("solid", fgColor="2F6FA3")
        for cell in ws[1]:
            cell.font = Font(color="FFFFFF", bold=True)
            cell.fill = header_fill
        for item in self.lote:
            prestacion = item.get("prestacion") or {}
            ruta_archivo = str(item.get("archivo", "") or "").strip()
            archivo = Path(ruta_archivo) if ruta_archivo else None
            estado = item.get("estado", "")
            motivo = item.get("motivo", "")
            ws.append(
                [
                    self._xlsx_safe(archivo.name if archivo else ""),
                    self._xlsx_safe(self._reporte_informe_label(item, archivo)),
                    self._xlsx_safe(self._reporte_fecha_informe(item, archivo)),
                    self._xlsx_safe(self._paciente_detectado_display(item)),
                    self._xlsx_safe(item.get("dni_pdf", "")),
                    self._xlsx_safe(prestacion.get("n_orden", "")),
                    self._xlsx_safe(prestacion.get("nombre", "")),
                    self._xlsx_safe(prestacion.get("beneficio", "")),
                    self._xlsx_safe(prestacion.get("practica", "")),
                    self._xlsx_safe(self._estado_display(estado, motivo)),
                    self._xlsx_safe(motivo),
                    self._xlsx_safe(prestacion.get("turno", "")),
                ]
            )
            if archivo:
                uri_archivo = self._file_uri(archivo)
                if uri_archivo:
                    ruta_cell = ws.cell(row=ws.max_row, column=2)
                    ruta_cell.hyperlink = uri_archivo
                    ruta_cell.style = "Hyperlink"
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        widths = [34, 24, 16, 28, 14, 16, 32, 18, 72, 58, 90, 22]
        for idx, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = width
            destino.parent.mkdir(parents=True, exist_ok=True)
        wb.save(destino)

    def _paciente_detectado_display(self, item: dict) -> str:
        raw = str(item.get("pdf_paciente", "") or item.get("paciente_detectado", "") or "")
        limpio = _normalizar_nombre_paciente_detectado_final(
            _limpiar_nombre_paciente_detectado(raw)
        )
        return limpio or self._xlsx_safe(raw)

    def _reporte_informe_label(self, item: dict, archivo: Path | None) -> str:
        prestacion = item.get("prestacion") or {}
        nombre = (
            self._paciente_detectado_display(item)
            or str(prestacion.get("nombre", "") or "")
            or (archivo.stem if archivo else "")
        )
        nombre = _normalizar_nombre_paciente_detectado_final(
            _limpiar_nombre_paciente_detectado(self._xlsx_safe(nombre))
        ) or self._xlsx_safe(nombre)
        if not nombre:
            return "INFORME"
        if "," in nombre:
            apellido = nombre.split(",", 1)[0].strip()
        else:
            apellido = nombre.split()[0].strip()
        return f"INFORME - {apellido.upper()}" if apellido else "INFORME"

    def _reporte_fecha_informe(self, item: dict, archivo: Path | None) -> str:
        fecha = self._xlsx_safe(item.get("fecha_informe", ""))
        if fecha:
            return fecha
        motivo = self._xlsx_safe(item.get("motivo", ""))
        match = re.search(r"fecha del informe\s+(\d{1,2}/\d{1,2}/\d{2,4})", motivo, flags=re.I)
        if match:
            return self._normalizar_fecha_reporte(match.group(1))
        if archivo:
            return self._extraer_fecha_archivo_reporte(archivo.name)
        return ""

    def _normalizar_fecha_reporte(self, value: str) -> str:
        match = re.search(r"\b(\d{1,2})/(\d{1,2})/(\d{2,4})\b", str(value or ""))
        if not match:
            return ""
        anio = int(match.group(3))
        if anio < 100:
            anio += 2000
        try:
            fecha = datetime(anio, int(match.group(2)), int(match.group(1)))
        except ValueError:
            return ""
        return fecha.strftime("%d/%m/%Y")

    def _extraer_fecha_archivo_reporte(self, value: str) -> str:
        return _extraer_fecha_archivo(value)

    def _xlsx_safe(self, value) -> str:
        text = str(value or "")
        return re.sub(r"[\x00-\x08\x0b-\x0c\x0e-\x1f]", " ", text).strip()

    def _file_uri(self, path: Path) -> str:
        try:
            return path.expanduser().resolve(strict=False).as_uri()
        except Exception:
            return ""

    def _estado_display(self, estado: str, motivo: str = "") -> str:
        estado = str(estado or "").strip()
        motivo = str(motivo or "").strip()
        mapping = {
            "listo": "Listo para subir",
            "cargado": "Cargado",
            "transmitido": "Transmitido",
            "ya_transmitido": "Ya transmitido",
            "faltante_ome": "Falta OME",
            "no_validada": "No validada",
            "disponible_activar": "Falta Activar OME",
            "falta_activar_ome": "Falta Activar OME",
            "pendiente_no_incluida": "Pendiente no incluida",
            "para_verificar_pami": "Para verificar en PAMI",
            "sin_coincidencia": "Sin paciente en XLS",
            "sin_coincidencia_pami": "Sin paciente en PAMI",
            "sin_datos_busqueda": "Sin datos de busqueda",
            "sin_practica_informe": "Sin practica detectada",
            "sin_documentacion_pendiente": "Sin doc. pendiente",
            "fecha_no_coincide": "Fecha no coincide",
            "duplicado": "Duplicado",
            "imagen_sin_ocr": "Imagen sin OCR",
            "archivo_blanco": "Archivo en blanco",
            "informe_viejo": "Informe viejo",
            "obra_social_no_pami": "No corresponde a PAMI",
            "error": "Error",
        }
        base = mapping.get(estado, estado)
        if not motivo:
            return base
        if estado == "obra_social_no_pami":
            return motivo or base
        motivo_lower = motivo.lower()
        if "la ome ya figura transmitida" in motivo_lower:
            return "Ya transmitido en PAMI"
        if "documentacion cargada y ome transmitida" in motivo_lower:
            return "Transmitido: documentacion cargada y OME transmitida"
        if "se transmitira sin volver a subir" in motivo_lower:
            return "Listo: documentacion ya cargada, falta transmitir"
        if "ome encontrada en panel de aceptacion" in motivo_lower:
            return "Falta Activar OME: encontrada en Panel de Aceptacion"
        if "otra ecografia pendiente" in motivo_lower:
            return "Pendiente no incluida: otra ecografia pendiente en PAMI"
        if "no se encontro el paciente/practica en pami" in motivo_lower:
            return "Sin paciente en PAMI: no se encontro paciente/practica"
        if "se encontro el paciente en pami, pero no la practica esperada" in motivo_lower:
            return "Sin practica en PAMI: paciente encontrado, practica no coincide"
        if "ya figura cargada" in motivo.lower():
            return "Ya cargada en PAMI"
        if len(motivo) > 115:
            motivo = motivo[:112].rstrip() + "..."
        return f"{base}: {motivo}" if base else motivo

    def _lote_key(self, item: dict) -> tuple[str, str]:
        prestacion = item.get("prestacion") or {}
        return (str(item.get("archivo", "")), str(prestacion.get("n_orden", "")))

    def _start_action_progress(self, target) -> None:
        if not hasattr(self, "action_progress"):
            return
        labels = {
            "_analyze_reports_only": "Analizando informes...",
            "_build_preview": "Previsualizando con XLSX...",
            "_build_preview_cima": "Previsualizando con Excel CIMA...",
            "_verify_lote_en_pami": "Verificando en PAMI...",
            "_upload_lote": "Subiendo docs pendientes...",
        }
        label = labels.get(getattr(target, "__name__", ""), "Procesando...")
        self.action_progress_label.configure(text=label)
        if getattr(target, "__name__", "") in {"_analyze_reports_only", "_build_preview", "_build_preview_cima"}:
            self.action_progress.configure(mode="determinate")
            self.action_progress.set(0)
        else:
            self.action_progress.configure(mode="indeterminate")
            self.action_progress.start()

    def _update_action_progress(self, current: int, total: int, message: str) -> None:
        if not hasattr(self, "action_progress"):
            return
        total = max(int(total or 0), 1)
        current = max(0, min(int(current or 0), total))
        self.action_progress.configure(mode="determinate")
        self.action_progress.set(current / total)
        self.action_progress_label.configure(text=message)
        self.status_label.configure(text=message)

    def _stop_action_progress(self) -> None:
        if not hasattr(self, "action_progress"):
            return
        self.action_progress.stop()
        self.action_progress.set(0)
        self.action_progress_label.configure(text="")

    def _set_controls_enabled(self, enabled: bool) -> None:
        state = "normal" if enabled else "disabled"
        for widget in (
            self.select_excel_button,
            self.clear_excel_button,
            self.select_pdfs_button,
            self.analyze_button,
            self.preview_button,
            self.preview_cima_button,
            self.clear_button,
            self.verify_pami_button,
            self.close_button,
            self.restart_button,
            self.gmail_desde_entry,
            self.gmail_desde_calendar_button,
            self.gmail_hasta_entry,
            self.gmail_hasta_calendar_button,
            self.gmail_destino_entry,
            self.gmail_browse_button,
            self.gmail_connect_button,
            self.profile_combo,
            self.client_entry,
            self.user_entry,
            self.password_entry,
        ):
            widget.configure(state=state)
        self.upload_button.configure(
            state="normal" if enabled and any(item.get("estado") == "listo" for item in self.lote) else "disabled"
        )
        self.verify_pami_button.configure(state="normal" if enabled and (self.lote or self.pdf_paths) else "disabled")
        self.gmail_download_button.configure(state="normal" if enabled and self.gmail_connected else "disabled")
        self._update_gmail_connection_ui()
        self._apply_cliente_mode()
        if hasattr(self, "stop_button"):
            self.stop_button.configure(state="normal" if self.action_running else "disabled")

    def _push_log(self, message: str) -> None:
        self.event_queue.put(("log", message))

    def _push_status(self, message: str) -> None:
        self.event_queue.put(("status", message))

    def _append_log(self, message: str) -> None:
        self.log_text.configure(state="normal")
        self.log_text.insert("end", str(message) + "\n")
        self.log_text.see("end")
        self.log_text.configure(state="disabled")

    def _load_saved_profiles(self) -> list[dict]:
        try:
            if self.profiles_file.exists():
                data = json.loads(self.profiles_file.read_text(encoding="utf-8"))
                return sync_profile_records([item for item in data.get("usuarios", []) if isinstance(item, dict)])
        except Exception:
            pass
        return []

    def _load_config(self) -> dict:
        try:
            if self.config_file.exists():
                data = json.loads(self.config_file.read_text(encoding="utf-8"))
                return data if isinstance(data, dict) else {}
        except Exception:
            pass
        return {}

    def _save_config(self, updates: dict) -> None:
        try:
            self.config.update(updates)
            self.config_file.parent.mkdir(parents=True, exist_ok=True)
            self.config_file.write_text(json.dumps(self.config, ensure_ascii=False, indent=2), encoding="utf-8")
        except Exception as exc:
            self._push_log(f"No se pudo guardar configuracion de documentacion: {exc}")

    def _load_saved_excel_path(self) -> None:
        ruta = Path(str(self.config.get("ruta_excel") or ""))
        if not str(ruta).strip() or str(ruta) == ".":
            return
        self.ruta_excel = ruta
        if ruta.exists():
            self.excel_label.configure(text=str(ruta))
        else:
            self.excel_label.configure(text=f"XLSX guardado no encontrado: {ruta}")

    def _save_saved_profiles(self) -> None:
        self.profiles_file.write_text(json.dumps({"usuarios": self.saved_profiles[:20]}, ensure_ascii=False, indent=2), encoding="utf-8")
        upsert_shared_credentials_from_records(self.saved_profiles[:20])

    def _profile_options(self) -> list[str]:
        return [self._format_profile(item) for item in self.saved_profiles]

    def _format_profile(self, item: dict) -> str:
        usuario = str(item.get("usuario", "")).strip()
        nombre = str(item.get("nombre", item.get("cliente", ""))).strip()
        return f"{usuario} - {nombre}" if usuario and nombre else (usuario or nombre)

    def _find_profile_by_usuario(self, usuario: str) -> dict | None:
        usuario = str(usuario or "").split(" - ", 1)[0].strip().lower()
        if not usuario:
            return None
        for profile in self.saved_profiles:
            if str(profile.get("usuario", "")).strip().lower() == usuario:
                return profile
        return None

    def _current_profile_from_form(self) -> dict:
        return {
            "nombre": self.profile_name_var.get().strip(),
            "usuario": self.profile_user_var.get().strip(),
            "clave": self.profile_password_var.get(),
        }

    def _on_profile_selected(self, selected: str) -> None:
        profile = self._find_profile_by_usuario(selected)
        if profile:
            self.profile_name_var.set(profile.get("nombre", profile.get("cliente", "")))
            self.profile_user_var.set(profile.get("usuario", ""))
            self.profile_password_var.set(profile.get("clave", ""))
            self._save_config({"perfil_usuario": str(profile.get("usuario", "")).strip()})
            return

    def _load_initial_profile_into_form(self) -> None:
        if self.saved_profiles:
            saved_usuario = str(self.config.get("perfil_usuario") or "").strip()
            profile = self._find_profile_by_usuario(saved_usuario) or self.saved_profiles[0]
            selected = self._format_profile(profile)
            self.profile_var.set(selected)
            self._on_profile_selected(selected)
        else:
            self._new_profile()

    def _save_current_profile(self) -> None:
        profile = self._current_profile_from_form()
        if not profile["usuario"]:
            messagebox.showwarning("Perfil", "Ingresa usuario PAMI antes de guardar.")
            return
        self.saved_profiles = [p for p in self.saved_profiles if str(p.get("usuario", "")).lower() != profile["usuario"].lower()]
        self.saved_profiles.insert(0, profile)
        self._save_saved_profiles()
        self.profile_combo.configure(values=self._profile_options() or [""])
        self.profile_var.set(self._format_profile(profile))
        self._save_config({"perfil_usuario": profile["usuario"]})
        self._push_log(f"Perfil guardado: {profile['usuario']}")

    def _new_profile(self) -> None:
        self.profile_var.set("")
        self.profile_name_var.set("")
        self.profile_user_var.set("")
        self.profile_password_var.set("")

    def _toggle_password(self) -> None:
        self.password_visible = not self.password_visible
        self.password_entry.configure(show="" if self.password_visible else "*")

    def _close_browser(self) -> None:
        if self.controller:
            self.controller.cerrar()
            self.controller = None

    def _restart_app(self) -> None:
        if self.action_running:
            messagebox.showwarning("Atencion", "Espera a que termine la accion actual antes de reiniciar.")
            return
        if not messagebox.askyesno("Reiniciar app", "Se cerrara y abrira nuevamente la app. Continuar?"):
            return
        self._push_log("Reiniciando app...")
        try:
            self._close_browser()
            self._clear_zip_extract_dir()
            preserved_args = [arg for arg in sys.argv[1:] if not arg.startswith("--module=")]
            if getattr(sys, "frozen", False):
                cmd = [sys.executable, *preserved_args, "--module=documentacion"]
            else:
                script_path = Path(__file__).with_name("app.py").resolve()
                cmd = [self._python_gui_executable(), str(script_path), *preserved_args, "--module=documentacion"]
            if sys.platform.startswith("win"):
                subprocess.Popen(
                    cmd,
                    cwd=str(Path(__file__).resolve().parent),
                    creationflags=subprocess.DETACHED_PROCESS | subprocess.CREATE_NEW_PROCESS_GROUP,
                    close_fds=True,
                )
            else:
                subprocess.Popen(cmd, cwd=str(Path(__file__).resolve().parent), start_new_session=True)
            self.after(500, self.winfo_toplevel().destroy)
        except Exception as exc:
            messagebox.showerror("Reiniciar app", f"No se pudo reiniciar la app:\n{exc}")

    def _python_gui_executable(self) -> str:
        executable = Path(sys.executable)
        if sys.platform.startswith("win") and executable.name.lower() == "python.exe":
            pythonw = executable.with_name("pythonw.exe")
            if pythonw.exists():
                return str(pythonw)
        return str(executable)

    def _go_home(self) -> None:
        if self.action_running:
            messagebox.showwarning("Atencion", "Espera a que termine la accion actual antes de volver.")
            return
        if self.on_back:
            self.on_back()

    def on_close(self) -> None:
        try:
            self._close_browser()
            self._clear_zip_extract_dir()
        finally:
            self.destroy()
