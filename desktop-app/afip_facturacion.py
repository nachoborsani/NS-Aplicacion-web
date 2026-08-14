import re
import threading
import time
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Callable, Optional

from openpyxl import Workbook
from playwright.sync_api import Browser, BrowserContext, Page, Playwright, TimeoutError, sync_playwright

from app_logging import log_message
from app_paths import get_output_dir
from pami_scraper import configurar_playwright


AFIP_LOGIN_URL = "https://auth.afip.gob.ar/contribuyente_/login.xhtml"
AFIP_PORTAL_URL = "https://auth.afip.gob.ar/contribuyente_/login.xhtml"
RCEL_BASE_URL = "https://fe.afip.gob.ar/rcel/jsp/"
RCEL_MENU_URL = "https://fe.afip.gob.ar/rcel/jsp/menu_ppal.jsp"
PAMI_EFECTORES_LOGIN_URL = "https://efectores.pami.org.ar/pami_efectores/login.php?xgap_historial=clear"
PAMI_FACTURACION_URL_HINT = "hb_prestadores.php"


@dataclass
class SesionAfip:
    playwright: Playwright
    browser: Browser
    context: BrowserContext
    page: Page


@dataclass
class SesionPamiLiquidacion:
    playwright: Playwright
    browser: Browser
    context: BrowserContext
    page: Page


def normalizar_importe_afip(value: str) -> str:
    raw = str(value or "").strip()
    if not raw:
        raise ValueError("Ingresa un importe.")

    cleaned = raw.replace("$", "").replace(" ", "")
    if "," in cleaned:
        cleaned = cleaned.replace(".", "").replace(",", ".")
    else:
        cleaned = cleaned.replace(".", "")

    try:
        number = float(cleaned)
    except ValueError as exc:
        raise ValueError(f"Importe no valido: {value}") from exc

    if number <= 0:
        raise ValueError("El importe debe ser mayor a cero.")

    if number.is_integer():
        return str(int(number))
    return f"{number:.2f}".rstrip("0").rstrip(".")


class AfipFacturacionController:
    def __init__(
        self,
        log_callback: Optional[Callable[[str], None]] = None,
        status_callback: Optional[Callable[[str], None]] = None,
    ) -> None:
        self.log_callback = log_callback or (lambda _: None)
        self.status_callback = status_callback or (lambda _: None)
        self._lock = threading.RLock()
        self._sesion: SesionAfip | None = None

    def abrir_afip(self, cuit: str | None = None, clave: str | None = None) -> None:
        with self._lock:
            self._cleanup_dead_session()
            if self._sesion is not None:
                self._dispose_session("Sesion anterior cerrada.", "Sesion AFIP anterior descartada.")

            configurar_playwright()
            playwright = sync_playwright().start()
            browser = playwright.chromium.launch(headless=False, args=["--window-size=1280,900"])
            context = browser.new_context(
                accept_downloads=True,
                ignore_https_errors=True,
                viewport={"width": 1280, "height": 900},
            )
            context.on("page", self._on_context_page)
            page = context.new_page()
            page.set_default_timeout(20000)
            page.on("console", self._on_console)
            page.on("pageerror", self._on_page_error)
            page.on("download", self._on_download)
            page.goto(AFIP_PORTAL_URL, wait_until="domcontentloaded")
            self._sesion = SesionAfip(playwright=playwright, browser=browser, context=context, page=page)
            self._status("AFIP abierto. Inicia sesion o usa autocompletar.")
            self._log(f"Navegador abierto en {AFIP_PORTAL_URL}")

            if cuit or clave:
                self.autocompletar_login(cuit or "", clave or "")

    def autocompletar_login(self, cuit: str, clave: str) -> None:
        with self._lock:
            page = self._get_page()
            cuit = re.sub(r"\D", "", cuit or "")
            clave = clave or ""
            if not cuit:
                raise RuntimeError("Ingresa CUIT/CUIL para autocompletar.")

            if "login.xhtml" not in (page.url or ""):
                page.goto(AFIP_LOGIN_URL, wait_until="domcontentloaded")

            self._log("Completando CUIT/CUIL en AFIP...")
            self._fill_first_visible(
                page,
                [
                    "input[name='F1:username']",
                    "input[id*='username']",
                    "input[name*='username']",
                    "input[type='number']",
                    "input[type='text']",
                ],
                cuit,
            )
            self._click_by_text(page, ["Siguiente", "Ingresar"], exact=False)
            page.wait_for_timeout(1200)

            if clave:
                self._log("Completando clave fiscal...")
                self._fill_first_visible(
                    page,
                    [
                        "input[name='F1:password']",
                        "input[id*='password']",
                        "input[type='password']",
                    ],
                    clave,
                )
                self._click_by_text(page, ["Ingresar"], exact=False)
                page.wait_for_load_state("domcontentloaded")
                page.wait_for_timeout(2500)

            self._status("Login autocompletado. Si AFIP pide validacion adicional, completala manualmente.")

    def abrir_comprobantes_en_linea(self, representado: str = "") -> None:
        with self._lock:
            page = self._get_best_page()
            self._log("Buscando servicio Comprobantes en linea...")
            nueva = self._click_service_and_capture_page(page, "Comprobantes en linea")
            if nueva is not None:
                page = nueva
                self._sesion.page = page
            page.wait_for_timeout(1800)

            if representado:
                self._seleccionar_representado(page, representado)

            self._status("Comprobantes en linea abierto.")

    def preparar_factura(self, profile: dict, invoice: dict) -> None:
        with self._lock:
            page = self._get_best_page()
            self._ensure_rcel_menu(page, profile)
            self._click_by_text(page, ["Generar Comprobantes", "Generar Comprobante"], exact=False)
            page.wait_for_timeout(1000)

            self._seleccionar_punto_venta_y_tipo(page, profile)
            self._click_by_text(page, ["Continuar"], exact=False)
            page.wait_for_timeout(1000)
            self._raise_if_afip_error(page)

            self._completar_periodo(page, profile, invoice)
            self._click_by_text(page, ["Continuar"], exact=False)
            page.wait_for_timeout(1000)
            self._raise_if_afip_error(page)

            self._completar_receptor(page, profile)
            self._click_by_text(page, ["Continuar"], exact=False)
            page.wait_for_timeout(1000)
            self._raise_if_afip_error(page)

            self._completar_detalle(page, invoice)
            self._click_by_text(page, ["Continuar"], exact=False)
            page.wait_for_timeout(1200)
            self._raise_if_afip_error(page)

            self._validar_resumen(page, profile, invoice)
            self._status("Factura preparada en resumen. Revisa la pantalla antes de confirmar.")

    def confirmar_y_descargar(self, profile: dict, invoice: dict) -> dict:
        with self._lock:
            page = self._get_best_page()
            self._validar_resumen(page, profile, invoice)
            output_dir = self._build_invoice_dir(profile, invoice)

            self._click_by_text(page, ["Confirmar Datos"], exact=False)
            page.wait_for_timeout(700)
            self._click_by_text(page, ["Confirmar"], exact=True)
            page.wait_for_timeout(2500)
            self._wait_for_generated(page)

            downloads = []
            for label, filename in (
                ("Imprimir", "factura.pdf"),
                ("Duplicados Electronicos", "duplicados_electronicos.zip"),
                ("Ventas", "ventas.zip"),
            ):
                saved = self._try_download_link(page, label, output_dir / filename)
                if saved:
                    downloads.append(str(saved))

            cae = self._extraer_cae_pagina(page)
            self._status(f"Comprobante generado. Archivos guardados: {len(downloads)}")
            self._log(f"Descargas AFIP guardadas en {output_dir}")
            return {"output_dir": str(output_dir), "downloads": downloads, "cae": cae}

    def cerrar_navegador(self) -> None:
        with self._lock:
            if self._sesion is None:
                self._status("No hay navegador AFIP abierto.")
                return
            self._dispose_session("Navegador AFIP cerrado.", "Sesion AFIP cerrada manualmente.")

    def cerrar(self) -> None:
        with self._lock:
            if self._sesion is None:
                return
            self._dispose_session("Navegador AFIP cerrado.", "Sesion AFIP cerrada.")

    def _ensure_rcel_menu(self, page: Page, profile: dict) -> None:
        url = page.url or ""
        if "menu_ppal.jsp" in url:
            return
        if "fe.afip.gob.ar/rcel" not in url:
            self.abrir_comprobantes_en_linea(profile.get("representado", ""))
            page = self._get_best_page()
        if "index_bis.jsp" in (page.url or "") and profile.get("representado"):
            self._seleccionar_representado(page, profile.get("representado", ""))
        if "menu_ppal.jsp" not in (page.url or ""):
            try:
                page.goto(RCEL_MENU_URL, wait_until="domcontentloaded")
                page.wait_for_timeout(1000)
            except Exception:
                pass

    def _seleccionar_representado(self, page: Page, representado: str) -> None:
        objetivo = self._norm(representado)
        if not objetivo:
            return
        self._log(f"Seleccionando representado: {representado}")
        clicked = page.evaluate(
            """
            (objetivo) => {
              const norm = (v) => (v || '').toString().normalize('NFD').replace(/[\\u0300-\\u036f]/g, '').toLowerCase();
              const candidatos = Array.from(document.querySelectorAll('input, button, a'));
              const item = candidatos.find((el) => norm(el.value || el.textContent).includes(objetivo));
              if (!item) return false;
              item.click();
              return true;
            }
            """,
            objetivo,
        )
        if not clicked:
            self._click_by_text(page, [representado], exact=False)
        page.wait_for_load_state("domcontentloaded")
        page.wait_for_timeout(1000)

    def _seleccionar_punto_venta_y_tipo(self, page: Page, profile: dict) -> None:
        punto = str(profile.get("punto_venta", "") or "").strip()
        tipo = str(profile.get("tipo_comprobante", "Factura C") or "Factura C").strip()
        self._select_option_matching(page, ["select[name*='punto']", "select"], punto)
        self._select_option_matching(page, ["select[name*='tipo']", "select"], tipo)
        self._log(f"Punto de venta/tipo seleccionados: {punto or '-'} | {tipo}")

    def _completar_periodo(self, page: Page, profile: dict, invoice: dict) -> None:
        self._validar_fechas_factura(invoice)
        self._select_option_matching(page, ["select"], invoice.get("concepto", "Productos y Servicios"))
        page.wait_for_timeout(900)

        self._fill_date_field_by_exact_label(page, "Fecha del Comprobante", invoice.get("fecha_comprobante", ""))
        self._fill_date_field_by_exact_label(page, "Desde", invoice.get("periodo_desde", ""))
        self._fill_date_field_by_exact_label(page, "Hasta", invoice.get("periodo_hasta", ""))
        self._fill_date_field_by_exact_label(page, "Vto. para el Pago", invoice.get("vencimiento_pago", ""))

        actividad = str(profile.get("actividad", "") or "").strip()
        if actividad:
            self._select_or_fill_contains(page, actividad)
        page.wait_for_timeout(500)
        self._log(
            "Fechas cargadas -> "
            f"comprobante={invoice.get('fecha_comprobante', '')} | "
            f"desde={invoice.get('periodo_desde', '')} | "
            f"hasta={invoice.get('periodo_hasta', '')} | "
            f"vto={invoice.get('vencimiento_pago', '')}"
        )

    def _completar_receptor(self, page: Page, profile: dict) -> None:
        self._select_option_matching(page, ["select"], profile.get("receptor_iva", "IVA Sujeto Exento"))
        self._select_option_matching(page, ["select"], profile.get("receptor_tipo_doc", "CUIT"))
        self._fill_first_visible(page, ["input[name*='nro_doc']", "input[name*='doc']", "input[type='text']"], profile.get("receptor_cuit", ""))
        page.wait_for_timeout(2200)
        self._wait_for_receptor_autocomplete(page)
        self._select_condicion_venta_cuenta_corriente(page)

    def _completar_detalle(self, page: Page, invoice: dict) -> None:
        descripcion = str(invoice.get("descripcion", "") or "").strip()
        importe = normalizar_importe_afip(str(invoice.get("importe", "") or ""))
        if not descripcion:
            raise RuntimeError("Ingresa una descripcion para la factura.")

        if not self._fill_detalle_factura_c(page, descripcion, importe):
            self._fill_input_by_column(page, "Producto/Servicio", descripcion)
            self._fill_input_by_column(page, "Cant.", "1")
            self._select_option_matching(page, ["select"], "unidades")
            self._fill_input_by_column(page, "Prec. Unitario", importe)
            self._fill_input_by_column(page, "% Bon.", "0")
        page.wait_for_timeout(1200)
        self._log(f"Detalle cargado: {descripcion} | {importe}")

    def _validar_resumen(self, page: Page, profile: dict, invoice: dict) -> None:
        body = self._norm(page.locator("body").inner_text(timeout=5000))
        checks = [
            ("CUIT receptor", re.sub(r"\D", "", profile.get("receptor_cuit", ""))),
            ("descripcion", invoice.get("descripcion", "")),
        ]
        importe = normalizar_importe_afip(str(invoice.get("importe", "")))
        importe_digits = re.sub(r"\D", "", importe)
        for label, expected in checks:
            expected_norm = self._norm(expected)
            if expected_norm and expected_norm not in body:
                raise RuntimeError(f"El resumen no coincide con {label}: {expected}")
        if importe_digits and importe_digits not in re.sub(r"\D", "", body):
            raise RuntimeError("El resumen no parece contener el importe esperado.")
        self._log("Resumen AFIP validado contra los datos cargados.")

    def _try_download_link(self, page: Page, label: str, destino: Path) -> Path | None:
        destino.parent.mkdir(parents=True, exist_ok=True)
        try:
            with page.expect_download(timeout=12000) as download_info:
                self._click_by_text(page, [label], exact=False)
            download = download_info.value
            download.save_as(str(destino))
            self._log(f"Descarga guardada: {destino}")
            return destino
        except Exception as exc:
            self._log(f"No se pudo descargar '{label}' automaticamente: {exc}")
            return None

    def _wait_for_generated(self, page: Page) -> None:
        try:
            page.get_by_text("Comprobante Generado", exact=False).wait_for(timeout=20000)
        except Exception as exc:
            raise RuntimeError("No pude confirmar que AFIP haya generado el comprobante.") from exc

    def _extraer_cae_pagina(self, page: Page) -> str:
        try:
            text = page.locator("body").inner_text(timeout=3000)
        except Exception:
            return ""
        match = re.search(r"\bCAE\b\D{0,20}(\d{10,20})", text, flags=re.IGNORECASE)
        return match.group(1) if match else ""

    def _build_invoice_dir(self, profile: dict, invoice: dict) -> Path:
        cuit = re.sub(r"\D", "", profile.get("usuario", "")) or "afip"
        periodo = re.sub(r"\D", "", invoice.get("periodo_desde", ""))[-6:] or time.strftime("%Y%m%d")
        stamp = time.strftime("%Y%m%d_%H%M%S")
        return get_output_dir() / "facturas_afip" / f"{cuit}_{periodo}_{stamp}"

    def _click_service_and_capture_page(self, page: Page, label: str) -> Page | None:
        try:
            with page.expect_popup(timeout=6000) as popup_info:
                self._click_by_text(page, [label], exact=False)
            popup = popup_info.value
            popup.set_default_timeout(20000)
            popup.on("console", self._on_console)
            popup.on("pageerror", self._on_page_error)
            popup.on("download", self._on_download)
            popup.wait_for_load_state("domcontentloaded")
            return popup
        except Exception:
            self._click_by_text(page, [label], exact=False)
            return None

    def _click_by_text(self, page: Page, labels: list[str], exact: bool = False) -> None:
        for label in labels:
            try:
                page.get_by_role("button", name=re.compile(re.escape(label), re.I) if not exact else label).first.click(timeout=1200)
                return
            except Exception:
                pass
            try:
                page.get_by_text(label, exact=exact).first.click(timeout=1200)
                return
            except Exception:
                pass
        clicked = page.evaluate(
            """
            ({ labels, exact }) => {
              const norm = (v) => (v || '').toString().normalize('NFD').replace(/[\\u0300-\\u036f]/g, '').toLowerCase().trim();
              const wanted = labels.map(norm);
              const candidatos = Array.from(document.querySelectorAll('button, input[type="button"], input[type="submit"], a'));
              const item = candidatos.find((el) => {
                const txt = norm(el.value || el.textContent || el.title);
                return wanted.some((w) => exact ? txt === w : txt.includes(w));
              });
              if (!item) return false;
              item.click();
              return true;
            }
            """,
            {"labels": labels, "exact": exact},
        )
        if not clicked:
            raise RuntimeError(f"No encontre boton/enlace: {', '.join(labels)}")

    def _fill_first_visible(self, page: Page, selectors: list[str], value: str) -> None:
        for selector in selectors:
            try:
                locator = page.locator(selector).first
                if locator.count() == 0:
                    continue
                locator.wait_for(state="visible", timeout=1000)
                locator.fill("")
                locator.type(str(value), delay=40)
                return
            except Exception:
                continue
        raise RuntimeError("No encontre un campo visible para completar.")

    def _select_option_matching(self, page: Page, selectors: list[str], value: str) -> bool:
        value = str(value or "").strip()
        if not value:
            return False
        for selector in selectors:
            try:
                result = page.evaluate(
                    """
                    ({ selector, value }) => {
                      const norm = (v) => (v || '').toString().normalize('NFD').replace(/[\\u0300-\\u036f]/g, '').toLowerCase();
                      const target = norm(value);
                      for (const sel of Array.from(document.querySelectorAll(selector))) {
                        const opts = Array.from(sel.options || []);
                        const opt = opts.find((o) => norm(o.textContent).includes(target) || norm(o.value).includes(target));
                        if (!opt) continue;
                        sel.value = opt.value;
                        sel.dispatchEvent(new Event('input', { bubbles: true }));
                        sel.dispatchEvent(new Event('change', { bubbles: true }));
                        return true;
                      }
                      return false;
                    }
                    """,
                    {"selector": selector, "value": value},
                )
                if result:
                    return True
            except Exception:
                continue
        return False

    def _select_or_fill_contains(self, page: Page, value: str) -> None:
        if self._select_option_matching(page, ["select"], value):
            return
        try:
            page.evaluate(
                """
                (value) => {
                  const norm = (v) => (v || '').toString().normalize('NFD').replace(/[\\u0300-\\u036f]/g, '').toLowerCase();
                  const input = Array.from(document.querySelectorAll('input[type="text"], input:not([type])'))
                    .find((el) => norm(el.value).includes(norm(value).slice(0, 8)) || norm(el.parentElement?.textContent).includes('actividad'));
                  if (!input) return false;
                  input.value = value;
                  input.dispatchEvent(new Event('input', { bubbles: true }));
                  input.dispatchEvent(new Event('change', { bubbles: true }));
                  return true;
                }
                """,
                value,
            )
        except Exception:
            pass

    def _fill_input_near_text(self, page: Page, label: str, value: str) -> None:
        if not value:
            return
        ok = page.evaluate(
            """
            ({ label, value }) => {
              const norm = (v) => (v || '').toString().normalize('NFD').replace(/[\\u0300-\\u036f]/g, '').toLowerCase();
              const target = norm(label);
              const nodes = Array.from(document.querySelectorAll('td, th, label, span, div'));
              const node = nodes.find((el) => norm(el.textContent).includes(target));
              if (!node) return false;
              const scope = node.closest('tr') || node.parentElement || document;
              const input = scope.querySelector('input[type="text"], input:not([type])') ||
                node.parentElement?.querySelector('input[type="text"], input:not([type])');
              if (!input) return false;
              input.value = value;
              input.dispatchEvent(new Event('input', { bubbles: true }));
              input.dispatchEvent(new Event('change', { bubbles: true }));
              return true;
            }
            """,
            {"label": label, "value": value},
        )
        if not ok:
            self._log(f"No pude completar automaticamente el campo cercano a '{label}'.")

    def _fill_date_field_by_exact_label(self, page: Page, label: str, value: str) -> None:
        value = str(value or "").strip()
        if not value:
            return
        ok = page.evaluate(
            """
            ({ label, value }) => {
              const norm = (v) => (v || '').toString().normalize('NFD').replace(/[\\u0300-\\u036f]/g, '').toLowerCase().trim();
              const target = norm(label);
              const visible = (el) => {
                const style = window.getComputedStyle(el);
                const rect = el.getBoundingClientRect();
                return style.display !== 'none'
                  && style.visibility !== 'hidden'
                  && rect.width > 0
                  && rect.height > 0
                  && !el.disabled
                  && !el.readOnly;
              };
              const labels = Array.from(document.querySelectorAll('td, th, label, span, div, b'))
                .filter((el) => norm(el.textContent) === target || norm(el.textContent).startsWith(target));
              for (const labelEl of labels) {
                const row = labelEl.closest('tr');
                const scopes = [row, labelEl.parentElement, labelEl.closest('table'), document].filter(Boolean);
                for (const scope of scopes) {
                  const inputs = Array.from(scope.querySelectorAll('input[type="text"], input:not([type])')).filter(visible);
                  if (inputs.length > 0) {
                    const input = inputs[0];
                    input.focus();
                    input.value = value;
                    input.dispatchEvent(new Event('input', { bubbles: true }));
                    input.dispatchEvent(new Event('change', { bubbles: true }));
                    input.dispatchEvent(new Event('blur', { bubbles: true }));
                    return true;
                  }
                }
              }
              return false;
            }
            """,
            {"label": label, "value": value},
        )
        if not ok:
            raise RuntimeError(f"No pude completar la fecha '{label}'.")

    def _validar_fechas_factura(self, invoice: dict) -> None:
        def parse(label: str, value: str) -> datetime:
            try:
                return datetime.strptime(str(value or "").strip(), "%d/%m/%Y")
            except Exception as exc:
                raise RuntimeError(f"La fecha {label} no tiene formato DD/MM/AAAA: {value}") from exc

        desde = parse("Desde", invoice.get("periodo_desde", ""))
        hasta = parse("Hasta", invoice.get("periodo_hasta", ""))
        vto = parse("Vto. pago", invoice.get("vencimiento_pago", ""))
        parse("Fecha comprobante", invoice.get("fecha_comprobante", ""))
        if hasta < desde:
            raise RuntimeError("La fecha Hasta no puede ser anterior a Desde.")
        if vto < desde:
            raise RuntimeError("El vencimiento de pago no puede ser anterior al inicio del periodo.")

    def _fill_date_inputs_in_order(self, page: Page, values: list[str]) -> bool:
        values = [str(item or "").strip() for item in values]
        if not any(values):
            return False
        try:
            return bool(
                page.evaluate(
                    """
                    (values) => {
                      const visible = (el) => {
                        const style = window.getComputedStyle(el);
                        const rect = el.getBoundingClientRect();
                        return style.display !== 'none'
                          && style.visibility !== 'hidden'
                          && rect.width > 0
                          && rect.height > 0
                          && !el.disabled
                          && !el.readOnly;
                      };
                      const inputs = Array.from(document.querySelectorAll('input[type="text"], input:not([type])'))
                        .filter(visible)
                        .filter((el) => {
                          const value = (el.value || '').trim();
                          const placeholder = (el.placeholder || '').trim();
                          const size = Number(el.getAttribute('size') || 0);
                          const max = Number(el.getAttribute('maxlength') || 0);
                          return /\\d{1,2}\\/\\d{1,2}\\/\\d{4}/.test(value)
                            || /DD\\/MM\\/AAAA/i.test(placeholder)
                            || max === 10
                            || (size >= 8 && size <= 12);
                        });
                      if (inputs.length < 4) return false;
                      values.slice(0, 4).forEach((value, index) => {
                        if (!value) return;
                        const el = inputs[index];
                        el.focus();
                        el.value = value;
                        el.dispatchEvent(new Event('input', { bubbles: true }));
                        el.dispatchEvent(new Event('change', { bubbles: true }));
                        el.dispatchEvent(new Event('blur', { bubbles: true }));
                      });
                      return true;
                    }
                    """,
                    values,
                )
            )
        except Exception:
            return False

    def _raise_if_afip_error(self, page: Page) -> None:
        try:
            text = page.locator("body").inner_text(timeout=3000)
        except Exception:
            return
        match = re.search(r"Error:\s*(.+)", text, flags=re.IGNORECASE | re.DOTALL)
        if match:
            message = " ".join(match.group(1).split())
            raise RuntimeError(f"AFIP devolvio un error: {message}")

    def _fill_input_by_column(self, page: Page, column: str, value: str) -> None:
        ok = page.evaluate(
            """
            ({ column, value }) => {
              const norm = (v) => (v || '').toString().normalize('NFD').replace(/[\\u0300-\\u036f]/g, '').toLowerCase();
              const headers = Array.from(document.querySelectorAll('th, td'))
                .filter((el) => norm(el.textContent).includes(norm(column)));
              for (const header of headers) {
                const table = header.closest('table');
                if (!table) continue;
                const headerRow = header.closest('tr');
                const cells = Array.from(headerRow?.children || []);
                const index = cells.indexOf(header);
                if (index < 0) continue;
                const dataRow = Array.from(table.querySelectorAll('tr')).find((tr) => tr !== headerRow && tr.querySelector('input, textarea, select'));
                const targetCell = dataRow?.children?.[index];
                const input = targetCell?.querySelector('input, textarea');
                if (!input) continue;
                input.value = value;
                input.dispatchEvent(new Event('input', { bubbles: true }));
                input.dispatchEvent(new Event('change', { bubbles: true }));
                return true;
              }
              return false;
            }
            """,
            {"column": column, "value": value},
        )
        if not ok:
            raise RuntimeError(f"No pude completar la columna {column}.")

    def _fill_detalle_factura_c(self, page: Page, descripcion: str, importe: str) -> bool:
        try:
            return bool(
                page.evaluate(
                    """
                    ({ descripcion, importe }) => {
                      const visible = (el) => {
                        const style = window.getComputedStyle(el);
                        const rect = el.getBoundingClientRect();
                        return style.display !== 'none'
                          && style.visibility !== 'hidden'
                          && rect.width > 0
                          && rect.height > 0
                          && !el.disabled
                          && !el.readOnly;
                      };
                      const setValue = (el, value) => {
                        el.focus();
                        el.value = value;
                        el.dispatchEvent(new Event('input', { bubbles: true }));
                        el.dispatchEvent(new Event('change', { bubbles: true }));
                        el.dispatchEvent(new Event('blur', { bubbles: true }));
                      };

                      const area = Array.from(document.querySelectorAll('textarea')).find(visible);
                      const inputs = Array.from(document.querySelectorAll('input[type="text"], input:not([type])'))
                        .filter(visible);
                      const selects = Array.from(document.querySelectorAll('select')).filter(visible);

                      if (area) {
                        setValue(area, descripcion);
                      } else {
                        const productInput = inputs.find((el) => {
                          const rect = el.getBoundingClientRect();
                          return rect.width > 180;
                        }) || inputs[1] || inputs[0];
                        if (!productInput) return false;
                        setValue(productInput, descripcion);
                      }

                      const numericInputs = inputs.filter((el) => el.value !== descripcion);
                      const cantidadInput = numericInputs.find((el) => {
                        const rect = el.getBoundingClientRect();
                        return rect.width <= 90;
                      }) || numericInputs[0];
                      if (cantidadInput) setValue(cantidadInput, '1');

                      const unitSelect = selects.find((sel) =>
                        Array.from(sel.options || []).some((opt) => /unidad/i.test(opt.textContent || opt.value || ''))
                      );
                      if (unitSelect) {
                        const opt = Array.from(unitSelect.options || []).find((item) => /unidad/i.test(item.textContent || item.value || ''));
                        if (opt) {
                          unitSelect.value = opt.value;
                          unitSelect.dispatchEvent(new Event('input', { bubbles: true }));
                          unitSelect.dispatchEvent(new Event('change', { bubbles: true }));
                        }
                      }

                      const afterProduct = Array.from(document.querySelectorAll('input[type="text"], input:not([type])'))
                        .filter(visible)
                        .filter((el) => el.value !== descripcion);
                      const priceInput = afterProduct.find((el) => {
                        const name = `${el.name || ''} ${el.id || ''}`.toLowerCase();
                        return /precio|unit/.test(name);
                      }) || afterProduct.find((el) => {
                        const rect = el.getBoundingClientRect();
                        return rect.width >= 80 && rect.width <= 150 && el !== cantidadInput;
                      }) || afterProduct[2];
                      if (!priceInput) return false;
                      setValue(priceInput, importe);

                      const bonusInput = afterProduct.find((el) => {
                        const name = `${el.name || ''} ${el.id || ''}`.toLowerCase();
                        return /bon/.test(name);
                      }) || afterProduct.find((el) => el !== cantidadInput && el !== priceInput && (el.value || '').trim() === '');
                      if (bonusInput) setValue(bonusInput, '0');
                      return true;
                    }
                    """,
                    {"descripcion": descripcion, "importe": importe},
                )
            )
        except Exception:
            return False

    def _check_label(self, page: Page, label: str) -> None:
        checked = page.evaluate(
            """
            (label) => {
              const norm = (v) => (v || '').toString().normalize('NFD').replace(/[\\u0300-\\u036f]/g, '').toLowerCase();
              const target = norm(label);
              const labels = Array.from(document.querySelectorAll('label, td, span, div'));
              const node = labels.find((el) => norm(el.textContent).includes(target));
              if (!node) return false;
              const input = node.querySelector('input[type="checkbox"]') || node.parentElement?.querySelector('input[type="checkbox"]');
              if (!input) return false;
              input.checked = true;
              input.dispatchEvent(new Event('click', { bubbles: true }));
              input.dispatchEvent(new Event('change', { bubbles: true }));
              return true;
            }
            """,
            label,
        )
        if not checked:
            self._log(f"No pude marcar condicion de venta: {label}")

    def _wait_for_receptor_autocomplete(self, page: Page) -> None:
        try:
            page.wait_for_function(
                """
                () => {
                  const inputs = Array.from(document.querySelectorAll('input[type="text"], input:not([type])'));
                  return inputs.some((el) => /INSTITUTO NACIONAL|JUBILADOS|PAMI/i.test(el.value || ''));
                }
                """,
                timeout=6000,
            )
            self._log("Datos del receptor autocompletados por AFIP.")
        except Exception:
            self._log("No pude confirmar autocompletado del receptor; continuo con Cuenta Corriente.")

    def _select_condicion_venta_cuenta_corriente(self, page: Page) -> None:
        selected = page.evaluate(
            """
            () => {
              const norm = (v) => (v || '').toString().normalize('NFD').replace(/[\\u0300-\\u036f]/g, '').toLowerCase();
              const heading = Array.from(document.querySelectorAll('td, th, div, span, legend'))
                .find((el) => norm(el.textContent).includes('condiciones de venta'));
              const container = heading
                ? (heading.closest('table') || heading.parentElement || document)
                : document;
              const checks = Array.from(container.querySelectorAll('input[type="checkbox"]'));
              let target = null;
              for (const input of checks) {
                const text = norm(input.closest('label')?.textContent || input.parentElement?.textContent || '');
                if (text.includes('cuenta corriente')) target = input;
              }
              if (!target) {
                for (const input of Array.from(document.querySelectorAll('input[type="checkbox"]'))) {
                  const text = norm(input.closest('label')?.textContent || input.parentElement?.textContent || '');
                  if (text.includes('cuenta corriente')) target = input;
                }
              }
              if (!target) return false;
              for (const input of checks) {
                if (input.checked) {
                  input.checked = false;
                  input.dispatchEvent(new Event('change', { bubbles: true }));
                }
              }
              target.checked = true;
              target.dispatchEvent(new Event('click', { bubbles: true }));
              target.dispatchEvent(new Event('change', { bubbles: true }));
              return true;
            }
            """
        )
        if selected:
            self._log("Condicion de venta marcada: Cuenta Corriente.")
        else:
            self._log("No pude marcar Cuenta Corriente automaticamente.")

    def _get_page(self) -> Page:
        self._cleanup_dead_session()
        if self._sesion is None:
            raise RuntimeError("Primero abre AFIP desde la app.")
        return self._sesion.page

    def _get_best_page(self) -> Page:
        page = self._get_page()
        if self._sesion is None:
            return page
        for candidate in reversed(self._sesion.context.pages):
            try:
                if candidate.is_closed():
                    continue
                url = candidate.url or ""
            except Exception:
                continue
            if "fe.afip.gob.ar/rcel" in url:
                self._sesion.page = candidate
                return candidate
        for candidate in reversed(self._sesion.context.pages):
            try:
                if not candidate.is_closed() and (candidate.url or "") != "about:blank":
                    self._sesion.page = candidate
                    return candidate
            except Exception:
                continue
        return page

    def _cleanup_dead_session(self) -> None:
        if self._sesion is None:
            return
        try:
            if self._sesion.page.is_closed() or not self._sesion.context.pages:
                self._dispose_session("Sesion AFIP finalizada.", "Sesion AFIP cerrada fuera de la app.")
        except Exception:
            self._sesion = None

    def _dispose_session(self, status_message: str, log_message_text: str) -> None:
        sesion = self._sesion
        self._sesion = None
        if sesion is None:
            return
        for closer in (sesion.context.close, sesion.browser.close, sesion.playwright.stop):
            try:
                closer()
            except Exception:
                pass
        self._status(status_message)
        self._log(log_message_text)

    def _on_context_page(self, page: Page) -> None:
        page.set_default_timeout(20000)
        page.on("console", self._on_console)
        page.on("pageerror", self._on_page_error)
        page.on("download", self._on_download)
        if self._sesion is not None:
            self._sesion.page = page

    def _on_download(self, download) -> None:
        try:
            destino = get_output_dir() / "facturas_afip" / "descargas_afip"
            destino.mkdir(parents=True, exist_ok=True)
            filename = (download.suggested_filename or "descarga_afip").strip()
            final_path = destino / filename
            if final_path.exists():
                stem = final_path.stem
                suffix = final_path.suffix
                final_path = destino / f"{stem}_{time.strftime('%Y%m%d_%H%M%S')}{suffix}"
            download.save_as(str(final_path))
            self._log(f"Descarga AFIP capturada: {final_path}")
            self._status(f"Descarga guardada en: {final_path}")
        except Exception as exc:
            self._log(f"No pude guardar automaticamente una descarga AFIP: {exc}")

    def _on_console(self, msg) -> None:
        try:
            self._log(msg.text)
        except Exception:
            self._log(str(msg))

    def _on_page_error(self, err) -> None:
        self._log(f"[PAGEERROR] {err}")

    def _norm(self, value: str) -> str:
        return " ".join(str(value or "").lower().split())

    def _log(self, message: str) -> None:
        log_message(f"[AFIP] {message}")
        self.log_callback(message)

    def _status(self, message: str) -> None:
        log_message(f"[AFIP] {message}")
        self.status_callback(message)


class PamiLiquidacionController:
    def __init__(
        self,
        log_callback: Optional[Callable[[str], None]] = None,
        status_callback: Optional[Callable[[str], None]] = None,
    ) -> None:
        self.log_callback = log_callback or (lambda _: None)
        self.status_callback = status_callback or (lambda _: None)
        self._lock = threading.RLock()
        self._sesion: SesionPamiLiquidacion | None = None

    def abrir_pami(self, usuario: str | None = None, clave: str | None = None) -> None:
        with self._lock:
            self._cleanup_dead_session()
            if self._sesion is not None:
                self._dispose_session("Sesion PAMI anterior cerrada.", "Sesion PAMI liquidacion anterior descartada.")

            configurar_playwright()
            playwright = sync_playwright().start()
            browser = playwright.chromium.launch(headless=False, args=["--window-size=1360,900"])
            context = browser.new_context(
                accept_downloads=True,
                ignore_https_errors=True,
                viewport={"width": 1360, "height": 900},
            )
            context.on("page", self._on_context_page)
            page = context.new_page()
            page.set_default_timeout(20000)
            page.on("console", self._on_console)
            page.on("pageerror", self._on_page_error)
            page.on("download", self._on_download)
            page.goto(PAMI_EFECTORES_LOGIN_URL, wait_until="domcontentloaded")
            self._sesion = SesionPamiLiquidacion(playwright=playwright, browser=browser, context=context, page=page)
            self._status("PAMI Efectores abierto.")
            self._log(f"Navegador abierto en {PAMI_EFECTORES_LOGIN_URL}")
            if usuario or clave:
                self.autocompletar_login(usuario or "", clave or "")

    def autocompletar_login(self, usuario: str, clave: str) -> None:
        with self._lock:
            page = self._get_page()
            usuario = (usuario or "").strip()
            clave = clave or ""
            if not usuario:
                raise RuntimeError("Ingresa usuario PAMI para autocompletar.")
            if "login.php" not in (page.url or ""):
                page.goto(PAMI_EFECTORES_LOGIN_URL, wait_until="domcontentloaded")
            self._fill_first_visible(
                page,
                [
                    "input[name='usuario']",
                    "input[name='user']",
                    "input[name='login']",
                    "input[type='text']",
                ],
                usuario,
            )
            if clave:
                self._fill_first_visible(
                    page,
                    [
                        "input[type='password']",
                        "input[name='clave']",
                        "input[name='password']",
                    ],
                    clave,
                )
            self._click_by_text(page, ["Ingresar", "Entrar", "Aceptar"], exact=False)
            page.wait_for_load_state("domcontentloaded")
            page.wait_for_timeout(1800)
            self._status("Login PAMI enviado. Si pide validacion adicional, completala manualmente.")

    def abrir_facturacion(self) -> None:
        with self._lock:
            page = self._get_best_page()
            if PAMI_FACTURACION_URL_HINT in (page.url or ""):
                self._status("Pantalla Facturacion PAMI lista.")
                return
            self._click_by_text(page, ["Facturacion", "Facturación"], exact=False)
            page.wait_for_load_state("domcontentloaded")
            page.wait_for_timeout(1500)
            self._status("Facturacion PAMI abierta.")

    def buscar_liquidacion(self, periodo: str, concepto: str = "Todos los Conceptos") -> dict:
        with self._lock:
            page = self._get_best_page()
            if PAMI_FACTURACION_URL_HINT not in (page.url or ""):
                self.abrir_facturacion()
                page = self._get_best_page()
            self._seleccionar_periodo_y_concepto(page, periodo, concepto)
            self._click_by_text(page, ["Buscar"], exact=False)
            page.wait_for_load_state("domcontentloaded")
            page.wait_for_timeout(1800)
            total = self._leer_total_liquidado(page)
            output_path = self._exportar_tablas_pagina(page, "liquidacion", periodo)
            self._status(f"Liquidacion PAMI consultada. Total: {total or 'no detectado'}")
            return {"total": total, "excel": str(output_path), "periodo": periodo}

    def descargar_detalle_liquidacion(self, periodo: str, detalle: str = "Errores de Transmision") -> dict:
        with self._lock:
            page = self._get_best_page()
            if "Resumen" not in self._safe_text(page) and PAMI_FACTURACION_URL_HINT not in (page.url or ""):
                self.abrir_facturacion()
                page = self._get_best_page()
            self._abrir_detalle_liquidacion(page)
            page.wait_for_timeout(1200)
            self._abrir_lupa_detalle(page, periodo, detalle)
            page.wait_for_load_state("domcontentloaded")
            page.wait_for_timeout(1500)
            output_path = self._exportar_tablas_pagina(page, self._slug(detalle), periodo)
            self._status(f"Detalle exportado: {output_path}")
            return {"excel": str(output_path), "periodo": periodo, "detalle": detalle}

    def cerrar_navegador(self) -> None:
        with self._lock:
            if self._sesion is None:
                self._status("No hay navegador PAMI abierto.")
                return
            self._dispose_session("Navegador PAMI cerrado.", "Sesion PAMI liquidacion cerrada manualmente.")

    def cerrar(self) -> None:
        with self._lock:
            if self._sesion is None:
                return
            self._dispose_session("Navegador PAMI cerrado.", "Sesion PAMI liquidacion cerrada.")

    def _seleccionar_periodo_y_concepto(self, page: Page, periodo: str, concepto: str) -> None:
        periodo = (periodo or "").strip()
        concepto = (concepto or "Todos los Conceptos").strip()
        if not periodo:
            raise RuntimeError("Ingresa un periodo PAMI, por ejemplo 06/2026.")
        if not self._select_option_matching(page, "select", periodo):
            raise RuntimeError(f"No encontre el periodo {periodo} en PAMI.")
        self._set_input_near_text(page, "Ajuste negativo no aplicado", "0")
        self._select_option_matching(page, "select", concepto)
        self._log(f"Filtros PAMI cargados: periodo={periodo} | concepto={concepto}")

    def _leer_total_liquidado(self, page: Page) -> str:
        try:
            return str(
                page.evaluate(
                    """
                    () => {
                      const rows = Array.from(document.querySelectorAll('tr'));
                      const row = rows.reverse().find((tr) => /TOTAL/i.test(tr.textContent || ''));
                      if (row) {
                        const cells = Array.from(row.querySelectorAll('td, th')).map((td) => (td.textContent || '').trim()).filter(Boolean);
                        if (cells.length) return cells[cells.length - 1];
                      }
                      const body = document.body.innerText || '';
                      const matches = Array.from(body.matchAll(/\\d{1,3}(?:\\.\\d{3})*(?:,\\d{2})|\\d+\\.\\d{1,2}/g)).map((m) => m[0]);
                      return matches.length ? matches[matches.length - 1] : '';
                    }
                    """
                )
            ).strip()
        except Exception:
            return ""

    def _abrir_detalle_liquidacion(self, page: Page) -> None:
        if "RESUMEN DE LIQUID" in self._safe_text(page).upper():
            return
        self._click_by_text(page, ["Detalle Liquidacion", "Detalle Liquidación"], exact=False)
        page.wait_for_load_state("domcontentloaded")

    def _abrir_lupa_detalle(self, page: Page, periodo: str, detalle: str) -> None:
        clicked = page.evaluate(
            """
            ({ periodo, detalle }) => {
              const norm = (v) => (v || '').toString().normalize('NFD').replace(/[\\u0300-\\u036f]/g, '').toLowerCase();
              const wantedPeriodo = norm(periodo);
              const wantedDetalle = norm(detalle);
              const rows = Array.from(document.querySelectorAll('tr'));
              let currentPeriodo = '';
              for (const row of rows) {
                const text = norm(row.textContent || '');
                const periodMatch = text.match(/\\d{2}\\/\\d{4}/);
                if (periodMatch) currentPeriodo = periodMatch[0];
                if (currentPeriodo === wantedPeriodo && text.includes(wantedDetalle)) {
                  const target = row.querySelector('a, button, input[type="button"], img');
                  if (target) {
                    target.click();
                    return true;
                  }
                }
              }
              return false;
            }
            """,
            {"periodo": periodo, "detalle": detalle},
        )
        if not clicked:
            raise RuntimeError(f"No encontre la lupa para {detalle} en {periodo}.")

    def _exportar_tablas_pagina(self, page: Page, prefix: str, periodo: str) -> Path:
        tables = page.evaluate(
            """
            () => Array.from(document.querySelectorAll('table')).map((table) =>
              Array.from(table.querySelectorAll('tr')).map((tr) =>
                Array.from(tr.querySelectorAll('th, td')).map((cell) => (cell.textContent || '').replace(/\\s+/g, ' ').trim())
              ).filter((row) => row.some(Boolean))
            ).filter((table) => table.length)
            """
        )
        if not tables:
            raise RuntimeError("No encontre tablas para exportar a Excel.")
        output_dir = get_output_dir() / "facturas_afip" / "pami_liquidaciones"
        output_dir.mkdir(parents=True, exist_ok=True)
        safe_period = re.sub(r"\D", "", periodo or time.strftime("%m%Y"))
        output_path = output_dir / f"{prefix}_{safe_period}_{time.strftime('%Y%m%d_%H%M%S')}.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "PAMI"
        row_index = 1
        for table_index, table in enumerate(tables, start=1):
            sheet.cell(row=row_index, column=1, value=f"Tabla {table_index}")
            row_index += 1
            for row in table:
                for col_index, value in enumerate(row, start=1):
                    sheet.cell(row=row_index, column=col_index, value=value)
                row_index += 1
            row_index += 1
        workbook.save(output_path)
        self._log(f"Excel PAMI exportado: {output_path}")
        return output_path

    def _fill_first_visible(self, page: Page, selectors: list[str], value: str) -> None:
        for selector in selectors:
            try:
                locator = page.locator(selector).first
                if locator.count() == 0:
                    continue
                locator.wait_for(state="visible", timeout=1000)
                locator.fill("")
                locator.type(str(value), delay=30)
                return
            except Exception:
                continue
        raise RuntimeError("No encontre un campo visible para completar.")

    def _click_by_text(self, page: Page, labels: list[str], exact: bool = False) -> None:
        for label in labels:
            try:
                page.get_by_role("button", name=re.compile(re.escape(label), re.I) if not exact else label).first.click(timeout=1000)
                return
            except Exception:
                pass
            try:
                page.get_by_text(label, exact=exact).first.click(timeout=1000)
                return
            except Exception:
                pass
        clicked = page.evaluate(
            """
            ({ labels, exact }) => {
              const norm = (v) => (v || '').toString().normalize('NFD').replace(/[\\u0300-\\u036f]/g, '').toLowerCase().trim();
              const wanted = labels.map(norm);
              const candidates = Array.from(document.querySelectorAll('button, input[type="button"], input[type="submit"], a, option, li'));
              const item = candidates.find((el) => {
                const txt = norm(el.value || el.textContent || el.title);
                return wanted.some((w) => exact ? txt === w : txt.includes(w));
              });
              if (!item) return false;
              item.click();
              return true;
            }
            """,
            {"labels": labels, "exact": exact},
        )
        if not clicked:
            raise RuntimeError(f"No encontre boton/enlace: {', '.join(labels)}")

    def _select_option_matching(self, page: Page, selector: str, value: str) -> bool:
        return bool(
            page.evaluate(
                """
                ({ selector, value }) => {
                  const norm = (v) => (v || '').toString().normalize('NFD').replace(/[\\u0300-\\u036f]/g, '').toLowerCase();
                  const target = norm(value);
                  for (const sel of Array.from(document.querySelectorAll(selector))) {
                    const opts = Array.from(sel.options || []);
                    const opt = opts.find((o) => norm(o.textContent).includes(target) || norm(o.value).includes(target));
                    if (!opt) continue;
                    sel.value = opt.value;
                    sel.dispatchEvent(new Event('input', { bubbles: true }));
                    sel.dispatchEvent(new Event('change', { bubbles: true }));
                    return true;
                  }
                  return false;
                }
                """,
                {"selector": selector, "value": value},
            )
        )

    def _set_input_near_text(self, page: Page, label: str, value: str) -> bool:
        return bool(
            page.evaluate(
                """
                ({ label, value }) => {
                  const norm = (v) => (v || '').toString().normalize('NFD').replace(/[\\u0300-\\u036f]/g, '').toLowerCase();
                  const target = norm(label);
                  const nodes = Array.from(document.querySelectorAll('td, th, label, span, div'));
                  const node = nodes.find((el) => norm(el.textContent).includes(target));
                  if (!node) return false;
                  const scope = node.closest('tr') || node.parentElement || document;
                  const input = scope.querySelector('input[type="text"], input:not([type])');
                  if (!input) return false;
                  input.value = value;
                  input.dispatchEvent(new Event('input', { bubbles: true }));
                  input.dispatchEvent(new Event('change', { bubbles: true }));
                  return true;
                }
                """,
                {"label": label, "value": value},
            )
        )

    def _safe_text(self, page: Page) -> str:
        try:
            return page.locator("body").inner_text(timeout=3000)
        except Exception:
            return ""

    def _get_page(self) -> Page:
        self._cleanup_dead_session()
        if self._sesion is None:
            raise RuntimeError("Primero abre PAMI desde la app.")
        return self._sesion.page

    def _get_best_page(self) -> Page:
        page = self._get_page()
        if self._sesion is None:
            return page
        for candidate in reversed(self._sesion.context.pages):
            try:
                if candidate.is_closed():
                    continue
                url = candidate.url or ""
            except Exception:
                continue
            if "efectores.pami.org.ar" in url:
                self._sesion.page = candidate
                return candidate
        return page

    def _cleanup_dead_session(self) -> None:
        if self._sesion is None:
            return
        try:
            if self._sesion.page.is_closed() or not self._sesion.context.pages:
                self._dispose_session("Sesion PAMI finalizada.", "Sesion PAMI liquidacion cerrada fuera de la app.")
        except Exception:
            self._sesion = None

    def _dispose_session(self, status_message: str, log_message_text: str) -> None:
        sesion = self._sesion
        self._sesion = None
        if sesion is None:
            return
        for closer in (sesion.context.close, sesion.browser.close, sesion.playwright.stop):
            try:
                closer()
            except Exception:
                pass
        self._status(status_message)
        self._log(log_message_text)

    def _on_context_page(self, page: Page) -> None:
        page.set_default_timeout(20000)
        page.on("console", self._on_console)
        page.on("pageerror", self._on_page_error)
        page.on("download", self._on_download)
        if self._sesion is not None:
            self._sesion.page = page

    def _on_download(self, download) -> None:
        try:
            destino = get_output_dir() / "facturas_afip" / "pami_liquidaciones"
            destino.mkdir(parents=True, exist_ok=True)
            filename = (download.suggested_filename or "descarga_pami").strip()
            final_path = destino / filename
            if final_path.exists():
                final_path = destino / f"{final_path.stem}_{time.strftime('%Y%m%d_%H%M%S')}{final_path.suffix}"
            download.save_as(str(final_path))
            self._log(f"Descarga PAMI capturada: {final_path}")
            self._status(f"Descarga PAMI guardada en: {final_path}")
        except Exception as exc:
            self._log(f"No pude guardar automaticamente una descarga PAMI: {exc}")

    def _on_console(self, msg) -> None:
        try:
            self._log(msg.text)
        except Exception:
            self._log(str(msg))

    def _on_page_error(self, err) -> None:
        self._log(f"[PAGEERROR] {err}")

    def _slug(self, value: str) -> str:
        slug = re.sub(r"[^a-zA-Z0-9]+", "_", value or "").strip("_").lower()
        return slug or "detalle"

    def _log(self, message: str) -> None:
        log_message(f"[PAMI-LIQ] {message}")
        self.log_callback(message)

    def _status(self, message: str) -> None:
        log_message(f"[PAMI-LIQ] {message}")
        self.status_callback(message)
