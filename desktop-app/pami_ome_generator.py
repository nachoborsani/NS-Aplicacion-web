import argparse
import asyncio
import csv
import os
import random
import re
import threading
import time
import unicodedata
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Callable, Iterable

from openpyxl import Workbook, load_workbook
from playwright.async_api import Browser, BrowserContext, Error, Locator, Page, Playwright, async_playwright

from app_paths import get_errors_dir, get_output_dir, get_resource_path


LOGIN_URL = "https://cup.pami.org.ar"
GENERAR_ORDEN_URL = "https://pe.pami.org.ar/controllers/generar_orden.php"
PANEL_ACEPTACION_URL = "https://pe.pami.org.ar/controllers/efector.php"
PATIENT_TIMEOUT_SECONDS = 120
RESULT_HEADERS = ["modo", "afiliado", "beneficio", "dni", "nombre", "diagnostico", "practica", "resultado", "nro_ome"]
INPUT_HEADER_ALIASES = {
    "modo": {"modo", "tipo", "busqueda", "searchmode"},
    "afiliado": {"afiliado", "beneficiario", "benef", "beneficio", "dni"},
    "beneficio": {"beneficio", "benef", "beneficiario", "nro beneficio", "numero beneficio"},
    "dni": {"dni", "documento", "nro dni", "numero dni"},
    "nombre": {"nombre", "apellido y nombre", "nombre y apellido", "paciente", "afiliado nombre"},
    "diagnostico": {"diagnostico", "diagnostico", "cie10", "cie-10", "cie_10"},
    "practica": {"practica", "practica"},
    "completar_benef": {"completar_benef", "completar benef", "solo completar benef", "benef only"},
    "completar_dni": {"completar_dni", "completar dni", "solo completar dni", "dni only"},
}

EXTRAER_OMES_EXISTENTES_SCRIPT = r"""
async (fechaHoy) => {
  if (!window.bandeja || window.bandeja.length === 0) return [];
  const domInfo = {};
  document.querySelectorAll('table tbody tr').forEach(tr => {
    const calBtn = tr.querySelector('.fa-calendar');
    const aceptarBtn = tr.querySelector(
      'i.boton-historial[data-estado="aceptar"], .boton-historial[data-estado="aceptar"], .fa-check'
    );
    const modificarBtn = tr.querySelector('i.boton-historial[data-estado="modificar"], .boton-historial[data-estado="modificar"]');
    const cancelarBtn = tr.querySelector(
      'i.boton-historial[data-estado="cancelar"], .boton-historial[data-estado="cancelar"], ' +
      'i.boton-historial[data-estado="anular"], .boton-historial[data-estado="anular"], .fa-ban, .fa-times, .fa-remove'
    );
    const infoBtn = tr.querySelector('.fa-info, .fa-info-circle');
    const accionBtn = aceptarBtn || modificarBtn || cancelarBtn || calBtn || infoBtn;
    if (!accionBtn) return;
    const celdas = Array.from(tr.querySelectorAll('td')).map(td => (td.textContent || '').trim());
    const orden = accionBtn.getAttribute('data-orden') || celdas[0] || '';
    if (!orden) return;
    let estadoAccion = 'activada_info';
    let mismoEfector = true;
    if (aceptarBtn) {
      estadoAccion = 'disponible_activar';
      mismoEfector = true;
    } else if (calBtn && cancelarBtn) {
      estadoAccion = 'activada_modificable';
      mismoEfector = true;
    } else if (calBtn && infoBtn) {
      estadoAccion = 'activada_otro_prestador';
      mismoEfector = false;
    } else if (modificarBtn || calBtn) {
      estadoAccion = 'activada_modificable';
      mismoEfector = true;
    } else if (infoBtn) {
      estadoAccion = 'transmitida';
      mismoEfector = true;
    }
    domInfo[orden] = { mismo_efector: mismoEfector, estado_accion: estadoAccion };
  });
  const hoy = new Date(fechaHoy);
  const parseFecha = (valor) => {
    if (!valor) return null;
    const sf = String(valor).split(' ')[0];
    const p = sf.split('/');
    if (p.length !== 3) return null;
    const ft = new Date(parseInt(p[2]), parseInt(p[1]) - 1, parseInt(p[0]));
    return isNaN(ft) ? null : ft;
  };
  const omes = [];
  for (const ome of window.bandeja) {
    const domData = domInfo[ome.n_orden];
    if (domData?.estado_accion === 'disponible_activar') {
      const fv = parseFecha(ome.f_vencimiento);
      if (fv && fv < hoy) continue;
      omes.push({
        n_orden: ome.n_orden,
        n_beneficio: ome.n_beneficio,
        practica: ome.d_practica,
        f_vencimiento: ome.f_vencimiento,
        f_agenda: '',
        mismo_efector: true,
        estado_accion: domData.estado_accion || '',
        turno_asignado: false,
        ya_transmitida: false,
        transmitida: false
      });
      continue;
    }
    if (ome.id_estado_efector !== '2' || !domData) continue;
    const resp = await fetch('ajax/efectores_detalle.php', {
      method: 'POST',
      headers: { 'Content-Type': 'application/x-www-form-urlencoded' },
      body: `orden=${ome.n_orden}&estado=info&bene=${ome.n_beneficio}&gp=${ome.c_grado_paren}`
    });
    const det = await resp.json();
    const agenda = det.prescripcion?.[0]?.agenda?.[0];
    if (!agenda?.f_agenda) continue;
    const sf = agenda.f_agenda.split(' ')[0];
    const ft = parseFecha(sf);
    const fv = parseFecha(ome.f_vencimiento);
    if (!ft) continue;
    if (ft < hoy && (!fv || fv < hoy)) continue;
    omes.push({
      n_orden: ome.n_orden,
      n_beneficio: ome.n_beneficio,
      practica: ome.d_practica,
      f_vencimiento: ome.f_vencimiento,
      f_agenda: sf,
      mismo_efector: domData.mismo_efector,
      estado_accion: domData.estado_accion || '',
      turno_asignado: domData.mismo_efector && domData.estado_accion !== 'transmitida',
      ya_transmitida: domData.estado_accion === 'transmitida',
      transmitida: domData.estado_accion === 'transmitida'
    });
  }
  return omes;
}
"""

# Lista CRUDA de window.bandeja (todas las OMEs del afiliado, sin el filtro de
# "activable/agenda futura" que aplica EXTRAER). Trae el CÓDIGO de práctica
# (c_practica) y el n_orden, que es justo lo que necesitamos para recuperar el
# número en el caso YA_TIENE_OME de cabecera (la OME recién generada queda
# "PENDIENTE DE ACEPTACIÓN" y EXTRAER la descarta).
RAW_OMES_AFILIADO_SCRIPT = r"""
() => (Array.isArray(window.bandeja) ? window.bandeja : []).map(o => ({
  n_orden: (o.n_orden || '').toString(),
  c_practica: (o.c_practica || '').toString(),
  c_practica_array: Array.isArray(o.c_practica_array) ? o.c_practica_array.map(String) : [],
  estado: (o.estado || '').toString(),
  f_emision_date: Number(o.f_emision_date) || 0,
  f_vencimiento_date: Number(o.f_vencimiento_date) || 0
}))
"""


def log(message: str) -> None:
    print(message, flush=True)


def configure_playwright_browsers() -> None:
    browser_dir = get_resource_path("playwright-browsers")
    if browser_dir.exists():
        os.environ["PLAYWRIGHT_BROWSERS_PATH"] = str(browser_dir)


def load_env_file(env_path: Path) -> dict[str, str]:
    values: dict[str, str] = {}
    if not env_path.exists():
        raise FileNotFoundError(f"No se encontro el archivo .env en {env_path}")

    for raw_line in env_path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip().strip('"').strip("'")
        if key:
            values[key] = value

    return values


def sanitize_text(value: str | None) -> str:
    return " ".join((value or "").split())


def combine_beneficio_gp(beneficio: str, gp: str = "") -> str:
    beneficio_digits = re.sub(r"\D+", "", sanitize_text(beneficio))
    gp_digits = re.sub(r"\D+", "", sanitize_text(gp))
    if not beneficio_digits:
        return ""
    if gp_digits and len(beneficio_digits) < 14:
        return f"{beneficio_digits}{gp_digits.zfill(2)[-2:]}"
    return beneficio_digits


def random_delay(min_seconds: float = 0.4, max_seconds: float = 1.0) -> float:
    return random.uniform(min_seconds, max_seconds)


async def human_pause(min_seconds: float = 0.4, max_seconds: float = 1.0) -> None:
    await asyncio.sleep(random_delay(min_seconds, max_seconds))


@dataclass(slots=True)
class PatientInput:
    modo: str
    afiliado: str
    diagnostico: str
    practica: str
    beneficio: str = ""
    dni: str = ""
    nombre: str = ""
    retry_capita_done: bool = False
    completar_benef: bool = False
    completar_dni: bool = False
    # Códigos de práctica equivalentes para la búsqueda de OME existente. En
    # cabecera las 4 consultas (427109/120/121/122) son "la misma OME": si PAMI
    # dice "ya tiene OME", el número puede estar bajo cualquiera. Cuando esto
    # viene cargado, la búsqueda del existente NO filtra por práctica y acepta
    # cualquier código de esta lista. Vacío = comportamiento viejo (1 práctica).
    practicas_equivalentes: tuple[str, ...] = ()


@dataclass(slots=True)
class PatientResult:
    modo: str
    afiliado: str
    beneficio: str
    dni: str
    nombre: str
    diagnostico: str
    practica: str
    resultado: str
    nro_ome: str

    def to_row(self) -> dict[str, str]:
        return {
            "modo": self.modo,
            "afiliado": self.afiliado,
            "beneficio": self.beneficio,
            "dni": self.dni,
            "nombre": self.nombre,
            "diagnostico": self.diagnostico,
            "practica": self.practica,
            "resultado": self.resultado,
            "nro_ome": self.nro_ome,
        }


@dataclass(slots=True)
class RunSummary:
    ok: int = 0
    generadas: int = 0
    limite: int = 0
    bajas: int = 0
    no_dni: int = 0
    doble_dni: int = 0
    benef_completados: int = 0
    dni_completados: int = 0
    errores: int = 0

    def register(self, result: PatientResult) -> None:
        if result.resultado == "OK":
            self.ok += 1
        elif result.resultado in {"GENERADA", "YA_TIENE_OME"}:
            self.generadas += 1
        elif result.resultado in {"LIMITE", "LIMITE_ANUAL"}:
            self.limite += 1
        elif result.resultado == "BAJA":
            self.bajas += 1
        elif result.resultado == "NO_DNI":
            self.no_dni += 1
        elif result.resultado == "DOBLE_DNI":
            self.doble_dni += 1
        elif result.resultado == "BENEF_COMPLETADO":
            self.benef_completados += 1
        elif result.resultado == "DNI_COMPLETADO":
            self.dni_completados += 1
        elif result.resultado.startswith("ERROR"):
            self.errores += 1
        else:
            self.errores += 1

    @property
    def total(self) -> int:
        return (
            self.ok
            + self.generadas
            + self.limite
            + self.bajas
            + self.no_dni
            + self.doble_dni
            + self.benef_completados
            + self.dni_completados
            + self.errores
        )

    def as_dict(self) -> dict[str, int]:
        return {
            "ok": self.ok,
            "generadas": self.generadas,
            "limite": self.limite,
            "bajas": self.bajas,
            "no_dni": self.no_dni,
            "doble_dni": self.doble_dni,
            "benef_completados": self.benef_completados,
            "dni_completados": self.dni_completados,
            "errores": self.errores,
            "total": self.total,
        }


class OutputStore:
    def __init__(self, output_path: Path) -> None:
        self.output_path = output_path
        self.output_path.parent.mkdir(parents=True, exist_ok=True)
        self.processed: dict[tuple[str, str], PatientResult] = {}
        self._load_existing()

    def _load_existing(self) -> None:
        if not self.output_path.exists():
            return

        with self.output_path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            for row in reader:
                modo = sanitize_text(row.get("modo") or "BENEF").upper() or "BENEF"
                afiliado = sanitize_text(row.get("afiliado") or row.get("beneficiario"))
                resultado = sanitize_text(row.get("resultado"))
                if not afiliado or not resultado:
                    continue
                self.processed[(modo, afiliado)] = PatientResult(
                    modo=modo,
                    afiliado=afiliado,
                    beneficio=sanitize_text(row.get("beneficio")),
                    dni=sanitize_text(row.get("dni")),
                    nombre=sanitize_text(row.get("nombre")),
                    diagnostico=sanitize_text(row.get("diagnostico")),
                    practica=sanitize_text(row.get("practica") or row.get("practica_usada")),
                    resultado=resultado,
                    nro_ome=sanitize_text(row.get("nro_ome")),
                )

    def is_processed(self, modo: str, afiliado: str) -> bool:
        return (modo, afiliado) in self.processed

    def write(self, result: PatientResult) -> None:
        needs_header = not self.output_path.exists()
        with self.output_path.open("a", encoding="utf-8", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=RESULT_HEADERS)
            if needs_header:
                writer.writeheader()
            writer.writerow(result.to_row())
        self.processed[(result.modo, result.afiliado)] = result


class PamiOmeGenerator:
    def __init__(self, *, user: str, password: str, headless: bool = False) -> None:
        self.user = user
        self.password = password
        self.headless = headless
        self.keep_browser_open = not headless
        self.playwright: Playwright | None = None
        self.browser: Browser | None = None
        self.context: BrowserContext | None = None
        self.last_visible_page: Page | None = None
        self.benef_lookup_page: Page | None = None

    async def __aenter__(self) -> "PamiOmeGenerator":
        configure_playwright_browsers()
        self.playwright = await async_playwright().start()
        self.browser = await self.playwright.chromium.launch(headless=self.headless, slow_mo=120)
        self.context = await self.browser.new_context(ignore_https_errors=True)
        self.context.set_default_timeout(15000)
        await self.ensure_logged_in()
        return self

    async def __aexit__(self, exc_type, exc, tb) -> None:
        if self.keep_browser_open:
            return
        if self.context is not None:
            await self.context.close()
        if self.browser is not None:
            await self.browser.close()
        if self.playwright is not None:
            await self.playwright.stop()

    async def new_page(self) -> Page:
        if self.context is None:
            raise RuntimeError("No hay contexto de navegador disponible.")
        page = await self.context.new_page()
        page.set_default_timeout(15000)
        return page

    async def ensure_logged_in(self, page: Page | None = None) -> None:
        current_page = page
        owns_page = False
        if current_page is None:
            current_page = await self.new_page()
            owns_page = True
        try:
            if await self._session_is_active(current_page):
                return

            log("Iniciando sesion en CUP PAMI...")
            await current_page.goto(LOGIN_URL, wait_until="domcontentloaded")
            await human_pause(0.8, 1.4)

            user_input = current_page.locator('input[name="usuario"], input[type="text"], #usuario').first
            pass_input = current_page.locator('input[name="password"], input[type="password"], #password').first
            await user_input.wait_for(state="visible")
            await user_input.fill("")
            await human_pause()
            await user_input.type(self.user, delay=90)
            await human_pause()
            await pass_input.fill("")
            await human_pause()
            await pass_input.type(self.password, delay=90)
            await human_pause(0.7, 1.2)

            submit = current_page.locator(
                'button[type="submit"], input[type="submit"], button:has-text("Ingresar"), button:has-text("Entrar")'
            ).first
            await submit.click()
            await current_page.wait_for_load_state("domcontentloaded")
            await human_pause(2.0, 3.0)

            if LOGIN_URL in current_page.url and not await self._session_is_active(current_page):
                raise RuntimeError("No se pudo iniciar sesion en CUP PAMI. Revisar credenciales.")
        finally:
            if owns_page:
                await current_page.close()

    async def _session_is_active(self, page: Page) -> bool:
        try:
            response = await page.goto(GENERAR_ORDEN_URL, wait_until="domcontentloaded")
            await human_pause(0.7, 1.3)
            if "cup.pami.org.ar" in page.url:
                return False
            if response and response.status >= 400:
                return False
            return await page.locator("#pe-n_afiliado").count() > 0
        except Error:
            return False

    async def process_patient(self, patient: PatientInput) -> PatientResult:
        try:
            if patient.completar_benef:
                return await asyncio.wait_for(
                    self._process_complete_benef_patient(patient),
                    timeout=PATIENT_TIMEOUT_SECONDS,
                )
            if patient.completar_dni:
                return await asyncio.wait_for(
                    self._process_complete_dni_patient(patient),
                    timeout=PATIENT_TIMEOUT_SECONDS,
                )
            return await asyncio.wait_for(self._process_patient_impl(patient), timeout=PATIENT_TIMEOUT_SECONDS)
        except asyncio.TimeoutError:
            log(
                f"[OME] {patient.modo} {patient.afiliado}: timeout de {PATIENT_TIMEOUT_SECONDS}s "
                f"antes de completar el procesamiento."
            )
            return PatientResult(
                modo=patient.modo,
                afiliado=patient.afiliado,
                beneficio=patient.afiliado if patient.modo != "DNI" else "",
                dni=patient.afiliado if patient.modo == "DNI" else "",
                nombre="",
                diagnostico=patient.diagnostico,
                practica=patient.practica,
                resultado="ERROR_TIMEOUT",
                nro_ome="",
            )

    async def _process_patient_impl(self, patient: PatientInput) -> PatientResult:
        page: Page
        effective_patient = PatientInput(
            modo=patient.modo,
            afiliado=patient.afiliado,
            diagnostico=patient.diagnostico,
            practica=patient.practica,
            beneficio=patient.beneficio,
            dni=patient.dni,
            nombre=patient.nombre,
            retry_capita_done=patient.retry_capita_done,
            completar_benef=patient.completar_benef,
            completar_dni=patient.completar_dni,
            practicas_equivalentes=patient.practicas_equivalentes,
        )
        nombre = ""
        expected_beneficio = sanitize_text(patient.beneficio)
        expected_dni = sanitize_text(patient.dni)
        lookup_mode = patient.modo
        lookup_value = patient.afiliado
        if expected_beneficio and not patient.completar_benef:
            lookup_mode = "BENEF"
            lookup_value = expected_beneficio
            effective_patient.modo = "BENEF"
            effective_patient.afiliado = expected_beneficio
        elif lookup_mode == "DNI" and expected_dni:
            lookup_value = expected_dni
        n_afiliado_busqueda = lookup_value
        beneficio_actual = lookup_value if lookup_mode != "DNI" else ""
        dni_actual = lookup_value if lookup_mode == "DNI" else expected_dni
        if self.keep_browser_open and self.last_visible_page is not None:
            try:
                if not self.last_visible_page.is_closed():
                    page = self.last_visible_page
                else:
                    self.last_visible_page = None
                    page = await self.new_page()
            except Exception:
                self.last_visible_page = None
                page = await self.new_page()
        else:
            page = await self.new_page()
        try:
            await self.ensure_logged_in(page)
            await self._prepare_clean_form(page)
            await human_pause(0.8, 1.2)

            if lookup_mode == "DNI":
                lookup = await self._load_patient_by_dni(
                    page,
                    lookup_value,
                    patient.nombre,
                    allow_name_match_on_multiple=patient.completar_benef,
                )
                if lookup == "BAJA":
                    return PatientResult(
                        modo=patient.modo,
                        afiliado=patient.afiliado,
                        beneficio="",
                        dni=lookup_value,
                        nombre="",
                        diagnostico=patient.diagnostico,
                        practica=patient.practica,
                        resultado="BAJA",
                        nro_ome="",
                    )
                if lookup == "__ERROR__DNI_MULTIPLE":
                    return PatientResult(
                        modo=patient.modo,
                        afiliado=patient.afiliado,
                        beneficio="",
                        dni=lookup_value,
                        nombre="",
                        diagnostico=patient.diagnostico,
                        practica=patient.practica,
                        resultado="DOBLE_DNI",
                        nro_ome="",
                    )
                if lookup == "__ERROR__DNI_SIN_RESULTADOS":
                    return PatientResult(
                        modo=patient.modo,
                        afiliado=patient.afiliado,
                        beneficio="",
                        dni=lookup_value,
                        nombre="",
                        diagnostico=patient.diagnostico,
                        practica=patient.practica,
                        resultado="NO_DNI",
                        nro_ome="",
                    )
                if lookup.startswith("__ERROR__"):
                    return PatientResult(
                        modo=patient.modo,
                        afiliado=patient.afiliado,
                        beneficio="",
                        dni=lookup_value,
                        nombre="",
                        diagnostico=patient.diagnostico,
                        practica=patient.practica,
                        resultado=lookup.replace("__ERROR__", "ERROR_", 1),
                        nro_ome="",
                    )
                nombre = lookup
                n_afiliado_busqueda = await self._read_benefit_value(page) or lookup_value
            else:
                nombre = await self._load_patient_by_benefit(page, lookup_value)

            if not nombre:
                await self._save_error_screenshot(page, patient.afiliado, "afiliado_inexistente")
                fallback_result = "NO_DNI" if lookup_mode == "DNI" else "NO_PAMI"
                return PatientResult(
                    modo=patient.modo,
                    afiliado=patient.afiliado,
                    beneficio=expected_beneficio if lookup_mode != "DNI" else "",
                    dni=lookup_value if lookup_mode == "DNI" else expected_dni,
                    nombre="",
                    diagnostico=patient.diagnostico,
                    practica=patient.practica,
                    resultado=fallback_result,
                    nro_ome="",
                )

            beneficio_actual = await self._read_benefit_value(page) or beneficio_actual
            dni_actual = await self._read_document_value(page) or dni_actual

            if (
                lookup_mode == "DNI"
                and expected_beneficio
                and beneficio_actual
                and not patient.completar_benef
                and re.sub(r"\D+", "", beneficio_actual) != re.sub(r"\D+", "", expected_beneficio)
            ):
                log(
                    f"[OME] {patient.modo} {patient.afiliado}: BENEF distinto al esperado por DNI. "
                    f"esperado='{expected_beneficio}' obtenido='{beneficio_actual}'. Reintentando por BENEF."
                )
                await self._prepare_clean_form(page)
                nombre = await self._load_patient_by_benefit(page, expected_beneficio)
                if not nombre:
                    return PatientResult(
                        modo=patient.modo,
                        afiliado=patient.afiliado,
                        beneficio="",
                        dni=lookup_value,
                        nombre="",
                        diagnostico=patient.diagnostico,
                        practica=patient.practica,
                        resultado="NO_PAMI",
                        nro_ome="",
                    )
                effective_patient.modo = "BENEF"
                effective_patient.afiliado = expected_beneficio
                beneficio_actual = await self._read_benefit_value(page) or expected_beneficio
                dni_actual = await self._read_document_value(page) or expected_dni
                n_afiliado_busqueda = beneficio_actual or expected_beneficio

            if patient.completar_benef:
                log(
                    f"[OME] {patient.modo} {patient.afiliado}: BENEF_COMPLETADO "
                    f"beneficio='{beneficio_actual}' dni='{dni_actual}'."
                )
                return PatientResult(
                    modo=effective_patient.modo,
                    afiliado=effective_patient.afiliado,
                    beneficio=beneficio_actual,
                    dni=dni_actual,
                    nombre=nombre,
                    diagnostico=effective_patient.diagnostico,
                    practica=effective_patient.practica,
                    resultado="BENEF_COMPLETADO",
                    nro_ome="",
                )

            await self._fill_code_with_events(page, "#pe-diagnostico_cod", patient.diagnostico)
            await asyncio.sleep(1.5)
            await self._click_first_dropdown_starting_with(page, patient.diagnostico)

            await self._cleanup_modals(page)
            retry_capita_done = patient.retry_capita_done
            while True:
                current_practice = effective_patient.practica
                log(f"[OME] {patient.modo} {patient.afiliado}: intentando practica {current_practice}.")
                rows_before_clear = await self._count_prescribed_practice_rows(page)
                log(f"[OME] {patient.modo} {patient.afiliado}: filas de practica antes de limpiar={rows_before_clear}.")
                await self._clear_prescribed_practices(page)
                rows_after_clear = await self._count_prescribed_practice_rows(page)
                input_after_clear = await self._read_input_value(page, "#pe-practica")
                log(
                    f"[OME] {patient.modo} {patient.afiliado}: despues de limpiar filas={rows_after_clear} input_practica='{input_after_clear}'."
                )
                await self._cleanup_modals(page)
                await self._set_practice_code(page, current_practice)
                practice_input_after_set = await self._read_input_value(page, "#pe-practica")
                practice_rows_after_set = await self._count_prescribed_practice_rows(page)
                log(
                    f"[OME] {patient.modo} {patient.afiliado}: despues de cargar practica input_practica='{practice_input_after_set}' filas={practice_rows_after_set}."
                )
                practice_input_before_generate = await self._read_input_value(page, "#pe-practica")
                practice_rows_before_generate = await self._count_prescribed_practice_rows(page)
                log(
                    f"[OME] {patient.modo} {patient.afiliado}: antes de generar input_practica='{practice_input_before_generate}' filas={practice_rows_before_generate}."
                )
                await human_pause()
                await page.click("#pe-btn-generar-orden")
                outcome = await self._wait_for_generation_outcome(page, current_practice, timeout_ms=5000)
                log(f"[OME] {patient.modo} {patient.afiliado}: outcome detectado={outcome}.")
                if outcome == "confirm":
                    await human_pause(0.4, 0.8)
                    await self._click_confirm_in_modal(page)
                    await asyncio.sleep(5.0)
                elif outcome == "capita":
                    await asyncio.sleep(0.3)
                else:
                    await asyncio.sleep(5.0)

                feedback_text = await self._read_visible_feedback_text(page)
                if feedback_text:
                    log(f"[OME] {patient.modo} {patient.afiliado}: feedback visible='{self._summarize_text(feedback_text)}'.")
                body_text = sanitize_text((await page.locator("body").inner_text(timeout=5000)).lower())
                retry_feedback = "\n".join(part for part in [feedback_text, body_text] if part)
                if retry_feedback and self._should_retry_with_capita(retry_feedback, current_practice):
                    await self._close_capita_alert(page)
                    await self._cleanup_modals(page)
                    if retry_capita_done:
                        log(f"[OME] {patient.modo} {patient.afiliado}: el aviso de capita persiste despues del reintento con 427109.")
                        return PatientResult(
                            modo=effective_patient.modo,
                            afiliado=effective_patient.afiliado,
                            beneficio=beneficio_actual,
                            dni=dni_actual,
                            nombre=nombre,
                            diagnostico=effective_patient.diagnostico,
                            practica=effective_patient.practica,
                            resultado="ERROR_RESULTADO_NO_CONFIRMADO",
                            nro_ome="",
                        )
                    log(f"[OME] {patient.modo} {patient.afiliado}: practica {current_practice} no corresponde por capita. Reintentando con 427109 sobre el mismo paciente.")
                    effective_patient = PatientInput(
                        modo=effective_patient.modo,
                        afiliado=effective_patient.afiliado,
                        diagnostico=effective_patient.diagnostico,
                        practica="427109",
                        retry_capita_done=True,
                        practicas_equivalentes=effective_patient.practicas_equivalentes,
                    )
                    retry_capita_done = True
                    continue
                break
            combined_feedback = "\n".join(part for part in [feedback_text, body_text] if part)
            if self._contains_annual_limit_message(combined_feedback):
                log(f"[OME] {patient.modo} {patient.afiliado}: clasificado como LIMITE_ANUAL.")
                return PatientResult(
                    modo=effective_patient.modo,
                    afiliado=effective_patient.afiliado,
                    beneficio=beneficio_actual,
                    dni=dni_actual,
                    nombre=nombre,
                    diagnostico=effective_patient.diagnostico,
                    practica=effective_patient.practica,
                    resultado="LIMITE_ANUAL",
                    nro_ome="",
                )

            if self._contains_limit_message(combined_feedback):
                log(f"[OME] {patient.modo} {patient.afiliado}: clasificado como LIMITE.")
                return PatientResult(
                    modo=effective_patient.modo,
                    afiliado=effective_patient.afiliado,
                    beneficio=beneficio_actual,
                    dni=dni_actual,
                    nombre=nombre,
                    diagnostico=effective_patient.diagnostico,
                    practica=effective_patient.practica,
                    resultado="LIMITE",
                    nro_ome="",
                )

            if self._contains_existing_ome_message(combined_feedback):
                nro_ome_existente = ""
                try:
                    nro_ome_existente = await self._extract_ome_number_from_detail(page)
                except Exception as exc:
                    log(
                        f"[OME] {patient.modo} {patient.afiliado}: no se pudo extraer nro_ome directo "
                        f"desde modal GENERADA: {exc}"
                    )
                if not nro_ome_existente:
                    try:
                        await self._close_feedback_modal(page)
                    except Exception as exc:
                        log(f"[OME] {patient.modo} {patient.afiliado}: no se pudo cerrar el modal de OME existente: {exc}")
                try:
                    if not nro_ome_existente:
                        nro_ome_existente = await self._lookup_existing_generated_ome(
                            page, n_afiliado_busqueda, effective_patient.practica,
                            practicas_equivalentes=effective_patient.practicas_equivalentes,
                        )
                except Exception as exc:
                    log(
                        f"[OME] {patient.modo} {patient.afiliado}: no se pudo recuperar el Nro OME existente, "
                        f"pero el caso queda como YA_TIENE_OME: {exc}"
                    )
                log(
                    f"[OME] {patient.modo} {patient.afiliado}: clasificado como YA_TIENE_OME. nro_ome_existente='{nro_ome_existente}'."
                )
                return PatientResult(
                    modo=effective_patient.modo,
                    afiliado=effective_patient.afiliado,
                    beneficio=beneficio_actual,
                    dni=dni_actual,
                    nombre=nombre,
                    diagnostico=effective_patient.diagnostico,
                    practica=effective_patient.practica,
                    resultado="YA_TIENE_OME",
                    nro_ome=nro_ome_existente,
                )
            if self._contains_success_message(combined_feedback):
                nro_ome = await self._extract_ome_number_from_detail(page)
                log(f"[OME] {patient.modo} {patient.afiliado}: clasificado como OK. nro_ome='{nro_ome}'.")
                return PatientResult(
                    modo=effective_patient.modo,
                    afiliado=effective_patient.afiliado,
                    beneficio=beneficio_actual,
                    dni=dni_actual,
                    nombre=nombre,
                    diagnostico=effective_patient.diagnostico,
                    practica=effective_patient.practica,
                    resultado="OK",
                    nro_ome=nro_ome,
                )

            nro_ome = await self._extract_ome_number_from_detail(page)
            if nro_ome:
                log(f"[OME] {patient.modo} {patient.afiliado}: clasificado como OK por nro_ome visible. nro_ome='{nro_ome}'.")
                return PatientResult(
                    modo=effective_patient.modo,
                    afiliado=effective_patient.afiliado,
                    beneficio=beneficio_actual,
                    dni=dni_actual,
                    nombre=nombre,
                    diagnostico=effective_patient.diagnostico,
                    practica=effective_patient.practica,
                    resultado="OK",
                    nro_ome=nro_ome,
                )

            await self._save_error_screenshot(page, patient.afiliado, "sin_resultado")
            return PatientResult(
                modo=effective_patient.modo,
                afiliado=effective_patient.afiliado,
                beneficio=beneficio_actual,
                dni=dni_actual,
                nombre=nombre,
                diagnostico=effective_patient.diagnostico,
                practica=effective_patient.practica,
                resultado="ERROR_RESULTADO_NO_CONFIRMADO",
                nro_ome="",
            )
        except Exception as exc:
            log(f"[OME] {patient.modo} {patient.afiliado}: ERROR_PROCESO -> {type(exc).__name__}: {exc}")
            try:
                await self._save_error_screenshot(page, patient.afiliado, "exception")
            except Exception:
                pass
            return PatientResult(
                modo=effective_patient.modo,
                afiliado=effective_patient.afiliado,
                beneficio=beneficio_actual,
                dni=dni_actual,
                nombre=nombre,
                diagnostico=effective_patient.diagnostico,
                practica=effective_patient.practica,
                resultado="ERROR_PROCESO",
                nro_ome="",
            )
        finally:
            if self.keep_browser_open:
                self.last_visible_page = page
            else:
                await page.close()

    async def _process_complete_benef_patient(self, patient: PatientInput) -> PatientResult:
        lookup_value = sanitize_text(patient.dni or patient.afiliado)
        if not lookup_value:
            return PatientResult(
                modo=patient.modo,
                afiliado=patient.afiliado,
                beneficio="",
                dni="",
                nombre="",
                diagnostico=patient.diagnostico,
                practica=patient.practica,
                resultado="NO_DNI",
                nro_ome="",
            )

        page, modal = await self._benef_lookup_modal()
        lookup = await self._lookup_benefit_in_modal(
            page,
            modal,
            lookup_value,
            expected_name=patient.nombre,
            allow_name_match_on_multiple=True,
        )
        status = lookup.get("status", "")
        if status == "OK":
            beneficio = sanitize_text(lookup.get("beneficio"))
            dni = sanitize_text(lookup.get("dni")) or lookup_value
            nombre = sanitize_text(lookup.get("nombre"))
            log(
                f"[OME] DNI {lookup_value}: BENEF_COMPLETADO desde modal "
                f"beneficio='{beneficio}' nombre='{nombre}'."
            )
            return PatientResult(
                modo=patient.modo,
                afiliado=patient.afiliado,
                beneficio=beneficio,
                dni=dni,
                nombre=nombre,
                diagnostico=patient.diagnostico,
                practica=patient.practica,
                resultado="BENEF_COMPLETADO",
                nro_ome="",
            )
        if status == "BAJA":
            result = "BAJA"
        elif status == "__ERROR__DNI_MULTIPLE":
            result = "DOBLE_DNI"
        elif status == "__ERROR__DNI_MODAL":
            result = "ERROR_DNI_MODAL"
        else:
            result = "NO_DNI"
        return PatientResult(
            modo=patient.modo,
            afiliado=patient.afiliado,
            beneficio="",
            dni=lookup_value,
            nombre="",
            diagnostico=patient.diagnostico,
            practica=patient.practica,
            resultado=result,
            nro_ome="",
        )

    async def _process_complete_dni_patient(self, patient: PatientInput) -> PatientResult:
        lookup_value = sanitize_text(patient.beneficio or patient.afiliado)
        if not lookup_value:
            return PatientResult(
                modo=patient.modo,
                afiliado=patient.afiliado,
                beneficio="",
                dni="",
                nombre="",
                diagnostico=patient.diagnostico,
                practica=patient.practica,
                resultado="ERROR_SIN_BENEF",
                nro_ome="",
            )

        if self.keep_browser_open and self.last_visible_page is not None:
            try:
                if not self.last_visible_page.is_closed():
                    page = self.last_visible_page
                else:
                    self.last_visible_page = None
                    page = await self.new_page()
            except Exception:
                self.last_visible_page = None
                page = await self.new_page()
        else:
            page = await self.new_page()

        try:
            await self.ensure_logged_in(page)
            await self._prepare_clean_form(page)
            await human_pause(0.8, 1.2)
            nombre = await self._load_patient_by_benefit(page, lookup_value)
            if not nombre:
                await self._save_error_screenshot(page, lookup_value, "beneficio_inexistente")
                return PatientResult(
                    modo="BENEF",
                    afiliado=lookup_value,
                    beneficio=lookup_value,
                    dni="",
                    nombre="",
                    diagnostico=patient.diagnostico,
                    practica=patient.practica,
                    resultado="NO_PAMI",
                    nro_ome="",
                )
            beneficio = await self._read_benefit_value(page) or lookup_value
            dni = await self._read_document_value(page)
            if not dni:
                return PatientResult(
                    modo="BENEF",
                    afiliado=lookup_value,
                    beneficio=beneficio,
                    dni="",
                    nombre=nombre,
                    diagnostico=patient.diagnostico,
                    practica=patient.practica,
                    resultado="ERROR_DNI_VACIO",
                    nro_ome="",
                )
            log(
                f"[OME] BENEF {lookup_value}: DNI_COMPLETADO desde formulario "
                f"dni='{dni}' nombre='{nombre}'."
            )
            return PatientResult(
                modo="BENEF",
                afiliado=lookup_value,
                beneficio=beneficio,
                dni=dni,
                nombre=nombre,
                diagnostico=patient.diagnostico,
                practica=patient.practica,
                resultado="DNI_COMPLETADO",
                nro_ome="",
            )
        except Exception as exc:
            log(f"[OME] BENEF {lookup_value}: ERROR_PROCESO_COMPLETAR_DNI -> {type(exc).__name__}: {exc}")
            try:
                await self._save_error_screenshot(page, lookup_value, "completar_dni_exception")
            except Exception:
                pass
            return PatientResult(
                modo="BENEF",
                afiliado=lookup_value,
                beneficio=lookup_value,
                dni="",
                nombre="",
                diagnostico=patient.diagnostico,
                practica=patient.practica,
                resultado="ERROR_PROCESO",
                nro_ome="",
            )
        finally:
            if self.keep_browser_open:
                self.last_visible_page = page
            else:
                await page.close()

    async def _fill_input(self, page: Page, selector: str, value: str) -> None:
        locator = page.locator(selector).first
        await locator.wait_for(state="visible")
        await locator.click()
        await locator.fill("")
        await human_pause(0.2, 0.4)
        await locator.type(value, delay=80)

    async def _load_patient_by_benefit(self, page: Page, afiliado: str) -> str:
        previous_name = await self._read_patient_name(page)
        await self._fill_input(page, "#pe-n_afiliado", afiliado)
        await human_pause()
        await page.click("#pe-btn-bsq-afiliado")
        return await self._wait_for_patient_name(page, timeout_ms=5000, previous_value=previous_name)

    async def _open_patient_search_modal(self, page: Page) -> Locator | None:
        button = page.locator("#pe-btn-bsq-afiliado").first
        last_error: Exception | None = None
        for attempt in range(3):
            try:
                await self._cleanup_modals(page)
                await asyncio.sleep(0.2)
                await button.scroll_into_view_if_needed()
                if attempt == 0:
                    await button.click(timeout=2500)
                elif attempt == 1:
                    await button.click(timeout=2500, force=True)
                else:
                    await page.evaluate(
                        """
                        () => {
                          const btn = document.querySelector('#pe-btn-bsq-afiliado');
                          if (!btn) return false;
                          btn.dispatchEvent(new MouseEvent('click', { bubbles: true, cancelable: true, view: window }));
                          btn.click?.();
                          return true;
                        }
                        """
                    )
                await asyncio.sleep(0.65)
                modal = page.locator(".modal.show, .modal.in, .modal[style*='display: block']").last
                await modal.wait_for(state="visible", timeout=4000)
                return modal
            except Exception as exc:
                last_error = exc
                await self._cleanup_modals(page)
                await asyncio.sleep(0.3)
        log(f"[OME] No se pudo abrir el modal de busqueda: {last_error}")
        return None

    async def _load_patient_by_dni(
        self,
        page: Page,
        dni: str,
        expected_name: str = "",
        *,
        allow_name_match_on_multiple: bool = False,
    ) -> str:
        previous_name = await self._read_patient_name(page)
        modal = await self._open_patient_search_modal(page)
        if modal is None:
            log(f"[OME] No se pudo abrir el modal de busqueda por DNI {dni}.")
            return "__ERROR__DNI_MODAL"

        search_select = modal.locator("select").first
        await search_select.wait_for(state="visible", timeout=3000)
        await search_select.select_option("n_documento")
        await asyncio.sleep(0.35)

        primary_input = modal.locator("#bsq_avz_afiliado_valor").first
        filled = False
        try:
            await primary_input.wait_for(state="visible", timeout=2000)
            filled = await self._force_fill_visible_input(page, primary_input, dni)
        except Exception:
            filled = False

        if not filled:
            filled = await self._force_fill_exact_dni_input(page, dni)

        if not filled:
            filled = await self._force_fill_dni_modal_inputs(page, modal, dni)

        if not filled:
            log(f"[OME] No se pudo cargar el DNI {dni} en el modal de busqueda.")
            return "__ERROR__DNI_CARGA"
        log(f"[OME] DNI modal cargado con valor final: {await self._read_dni_modal_value(page)!r}")
        await human_pause(0.4, 0.8)

        buscar_button = modal.get_by_role("button", name="Buscar").first
        if await buscar_button.count() == 0:
            buscar_button = modal.locator("button:has-text('Buscar'), input[value='Buscar']").first
        await buscar_button.click()
        await human_pause(1.2, 1.8)

        selection = await self._select_active_affiliate_result(
            modal,
            expected_name,
            allow_name_match_on_multiple=allow_name_match_on_multiple,
        )
        if selection == "BAJA":
            return "BAJA"
        if selection == "__ERROR__DNI_MULTIPLE":
            return "__ERROR__DNI_MULTIPLE"
        if selection is None:
            return "__ERROR__DNI_SIN_RESULTADOS"
        patient_name = await self._wait_for_patient_name(page, timeout_ms=5000, previous_value=previous_name)
        return patient_name or "__ERROR__DNI_SELECCION"

    async def _benef_lookup_modal(self) -> tuple[Page, Locator]:
        page: Page | None = self.benef_lookup_page
        if page is None or page.is_closed():
            if self.last_visible_page is not None and not self.last_visible_page.is_closed():
                page = self.last_visible_page
            else:
                page = await self.new_page()
            self.benef_lookup_page = page
            self.last_visible_page = page
            await self.ensure_logged_in(page)
            await page.goto(GENERAR_ORDEN_URL, wait_until="domcontentloaded")
            await page.wait_for_selector("#pe-btn-bsq-afiliado")
            await self._wait_for_busy_overlay_to_clear(page)

        modal = page.locator(".modal.show, .modal.in, .modal[style*='display: block']").last
        try:
            if await modal.count() > 0 and await modal.is_visible(timeout=500):
                return page, modal
        except Exception:
            pass

        modal = await self._open_patient_search_modal(page)
        if modal is None:
            return page, page.locator("__modal_no_encontrado__")
        return page, modal

    async def _lookup_benefit_in_modal(
        self,
        page: Page,
        modal: Locator,
        dni: str,
        *,
        expected_name: str = "",
        allow_name_match_on_multiple: bool = False,
    ) -> dict[str, str]:
        dni = sanitize_text(dni)
        if not dni:
            return {"status": "__ERROR__DNI_SIN_RESULTADOS"}
        if await modal.count() == 0:
            return {"status": "__ERROR__DNI_MODAL"}

        try:
            search_select = modal.locator("select").first
            await search_select.wait_for(state="visible", timeout=3000)
            await search_select.select_option("n_documento")
            await asyncio.sleep(0.2)
        except Exception:
            return {"status": "__ERROR__DNI_MODAL"}

        filled = False
        primary_input = modal.locator("#bsq_avz_afiliado_valor").first
        try:
            await primary_input.wait_for(state="visible", timeout=2000)
            filled = await self._force_fill_visible_input(page, primary_input, dni)
        except Exception:
            filled = False
        if not filled:
            filled = await self._force_fill_exact_dni_input(page, dni)
        if not filled:
            filled = await self._force_fill_dni_modal_inputs(page, modal, dni)
        if not filled:
            return {"status": "__ERROR__DNI_CARGA"}

        buscar_button = modal.get_by_role("button", name="Buscar").first
        if await buscar_button.count() == 0:
            buscar_button = modal.locator("button:has-text('Buscar'), input[value='Buscar']").first
        await buscar_button.click()

        expected_digits = re.sub(r"\D+", "", dni)
        for _ in range(12):
            await asyncio.sleep(0.25)
            result = await self._read_benefit_modal_result(
                modal,
                expected_digits,
                expected_name=expected_name,
                allow_name_match_on_multiple=allow_name_match_on_multiple,
            )
            if result.get("status") != "__WAIT__":
                return result
        return {"status": "__ERROR__DNI_SIN_RESULTADOS"}

    async def _read_benefit_modal_result(
        self,
        modal: Locator,
        expected_digits: str,
        *,
        expected_name: str = "",
        allow_name_match_on_multiple: bool = False,
    ) -> dict[str, str]:
        rows = modal.locator("table tbody tr")
        count = await rows.count()
        if count == 0:
            return {"status": "__WAIT__"}

        active: list[dict[str, str]] = []
        saw_matching_inactive = False
        for index in range(count):
            row = rows.nth(index)
            cells = row.locator("td")
            cell_count = await cells.count()
            if cell_count < 4:
                continue
            beneficio_base = sanitize_text(await cells.nth(0).inner_text())
            gp = sanitize_text(await cells.nth(1).inner_text()) if cell_count > 1 else ""
            beneficio = combine_beneficio_gp(beneficio_base, gp)
            nombre = sanitize_text(await cells.nth(2).inner_text())
            documento = sanitize_text(await cells.nth(3).inner_text())
            doc_digits = re.sub(r"\D+", "", documento)
            if expected_digits and doc_digits != expected_digits:
                continue
            fecha_baja = sanitize_text(await cells.nth(cell_count - 1).inner_text())
            item = {
                "status": "OK",
                "beneficio": beneficio,
                "dni": documento,
                "nombre": nombre,
            }
            if fecha_baja:
                saw_matching_inactive = True
            else:
                active.append(item)

        if not active:
            return {"status": "BAJA"} if saw_matching_inactive else {"status": "__WAIT__"}
        if len(active) == 1:
            return active[0]
        if allow_name_match_on_multiple and expected_name:
            scored = [
                (self._name_match_score(expected_name, item.get("nombre", "")), item)
                for item in active
            ]
            scored.sort(key=lambda item: item[0], reverse=True)
            if scored and scored[0][0] >= 0.5:
                top_score = scored[0][0]
                top_matches = [item for score, item in scored if score == top_score]
                if len(top_matches) == 1:
                    log(
                        "[OME] DNI multiple resuelto solo para completar BENEF por nombre de hoja: "
                        f"'{expected_name}' -> '{top_matches[0].get('nombre', '')}'."
                    )
                    return top_matches[0]
        return {"status": "__ERROR__DNI_MULTIPLE"}

    async def _prepare_clean_form(self, page: Page) -> None:
        await page.goto("about:blank", wait_until="domcontentloaded")
        await page.goto(GENERAR_ORDEN_URL, wait_until="domcontentloaded")
        await page.wait_for_selector("#pe-btn-bsq-afiliado")
        await self._wait_for_busy_overlay_to_clear(page)
        await self._cleanup_modals(page)
        await self._hard_reset_generate_form(page)

    async def _read_benefit_value(self, page: Page) -> str:
        try:
            return sanitize_text(await page.locator("#pe-n_afiliado").first.input_value())
        except Exception:
            return ""

    async def _hard_reset_generate_form(self, page: Page) -> None:
        try:
            await page.evaluate(
                """
                () => {
                  const clearValue = (selector) => {
                    const el = document.querySelector(selector);
                    if (!el) return;
                    try {
                      el.value = '';
                      el.setAttribute('value', '');
                      el.dispatchEvent(new Event('input', { bubbles: true }));
                      el.dispatchEvent(new Event('change', { bubbles: true }));
                    } catch (e) {}
                  };

                  ['#pe-n_afiliado', '#pe-nomyap_afiliado', '#pe-n_doc', '#pe-nrodoc_afiliado', '#pe-diagnostico_cod', '#pe-practica']
                    .forEach(clearValue);

                  document.querySelectorAll('.modal.show, .modal.in, .modal[style*="display: block"], .modal-backdrop, .swal2-container').forEach((node) => {
                    try { node.remove(); } catch (e) {}
                  });
                  document.body.classList.remove('modal-open');
                  document.body.style.removeProperty('padding-right');

                  document.querySelectorAll('tr .fa-trash, tr .fa-remove, tr .fa-times').forEach((icon) => {
                    const action = icon.closest('button, a, span');
                    try { action?.click(); } catch (e) {}
                  });
                }
                """
            )
            await asyncio.sleep(0.35)
            await self._wait_for_busy_overlay_to_clear(page)
        except Exception:
            pass

    async def _read_document_value(self, page: Page) -> str:
        selectors = [
            "#pe-n_doc",
            "#pe-nrodoc_afiliado",
            "#pe-n_doc_afiliado",
            "#pe-n_documento",
            "input[id*='doc']",
            "input[name*='doc']",
        ]
        for selector in selectors:
            try:
                locator = page.locator(selector).first
                if await locator.count() == 0:
                    continue
                value = sanitize_text(await locator.input_value())
                if value:
                    return value
            except Exception:
                continue
        return ""

    async def _lookup_ome_cabecera_raw(self, page: Page, n_afiliado: str,
                                       equivalentes: tuple[str, ...]) -> str:
        """Recupera el Nro OME de cabecera leyendo la bandeja CRUDA del panel.

        Busca al afiliado (sin filtro de práctica), lee todas sus OMEs de
        window.bandeja y elige la de código de práctica dentro de la cascada
        (`equivalentes`), la de emisión MÁS RECIENTE. Sirve para YA_TIENE_OME:
        la OME recién generada queda 'PENDIENTE DE ACEPTACIÓN' y el extractor
        normal (activables/agenda futura) la descarta, pero acá aparece igual."""
        equiv = set(equivalentes or ())
        if not equiv:
            return ""
        try:
            await self._search_existing_ome_in_panel(page, n_afiliado=n_afiliado, practica="")
            await page.wait_for_function("window.bandeja !== undefined", timeout=8000)
            raw = await page.evaluate(RAW_OMES_AFILIADO_SCRIPT) or []
        except Exception as exc:
            log(f"[OME] No pude leer la bandeja cruda para {n_afiliado}: {exc}")
            return ""

        def codigos(o: dict) -> set:
            cs = {self._extract_practice_code(o.get("c_practica", ""))}
            for c in (o.get("c_practica_array") or []):
                cs.add(self._extract_practice_code(c))
            return {c for c in cs if c}

        candidatos = [o for o in raw if codigos(o) & equiv]
        if not candidatos:
            log(f"[OME] Bandeja cruda: {len(raw)} OME(s) pero ninguna de cabecera "
                f"({'/'.join(sorted(equiv))}) para {n_afiliado}.")
            return ""
        candidatos.sort(key=lambda o: o.get("f_emision_date") or 0, reverse=True)
        elegida = candidatos[0]
        log(f"[OME] Bandeja cruda: elegida n_orden={elegida.get('n_orden')} "
            f"practica={elegida.get('c_practica')} estado={elegida.get('estado')} "
            f"(de {len(candidatos)} de cabecera).")
        return sanitize_text(elegida.get("n_orden"))

    async def _lookup_existing_generated_ome(self, page: Page, n_afiliado: str, practica: str,
                                             practicas_equivalentes: tuple[str, ...] = ()) -> str:
        n_afiliado = sanitize_text(n_afiliado)
        practica = sanitize_text(practica)
        # Códigos equivalentes normalizados a 6 dígitos (cabecera: las 4 consultas).
        equivalentes = tuple(c for c in (self._extract_practice_code(p) for p in (practicas_equivalentes or ())) if c)
        if not n_afiliado:
            return ""
        try:
            log(f"[OME] Buscando OME existente para {n_afiliado} con practica='{practica}'"
                + (f" (equivalentes {'/'.join(equivalentes)})" if equivalentes else "") + ".")
            nro_visible = await self._extract_ome_number_from_detail(page)
            if nro_visible:
                log(f"[OME] OME existente recuperada desde pantalla actual para {n_afiliado}: {nro_visible}")
                return nro_visible
            await page.goto(PANEL_ACEPTACION_URL, wait_until="domcontentloaded")
            await human_pause(1.0, 1.4)
            # En modo equivalentes (cabecera) buscamos SIN filtro de práctica para
            # que PAMI liste todas las OMEs del afiliado; el filtrado lo hace
            # _pick_best_existing_ome contra el set equivalente.
            practica_busqueda = "" if equivalentes else practica
            # Cabecera: leer la lista CRUDA de window.bandeja y elegir por código de
            # práctica (la recién generada queda "PENDIENTE DE ACEPTACIÓN" y el
            # extractor normal la descarta). Es lo más confiable para YA_TIENE_OME.
            if equivalentes:
                nro_raw = await self._lookup_ome_cabecera_raw(page, n_afiliado, equivalentes)
                if nro_raw:
                    log(f"[OME] OME de cabecera recuperada (bandeja cruda) para {n_afiliado}: {nro_raw}")
                    return nro_raw
            omes = await self._search_existing_ome_with_fallback(page, n_afiliado=n_afiliado, practica=practica_busqueda)
            nro_ome = self._pick_best_existing_ome(omes, practica, equivalentes=equivalentes)
            if nro_ome:
                log(f"[OME] OME existente recuperada para {n_afiliado}: {nro_ome}")
            else:
                log(f"[OME] No se encontro Nro OME existente para {n_afiliado}, pero el caso queda como YA_TIENE_OME.")
            return nro_ome
        except Exception as exc:
            log(f"[OME] No se pudo recuperar OME existente para {n_afiliado}: {exc}")
            return ""

    async def _search_existing_ome_with_fallback(self, page: Page, n_afiliado: str, practica: str) -> list[dict]:
        attempts = [practica]
        if practica and not self._extract_practice_code(practica):
            attempts.append("")
        last_error = ""
        for current_practice in attempts:
            try:
                await self._search_existing_ome_in_panel(page, n_afiliado=n_afiliado, practica=current_practice)
                await page.wait_for_function("window.bandeja !== undefined", timeout=8000)
                omes = await page.evaluate(EXTRAER_OMES_EXISTENTES_SCRIPT, time.strftime("%d/%m/%Y")) or []
                if omes:
                    log(
                        f"[OME] Panel de aceptacion devolvio {len(omes)} OME(s) para {n_afiliado} "
                        f"con practica='{current_practice}'."
                    )
                    return omes
                log(
                    f"[OME] Panel de aceptacion sin resultados para {n_afiliado} "
                    f"con practica='{current_practice}'."
                )
            except Exception as exc:
                last_error = str(exc)
                log(
                    f"[OME] Fallo buscando OME existente para {n_afiliado} "
                    f"con practica='{current_practice}': {exc}"
                )
        if last_error:
            log(f"[OME] Se agotaron los intentos de busqueda de OME existente para {n_afiliado}: {last_error}")
        return []

    async def _search_existing_ome_in_panel(self, page: Page, n_afiliado: str, practica: str) -> None:
        result = await page.evaluate(
            """
            ({ afiliado, practica }) => {
              try {
                window.bandeja = undefined;
                var fd = document.getElementById('f_emision_desde');
                if (fd && typeof $ !== 'undefined') {
                  $(fd).datepicker('setDate', '01/01/2025');
                }
              } catch (e) {}
              var inp = document.querySelector('input[name="n_afiliado"]');
              var practicaInput =
                document.querySelector('input[name="practica"]') ||
                document.querySelector('input[name="c_practica"]') ||
                document.querySelector('input[id="practica"]') ||
                document.querySelector('input[id="c_practica"]') ||
                document.querySelector('input[placeholder*="Práctica"]') ||
                document.querySelector('input[placeholder*="Practica"]');
              if (!inp) return 'NO_CAMPO';
              inp.value = afiliado || '';
              inp.dispatchEvent(new Event('input', { bubbles: true }));
              inp.dispatchEvent(new Event('change', { bubbles: true }));
              if (practicaInput) {
                practicaInput.value = practica || '';
                practicaInput.dispatchEvent(new Event('input', { bubbles: true }));
                practicaInput.dispatchEvent(new Event('change', { bubbles: true }));
              }
              var btn = document.querySelector('input[type="submit"][value="Buscar"]');
              if (!btn) return 'NO_BTN';
              btn.click();
              return 'OK';
            }
            """,
            {"afiliado": n_afiliado, "practica": practica},
        )
        if result != "OK":
            raise RuntimeError(f"No se pudo buscar OME existente en panel: {result}")
        await human_pause(1.5, 2.0)

    def _pick_best_existing_ome(self, omes: list[dict], practica: str,
                                equivalentes: tuple[str, ...] = ()) -> str:
        practica_codigo = self._extract_practice_code(practica)
        # Set de códigos aceptables: si vienen equivalentes (cabecera), cualquiera
        # de esos códigos vale; si no, solo el código exacto de la práctica.
        codigos_ok = set(equivalentes) if equivalentes else ({practica_codigo} if practica_codigo else set())
        candidatos = []
        for ome in omes or []:
            practica_item = sanitize_text(ome.get("practica"))
            if codigos_ok:
                if self._extract_practice_code(practica_item) not in codigos_ok:
                    continue
            elif practica and practica.upper() not in practica_item.upper():
                continue
            candidatos.append(ome)
        if not candidatos:
            if not practica_codigo:
                # Some PAMI listings expose only the practice description, not the code.
                # This fallback is safe only when the requested practice was not a concrete code.
                candidatos = [ome for ome in (omes or []) if sanitize_text(ome.get("n_orden"))]
        if not candidatos:
            return ""

        def prioridad(ome: dict) -> tuple[int, int, int, str]:
            estado = sanitize_text(ome.get("estado_accion")).lower()
            turno_asignado = bool(ome.get("turno_asignado"))
            transmitida = bool(ome.get("ya_transmitida")) or bool(ome.get("transmitida"))
            mismo_efector = bool(ome.get("mismo_efector"))
            if estado == "disponible_activar" and not transmitida:
                base = 0
            elif turno_asignado and not transmitida:
                base = 1
            elif not transmitida:
                base = 2
            else:
                base = 3
            return (base, 0 if mismo_efector else 1, 0 if turno_asignado else 1, sanitize_text(ome.get("n_orden")))

        candidatos.sort(key=prioridad)
        return sanitize_text((candidatos[0] if candidatos else {}).get("n_orden"))

    def _extract_practice_code(self, practice: str) -> str:
        digits = "".join(ch for ch in str(practice or "") if ch.isdigit())
        return digits[:6] if len(digits) >= 6 else ""

    def _normalize_modal_text(self, value: str) -> str:
        normalized = unicodedata.normalize("NFD", str(value or ""))
        normalized = "".join(ch for ch in normalized if unicodedata.category(ch) != "Mn")
        return " ".join(normalized.lower().split())

    def _name_match_score(self, expected_name: str, candidate_name: str) -> float:
        expected = self._normalize_modal_text(expected_name)
        candidate = self._normalize_modal_text(candidate_name)
        if not expected or not candidate:
            return 0.0
        expected_parts = [part for part in expected.split() if len(part) > 1]
        candidate_parts = set(part for part in candidate.split() if len(part) > 1)
        if not expected_parts or not candidate_parts:
            return 0.0
        hits = sum(1 for part in expected_parts if part in candidate_parts)
        return hits / len(expected_parts)

    def _contains_existing_ome_message(self, text: str) -> bool:
        normalized = self._normalize_modal_text(text)
        return "ya cuenta con una orden medica electronica" in normalized

    def _contains_limit_message(self, text: str) -> bool:
        normalized = self._normalize_modal_text(text)
        return "alcanzo la cantidad mensual" in normalized or (
            "cantidad mensual" in normalized and "prestaciones realizadas" in normalized
        )

    def _contains_annual_limit_message(self, text: str) -> bool:
        normalized = self._normalize_modal_text(text)
        return "alcanzo la cantidad anual" in normalized or (
            "cantidad anual" in normalized and "prestaciones realizadas" in normalized
        )

    def _contains_success_message(self, text: str) -> bool:
        normalized = self._normalize_modal_text(text)
        return "se ha generado una orden" in normalized

    def _should_retry_with_capita(self, body_text: str, practica: str) -> bool:
        practice_code = self._extract_practice_code(practica)
        normalized = self._normalize_modal_text(body_text)
        return (
            practice_code == "427122"
            and f"no se puede prescribir la practica {practice_code}" in normalized
            and "forma parte de su capita" in normalized
        )

    async def _read_patient_name(self, page: Page) -> str:
        selectors = [
            "#pe-nomyap_afiliado",
            "#pe-nombre-afiliado",
            "#pe-afiliado-nombre",
            "#pe_nombre_afiliado",
        ]
        for selector in selectors:
            locator = page.locator(selector).first
            if await locator.count() == 0:
                continue
            try:
                if await locator.evaluate("el => el.tagName.toLowerCase() === 'input'"):
                    return sanitize_text(await locator.input_value())
                return sanitize_text(await locator.inner_text())
            except Error:
                continue
        return ""

    async def _read_visible_feedback_text(self, page: Page) -> str:
        try:
            alert = page.locator(".alert.rin-alert.alert-danger").last
            if await alert.count() > 0 and await alert.is_visible():
                return sanitize_text(await alert.inner_text())
            text = await page.evaluate(
                """
                () => {
                  const normalizedVisible = (el) => {
                    if (!el) return false;
                    const style = window.getComputedStyle(el);
                    const rect = el.getBoundingClientRect();
                    return style.display !== 'none' && style.visibility !== 'hidden' && rect.width > 40 && rect.height > 20;
                  };
                  const selectors = [
                    '.modal.show',
                    '.modal.in',
                    '.modal[style*="display: block"]',
                    '.sweet-alert',
                    '.swal2-popup',
                    '.swal2-container',
                    '.alert',
                    '[role="dialog"]'
                  ];
                  const candidates = Array.from(document.querySelectorAll(selectors.join(',')))
                    .filter(normalizedVisible)
                    .map((node) => String(node.innerText || node.textContent || '').trim())
                    .filter(Boolean)
                    .sort((a, b) => b.length - a.length);
                  for (const text of candidates) {
                    if (text) return text;
                  }
                  return '';
                }
                """
            )
            return sanitize_text(text)
        except Exception:
            pass
        return ""

    async def _force_fill_visible_input(self, page: Page, locator: Locator, value: str) -> bool:
        desired = sanitize_text(value)
        try:
            await locator.click()
            await asyncio.sleep(0.08)
        except Exception:
            pass
        try:
            await locator.fill("")
        except Exception:
            pass
        await asyncio.sleep(0.12)
        try:
            await locator.type(desired, delay=80)
        except Exception:
            pass
        await asyncio.sleep(0.12)
        try:
            current_value = sanitize_text(await locator.input_value())
        except Exception:
            current_value = ""
        if current_value == desired:
            return True
        try:
            handle = await locator.element_handle()
            if handle is not None:
                await page.evaluate(
                    """
                    ([el, val]) => {
                      if (!el) return;
                      el.focus();
                      const setter = Object.getOwnPropertyDescriptor(HTMLInputElement.prototype, 'value')?.set;
                      if (setter) {
                        setter.call(el, val);
                      } else {
                        el.value = val;
                      }
                      el.dispatchEvent(new Event('input', { bubbles: true }));
                      el.dispatchEvent(new KeyboardEvent('keydown', { bubbles: true, key: '0' }));
                      el.dispatchEvent(new KeyboardEvent('keyup', { bubbles: true, key: '0' }));
                      el.dispatchEvent(new Event('change', { bubbles: true }));
                    }
                    """,
                    [handle, desired],
                )
                await asyncio.sleep(0.18)
        except Exception:
            pass
        try:
            current_value = sanitize_text(await locator.input_value())
        except Exception:
            current_value = ""
        return current_value == desired

    async def _force_fill_exact_dni_input(self, page: Page, value: str) -> bool:
        desired = sanitize_text(value)
        try:
            result = await page.evaluate(
                """
                (dni) => {
                  const el = document.getElementById('bsq_avz_afiliado_valor');
                  if (!el || el.disabled || el.readOnly) {
                    return { ok: false, value: el ? (el.value || '') : '' };
                  }
                  el.focus();
                  el.click();
                  const setter = Object.getOwnPropertyDescriptor(HTMLInputElement.prototype, 'value')?.set;
                  if (setter) {
                    setter.call(el, '');
                    setter.call(el, dni);
                  } else {
                    el.value = '';
                    el.value = dni;
                  }
                  el.setAttribute('value', dni);
                  el.dispatchEvent(new Event('input', { bubbles: true }));
                  el.dispatchEvent(new KeyboardEvent('keydown', { bubbles: true, key: '0' }));
                  el.dispatchEvent(new KeyboardEvent('keypress', { bubbles: true, key: '0' }));
                  el.dispatchEvent(new KeyboardEvent('keyup', { bubbles: true, key: '0' }));
                  el.dispatchEvent(new Event('change', { bubbles: true }));
                  el.dispatchEvent(new Event('blur', { bubbles: true }));
                  return { ok: String((el.value || '').trim()) === String(dni).trim(), value: el.value || '' };
                }
                """,
                desired,
            )
            await asyncio.sleep(0.22)
            return bool(result and result.get("ok"))
        except Exception:
            return False

    async def _read_dni_modal_value(self, page: Page) -> str:
        try:
            return sanitize_text(
                await page.evaluate(
                    """
                    () => {
                      const el = document.getElementById('bsq_avz_afiliado_valor');
                      return el ? (el.value || '') : '';
                    }
                    """
                )
            )
        except Exception:
            return ""

    async def _force_fill_dni_modal_inputs(self, page: Page, modal: Locator, value: str) -> bool:
        desired = sanitize_text(value)
        try:
            handle = await modal.element_handle()
            if handle is None:
                return False
            result = await page.evaluate(
                """
                ([modalEl, dni]) => {
                  if (!modalEl) return { ok: false, values: [] };
                  const visibleInputs = Array.from(
                    modalEl.querySelectorAll("input[type='text'], input[type='type'], input:not([type]), input#bsq_avz_afiliado_valor")
                  ).filter((el) => {
                    if (!el || el.disabled || el.readOnly) return false;
                    const style = window.getComputedStyle(el);
                    const rect = el.getBoundingClientRect();
                    return style.display !== 'none' && style.visibility !== 'hidden' && rect.width > 40 && rect.height > 10;
                  });
                  const setter = Object.getOwnPropertyDescriptor(HTMLInputElement.prototype, 'value')?.set;
                  const values = [];
                  for (const el of visibleInputs) {
                    el.focus();
                    el.click();
                    if (setter) {
                      setter.call(el, dni);
                    } else {
                      el.value = dni;
                    }
                    el.setAttribute('value', dni);
                    el.dispatchEvent(new Event('input', { bubbles: true }));
                    el.dispatchEvent(new KeyboardEvent('keydown', { bubbles: true, key: '0' }));
                    el.dispatchEvent(new KeyboardEvent('keypress', { bubbles: true, key: '0' }));
                    el.dispatchEvent(new KeyboardEvent('keyup', { bubbles: true, key: '0' }));
                    el.dispatchEvent(new Event('change', { bubbles: true }));
                    el.dispatchEvent(new Event('blur', { bubbles: true }));
                    values.push((el.value || '').trim());
                  }
                  return { ok: values.includes(String(dni).trim()), values };
                }
                """,
                [handle, desired],
            )
            await asyncio.sleep(0.22)
            return bool(result and result.get("ok"))
        except Exception:
            return False

    async def _select_active_affiliate_result(
        self,
        modal: Locator,
        expected_name: str = "",
        *,
        allow_name_match_on_multiple: bool = False,
    ) -> str | None:
        rows = modal.locator("table tbody tr")
        await rows.first.wait_for(state="visible", timeout=5000)
        count = await rows.count()
        if count == 0:
            return None

        active_indexes: list[int] = []
        active_candidates: list[tuple[int, str]] = []
        for index in range(count):
            row = rows.nth(index)
            cell_count = await row.locator("td").count()
            if cell_count == 0:
                continue
            fecha_baja = sanitize_text(await row.locator("td").nth(cell_count - 1).inner_text())
            if not fecha_baja:
                active_indexes.append(index)
                candidate_name = ""
                if cell_count >= 3:
                    candidate_name = sanitize_text(await row.locator("td").nth(2).inner_text())
                active_candidates.append((index, candidate_name))

        if not active_indexes:
            return "BAJA"
        if len(active_indexes) > 1:
            if allow_name_match_on_multiple and expected_name:
                scored = [
                    (self._name_match_score(expected_name, candidate_name), index, candidate_name)
                    for index, candidate_name in active_candidates
                ]
                scored.sort(reverse=True)
                if scored and scored[0][0] >= 0.5:
                    top_score = scored[0][0]
                    top_matches = [item for item in scored if item[0] == top_score]
                    if len(top_matches) == 1:
                        index = top_matches[0][1]
                        log(
                            "[OME] DNI multiple resuelto solo para completar BENEF por nombre de hoja: "
                            f"'{expected_name}' -> '{top_matches[0][2]}'."
                        )
                        await rows.nth(index).click()
                        await human_pause(0.6, 1.0)
                        return "OK"
            log(
                "[OME] DNI multiple: se encontraron "
                f"{len(active_indexes)} afiliados activos para el mismo documento. "
                "No se selecciona automaticamente."
            )
            return "__ERROR__DNI_MULTIPLE"

        await rows.nth(active_indexes[0]).click()
        await human_pause(0.6, 1.0)
        return "OK"

    async def _fill_code_with_events(self, page: Page, selector: str, value: str) -> None:
        await self._fill_input(page, selector, value)
        await page.eval_on_selector(
            selector,
            """
            (element, val) => {
                element.value = val;
                element.dispatchEvent(new Event('input', { bubbles: true }));
                element.dispatchEvent(new KeyboardEvent('keyup', { bubbles: true, key: val.slice(-1) || '' }));
            }
            """,
            value,
        )

    async def _force_fill_text_input(self, page: Page, selector: str, value: str) -> bool:
        desired = sanitize_text(value)
        try:
            result = await page.evaluate(
                """
                ({ selector, value }) => {
                  const el = document.querySelector(selector);
                  if (!el) return { ok: false, value: '' };
                  const setter = Object.getOwnPropertyDescriptor(HTMLInputElement.prototype, 'value')?.set;
                  el.focus();
                  el.click();
                  if (setter) {
                    setter.call(el, '');
                    setter.call(el, value);
                  } else {
                    el.value = '';
                    el.value = value;
                  }
                  el.setAttribute('value', value);
                  el.dispatchEvent(new Event('input', { bubbles: true }));
                  el.dispatchEvent(new KeyboardEvent('keydown', { bubbles: true, key: value.slice(-1) || '0' }));
                  el.dispatchEvent(new KeyboardEvent('keyup', { bubbles: true, key: value.slice(-1) || '0' }));
                  el.dispatchEvent(new Event('change', { bubbles: true }));
                  return { ok: String(el.value || '').trim() === String(value || '').trim(), value: String(el.value || '').trim() };
                }
                """,
                {"selector": selector, "value": desired},
            )
            await asyncio.sleep(0.25)
            return bool(result and result.get("ok"))
        except Exception:
            return False

    async def _set_practice_code(self, page: Page, code: str) -> None:
        last_value = ""
        for attempt in range(1, 4):
            await self._wait_for_busy_overlay_to_clear(page)
            try:
                practice_tab = page.locator(
                    "button:has-text('Pr??cticas'), button:has-text('Practicas'), a:has-text('Pr??cticas'), a:has-text('Practicas')"
                ).first
                if await practice_tab.count() > 0:
                    await practice_tab.click()
                    await asyncio.sleep(0.15)
            except Exception:
                pass
            locator = page.locator("#pe-practica").first
            await locator.wait_for(state="visible", timeout=5000)
            try:
                await locator.click()
                await asyncio.sleep(0.08)
                await locator.fill("")
                await asyncio.sleep(0.08)
                await locator.type(code, delay=70)
            except Exception:
                pass
            current_value = await self._read_input_value(page, "#pe-practica")
            if current_value != code:
                await self._force_fill_text_input(page, "#pe-practica", code)
                current_value = await self._read_input_value(page, "#pe-practica")
            last_value = current_value
            if current_value == code:
                await asyncio.sleep(1.2)
                try:
                    await self._click_first_dropdown_starting_with(page, code)
                except Exception:
                    await self._click_first_dropdown_containing(page, code)
                await self._wait_for_busy_overlay_to_clear(page)
                return
            log(f"[OME] No se pudo fijar practica {code} en intento {attempt}. Valor actual='{current_value}'.")
            await asyncio.sleep(0.4)
        raise RuntimeError(
            f"No se pudo verificar el codigo de practica {code} en el input. Valor actual='{last_value}'."
        )

    async def _read_input_value(self, page: Page, selector: str) -> str:
        try:
            value = await page.locator(selector).first.input_value(timeout=1000)
            return (value or "").strip()
        except Exception:
            try:
                value = await page.evaluate(
                    """
                    (sel) => {
                      const el = document.querySelector(sel);
                      return el ? String(el.value || '').trim() : '';
                    }
                    """,
                    selector,
                )
                return (value or "").strip()
            except Exception:
                return ""

    async def _cleanup_modals(self, page: Page) -> None:
        await page.evaluate(
            """
            () => {
                document.querySelectorAll('.modal.show, .modal.in, .modal[style*="display: block"]').forEach((modal) => {
                    modal.style.display = 'none';
                    modal.classList.remove('show', 'in');
                    modal.setAttribute('aria-hidden', 'true');
                });
                document.querySelectorAll('.modal-backdrop').forEach((node) => node.remove());
                document.body.classList.remove('modal-open');
                document.body.style.removeProperty('padding-right');
                document.body.style.overflow = 'auto';
            }
            """
        )
        await human_pause(0.2, 0.4)

    async def _click_first_dropdown_starting_with(self, page: Page, prefix: str) -> None:
        await self._click_dropdown_item(
            page,
            """
            (items, wanted) => {
                const normalized = wanted.trim().toLowerCase();
                return items.find((text) => text.trim().toLowerCase().startsWith(normalized)) || null;
            }
            """,
            prefix,
        )

    async def _click_first_dropdown_containing(self, page: Page, fragment: str) -> None:
        await self._click_dropdown_item(
            page,
            """
            (items, wanted) => {
                const normalized = wanted.trim().toLowerCase();
                return items.find((text) => text.trim().toLowerCase().includes(normalized)) || null;
            }
            """,
            fragment,
        )

    async def _count_prescribed_practice_rows(self, page: Page) -> int:
        try:
            count = await page.evaluate(
                """
                () => {
                  const normalized = (value) => String(value || '').toLowerCase().replace(/\\s+/g, ' ').trim();
                  const tables = Array.from(document.querySelectorAll('table'));
                  const targetTable = tables.find((table) => {
                    const text = normalized(table.textContent || '');
                    return text.includes('cód. práctica') && text.includes('acciones');
                  });
                  if (!targetTable) return 0;
                  const rows = Array.from(targetTable.querySelectorAll('tbody tr, tr')).filter((row) => {
                    const cells = row.querySelectorAll('td');
                    if (!cells.length) return false;
                    const text = normalized(row.textContent || '');
                    return text && !text.includes('cód. práctica');
                  });
                  return rows.length;
                }
                """
            )
            return int(count or 0)
        except Exception:
            return -1

    async def _clear_prescribed_practices(self, page: Page) -> None:
        try:
            await page.evaluate(
                """
                () => {
                  const normalized = (value) => String(value || '').toLowerCase().replace(/\\s+/g, ' ').trim();
                  const tables = Array.from(document.querySelectorAll('table'));
                  const targetTable = tables.find((table) => {
                    const text = normalized(table.textContent || '');
                    return text.includes('cód. práctica') && text.includes('acciones');
                  });
                  const practiceRows = targetTable
                    ? Array.from(targetTable.querySelectorAll('tbody tr, tr')).filter((row) => {
                        const cells = row.querySelectorAll('td');
                        if (!cells.length) return false;
                        const text = normalized(row.textContent || '');
                        return text && !text.includes('cód. práctica');
                      })
                    : [];
                  for (const row of practiceRows) {
                    const actionCell = row.querySelector('td:last-child');
                    const candidates = actionCell
                      ? Array.from(actionCell.querySelectorAll('button, a, span, i'))
                      : [];
                    const deleteCandidate =
                      candidates.find((el) => /trash|remove|times|eliminar|borrar|anular/i.test(String(el.className || '') + ' ' + String(el.getAttribute?.('title') || ''))) ||
                      candidates[candidates.length - 1];
                    const button = deleteCandidate?.closest?.('button, a, span') || deleteCandidate;
                    try { button?.click(); } catch (e) {}
                  }
                  const practicaInput = document.querySelector('#pe-practica');
                  if (practicaInput) {
                    practicaInput.value = '';
                    practicaInput.dispatchEvent(new Event('input', { bubbles: true }));
                    practicaInput.dispatchEvent(new Event('change', { bubbles: true }));
                  }
                }
                """
            )
            await asyncio.sleep(0.8)
            await self._wait_for_busy_overlay_to_clear(page)
        except Exception:
            pass


    async def _wait_for_busy_overlay_to_clear(self, page: Page, timeout_ms: int = 6000) -> None:
        deadline = asyncio.get_running_loop().time() + (timeout_ms / 1000)
        while asyncio.get_running_loop().time() < deadline:
            try:
                busy = await page.evaluate(
                    """
                    () => {
                      const visible = (el) => {
                        if (!el) return false;
                        const style = window.getComputedStyle(el);
                        const rect = el.getBoundingClientRect();
                        return style.display !== 'none' && style.visibility !== 'hidden' && rect.width > 40 && rect.height > 20;
                      };
                      const texts = Array.from(document.querySelectorAll('div, span'))
                        .filter(visible)
                        .map((el) => String(el.innerText || el.textContent || '').trim().toLowerCase())
                        .filter(Boolean);
                      return texts.some((text) =>
                        text.includes('buscando afiliado') ||
                        text.includes('buscando afiliado/s') ||
                        text.includes('procesando')
                      );
                    }
                    """
                )
                if not busy:
                    return
            except Exception:
                return
            await asyncio.sleep(0.2)

    def _summarize_text(self, text: str, limit: int = 180) -> str:
        normalized = " ".join(str(text or "").split())
        if len(normalized) <= limit:
            return normalized
        return normalized[: limit - 3] + "..."

    async def _click_dropdown_item(self, page: Page, matcher_script: str, value: str) -> None:
        option_text = await page.evaluate(
            f"""
            (wanted) => {{
                const candidates = Array.from(document.querySelectorAll(
                    '.ui-menu-item, .ui-menu-item-wrapper, .dropdown-item, .dropdown-menu li, .dropdown-menu li a, li[role="option"], .tt-suggestion, a.dropdown-item'
                ))
                    .filter((node) => node.offsetParent !== null)
                    .map((node) => (node.textContent || '').trim())
                    .filter(Boolean);
                const resolver = {matcher_script};
                return resolver(candidates, wanted);
            }}
            """,
            value,
        )
        if not option_text:
            raise RuntimeError(f"No se encontro una opcion del dropdown para '{value}'.")

        clicked = await page.evaluate(
            """
            (wanted) => {
                const nodes = Array.from(document.querySelectorAll(
                    '.ui-menu-item, .ui-menu-item-wrapper, .dropdown-item, .dropdown-menu li, .dropdown-menu li a, li[role="option"], .tt-suggestion, a.dropdown-item'
                )).filter((node) => node.offsetParent !== null);
                const target = nodes.find((node) => (node.textContent || '').trim() === wanted);
                if (!target) return false;
                target.click();
                return true;
            }
            """,
            option_text,
        )
        if not clicked:
            raise RuntimeError(f"No se pudo hacer click en la opcion '{option_text}'.")
        await human_pause(0.5, 0.9)

    async def _click_confirm_in_modal(self, page: Page) -> None:
        confirm_button = page.locator('#modal-generar-orden button:has-text("Confirmar")').first
        await confirm_button.wait_for(state="visible")
        await human_pause(0.4, 0.8)
        await confirm_button.click()

    async def _wait_for_generation_outcome(self, page: Page, practica: str = "427122", timeout_ms: int = 5000) -> str:
        deadline = asyncio.get_running_loop().time() + (timeout_ms / 1000)
        while asyncio.get_running_loop().time() < deadline:
            try:
                alert = page.locator(".alert.rin-alert.alert-danger").last
                if await alert.count() > 0 and await alert.is_visible():
                    text = sanitize_text(await alert.inner_text())
                    if text and self._should_retry_with_capita(text, practica):
                        return "capita"
                feedback_text = await self._read_visible_feedback_text(page)
                if feedback_text and self._should_retry_with_capita(feedback_text, practica):
                    return "capita"
            except Exception:
                pass

            try:
                confirm_button = page.locator(
                    "#modal-generar-orden button:has-text('Confirmar'), .modal.show button:has-text('Confirmar')"
                ).first
                if await confirm_button.count() > 0 and await confirm_button.is_visible():
                    return "confirm"
            except Exception:
                pass

            try:
                detail_button = page.locator(
                    "button:has-text('Detalle'), a:has-text('Detalle'), button[data-n_orden]"
                ).first
                if await detail_button.count() > 0 and await detail_button.is_visible():
                    return "detail"
            except Exception:
                pass

            try:
                body_text = sanitize_text((await page.locator("body").inner_text(timeout=1500)).lower())
                if (
                    "ya cuenta con una orden" in body_text
                    or "se ha generado una orden" in body_text
                    or ("alcanz" in body_text and "cantidad mensual" in body_text)
                    or ("alcanz" in body_text and "cantidad anual" in body_text)
                ):
                    return "body"
            except Exception:
                pass

            await asyncio.sleep(0.25)

        raise RuntimeError("No se pudo confirmar el resultado luego de pulsar Generar.")

    async def _extract_ome_number_from_detail(self, page: Page) -> str:
        fallback = ""
        try:
            fallback = sanitize_text(await page.locator("button[data-n_orden]").first.get_attribute("data-n_orden") or "")
        except Exception:
            fallback = ""

        try:
            detail_button = page.locator(
                "button:has-text('Detalle'), a:has-text('Detalle'), button[data-n_orden]"
            ).first
            if await detail_button.count() == 0:
                return sanitize_text(fallback or await self._extract_order_number_from_visible_content(page))

            await detail_button.click()
            await human_pause(0.9, 1.3)

            await human_pause(0.3, 0.5)
            nro_ome = await self._extract_order_number_from_visible_content(page)

            try:
                modal = page.locator(".modal.show, .modal.in, .modal[style*='display: block']").last
                close_button = modal.locator("button:has-text('Cerrar'), .close, button[data-dismiss='modal']").first
                if await close_button.count() > 0:
                    await close_button.click()
                    await human_pause(0.2, 0.4)
            except Exception:
                pass

            return sanitize_text(nro_ome) or fallback
        except Exception:
            return sanitize_text(fallback or await self._extract_order_number_from_visible_content(page))

    async def _extract_order_number_from_visible_content(self, page: Page) -> str:
        try:
            return sanitize_text(
                await page.evaluate(
                    """
                    () => {
                      const clean = (value) => String(value || '').trim();
                      const visible = (el) => {
                        if (!el) return false;
                        const style = window.getComputedStyle(el);
                        const rect = el.getBoundingClientRect();
                        return style.display !== 'none' && style.visibility !== 'hidden' && rect.width > 10 && rect.height > 10;
                      };
                      const pickNumber = (text) => {
                        const match = clean(text).match(/\\b\\d{8,}\\b/);
                        return match ? match[0] : '';
                      };

                      for (const node of document.querySelectorAll('[data-n_orden], [data-orden]')) {
                        const value = clean(node.getAttribute('data-n_orden') || node.getAttribute('data-orden'));
                        if (/^\\d{8,}$/.test(value)) return value;
                      }

                      const scopes = Array.from(
                        document.querySelectorAll('.modal.show, .modal.in, .modal[style*="display: block"], .alert, [role="dialog"], .swal2-popup, .sweet-alert')
                      ).filter(visible);

                      for (const scope of scopes) {
                        for (const node of scope.querySelectorAll('input, textarea, [title], [data-original-title], [data-n_orden], [data-orden], button, a, span, div, td, strong, b')) {
                          const value =
                            clean(node.value) ||
                            clean(node.getAttribute('data-n_orden')) ||
                            clean(node.getAttribute('data-orden')) ||
                            clean(node.getAttribute('title')) ||
                            clean(node.getAttribute('data-original-title')) ||
                            clean(node.textContent);
                          const picked = pickNumber(value);
                          if (picked) return picked;
                        }
                        const picked = pickNumber(scope.textContent);
                        if (picked) return picked;
                      }

                      return pickNumber(document.body ? document.body.innerText : '');
                    }
                    """
                )
            )
        except Exception:
            return ""

    async def _wait_for_patient_name(self, page: Page, timeout_ms: int = 5000, previous_value: str = "") -> str:
        selectors = [
            "#pe-nomyap_afiliado",
            "#pe-nombre-afiliado",
            "#pe-afiliado-nombre",
            "#pe_nombre_afiliado",
        ]
        previous_normalized = sanitize_text(previous_value)
        deadline = asyncio.get_running_loop().time() + (timeout_ms / 1000)
        while asyncio.get_running_loop().time() < deadline:
            for selector in selectors:
                locator = page.locator(selector).first
                if await locator.count() == 0:
                    continue
                try:
                    if await locator.evaluate("el => el.tagName.toLowerCase() === 'input'"):
                        value = sanitize_text(await locator.input_value())
                    else:
                        value = sanitize_text(await locator.inner_text())
                    if value and value != previous_normalized:
                        return value
                except Error:
                    continue
            await asyncio.sleep(0.25)
        return ""

    async def _close_feedback_modal(self, page: Page) -> None:
        try:
            modal = page.locator(".modal.show, .modal.in, .modal[style*='display: block']").last
            if await modal.count() == 0:
                return
            close_button = modal.locator("button:has-text('Cerrar'), .close, button[data-dismiss='modal'], [aria-label='Close']").first
            if await close_button.count() > 0:
                await close_button.click()
                await asyncio.sleep(0.4)
        except Exception:
            pass
        await self._cleanup_modals(page)

    async def _close_capita_alert(self, page: Page) -> None:
        try:
            alert = page.locator(".alert.rin-alert.alert-danger").last
            if await alert.count() > 0 and await alert.is_visible():
                for selector in (".close", "[data-dismiss='modal']", "[aria-label='Close']", "button", "a", "span", "i"):
                    try:
                        button = alert.locator(selector).first
                        if await button.count() > 0 and await button.is_visible():
                            await button.click()
                            await asyncio.sleep(0.25)
                            break
                    except Exception:
                        continue
            await page.evaluate(
                """
                () => {
                  const normalized = (value) =>
                    String(value || '')
                      .normalize('NFD')
                      .replace(/[\\u0300-\\u036f]/g, '')
                      .toLowerCase()
                      .replace(/\\s+/g, ' ')
                      .trim();
                  const matchesText = (el) => {
                    const text = normalized(el?.innerText || el?.textContent || '');
                    return text.includes('no se puede prescribir la practica') && text.includes('forma parte de su capita');
                  };
                  const candidates = Array.from(document.querySelectorAll('.modal, .alert, .ui-pnotify, .sweet-alert, .swal2-container, .swal2-popup, div'));
                  for (const node of candidates) {
                    if (!matchesText(node)) continue;
                    const closer = node.querySelector(
                      "button, .close, [data-dismiss='modal'], [aria-label='Close'], .ui-pnotify-closer, .swal2-close, a"
                    );
                    if (closer) {
                      closer.click();
                    }
                    node.remove();
                  }
                  document.querySelectorAll('.modal-backdrop, .swal2-container').forEach((node) => node.remove());
                  document.body.classList.remove('modal-open');
                  document.body.style.removeProperty('padding-right');
                }
                """
            )
            await asyncio.sleep(0.3)
        except Exception:
            pass
        await self._cleanup_modals(page)

    async def _save_error_screenshot(self, page: Page, afiliado: str, reason: str) -> Path:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_afiliado = "".join(ch for ch in afiliado if ch.isdigit()) or "sin_afiliado"
        path = get_errors_dir() / f"ome_{safe_afiliado}_{reason}_{timestamp}.png"
        await page.screenshot(path=str(path), full_page=True)
        return path


def normalize_header(value: str) -> str:
    return sanitize_text(value).lower().replace("_", "").replace("-", "").replace(" ", "")


def detect_input_columns(headers: list[str]) -> dict[str, int]:
    normalized = [normalize_header(header) for header in headers]
    indices: dict[str, int] = {}

    for target in ("afiliado", "diagnostico", "practica"):
        alias_set = {normalize_header(alias) for alias in INPUT_HEADER_ALIASES[target]}
        for index, header in enumerate(normalized):
            if header in alias_set:
                indices[target] = index
                break

    missing = [key for key in ("afiliado", "diagnostico", "practica") if key not in indices]
    if missing:
        raise ValueError(f"Faltan columnas requeridas: {', '.join(missing)}")

    alias_set_modo = {normalize_header(alias) for alias in INPUT_HEADER_ALIASES["modo"]}
    for index, header in enumerate(normalized):
        if header in alias_set_modo:
            indices["modo"] = index
            break

    alias_set_nombre = {normalize_header(alias) for alias in INPUT_HEADER_ALIASES["nombre"]}
    for index, header in enumerate(normalized):
        if header in alias_set_nombre:
            indices["nombre"] = index
            break

    for target in ("beneficio", "dni"):
        alias_set = {normalize_header(alias) for alias in INPUT_HEADER_ALIASES[target]}
        for index, header in enumerate(normalized):
            if header in alias_set:
                indices[target] = index
                break

    alias_set_completar = {normalize_header(alias) for alias in INPUT_HEADER_ALIASES["completar_benef"]}
    for index, header in enumerate(normalized):
        if header in alias_set_completar:
            indices["completar_benef"] = index
            break

    alias_set_completar_dni = {normalize_header(alias) for alias in INPUT_HEADER_ALIASES["completar_dni"]}
    for index, header in enumerate(normalized):
        if header in alias_set_completar_dni:
            indices["completar_dni"] = index
            break

    return indices


def _truthy_cell(value: object) -> bool:
    normalized = sanitize_text(str(value or "")).lower()
    return normalized in {"1", "si", "sí", "true", "x", "yes", "y"}


def parse_input_excel(path: Path) -> list[PatientInput]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    rows = list(workbook.active.iter_rows(values_only=True))
    if not rows:
        raise ValueError("El archivo Excel esta vacio.")

    headers = [str(value or "").strip() for value in rows[0]]
    indices = detect_input_columns(headers)
    records: list[PatientInput] = []

    for row_index, row in enumerate(rows[1:], start=2):
        modo = sanitize_text(
            str(row[indices["modo"]]) if "modo" in indices and indices["modo"] < len(row) and row[indices["modo"]] is not None else "BENEF"
        ).upper() or "BENEF"
        afiliado = sanitize_text(
            str(row[indices["afiliado"]]) if indices["afiliado"] < len(row) and row[indices["afiliado"]] is not None else ""
        )
        diagnostico = sanitize_text(
            str(row[indices["diagnostico"]]) if indices["diagnostico"] < len(row) and row[indices["diagnostico"]] is not None else ""
        )
        practica = sanitize_text(
            str(row[indices["practica"]]) if indices["practica"] < len(row) and row[indices["practica"]] is not None else ""
        )
        if not afiliado:
            continue
        completar_benef = _truthy_cell(
            row[indices["completar_benef"]]
            if "completar_benef" in indices and indices["completar_benef"] < len(row)
            else ""
        )
        completar_dni = _truthy_cell(
            row[indices["completar_dni"]]
            if "completar_dni" in indices and indices["completar_dni"] < len(row)
            else ""
        )
        if not (completar_benef or completar_dni) and (not diagnostico or not practica):
            raise ValueError(f"Fila {row_index}: diagnostico y practica son obligatorios.")
        nombre = sanitize_text(
            str(row[indices["nombre"]]) if "nombre" in indices and indices["nombre"] < len(row) and row[indices["nombre"]] is not None else ""
        )
        beneficio = sanitize_text(
            str(row[indices["beneficio"]]) if "beneficio" in indices and indices["beneficio"] < len(row) and row[indices["beneficio"]] is not None else ""
        )
        dni = sanitize_text(
            str(row[indices["dni"]]) if "dni" in indices and indices["dni"] < len(row) and row[indices["dni"]] is not None else ""
        )

        records.append(
            PatientInput(
                modo=modo,
                afiliado=afiliado,
                diagnostico=diagnostico,
                practica=practica,
                beneficio=beneficio,
                dni=dni,
                nombre=nombre,
                completar_benef=completar_benef,
                completar_dni=completar_dni,
            )
        )

    return records


def parse_input_csv(csv_path: Path) -> list[PatientInput]:
    with csv_path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        if not reader.fieldnames:
            raise ValueError("El archivo de entrada no tiene encabezados.")

        fieldnames = [str(field or "").strip() for field in reader.fieldnames]
        indices = detect_input_columns(fieldnames)
        rows: list[PatientInput] = []

        for row_index, row in enumerate(reader, start=2):
            values = [row.get(name, "") for name in fieldnames]
            modo = sanitize_text(values[indices["modo"]] if "modo" in indices and indices["modo"] < len(values) else "BENEF").upper() or "BENEF"
            afiliado = sanitize_text(values[indices["afiliado"]] if indices["afiliado"] < len(values) else "")
            diagnostico = sanitize_text(values[indices["diagnostico"]] if indices["diagnostico"] < len(values) else "")
            practica = sanitize_text(values[indices["practica"]] if indices["practica"] < len(values) else "")
            if not afiliado:
                continue
            completar_benef = _truthy_cell(
                values[indices["completar_benef"]]
                if "completar_benef" in indices and indices["completar_benef"] < len(values)
                else ""
            )
            completar_dni = _truthy_cell(
                values[indices["completar_dni"]]
                if "completar_dni" in indices and indices["completar_dni"] < len(values)
                else ""
            )
            if not (completar_benef or completar_dni) and (not diagnostico or not practica):
                raise ValueError(f"Fila {row_index}: diagnostico y practica son obligatorios.")
            nombre = sanitize_text(values[indices["nombre"]] if "nombre" in indices and indices["nombre"] < len(values) else "")
            beneficio = sanitize_text(values[indices["beneficio"]] if "beneficio" in indices and indices["beneficio"] < len(values) else "")
            dni = sanitize_text(values[indices["dni"]] if "dni" in indices and indices["dni"] < len(values) else "")
            rows.append(
                PatientInput(
                    modo=modo,
                    afiliado=afiliado,
                    diagnostico=diagnostico,
                    practica=practica,
                    beneficio=beneficio,
                    dni=dni,
                    nombre=nombre,
                    completar_benef=completar_benef,
                    completar_dni=completar_dni,
                )
            )

        return rows


def parse_input_file(path: Path) -> list[PatientInput]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return parse_input_csv(path)
    if suffix in {".xlsx", ".xlsm"}:
        return parse_input_excel(path)
    raise ValueError("Formato no soportado. Usa CSV o Excel .xlsx/.xlsm.")


def build_default_output_path(input_path: Path) -> Path:
    return get_output_dir() / f"{input_path.stem}_ome_resultados.csv"


def build_default_report_path(input_path: Path) -> Path:
    return get_output_dir() / f"{input_path.stem}_ome_resultados.xlsx"


def _visible_result_value(resultado: str) -> str:
    normalized = sanitize_text(resultado).upper()
    if normalized.startswith("ERROR"):
        return "ERROR"
    if normalized == "NO_DNI":
        return "NO DNI"
    if normalized == "NO_PAMI":
        return "NO PAMI"
    if normalized == "DOBLE_DNI":
        return "DOBLE DNI"
    if normalized == "YA_TIENE_OME":
        return "YA TIENE OME"
    if normalized == "LIMITE_ANUAL":
        return "LIMITE ANUAL"
    if normalized == "VERIFICAR_CREDENCIAL":
        return "Verificar credencial"
    return normalized


def _visible_nro_ome_value(resultado: str, nro_ome: str) -> str:
    visible_result = _visible_result_value(resultado)
    normalized_nro_ome = sanitize_text(nro_ome)
    if normalized_nro_ome:
        return normalized_nro_ome
    if visible_result == "YA TIENE OME":
        return visible_result
    if visible_result in {"GENERADA", "YA TIENE OME", "LIMITE", "LIMITE ANUAL", "BAJA", "ERROR", "NO DNI", "NO PAMI", "DOBLE DNI", "Verificar credencial"}:
        return visible_result
    return ""


def export_results_csv_to_excel(csv_path: Path, excel_path: Path) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "OME"

    sheet.append(["fila", "modo", "afiliado", "beneficio", "dni", "nombre", "diagnostico", "practica", "resultado", "nro_ome"])

    with csv_path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            sheet.append(
                [
                    sanitize_text(row.get("sheet_row")),
                    sanitize_text(row.get("modo")),
                    sanitize_text(row.get("afiliado")),
                    sanitize_text(row.get("beneficio")),
                    sanitize_text(row.get("dni")),
                    sanitize_text(row.get("nombre")),
                    sanitize_text(row.get("diagnostico")),
                    sanitize_text(row.get("practica")),
                    _visible_result_value(row.get("resultado", "")),
                    _visible_nro_ome_value(row.get("resultado", ""), row.get("nro_ome", "")),
                ]
            )

    widths = {
        "A": 10,
        "B": 12,
        "C": 18,
        "D": 20,
        "E": 16,
        "F": 34,
        "G": 16,
        "H": 18,
        "I": 18,
        "J": 16,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    excel_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(excel_path)
    return excel_path


async def run_batch(
    *,
    input_path: Path,
    output_path: Path,
    user: str,
    password: str,
    dry_run: bool,
    headless: bool,
    log_callback: Callable[[str], None] | None = None,
    progress_callback: Callable[[dict], None] | None = None,
    stop_requested: Callable[[], bool] | None = None,
) -> dict[str, int | str]:
    logger = log_callback or log
    patients = parse_input_file(input_path)
    store = OutputStore(output_path)
    pending = [patient for patient in patients if not store.is_processed(patient.modo, patient.afiliado)]

    logger(f"Entrada leida: {len(patients)} fila(s)")
    logger(f"Ya resueltas en salida previa: {len(patients) - len(pending)}")
    logger(f"Pendientes para esta corrida: {len(pending)}")

    if dry_run:
        logger("Modo dry-run activo: no se ejecuta ninguna accion sobre PAMI.")
        result = RunSummary().as_dict()
        result["output_path"] = str(output_path)
        return result

    if not user or not password:
        raise ValueError("Debes indicar usuario y clave para ejecutar el lote OME.")

    summary = RunSummary()
    async with PamiOmeGenerator(user=user, password=password, headless=headless) as generator:
        total = len(pending)
        for index, patient in enumerate(pending, start=1):
            if stop_requested and stop_requested():
                logger("Detencion solicitada. Se corta el lote antes de iniciar la siguiente fila.")
                break
            logger(f"[{index}/{total}] Procesando {patient.modo} {patient.afiliado}")
            result = await generator.process_patient(patient)
            store.write(result)
            summary.register(result)
            progress = {
                "current": index,
                "total": total,
                "modo": patient.modo,
                "afiliado": patient.afiliado,
                "nombre": result.nombre,
                "resultado": result.resultado,
                "nro_ome": result.nro_ome,
            }
            if progress_callback:
                progress_callback(progress)
            name_piece = f"{result.nombre} -> " if result.nombre else ""
            ome_piece = f" | OME {result.nro_ome}" if result.nro_ome else ""
            logger(f"[{index}/{total}] {patient.modo} {name_piece}{patient.afiliado} | {result.resultado}{ome_piece}")
            if result.resultado == "ERROR_DNI_MODAL":
                logger(
                    f"[{index}/{total}] Lote interrumpido: no se pudo abrir el modal de DNI para {patient.afiliado}."
                )
                break

    logger("")
    logger("Resumen final")
    logger(f"Total procesados: {summary.total}")
    logger(f"OMEs OK: {summary.ok}")
    logger(f"Ya existentes: {summary.generadas}")
    logger(f"Con limite mensual: {summary.limite}")
    logger(f"Bajas: {summary.bajas}")
    logger(f"Doble DNI: {summary.doble_dni}")
    logger(f"BENEF completados: {summary.benef_completados}")
    logger(f"DNI completados: {summary.dni_completados}")
    logger(f"Errores: {summary.errores}")
    logger(f"CSV de salida: {output_path}")

    result = summary.as_dict()
    result["output_path"] = str(output_path)
    return result


def run_batch_sync(
    *,
    input_path: Path,
    output_path: Path,
    user: str,
    password: str,
    dry_run: bool = False,
    headless: bool = False,
    log_callback: Callable[[str], None] | None = None,
    progress_callback: Callable[[dict], None] | None = None,
    stop_requested: Callable[[], bool] | None = None,
) -> dict[str, int | str]:
    try:
        asyncio.get_running_loop()
    except RuntimeError:
        return asyncio.run(
            run_batch(
                input_path=input_path,
                output_path=output_path,
                user=user,
                password=password,
                dry_run=dry_run,
                headless=headless,
                log_callback=log_callback,
                progress_callback=progress_callback,
                stop_requested=stop_requested,
            )
        )

    result_holder: dict[str, int | str] = {}
    error_holder: list[BaseException] = []
    done = threading.Event()

    def _runner() -> None:
        try:
            result_holder.update(
                asyncio.run(
                    run_batch(
                        input_path=input_path,
                        output_path=output_path,
                        user=user,
                        password=password,
                        dry_run=dry_run,
                        headless=headless,
                        log_callback=log_callback,
                        progress_callback=progress_callback,
                        stop_requested=stop_requested,
                    )
                )
            )
        except BaseException as exc:
            error_holder.append(exc)
        finally:
            done.set()

    worker = threading.Thread(target=_runner, name="ome-batch-runner", daemon=True)
    worker.start()
    done.wait()

    if error_holder:
        raise error_holder[0]

    return result_holder


def parse_args(argv: Iterable[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Automatiza la generacion de OMEs en PAMI con Playwright.")
    parser.add_argument("input_file", help="Ruta al CSV o Excel con columnas modo, afiliado, diagnostico y practica.")
    parser.add_argument(
        "--output",
        dest="output_csv",
        help="Ruta al CSV de salida. Si no se indica, se genera en la carpeta salidas/ del proyecto.",
    )
    parser.add_argument(
        "--env-file",
        default=".env",
        help="Ruta al archivo .env con PAMI_USER y PAMI_PASS. Default: .env",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Valida el archivo de entrada y muestra el plan de trabajo sin abrir el navegador.",
    )
    parser.add_argument(
        "--headless",
        action="store_true",
        help="Permite ejecutar sin navegador visible. Por defecto se abre en modo visible.",
    )
    return parser.parse_args(argv)


def main(argv: Iterable[str] | None = None) -> None:
    args = parse_args(argv)
    input_path = Path(args.input_file).expanduser().resolve()
    output_path = Path(args.output_csv).expanduser().resolve() if args.output_csv else build_default_output_path(input_path)
    env_path = Path(args.env_file).expanduser().resolve()

    env_values = load_env_file(env_path)
    user = env_values.get("PAMI_USER", "").strip()
    password = env_values.get("PAMI_PASS", "").strip()
    if not user or not password:
        raise ValueError("El archivo .env debe contener PAMI_USER y PAMI_PASS.")

    run_batch_sync(
        input_path=input_path,
        output_path=output_path,
        user=user,
        password=password,
        dry_run=args.dry_run,
        headless=args.headless,
    )


if __name__ == "__main__":
    main()
