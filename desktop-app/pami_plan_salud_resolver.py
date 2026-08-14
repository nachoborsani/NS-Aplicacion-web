from __future__ import annotations

import re
import unicodedata
from dataclasses import dataclass
from functools import lru_cache
from pathlib import Path

from openpyxl import load_workbook

from app_paths import get_resource_path


DEFAULT_NOMENCLADOR_FILENAME = "nomenclador_pami.xlsx"
LEGACY_NOMENCLADOR_FILENAME = "73-Nomenclador Pami valorizado 04-2026.xlsx"

DEFAULT_NOMENCLADOR_PATH = get_resource_path(DEFAULT_NOMENCLADOR_FILENAME)
LEGACY_NOMENCLADOR_PATH = get_resource_path(LEGACY_NOMENCLADOR_FILENAME)
DESKTOP_NOMENCLADOR_PATH = Path.home() / "Desktop" / DEFAULT_NOMENCLADOR_FILENAME


PRACTICE_CODE_HINTS = {
    "alergia": "820111",
    "cardiologia": "570129",
    "dermatologia": "820116",
    "diabetologia": "820171",
    "doppler cardiaco": "180301",
    "doppler": "180301",
    "ecocardiograma doppler": "180301",
    "ecodoppler cardiaco": "180301",
    "ecografia adbdominal": "180112",
    "ecografia abdominal": "180112",
    "ecografia abdominal completa": "180112",
    "ecografia ginecologica": "180104",
    "ecografia de partes blandas": "186001",
    "ecografia mamaria": "180106",
    "ecografia partes blandas": "186001",
    "ecografia prostatica": "180114",
    "ecografia renal": "180116",
    "ecografia tocoginecologica": "180104",
    "ecografia transvaginal": "180128",
    "ecografia trasnvaginal": "180128",
    "ecografia vesical": "180123",
    "endocrinologia": "820118",
    "ergometria": "570124",
    "espirometria": "687114",
    "flebologia": "820143",
    "gastroenterologia": "820139",
    "gatroenterologia": "820139",
    "ginecologia": "820145",
    "hematologia": "820121",
    "holter": "570121",
    "m a p a": "570120",
    "mamografia": "340622",
    "mapa": "570120",
    "nefrologia": "820155",
    "neumonologia": "820157",
    "otorrinolaringologia": "820168",
    "presurometria": "570120",
    "prostatica": "180114",
    "radiografia de torax": "340301",
    "reumatologia": "820163",
    "rpm": "180123",
    "rx torax": "340301",
    "traumatologia y ortopedia": "820165",
    "traumatologia": "820165",
    "urologia": "820167",
    "vesical": "180123",
}

MULTI_PRACTICE_SEPARATORS = (",", " + ")
NOISE_TOKENS = (
    "plan de salud",
    "plan salud",
    "pami",
    "consulta",
    "medico",
    "especialista",
    "control",
    "primera vez",
    "hacer alguna",
    "1 x",
)


@dataclass(frozen=True)
class PracticeCatalogItem:
    module_id: str
    module_name: str
    code: str
    description: str


def normalize_code(value: str) -> str:
    text = (value or "").strip()
    if not text:
        return ""
    digit_tokens = re.findall(r"\d+", text)
    if digit_tokens:
        longest = max(digit_tokens, key=len)
        if len(longest) >= 6:
            return longest[:6]
    if " - " in text:
        return text.split(" - ", 1)[0].strip().upper()
    return text.split()[0].strip().upper()


def normalize_search_text(value: str) -> str:
    text = unicodedata.normalize("NFKD", value or "")
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    lowered = text.lower()
    for token in ("(", ")", "/", "\\", ";", "-", "_", ".", ":"):
        lowered = lowered.replace(token, " ")
    return " ".join(lowered.split())


def _default_nomenclador_candidates() -> list[Path]:
    candidates = [
        DEFAULT_NOMENCLADOR_PATH,
        DESKTOP_NOMENCLADOR_PATH,
        LEGACY_NOMENCLADOR_PATH,
    ]
    seen: set[str] = set()
    result: list[Path] = []
    for path in candidates:
        resolved = str(path)
        if resolved in seen:
            continue
        seen.add(resolved)
        if path.exists():
            result.append(path)
    return result


def _build_items_from_rows(rows) -> tuple[PracticeCatalogItem, ...]:
    """Arma los items desde filas posicionales (row[0..4]). Sirve tanto para el
    Excel como para el nomenclador que baja la web (mismo formato de fila)."""
    items: list[PracticeCatalogItem] = []
    seen: set[tuple[str, str, str]] = set()
    for row in rows:
        if not row or len(row) < 4:
            continue
        module_code = str(row[0] or "").strip()
        module_name = " ".join(str(row[1] or "").split())
        practice_code = normalize_code(str(row[2] or ""))
        descriptions = [
            " ".join(str(row[index] or "").split())
            for index in (3, 4)
            if index < len(row)
        ]
        description = next((value for value in descriptions if value and normalize_code(value) != practice_code), "")
        if not module_code or not module_name or not practice_code or not description:
            continue
        module_id = f"{module_code}::{module_name}"
        key = (module_id, practice_code, description)
        if key in seen:
            continue
        seen.add(key)
        items.append(
            PracticeCatalogItem(
                module_id=module_id,
                module_name=module_name,
                code=practice_code,
                description=description,
            )
        )
    return tuple(items)


@lru_cache(maxsize=4)
def _read_catalog(catalog_path: str) -> tuple[PracticeCatalogItem, ...]:
    workbook = load_workbook(catalog_path, read_only=True, data_only=True)
    if "Nomenclador" not in workbook.sheetnames:
        return ()
    ws = workbook["Nomenclador"]
    return _build_items_from_rows(ws.iter_rows(min_row=11, values_only=True))


def _web_catalog_items() -> tuple[PracticeCatalogItem, ...]:
    """Nomenclador del mes de trabajo bajado de la web (fuente principal)."""
    try:
        import ns_catalog

        rows = ns_catalog.catalog_rows()
    except Exception:
        return ()
    return _build_items_from_rows(rows) if rows else ()


def _catalog_items() -> tuple[PracticeCatalogItem, ...]:
    # Opción A: la web es la fuente. El Excel local queda solo de respaldo por si
    # todavía no se descargó ningún mes en esta PC.
    web = _web_catalog_items()
    if web:
        return web
    for path in _default_nomenclador_candidates():
        items = _read_catalog(str(path))
        if items:
            return items
    return ()


def _hint_codes(normalized_text: str) -> set[str]:
    codes: set[str] = set()
    for alias, code in sorted(PRACTICE_CODE_HINTS.items(), key=lambda item: len(item[0]), reverse=True):
        if alias in normalized_text:
            codes.add(code)
    return codes


def _clean_hint(normalized_text: str) -> str:
    hint = normalized_text
    for token in NOISE_TOKENS:
        hint = hint.replace(token, " ")
    return " ".join(hint.split())


def _is_multi_practice(raw_text: str, codes: set[str]) -> bool:
    if len(codes) > 1:
        return True
    normalized = normalize_search_text(raw_text)
    if normalized.count("1 x") > 1:
        return True
    return any(separator in raw_text for separator in MULTI_PRACTICE_SEPARATORS) and len(codes) != 1


def resolve_plan_salud_practice(raw_text: str) -> str:
    text = (raw_text or "").strip()
    if not text:
        return ""
    code = normalize_code(text)
    if code.isdigit() and len(code) >= 6:
        return code

    normalized = normalize_search_text(text)
    codes = _hint_codes(normalized)
    if _is_multi_practice(text, codes):
        return ""
    if len(codes) == 1:
        return next(iter(codes))

    hint = _clean_hint(normalized)
    if not hint:
        return ""
    preferred: list[str] = []
    fallback: list[str] = []
    for item in _catalog_items():
        haystack = normalize_search_text(f"{item.module_name} {item.description}")
        if hint and hint in haystack:
            if "consulta con especialista" in normalize_search_text(item.description):
                preferred.append(item.code)
            else:
                fallback.append(item.code)
    if preferred:
        return preferred[0]
    if fallback:
        return fallback[0]
    return ""


def resolve_plan_salud_practices(raw_text: str) -> list[str]:
    text = (raw_text or "").strip()
    if not text:
        return []

    code = normalize_code(text)
    if code.isdigit() and len(code) >= 6:
        return [code]

    normalized = normalize_search_text(text)
    codes = _hint_codes(normalized)
    if not _is_multi_practice(text, codes):
        single = resolve_plan_salud_practice(text)
        return [single] if single else []

    parts = [part.strip(" ;") for part in re.split(r"\s*,\s*|\s+\+\s+", text) if part.strip(" ;")]
    resolved: list[str] = []
    seen: set[str] = set()
    for part in parts:
        code_part = resolve_plan_salud_practice(part)
        if not code_part or code_part in seen:
            continue
        resolved.append(code_part)
        seen.add(code_part)
    return resolved


def is_skippable_plan_salud_practice(raw_text: str) -> bool:
    normalized = normalize_search_text(raw_text)
    if not normalized:
        return False
    cleaned = normalized
    for token in (",", "+", "1 x", "radiografias", "radiografia", "rx", "y"):
        cleaned = cleaned.replace(token, " ")
    return not " ".join(cleaned.split())


def explain_unresolved_plan_salud_practice(raw_text: str) -> str:
    text = (raw_text or "").strip()
    if not text:
        return "sin especialidad/practica"
    normalized = normalize_search_text(text)
    codes = _hint_codes(normalized)
    if _is_multi_practice(text, codes):
        return "multiples practicas en una fila"
    return "sin codigo resoluble"
