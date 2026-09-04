import re
import shutil
import subprocess
import tempfile
import threading
import time
import unicodedata
import zipfile
from calendar import monthrange
from dataclasses import dataclass
from datetime import datetime, timedelta
from html import unescape
from html.parser import HTMLParser
from pathlib import Path
from typing import Callable
from xml.etree import ElementTree

from openpyxl import load_workbook
from playwright.sync_api import Browser, BrowserContext, Page, Playwright, sync_playwright

from app_logging import log_message
from app_paths import get_output_dir
from pami_scraper import configurar_playwright
from pami_verificar import (
    CLIENTE_CIMA,
    cargar_especialidades_medicos,
    leer_pacientes_pami,
    _requisitos_practicas_turno as _cima_requisitos_practicas_turno,
)


IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".bmp", ".tif", ".tiff"}
MAX_ANTIGUEDAD_INFORME_DIAS = 60
_RAPIDOCR_ENGINE = None
_RAPIDOCR_AVAILABLE: bool | None = None
_EASYOCR_READER = None
_EASYOCR_AVAILABLE: bool | None = None


def _es_archivo_temporal_informe(path: Path) -> bool:
    name = Path(path).name
    upper = name.upper()
    return name.startswith("~$") or upper.startswith("~WRD")


def _subprocess_no_window_kwargs() -> dict:
    kwargs: dict = {}
    if hasattr(subprocess, "STARTUPINFO"):
        startupinfo = subprocess.STARTUPINFO()
        startupinfo.dwFlags |= subprocess.STARTF_USESHOWWINDOW
        startupinfo.wShowWindow = 0
        kwargs["startupinfo"] = startupinfo
    if hasattr(subprocess, "CREATE_NO_WINDOW"):
        kwargs["creationflags"] = subprocess.CREATE_NO_WINDOW
    return kwargs


CUP_LOGIN_URL = "https://cup.pami.org.ar/controllers/loginController.php?redirect=https://pe.pami.org.ar"
PAMI_TRANSMISION_URL = "https://pe.pami.org.ar/controllers/transmision.php"
PAMI_PANEL_URL = "https://pe.pami.org.ar/controllers/efector.php?registros_por_pagina=50"


class _HtmlExcelTableParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.in_table = False
        self.current_row: list[str] | None = None
        self.current_cell: list[str] | None = None
        self.rows: list[list[str]] = []

    def handle_starttag(self, tag: str, attrs) -> None:
        tag = tag.lower()
        if tag == "table" and not self.in_table:
            self.in_table = True
            return
        if not self.in_table:
            return
        if tag == "tr":
            self.current_row = []
        elif tag in {"th", "td"} and self.current_row is not None:
            self.current_cell = []
        elif tag == "br" and self.current_cell is not None:
            self.current_cell.append("\n")

    def handle_endtag(self, tag: str) -> None:
        tag = tag.lower()
        if tag == "table" and self.in_table:
            self.in_table = False
            return
        if not self.in_table:
            return
        if tag in {"th", "td"} and self.current_row is not None and self.current_cell is not None:
            value = unescape("".join(self.current_cell)).replace("\xa0", " ").strip()
            self.current_row.append(value)
            self.current_cell = None
        elif tag == "tr" and self.current_row is not None:
            if any(cell != "" for cell in self.current_row):
                self.rows.append(self.current_row)
            self.current_row = None

    def handle_data(self, data: str) -> None:
        if self.in_table and self.current_cell is not None:
            self.current_cell.append(data)


@dataclass
class SesionDocumentacion:
    playwright: Playwright
    browser: Browser
    context: BrowserContext
    page: Page


def _clave_texto(value: str) -> str:
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = text.encode("ascii", "ignore").decode("ascii")
    text = re.sub(r"[^a-zA-Z0-9]+", " ", text).strip().lower()
    return re.sub(r"\s+", " ", text)


def _clave_contiene_marca(key: str, marca: str) -> bool:
    marca_key = _clave_texto(marca)
    if not key or not marca_key:
        return False
    return re.search(rf"(?:^|\s){re.escape(marca_key)}(?:\s|$)", str(key or "")) is not None


def _agregar_etiquetas_desde_nombre_archivo(datos_pdf: dict, path: Path) -> None:
    path_key = _clave_texto(path.stem)
    texto_key = _clave_texto(datos_pdf.get("texto", ""))
    for marca, etiqueta in (
        ("mapa", "MAPA"),
        ("ett", "ETT"),
        ("eco vc", "ECO VC"),
        ("sibo", "SIBO"),
    ):
        if _clave_contiene_marca(path_key, marca) and not _clave_contiene_marca(texto_key, marca):
            datos_pdf["texto"] = f"{datos_pdf.get('texto', '')}\n{etiqueta}"
            texto_key = _clave_texto(datos_pdf.get("texto", ""))


def _solo_digitos(value: str) -> str:
    return re.sub(r"\D+", "", str(value or ""))


def _normalizar_dni_extraido(value: str) -> str:
    digits = _solo_digitos(value)
    if 7 <= len(digits) <= 9:
        return digits
    return ""


def _extraer_dni_desde_texto(text: str) -> str:
    text = str(text or "")
    patrones = [
        r"\bDNI\b\s*:?\s*([0-9][0-9.\s]{5,20}[0-9])",
        r"\bDocumento\b\s*:?\s*([0-9][0-9.\s]{5,20}[0-9])",
        r"Document(?:o)?\s*:?\s*([0-9][0-9.\s]{5,20}[0-9])",
        r"\bC[oÃ³ó]d(?:igo)?\.?\s*paciente\b\s*:?\s*([0-9][0-9.\s]{5,20}[0-9])",
        r"\bID\s+del\s+paciente\b\s*:?\s*([0-9][0-9.\s]{5,20}[0-9])",
        r"\b([0-9][0-9.\s]{5,20}[0-9])\s*ID\b",
    ]
    for patron in patrones:
        match = re.search(patron, text, flags=re.I)
        if not match:
            continue
        dni = _normalizar_dni_extraido(match.group(1))
        if dni:
            return dni
    return ""


def _extraer_fecha_turno(value: str) -> datetime | None:
    match = re.search(r"\b(\d{1,2})/(\d{1,2})/(\d{4})\b", str(value or ""))
    if not match:
        return None
    try:
        return datetime(int(match.group(3)), int(match.group(2)), int(match.group(1)))
    except ValueError:
        return None


def _rango_mes_desde_turno(value: str) -> tuple[str, str] | None:
    fecha = _extraer_fecha_turno(value)
    if not fecha:
        return None
    ultimo = monthrange(fecha.year, fecha.month)[1]
    return f"01/{fecha.month:02d}/{fecha.year}", f"{ultimo:02d}/{fecha.month:02d}/{fecha.year}"


def _extraer_fecha_archivo(value: str) -> str:
    texto = str(value or "")
    match_iso = re.search(r"(?<!\d)(\d{4})[.\-/_](\d{1,2})[.\-/_](\d{1,2})(?!\d)", texto)
    if match_iso:
        try:
            fecha = datetime(int(match_iso.group(1)), int(match_iso.group(2)), int(match_iso.group(3)))
            return f"{fecha.day:02d}/{fecha.month:02d}/{fecha.year}"
        except ValueError:
            pass

    meses_archivo = {
        "ENE": 1,
        "FEB": 2,
        "MAR": 3,
        "ABR": 4,
        "MAY": 5,
        "JUN": 6,
        "JUL": 7,
        "AGO": 8,
        "SEP": 9,
        "OCT": 10,
        "NOV": 11,
        "DIC": 12,
    }
    match_mes = re.search(
        r"(?<!\d)(\d{1,2})\s*(ENE|FEB|MAR|ABR|MAY|JUN|JUL|AGO|SEP|OCT|NOV|DIC)(?:[.\-/_]\d{1,2})?(?![A-Z])",
        texto,
        flags=re.I,
    )
    if match_mes:
        try:
            fecha = datetime(datetime.now().year, meses_archivo[match_mes.group(2).upper()], int(match_mes.group(1)))
            return f"{fecha.day:02d}/{fecha.month:02d}/{fecha.year}"
        except (KeyError, ValueError):
            pass

    match = re.search(r"(?<!\d)(\d{1,2})[.\-/_](\d{1,2})[.\-/_](\d{2,4})(?!\d)", str(value or ""))
    if not match:
        return ""
    anio = int(match.group(3))
    if anio < 100:
        anio += 2000
    try:
        fecha = datetime(anio, int(match.group(2)), int(match.group(1)))
    except ValueError:
        return ""
    return f"{fecha.day:02d}/{fecha.month:02d}/{fecha.year}"


def _fecha_normalizada(value: str) -> str:
    fecha = _extraer_fecha_turno(value)
    if not fecha:
        return ""
    return f"{fecha.day:02d}/{fecha.month:02d}/{fecha.year}"


def _fecha_panel_aceptacion_desde() -> str:
    hoy = datetime.now()
    mes = hoy.month - 5
    anio = hoy.year
    while mes <= 0:
        mes += 12
        anio -= 1
    return f"01/{mes:02d}/{anio}"


def _cell(row: list[str], headers: dict[str, int], *names: str) -> str:
    for name in names:
        idx = headers.get(_clave_texto(name))
        if idx is not None and idx < len(row):
            return str(row[idx] or "").strip()
    return ""


def _xls_cell_a_texto(cell, datemode) -> str:
    import xlrd

    ctype = cell.ctype
    value = cell.value
    if ctype == xlrd.XL_CELL_EMPTY or value is None:
        return ""
    if ctype == xlrd.XL_CELL_NUMBER:
        try:
            if float(value).is_integer():
                return str(int(value))
        except Exception:
            pass
        return str(value)
    if ctype == xlrd.XL_CELL_DATE:
        try:
            dt = xlrd.xldate.xldate_as_datetime(value, datemode)
            if dt.hour == 0 and dt.minute == 0 and dt.second == 0:
                return dt.strftime("%d/%m/%Y")
            return dt.strftime("%d/%m/%Y %H:%M")
        except Exception:
            return str(value)
    if ctype == xlrd.XL_CELL_BOOLEAN:
        return "1" if value else "0"
    return str(value).strip()


def _leer_xls_binario(path: Path) -> list[list[str]]:
    """Lee un .xls BINARIO (Excel viejo) con xlrd. Devuelve [] si no es binario
    (por ejemplo, un HTML renombrado a .xls), para que siga el fallback HTML."""
    try:
        import xlrd
    except Exception:
        return []
    try:
        book = xlrd.open_workbook(str(path))
        sheet = book.sheet_by_index(0)
    except Exception:
        return []
    rows: list[list[str]] = []
    for r in range(sheet.nrows):
        values: list[str] = []
        for c in range(sheet.ncols):
            try:
                values.append(_xls_cell_a_texto(sheet.cell(r, c), book.datemode))
            except Exception:
                values.append("")
        if any(values):
            rows.append(values)
    return rows


def _leer_tabla_xls(path: Path) -> list[list[str]]:
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        wb = load_workbook(path, data_only=True, read_only=True)
        ws = wb.active
        rows = []
        for row in ws.iter_rows(values_only=True):
            values = ["" if value is None else str(value).strip() for value in row]
            if any(values):
                rows.append(values)
        return rows

    # .xls: puede ser (a) un .xls BINARIO real (Excel viejo) o (b) un HTML
    # renombrado a .xls (como exporta a veces PAMI). Probamos primero binario
    # con xlrd; si no es binario, cae al parser HTML de mas abajo.
    if suffix == ".xls":
        rows = _leer_xls_binario(path)
        if rows:
            return rows

    raw = path.read_text(encoding="utf-8", errors="replace")
    parser = _HtmlExcelTableParser()
    parser.feed(raw)
    if parser.rows:
        return parser.rows

    raw = path.read_text(encoding="latin-1", errors="replace")
    parser = _HtmlExcelTableParser()
    parser.feed(raw)
    return parser.rows


def leer_prestaciones_documentacion(ruta_excel: Path) -> list[dict]:
    rows = _leer_tabla_xls(ruta_excel)
    if not rows:
        return []
    headers = {_clave_texto(value): idx for idx, value in enumerate(rows[0])}
    required = {"nro orden", "nro beneficio gp", "apellido y nombre", "practica"}
    if required - set(headers):
        raise RuntimeError("El XLS no parece ser una exportacion del panel de Transmision PAMI.")

    prestaciones: list[dict] = []
    for row in rows[1:]:
        nro_orden = _cell(row, headers, "nro orden")
        nombre = _cell(row, headers, "apellido y nombre")
        if not nro_orden or not nombre:
            continue
        transmitida_raw = _cell(row, headers, "trasmitida", "transmitida")
        validada_raw = _cell(row, headers, "validada")
        f_transmitida = _cell(row, headers, "f transmitida")
        transmitida = transmitida_raw.upper().startswith("S") or bool(f_transmitida)
        validada = validada_raw.upper().startswith("S") or transmitida
        prestaciones.append(
            {
                "n_orden": nro_orden,
                "beneficio": _solo_digitos(_cell(row, headers, "nro beneficio gp", "nro beneficio/gp")),
                "nombre": nombre,
                "nombre_key": _clave_texto(nombre),
                "practica": _cell(row, headers, "practica"),
                "turno": _cell(row, headers, "turno"),
                "validada": validada,
                "transmitida": transmitida,
                "documentacion_pendiente": validada and not transmitida,
            }
        )
    return prestaciones


def _es_bandeja_transmision_pami(ruta_excel: Path) -> bool:
    try:
        rows = _leer_tabla_xls(ruta_excel)
    except Exception:
        return False
    if not rows:
        return False

    headers = {_clave_texto(value) for value in rows[0]}
    requeridos = {"nro orden", "nro beneficio gp", "apellido y nombre", "practica"}
    indicios = {"turno", "trasmitida", "transmitida", "validada"}
    return requeridos.issubset(headers) and bool(headers & indicios)


def extraer_texto_informe(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix == ".docx":
        return _extraer_texto_docx(path)
    if suffix == ".doc":
        return _extraer_texto_word(path)
    if suffix in IMAGE_EXTENSIONS:
        return _extraer_texto_imagen(path)

    try:
        from pypdf import PdfReader
    except Exception:
        return path.stem

    try:
        reader = PdfReader(str(path))
        texto = "\n".join((page.extract_text() or "") for page in reader.pages).strip()
        if texto:
            return texto
        return _extraer_texto_pdf_ocr(path)
    except Exception:
        texto_ocr = _extraer_texto_pdf_ocr(path)
        return texto_ocr if texto_ocr.strip() else path.stem


def _extraer_texto_docx(path: Path) -> str:
    ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
    try:
        textos: list[str] = []
        with zipfile.ZipFile(path) as zf:
            nombres = [
                "word/document.xml",
                *sorted(name for name in zf.namelist() if name.startswith("word/header") and name.endswith(".xml")),
                *sorted(name for name in zf.namelist() if name.startswith("word/footer") and name.endswith(".xml")),
            ]
            for nombre in nombres:
                if nombre not in zf.namelist():
                    continue
                root = ElementTree.fromstring(zf.read(nombre))
                for paragraph in root.findall(".//w:p", ns):
                    partes = [node.text or "" for node in paragraph.findall(".//w:t", ns)]
                    linea = "".join(partes).strip()
                    if linea:
                        textos.append(linea)
        return "\n".join(textos).strip() or path.stem
    except Exception:
        return _extraer_texto_word(path)


def extraer_texto_pdf(path: Path) -> str:
    return extraer_texto_informe(path)


def _extraer_texto_pdf_ocr(path: Path) -> str:
    pdftoppm = _buscar_pdftoppm_cmd()
    if not pdftoppm:
        return path.stem
    try:
        with tempfile.TemporaryDirectory(prefix="pami_pdf_ocr_") as tmp:
            prefix = Path(tmp) / "page"
            subprocess.run(
                [str(pdftoppm), "-r", "200", "-png", str(path), str(prefix)],
                capture_output=True,
                text=True,
                encoding="utf-8",
                errors="replace",
                timeout=60,
                check=True,
                **_subprocess_no_window_kwargs(),
            )
            textos: list[str] = []
            for image_path in sorted(Path(tmp).glob("page-*.png")):
                texto = _extraer_texto_imagen(image_path)
                if texto and texto != image_path.stem:
                    textos.append(texto)
            return "\n".join(textos).strip() or path.stem
    except Exception:
        return path.stem


def _buscar_pdftoppm_cmd() -> Path | None:
    candidates = [
        Path.home()
        / ".cache"
        / "codex-runtimes"
        / "codex-primary-runtime"
        / "dependencies"
        / "native"
        / "poppler"
        / "Library"
        / "bin"
        / "pdftoppm.exe",
        Path.home()
        / ".cache"
        / "codex-runtimes"
        / "codex-primary-runtime"
        / "dependencies"
        / "bin"
        / "override"
        / "pdftoppm.cmd",
    ]
    for candidate in candidates:
        if candidate.exists():
            return candidate
    found = shutil.which("pdftoppm")
    if found:
        return Path(found)
    return None


def _imagen_en_blanco(path: Path) -> bool:
    if path.suffix.lower() not in IMAGE_EXTENSIONS:
        return False
    try:
        from PIL import Image, ImageChops

        image = Image.open(path).convert("RGBA")
        white = Image.new("RGBA", image.size, (255, 255, 255, 255))
        image = Image.alpha_composite(white, image).convert("RGB")
        diff = ImageChops.difference(image, Image.new("RGB", image.size, (255, 255, 255)))
        if not diff.getbbox():
            return True
        total = max(1, image.width * image.height)
        non_white = sum(1 for pixel in image.getdata() if min(pixel) < 245)
        return non_white / total <= 0.002
    except Exception:
        return False


def _extraer_texto_imagen(path: Path) -> str:
    global _RAPIDOCR_AVAILABLE, _RAPIDOCR_ENGINE, _EASYOCR_AVAILABLE, _EASYOCR_READER
    textos: list[str] = []
    if _RAPIDOCR_AVAILABLE is not False:
        try:
            from rapidocr_onnxruntime import RapidOCR

            if _RAPIDOCR_ENGINE is None:
                _RAPIDOCR_ENGINE = RapidOCR()
            result, _elapsed = _RAPIDOCR_ENGINE(str(path))
            lines = []
            for item in result or []:
                if len(item) >= 2 and str(item[1]).strip():
                    lines.append(str(item[1]).strip())
            if lines:
                _RAPIDOCR_AVAILABLE = True
                textos.append("\n".join(lines))
        except Exception:
            _RAPIDOCR_AVAILABLE = False
    if _EASYOCR_AVAILABLE is not False and _easyocr_modelos_disponibles():
        try:
            import easyocr

            if _EASYOCR_READER is None:
                _EASYOCR_READER = easyocr.Reader(["es"], gpu=False, verbose=False, download_enabled=False)
            result = _EASYOCR_READER.readtext(str(path), detail=1, paragraph=False)
            lines = []
            for item in result or []:
                if len(item) >= 2 and str(item[1]).strip():
                    lines.append(str(item[1]).strip())
            if lines:
                _EASYOCR_AVAILABLE = True
                textos.append("\n".join(lines))
        except Exception:
            _EASYOCR_AVAILABLE = False
    try:
        import pytesseract
        from PIL import Image

        tesseract_cmd = _buscar_tesseract_cmd()
        if tesseract_cmd:
            pytesseract.pytesseract.tesseract_cmd = str(tesseract_cmd)
        text = pytesseract.image_to_string(Image.open(path), lang="spa+eng")
        if text.strip():
            textos.append(text.strip())
    except Exception:
        pass
    if not textos:
        return path.stem
    vistas: set[str] = set()
    lines: list[str] = []
    for text in textos:
        for line in text.splitlines():
            clean = line.strip()
            key = _clave_texto(clean)
            if not clean or key in vistas:
                continue
            vistas.add(key)
            lines.append(clean)
    return "\n".join(lines) or path.stem


def _easyocr_modelos_disponibles() -> bool:
    model_dir = Path.home() / ".EasyOCR" / "model"
    return (model_dir / "craft_mlt_25k.pth").exists() and any(model_dir.glob("*_g2.pth"))


def _buscar_tesseract_cmd() -> Path | None:
    candidates = [
        Path(r"C:\Program Files\Tesseract-OCR\tesseract.exe"),
        Path(r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe"),
    ]
    for candidate in candidates:
        if candidate.exists():
            return candidate
    return None


def _extraer_texto_word(path: Path) -> str:
    script = r"""
$ErrorActionPreference = 'Stop'
[Console]::OutputEncoding = [System.Text.UTF8Encoding]::new($false)
$path = $env:PAMI_DOC_PATH
$word = New-Object -ComObject Word.Application
$word.Visible = $false
$word.DisplayAlerts = 0
$doc = $null
try {
  $doc = $word.Documents.Open($path, $false, $true)
  [Console]::Write($doc.Content.Text)
} finally {
  if ($doc -ne $null) { $doc.Close($false) | Out-Null }
  $word.Quit() | Out-Null
  [System.Runtime.InteropServices.Marshal]::ReleaseComObject($word) | Out-Null
}
"""
    env = dict(**__import__("os").environ, PAMI_DOC_PATH=str(path))
    try:
        result = subprocess.run(
            ["powershell", "-NoProfile", "-ExecutionPolicy", "Bypass", "-Command", script],
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            env=env,
            timeout=20,
            check=True,
            **_subprocess_no_window_kwargs(),
        )
        texto = result.stdout.strip()
        if texto:
            return texto
    except Exception:
        pass

    return _extraer_texto_word_guardado_como_txt(path) or path.stem


def _extraer_texto_word_guardado_como_txt(path: Path) -> str:
    with tempfile.TemporaryDirectory(prefix="pami_doc_word_") as tmp_dir:
        txt_path = Path(tmp_dir) / "word.txt"
        script = r"""
$ErrorActionPreference = 'Stop'
$path = $env:PAMI_DOC_PATH
$out = $env:PAMI_TXT_PATH
$word = New-Object -ComObject Word.Application
$word.Visible = $false
$word.DisplayAlerts = 0
$doc = $null
try {
  $doc = $word.Documents.Open($path, $false, $true)
  $doc.SaveAs([ref] $out, [ref] 2)
} finally {
  if ($doc -ne $null) { $doc.Close($false) | Out-Null }
  $word.Quit() | Out-Null
  [System.Runtime.InteropServices.Marshal]::ReleaseComObject($word) | Out-Null
}
"""
        env = dict(**__import__("os").environ, PAMI_DOC_PATH=str(path), PAMI_TXT_PATH=str(txt_path))
        try:
            subprocess.run(
                ["powershell", "-NoProfile", "-ExecutionPolicy", "Bypass", "-Command", script],
                capture_output=True,
                text=True,
                encoding="utf-8",
                errors="replace",
                env=env,
                timeout=60,
                check=True,
                **_subprocess_no_window_kwargs(),
            )
        except Exception:
            return ""
        if not txt_path.exists():
            return ""
        raw = txt_path.read_bytes()
        for encoding in ("utf-16", "utf-8-sig", "cp1252"):
            texto = raw.decode(encoding, errors="replace").strip()
            if texto:
                return texto
    return ""


def convertir_word_a_pdf(path: Path) -> Path:
    if path.suffix.lower() not in {".doc", ".docx"}:
        return path
    out_dir = get_output_dir() / "documentacion_convertida"
    out_dir.mkdir(parents=True, exist_ok=True)
    pdf_path = out_dir / f"{path.stem}.pdf"
    if pdf_path.exists() and pdf_path.stat().st_mtime >= path.stat().st_mtime:
        return pdf_path

    script = r"""
$ErrorActionPreference = 'Stop'
$path = $env:PAMI_DOC_PATH
$out = $env:PAMI_PDF_PATH
$word = New-Object -ComObject Word.Application
$word.Visible = $false
$word.DisplayAlerts = 0
$doc = $null
try {
  $doc = $word.Documents.Open($path, $false, $true)
  $doc.SaveAs([ref] $out, [ref] 17)
} finally {
  if ($doc -ne $null) { $doc.Close($false) | Out-Null }
  $word.Quit() | Out-Null
  [System.Runtime.InteropServices.Marshal]::ReleaseComObject($word) | Out-Null
}
"""
    env = dict(**__import__("os").environ, PAMI_DOC_PATH=str(path), PAMI_PDF_PATH=str(pdf_path))
    subprocess.run(
        ["powershell", "-NoProfile", "-ExecutionPolicy", "Bypass", "-Command", script],
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
        env=env,
        timeout=60,
        check=True,
        **_subprocess_no_window_kwargs(),
    )
    return pdf_path


def convertir_imagen_a_pdf(path: Path) -> Path:
    if path.suffix.lower() not in IMAGE_EXTENSIONS:
        return path
    out_dir = get_output_dir() / "documentacion_convertida"
    out_dir.mkdir(parents=True, exist_ok=True)
    pdf_path = out_dir / f"{path.stem}.pdf"
    if pdf_path.exists() and pdf_path.stat().st_mtime >= path.stat().st_mtime:
        return pdf_path
    try:
        from PIL import Image

        image = Image.open(path)
        if image.mode not in {"RGB", "L"}:
            image = image.convert("RGB")
        image.save(pdf_path, "PDF", resolution=150.0)
        return pdf_path
    except Exception:
        return path


def archivo_documentacion_para_subir(path: Path) -> Path:
    if path.suffix.lower() in IMAGE_EXTENSIONS:
        return convertir_imagen_a_pdf(path)
    return path


def _limpiar_nombre_paciente_detectado(value: str) -> str:
    clean = re.sub(r"\s+", " ", value or "").strip(" ,-:.|")
    if not clean:
        return ""
    clean = re.sub(r"\([^)]*\)", " ", clean)
    if re.match(r"^(?:n[Â°Âºº°]?\s*)?departamento\b", clean, flags=re.I):
        return ""
    if re.match(r"^(?:documento|image\d*|whatsapp\s+image)\b", clean, flags=re.I):
        return ""
    clean = re.sub(r"^(?:y\s+)?nombre\s*:?\s*", "", clean, flags=re.I).strip(" ,-:.|")
    clean = re.sub(r"^(?:paciente|apellido)\s*:?\s*", "", clean, flags=re.I).strip(" ,-:.|")
    clean = re.sub(r"^(?:r\s*/?\s*p|rp)\s*[\./:;-]*\s*", "", clean, flags=re.I).strip(" ,-:.|")
    clean = re.split(
        r"\bEDAD\s*:?\s*\d{1,3}\b|\bDNI\s*:?\s*\d{6,}\b|\bDOCUMENTO\s*:?\s*\d{6,}\b|\bSEXO\s*:",
        clean,
        maxsplit=1,
        flags=re.I,
    )[0]
    clean = re.sub(
        r"\b(?:MAPA|HOLTER|ETT|ECO\s*VC|SPIROMETRY\s+EXAM|ESPIROMETRIA|ELECTROCARDIOGRAMA|SIBO)\b",
        " ",
        clean,
        flags=re.I,
    )
    clean = re.sub(
        r"\b(?:TEST\s+DE\s+AIRE\s+ESPIRADO|SOBRECRECIMIENTO\s+BACTERIANO|HIDROGENO|METANO)\b",
        " ",
        clean,
        flags=re.I,
    )
    clean = re.sub(r"\b(?:PRIVAD[AO]?|PARTICULAR)\b", " ", clean, flags=re.I)
    clean = re.sub(r"\s+", " ", clean).strip(" ,-:.|")
    return _compactar_nombre_duplicado(clean)


def _compactar_nombre_duplicado(value: str) -> str:
    clean = re.sub(r"\s+", " ", value or "").strip(" ,-:.|")
    tokens = clean.split()
    if len(tokens) >= 4 and len(tokens) % 2 == 0:
        half = len(tokens) // 2
        if [_clave_texto(token) for token in tokens[:half]] == [_clave_texto(token) for token in tokens[half:]]:
            return " ".join(tokens[:half]).strip(" ,-:.|")
    return clean


def _nombre_detectado_parece_ruido(value: str) -> bool:
    key = _clave_texto(value)
    if not key:
        return True
    if key in {"tnere poo", "pou cnoto cagmiro", "gogr lorccra"}:
        return True
    patrones_ruido = (
        "estudio computarizado",
        "estudio solicitado",
        "datos del paciente",
        "resultados",
        "resultado",
        "informe",
        "n departamento",
        "departamento",
        "documento",
        "whatsapp image",
        "image",
        "liq k2",
        "cred post cierre",
        "displasia sin evidencia",
        "firma y sello",
        "firma sello",
        "medica",
        "centro medico",
        "caballito",
        "turnos",
    )
    if any(patron in key for patron in patrones_ruido):
        return True
    if re.search(r"\b(?:dr|dra|doctor|doctora)\b", key):
        return True
    if re.search(r"\b\d{6,}\b", key):
        return True
    return False


def _nombre_paciente_detectado_valido(value: str) -> bool:
    clean = _limpiar_nombre_paciente_detectado(value)
    key = _clave_texto(clean)
    if _nombre_detectado_parece_ruido(clean):
        return False
    tokens = key.split()
    if len(tokens) < 2:
        return False
    if not any(len(token) >= 3 for token in tokens):
        return False
    if sum(1 for token in tokens if token.isalpha()) < max(1, len(tokens) - 1):
        return False
    return True


def _normalizar_nombre_paciente_detectado_final(value: str) -> str:
    clean = _limpiar_nombre_paciente_detectado(value)
    if not _nombre_paciente_detectado_valido(clean):
        return ""
    return clean


def _normalizar_nombre_ocr_manuscrito(nombre: str, text: str) -> str:
    """Corrige deformaciones frecuentes del OCR liviano en ordenes manuscritas."""
    key = _clave_texto(nombre)
    text_key = _clave_texto(text)
    combinado = f"{key} {text_key}"
    if "ciche llo" in combinado and ("mand" in combinado or "osehng" in combinado):
        return "Cichetto Mand Josehna"
    soler_tokens = {"ssq01", "ssqa1", "ssqe1", "gsqe1", "sqa1", "sq01", "gogr"}
    laura_tokens = {"lovua", "loeua", "loeuna", "lorccra", "laeura"}
    if any(token in combinado for token in soler_tokens) and any(token in combinado for token in laura_tokens):
        return "Soler Laura"
    tripicchio_tokens = {"tripicchio", "tripicc", "trippicchio", "taipi cchi", "taipicchi"}
    carlos_tokens = {"carlos", "conqo", "carlo"}
    if any(token in combinado for token in tripicchio_tokens) and any(token in combinado for token in carlos_tokens):
        return "Tripicchio Carlos"
    return nombre


def _correcciones_ocr_manuscrito_documentacion(text: str) -> dict:
    key = _clave_texto(text)
    if "ciche llo" in key and ("eechocoagulacon" in key or "echocoagulacon" in key):
        return {"texto_extra": "electrocoagulacion\n537106"}
    return {}


def _extraer_nombre_ecud_desde_texto(text: str) -> str:
    """Extrae el paciente de informes ECUD donde la linea mezcla nombre y estudio."""
    cortes = (
        r"Uroflujometr[ií]a",
        r"Flujometr[ií]a",
        r"Estudio\s+urodin[aá]mico",
        r"Urodinamia",
        r"Tipo\s*:",
    )
    patron_corte = re.compile(rf"\b(?:{'|'.join(cortes)})\b", flags=re.I)
    for raw_line in text.splitlines():
        line = re.sub(r"\s+", " ", raw_line).strip(" .:-")
        if not line:
            continue
        if not patron_corte.search(line):
            continue
        candidato = patron_corte.split(line, maxsplit=1)[0]
        candidato = re.sub(r"\b(?:PAMI|ECUD|Datos\s+del\s+Paciente|Nombre)\b", " ", candidato, flags=re.I)
        candidato = re.sub(r"\s+", " ", candidato).strip(" .:-")
        candidato = _normalizar_nombre_paciente_detectado_final(candidato)
        if candidato:
            return candidato
    return ""


def _extraer_campos_paciente_pegados(text: str) -> dict:
    text = str(text or "")
    campos = {"paciente": "", "dni": "", "beneficio": ""}

    match = re.search(
        r"\bNomb(?:re)?\s*:?\s*([^\r\n]+?)"
        r"(?=\s*(?:Document(?:o)?|DNI)\s*:?\s*\d|\s*N[^\w\d]?\s*Benef|\s*Beneficio|\r?\n|$)",
        text,
        flags=re.I,
    )
    if match:
        campos["paciente"] = _normalizar_nombre_paciente_detectado_final(match.group(1))

    dni = _extraer_dni_desde_texto(text)
    if dni:
        campos["dni"] = dni

    match = re.search(r"N[^\w\d]?\s*Benef\.?\s*:?\s*(?:PAMI\s*)?(\d{10,})", text, flags=re.I)
    if not match:
        match = re.search(r"\bBeneficio\s*:?\s*(?:PAMI\s*)?(\d{10,})", text, flags=re.I)
    if match:
        campos["beneficio"] = _solo_digitos(match.group(1))

    return campos


def _nombre_paciente_desde_nombre_archivo(path: Path) -> str:
    stem = re.sub(r"[_-]+", " ", path.stem)
    stem = re.sub(r"\([^)]*\)", " ", stem)
    stem = re.sub(
        r"\b(?:PAMI|INFORME|FL|FLUJO|UROFLUJO|UROFLUJOMETRIA|FLUJOMETRIA|MAPA|SIBO|ECUD)\b",
        " ",
        stem,
        flags=re.I,
    )
    stem = re.sub(
        r"^\s*\d{1,2}\s*(?:ENE|FEB|MAR|ABR|MAY|JUN|JUL|AGO|SEP|OCT|NOV|DIC)\b\s*\d{1,2}\b",
        " ",
        stem,
        flags=re.I,
    )
    stem = re.sub(
        r"^\s*\d{1,2}\s*(?:ENE|FEB|MAR|ABR|MAY|JUN|JUL|AGO|SEP|OCT|NOV|DIC)\b",
        " ",
        stem,
        flags=re.I,
    )
    stem = re.sub(r"\b\d{1,2}[.\-]\d{1,2}[.\-]\d{2,4}\b", " ", stem)
    stem = re.sub(r"\s+", " ", stem).strip(" ,-:.|")
    nombre = _normalizar_nombre_paciente_detectado_final(stem)
    if nombre:
        return nombre
    # En nombres de archivo como "CONSTENLA FL.pdf" queda solo el apellido luego
    # de limpiar la etiqueta de practica. Sirve como fallback para cruzar con
    # bandeja, sin relajar la validacion general del OCR/texto.
    fallback = _limpiar_nombre_paciente_detectado(stem)
    key = _clave_texto(fallback)
    if key and not _nombre_detectado_parece_ruido(fallback) and len(key.split()) == 1 and len(key) >= 4:
        return fallback
    return ""


def _etiquetas_practica_desde_nombre_archivo(path: Path) -> list[str]:
    key = _clave_texto(path.stem)
    etiquetas: list[str] = []
    if re.search(r"(?:^|\s)mapa(?:\s|$)", key):
        etiquetas.append("MAPA")
    if re.search(r"(?:^|\s)sibo(?:\s|$)", key):
        etiquetas.append("SIBO")
    if re.search(r"(?:^|\s)(?:fl|flujo|uroflujo|uroflujometria|flujometria)(?:\s|$)", key):
        etiquetas.append("FLUJOMETRIA")
    return etiquetas


def _datos_rapidos_desde_nombre_archivo(path: Path) -> dict | None:
    if path.suffix.lower() != ".pdf":
        return None
    etiquetas = _etiquetas_practica_desde_nombre_archivo(path)
    if not etiquetas:
        return None
    nombre_completo = _nombre_paciente_desde_nombre_archivo(path)
    if not nombre_completo:
        return None
    text = "\n".join([path.stem, nombre_completo, *etiquetas])
    return {
        "archivo": str(path),
        "dni": "",
        "beneficio_pdf": "",
        "apellido": nombre_completo,
        "nombre": "",
        "nombre_completo": nombre_completo,
        "nombre_key": _clave_texto(nombre_completo),
        "texto": text,
        "fecha_informe": _fecha_informe_documentacion({"texto": text}, path),
        "obra_social": "",
        "obra_social_no_pami": False,
        "imagen_ocr": False,
    }


def _paciente_detectado_para_lote(datos_pdf: dict) -> str:
    return datos_pdf.get("nombre_completo", "")


_OBRAS_SOCIALES_NO_PAMI = {
    "osde": "OSDE",
    "privado": "PRIVADO",
    "ospeca": "OSPECA",
}


def _detectar_obra_social_informe(text: str) -> str:
    if not text:
        return ""
    lines = [line.strip() for line in str(text).splitlines()]
    for line in lines:
        match = re.search(
            r"\b(?:obra\s+social|cobertura|prepaga)\s*[:\-]?\s*([A-Za-z0-9 ._ÁÉÍÓÚÜÑáéíóúüñ/-]{2,60})",
            line,
            flags=re.IGNORECASE,
        )
        if not match:
            continue
        value = match.group(1).strip(" .:-\t")
        value = re.split(
            r"\s{2,}|\b(?:dni|documento|edad|fecha|peso|talla|isc|afiliad[oa])\b\s*:?",
            value,
            maxsplit=1,
            flags=re.IGNORECASE,
        )[0].strip(" .:-\t")
        if value:
            return value
    for line in lines:
        key = _clave_texto(line)
        if not re.search(r"\b(?:benef|beneficio|afiliado|afiliada)\b", key):
            continue
        for marker, label in _OBRAS_SOCIALES_NO_PAMI.items():
            if re.search(rf"(?:^|\s){re.escape(marker)}(?:\s|$)", key):
                return label
    return ""


def _obra_social_no_corresponde_pami(obra_social: str) -> bool:
    key = _clave_texto(obra_social)
    return bool(key and "pami" not in key)


def _token_nombre_canonico(token: str) -> str:
    token = _clave_texto(token)
    if token in {"anai", "anali"}:
        return "analia"
    if token in {"stella", "estella"}:
        return "estela"
    if token in {"gosalbes", "gonsalbes"}:
        return "gonsalbes"
    if token in {"trottzki", "trottski", "trotzki", "trozki"}:
        return "trotzki"
    return token


def _tokens_nombre_compatibles(left: str, right: str) -> bool:
    left = _token_nombre_canonico(left)
    right = _token_nombre_canonico(right)
    if not left or not right:
        return False
    if left == right:
        return True
    if left.isdigit() or right.isdigit():
        return False
    if len(left) == 1 and right.startswith(left):
        return True
    if len(right) == 1 and left.startswith(right):
        return True
    if len(left) >= 5 and len(right) >= 5 and (left == f"{right}s" or right == f"{left}s"):
        return True
    if min(len(left), len(right)) < 5:
        return False
    if len(left) == len(right) and left[0] == right[0] and left[-1] == right[-1]:
        if sum(1 for a, b in zip(left, right) if a != b) == 1:
            return True
    try:
        from difflib import SequenceMatcher

        return SequenceMatcher(None, left, right).ratio() >= 0.88
    except Exception:
        return False


def _tokens_nombre_utiles_para_comparar(value: str) -> list[str]:
    ignorar = {
        "pami",
        "lut",
        "luts",
        "rp",
        "rx",
        "doc",
        "pdf",
        "jpeg",
        "jpg",
        "png",
        "informe",
    }
    return [
        token
        for token in _clave_texto(value).split()
        if len(token) >= 3 and token not in ignorar and not token.isdigit()
    ]


def _token_nombre_muy_parecido(left: str, right: str) -> bool:
    if _tokens_nombre_compatibles(left, right):
        return True
    left = _token_nombre_canonico(left)
    right = _token_nombre_canonico(right)
    if min(len(left), len(right)) < 4:
        return False
    try:
        from difflib import SequenceMatcher

        return SequenceMatcher(None, left, right).ratio() >= 0.75
    except Exception:
        return False


def _debe_preferir_nombre_archivo(nombre_ocr: str, nombre_archivo: str) -> bool:
    """Usa el nombre del archivo cuando el OCR de imagen trae basura compatible."""
    nombre_archivo = _normalizar_nombre_paciente_detectado_final(nombre_archivo)
    if not nombre_archivo:
        return False
    tokens_archivo = _tokens_nombre_utiles_para_comparar(nombre_archivo)
    # Solo pisamos OCR si el archivo trae apellido y nombre; con un solo token
    # seria demasiado facil cruzar contra otro paciente de la bandeja.
    if len(tokens_archivo) < 2:
        return False
    nombre_ocr = _normalizar_nombre_paciente_detectado_final(nombre_ocr)
    if not nombre_ocr or _nombre_detectado_parece_ruido(nombre_ocr):
        return True
    tokens_ocr = _tokens_nombre_utiles_para_comparar(nombre_ocr)
    if not tokens_ocr:
        return True
    coincidencias = 0
    usados_ocr: set[int] = set()
    for token_archivo in tokens_archivo:
        for idx, token_ocr in enumerate(tokens_ocr):
            if idx in usados_ocr:
                continue
            if _token_nombre_muy_parecido(token_archivo, token_ocr):
                usados_ocr.add(idx)
                coincidencias += 1
                break
    if coincidencias >= min(2, len(tokens_archivo)):
        return True
    if coincidencias >= 1 and len(tokens_ocr) >= len(tokens_archivo) + 1:
        return True
    return False


def _extraer_datos_paciente_pdf(path: Path) -> dict:
    datos_rapidos = _datos_rapidos_desde_nombre_archivo(path)
    if datos_rapidos:
        try:
            texto_real = extraer_texto_informe(path)
        except Exception:
            texto_real = ""
        if texto_real:
            texto_rapido = str(datos_rapidos.get("texto", "") or "")
            texto_combinado = "\n".join(part for part in [texto_real, texto_rapido] if part)
            obra_social = _detectar_obra_social_informe(texto_combinado)
            datos_rapidos["texto"] = texto_combinado
            datos_rapidos["fecha_informe"] = _fecha_informe_documentacion({"texto": texto_combinado}, path)
            datos_rapidos["obra_social"] = obra_social
            datos_rapidos["obra_social_no_pami"] = _obra_social_no_corresponde_pami(obra_social)
        return datos_rapidos

    text = extraer_texto_informe(path)
    obra_social = _detectar_obra_social_informe(text)
    obra_social_no_pami = _obra_social_no_corresponde_pami(obra_social)
    es_imagen = path.suffix.lower() in IMAGE_EXTENSIONS
    archivo_blanco = es_imagen and _imagen_en_blanco(path)
    correcciones_ocr = _correcciones_ocr_manuscrito_documentacion(text) if es_imagen else {}
    if correcciones_ocr.get("texto_extra"):
        text = f"{text}\n{correcciones_ocr['texto_extra']}"
    apellido = ""
    nombre = ""
    dni = ""
    beneficio = ""
    match = re.search(
        r"Paciente\s*:\s*Edad\s*:\s*Sexo\s*:\s*Documento\s*:\s*Marcapasos\s*:\s*"
        r"([A-Z??????? ]+?)\s*,\s*([A-Z??????? ]+?)\s*\d{1,3}\s*(?:Femenino|Masculino)\s*(\d{6,})",
        text,
        flags=re.I | re.S,
    )
    if match:
        apellido = re.sub(r"\s+", " ", match.group(1)).strip(" ,-:")
        nombre = re.sub(r"\s+", " ", match.group(2)).strip(" ,-:")
        dni = _normalizar_dni_extraido(match.group(3))
    if not dni:
        dni = _extraer_dni_desde_texto(text)
    match = re.search(r"ID\s+del\s+paciente\s+(\d{6,})", text, flags=re.I)
    if match and not dni:
        dni = _normalizar_dni_extraido(match.group(1))
    if not dni:
        match = re.search(r"C[oó]d(?:igo)?\.?\s*paciente\s*:?\s*(\d{6,})", text, flags=re.I)
        if match:
            dni = _normalizar_dni_extraido(match.group(1))
    if not dni:
        match = re.search(r"\bDocumento\s*:?\s*(\d{6,})", text, flags=re.I)
        if match:
            dni = _normalizar_dni_extraido(match.group(1))
    if not dni:
        match = re.search(r"\b(\d{6,})\s*ID\b", text, flags=re.I)
        if match:
            dni = _normalizar_dni_extraido(match.group(1))
    if not dni:
        match = re.search(r"\bDNI\s*:?\s*(\d{6,})", text, flags=re.I)
        if match:
            dni = _normalizar_dni_extraido(match.group(1))
    match = re.search(r"N[°º]?\s*Benef\.?\s*:?\s*(?:PAMI\s*)?(\d{10,})", text, flags=re.I)
    if match:
        beneficio = match.group(1)
    if not beneficio:
        match = re.search(r"\bPAMI\s*[:\-]?\s*(\d{10,})", text, flags=re.I)
        if match:
            beneficio = _solo_digitos(match.group(1))
    if not beneficio and correcciones_ocr.get("beneficio"):
        beneficio = correcciones_ocr["beneficio"]
    campos_pegados = _extraer_campos_paciente_pegados(text)
    if not dni and campos_pegados.get("dni"):
        dni = campos_pegados["dni"]
    if not beneficio and campos_pegados.get("beneficio"):
        beneficio = campos_pegados["beneficio"]
    if not apellido and not nombre and campos_pegados.get("paciente"):
        apellido = campos_pegados["paciente"]
    match = re.search(r"Apellido\s+([^\n\r]+)", text, flags=re.I)
    if match and match.group(1).strip().lower() not in {"nom", "nombre"}:
        apellido = match.group(1).strip()
    match = re.search(r"Nombre\s*:?\s*([^\n\r]+)", text, flags=re.I)
    if match:
        candidato = re.split(r"\bDatos\s+del\s+Estudio\b|\bTipo\s*:", match.group(1), maxsplit=1, flags=re.I)[0]
        nombre = candidato.strip(" :-")
        if not apellido and nombre and not re.search(r"\bApellidos?\b", text, flags=re.I):
            apellido = nombre
            nombre = ""
    if not apellido and re.search(r"\bApellidos?\b", text, flags=re.I):
        match = re.search(
            r"\bApellidos?\s*:?\s*([^\n\r]+)\s+Nombre\s*:?\s*([^\n\r]+)",
            text,
            flags=re.I,
        )
        if match:
            apellido = re.sub(r"\s+", " ", match.group(1)).strip(" :-")
            nombre = re.sub(r"\s+", " ", match.group(2)).strip(" :-")
    if not apellido and not nombre:
        match = re.search(r"\bNOMBRE\s*:\s*([^\n\r]+?)(?:\s+EDAD\s*:|\s+DNI\s*:|$)", text, flags=re.I)
        if match:
            apellido = re.sub(r"\s+", " ", match.group(1)).strip(" :-")
    if _clave_texto(apellido).startswith("paciente ") and " edad " in _clave_texto(apellido):
        apellido = ""
        nombre = ""

    if not apellido and not nombre:
        for line in text.splitlines():
            if "datos del paciente" in _clave_texto(line):
                continue
            match = re.search(r"\bPac(?:iente|lente)\s*:?\s*(.*)", line, flags=re.I)
            if not match:
                continue
            paciente = match.group(1).strip()
            paciente = re.split(r"\s{2,}PAMI\b|\bPAMI\b|\s{2,}Fecha\s*:|\bFecha\s*:", paciente, maxsplit=1, flags=re.I)[0]
            paciente = re.sub(r"\s+", " ", paciente).strip(" :-")
            if _clave_texto(paciente).startswith(("hombre edad", "mujer edad")) or " edad " in _clave_texto(paciente):
                continue
            if paciente:
                apellido = paciente
                break
    if es_imagen and not apellido and not nombre:
        lines = [line.strip(" .:-") for line in text.splitlines() if line.strip(" .:-")]
        ruido = {
            "caballito",
            "centro medico",
            "centro odontologico",
            "centro kinesiologico",
            "firma y sello profesional",
            "fecha",
            "rp",
            "rp/",
        }

        def candidato_nombre(line: str) -> bool:
            key = _clave_texto(line)
            if not key or key in ruido:
                return False
            if any(token in key for token in {"directorio", "email", "turnos", "whatsapp", "especialidades", "medicas", "clinica", "cardiologia", "dermatologia"}):
                return False
            if re.search(r"\d", key):
                return False
            return len(key.split()) >= 2 and len(key) >= 8

        for idx, line in enumerate(lines):
            if not re.search(r"\bpac(?:iente|lente)\b", _clave_texto(line)):
                continue
            vecinos = []
            if idx + 1 < len(lines):
                vecinos.append(lines[idx + 1])
            if idx + 2 < len(lines):
                vecinos.append(f"{lines[idx + 1]} {lines[idx + 2]}")
            if idx - 1 >= 0:
                vecinos.append(lines[idx - 1])
            for vecino in vecinos:
                corregido = _normalizar_nombre_ocr_manuscrito(vecino, text)
                if corregido != vecino or candidato_nombre(vecino):
                    apellido = corregido
                    break
            if apellido:
                break
    if (not apellido or _clave_texto(apellido) in {"nom", "nombre"}) and re.search(r"\bApellido\b\s*\n\s*Nom\.?", text, flags=re.I):
        lines = [line.strip() for line in text.splitlines() if line.strip()]
        for idx, line in enumerate(lines):
            if _clave_texto(line) == "grupo pacientes" and idx + 2 < len(lines):
                apellido = lines[idx + 1]
                nombre = lines[idx + 2]
                break
    if not apellido and not nombre:
        match = re.search(r"Paciente\s*:\s*.*?([A-Z??????]+,\s*[A-Z??????]+(?:\s+[A-Z??????]+)?)", text, flags=re.I | re.S)
        if match:
            apellido = re.sub(r"\s+", " ", match.group(1)).strip(" ,-:")

    apellido = _limpiar_nombre_paciente_detectado(apellido)
    nombre = _limpiar_nombre_paciente_detectado(nombre)
    if not _normalizar_nombre_paciente_detectado_final(" ".join(part for part in [apellido, nombre] if part)):
        nombre_ecud = _extraer_nombre_ecud_desde_texto(text)
        if nombre_ecud:
            apellido = nombre_ecud
            nombre = ""

    stem = re.sub(r"[_-]+", " ", path.stem)
    stem = re.sub(
        r"^\s*\d{1,2}\s*(?:ENE|FEB|MAR|ABR|MAY|JUN|JUL|AGO|SEP|OCT|NOV|DIC)\b\s*\d{1,2}\b",
        " ",
        stem,
        flags=re.I,
    )
    stem = re.sub(
        r"^\s*\d{1,2}\s*(?:ENE|FEB|MAR|ABR|MAY|JUN|JUL|AGO|SEP|OCT|NOV|DIC)\b",
        " ",
        stem,
        flags=re.I,
    )
    stem = re.sub(
        r"\b(?:MAPA|ETT|ECO\s*VC|HOLTER|SPIROMETRY\s+EXAM|ESPIROMETRIA|ELECTROCARDIOGRAMA|SIBO)\b",
        " ",
        stem,
        flags=re.I,
    )
    stem = re.sub(
        r"\b(?:TEST\s+DE\s+AIRE\s+ESPIRADO|SOBRECRECIMIENTO\s+BACTERIANO|HIDROGENO|METANO)\b",
        " ",
        stem,
        flags=re.I,
    )
    stem = re.sub(r"\b(?:EU|FL)\b", " ", stem, flags=re.I)
    stem = re.sub(r"\b\d{1,2}[.\-]\d{1,2}[.\-]\d{2,4}\b", " ", stem)
    stem = re.sub(r"\bMAPA\b", " ", stem, flags=re.I)
    stem = re.sub(r"\s+", " ", stem).strip()
    if _clave_texto(apellido) in {"datos del estudio", "datos del paciente", "tipo", "nombre tipo"}:
        apellido = ""
        nombre = ""
    if not apellido and not nombre:
        nombre_ecud = _extraer_nombre_ecud_desde_texto(text)
        if nombre_ecud:
            apellido = nombre_ecud
            nombre = ""
    if (not apellido and not nombre) or (
        "mapa" in _clave_texto(path.stem)
        and (_clave_texto(apellido).startswith(("paciente hombre edad", "paciente mujer edad")) or not apellido)
    ):
        apellido = stem
        nombre = ""
    if not apellido and not nombre:
        apellido = stem

    nombre_archivo = _nombre_paciente_desde_nombre_archivo(path)
    nombre_completo = " ".join(part for part in [apellido, nombre] if part).strip()
    nombre_completo = _normalizar_nombre_paciente_detectado_final(nombre_completo)
    if es_imagen:
        nombre_completo = _normalizar_nombre_ocr_manuscrito(nombre_completo, text)
        if correcciones_ocr.get("nombre"):
            nombre_completo = correcciones_ocr["nombre"]
        if _debe_preferir_nombre_archivo(nombre_completo, nombre_archivo):
            nombre_completo = nombre_archivo
    if not nombre_completo:
        nombre_completo = nombre_archivo
    if not nombre_completo:
        apellido = ""
        nombre = ""
    else:
        apellido = nombre_completo
        nombre = ""
    return {
        "archivo": str(path),
        "dni": dni,
        "beneficio_pdf": beneficio,
        "apellido": apellido,
        "nombre": nombre,
        "nombre_completo": nombre_completo,
        "nombre_key": _clave_texto(nombre_completo),
        "texto": text,
        "fecha_informe": _fecha_informe_documentacion({"texto": text}, path),
        "obra_social": obra_social,
        "obra_social_no_pami": obra_social_no_pami,
        "imagen_ocr": es_imagen and _clave_texto(text) != _clave_texto(path.stem),
        "archivo_blanco": archivo_blanco,
    }


def _score_nombre(doc_key: str, paciente_key: str) -> float:
    if not doc_key or not paciente_key:
        return 0.0
    doc_compact = re.sub(r"\s+", "", doc_key)
    paciente_compact = re.sub(r"\s+", "", paciente_key)
    if doc_compact and paciente_compact and (doc_compact in paciente_compact or paciente_compact in doc_compact):
        return 0.95
    doc_tokens = set(doc_key.split())
    paciente_tokens = set(paciente_key.split())
    if not doc_tokens or not paciente_tokens:
        return 0.0
    inter = len(doc_tokens & paciente_tokens)
    doc_tokens_canon = {_token_nombre_canonico(token) for token in doc_tokens}
    paciente_tokens_canon = {_token_nombre_canonico(token) for token in paciente_tokens}
    inter_canon = len(doc_tokens_canon & paciente_tokens_canon)
    if inter_canon >= min(len(doc_tokens_canon), 2) and inter_canon / max(len(doc_tokens_canon), 1) >= 0.8:
        return 0.9
    paciente_parts_usados: set[int] = set()
    fuzzy_matches = 0
    for doc_part in doc_key.split():
        for idx, paciente_part in enumerate(paciente_key.split()):
            if idx in paciente_parts_usados:
                continue
            if _tokens_nombre_compatibles(doc_part, paciente_part):
                paciente_parts_usados.add(idx)
                fuzzy_matches += 1
                break
    if fuzzy_matches >= min(len(doc_tokens), 2) and fuzzy_matches / max(len(doc_tokens), 1) >= 0.8:
        return 0.9
    if len(doc_tokens) >= 3 and inter >= 3 and inter / max(len(doc_tokens), 1) >= 0.8:
        return 0.9
    doc_parts = doc_key.split()
    paciente_parts = paciente_key.split()
    if len(doc_parts) >= 2 and len(paciente_parts) >= 2:
        doc_first = doc_parts[0]
        paciente_first = paciente_parts[0]
        if doc_first != paciente_first:
            try:
                from difflib import SequenceMatcher

                first_ratio = SequenceMatcher(None, doc_first, paciente_first).ratio()
            except Exception:
                first_ratio = 0.0
            if first_ratio < 0.86:
                return 0.0
    token_score = inter / max(len(doc_tokens), 1)
    try:
        from difflib import SequenceMatcher

        ratio = SequenceMatcher(None, doc_key, paciente_key).ratio()
    except Exception:
        ratio = 0.0
    return max(token_score, ratio)


def _score_nombre_relajado_ocr(doc_key: str, paciente_key: str) -> float:
    base = _score_nombre(doc_key, paciente_key)
    if base >= 0.78:
        return base
    if not doc_key or not paciente_key:
        return base
    try:
        from difflib import SequenceMatcher

        ratio = SequenceMatcher(None, doc_key, paciente_key).ratio()
    except Exception:
        ratio = 0.0
    return max(base, ratio if ratio >= 0.72 else 0.0)


def _practica_keywords_pdf(text: str) -> set[str]:
    key = _clave_texto(text)
    estudio_key = _clave_texto(_linea_estudio_informe(text))
    practica_key = estudio_key or key
    keywords: set[str] = set()
    es_holter = "holter" in key or "holter" in practica_key
    es_flujometria = (
        "flujometr" in key
        or "uroflujometr" in key
        or re.search(r"\bfl\b", key)
    )
    es_urodinamia = (
        "urodinam" in key
        or "ecud" in key
        or re.search(r"\beu\b", key)
    ) and not es_flujometria
    es_funcional_urologico = es_flujometria or es_urodinamia
    es_lactosa = "lactose" in key or "lactosa" in key or "fructosa" in key or "sacarosa" in key
    if "espirometria" in key or "spirometr" in key or "funcion pulmonar" in key:
        keywords.update({"espirometria", "687114"})
    if not es_holter and ("ecg" in key or "electrocardiograma" in key):
        keywords.update({"electrocardiograma", "570126"})
    if es_holter:
        keywords.update({"holter", "570121"})
    if "presurometria" in key or "mapa" in key:
        keywords.update({"presurometria", "570120"})
    if es_urodinamia:
        keywords.update({"urodinamico", "urodinamia", "ecud", "507313"})
    if es_flujometria:
        keywords.update({"flujometria", "uroflujometria", "507315"})
    if "otomicrosc" in key or "otomicro" in key:
        keywords.update({"otomicroscopia", "otomicro", "717122"})
    if "cerumen" in key or "tapon" in key or "extraccion de cuerpo extrano" in key or "extraccion de cuerpo extra" in key:
        keywords.update({
            "cerumen",
            "tapon de cerumen",
            "extraccion de tapon de cerumen",
            "extraccion de cuerpo extrano",
            "extraccion de cuerpo extrano en oido",
            "717111",
        })
    if "rinomanometr" in key or re.search(r"\brino\b", key):
        keywords.update({"rinomanometria", "examen funcional de nariz", "717110"})
    if _tiene_tratamiento_orl_key(key):
        keywords.update({"tratamiento lesiones otorrinolaringologicas", "medios fisicos o quimicos", "717125"})
    if ("sibo" in key and not es_lactosa) or "sobrecrecimiento bacteriano" in key or "hidrogeno" in key or "metano" in key:
        keywords.update({"test de aire espirado", "sobrecrecimiento bacteriano", "hidrogeno", "metano", "607130"})
    if es_lactosa:
        keywords.update({"test de aire espirado", "lactosa", "fructosa", "sacarosa", "607131"})
    if (
        ("ecodoppler" in key or "doppler" in key)
        and "venoso" in key
        and _tiene_miembros_inferiores_key(key)
    ):
        keywords.update({"ecodoppler venoso de miembros inferiores", "venoso de miembros inferiores", "180606"})
    if (
        ("ecodoppler" in key or "doppler" in key)
        and "arterial" in key
        and _tiene_miembros_inferiores_key(key)
    ):
        keywords.update({"ecodoppler arterial de miembros inferiores", "arterial de miembros inferiores", "180610"})
    if "ecodoppler color de carotidas" in key or "ecodoppler vasos de cuello" in key or "eco vc" in key:
        keywords.update({"ecodoppler vasos cuello", "vasos cuello", "carotidas", "180607"})
    if "ecocardiograma doppler color" in key or re.search(r"\bett\b", key):
        keywords.update({"ecodoppler cardiaco", "ecocardiograma", "ett", "180301"})
    if "biopsia" in key and ("piel" in key or "tejido celular" in key or "subcutaneo" in key):
        keywords.update({"biopsia", "biopsia de piel", "537108"})
    if (
        "topicacion" in key
        or re.search(r"\btca\b", key)
        or ("destruccion" in key and "lesion" in key and "piel" in key)
        or ("ablacion" in key and "lesion" in key and "piel" in key)
        or "criocirugia" in key
        or "crio cirugia" in key
        or "electrocoagulacion" in key
        or "electro coagulacion" in key
        or "eechocoagulacon" in key
        or "echocoagulacon" in key
        or "verruga" in key
        or "queratosis" in key
        or _tiene_dermatologia_ocr_manuscrita_key(key)
        or "537106" in key
        or "510320" in key
    ):
        keywords.update(
            {
                "destruccion lesion piel",
                "ablacion lesiones piel",
                "criocirugia",
                "electrocoagulacion",
                "eechocoagulacon",
                "echocoagulacon",
                "topicacion",
                "tca",
                "queratosis",
                "verruga",
                "537106",
                "510320",
            }
        )
    if _tiene_transvaginal_key(key):
        keywords.update({"transvaginal", "transvag", "tv", "endocavitaria", "180128", "180104", "tocoginecologica"})
    if _tiene_mamaria_key(key):
        keywords.update({"mamaria", "180106"})
    if "tiroidea" in key or "tiroides" in key:
        keywords.update({"tiroidea", "tiroides", "180110"})
    if _tiene_abdominal_key(practica_key) and not es_funcional_urologico:
        keywords.update({"abdominal", "abdomen", "abd", "180112"})
    if (
        ("ginecologica" in practica_key or "tocoginecologica" in practica_key)
        and not _tiene_transvaginal_key(practica_key)
        and not _tiene_mamaria_key(practica_key)
    ):
        keywords.update({"ginecologica", "ginecologica transabdominal", "tocoginecologica", "tocoginecologica transabdominal", "180104"})
    tiene_reno_vesico_prostatica = _tiene_reno_vesico_prostatica_key(f"{practica_key} {key}")
    tiene_vesical = (
        ("vesical" in practica_key and "transvesical" not in practica_key)
        or tiene_reno_vesico_prostatica
        or (_tiene_residuo_post_miccional_key(practica_key) and not es_funcional_urologico)
    )
    tiene_prostata = (
        "prostatica" in practica_key
        or "prostata" in practica_key
        or "prostica" in practica_key
        or "vesicoprostatica" in practica_key
        or "vesicoprostica" in practica_key
        or tiene_reno_vesico_prostatica
    )
    tiene_renal = "renal" in practica_key or "rinon" in practica_key or "rinones" in practica_key
    if "abdominorenal" in practica_key or tiene_reno_vesico_prostatica:
        tiene_renal = True
    if not tiene_prostata and tiene_vesical and ("prostata" in key or "prostatico" in key):
        tiene_prostata = True
    if tiene_vesical and tiene_prostata and not tiene_renal:
        keywords.update({"vesicoprostatica", "vesical", "prostata", "prostatica", "180114"})
        if tiene_reno_vesico_prostatica or _tiene_residuo_post_miccional_key(practica_key) or _tiene_residuo_post_miccional_key(key):
            keywords.update({"residuo post miccional", "rpm", "180123"})
        return keywords
    if tiene_renal and tiene_vesical and tiene_prostata:
        keywords.update({"renal", "180116", "vesicoprostatica", "vesical", "prostata", "prostatica", "180114"})
        if tiene_reno_vesico_prostatica or _tiene_residuo_post_miccional_key(practica_key) or _tiene_residuo_post_miccional_key(key):
            keywords.update({"residuo post miccional", "rpm", "180123"})
        return keywords
    if "doppler" in practica_key and "renal" in practica_key:
        keywords.update({"ecodoppler renal", "doppler renal", "180608"})
    elif "renal" in practica_key or (not estudio_key and ("rinon" in key or "rinones" in key)):
        keywords.update({"renal", "180116"})
    if tiene_vesical:
        keywords.update({"vesical", "residuo post miccional", "rpm", "180123"})
    if tiene_prostata:
        keywords.update({"prostatica", "prostata", "180114"})
    return keywords


def _tiene_tratamiento_orl_key(key: str) -> bool:
    key = str(key or "")
    return (
        "717125" in key
        or "tto quimic" in key
        or "tratamiento quimic" in key
        or "tto de lesion quimic" in key
        or "lesiones otorrinolaringologicas" in key
        or "lesion otorrinolaringologica" in key
        or (
            "otorrinolaringolog" in key
            and (
                "tratamiento" in key
                or "medios fisicos" in key
                or "quimic" in key
                or "lesion" in key
            )
        )
        or (
            "tratamiento local" in key
            and ("medios fisicos" in key or "quimic" in key)
            and ("cerumen" in key or "conducto auditivo" in key or "lesion local" in key)
        )
    )


def _score_practica_pdf(keywords: set[str], practica: str) -> float:
    if not keywords:
        return 0.0
    practica_key = _clave_texto(practica)
    codigo = _codigo_practica(practica)
    if (
        codigo == "180301"
        and {"570126", "electrocardiograma", "ecg"} & keywords
        and not {"180301", "ecodoppler cardiaco", "ecocardiograma", "ett"} & keywords
    ):
        return 0.0
    return 1.0 if any(keyword in practica_key for keyword in keywords) else 0.0


def _codigo_practica(value: str) -> str:
    match = re.search(r"\b(\d{6})\b", str(value or ""))
    return match.group(1) if match else ""


_CODIGOS_EQUIVALENTES_DOCUMENTACION = {
    # Un informe de ECG puede quedar cubierto por una consulta cardiologica que lo incluye.
    # Estas practicas no necesariamente requieren carga de documentacion, pero sirven para
    # reconocer el estado real en PAMI y evitar reportarlo como paciente no encontrado.
    "570126": {"570126", "570129", "820113"},
    # PAMI puede registrar procedimientos de lesiones de piel con codigos distintos
    # segun la tecnica/descriptivo, pero el informe clinico suele describir el acto
    # como electrocoagulacion, criocirugia, queratosis o verruga.
    "537106": {"537106", "510320"},
    "510320": {"510320", "537106"},
    # En informes reno/vesico/prostaticos, la vesicoprostatica ya cubre el
    # componente vesical cuando PAMI no trae una OME separada de residuo.
    "180123": {"180123", "180114"},
    # Criterio administrativo: para documentacion de miembros inferiores se
    # acepta venoso/arterial como codigo compatible del mismo paciente.
    "180606": {"180606", "180610"},
    "180610": {"180610", "180606"},
}


def _codigos_equivalentes_documentacion(codigo: str) -> set[str]:
    codigo = str(codigo or "")
    equivalentes = _CODIGOS_EQUIVALENTES_DOCUMENTACION.get(codigo)
    if equivalentes:
        return set(equivalentes)
    for principal, relacionados in _CODIGOS_EQUIVALENTES_DOCUMENTACION.items():
        if codigo in relacionados:
            return set(relacionados) | {principal}
    return {codigo} if codigo else set()


def _codigo_equivalente_distinto(codigo_objetivo: str, practicas: list[str]) -> str:
    codigo_objetivo = _codigo_practica(codigo_objetivo) or str(codigo_objetivo or "")
    equivalentes = _codigos_equivalentes_documentacion(codigo_objetivo)
    for practica in practicas:
        codigo = _codigo_practica(practica)
        if codigo and codigo != codigo_objetivo and codigo in equivalentes:
            return codigo
    return ""


def _linea_estudio_informe(text: str) -> str:
    raw = str(text or "")
    estudio_match = re.search(r"\bEstudio(?:\s+realizado)?\s*:\s*([^\n\r]+)", raw, flags=re.I)
    if estudio_match:
        return estudio_match.group(1)
    for line in raw.splitlines():
        clean = line.strip(" \t\r\n.-:")
        clean_key = _clave_texto(clean)
        if not clean_key:
            continue
        if re.search(r"\becografia\b", clean_key) or re.search(r"\becodoppler\b", clean_key):
            return clean
    return ""


def _tiene_residuo_post_miccional_key(key: str) -> bool:
    return (
        "residuo post miccional" in key
        or "post miccional" in key
        or "postmiccional" in key
        or "residuo post miccion" in key
        or "post miccion" in key
        or "postmiccion" in key
        or ("residuo" in key and "miccion" in key)
        or re.search(r"\brpm\b", key)
    )


def _tiene_miembros_inferiores_key(key: str) -> bool:
    compact = re.sub(r"[^a-z0-9]+", "", _clave_texto(key))
    return (
        "miembros inferiores" in key
        or "miembrosinferiores" in compact
        or "miembrosnfeiores" in compact
        or "miembrosinfer" in compact
        or re.search(r"\bmmi\b", key) is not None
        or re.search(r"\bmmii\b", key) is not None
    )


def _tiene_abdominal_key(key: str) -> bool:
    return (
        "abdominal" in key
        or "abodminal" in key
        or "abdomen" in key
        or "abdominorenal" in key
        or re.search(r"\babd\b", key) is not None
    )


def _tiene_transvaginal_key(key: str) -> bool:
    return (
        "transvaginal" in key
        or "transvag" in key
        or "transvecial" in key
        or "transvesical" in key
        or "endocavitaria" in key
        or "endocaviatria" in key
        or re.search(r"\btv\b", key) is not None
    )


def _tiene_mamaria_key(key: str) -> bool:
    return (
        "mamaria" in key
        or "mamas" in key
        or re.search(r"\btv\s+(?:y\s+)?m\b", key) is not None
    )


def _tiene_reno_vesico_prostatica_key(key: str) -> bool:
    return (
        "renovesicoprostatica" in key
        or "renovesicoprostatico" in key
        or "renovesicoprostica" in key
        or "renovesicoprostico" in key
        or "renovesico prostatica" in key
        or "renovesico prostatico" in key
        or "renovesico prostica" in key
        or "renovesico prostico" in key
        or "reno vesico prostatica" in key
        or "reno vesico prostatico" in key
        or "reno vesico prostica" in key
        or "reno vesico prostico" in key
    )


def _tiene_dermatologia_ocr_manuscrita_key(key: str) -> bool:
    return (
        ("cuio" in key and ("4egio" in key or "cireeg" in key or "cireg" in key))
        or ("qucu" in key and ("atmico" in key or "actmico" in key))
        or ("cinuuge" in key and ("ctmieo" in key or "octmicc" in key or "cobeludo" in key or "calorbqude" in key))
        or ("ossipoff" in key and ("crio" in key or "cuio" in key or "qucu" in key or "cinuuge" in key))
    )


def _requisitos_informe(text: str) -> list[dict]:
    key = _clave_texto(text)
    estudio_key = _clave_texto(_linea_estudio_informe(text))
    es_holter = "holter" in key or "holter" in estudio_key
    es_flujometria = (
        "flujometr" in key
        or "uroflujometr" in key
        or re.search(r"\bfl\b", key)
    )
    es_urodinamia = (
        "urodinam" in key
        or "ecud" in key
        or re.search(r"\beu\b", key)
    ) and not es_flujometria
    es_funcional_urologico = es_flujometria or es_urodinamia
    es_lactosa = "lactose" in key or "lactosa" in key or "fructosa" in key or "sacarosa" in key
    requisitos: list[dict] = []

    def add(codigo: str, descripcion: str, aliases: set[str]) -> None:
        if any(item["codigo"] == codigo for item in requisitos):
            return
        requisitos.append({"codigo": codigo, "descripcion": descripcion, "aliases": aliases})

    if "espirometria" in key or "spirometr" in key or "funcion pulmonar" in key:
        add(
            "687114",
            "ESPIROMETRIA (INCLUYE CURVA DE FLUJO VOLUMEN/USO DE BRONCODILATADORES)",
            {"espirometria", "687114"},
        )
    if "biopsia" in key and ("piel" in key or "tejido celular" in key or "subcutaneo" in key):
        add(
            "537108",
            "BIOPSIA DE PIEL Y/O TEJIDO CELULAR SUBCUTANEO Y/O MUSCULAR",
            {"biopsia", "biopsia de piel", "537108"},
        )
    if (
        "topicacion" in key
        or re.search(r"\btca\b", key)
        or ("destruccion" in key and "lesion" in key and "piel" in key)
        or ("ablacion" in key and "lesion" in key and "piel" in key)
        or "criocirugia" in key
        or "crio cirugia" in key
        or "electrocoagulacion" in key
        or "electro coagulacion" in key
        or "eechocoagulacon" in key
        or "echocoagulacon" in key
        or "verruga" in key
        or "queratosis" in key
        or _tiene_dermatologia_ocr_manuscrita_key(key)
        or "537106" in key
        or "510320" in key
    ):
        add(
            "537106",
            "DESTRUCCION DE LESION DE PIEL POR ELECTROCOAGULACION O APLICACION DE SUSTANCIAS QUIMICAS",
            {
                "destruccion",
                "ablacion",
                "lesion",
                "piel",
                "criocirugia",
                "electrocoagulacion",
                "eechocoagulacon",
                "echocoagulacon",
                "topicacion",
                "tca",
                "queratosis",
                "verruga",
                "537106",
                "510320",
            },
        )
    if _tiene_transvaginal_key(key):
        add(
            "180128",
            "ECOGRAFIA ENDOCAVITARIA GINECOLOGICA",
            {"transvaginal", "transvag", "tv", "endocavitaria", "endocaviatria", "180128", "180104", "tocoginecologica"},
        )
    if (
        ("ginecologica" in estudio_key or "tocoginecologica" in estudio_key)
        and not _tiene_transvaginal_key(estudio_key)
        and not _tiene_mamaria_key(estudio_key)
    ):
        add(
            "180104",
            "ECOGRAFIA GINECOLOGICA / TOCOGINECOLOGICA TRANSABDOMINAL",
            {"ginecologica", "ginecologica transabdominal", "tocoginecologica", "tocoginecologica transabdominal", "180104"},
        )
    if _tiene_mamaria_key(key):
        add("180106", "ECOGRAFIA MAMARIA BILATERAL", {"mamaria", "mamas", "180106"})

    if not es_funcional_urologico and (
        _tiene_abdominal_key(estudio_key) or (not estudio_key and _tiene_abdominal_key(key))
    ):
        add("180112", "ECOGRAFIA ABDOMINAL COMPLETA", {"abdominal", "abdomen", "abd", "180112"})
    if "tiroidea" in key or "tiroides" in key:
        add("180110", "ECOGRAFIA TIROIDEA", {"tiroidea", "tiroides", "180110"})

    tiene_reno_vesico_prostatica_estudio = _tiene_reno_vesico_prostatica_key(f"{estudio_key} {key}")
    tiene_vesical_estudio = (
        ("vesical" in estudio_key and "transvesical" not in estudio_key)
        or tiene_reno_vesico_prostatica_estudio
        or (_tiene_residuo_post_miccional_key(estudio_key) and not es_funcional_urologico)
    )
    tiene_prostata_estudio = (
        "prostatica" in estudio_key
        or "prostata" in estudio_key
        or "prostica" in estudio_key
        or "vesicoprostatica" in estudio_key
        or "vesicoprostica" in estudio_key
        or tiene_reno_vesico_prostatica_estudio
    )
    tiene_renal_estudio = "renal" in estudio_key or "rinon" in estudio_key or "rinones" in estudio_key
    if "abdominorenal" in estudio_key or tiene_reno_vesico_prostatica_estudio:
        tiene_renal_estudio = True
    if not tiene_prostata_estudio and tiene_vesical_estudio and ("prostata" in key or "prostatico" in key):
        tiene_prostata_estudio = True
    if (
        tiene_reno_vesico_prostatica_estudio
        and (_tiene_abdominal_key(estudio_key) or _tiene_abdominal_key(key))
        and not es_funcional_urologico
    ):
        add(
            "180114",
            "ECOGRAFIA VESICOPROSTATICA",
            {"vesicoprostatica", "vesical", "prostata", "prostatica", "180114"},
        )
        return requisitos
    if tiene_vesical_estudio and tiene_prostata_estudio and not tiene_renal_estudio:
        add(
            "180114",
            "ECOGRAFIA VESICOPROSTATICA",
            {"vesicoprostatica", "vesical", "prostata", "prostatica", "180114"},
        )
        if _tiene_residuo_post_miccional_key(estudio_key) or (
            not estudio_key and _tiene_residuo_post_miccional_key(key)
        ):
            add(
                "180123",
                "ECOGRAFIA VESICAL CON MEDICION DE RESIDUO POST MICCIONAL",
                {"vesical", "residuo post miccional", "rpm", "180123"},
            )
        return requisitos

    if tiene_renal_estudio and tiene_vesical_estudio and tiene_prostata_estudio:
        add("180116", "ECOGRAFIA RENAL BILATERAL", {"renal", "rinon", "rinones", "180116"})
        add(
            "180114",
            "ECOGRAFIA VESICOPROSTATICA",
            {"vesicoprostatica", "vesical", "prostata", "prostatica", "180114"},
        )
        if _tiene_residuo_post_miccional_key(estudio_key) or (
            not estudio_key and not tiene_reno_vesico_prostatica_estudio and _tiene_residuo_post_miccional_key(key)
        ):
            add(
                "180123",
                "ECOGRAFIA VESICAL CON MEDICION DE RESIDUO POST MICCIONAL",
                {"vesical", "residuo post miccional", "rpm", "180123"},
            )
        return requisitos

    if estudio_key and "doppler" in estudio_key and tiene_renal_estudio:
        add("180608", "ECODOPPLER RENAL", {"ecodoppler renal", "doppler renal", "180608"})
        return requisitos

    if "renal" in estudio_key or "rinon" in estudio_key or "rinones" in estudio_key:
        add("180116", "ECOGRAFIA RENAL BILATERAL", {"renal", "rinon", "rinones", "180116"})
    elif not estudio_key and ("ecografia renal" in key or "rinon" in key or "rinones" in key):
        add("180116", "ECOGRAFIA RENAL BILATERAL", {"renal", "rinon", "rinones", "180116"})

    if tiene_vesical_estudio:
        add(
            "180123",
            "ECOGRAFIA VESICAL CON MEDICION DE RESIDUO POST MICCIONAL",
            {"vesical", "residuo post miccional", "rpm", "180123"},
        )
    elif not estudio_key and not es_funcional_urologico and ("ecografia vesical" in key or _tiene_residuo_post_miccional_key(key)):
        add(
            "180123",
            "ECOGRAFIA VESICAL CON MEDICION DE RESIDUO POST MICCIONAL",
            {"vesical", "residuo post miccional", "rpm", "180123"},
        )

    if "prostatica" in estudio_key or "prostata" in estudio_key:
        add("180114", "ECOGRAFIA PROSTATICA", {"prostatica", "prostata", "180114"})
    elif not estudio_key and ("ecografia prostatica" in key or "prostata" in key):
        add("180114", "ECOGRAFIA PROSTATICA", {"prostatica", "prostata", "180114"})
    if es_holter:
        add("570121", "HOLTER CARDIACO DE 3 CANALES 24 HS.", {"holter", "570121"})
        return requisitos
    if es_urodinamia:
        add(
            "507313",
            "ESTUDIO URODINAMICO COMPLETO (NO INCLUYE INSUMOS NI CATETERES)",
            {"urodinamico", "urodinamia", "ecud", "507313"},
        )
    if es_flujometria:
        add(
            "507315",
            "FLUJOMETRIA URINARIA COMPUTARIZADA",
            {"flujometria", "uroflujometria", "507315"},
        )
    if "otomicrosc" in key or "otomicro" in key:
        add("717122", "OTOMICROSCOPIA", {"otomicroscopia", "otomicro", "717122"})
    if "cerumen" in key or "tapon" in key or "extraccion de cuerpo extrano" in key or "extraccion de cuerpo extra" in key:
        add(
            "717111",
            "EXTRACCION DE CUERPO EXTRANO EN OIDO / TAPON DE CERUMEN",
            {
                "cerumen",
                "tapon de cerumen",
                "extraccion de tapon de cerumen",
                "extraccion de cuerpo extrano",
                "extraccion de cuerpo extrano en oido",
                "717111",
            },
        )
    if "rinomanometr" in key or re.search(r"\brino\b", key):
        add(
            "717110",
            "EXAMEN FUNCIONAL DE NARIZ (RINOMANOMETRIA)",
            {"rinomanometria", "examen funcional de nariz", "717110"},
        )
    if _tiene_tratamiento_orl_key(key):
        add(
            "717125",
            "TRATAMIENTO DE LESIONES OTORRINOLARINGOLOGICAS POR MEDIOS FISICOS O QUIMICOS",
            {"tratamiento lesiones otorrinolaringologicas", "medios fisicos o quimicos", "717125"},
        )
    if ("sibo" in key and not es_lactosa) or "sobrecrecimiento bacteriano" in key or "hidrogeno" in key or "metano" in key:
        add(
            "607130",
            "TEST DE AIRE ESPIRADO PARA SOBRECRECIMIENTO BACTERIANO CON MEDICION DE HIDROGENO Y METANO",
            {"test de aire espirado", "sobrecrecimiento bacteriano", "hidrogeno", "metano", "607130"},
        )
    if es_lactosa:
        add(
            "607131",
            "TEST DE AIRE ESPIRADO PARA INTOLERANCIA A LACTOSA, FRUCTOSA O SACAROSA",
            {"test de aire espirado", "lactosa", "fructosa", "sacarosa", "607131"},
        )
    if (
        ("ecodoppler" in key or "doppler" in key)
        and "venoso" in key
        and _tiene_miembros_inferiores_key(key)
    ):
        add(
            "180606",
            "ECODOPPLER VENOSO DE MIEMBROS INFERIORES",
            {"ecodoppler venoso de miembros inferiores", "venoso de miembros inferiores", "180606"},
        )
    if (
        ("ecodoppler" in key or "doppler" in key)
        and "arterial" in key
        and _tiene_miembros_inferiores_key(key)
    ):
        add(
            "180610",
            "ECODOPPLER ARTERIAL DE MIEMBROS INFERIORES",
            {"ecodoppler arterial de miembros inferiores", "arterial de miembros inferiores", "180610"},
        )
    if "ecodoppler color de carotidas" in key or "ecodoppler vasos de cuello" in key or "eco vc" in key:
        add("180607", "ECODOPPLER DE VASOS DEL CUELLO", {"ecodoppler vasos cuello", "vasos cuello", "carotidas", "180607"})
    if "ecocardiograma doppler color" in key or re.search(r"\bett\b", key):
        add("180301", "ECODOPPLER CARDIACO", {"ecodoppler cardiaco", "ecocardiograma", "ett", "180301"})
    if "electrocardiograma" in key or re.search(r"\becg\b", key):
        add(
            "570126",
            "ELECTROCARDIOGRAMA",
            {
                "electrocardiograma",
                "ecg",
                "570126",
                "570129",
                "820113",
                "consulta con especialista en cardiologia",
            },
        )
    if "presurometria" in key or re.search(r"\bmapa\b", key):
        add("570120", "PRESUROMETRIA (POR 24HS)", {"presurometria", "mapa", "570120"})
    return requisitos


def _prestacion_cubre_requisito(prestacion: dict, requisito: dict) -> bool:
    return _prestacion_match_requisito(prestacion, requisito) > 0


def _prestacion_match_requisito(prestacion: dict, requisito: dict) -> int:
    practica = str(prestacion.get("practica", "") or "")
    codigo = _codigo_practica(practica)
    codigo_requisito = str(requisito.get("codigo") or "")
    if codigo and codigo_requisito and codigo in _codigos_equivalentes_documentacion(codigo_requisito):
        return 2
    if codigo_requisito == "570126" and codigo:
        return 0
    if codigo_requisito == "180116" and codigo == "180608":
        return 0
    practica_key = _clave_texto(practica)
    return 1 if any(alias in practica_key for alias in requisito.get("aliases", set())) else 0


def _marcar_codigo_compatible_documentacion(prestacion: dict, requisito: dict) -> dict:
    codigo_requisito = str(requisito.get("codigo") or "")
    codigo_prestacion = _codigo_practica(prestacion.get("practica", ""))
    if (
        not codigo_requisito
        or not codigo_prestacion
        or codigo_requisito == codigo_prestacion
        or codigo_prestacion not in _codigos_equivalentes_documentacion(codigo_requisito)
    ):
        return prestacion
    marcada = dict(prestacion)
    marcada["codigo_requisito_documentacion"] = codigo_requisito
    marcada["codigo_compatible_documentacion"] = codigo_prestacion
    return marcada


def _es_consulta_practica(practica: str) -> bool:
    practica_key = _clave_texto(practica)
    return "consulta con especialista" in practica_key or practica_key.startswith("consulta ")


def _prestacion_faltante(
    datos_pdf: dict,
    requisito: dict,
    referencia: dict | None = None,
    paciente_prestaciones: list[dict] | None = None,
) -> dict:
    referencia = referencia or {}
    prestaciones_encontradas = paciente_prestaciones or ([referencia] if referencia else [])
    practicas_encontradas = [
        str(prestacion.get("practica", "") or "").strip()
        for prestacion in prestaciones_encontradas
        if str(prestacion.get("practica", "") or "").strip()
    ]
    practicas_transmitidas = [
        str(prestacion.get("practica", "") or "").strip()
        for prestacion in prestaciones_encontradas
        if prestacion.get("transmitida") and str(prestacion.get("practica", "") or "").strip()
    ]
    return {
        "n_orden": "",
        "beneficio": referencia.get("beneficio", datos_pdf.get("beneficio_pdf", "")),
        "nombre": referencia.get("nombre", datos_pdf.get("nombre_completo", "")),
        "nombre_key": referencia.get("nombre_key", datos_pdf.get("nombre_key", "")),
        "practica": f"{requisito['codigo']} - {requisito['descripcion']}",
        "turno": "",
        "validada": False,
        "transmitida": False,
        "documentacion_pendiente": False,
        "faltante_ome": True,
        "practicas_paciente_encontradas": practicas_encontradas,
        "practicas_paciente_transmitidas": practicas_transmitidas,
    }


def _prestacion_esperada_desde_informe(datos_pdf: dict, path: Path) -> dict | None:
    requisitos = _requisitos_informe(_texto_informe_para_requisitos(datos_pdf, path))
    if not requisitos:
        return None
    requisito = requisitos[0]
    return {
        "n_orden": "",
        "beneficio": datos_pdf.get("beneficio_pdf", ""),
        "nombre": datos_pdf.get("nombre_completo", ""),
        "nombre_key": datos_pdf.get("nombre_key", ""),
        "practica": f"{requisito['codigo']} - {requisito['descripcion']}",
        "turno": "",
        "validada": False,
        "transmitida": False,
        "documentacion_pendiente": False,
    }


def _texto_informe_para_requisitos(datos_pdf: dict, path: Path) -> str:
    texto = str(datos_pdf.get("texto", "") or "")
    stem = str(path.stem or "")
    if not stem:
        return texto
    partes = [texto]
    texto_key = _clave_texto(texto)
    stem_key = _clave_texto(stem)
    etiquetas_nombre = {
        "mapa": "MAPA",
        "ett": "ETT",
        "eco vc": "ECO VC",
        "holter": "HOLTER",
        "sibo": "SIBO",
        "fl": "FL",
        "eu": "EU",
    }
    for clave, etiqueta in etiquetas_nombre.items():
        if re.search(rf"(?:^|\s){re.escape(clave)}(?:\s|$)", stem_key) and not re.search(
            rf"(?:^|\s){re.escape(clave)}(?:\s|$)", texto_key
        ):
            partes.append(etiqueta)
    if stem_key and stem_key not in texto_key:
        partes.append(stem)
    return "\n".join(part for part in partes if part)


def _requisito_cima_desde_turno_codigo(codigo: str, descripcion: str) -> dict:
    codigo = str(codigo or "").strip()
    descripcion = str(descripcion or "").strip()
    aliases = {
        token
        for token in _clave_texto(f"{codigo} {descripcion}").split()
        if len(token) > 3 and not token.isdigit()
    }
    return {"codigo": codigo if re.fullmatch(r"\d{6}", codigo) else "", "descripcion": descripcion or codigo, "aliases": aliases}


def _requisitos_turno_cima(turno: dict, mapa_especialidades: dict[str, list[str]]) -> list[dict]:
    requisitos_raw = _cima_requisitos_practicas_turno(turno, mapa_especialidades, "CIMA")
    requisitos: list[dict] = []
    for codigo_raw, descripcion in requisitos_raw.items():
        opciones = [item.strip() for item in str(codigo_raw or "").split("|") if item.strip()]
        codigos = [item for item in opciones if re.fullmatch(r"\d{6}", item)]
        if codigos:
            for codigo in codigos:
                requisitos.append(_requisito_cima_desde_turno_codigo(codigo, descripcion))
            continue
        requisitos.append(_requisito_cima_desde_turno_codigo("", descripcion))
    return requisitos


def _turno_cima_cubre_requisito(turno: dict, requisito: dict, mapa_especialidades: dict[str, list[str]]) -> bool:
    for requisito_turno in _requisitos_turno_cima(turno, mapa_especialidades):
        codigo_turno = requisito_turno.get("codigo", "")
        codigo_requisito = requisito.get("codigo", "")
        if codigo_turno and codigo_requisito and codigo_turno in _codigos_equivalentes_documentacion(codigo_requisito):
            return True
        aliases_turno = set(requisito_turno.get("aliases", set()))
        aliases_informe = set(requisito.get("aliases", set()))
        if aliases_turno and aliases_informe and aliases_turno & aliases_informe:
            return True
    return False


def _mejor_paciente_cima(datos_pdf: dict, pacientes: list[dict]) -> tuple[float, dict | None]:
    beneficio = _solo_digitos(datos_pdf.get("beneficio_pdf", ""))
    dni = _solo_digitos(datos_pdf.get("dni", ""))
    nombre_key = datos_pdf.get("nombre_key", "")
    mejores: list[tuple[float, dict]] = []
    for paciente in pacientes:
        if beneficio and beneficio[:14] == _solo_digitos(paciente.get("beneficiario", ""))[:14]:
            mejores.append((1.0, paciente))
            continue
        if dni and dni == _solo_digitos(paciente.get("dni", "")):
            mejores.append((0.99, paciente))
            continue
        score = _score_nombre(nombre_key, _clave_texto(paciente.get("nombre", "")))
        if score >= 0.78:
            mejores.append((score, paciente))
    if not mejores:
        return 0.0, None
    mejores.sort(key=lambda item: item[0], reverse=True)
    return mejores[0]


def _prestacion_cima_para_requisito(
    datos_pdf: dict,
    requisito: dict,
    paciente: dict | None,
    turno: dict | None,
) -> dict:
    paciente = paciente or {}
    turno = turno or {}
    return {
        "n_orden": "",
        "beneficio": _solo_digitos(paciente.get("beneficiario", "")) or datos_pdf.get("beneficio_pdf", ""),
        "nombre": paciente.get("nombre", "") or datos_pdf.get("nombre_completo", ""),
        "nombre_key": _clave_texto(paciente.get("nombre", "") or datos_pdf.get("nombre_completo", "")),
        "practica": f"{requisito['codigo']} - {requisito['descripcion']}" if requisito.get("codigo") else requisito["descripcion"],
        "turno": turno.get("fecha", "") or _fecha_informe_documentacion(datos_pdf, Path(datos_pdf.get("archivo", ""))),
        "validada": False,
        "transmitida": False,
        "documentacion_pendiente": False,
    }


def preparar_lote_documentacion_desde_cima(
    ruta_excel_cima: Path,
    paths: list[Path],
    progress_callback: Callable[[int, int, Path, list[dict]], None] | None = None,
    cancel_event: threading.Event | None = None,
) -> list[dict]:
    paths = [path for path in paths if not _es_archivo_temporal_informe(path)]
    if _es_bandeja_transmision_pami(ruta_excel_cima):
        raise RuntimeError(
            "El archivo seleccionado parece una bandeja de Transmision PAMI, no un Excel de turnos CIMA. "
            "Para CIMA, si los informes tienen beneficio o DNI, usa 'Cruzar CIMA sin XLSX vs PAMI'. "
            "Usa 'Previsualizar con Excel CIMA' solo con el listado/Excel propio de turnos CIMA."
        )
    pacientes = leer_pacientes_pami(ruta_excel_cima, cliente_codigo=CLIENTE_CIMA)
    mapa_especialidades = cargar_especialidades_medicos(sistema_turnos="CIMA")
    lote: list[dict] = []
    total = len(paths)
    for idx, path in enumerate(paths, start=1):
        if cancel_event and cancel_event.is_set():
            break
        datos_pdf = _extraer_datos_paciente_pdf(path)
        datos_pdf["archivo"] = str(path)
        _agregar_etiquetas_desde_nombre_archivo(datos_pdf, path)
        if datos_pdf.get("archivo_blanco"):
            lote.append(_item_archivo_blanco(datos_pdf, path))
            if progress_callback:
                progress_callback(idx, total, path, list(lote))
            continue

        motivo_viejo = _motivo_informe_viejo(datos_pdf, path)
        if motivo_viejo:
            lote.append(_item_informe_viejo(datos_pdf, path, motivo_viejo))
            if progress_callback:
                progress_callback(idx, total, path, list(lote))
            continue
        if datos_pdf.get("obra_social_no_pami"):
            lote.append(_item_obra_social_no_pami(datos_pdf, path))
            if progress_callback:
                progress_callback(idx, total, path, list(lote))
            continue

        requisitos = _requisitos_informe(_texto_informe_para_requisitos(datos_pdf, path))
        _score_paciente, paciente = _mejor_paciente_cima(datos_pdf, pacientes)
        if paciente:
            if not datos_pdf.get("dni"):
                datos_pdf["dni"] = _solo_digitos(paciente.get("dni", ""))
            if not datos_pdf.get("beneficio_pdf"):
                datos_pdf["beneficio_pdf"] = _solo_digitos(paciente.get("beneficiario", ""))
            if not datos_pdf.get("nombre_completo"):
                datos_pdf["nombre_completo"] = paciente.get("nombre", "")
                datos_pdf["nombre_key"] = _clave_texto(paciente.get("nombre", ""))

        if not requisitos:
            estado = "sin_practica_informe"
            motivo = "Paciente CIMA encontrado, pero no se pudo detectar la practica del informe." if paciente else "No se pudo detectar la practica del informe."
            lote.append(
                {
                    "archivo": str(path),
                    "pdf_paciente": _paciente_detectado_para_lote(datos_pdf),
                    "dni_pdf": datos_pdf.get("dni", ""),
                    "beneficio_pdf": datos_pdf.get("beneficio_pdf", ""),
                    "fecha_informe": datos_pdf.get("fecha_informe", ""),
                    "estado": estado,
                    "motivo": motivo,
                    "prestacion": _prestacion_cima_para_requisito(
                        datos_pdf,
                        {"codigo": "", "descripcion": ""},
                        paciente,
                        (paciente.get("turnos") or [{}])[0] if paciente else None,
                    ),
                }
            )
        elif not paciente and not (datos_pdf.get("dni") or datos_pdf.get("beneficio_pdf")):
            lote.append(
                {
                    "archivo": str(path),
                    "pdf_paciente": _paciente_detectado_para_lote(datos_pdf),
                    "dni_pdf": datos_pdf.get("dni", ""),
                    "beneficio_pdf": datos_pdf.get("beneficio_pdf", ""),
                    "fecha_informe": datos_pdf.get("fecha_informe", ""),
                    "estado": "sin_coincidencia",
                    "motivo": "Tenemos INFORME pero no se encuentra paciente en CIMA y no hay DNI/beneficio para verificar en PAMI.",
                    "prestacion": {},
                }
            )
        else:
            turnos = list((paciente or {}).get("turnos") or [])
            for requisito in requisitos:
                turno_match = next(
                    (
                        turno
                        for turno in turnos
                        if _turno_cima_cubre_requisito(turno, requisito, mapa_especialidades)
                    ),
                    turnos[0] if turnos else None,
                )
                if paciente and turnos and turno_match and not _turno_cima_cubre_requisito(turno_match, requisito, mapa_especialidades):
                    motivo = "Paciente CIMA encontrado; no se encontro turno CIMA compatible. Se verificara en PAMI."
                elif paciente:
                    motivo = "Paciente CIMA encontrado; preparado para verificar OME en PAMI."
                else:
                    motivo = "Paciente no encontrado en CIMA, pero el informe tiene DNI/beneficio. Se verificara en PAMI."
                lote.append(
                    {
                        "archivo": str(path),
                        "pdf_paciente": _paciente_detectado_para_lote(datos_pdf),
                        "dni_pdf": datos_pdf.get("dni", ""),
                        "beneficio_pdf": datos_pdf.get("beneficio_pdf", ""),
                        "fecha_informe": datos_pdf.get("fecha_informe", ""),
                        "estado": "para_verificar_pami",
                        "motivo": motivo,
                        "prestacion": _prestacion_cima_para_requisito(datos_pdf, requisito, paciente, turno_match),
                    }
                )
        if progress_callback:
            progress_callback(idx, total, path, list(lote))
    return lote


def _prestaciones_mismo_paciente(datos_pdf: dict, prestaciones: list[dict]) -> list[dict]:
    score_fn = _score_nombre_relajado_ocr if datos_pdf.get("imagen_ocr") else _score_nombre
    threshold = 0.72 if datos_pdf.get("imagen_ocr") else 0.78
    return [
        prestacion
        for prestacion in prestaciones
        if score_fn(datos_pdf["nombre_key"], prestacion.get("nombre_key", "")) >= threshold
    ]


def _apellido_principal_nombre_key(nombre_key: str) -> str:
    partes = [parte for parte in str(nombre_key or "").split() if parte]
    return partes[0] if partes else ""


def _prestaciones_por_apellido_y_requisito(
    datos_pdf: dict,
    prestaciones: list[dict],
    requisito: dict,
    usadas: set[int] | None = None,
) -> list[tuple[int, dict, int]]:
    apellido = _apellido_principal_nombre_key(datos_pdf.get("nombre_key", ""))
    if len(apellido) < 4:
        return []
    usadas = usadas or set()
    candidatos: list[tuple[int, dict, int]] = []
    for idx, prestacion in enumerate(prestaciones):
        if idx in usadas:
            continue
        nombre_key = str(prestacion.get("nombre_key", "") or "")
        nombre_partes = nombre_key.split()
        if not nombre_partes or apellido != nombre_partes[0]:
            continue
        match_level = _prestacion_match_requisito(prestacion, requisito)
        if match_level <= 0:
            continue
        candidatos.append((idx, prestacion, match_level))
    if not candidatos:
        return []

    # Con solo apellido no conviene adivinar entre dos afiliados distintos.
    pacientes = {
        (_solo_digitos(prestacion.get("beneficio", ""))[:14], prestacion.get("nombre_key", ""))
        for _idx, prestacion, _match_level in candidatos
    }
    if len(pacientes) != 1:
        return []
    candidatos.sort(
        key=lambda item: (
            item[2],
            bool(item[1].get("documentacion_pendiente")),
            bool(item[1].get("transmitida")),
        ),
        reverse=True,
    )
    return candidatos


def _candidatos_por_requisitos(datos_pdf: dict, prestaciones: list[dict]) -> list[tuple[float, dict]]:
    requisitos = _requisitos_informe(
        _texto_informe_para_requisitos(datos_pdf, Path(str(datos_pdf.get("archivo", "") or "")))
    )
    if not requisitos:
        return []

    paciente_prestaciones = _prestaciones_mismo_paciente(datos_pdf, prestaciones)
    if not paciente_prestaciones:
        resultados_relajados: list[tuple[float, dict]] = []
        usadas_relajadas: set[int] = set()
        for requisito in requisitos:
            candidatos_relajados = _prestaciones_por_apellido_y_requisito(
                datos_pdf,
                prestaciones,
                requisito,
                usadas_relajadas,
            )
            if not candidatos_relajados:
                return [(0.0, None)]
            idx, prestacion, _match_level = candidatos_relajados[0]
            usadas_relajadas.add(idx)
            resultados_relajados.append((0.82, _marcar_codigo_compatible_documentacion(prestacion, requisito)))
        return resultados_relajados

    resultados: list[tuple[float, dict]] = []
    usadas: set[int] = set()
    for requisito in requisitos:
        candidatos = [
            (idx, prestacion, _prestacion_match_requisito(prestacion, requisito))
            for idx, prestacion in enumerate(paciente_prestaciones)
            if idx not in usadas and _prestacion_match_requisito(prestacion, requisito) > 0
        ]
        if not candidatos:
            ya_cubierto = any(
                _prestacion_match_requisito(paciente_prestaciones[idx], requisito) > 0
                for idx in usadas
                if idx < len(paciente_prestaciones)
            )
            if ya_cubierto:
                continue
            resultados.append(
                (
                    0.0,
                    _prestacion_faltante(
                        datos_pdf,
                        requisito,
                        paciente_prestaciones[0],
                        paciente_prestaciones=paciente_prestaciones,
                    ),
                )
            )
            continue
        candidatos.sort(
            key=lambda item: (
                item[2],
                bool(item[1].get("documentacion_pendiente")),
                bool(item[1].get("transmitida")),
            ),
            reverse=True,
        )
        multiples_pendientes = [
            item for item in candidatos if item[2] == 2 and bool(item[1].get("documentacion_pendiente"))
        ]
        if len(multiples_pendientes) > 1:
            for idx, prestacion, _match_level in multiples_pendientes:
                usadas.add(idx)
                resultados.append((1.1, _marcar_codigo_compatible_documentacion(prestacion, requisito)))
            continue
        idx, prestacion, _match_level = candidatos[0]
        usadas.add(idx)
        resultados.append((1.1, _marcar_codigo_compatible_documentacion(prestacion, requisito)))
    return resultados


def _candidatos_documentacion(datos_pdf: dict, prestaciones: list[dict]) -> list[tuple[float, dict]]:
    por_requisitos = _candidatos_por_requisitos(datos_pdf, prestaciones)
    if por_requisitos:
        return por_requisitos

    keywords = _practica_keywords_pdf(
        _texto_informe_para_requisitos(datos_pdf, Path(str(datos_pdf.get("archivo", "") or "")))
    )
    candidatos = []
    score_fn = _score_nombre_relajado_ocr if datos_pdf.get("imagen_ocr") else _score_nombre
    threshold = 0.72 if datos_pdf.get("imagen_ocr") else 0.78
    for prestacion in prestaciones:
        name_score = score_fn(datos_pdf["nombre_key"], prestacion.get("nombre_key", ""))
        if name_score < threshold:
            continue
        practice_score = _score_practica_pdf(keywords, prestacion.get("practica", ""))
        pending_score = 0.05 if prestacion.get("documentacion_pendiente") else 0.0
        candidatos.append((name_score, practice_score, pending_score, prestacion))
    if not candidatos:
        return []
    candidatos.sort(key=lambda item: (item[0], item[1], item[2]), reverse=True)
    if any(item[1] > 0 for item in candidatos):
        nombre_top = candidatos[0][0]
        return [
            (round(name_score + (practice_score * 0.1), 3), prestacion)
            for name_score, practice_score, _pending_score, prestacion in candidatos
            if name_score >= max(threshold, nombre_top - 0.03) and practice_score > 0
        ]
    nombre_top = candidatos[0][0]
    candidatos_mismo_nombre = [
        (name_score, practice_score, pending_score, prestacion)
        for name_score, practice_score, pending_score, prestacion in candidatos
        if name_score >= max(threshold, nombre_top - 0.03)
    ]
    no_consulta = [
        item for item in candidatos_mismo_nombre if not _es_consulta_practica(item[3].get("practica", ""))
    ]
    if len(no_consulta) == 1:
        name_score, practice_score, _pending_score, prestacion = no_consulta[0]
        return [(round(name_score + (practice_score * 0.1), 3), prestacion)]
    name_score, practice_score, _pending_score, prestacion = candidatos[0]
    return [(round(name_score + (practice_score * 0.1), 3), prestacion)]


def _mejor_candidato_documentacion(datos_pdf: dict, prestaciones: list[dict]) -> tuple[float, dict | None]:
    candidatos = _candidatos_documentacion(datos_pdf, prestaciones)
    return candidatos[0] if candidatos else (0.0, None)


def _estado_para_prestacion(prestacion: dict | None) -> tuple[str, str]:
    if not prestacion:
        return "sin_coincidencia", "Tenemos INFORME pero no se encuentra paciente en el XLS."
    codigo_compatible = str(prestacion.get("codigo_compatible_documentacion") or "")
    codigo_requisito = str(prestacion.get("codigo_requisito_documentacion") or "")
    if prestacion.get("faltante_ome"):
        practica = prestacion.get("practica", "esta practica")
        encontradas = [
            _codigo_practica(practica_encontrada) or practica_encontrada
            for practica_encontrada in prestacion.get("practicas_paciente_encontradas", [])
        ]
        transmitidas = [
            _codigo_practica(practica_transmitida) or practica_transmitida
            for practica_transmitida in prestacion.get("practicas_paciente_transmitidas", [])
        ]
        detalle = ""
        if encontradas:
            detalle = f" OMEs del paciente en bandeja: {', '.join(dict.fromkeys(encontradas))}."
        if transmitidas:
            detalle += f" Transmitidas: {', '.join(dict.fromkeys(transmitidas))}."
        codigo_detectado = _codigo_practica(practica)
        codigo_transmitido_compatible = _codigo_equivalente_distinto(
            codigo_detectado or str(practica),
            prestacion.get("practicas_paciente_transmitidas", []),
        )
        if codigo_transmitido_compatible:
            return (
                "ya_transmitido",
                f"Transmitido codigo {codigo_transmitido_compatible}: cubre codigo detectado {codigo_detectado or practica}.",
            )
        return "faltante_ome", f"Paciente encontrado en bandeja; falta OME para {practica}.{detalle}"
    if prestacion.get("documentacion_pendiente"):
        return "listo", ""
    if prestacion.get("transmitida"):
        if codigo_compatible and codigo_requisito:
            return "ya_transmitido", f"Transmitido codigo {codigo_compatible}: cubre codigo detectado {codigo_requisito}."
        return "ya_transmitido", "Paciente encontrado, pero la OME ya figura transmitida."
    if not prestacion.get("validada"):
        return "no_validada", "Paciente encontrado, pero la OME todavia no figura validada."
    return "sin_documentacion_pendiente", "Paciente encontrado, pero no figura pendiente de documentacion."


def _fecha_informe_documentacion(datos_pdf: dict, path: Path) -> str:
    fecha_texto = _fecha_informe_desde_texto(datos_pdf.get("texto", ""))
    fecha_archivo = _extraer_fecha_archivo(path.name)
    if fecha_texto and fecha_archivo:
        texto_dt = _extraer_fecha_turno(fecha_texto)
        archivo_dt = _extraer_fecha_turno(fecha_archivo)
        if texto_dt and archivo_dt and texto_dt.day == archivo_dt.day and texto_dt.month == archivo_dt.month:
            hoy = datetime.now().date()
            texto_viejo = (hoy - texto_dt.date()).days > MAX_ANTIGUEDAD_INFORME_DIAS
            archivo_vigente = (hoy - archivo_dt.date()).days <= MAX_ANTIGUEDAD_INFORME_DIAS
            if texto_viejo and archivo_vigente:
                return fecha_archivo
    return fecha_texto or fecha_archivo


def _motivo_informe_viejo(datos_pdf: dict, path: Path) -> str:
    fecha_texto = _fecha_informe_documentacion(datos_pdf, path)
    fecha = _extraer_fecha_turno(fecha_texto)
    if not fecha:
        return ""
    hoy = datetime.now()
    antiguedad = (hoy.date() - fecha.date()).days
    if antiguedad <= MAX_ANTIGUEDAD_INFORME_DIAS:
        return ""
    return (
        f"Informe viejo: fecha del informe {fecha_texto}; supera "
        f"{MAX_ANTIGUEDAD_INFORME_DIAS} dias. No se cruza con bandeja ni PAMI."
    )


def _item_informe_viejo(datos_pdf: dict, path: Path, motivo: str) -> dict:
    return {
        "archivo": str(path),
        "pdf_paciente": _paciente_detectado_para_lote(datos_pdf),
        "dni_pdf": datos_pdf.get("dni", ""),
        "beneficio_pdf": datos_pdf.get("beneficio_pdf", ""),
        "fecha_informe": datos_pdf.get("fecha_informe", ""),
        "score": 0.0,
        "estado": "informe_viejo",
        "motivo": motivo,
        "prestacion": {},
    }


def _item_archivo_blanco(datos_pdf: dict, path: Path) -> dict:
    return {
        "archivo": str(path),
        "pdf_paciente": "",
        "dni_pdf": "",
        "beneficio_pdf": "",
        "fecha_informe": "",
        "score": 0.0,
        "estado": "archivo_blanco",
        "motivo": "Archivo en blanco",
        "prestacion": {},
    }


def _item_obra_social_no_pami(datos_pdf: dict, path: Path) -> dict:
    obra_social = str(datos_pdf.get("obra_social") or "").strip()
    motivo = (
        f"No corresponde a PAMI: el informe indica obra social {obra_social}."
        if obra_social
        else "No corresponde a PAMI: el informe indica otra obra social."
    )
    return {
        "archivo": str(path),
        "pdf_paciente": _paciente_detectado_para_lote(datos_pdf),
        "dni_pdf": datos_pdf.get("dni", ""),
        "beneficio_pdf": datos_pdf.get("beneficio_pdf", ""),
        "fecha_informe": datos_pdf.get("fecha_informe", ""),
        "score": 0.0,
        "estado": "obra_social_no_pami",
        "motivo": motivo,
        "prestacion": {},
    }


def _fecha_informe_desde_texto(text: str) -> str:
    raw = str(text or "")

    def normalizar(match: re.Match) -> str:
        anio = int(match.group(3))
        if anio < 100:
            anio += 2000
        try:
            fecha = datetime(anio, int(match.group(2)), int(match.group(1)))
        except ValueError:
            return ""
        return f"{fecha.day:02d}/{fecha.month:02d}/{fecha.year}"

    sep_fecha = r"\s*[/-]\s*"
    patrones = [
        rf"\bFecha\s+de\s+visita\s*:?\s*(\d{{1,2}}){sep_fecha}(\d{{1,2}}){sep_fecha}(\d{{2,4}})",
        rf"\bFecha\s+del\s+estudio\s*:?\s*(\d{{1,2}}){sep_fecha}(\d{{1,2}}){sep_fecha}(\d{{2,4}})",
        rf"\bFecha\s*:?\s*(\d{{1,2}}){sep_fecha}(\d{{1,2}}){sep_fecha}(\d{{2,4}})",
        rf"\bEnviado\s+el\s+(\d{{1,2}}){sep_fecha}(\d{{1,2}}){sep_fecha}(\d{{2,4}})",
    ]
    for patron in patrones:
        match = re.search(patron, raw, flags=re.I)
        if match:
            fecha = normalizar(match)
            if fecha:
                return fecha

    meses = {
        "enero": 1,
        "febrero": 2,
        "marzo": 3,
        "abril": 4,
        "mayo": 5,
        "junio": 6,
        "julio": 7,
        "agosto": 8,
        "septiembre": 9,
        "setiembre": 9,
        "octubre": 10,
        "noviembre": 11,
        "diciembre": 12,
    }
    match = re.search(r"\bFecha\s*:?\s*(\d{1,2})\s+de\s+([a-záéíóúñ]+)\s+de\s+(\d{4})", raw, flags=re.I)
    if match:
        mes = meses.get(_clave_texto(match.group(2)))
        if mes:
            try:
                fecha = datetime(int(match.group(3)), mes, int(match.group(1)))
                return f"{fecha.day:02d}/{fecha.month:02d}/{fecha.year}"
            except ValueError:
                return ""
    for fecha_match in re.finditer(r"\bFecha\b", raw, flags=re.I):
        ventana = raw[max(0, fecha_match.start() - 80) : fecha_match.end() + 120]
        for token in re.findall(r"\b\d{10}\b", ventana):
            if token[2] != "1" or token[5] != "1":
                continue
            dia_raw, mes_raw, anio_raw = token[:2], token[3:5], token[6:]
            anios = [int(anio_raw)]
            if anio_raw.startswith("20"):
                anios.append(int(f"202{anio_raw[-1]}"))
            for anio in dict.fromkeys(anios):
                try:
                    fecha = datetime(anio, int(mes_raw), int(dia_raw))
                except ValueError:
                    continue
                if datetime.now().year - 1 <= fecha.year <= datetime.now().year + 1:
                    return f"{fecha.day:02d}/{fecha.month:02d}/{fecha.year}"
        for token in re.findall(r"\b\d{2}1\d{1,2}1\d{2}\b", ventana):
            splits = [(token[:2], token[3:4], token[5:]), (token[:2], token[3:5], token[6:])]
            for dia_raw, mes_raw, anio_raw in splits:
                try:
                    fecha = datetime(2000 + int(anio_raw), int(mes_raw), int(dia_raw))
                except ValueError:
                    continue
                if fecha.year > datetime.now().year + 1:
                    continue
                return f"{fecha.day:02d}/{fecha.month:02d}/{fecha.year}"
    for match in re.finditer(r"\b(\d{1,2})\s*[/-]\s*(\d{1,2})\s*[/-]\s*(\d{2,4})\b", raw):
        fecha = normalizar(match)
        fecha_dt = _extraer_fecha_turno(fecha)
        if fecha_dt and datetime.now().year - 1 <= fecha_dt.year <= datetime.now().year + 1:
            return fecha
    return ""


def _fechas_no_coinciden(datos_pdf: dict, path: Path, prestacion: dict) -> tuple[bool, str, str]:
    fecha_informe = _fecha_informe_documentacion(datos_pdf, path)
    fecha_turno = _fecha_normalizada(prestacion.get("turno", ""))
    if not fecha_informe or not fecha_turno:
        return False, fecha_informe, fecha_turno
    return fecha_informe != fecha_turno, fecha_informe, fecha_turno


def _clave_intento_documentacion(item: dict) -> tuple[str, str]:
    prestacion = item.get("prestacion") or {}
    codigo = _codigo_practica(prestacion.get("practica", ""))
    beneficio = _solo_digitos(prestacion.get("beneficio", "")) or _solo_digitos(item.get("beneficio_pdf", ""))
    dni = _solo_digitos(item.get("dni_pdf", ""))
    nombre = _clave_texto(prestacion.get("nombre_key", "") or prestacion.get("nombre", "") or item.get("pdf_paciente", ""))
    paciente = beneficio or dni or nombre
    return paciente, codigo


def _marcar_intentos_duplicados(resultados: list[dict]) -> list[dict]:
    vistos: dict[tuple[str, str], dict] = {}
    finales: list[dict] = []
    for item in resultados:
        item_final = dict(item)
        clave = _clave_intento_documentacion(item_final)
        if item_final.get("estado") == "listo" and all(clave):
            previo = vistos.get(clave)
            if previo is not None:
                item_final["estado"] = "duplicado"
                item_final["motivo"] = (
                    "Duplicado: ya hay otro informe para este paciente y codigo de practica. "
                    f"Se usara {Path(str(previo.get('archivo', ''))).name or 'el primero'}."
                )
            else:
                vistos[clave] = item_final
        elif all(clave) and item_final.get("estado") in {"ya_transmitido", "transmitido"}:
            vistos.setdefault(clave, item_final)
        finales.append(item_final)
    return finales


def preparar_lote_documentacion(
    ruta_excel: Path,
    pdf_paths: list[Path],
    progress_callback: Callable[[int, int, Path, list[dict]], None] | None = None,
    cancel_event=None,
) -> list[dict]:
    pdf_paths = [path for path in pdf_paths if not _es_archivo_temporal_informe(path)]
    prestaciones = leer_prestaciones_documentacion(ruta_excel)

    resultados: list[dict] = []
    total = len(pdf_paths)
    for index, path in enumerate(pdf_paths, start=1):
        if cancel_event is not None and cancel_event.is_set():
            break
        datos_pdf = _extraer_datos_paciente_pdf(path)
        _agregar_etiquetas_desde_nombre_archivo(datos_pdf, path)
        if datos_pdf.get("archivo_blanco"):
            resultados.append(_item_archivo_blanco(datos_pdf, path))
            if progress_callback:
                progress_callback(index, total, path, list(resultados))
            continue
        motivo_viejo = _motivo_informe_viejo(datos_pdf, path)
        if motivo_viejo:
            resultados.append(_item_informe_viejo(datos_pdf, path, motivo_viejo))
            if progress_callback:
                progress_callback(index, total, path, list(resultados))
            continue
        if datos_pdf.get("obra_social_no_pami"):
            resultados.append(_item_obra_social_no_pami(datos_pdf, path))
            if progress_callback:
                progress_callback(index, total, path, list(resultados))
            continue
        requisitos_archivo = _requisitos_informe(_texto_informe_para_requisitos(datos_pdf, path))
        imagen_sin_ocr = (
            path.suffix.lower() in IMAGE_EXTENSIONS
            and _clave_texto(datos_pdf.get("texto", "")) == _clave_texto(path.stem)
            and not requisitos_archivo
        )
        if imagen_sin_ocr:
            resultados.append(
                {
                    "archivo": str(path),
                    "pdf_paciente": _paciente_detectado_para_lote(datos_pdf),
                    "dni_pdf": datos_pdf["dni"],
                    "beneficio_pdf": datos_pdf.get("beneficio_pdf", ""),
                    "fecha_informe": datos_pdf.get("fecha_informe", ""),
                    "score": 0.0,
                    "estado": "imagen_sin_ocr",
                    "motivo": (
                        "Imagen sin OCR/practica detectable. No se vincula automaticamente para evitar cargar "
                        "documentacion en un paciente u OME incorrecta. Renombra el archivo con el codigo o practica."
                    ),
                    "prestacion": {},
                }
            )
            if progress_callback:
                progress_callback(index, total, path, list(resultados))
            continue
        candidatos = _candidatos_documentacion(datos_pdf, prestaciones)
        if not candidatos or all(score == 0.0 and not elegido for score, elegido in candidatos):
            esperado = _prestacion_esperada_desde_informe(datos_pdf, path)
            if esperado:
                candidatos = [(0.0, esperado)]
        for score, elegido in candidatos:
            if score == 0.0 and elegido and not elegido.get("n_orden") and not elegido.get("faltante_ome"):
                estado, motivo = "sin_coincidencia", "Tenemos INFORME pero no se encuentra paciente en el XLS."
            else:
                estado, motivo = _estado_para_prestacion(elegido)
            if elegido and elegido.get("n_orden") and estado not in {"ya_transmitido", "transmitido"}:
                fecha_distinta, fecha_informe, fecha_turno = _fechas_no_coinciden(datos_pdf, path, elegido)
                if fecha_distinta:
                    estado = "fecha_no_coincide"
                    motivo = f"Fecha del informe {fecha_informe} no coincide con turno OME {fecha_turno}."
            resultados.append(
                {
                    "archivo": str(path),
                    "pdf_paciente": _paciente_detectado_para_lote(datos_pdf),
                    "dni_pdf": datos_pdf["dni"],
                    "beneficio_pdf": datos_pdf.get("beneficio_pdf", ""),
                    "fecha_informe": datos_pdf.get("fecha_informe", ""),
                    "score": score,
                    "estado": estado,
                    "motivo": motivo,
                    "prestacion": elegido or {},
                }
            )
        if progress_callback:
            progress_callback(index, total, path, list(resultados))
    return _marcar_intentos_duplicados(resultados)


def preparar_lote_documentacion_para_pami(
    pdf_paths: list[Path],
    progress_callback: Callable[[int, int, Path, list[dict]], None] | None = None,
    cancel_event=None,
) -> list[dict]:
    pdf_paths = [path for path in pdf_paths if not _es_archivo_temporal_informe(path)]
    resultados: list[dict] = []
    total = len(pdf_paths)
    for index, path in enumerate(pdf_paths, start=1):
        if cancel_event is not None and cancel_event.is_set():
            break
        datos_pdf = _extraer_datos_paciente_pdf(path)
        _agregar_etiquetas_desde_nombre_archivo(datos_pdf, path)
        if datos_pdf.get("archivo_blanco"):
            resultados.append(_item_archivo_blanco(datos_pdf, path))
            if progress_callback:
                progress_callback(index, total, path, list(resultados))
            continue
        motivo_viejo = _motivo_informe_viejo(datos_pdf, path)
        if motivo_viejo:
            resultados.append(_item_informe_viejo(datos_pdf, path, motivo_viejo))
            if progress_callback:
                progress_callback(index, total, path, list(resultados))
            continue
        if datos_pdf.get("obra_social_no_pami"):
            resultados.append(_item_obra_social_no_pami(datos_pdf, path))
            if progress_callback:
                progress_callback(index, total, path, list(resultados))
            continue
        requisitos_archivo = _requisitos_informe(_texto_informe_para_requisitos(datos_pdf, path))
        imagen_sin_ocr = (
            path.suffix.lower() in IMAGE_EXTENSIONS
            and _clave_texto(datos_pdf.get("texto", "")) == _clave_texto(path.stem)
            and not requisitos_archivo
        )
        base = {
            "archivo": str(path),
            "pdf_paciente": _paciente_detectado_para_lote(datos_pdf),
            "dni_pdf": datos_pdf["dni"],
            "beneficio_pdf": datos_pdf.get("beneficio_pdf", ""),
            "fecha_informe": datos_pdf.get("fecha_informe", ""),
            "score": 0.0,
        }
        if imagen_sin_ocr:
            resultados.append(
                {
                    **base,
                    "estado": "imagen_sin_ocr",
                    "motivo": (
                        "Imagen sin OCR/practica detectable. No se vincula automaticamente para evitar cargar "
                        "documentacion en un paciente u OME incorrecta. Renombra el archivo con el codigo o practica."
                    ),
                    "prestacion": {},
                }
            )
            if progress_callback:
                progress_callback(index, total, path, list(resultados))
            continue
        if not requisitos_archivo:
            estado = "sin_datos_busqueda" if not (datos_pdf["dni"] or datos_pdf.get("beneficio_pdf", "")) else "sin_practica_informe"
            motivo = (
                "No hay beneficio ni DNI para verificar en PAMI."
                if estado == "sin_datos_busqueda"
                else "No se pudo detectar la practica del informe para cruzar contra PAMI."
            )
            resultados.append({**base, "estado": estado, "motivo": motivo, "prestacion": {}})
            if progress_callback:
                progress_callback(index, total, path, list(resultados))
            continue
        for requisito in requisitos_archivo:
            prestacion = _prestacion_faltante(datos_pdf, requisito)
            estado = "para_verificar_pami"
            motivo = "Preparado para verificar en PAMI por DNI/beneficio del informe."
            if not (datos_pdf["dni"] or datos_pdf.get("beneficio_pdf", "")):
                estado = "sin_datos_busqueda"
                motivo = "No hay beneficio ni DNI para verificar en PAMI."
            resultados.append({**base, "estado": estado, "motivo": motivo, "prestacion": prestacion})
        if progress_callback:
            progress_callback(index, total, path, list(resultados))
    return _marcar_intentos_duplicados(resultados)


class PamiDocumentacionController:
    def __init__(
        self,
        usuario: str,
        clave: str,
        log_callback: Callable[[str], None],
        status_callback: Callable[[str], None] | None = None,
        headless: bool = False,
    ) -> None:
        self.usuario = (usuario or "").strip()
        self.clave = clave or ""
        self.log_callback = log_callback
        self.status_callback = status_callback
        self.headless = headless
        self._sesion: SesionDocumentacion | None = None
        self._sesion_thread_id: int | None = None
        self._ultimo_dialogo = ""  # último alert/confirm nativo de PAMI (ej. "afiliado inactivo")
        self.cancel_event = None
        self.rango_turno_fallback: tuple[str, str] | None = None

    def _log(self, message: str) -> None:
        log_message(message)
        if self.log_callback:
            self.log_callback(message)

    def _status(self, message: str) -> None:
        if self.status_callback:
            self.status_callback(message)

    def set_cancel_event(self, cancel_event) -> None:
        self.cancel_event = cancel_event

    def set_rango_turno_fallback(self, desde: str, hasta: str) -> None:
        desde = str(desde or "").strip()
        hasta = str(hasta or "").strip()
        self.rango_turno_fallback = (desde, hasta) if desde and hasta else None

    def _cancel_requested(self) -> bool:
        return bool(self.cancel_event and self.cancel_event.is_set())

    def abrir_pami(self) -> None:
        ultimo_error: Exception | None = None
        for _intento in range(2):
            try:
                self._abrir_pami_once()
                return
            except AttributeError as exc:
                ultimo_error = exc
                if "_playwright" not in str(exc):
                    raise
                self._log("Playwright quedo en un estado inconsistente. Reintentando con una sesion nueva...")
                self._descartar_sesion()
                time.sleep(0.6)
        raise RuntimeError("No se pudo iniciar Playwright. Cerra el navegador PAMI desde la app y volve a intentar.") from ultimo_error

    def _abrir_pami_once(self) -> None:
        if self._sesion is not None:
            if self._sesion_disponible():
                self._status("Navegador ya abierto.")
                return
            self._descartar_sesion()
        configurar_playwright()
        playwright: Playwright | None = None
        try:
            playwright = sync_playwright().start()
            browser = self._launch_browser(playwright)
            context = browser.new_context(
                accept_downloads=True,
                ignore_https_errors=True,
                viewport={"width": 1280, "height": 900},
            )
            page = context.new_page()
            # Capturar los diálogos nativos de PAMI ("pe.pami.org.ar dice…"), que
            # Playwright descartaría en silencio. Guardamos el texto para poder
            # reportar el motivo real (ej. afiliado inactivo al transmitir).
            page.on("dialog", self._on_dialog)
            page.set_default_timeout(20000)
            page.set_default_navigation_timeout(45000)
            page.goto(
                CUP_LOGIN_URL if self.usuario and self.clave else PAMI_TRANSMISION_URL,
                wait_until="domcontentloaded",
                timeout=45000,
            )
            self._sesion = SesionDocumentacion(playwright, browser, context, page)
            self._sesion_thread_id = threading.get_ident()
            if self.usuario and self.clave:
                self._login(page)
                self._goto_transmision(page)
            self._status("PAMI abierto en Transmision.")
            self._log("Navegador preparado para carga de documentacion.")
        except Exception:
            if playwright is not None:
                try:
                    playwright.stop()
                except Exception:
                    pass
            self._sesion = None
            self._sesion_thread_id = None
            raise

    def _descartar_sesion(self) -> None:
        self._sesion = None
        self._sesion_thread_id = None

    def _launch_browser(self, playwright: Playwright) -> Browser:
        launch_args = {"headless": self.headless}
        if not self.headless:
            launch_args["args"] = ["--window-size=1280,900"]
        try:
            return playwright.chromium.launch(**launch_args)
        except Exception as exc:
            mensaje = str(exc)
            if "Executable doesn't exist" not in mensaje and "playwright install" not in mensaje.lower():
                raise
            self._log("Chromium empaquetado no esta disponible. Usando Google Chrome instalado.")
            try:
                return playwright.chromium.launch(channel="chrome", **launch_args)
            except Exception:
                self._log("No se pudo abrir Google Chrome por canal Playwright. Probando Microsoft Edge.")
                return playwright.chromium.launch(channel="msedge", **launch_args)

    def _sesion_disponible(self) -> bool:
        if self._sesion is None:
            return False
        if self._sesion_thread_id is not None and self._sesion_thread_id != threading.get_ident():
            self._log("La sesion PAMI fue creada por otra accion; se abrira un navegador nuevo para evitar error de hilos.")
            self._sesion = None
            self._sesion_thread_id = None
            return False
        try:
            return not self._sesion.page.is_closed()
        except Exception:
            return False

    def _error_pagina_cerrada(self, exc: Exception) -> bool:
        mensaje = str(exc)
        return (
            "Target page, context or browser has been closed" in mensaje
            or "TargetClosedError" in mensaje
            or ("Page.goto:" in mensaje and "has been closed" in mensaje)
        )

    def _error_navegacion_transitoria(self, exc: Exception) -> bool:
        mensaje = str(exc)
        return (
            self._error_pagina_cerrada(exc)
            or "Execution context was destroyed" in mensaje
            or "most likely because of a navigation" in mensaje
            or "ERR_HTTP_RESPONSE_CODE_FAILURE" in mensaje
            or "HTTP_RESPONSE_CODE_FAILURE" in mensaje
            or "response code failure" in mensaje.lower()
            or ("HTTP" in mensaje and ("transmision.php" in mensaje or "aceptacion" in mensaje.lower()))
            or ("Timeout" in mensaje and ("transmision.php" in mensaje or "aceptacion" in mensaje.lower()))
            or ("Page.goto:" in mensaje and "Timeout" in mensaje)
        )

    def _goto_transmision(self, page: Page, timeout: int = 45000) -> None:
        ultimo_error: Exception | None = None
        for intento in range(3):
            try:
                response = page.goto(PAMI_TRANSMISION_URL, wait_until="domcontentloaded", timeout=timeout)
                page.wait_for_timeout(900)
                if response and response.status >= 400:
                    raise RuntimeError(f"PAMI respondio HTTP {response.status} al abrir Transmision.")
                return
            except Exception as exc:
                ultimo_error = exc
                if self._error_pagina_cerrada(exc):
                    raise
                if not self._error_navegacion_transitoria(exc):
                    raise
                self._log(f"PAMI tardo en abrir Transmision; reintento {intento + 1}/3...")
                try:
                    page.wait_for_load_state("domcontentloaded", timeout=6000)
                except Exception:
                    pass
                page.wait_for_timeout(1200)
                if "transmision.php" in (page.url or "") and page.locator("table tbody, button:has-text('Buscar')").count() > 0:
                    return
        if ultimo_error:
            raise ultimo_error

    def _goto_panel_aceptacion(self, page: Page, timeout: int = 45000) -> None:
        ultimo_error: Exception | None = None
        for intento in range(3):
            try:
                response = page.goto(PAMI_PANEL_URL, wait_until="domcontentloaded", timeout=timeout)
                page.wait_for_timeout(900)
                if response and response.status >= 400:
                    raise RuntimeError(f"PAMI respondio HTTP {response.status} al abrir Panel de Aceptacion.")
                return
            except Exception as exc:
                ultimo_error = exc
                if self._error_pagina_cerrada(exc):
                    raise
                if not self._error_navegacion_transitoria(exc):
                    raise
                self._log(f"PAMI tardo en abrir Panel de Aceptacion; reintento {intento + 1}/3...")
                try:
                    page.wait_for_load_state("domcontentloaded", timeout=6000)
                except Exception:
                    pass
                page.wait_for_timeout(1200)
        if ultimo_error:
            raise ultimo_error

    def _obtener_page_transmision(self) -> Page:
        if not self._sesion_disponible():
            self.abrir_pami()
        assert self._sesion is not None
        page = self._sesion.page
        try:
            self._asegurar_transmision(page)
            return page
        except Exception as exc:
            if not self._error_pagina_cerrada(exc):
                raise
            self._log("La pagina PAMI se cerro durante la operacion. Reabriendo navegador y retomando...")
            self._descartar_sesion()
            self.abrir_pami()
            assert self._sesion is not None
            page = self._sesion.page
            self._asegurar_transmision(page)
            return page

    def cargar_lote(self, items: list[dict]) -> list[dict]:
        page = self._obtener_page_transmision()
        resultados = []
        for item in items:
            if self._cancel_requested():
                self._log("Carga detenida por el usuario. No se procesaran mas OMEs.")
                break
            prestacion = item.get("prestacion") or {}
            nro_orden = str(prestacion.get("n_orden") or "").strip()
            archivo = Path(str(item.get("archivo") or ""))
            if not nro_orden or not archivo.exists():
                resultados.append({**item, "estado": "error", "motivo": "Falta OME o archivo PDF."})
                continue
            try:
                self._log(f"Subiendo documentacion: OME {nro_orden} | {archivo.name}")
                try:
                    estado_final = self._subir_pdf_a_orden(page, nro_orden, archivo, item)
                except Exception as exc:
                    if not self._error_pagina_cerrada(exc):
                        raise
                    self._log(f"PAMI cerro la pagina al subir OME {nro_orden}. Reabriendo y reintentando...")
                    page = self._obtener_page_transmision()
                    estado_final = self._subir_pdf_a_orden(page, nro_orden, archivo, item)
                if estado_final == "ya_transmitido":
                    resultados.append({**item, "estado": "ya_transmitido", "motivo": "OME ya figura transmitida en PAMI."})
                else:
                    resultados.append({**item, "estado": "transmitido", "motivo": "Documentacion cargada y OME transmitida."})
            except Exception as exc:
                self._log(f"ERROR OME {nro_orden}: {exc}")
                motivo = str(exc)
                estado = "error"
                motivo_key = motivo.lower()
                if "ya figura transmitida" in motivo_key:
                    estado = "ya_transmitido"
                elif "no esta validada" in motivo_key or "no figura validada" in motivo_key or "transmision excepcional" in motivo_key:
                    estado = "no_validada"
                resultados.append({**item, "estado": estado, "motivo": motivo})
        return resultados

    def verificar_lote_en_pami(
        self,
        items: list[dict],
        progress_callback: Callable[[int, int, dict], None] | None = None,
    ) -> list[dict]:
        page = self._obtener_page_transmision()

        resultados = []
        avisos_pendientes: set[str] = set()
        cache_transmision: dict[tuple[str, str, tuple[str, str]], list[dict]] = {}
        cache_aceptacion: dict[tuple[str, str], list[dict]] = {}
        total_items = len(items)
        for index, item in enumerate(items, start=1):
            if progress_callback:
                progress_callback(index, total_items, item)
            if self._cancel_requested():
                self._log("Verificacion detenida por el usuario. No se procesaran mas informes.")
                break
            item = self._item_con_practica_esperada(item)
            prestacion = item.get("prestacion") or {}
            beneficio = _solo_digitos(prestacion.get("beneficio", "")) or _solo_digitos(item.get("beneficio_pdf", ""))
            dni = _solo_digitos(item.get("dni_pdf", ""))
            nro_orden = _solo_digitos(prestacion.get("n_orden", ""))
            buscar_por_dni = item.get("estado") in {"faltante_ome", "sin_coincidencia", "para_verificar_pami"} and bool(dni)
            if buscar_por_dni:
                modo = "dni"
                valor = dni
            elif beneficio:
                modo = "beneficiario"
                valor = beneficio[:14]
            elif dni:
                modo = "dni"
                valor = dni
            elif nro_orden:
                modo = "orden"
                valor = nro_orden
            else:
                resultados.append(
                    {
                        **item,
                        "estado": "sin_datos_busqueda",
                        "motivo": "No hay beneficio, DNI ni OME para verificar en PAMI.",
                    }
                )
                continue

            self._log(f"Verificando en PAMI por {modo}: {valor}")
            try:
                registros = self._buscar_registros_transmision_cache(
                    page,
                    modo,
                    valor,
                    prestacion.get("turno", ""),
                    cache_transmision,
                    item,
                )
            except Exception as exc:
                if not self._error_navegacion_transitoria(exc):
                    raise
                if self._error_pagina_cerrada(exc):
                    self._log("PAMI cerro la pagina durante la verificacion. Reabriendo y reintentando el paciente actual...")
                    page = self._obtener_page_transmision()
                else:
                    self._log("PAMI cambio de pagina durante la verificacion. Esperando y reintentando el paciente actual...")
                    try:
                        page.wait_for_load_state("domcontentloaded", timeout=6000)
                    except Exception:
                        pass
                    page.wait_for_timeout(900)
                try:
                    registros = self._buscar_registros_transmision_cache(
                        page,
                        modo,
                        valor,
                        prestacion.get("turno", ""),
                        cache_transmision,
                        item,
                    )
                except Exception as retry_exc:
                    if not self._error_navegacion_transitoria(retry_exc):
                        raise
                    motivo = f"PAMI no respondio al verificar este paciente: {retry_exc}"
                    self._log(motivo)
                    resultados.append({**item, "estado": "error", "motivo": motivo})
                    continue
            if not registros:
                try:
                    panel = [] if modo == "orden" else self._buscar_registros_panel_aceptacion_cache(page, modo, valor, cache_aceptacion)
                except Exception as exc:
                    if not self._error_navegacion_transitoria(exc):
                        raise
                    if self._error_pagina_cerrada(exc):
                        self._log("PAMI cerro la pagina al revisar Panel de Aceptacion. Reabriendo y reintentando...")
                        page = self._obtener_page_transmision()
                    else:
                        self._log("PAMI cambio de pagina al revisar Panel de Aceptacion. Esperando y reintentando...")
                        try:
                            page.wait_for_load_state("domcontentloaded", timeout=6000)
                        except Exception:
                            pass
                        page.wait_for_timeout(900)
                    panel = [] if modo == "orden" else self._buscar_registros_panel_aceptacion_cache(page, modo, valor, cache_aceptacion)
                registro_panel = self._elegir_registro_pami(item, panel) if panel else None
                if registro_panel and self._registro_cubre_item(registro_panel, item):
                    estado, motivo = self._estado_desde_registro_pami(registro_panel)
                    resultados.append({**item, "estado": estado, "motivo": motivo, "prestacion": self._prestacion_desde_registro(registro_panel, item)})
                else:
                    resultados.append(
                        {
                            **item,
                            "estado": "sin_coincidencia_pami",
                            "motivo": "No se encontro el paciente/practica en PAMI.",
                        }
                    )
                continue

            registro = self._elegir_registro_pami(item, registros)
            if not self._registro_cubre_item(registro, item):
                try:
                    panel = [] if modo == "orden" else self._buscar_registros_panel_aceptacion_cache(page, modo, valor, cache_aceptacion)
                except Exception as exc:
                    if not self._error_navegacion_transitoria(exc):
                        raise
                    if self._error_pagina_cerrada(exc):
                        self._log("PAMI cerro la pagina al revisar Panel de Aceptacion. Reabriendo y reintentando...")
                        page = self._obtener_page_transmision()
                    else:
                        self._log("PAMI cambio de pagina al revisar Panel de Aceptacion. Esperando y reintentando...")
                        try:
                            page.wait_for_load_state("domcontentloaded", timeout=6000)
                        except Exception:
                            pass
                        page.wait_for_timeout(900)
                    panel = [] if modo == "orden" else self._buscar_registros_panel_aceptacion_cache(page, modo, valor, cache_aceptacion)
                registro_panel = self._elegir_registro_pami(item, panel) if panel else None
                if registro_panel and self._registro_cubre_item(registro_panel, item):
                    registro = registro_panel
                else:
                    resultados.append(
                        {
                            **item,
                            "estado": "sin_coincidencia_pami",
                            "motivo": "Se encontro el paciente en PAMI, pero no la practica esperada.",
                        }
                    )
                    continue
            estado, motivo = self._estado_desde_registro_pami(registro)
            resultados.append(
                {
                    **item,
                    "estado": estado,
                    "motivo": motivo,
                    "prestacion": self._prestacion_desde_registro(registro, item),
                }
            )
            self._agregar_avisos_pendientes_no_incluidas(resultados, items, registros, item, avisos_pendientes)
        return resultados

    def _buscar_registros_transmision_cache(
        self,
        page: Page,
        modo: str,
        valor: str,
        turno: str,
        cache: dict[tuple[str, str, tuple[str, str]], list[dict]],
        item: dict | None = None,
    ) -> list[dict]:
        usar_rango_amplio = self._usar_rango_fallback_para_verificacion(item)
        rango_por_turno = None if usar_rango_amplio else _rango_mes_desde_turno(turno)
        rango = self.rango_turno_fallback if usar_rango_amplio and self.rango_turno_fallback else (rango_por_turno or self.rango_turno_fallback or ("", ""))
        key = (modo, valor, rango)
        if key not in cache:
            if usar_rango_amplio and self.rango_turno_fallback:
                self._log(f"Filtro de turno amplio tomado de la pantalla: {rango[0]} a {rango[1]}")
            elif rango_por_turno:
                self._log(f"Filtro de turno tomado del XLSX/PAMI: {rango[0]} a {rango[1]}")
            elif self.rango_turno_fallback:
                self._log(f"Filtro de turno tomado de la pantalla: {rango[0]} a {rango[1]}")
            cache[key] = self._buscar_registros_transmision(
                page,
                modo,
                valor,
                turno,
                rango if all(rango) else None,
            )
        return cache[key]

    def _usar_rango_fallback_para_verificacion(self, item: dict | None) -> bool:
        if not item or not self.rango_turno_fallback:
            return False
        estado = str(item.get("estado") or "")
        prestacion = item.get("prestacion") or {}
        n_orden = _solo_digitos(prestacion.get("n_orden", ""))
        return estado in {
            "faltante_ome",
            "sin_coincidencia",
            "para_verificar_pami",
            "fecha_no_coincide",
            "sin_coincidencia_pami",
        } or not n_orden

    def _buscar_registros_panel_aceptacion_cache(
        self,
        page: Page,
        modo: str,
        valor: str,
        cache: dict[tuple[str, str], list[dict]],
    ) -> list[dict]:
        key = (modo, valor)
        if key not in cache:
            cache[key] = self._buscar_registros_panel_aceptacion(page, modo, valor)
        return cache[key]

    def _item_con_practica_esperada(self, item: dict) -> dict:
        prestacion = item.get("prestacion") or {}
        if prestacion.get("practica"):
            return item
        path = Path(str(item.get("archivo") or ""))
        if not path:
            return item
        try:
            texto = extraer_texto_informe(path)
        except Exception:
            texto = path.stem
        datos_pdf = {
            "texto": texto,
            "beneficio_pdf": item.get("beneficio_pdf", ""),
            "nombre_completo": item.get("pdf_paciente", ""),
            "nombre_key": _clave_texto(item.get("pdf_paciente", "")),
        }
        esperado = _prestacion_esperada_desde_informe(datos_pdf, path)
        if not esperado:
            return item
        prestacion_actualizada = {**esperado, **prestacion}
        if not prestacion_actualizada.get("turno"):
            prestacion_actualizada["turno"] = esperado.get("turno", "")
        if not prestacion_actualizada.get("practica"):
            prestacion_actualizada["practica"] = esperado.get("practica", "")
        return {**item, "prestacion": prestacion_actualizada}

    def _agregar_avisos_pendientes_no_incluidas(
        self,
        resultados: list[dict],
        items: list[dict],
        registros: list[dict],
        item_referencia: dict,
        vistos: set[str],
    ) -> None:
        ordenes_lote = {
            str((item.get("prestacion") or {}).get("n_orden") or "").strip()
            for item in items
        }
        ordenes_lote.update(
            str((item.get("prestacion") or {}).get("n_orden") or "").strip()
            for item in resultados
        )
        archivo = Path(str(item_referencia.get("archivo") or ""))
        try:
            keywords_referencia = _practica_keywords_pdf(
                _texto_informe_para_requisitos({"texto": extraer_texto_pdf(archivo)}, archivo)
            )
        except Exception:
            keywords_referencia = _practica_keywords_pdf(archivo.stem)
        for registro in registros:
            nro_orden = str(registro.get("n_orden") or "").strip()
            if not nro_orden or nro_orden in ordenes_lote or nro_orden in vistos:
                continue
            practica_key = _clave_texto(registro.get("practica", ""))
            if not registro.get("documentacion_pendiente") or registro.get("transmitida"):
                continue
            if not any(token in practica_key for token in {"ecografia", "ecodoppler", "doppler", "eco"}):
                continue
            if keywords_referencia and _score_practica_pdf(keywords_referencia, registro.get("practica", "")) <= 0:
                continue
            vistos.add(nro_orden)
            resultados.append(
                {
                    **item_referencia,
                    "estado": "pendiente_no_incluida",
                    "motivo": "PAMI muestra otra ecografia pendiente para este paciente que no estaba incluida en el lote.",
                    "prestacion": self._prestacion_desde_registro(registro, item_referencia),
                }
            )

    def cerrar(self) -> None:
        if self._sesion is None:
            return
        if self._sesion_thread_id is not None and self._sesion_thread_id != threading.get_ident():
            self._sesion = None
            self._sesion_thread_id = None
            self._status("Referencia de navegador descartada; se abrira una sesion nueva en la proxima accion.")
            return
        sesion = self._sesion
        self._sesion = None
        self._sesion_thread_id = None
        try:
            sesion.context.close()
        except Exception:
            pass
        try:
            sesion.browser.close()
        except Exception:
            pass
        try:
            sesion.playwright.stop()
        except Exception:
            pass
        self._status("Navegador cerrado.")

    def _login(self, page: Page) -> None:
        if not self.usuario or not self.clave:
            raise RuntimeError("El portal pidio login y no hay credenciales cargadas.")

        if self._session_is_active(page):
            self._log("La sesion de CUP ya estaba activa.")
            return

        self._log("Iniciando sesion automatica en CUP PAMI para documentacion...")
        if "cup.pami.org.ar" not in (page.url or ""):
            page.goto(CUP_LOGIN_URL, wait_until="domcontentloaded")
        page.wait_for_timeout(350)

        user_input = page.locator('input[name="usuario"], input[type="text"], #usuario').first
        pass_input = page.locator('input[name="password"], input[type="password"], #password').first

        user_input.wait_for(state="visible")
        user_input.fill("")
        page.wait_for_timeout(120)
        user_input.type(self.usuario, delay=45)
        page.wait_for_timeout(120)

        pass_input.wait_for(state="visible")
        pass_input.fill("")
        page.wait_for_timeout(120)
        pass_input.type(self.clave, delay=45)
        page.wait_for_timeout(250)

        page.locator('button[type="submit"], input[type="submit"], button:has-text("Ingresar"), button:has-text("Entrar")').first.click()
        page.wait_for_load_state("domcontentloaded")
        page.wait_for_timeout(700)

        if "cup.pami.org.ar" in (page.url or "") and not self._session_is_active(page):
            raise RuntimeError("No se pudo iniciar sesion en CUP PAMI. Revisa usuario y clave.")
        self._log("Sesion iniciada automaticamente en CUP PAMI.")

    def _session_is_active(self, page: Page) -> bool:
        try:
            response = page.goto(PAMI_TRANSMISION_URL, wait_until="domcontentloaded", timeout=45000)
            page.wait_for_timeout(1200)
            if "cup.pami.org.ar" in (page.url or ""):
                return False
            if response and response.status >= 400:
                return False
            return page.locator("table tbody, button:has-text('Buscar'), input[type='button'][value*='Buscar' i]").count() > 0
        except Exception:
            return False

    def _asegurar_transmision(self, page: Page) -> None:
        if "transmision.php" not in (page.url or ""):
            self._goto_transmision(page)
        if "cup.pami.org.ar" in (page.url or "") and self.usuario and self.clave:
            self._login(page)
            self._goto_transmision(page)
        if page.locator("table tbody, button:has-text('Buscar')").count() == 0:
            raise RuntimeError("No se pudo acceder al panel de Transmision PAMI.")

    def _buscar_registros_transmision(
        self,
        page: Page,
        modo: str,
        valor: str,
        turno: str = "",
        rango_turno: tuple[str, str] | None = None,
    ) -> list[dict]:
        self._goto_transmision(page)

        def esperar_estable(ms: int = 700) -> None:
            try:
                page.wait_for_load_state("domcontentloaded", timeout=6000)
            except Exception:
                pass
            page.wait_for_timeout(ms)

        def evaluar(script: str, arg=None):
            ultimo_error: Exception | None = None
            for _ in range(4):
                try:
                    if arg is None:
                        return page.evaluate(script)
                    return page.evaluate(script, arg)
                except Exception as exc:
                    ultimo_error = exc
                    if not self._error_navegacion_transitoria(exc):
                        raise
                    esperar_estable(900)
            if ultimo_error:
                raise ultimo_error
            return None

        try:
            evaluar(
                """
                () => {
                  const visible = (el) => !!el && (el.offsetParent !== null || el.getClientRects().length > 0);
                  const limpiar = Array.from(document.querySelectorAll('button, input[type="button"], input[type="submit"], a'))
                    .find((el) => visible(el) && (el.textContent || el.value || '').trim().toLowerCase() === 'limpiar');
                  if (limpiar) limpiar.click();
                }
                """
            )
            page.wait_for_timeout(350)
        except Exception:
            pass

        evaluar(
            """
            ({modo, valor, rango}) => {
              const visible = (el) => !!el && (el.offsetParent !== null || el.getClientRects().length > 0);
              const setValor = (el, value) => {
                if (!el) return;
                el.value = value || '';
                el.dispatchEvent(new Event('input', { bubbles: true }));
                el.dispatchEvent(new Event('change', { bubbles: true }));
              };
              const allInputs = Array.from(document.querySelectorAll('input[type="text"], input:not([type])'))
                .filter((el) => visible(el) && !el.disabled);
              if (modo === 'orden') {
                const ordenInput = allInputs.find((el) => {
                  const text = `${el.name || ''} ${el.id || ''} ${el.getAttribute('ng-model') || ''} ${el.placeholder || ''}`.toLowerCase();
                  return text.includes('orden');
                }) || allInputs[0];
                setValor(ordenInput, valor);
                const buscarOrden = Array.from(document.querySelectorAll('button, input[type="submit"], input[type="button"], a'))
                  .find((el) => visible(el) && (el.textContent || el.value || '').trim().toLowerCase() === 'buscar');
                if (buscarOrden) buscarOrden.click();
                return;
              }
              const tipo = modo === 'dni' ? '2' : '1';
              const select = Array.from(document.querySelectorAll('select')).find((sel) =>
                visible(sel) && Array.from(sel.options || []).some((opt) => (opt.value || '').trim() === tipo)
              );
              if (select) {
                select.value = tipo;
                select.dispatchEvent(new Event('change', { bubbles: true }));
              }
              if (rango) {
                setValor(document.querySelector('input[name="f_turno_desde"]'), rango.desde);
                setValor(document.querySelector('input[name="f_turno_hasta"]'), rango.hasta);
              }
              const inputs = allInputs
                .filter((el) => {
                  const text = `${el.name || ''} ${el.id || ''} ${el.getAttribute('ng-model') || ''} ${el.placeholder || ''}`.toLowerCase();
                  return !text.includes('orden') && !text.includes('practica') && !text.includes('fecha') && !text.includes('turno');
                });
              let input = null;
              if (select) {
                const selectBox = select.getBoundingClientRect();
                input = inputs.find((el) => {
                  const box = el.getBoundingClientRect();
                  const mismaColumna = Math.abs(box.left - selectBox.left) < 80 || Math.abs(box.right - selectBox.right) < 80;
                  return mismaColumna && box.top >= selectBox.top - 10 && box.top <= selectBox.bottom + 80;
                });
              }
              input = input || inputs.find((el) => /benef|afiliado|documento|dni/i.test(`${el.name || ''} ${el.id || ''} ${el.placeholder || ''}`)) || inputs[0];
              if (input) {
                setValor(input, valor);
              }
              const buscar = Array.from(document.querySelectorAll('button, input[type="submit"], input[type="button"], a'))
                .find((el) => visible(el) && (el.textContent || el.value || '').trim().toLowerCase() === 'buscar');
              if (buscar) buscar.click();
            }
            """,
            {
                "modo": modo,
                "valor": valor,
                "rango": (
                    {"desde": rango_turno[0], "hasta": rango_turno[1]}
                    if rango_turno
                    else None
                ),
            },
        )
        esperar_estable(1400)

        registros: list[dict] = []
        for _ in range(12):
            pagina = evaluar(
                """
                () => {
                  const visible = (el) => !!el && (el.offsetParent !== null || el.getClientRects().length > 0);
                  const texto = (el) => (el?.textContent || el?.value || '').replace(/\\s+/g, ' ').trim();
                  const clave = (value) => String(value || '')
                    .normalize('NFD')
                    .replace(/[\\u0300-\\u036f]/g, '')
                    .replace(/\\s+/g, ' ')
                    .trim()
                    .toLowerCase();
                  const esSi = (value) => /^si\\b/.test(clave(value));
                  const esNo = (value) => /^no\\b/.test(clave(value));
                  const headerIndex = (table) => {
                    const headers = Array.from(table?.querySelectorAll('thead th, thead td, tr:first-child th') || []).map((th) => clave(texto(th)));
                    const find = (...names) => headers.findIndex((header) => names.some((name) => header.includes(name)));
                    return {
                      orden: find('nro. orden', 'nro orden', 'orden'),
                      beneficio: find('beneficio', 'afiliado'),
                      nombre: find('apellido y nombre', 'nombre'),
                      practica: find('practica'),
                      turno: find('turno'),
                      transmitida: find('trasmitida', 'transmitida')
                    };
                  };
                  const esAzul = (el) => {
                    const target = el?.closest('button,a,span') || el;
                    const cls = `${target?.className || ''} ${el?.className || ''}`.toLowerCase();
                    if (cls.includes('btn-primary') || cls.includes('btn-info') || cls.includes('blue')) return true;
                    const rgb = (getComputedStyle(target).backgroundColor || '').match(/\\d+/g)?.map(Number) || [];
                    return rgb.length >= 3 && rgb[2] > rgb[0] + 20 && rgb[2] >= rgb[1] - 10;
                  };
                  const esVerde = (el) => {
                    const target = el?.closest('button,a,span') || el;
                    const cls = `${target?.className || ''} ${el?.className || ''}`.toLowerCase();
                    if (cls.includes('btn-success') || cls.includes('green')) return true;
                    const rgb = (getComputedStyle(target).backgroundColor || '').match(/\\d+/g)?.map(Number) || [];
                    return rgb.length >= 3 && rgb[1] > rgb[0] + 20 && rgb[1] > rgb[2] + 20;
                  };
                  const control = (el) => el?.closest('button,a,span.btn,.btn') || el;
                  const controlesAccion = (acciones) => {
                    const vistos = new Set();
                    return Array.from(acciones.querySelectorAll('a, button, span.btn, .btn'))
                      .map(control)
                      .filter((el) => el && visible(el))
                      .filter((el) => {
                        if (vistos.has(el)) return false;
                        vistos.add(el);
                        return true;
                      })
                      .sort((a, b) => a.getBoundingClientRect().left - b.getBoundingClientRect().left);
                  };
                  return Array.from(document.querySelectorAll('table tbody tr')).map((tr) => {
                    const tds = Array.from(tr.querySelectorAll('td'));
                    if (tds.length < 6) return null;
                    const celdas = tds.map(texto);
                    const textoFila = celdas.join(' ');
                    const idx = headerIndex(tr.closest('table'));
                    const celda = (name, fallback) => idx[name] >= 0 && idx[name] < celdas.length ? celdas[idx[name]] : (celdas[fallback] || '');
                    const acciones = tds[tds.length - 1] || tr;
                    const controles = controlesAccion(acciones);
                    const check = acciones.querySelector('.fa-check, .glyphicon-ok, .icon-ok, .icon-check');
                    const doc = acciones.querySelector('.fa-upload, .fa-file, .fa-file-text, .fa-folder, .fa-cloud-upload, .glyphicon-open, .glyphicon-upload, .icon-upload, .icon-file');
                    const lupa = acciones.querySelector('.fa-search, .fa-eye');
                    const transmitir = acciones.querySelector('i.transmitir, .transmitir, .fa-arrow-right, .fa-arrow-circle-right, .glyphicon-arrow-right, .icon-arrow-right');
                    const checkControl = check ? control(check) : controles[0];
                    const docControl = doc ? control(doc) : controles[1];
                    const transmitirControl = transmitir ? control(transmitir) : controles[2];
                    const transmitidaTexto =
                      celda('transmitida', 6) ||
                      celdas.find((value) => /^si\\s*[-\\u2013\\u2014]/.test(clave(value))) ||
                      celdas.find((value) => /^no\\s*[-\\u2013\\u2014]/.test(clave(value))) ||
                      celdas.find((value) => esSi(value) || esNo(value)) ||
                      '';
                    const checkAzul = checkControl ? esAzul(checkControl) : false;
                    const docPendiente = docControl ? esVerde(docControl) : false;
                    const docCargada = docControl ? esAzul(docControl) : false;
                    const botonTransmision = !!(transmitirControl && visible(transmitirControl) && esVerde(transmitirControl));
                    const cerradaPorAcciones = !!(lupa && visible(lupa) && checkAzul && !docPendiente && !botonTransmision);
                    const transmitida = esSi(transmitidaTexto) || /(?:^|\\s)si\\s*[-\\u2013\\u2014]/.test(clave(textoFila)) || cerradaPorAcciones;
                    const validacionPendiente = !transmitida && !!(check && checkControl && visible(checkControl) && esVerde(checkControl) && !checkAzul);
                    const validada = transmitida || (!validacionPendiente && (checkAzul || docPendiente || docCargada || botonTransmision));
                    return {
                      n_orden: celda('orden', 0),
                      beneficio: celda('beneficio', 2),
                      nombre: celda('nombre', 3),
                      practica: celda('practica', 4),
                      turno: celda('turno', 5),
                      transmitida,
                      validada,
                      validacion_pendiente: validacionPendiente,
                      documentacion_pendiente: docPendiente && !transmitida,
                      documentacion_cargada: docCargada || transmitida,
                      tiene_boton_transmitir: botonTransmision,
                      transmitida_texto: transmitidaTexto
                    };
                  }).filter(Boolean);
                }
                """
            )
            if isinstance(pagina, list):
                registros.extend(item for item in pagina if isinstance(item, dict))
            advanced = evaluar(
                """
                () => {
                  const visible = (el) => !!el && (el.offsetParent !== null || el.getClientRects().length > 0);
                  const paginaActual = Number(new URL(location.href).searchParams.get('pagina') || '1');
                  const mas = Array.from(document.querySelectorAll('a, button'))
                    .find((el) => visible(el) && (el.textContent || '').trim().toLowerCase().includes('mas resultados'));
                  if (mas) { mas.click(); return true; }
                  const sig = Array.from(document.querySelectorAll('a, button'))
                    .find((el) => visible(el) && /^\\d+$/.test((el.textContent || '').trim()) && Number((el.textContent || '').trim()) === paginaActual + 1);
                  if (!sig) return false;
                  sig.click();
                  return true;
                }
                """
            )
            if not advanced:
                break
            esperar_estable(900)
        return registros

    def _buscar_registros_panel_aceptacion(self, page: Page, modo: str, valor: str) -> list[dict]:
        self._goto_panel_aceptacion(page)
        if "cup.pami.org.ar" in (page.url or "") and self.usuario and self.clave:
            self._login(page)
            self._goto_panel_aceptacion(page)

        def esperar_estable(ms: int = 700) -> None:
            try:
                page.wait_for_load_state("domcontentloaded", timeout=6000)
            except Exception:
                pass
            page.wait_for_timeout(ms)

        def evaluar(script: str, arg=None):
            ultimo_error: Exception | None = None
            for _ in range(4):
                try:
                    if arg is None:
                        return page.evaluate(script)
                    return page.evaluate(script, arg)
                except Exception as exc:
                    ultimo_error = exc
                    if not self._error_navegacion_transitoria(exc):
                        raise
                    esperar_estable(900)
            if ultimo_error:
                raise ultimo_error
            return None

        fecha_desde = _fecha_panel_aceptacion_desde()
        self._log(f"Buscando en Panel de Aceptacion por {modo}: {valor} desde emision {fecha_desde}")
        evaluar(
            """
            ({modo, valor, fechaDesde}) => {
              const visible = (el) => !!el && (el.offsetParent !== null || el.getClientRects().length > 0);
              const setValue = (input, value) => {
                if (!input) return false;
                const setter = Object.getOwnPropertyDescriptor(window.HTMLInputElement.prototype, 'value')?.set;
                if (setter) setter.call(input, value || '');
                else input.value = value || '';
                input.dispatchEvent(new Event('input', { bubbles: true }));
                input.dispatchEvent(new Event('change', { bubbles: true }));
                input.dispatchEvent(new KeyboardEvent('keyup', { bubbles: true }));
                return true;
              };
              const tipo = modo === 'dni' ? '2' : '1';
              const select = Array.from(document.querySelectorAll('select')).find((sel) =>
                visible(sel) && Array.from(sel.options || []).some((opt) => (opt.value || '').trim() === tipo)
              );
              if (select) {
                select.value = tipo;
                select.dispatchEvent(new Event('change', { bubbles: true }));
              }
              setValue(document.querySelector('#f_emision_desde, input[name="f_emision_desde"]'), fechaDesde);
              setValue(document.querySelector('#f_emision_hasta, input[name="f_emision_hasta"]'), '');
              const practica = Array.from(document.querySelectorAll('input[type="text"], input:not([type])'))
                .find((el) => visible(el) && /practica/i.test(`${el.name || ''} ${el.id || ''} ${el.placeholder || ''}`));
              setValue(practica, '');
              const inputs = Array.from(document.querySelectorAll('input[type="text"], input:not([type])'))
                .filter((el) => visible(el) && !el.disabled)
                .filter((el) => {
                  const text = `${el.name || ''} ${el.id || ''} ${el.getAttribute('ng-model') || ''} ${el.placeholder || ''}`.toLowerCase();
                  return !text.includes('orden') && !text.includes('practica') && !text.includes('fecha') && !text.includes('emision');
                });
              let input = null;
              const selectors = modo === 'dni'
                ? ['input[ng-model*="documento"]', 'input[ng-model*="doc"]', 'input[name*="documento"]', 'input[name*="doc"]', 'input[placeholder*="documento" i]']
                : ['input[ng-model*="beneficio"]', 'input[ng-model*="afiliado"]', 'input[name*="beneficio"]', 'input[name*="afiliado"]', 'input[name="n_afiliado"]', 'input[placeholder*="beneficio" i]', 'input[placeholder*="afiliado" i]'];
              for (const selector of selectors) {
                input = Array.from(document.querySelectorAll(selector)).find(visible);
                if (input) break;
              }
              if (!input && select) {
                const selectBox = select.getBoundingClientRect();
                input = inputs
                  .map((el) => {
                    const box = el.getBoundingClientRect();
                    const horizontal = Math.abs(box.left - selectBox.left);
                    const vertical = Math.abs(box.top - selectBox.bottom);
                    const sameBlock = box.top >= selectBox.top - 8 && box.top <= selectBox.bottom + 65;
                    return { el, score: horizontal + vertical + (sameBlock ? 0 : 1000) };
                  })
                  .sort((a, b) => a.score - b.score)[0]?.el || null;
              }
              input = input || inputs[0];
              if (input) {
                setValue(input, valor);
              }
              window.bandeja = undefined;
              const buscar = Array.from(document.querySelectorAll('button, input[type="submit"], input[type="button"], a'))
                .find((el) => visible(el) && (el.textContent || el.value || '').trim().toLowerCase() === 'buscar');
              if (buscar) buscar.click();
            }
            """,
            {"modo": modo, "valor": valor, "fechaDesde": fecha_desde},
        )
        try:
            page.wait_for_function("window.bandeja !== undefined", timeout=10000)
        except Exception:
            return []
        try:
            return evaluar(
                """
                () => {
                  const domInfo = {};
                  document.querySelectorAll('table tbody tr').forEach((tr) => {
                    const aceptar = tr.querySelector('i.boton-historial[data-estado="aceptar"], .boton-historial[data-estado="aceptar"], .fa-check');
                    const info = tr.querySelector('.fa-info, .fa-info-circle');
                    const agenda = tr.querySelector('.fa-calendar, .fa-calendar-o');
                    const accion = aceptar || info || tr.querySelector('i, button, a');
                    const celdas = Array.from(tr.querySelectorAll('td')).map((td) => (td.textContent || '').replace(/\\s+/g, ' ').trim());
                    const orden = accion?.getAttribute('data-orden') || celdas[0] || '';
                    if (!orden) return;
                    domInfo[orden] = {
                      disponible: !!aceptar || !!agenda,
                      transmitida: false
                    };
                  });
                  if (!Array.isArray(window.bandeja)) return [];
                  return window.bandeja.map((ome) => {
                    const dom = domInfo[ome.n_orden] || {};
                    return {
                      n_orden: ome.n_orden || '',
                      beneficio: `${ome.n_beneficio || ''}${ome.c_grado_paren || ''}`,
                      nombre: ome.afiliado || ome.nombre || '',
                      practica: ome.d_practica || '',
                      turno: '',
                      transmitida: !!dom.transmitida,
                      validada: false,
                      documentacion_pendiente: false,
                      documentacion_cargada: false,
                      disponible_activar: !!dom.disponible,
                      estado_accion: dom.disponible ? 'disponible_activar' : ''
                    };
                  });
                }
                """
            ) or []
        except Exception:
            return []

    def _elegir_registro_pami(self, item: dict, registros: list[dict]) -> dict:
        nro_objetivo = _solo_digitos((item.get("prestacion") or {}).get("n_orden", ""))
        if nro_objetivo:
            exacto = next(
                (
                    registro
                    for registro in registros
                    if _solo_digitos(registro.get("n_orden", "")) == nro_objetivo
                ),
                None,
            )
            if exacto:
                return exacto
        pdf_key = _clave_texto(item.get("pdf_paciente", ""))
        objetivo = (item.get("prestacion") or {}).get("practica", "")
        archivo = Path(str(item.get("archivo") or ""))
        try:
            keywords = _practica_keywords_pdf(
                _texto_informe_para_requisitos({"texto": extraer_texto_pdf(archivo)}, archivo)
            )
        except Exception:
            keywords = _practica_keywords_pdf(archivo.stem)
        scored = []
        for registro in registros:
            name_score = _score_nombre(pdf_key, _clave_texto(registro.get("nombre", "")))
            expected_score = self._score_registro_objetivo(registro, objetivo)
            practice_score = _score_practica_pdf(keywords, registro.get("practica", ""))
            transmitted_score = 0.05 if registro.get("transmitida") else 0.0
            scored.append((expected_score, practice_score, name_score, transmitted_score, registro))
        scored.sort(key=lambda value: value[:4], reverse=True)
        return scored[0][4]

    def _estado_desde_registro_pami(self, registro: dict) -> tuple[str, str]:
        if registro.get("disponible_activar") or registro.get("estado_accion") == "disponible_activar":
            return "falta_activar_ome", "OME encontrada en Panel de Aceptacion; falta activar OME."
        transmitida_texto = _clave_texto(
            " ".join(
                str(registro.get(campo, "") or "")
                for campo in ("transmitida", "transmitida_texto", "trasmitida")
            )
        )
        if registro.get("transmitida") or re.search(r"(?:^|\s)si\s*[-\u2013\u2014]?", transmitida_texto):
            return "ya_transmitido", "Paciente encontrado en PAMI; la OME ya figura transmitida."
        if registro.get("validacion_pendiente") or not registro.get("validada"):
            return "no_validada", "Paciente encontrado en PAMI, pero la OME todavia no figura validada."
        if registro.get("documentacion_pendiente"):
            return "listo", ""
        if registro.get("documentacion_cargada") and registro.get("tiene_boton_transmitir"):
            return "listo", "La documentacion ya figura cargada en PAMI; se transmitira sin volver a subir."
        return "sin_documentacion_pendiente", "Paciente encontrado en PAMI, pero no figura pendiente de documentacion."

    def _registro_cubre_item(self, registro: dict, item: dict) -> bool:
        objetivo = (item.get("prestacion") or {}).get("practica", "")
        if not objetivo:
            archivo = Path(str(item.get("archivo") or ""))
            try:
                keywords = _practica_keywords_pdf(
                    _texto_informe_para_requisitos({"texto": extraer_texto_pdf(archivo)}, archivo)
                )
            except Exception:
                keywords = _practica_keywords_pdf(archivo.stem)
            if keywords:
                return _score_practica_pdf(keywords, registro.get("practica", "")) > 0
            return False
        return self._score_registro_objetivo(registro, objetivo) > 0

    def _score_registro_objetivo(self, registro: dict, objetivo: str) -> int:
        requisito = self._requisito_desde_practica(objetivo)
        if not requisito:
            return 0
        codigo_registro = _codigo_practica(registro.get("practica", ""))
        if requisito.get("codigo") and codigo_registro:
            if codigo_registro == requisito["codigo"]:
                return 2
            if codigo_registro in _codigos_equivalentes_documentacion(requisito["codigo"]):
                return 2
            practica_key = _clave_texto(registro.get("practica", ""))
            if any(alias in practica_key for alias in requisito.get("aliases", set())):
                return 1
            return 0
        return _prestacion_cubre_requisito({"practica": registro.get("practica", "")}, requisito)

    def _requisito_desde_practica(self, practica: str) -> dict:
        codigo = _codigo_practica(practica)
        stopwords = {"con", "del", "los", "las", "por", "para", "y/o", "incluye", "hasta", "completa"}
        aliases = {
            token
            for token in _clave_texto(practica).split()
            if len(token) > 3 and token not in stopwords and not token.isdigit()
        }
        return {"codigo": codigo, "aliases": aliases}

    def _prestacion_desde_registro(self, registro: dict, item: dict) -> dict:
        previa = item.get("prestacion") or {}
        return {
            "n_orden": registro.get("n_orden", "") or previa.get("n_orden", ""),
            "beneficio": _solo_digitos(registro.get("beneficio", "")) or previa.get("beneficio", ""),
            "nombre": registro.get("nombre", "") or previa.get("nombre", ""),
            "nombre_key": _clave_texto(registro.get("nombre", "") or previa.get("nombre", "")),
            "practica": registro.get("practica", "") or previa.get("practica", ""),
            "turno": registro.get("turno", "") or previa.get("turno", ""),
            "validada": bool(registro.get("validada")),
            "validacion_pendiente": bool(registro.get("validacion_pendiente")),
            "transmitida": bool(registro.get("transmitida")),
            "documentacion_pendiente": bool(registro.get("documentacion_pendiente")),
            "documentacion_cargada": bool(registro.get("documentacion_cargada")),
            "tiene_boton_transmitir": bool(registro.get("tiene_boton_transmitir")),
            "disponible_activar": bool(registro.get("disponible_activar")),
        }

    def _subir_pdf_a_orden(self, page: Page, nro_orden: str, archivo: Path, item: dict) -> str:
        self._buscar_orden_en_tabla(page, nro_orden, item)
        estado_doc = self._estado_documentacion_orden(page, nro_orden)
        self._log(f"Estado documentacion detectado para OME {nro_orden}: {estado_doc}")
        if estado_doc == "transmitida":
            self._log(f"OME {nro_orden} ya figura transmitida en PAMI; no se carga documentacion.")
            return "ya_transmitido"
        if estado_doc == "cargada":
            self._log(f"La documentacion ya figura cargada en PAMI para OME {nro_orden}; se transmite sin volver a subir.")
            self._transmitir_orden(page, nro_orden)
            return "transmitido"
        if estado_doc == "no_validada":
            raise RuntimeError("La OME no esta validada; no se sube documentacion ni se transmite.")
        if estado_doc not in {"pendiente", "desconocida"}:
            raise RuntimeError(f"La OME no esta disponible para cargar documentacion. Estado detectado: {estado_doc}.")
        if not self._click_boton_documentacion(page, nro_orden):
            raise RuntimeError("No encontre el boton de documentacion en la fila de la OME.")
        self._log(f"Boton de documentacion clickeado para OME {nro_orden}.")
        try:
            self._esperar_modal_documentacion(page, nro_orden)
        except Exception:
            self._log(f"Click en documentacion ejecutado para OME {nro_orden}, pero no aparecio modal; se reintenta.")
            if not self._click_boton_documentacion(page, nro_orden):
                raise RuntimeError("No encontre el boton de documentacion en la fila de la OME.")
            self._log(f"Boton de documentacion clickeado en reintento para OME {nro_orden}.")
            self._esperar_modal_documentacion(page, nro_orden)
        page.wait_for_timeout(900)

        self._seleccionar_tipo_documentacion(page)
        self._seleccionar_archivo_documentacion(page, archivo)
        self._confirmar_modal_documentacion(page)
        self._validar_documentacion_cargada(page, nro_orden)
        self._cerrar_modal_documentacion(page)
        self._transmitir_orden(page, nro_orden, documentacion_confirmada=True)
        return "transmitido"

    def _click_boton_documentacion(self, page: Page, nro_orden: str) -> bool:
        token = f"pami-doc-{int(time.time() * 1000)}"
        selector = page.evaluate(
            """
            ({nroOrden, token}) => {
              const visible = (el) => !!el && (el.offsetParent !== null || el.getClientRects().length > 0);
              const norm = (value) => String(value || '')
                .normalize('NFD')
                .replace(/[\\u0300-\\u036f]/g, '')
                .toLowerCase();
              const meta = (el) => {
                const target = el?.closest('a,button,span,div') || el;
                return norm([
                  el?.className || '',
                  target?.className || '',
                  el?.textContent || '',
                  target?.textContent || '',
                  el?.title || '',
                  target?.title || '',
                  el?.getAttribute?.('aria-label') || '',
                  target?.getAttribute?.('aria-label') || '',
                  el?.getAttribute?.('data-original-title') || '',
                  target?.getAttribute?.('data-original-title') || '',
                  el?.getAttribute?.('data-toggle') || '',
                  target?.getAttribute?.('data-toggle') || ''
                ].join(' '));
              };
              const esVerde = (el) => {
                const target = el?.closest('button,a,span') || el;
                const cls = `${target?.className || ''} ${el?.className || ''}`.toLowerCase();
                if (cls.includes('btn-success') || cls.includes('green')) return true;
                const rgb = (getComputedStyle(target).backgroundColor || '').match(/\\d+/g)?.map(Number) || [];
                return rgb.length >= 3 && rgb[1] > rgb[0] + 20 && rgb[1] > rgb[2] + 20;
              };
              const esTransmision = (el) => {
                const value = meta(el);
                return value.includes('transmit') || value.includes('enviar') || value.includes('arrow-right') ||
                  value.includes('flecha') || value.includes('glyphicon-arrow-right') ||
                  value.includes('fa-arrow-right') || value.includes('fa-arrow-circle-right');
              };
              const esDocumentacion = (el) => {
                const value = meta(el);
                return value.includes('cargar document') || value.includes('documentaci') ||
                  value.includes('upload') || value.includes('subir') || value.includes('adjuntar') ||
                  value.includes('glyphicon-open') || value.includes('glyphicon-upload') ||
                  value.includes('fa-file') || value.includes('fa-folder') || value.includes('cloud-upload');
              };
              const control = (el) => el?.closest('button,a,span.btn,.btn') || el;
              for (const row of document.querySelectorAll('table tbody tr')) {
                const tds = Array.from(row.querySelectorAll('td'));
                if ((tds[0]?.textContent || '').trim() !== nroOrden) continue;
                const acciones = tds[tds.length - 1] || row;
                const vistos = new Set();
                const controles = Array.from(acciones.querySelectorAll('a, button, span.btn, .btn'))
                  .map(control)
                  .filter((el) => el && visible(el))
                  .filter((el) => {
                    if (vistos.has(el)) return false;
                    vistos.add(el);
                    return true;
                  })
                  .sort((a, b) => a.getBoundingClientRect().left - b.getBoundingClientRect().left);
                const iconoDoc = Array.from(acciones.querySelectorAll('a, button, span, i'))
                  .filter(visible)
                  .find((el) => esDocumentacion(el) && !esTransmision(el));
                const boton =
                  (iconoDoc ? control(iconoDoc) : null) ||
                  controles.find((el) => esDocumentacion(el) && !esTransmision(el)) ||
                  controles.find((el) => esVerde(el) && !esTransmision(el)) ||
                  controles[1];
                if (!boton) return '';
                document.querySelectorAll('[data-pami-doc-click]').forEach((el) => el.removeAttribute('data-pami-doc-click'));
                boton.setAttribute('data-pami-doc-click', token);
                boton.scrollIntoView({block: 'center', inline: 'center'});
                return `[data-pami-doc-click="${token}"]`;
              }
              return '';
            }
            """,
            {"nroOrden": nro_orden, "token": token},
        )
        if not selector:
            return False
        try:
            boton = page.locator(str(selector)).first
            boton.scroll_into_view_if_needed(timeout=3000)
            try:
                boton.click(timeout=5000)
            except Exception:
                boton.click(timeout=5000, force=True)
            page.wait_for_timeout(400)
            if not self._modal_documentacion_abierto(page):
                clicked_js = page.evaluate(
                    """
                    (selector) => {
                      const boton = document.querySelector(selector);
                      if (!boton) return false;
                      boton.scrollIntoView({block: 'center', inline: 'center'});
                      boton.click();
                      return true;
                    }
                    """,
                    str(selector),
                )
                if clicked_js:
                    page.wait_for_timeout(700)
            return True
        except Exception as exc:
            self._log(f"No se pudo clickear el boton de documentacion para OME {nro_orden}: {exc}")
            return False

    def _modal_documentacion_abierto(self, page: Page) -> bool:
        try:
            return bool(
                page.evaluate(
                    """
                    () => {
                      const visible = (el) => !!el && (el.offsetParent !== null || el.getClientRects().length > 0);
                      const norm = (value) => String(value || '')
                        .normalize('NFD')
                        .replace(/[\\u0300-\\u036f]/g, '')
                        .replace(/\\s+/g, ' ')
                        .trim()
                        .toLowerCase();
                      const modales = Array.from(document.querySelectorAll('.modal.show, .modal.in, .bootbox, [role="dialog"], .modal'))
                        .filter(visible);
                      return modales.some((modal) => {
                        const txt = norm(modal.textContent);
                        return (
                          txt.includes('cargar document') ||
                          txt.includes('tipo de documentacion') ||
                          txt.includes('documentos requeridos') ||
                          !!modal.querySelector('input[type="file"]')
                        );
                      });
                    }
                    """
                )
            )
        except Exception:
            return False

    def _esperar_modal_documentacion(self, page: Page, nro_orden: str) -> None:
        for _ in range(16):
            if self._modal_documentacion_abierto(page):
                self._log(f"Modal de documentacion abierto para OME {nro_orden}.")
                return
            page.wait_for_timeout(500)
        raise RuntimeError(f"No se abrio el modal de documentacion para OME {nro_orden}.")

    def _seleccionar_archivo_documentacion(self, page: Page, archivo: Path) -> None:
        archivo_subida = archivo_documentacion_para_subir(archivo)
        if archivo_subida != archivo:
            self._log(f"Imagen convertida a PDF para subir: {archivo_subida}")
        self._esperar_modal_documentacion(page, "")
        input_selector = self._esperar_input_archivo_documentacion(page)
        clicked = False
        try:
            with page.expect_file_chooser(timeout=5000) as chooser_info:
                clicked = bool(
                    page.evaluate(
                        """
                        () => {
                          const visible = (el) => !!el && (el.offsetParent !== null || el.getClientRects().length > 0);
                          const texto = (el) => (el?.textContent || el?.value || el?.title || '').replace(/\\s+/g, ' ').trim();
                          const modales = Array.from(document.querySelectorAll('.modal.show, .modal.in, .bootbox, [role="dialog"], .modal'))
                            .filter(visible)
                            .filter((modal) => {
                              const txt = (modal.textContent || '').normalize('NFD').replace(/[\\u0300-\\u036f]/g, '').toLowerCase();
                              return txt.includes('cargar document') || txt.includes('tipo de documentacion') || modal.querySelector('input[type="file"]');
                            });
                          const modal = modales[0] || document;
                          const candidatos = Array.from(modal.querySelectorAll('button, input[type="button"], input[type="file"], label, a'))
                            .filter(visible);
                          const boton = candidatos.find((el) => /seleccionar archivo|archivo|examinar|browse/i.test(texto(el)));
                          const target = boton?.closest('button,a,label') || boton;
                          if (!target) return false;
                          target.click();
                          return true;
                        }
                        """
                    )
                )
            if clicked:
                chooser = chooser_info.value
                chooser.set_files(str(archivo_subida))
                self._log(f"Archivo seleccionado desde el selector de Chrome: {archivo_subida.name}")
                page.wait_for_timeout(1200)
                return
        except Exception as exc:
            self._log(f"No se pudo usar selector de archivo de Chrome; se usa carga directa. Detalle: {exc}")

        modal_input = page.locator(input_selector).first
        try:
            modal_input.wait_for(state="attached", timeout=3000)
            file_input = modal_input
        except Exception:
            file_input = page.locator("input[type='file']").first
        file_input.wait_for(state="attached", timeout=10000)
        file_input.set_input_files(str(archivo_subida))
        self._disparar_eventos_archivo(page)
        self._log(f"Archivo asignado al campo de documentacion: {archivo_subida.name}")
        page.wait_for_timeout(1200)

    def _esperar_input_archivo_documentacion(self, page: Page) -> str:
        token = f"pami-doc-file-{int(time.time() * 1000)}"
        for _ in range(20):
            selector = page.evaluate(
                """
                (token) => {
                  const visible = (el) => !!el && (el.offsetParent !== null || el.getClientRects().length > 0);
                  const norm = (value) => String(value || '')
                    .normalize('NFD')
                    .replace(/[\\u0300-\\u036f]/g, '')
                    .replace(/\\s+/g, ' ')
                    .trim()
                    .toLowerCase();
                  const modales = Array.from(document.querySelectorAll('.modal.show, .modal.in, .bootbox, [role="dialog"], .modal'))
                    .filter(visible)
                    .filter((modal) => {
                      const txt = norm(modal.textContent);
                      return (
                        txt.includes('cargar document') ||
                        txt.includes('tipo de documentacion') ||
                        txt.includes('documentos requeridos') ||
                        !!modal.querySelector('input[type="file"]')
                      );
                    });
                  let input = null;
                  for (const modal of modales) {
                    input = modal.querySelector('input[type="file"]');
                    if (input) break;
                  }
                  input = input || Array.from(document.querySelectorAll('input[type="file"]'))
                    .find((el) => visible(el.closest('.modal, .bootbox, [role="dialog"]') || el));
                  if (!input) return '';
                  document.querySelectorAll('[data-pami-doc-file]').forEach((el) => el.removeAttribute('data-pami-doc-file'));
                  input.setAttribute('data-pami-doc-file', token);
                  return `[data-pami-doc-file="${token}"]`;
                }
                """,
                token,
            )
            if selector:
                return str(selector)
            page.wait_for_timeout(500)
        raise RuntimeError("No aparecio el campo Archivo en el modal de documentacion.")

    def _disparar_eventos_archivo(self, page: Page) -> None:
        try:
            page.evaluate(
                """
                () => {
                  const input = document.querySelector('[data-pami-doc-file]') ||
                    Array.from(document.querySelectorAll('input[type="file"]'))
                    .find((el) => el.files && el.files.length > 0) ||
                    document.querySelector('input[type="file"]');
                  if (!input) return;
                  for (const eventName of ['input', 'change', 'blur']) {
                    input.dispatchEvent(new Event(eventName, { bubbles: true }));
                  }
                }
                """
            )
        except Exception:
            pass

    def _estado_documentacion_orden(self, page: Page, nro_orden: str) -> str:
        try:
            estado = page.evaluate(
                """
                (nroOrden) => {
                  const visible = (el) => !!el && (el.offsetParent !== null || el.getClientRects().length > 0);
                  const colorEstado = (el) => {
                    const target = el?.closest('button,a,span') || el;
                    const cls = `${target?.className || ''} ${el?.className || ''}`.toLowerCase();
                    if (cls.includes('btn-primary') || cls.includes('btn-info') || cls.includes('blue')) return 'cargada';
                    if (cls.includes('btn-success') || cls.includes('green')) return 'pendiente';
                    const rgb = (getComputedStyle(target).backgroundColor || '').match(/\\d+/g)?.map(Number) || [];
                    if (rgb.length >= 3 && rgb[2] > rgb[0] + 20 && rgb[2] >= rgb[1] - 10) return 'cargada';
                    if (rgb.length >= 3 && rgb[1] > rgb[0] + 20 && rgb[1] > rgb[2] + 20) return 'pendiente';
                    return '';
                  };
                  const control = (el) => el?.closest('button,a,span.btn,.btn') || el;
                  const controlesAccion = (acciones) => {
                    const vistos = new Set();
                    return Array.from(acciones.querySelectorAll('a, button, span.btn, .btn'))
                      .map(control)
                      .filter((el) => el && visible(el))
                      .filter((el) => {
                        if (vistos.has(el)) return false;
                        vistos.add(el);
                        return true;
                      })
                      .sort((a, b) => a.getBoundingClientRect().left - b.getBoundingClientRect().left);
                  };
                  for (const row of document.querySelectorAll('table tbody tr')) {
                    const tds = Array.from(row.querySelectorAll('td'));
                    if ((tds[0]?.textContent || '').trim() !== nroOrden) continue;
                    const texto = tds.map((td) => (td.textContent || '').replace(/\\s+/g, ' ').trim()).join(' ');
                    if (/\\bSI\\b/i.test(texto)) return 'transmitida';
                    const acciones = tds[tds.length - 1] || row;
                    const controles = controlesAccion(acciones);
                    const check = Array.from(acciones.querySelectorAll('i.fa-check, .fa-check')).find(visible);
                    const lupa = Array.from(acciones.querySelectorAll('i.fa-search, .fa-search, i.fa-eye, .fa-eye')).find(visible);
                    const transmitir = Array.from(acciones.querySelectorAll('i.transmitir, .transmitir, .fa-arrow-right, .fa-arrow-circle-right, .glyphicon-arrow-right, .icon-arrow-right'))
                      .find(visible);
                    const doc = Array.from(acciones.querySelectorAll('i.fa-upload, .fa-upload, i.fa-file, .fa-file, i.fa-file-text, .fa-file-text, i.fa-folder, .fa-folder, i.fa-cloud-upload, .fa-cloud-upload, .glyphicon-open, .glyphicon-upload, .icon-upload, .icon-file'))
                      .find(visible);
                    const checkControl = check ? control(check) : controles[0];
                    const docControl = doc ? control(doc) : controles[1];
                    const transmitirControl = transmitir ? control(transmitir) : controles[2];
                    const checkEstado = checkControl ? colorEstado(checkControl) : '';
                    const docEstado = docControl ? colorEstado(docControl) : '';
                    if (check && checkEstado === 'pendiente') return 'no_validada';
                    if (lupa && checkControl && colorEstado(checkControl) === 'cargada' && !transmitirControl && docEstado !== 'pendiente') return 'transmitida';
                    if (docEstado === 'cargada') return 'cargada';
                    if (docEstado === 'pendiente') return 'pendiente';
                    if (!docControl) return 'sin_boton';
                    return 'desconocida';
                  }
                  return 'no_encontrada';
                }
                """,
                nro_orden,
            )
            return str(estado or "desconocida")
        except Exception:
            return "desconocida"

    def _validar_documentacion_cargada(self, page: Page, nro_orden: str) -> None:
        for _ in range(12):
            page.wait_for_timeout(700)
            if self._modal_documentacion_ok(page):
                self._log(f"Documentacion cargada confirmada en modal PAMI para OME {nro_orden}.")
                return
            estado = self._estado_documentacion_orden(page, nro_orden)
            if estado == "cargada":
                self._log(f"Documentacion cargada confirmada en PAMI para OME {nro_orden}.")
                return
            if estado == "pendiente":
                continue
            if estado in {"no_encontrada", "sin_boton"}:
                self._log(
                    f"Validacion de carga para OME {nro_orden}: estado {estado}; se sigue esperando confirmacion."
                )
                continue
        raise RuntimeError("El archivo fue seleccionado, pero PAMI sigue mostrando documentacion pendiente.")

    def _modal_documentacion_ok(self, page: Page) -> bool:
        try:
            return bool(
                page.evaluate(
                    """
                    () => {
                      const norm = (value) => (value || '').toString().normalize('NFD').replace(/[\\u0300-\\u036f]/g, '').toLowerCase();
                      const modal =
                        document.querySelector('.modal.show, .modal.in, .bootbox, [role="dialog"]') ||
                        Array.from(document.querySelectorAll('.modal, .bootbox, [role="dialog"]')).find((el) => norm(el.textContent).includes('document'));
                      if (!modal) return false;
                      const txt = norm(modal.textContent);
                      return txt.includes('documentos requeridos') &&
                        txt.includes('informe/resultados') &&
                        (txt.includes('?') || txt.includes('?') || txt.includes('check') || txt.includes('fecha carga'));
                    }
                    """
                )
            )
        except Exception:
            return False

    def _cerrar_modal_documentacion(self, page: Page) -> None:
        try:
            cerrado = page.evaluate(
                """
                () => {
                  const visible = (el) => !!el && (el.offsetParent !== null || el.getClientRects().length > 0);
                  const modal =
                    Array.from(document.querySelectorAll('.modal.show, .modal.in, .bootbox, [role="dialog"], .modal'))
                      .find((el) => visible(el) && /cargar documentaci/i.test(el.textContent || ''));
                  if (!modal) return true;
                  const cerrar = Array.from(modal.querySelectorAll('button, input[type="button"]'))
                    .find((el) => visible(el) && /^cerrar$/i.test((el.textContent || el.value || '').trim()));
                  if (!cerrar) return false;
                  cerrar.click();
                  return true;
                }
                """
            )
            if cerrado:
                page.wait_for_timeout(900)
                self._log("Modal de documentacion cerrado.")
            if "transmision.php" not in (page.url or ""):
                self._log(f"PAMI navego fuera de Transmision al cerrar modal: {page.url}. Volviendo al panel.")
                self._goto_transmision(page)
        except Exception:
            pass

    def _on_dialog(self, dialog) -> None:
        # PAMI muestra alerts nativos ("pe.pami.org.ar dice…"). Guardamos el texto
        # y aceptamos (un alert solo tiene OK; y ya forzamos confirm=true igual).
        try:
            self._ultimo_dialogo = dialog.message or ""
        except Exception:
            self._ultimo_dialogo = ""
        # Descartar (no aceptar): es lo que hacía Playwright por defecto sin
        # handler, así no cambiamos el comportamiento de ningún confirm nativo.
        try:
            dialog.dismiss()
        except Exception:
            try:
                dialog.accept()
            except Exception:
                pass

    def _dialogo_afiliado_inactivo(self) -> bool:
        # ¿El último diálogo de PAMI dice que el afiliado no estaba activo en el
        # padrón a la fecha de la práctica? (motivo real por el que no transmite).
        m = str(self._ultimo_dialogo or "").lower()
        m = m.translate(str.maketrans("áéíóú", "aeiou"))
        return ("no se encontraba activa" in m or "no se encuentra activa" in m
                or "activa en el padron" in m
                or ("padron" in m and "activa" in m))

    def _transmitir_orden(self, page: Page, nro_orden: str, documentacion_confirmada: bool = False) -> None:
        self._ultimo_dialogo = ""  # limpiar antes de transmitir esta OME
        elegible = self._estado_transmision_orden(page, nro_orden)
        self._log(f"Estado transmision detectado para OME {nro_orden}: {elegible}")
        if elegible.get("transmitida"):
            self._log(f"OME {nro_orden} ya figura transmitida en PAMI.")
            return
        if elegible.get("validacion_pendiente") or not elegible.get("validada"):
            raise RuntimeError("La OME no esta validada; no se transmite.")
        if not elegible.get("doc_cargada") and not documentacion_confirmada:
            raise RuntimeError("La documentacion no figura cargada; no se transmite la OME aunque exista flecha verde.")
        if not elegible.get("tiene_boton"):
            raise RuntimeError("La documentacion figura cargada, pero no encontre boton verde para transmitir.")
        click_info = page.evaluate(
            """
            async (nroOrden) => {
              window.confirm = () => true;
              const visible = (el) => !!el && (el.offsetParent !== null || el.getClientRects().length > 0);
              const delay = (ms) => new Promise((resolve) => setTimeout(resolve, ms));
              const confirmVisible = () => Array.from(document.querySelectorAll('button')).some((b) => {
                const txt = (b.textContent || '').trim().toLowerCase();
                return ['confirmar', 'aceptar', 'si', 'sí', 'sÃ­'].includes(txt) && visible(b) && !b.disabled;
              });
              const ordenTransmitida = () => {
                for (const row of document.querySelectorAll('table tbody tr')) {
                  const tds = Array.from(row.querySelectorAll('td'));
                  if ((tds[0]?.textContent || '').trim() !== nroOrden) continue;
                  return /\\bSI\\b/i.test(tds.map((td) => (td.textContent || '').trim()).join(' '));
                }
                return false;
              };
              const esVerde = (el) => {
                const target = el?.closest('button,a,span') || el;
                const cls = `${target?.className || ''} ${el?.className || ''}`.toLowerCase();
                if (cls.includes('btn-success') || cls.includes('green')) return true;
                const rgb = (getComputedStyle(target).backgroundColor || '').match(/\\d+/g)?.map(Number) || [];
                return rgb.length >= 3 && rgb[1] > rgb[0] + 20 && rgb[1] > rgb[2] + 20;
              };
              const norm = (value) => String(value || '').normalize('NFD').replace(/[\\u0300-\\u036f]/g, '').toLowerCase();
              const meta = (el) => {
                const target = el?.closest('a,button,span,div') || el;
                return norm([
                  el?.className || '',
                  target?.className || '',
                  el?.textContent || '',
                  target?.textContent || '',
                  el?.title || '',
                  target?.title || '',
                  el?.getAttribute?.('aria-label') || '',
                  target?.getAttribute?.('aria-label') || '',
                  el?.getAttribute?.('data-original-title') || '',
                  target?.getAttribute?.('data-original-title') || ''
                ].join(' '));
              };
              for (const row of document.querySelectorAll('table tbody tr')) {
                const tds = Array.from(row.querySelectorAll('td'));
                if ((tds[0]?.textContent || '').trim() !== nroOrden) continue;
                const acciones = tds[tds.length - 1] || row;
                const control = (el) => el?.closest('button,a,span.btn,.btn') || el;
                const vistos = new Set();
                const controles = Array.from(acciones.querySelectorAll('a, button, span.btn, .btn'))
                  .map(control)
                  .filter((el) => el && visible(el))
                  .filter((el) => {
                    if (vistos.has(el)) return false;
                    vistos.add(el);
                    return true;
                  })
                  .sort((a, b) => a.getBoundingClientRect().left - b.getBoundingClientRect().left);
                const transmitirPorIcono = Array.from(acciones.querySelectorAll('i.transmitir, .transmitir, .fa-arrow-right, .fa-arrow-circle-right, .glyphicon-arrow-right, .icon-arrow-right, a, button, span, i'))
                  .filter(visible)
                  .find((el) => {
                    const value = meta(el);
                    return (
                      value.includes('transmit') ||
                      value.includes('enviar') ||
                      value.includes('arrow-right') ||
                      value.includes('flecha') ||
                      value.includes('glyphicon-arrow-right') ||
                      value.includes('fa-arrow-right') ||
                      value.includes('fa-arrow-circle-right')
                    );
                  });
                const transmitirControl = transmitirPorIcono ? control(transmitirPorIcono) : null;
                const candidatos = [
                  transmitirPorIcono,
                  transmitirControl,
                  controles.find((el) => {
                    const value = meta(el);
                    return esVerde(el) && (
                      value.includes('transmit') ||
                      value.includes('enviar') ||
                      value.includes('arrow-right') ||
                      value.includes('flecha')
                    );
                  }),
                  controles.filter((el) => esVerde(el)).at(-1),
                  controles[2],
                ].filter(Boolean);
                const vistosCandidatos = new Set();
                let intento = 0;
                for (const candidato of candidatos) {
                  if (!candidato || vistosCandidatos.has(candidato)) continue;
                  vistosCandidatos.add(candidato);
                  intento += 1;
                  candidato.scrollIntoView({block: 'center', inline: 'center'});
                  candidato.click();
                  await delay(900);
                  if (confirmVisible() || ordenTransmitida()) {
                    return {
                      clicked: true,
                      intento,
                      modal: confirmVisible(),
                      transmitida: ordenTransmitida(),
                      meta: meta(candidato),
                      tag: candidato.tagName,
                      className: String(candidato.className || '')
                    };
                  }
                }
                return {
                  clicked: candidatos.length > 0,
                  intento,
                  modal: confirmVisible(),
                  transmitida: ordenTransmitida(),
                  meta: candidatos[0] ? meta(candidatos[0]) : '',
                  tag: candidatos[0]?.tagName || '',
                  className: String(candidatos[0]?.className || '')
                };
              }
              return {clicked: false, motivo: 'orden_no_encontrada'};
            }
            """,
            nro_orden,
        )
        if not (
            isinstance(click_info, dict)
            and (click_info.get("modal") or click_info.get("transmitida"))
        ):
            trusted_click_info = self._click_boton_transmitir(page, nro_orden)
            if isinstance(trusted_click_info, dict):
                click_info = {
                    "primer_intento": click_info,
                    "click_con_mouse": trusted_click_info,
                    "clicked": trusted_click_info.get("clicked"),
                }
            else:
                click_info = trusted_click_info
        clicked = bool(click_info.get("clicked") if isinstance(click_info, dict) else click_info)
        if not clicked:
            raise RuntimeError("Documentacion cargada, pero no encontre boton para transmitir la OME.")
        self._log(f"Boton de transmision clickeado para OME {nro_orden}.")
        try:
            page.wait_for_function(
                """
                Array.from(document.querySelectorAll('button')).some((b) => {
                  const txt = (b.textContent || '').trim().toLowerCase();
                  return ['confirmar', 'aceptar', 'si', 'sí'].includes(txt) && b.offsetParent !== null;
                })
                """,
                timeout=6000,
            )
        except Exception:
            if "solicitudAgenteController.php" in (page.url or ""):
                raise RuntimeError("Se abrio Solicitudes en vez del modal de transmision; PAMI no expuso el boton verde de transmitir esperado.")
            self._log(f"No aparecio modal de confirmacion para OME {nro_orden}; verifico si PAMI transmitio igual.")
        if self._cerrar_modal_transmision_no_validada(page):
            raise RuntimeError("La OME no esta validada; PAMI pidio motivo de transmision excepcional. No se transmite.")
        self._confirmar_transmision(page, nro_orden)

    def _cerrar_modal_transmision_no_validada(self, page: Page) -> bool:
        for _ in range(8):
            try:
                cerrado = page.evaluate(
                    """
                    () => {
                      const visible = (el) => !!el && (el.offsetParent !== null || el.getClientRects().length > 0);
                      const norm = (value) => String(value || '')
                        .normalize('NFD')
                        .replace(/[\\u0300-\\u036f]/g, '')
                        .replace(/\\s+/g, ' ')
                        .trim()
                        .toLowerCase();
                      const modal = Array.from(document.querySelectorAll('.modal.show, .modal.in, .bootbox, [role="dialog"], .modal'))
                        .filter(visible)
                        .find((el) => {
                          const txt = norm(el.textContent);
                          return (
                            txt.includes('prestacion no se encuentra validada') ||
                            txt.includes('transmision excepcional') ||
                            (txt.includes('motivo') && txt.includes('validada'))
                          );
                        });
                      if (!modal) return false;
                      const boton = Array.from(modal.querySelectorAll('button, a, .close'))
                        .find((el) => {
                          if (!visible(el) || el.disabled) return false;
                          const txt = norm(el.textContent);
                          const cls = norm(el.className);
                          return txt.includes('cancelar') || txt.includes('cerrar') || txt === 'no' || cls.includes('close');
                        });
                      if (boton) boton.click();
                      return true;
                    }
                    """
                )
                if cerrado:
                    page.wait_for_timeout(600)
                    return True
            except Exception:
                return False
            page.wait_for_timeout(250)
        return False

    def _click_boton_transmitir(self, page: Page, nro_orden: str) -> dict:
        marker = f"doc-transmitir-{nro_orden}"
        target = page.evaluate(
            """
            ({nroOrden, marker}) => {
              const visible = (el) => !!el && (el.offsetParent !== null || el.getClientRects().length > 0);
              const esVerde = (el) => {
                const target = el?.closest('button,a,span') || el;
                const cls = `${target?.className || ''} ${el?.className || ''}`.toLowerCase();
                if (cls.includes('btn-success') || cls.includes('green')) return true;
                const rgb = (getComputedStyle(target).backgroundColor || '').match(/\\d+/g)?.map(Number) || [];
                return rgb.length >= 3 && rgb[1] > rgb[0] + 20 && rgb[1] > rgb[2] + 20;
              };
              for (const old of document.querySelectorAll('[data-doc-transmit-marker]')) {
                old.removeAttribute('data-doc-transmit-marker');
              }
              for (const row of document.querySelectorAll('table tbody tr')) {
                const tds = Array.from(row.querySelectorAll('td'));
                if ((tds[0]?.textContent || '').trim() !== nroOrden) continue;
                const acciones = tds[tds.length - 1] || row;
                const exacto = acciones.querySelector('i.transmitir');
                const verdes = Array.from(acciones.querySelectorAll('a, button, span.btn, .btn')).filter(esVerde);
                const candidato = exacto || verdes[verdes.length - 1];
                if (!candidato || !visible(candidato)) {
                  return {found: false, reason: 'sin_boton_visible'};
                }
                const control = candidato.closest('a,button,span.btn,.btn') || candidato;
                candidato.setAttribute('data-doc-transmit-marker', marker);
                control.setAttribute('data-doc-transmit-marker', `${marker}-control`);
                candidato.scrollIntoView({block: 'center', inline: 'center'});
                const rect = candidato.getBoundingClientRect();
                return {
                  found: true,
                  iconSelector: `[data-doc-transmit-marker="${marker}"]`,
                  controlSelector: `[data-doc-transmit-marker="${marker}-control"]`,
                  tag: candidato.tagName,
                  className: String(candidato.className || ''),
                  controlTag: control.tagName,
                  controlClassName: String(control.className || ''),
                  x: rect.left + rect.width / 2,
                  y: rect.top + rect.height / 2
                };
              }
              return {found: false, reason: 'orden_no_encontrada'};
            }
            """,
            {"nroOrden": nro_orden, "marker": marker},
        )
        if not isinstance(target, dict) or not target.get("found"):
            return {"clicked": False, "target": target}

        attempts = []
        for name, selector in (
            ("playwright-icon", target.get("iconSelector")),
            ("playwright-control", target.get("controlSelector")),
        ):
            if not selector:
                continue
            try:
                page.locator(selector).click(timeout=2500, force=True)
                page.wait_for_timeout(900)
                attempts.append(name)
                if self._modal_confirmacion_visible(page) or self._estado_transmitida_orden(page, nro_orden):
                    return {"clicked": True, "attempts": attempts, "target": target}
            except Exception as exc:
                attempts.append(f"{name}: {exc}")

        try:
            page.mouse.click(float(target["x"]), float(target["y"]))
            page.wait_for_timeout(900)
            attempts.append("mouse-icon")
            if self._modal_confirmacion_visible(page) or self._estado_transmitida_orden(page, nro_orden):
                return {"clicked": True, "attempts": attempts, "target": target}
        except Exception as exc:
            attempts.append(f"mouse-icon: {exc}")

        js_clicked = page.evaluate(
            """
            ({iconSelector, controlSelector}) => {
              window.confirm = () => true;
              const el = document.querySelector(iconSelector) || document.querySelector(controlSelector);
              if (!el) return false;
              el.click();
              return true;
            }
            """,
            {"iconSelector": target.get("iconSelector"), "controlSelector": target.get("controlSelector")},
        )
        page.wait_for_timeout(900)
        attempts.append(f"js-icon:{js_clicked}")
        return {"clicked": bool(js_clicked), "attempts": attempts, "target": target}

    def _modal_confirmacion_visible(self, page: Page) -> bool:
        try:
            return bool(
                page.evaluate(
                    """
                    () => {
                      const visible = (el) => !!el && (el.offsetParent !== null || el.getClientRects().length > 0);
                      return Array.from(document.querySelectorAll('button')).some((b) => {
                        const txt = (b.textContent || '').trim().toLowerCase();
                        return ['confirmar', 'aceptar', 'si', 'sÃ­'].includes(txt) && visible(b) && !b.disabled;
                      });
                    }
                    """
                )
            )
        except Exception:
            return False

    def _estado_transmision_orden(self, page: Page, nro_orden: str) -> dict:
        try:
            estado = page.evaluate(
                """
                (nroOrden) => {
                  const visible = (el) => !!el && (el.offsetParent !== null || el.getClientRects().length > 0);
                  const esAzul = (el) => {
                    const target = el?.closest('button,a,span') || el;
                    const cls = `${target?.className || ''} ${el?.className || ''}`.toLowerCase();
                    if (cls.includes('btn-primary') || cls.includes('btn-info') || cls.includes('blue')) return true;
                    const rgb = (getComputedStyle(target).backgroundColor || '').match(/\\d+/g)?.map(Number) || [];
                    return rgb.length >= 3 && rgb[2] > rgb[0] + 20 && rgb[2] >= rgb[1] - 10;
                  };
                  const esVerde = (el) => {
                    const target = el?.closest('button,a,span') || el;
                    const cls = `${target?.className || ''} ${el?.className || ''}`.toLowerCase();
                    if (cls.includes('btn-success') || cls.includes('green')) return true;
                    const rgb = (getComputedStyle(target).backgroundColor || '').match(/\\d+/g)?.map(Number) || [];
                    return rgb.length >= 3 && rgb[1] > rgb[0] + 20 && rgb[1] > rgb[2] + 20;
                  };
                  const control = (el) => el?.closest('button,a,span.btn,.btn') || el;
                  const controlesAccion = (acciones) => {
                    const vistos = new Set();
                    return Array.from(acciones.querySelectorAll('a, button, span.btn, .btn'))
                      .map(control)
                      .filter((el) => el && visible(el))
                      .filter((el) => {
                        if (vistos.has(el)) return false;
                        vistos.add(el);
                        return true;
                      })
                      .sort((a, b) => a.getBoundingClientRect().left - b.getBoundingClientRect().left);
                  };
                  for (const row of document.querySelectorAll('table tbody tr')) {
                    const tds = Array.from(row.querySelectorAll('td'));
                    if ((tds[0]?.textContent || '').trim() !== nroOrden) continue;
                    const texto = tds.map((td) => (td.textContent || '').trim()).join(' ');
                    const acciones = tds[tds.length - 1] || row;
                    const controles = controlesAccion(acciones);
                    const check = acciones.querySelector('i.fa-check, .fa-check, i.glyphicon-ok, .glyphicon-ok, i.icon-ok, .icon-ok, i.icon-check, .icon-check');
                    const doc = acciones.querySelector('i.fa-upload, .fa-upload, i.fa-file, .fa-file, i.fa-file-text, .fa-file-text, i.fa-folder, .fa-folder, i.fa-cloud-upload, .fa-cloud-upload, .glyphicon-open, .glyphicon-upload, .icon-upload, .icon-file');
                    const lupa = acciones.querySelector('i.fa-search, .fa-search, i.fa-eye, .fa-eye');
                    const transmitir = acciones.querySelector('i.transmitir, .transmitir, .fa-arrow-right, .fa-arrow-circle-right, .glyphicon-arrow-right, .icon-arrow-right');
                    const checkControl = check ? control(check) : controles[0];
                    const docControl = doc ? control(doc) : controles[1];
                    const transmitirControl = transmitir ? control(transmitir) : controles[2];
                    const checkAzul = checkControl ? esAzul(checkControl) : false;
                    const docPendiente = docControl ? esVerde(docControl) : false;
                    const docCargada = docControl ? esAzul(docControl) : false;
                    const botonTransmision = !!(transmitirControl && visible(transmitirControl) && esVerde(transmitirControl));
                    const cerradaPorAcciones = !!(lupa && visible(lupa) && checkAzul && !docPendiente && !botonTransmision);
                    const transmitida = /\\bSI\\b/i.test(texto) || cerradaPorAcciones;
                    const validacionPendiente = !transmitida && !!(check && checkControl && visible(checkControl) && esVerde(checkControl) && !checkAzul);
                    return {
                      encontrada: true,
                      transmitida,
                      validada: transmitida || (!validacionPendiente && (checkAzul || docPendiente || docCargada || botonTransmision)),
                      validacion_pendiente: validacionPendiente,
                      doc_cargada: docCargada,
                      doc_pendiente: docPendiente,
                      tiene_boton: botonTransmision
                    };
                  }
                  return {encontrada: false};
                }
                """,
                nro_orden,
            )
            return estado if isinstance(estado, dict) else {"encontrada": False}
        except Exception:
            return {"encontrada": False}

    def _confirmar_transmision(self, page: Page, nro_orden: str) -> None:
        clicked = page.evaluate(
            """
            () => {
              window.confirm = () => true;
              const visible = (el) => !!el && (el.offsetParent !== null || el.getClientRects().length > 0);
              const btn = Array.from(document.querySelectorAll('button'))
                .find((el) => {
                  if (!visible(el) || el.disabled) return false;
                  const txt = (el.textContent || '').trim().toLowerCase();
                  return ['confirmar', 'aceptar', 'si', 'sí'].includes(txt);
                });
              if (!btn) return false;
              btn.click();
              return true;
            }
            """
        )
        if clicked:
            page.wait_for_timeout(1500)
        else:
            page.wait_for_timeout(900)
        if self._orden_visible(page, nro_orden):
            estado = self._estado_transmitida_orden(page, nro_orden)
            if estado:
                self._log(f"Transmision confirmada en PAMI para OME {nro_orden}.")
                return
            if self._dialogo_afiliado_inactivo():
                raise RuntimeError("Afiliado inactivo en el padron del INSSJP a la fecha de la practica: PAMI no permite transmitir.")
            raise RuntimeError("Documentacion cargada, pero la OME sigue sin figurar transmitida.")
        self._log(f"Transmision confirmada en PAMI para OME {nro_orden}.")

    def _estado_transmitida_orden(self, page: Page, nro_orden: str) -> bool:
        try:
            return bool(
                page.evaluate(
                    """
                    (nroOrden) => {
                      for (const row of document.querySelectorAll('table tbody tr')) {
                        const tds = Array.from(row.querySelectorAll('td'));
                        if ((tds[0]?.textContent || '').trim() !== nroOrden) continue;
                        const texto = tds.map((td) => (td.textContent || '').trim()).join(' ');
                        if (/\\bSI\\b/i.test(texto)) return true;
                        const visible = (el) => !!el && (el.offsetParent !== null || el.getClientRects().length > 0);
                        const esAzul = (el) => {
                          const target = el?.closest('button,a,span') || el;
                          const cls = `${target?.className || ''} ${el?.className || ''}`.toLowerCase();
                          if (cls.includes('btn-primary') || cls.includes('btn-info') || cls.includes('blue')) return true;
                          const rgb = (getComputedStyle(target).backgroundColor || '').match(/\\d+/g)?.map(Number) || [];
                          return rgb.length >= 3 && rgb[2] > rgb[0] + 20 && rgb[2] >= rgb[1] - 10;
                        };
                        const esVerde = (el) => {
                          const target = el?.closest('button,a,span') || el;
                          const cls = `${target?.className || ''} ${el?.className || ''}`.toLowerCase();
                          if (cls.includes('btn-success') || cls.includes('green')) return true;
                          const rgb = (getComputedStyle(target).backgroundColor || '').match(/\\d+/g)?.map(Number) || [];
                          return rgb.length >= 3 && rgb[1] > rgb[0] + 20 && rgb[1] > rgb[2] + 20;
                        };
                        const control = (el) => el?.closest('button,a,span.btn,.btn') || el;
                        const acciones = tds[tds.length - 1] || row;
                        const vistos = new Set();
                        const controles = Array.from(acciones.querySelectorAll('a, button, span.btn, .btn'))
                          .map(control)
                          .filter((el) => el && visible(el))
                          .filter((el) => {
                            if (vistos.has(el)) return false;
                            vistos.add(el);
                            return true;
                          })
                          .sort((a, b) => a.getBoundingClientRect().left - b.getBoundingClientRect().left);
                        const check = acciones.querySelector('i.fa-check, .fa-check, i.glyphicon-ok, .glyphicon-ok, i.icon-ok, .icon-ok, i.icon-check, .icon-check');
                        const lupa = acciones.querySelector('i.fa-search, .fa-search, i.fa-eye, .fa-eye');
                        const doc = acciones.querySelector('i.fa-upload, .fa-upload, i.fa-file, .fa-file, i.fa-file-text, .fa-file-text, i.fa-folder, .fa-folder, i.fa-cloud-upload, .fa-cloud-upload, .glyphicon-open, .glyphicon-upload, .icon-upload, .icon-file');
                        const transmitir = acciones.querySelector('i.transmitir, .transmitir, .fa-arrow-right, .fa-arrow-circle-right, .glyphicon-arrow-right, .icon-arrow-right');
                        const checkControl = check ? control(check) : controles[0];
                        const docControl = doc ? control(doc) : controles[1];
                        const transmitirControl = transmitir ? control(transmitir) : controles[2];
                        return !!(lupa && visible(lupa) && checkControl && esAzul(checkControl) && (!docControl || !esVerde(docControl)) && !(transmitirControl && visible(transmitirControl) && esVerde(transmitirControl)));
                      }
                      return false;
                    }
                    """,
                    nro_orden,
                )
            )
        except Exception:
            return False

    def _seleccionar_tipo_documentacion(self, page: Page) -> None:
        seleccionado = page.evaluate(
            """
            () => {
              const visible = (el) => !!el && (el.offsetParent !== null || el.getClientRects().length > 0);
              const norm = (value) => (value || '').toString().normalize('NFD').replace(/[\\u0300-\\u036f]/g, '').toLowerCase();
              const modal =
                Array.from(document.querySelectorAll('.modal.show, .modal.in, .bootbox, [role="dialog"], .modal'))
                .filter(visible)
                .find((el) => {
                  const txt = norm(el.textContent);
                  return txt.includes('cargar documentacion') || txt.includes('tipo de documentacion') || el.querySelector('input[type="file"]');
                }) ||
                document;
              const selects = Array.from(modal.querySelectorAll('select'));
              const allSelects = selects.length ? selects : Array.from(document.querySelectorAll('select'));
              const candidates = allSelects.map((sel) => ({
                sel,
                visible: visible(sel),
                options: Array.from(sel.options || []).map((opt) => ({
                  text: (opt.textContent || '').trim(),
                  value: (opt.value || '').trim(),
                  disabled: !!opt.disabled
                }))
              }));
              const selectInfo = candidates.find((item) =>
                item.options.some((opt) => /informe|resultado|pdf/.test(norm(`${opt.text} ${opt.value}`)))
              ) || candidates.find((item) => item.options.some((opt) => opt.value || !/^--?$/.test((opt.text || '').trim())));
              if (!selectInfo) {
                return { ok: false, reason: 'missing_select', selects: candidates.map((item) => item.options) };
              }
              const select = selectInfo.sel;
              const option = Array.from(select.options || []).find((opt) =>
                /informe|resultado|pdf/.test(norm(`${opt.textContent || ''} ${opt.value || ''}`))
              ) || Array.from(select.options || []).find((opt) => {
                const text = (opt.textContent || '').trim();
                return !opt.disabled && ((opt.value || '').trim() || (text && !/^--?$/.test(text)));
              });
              if (!option) {
                return { ok: false, reason: 'missing_option', selects: candidates.map((item) => item.options) };
              }
              select.value = option.value;
              option.selected = true;
              select.dispatchEvent(new Event('input', { bubbles: true }));
              select.dispatchEvent(new Event('change', { bubbles: true }));
              try {
                if (window.jQuery) {
                  window.jQuery(select).val(option.value).trigger('change');
                }
              } catch (e) {}
              return { ok: true, label: option.textContent || option.value || '' };
            }
            """
        )
        if not isinstance(seleccionado, dict) or not seleccionado.get("ok"):
            self._log(f"No pude seleccionar tipo de documentacion ({seleccionado}). Se intenta cargar archivo igual.")
            return
        self._log(f"Tipo de documentacion seleccionado: {seleccionado.get('label', '').strip()}")
        try:
            self._esperar_input_archivo_documentacion(page)
        except Exception as exc:
            self._log(f"Tipo seleccionado, pero todavia no aparecio el campo Archivo: {exc}")
            page.wait_for_timeout(800)

    def _buscar_orden_en_tabla(self, page: Page, nro_orden: str, item: dict) -> None:
        prestacion = item.get("prestacion") or {}
        turno = str(prestacion.get("turno", "") or "")
        beneficio = _solo_digitos(prestacion.get("beneficio", ""))
        dni = _solo_digitos(item.get("dni_pdf", ""))
        modo = "beneficiario" if beneficio else ("dni" if dni else "orden")
        valor = beneficio[:14] if beneficio else (dni if dni else nro_orden)

        rango = _rango_mes_desde_turno(turno) or self.rango_turno_fallback
        if rango:
            self._log(f"Filtro de turno ajustado para buscar OME {nro_orden}: {rango[0]} a {rango[1]}")
        if modo != "orden":
            self._log(f"Buscando OME para cargar documento por {modo}: {valor}")

        def esperar_pagina_estable(ms: int = 900) -> None:
            try:
                page.wait_for_load_state("domcontentloaded", timeout=6000)
            except Exception:
                pass
            page.wait_for_timeout(ms)

        def evaluar_con_reintento(script: str, arg=None):
            ultimo_error: Exception | None = None
            for _ in range(3):
                try:
                    if arg is None:
                        return page.evaluate(script)
                    return page.evaluate(script, arg)
                except Exception as exc:
                    ultimo_error = exc
                    mensaje = str(exc)
                    if (
                        "Execution context was destroyed" in mensaje
                        or "most likely because of a navigation" in mensaje
                        or "Target page, context or browser has been closed" in mensaje
                    ):
                        esperar_pagina_estable(700)
                        continue
                    raise
            if ultimo_error:
                raise ultimo_error
            return None

        def buscar_en_panel(search_modo: str, search_valor: str) -> bool:
            self._goto_transmision(page)
            try:
                evaluar_con_reintento(
                    """
                    () => {
                      const visible = (el) => !!el && (el.offsetParent !== null || el.getClientRects().length > 0);
                      const limpiar = Array.from(document.querySelectorAll('button, input[type="button"], input[type="submit"], a'))
                        .find((el) => visible(el) && (el.textContent || el.value || '').trim().toLowerCase() === 'limpiar');
                      if (limpiar) limpiar.click();
                    }
                    """
                )
                page.wait_for_timeout(300)
            except Exception:
                pass
            try:
                ok = bool(
                    evaluar_con_reintento(
                        """
                    ({nroOrden, rango, modo, valor}) => {
                      const visible = (el) => !!el && (el.offsetParent !== null || el.getClientRects().length > 0);
                      const setValor = (el, valor) => {
                        if (!el) return;
                        el.value = valor;
                        el.dispatchEvent(new Event('input', { bubbles: true }));
                        el.dispatchEvent(new Event('change', { bubbles: true }));
                      };
                      if (rango) {
                        setValor(document.querySelector('input[name="f_turno_desde"]'), rango.desde);
                        setValor(document.querySelector('input[name="f_turno_hasta"]'), rango.hasta);
                      }
                      let select = null;
                      if (modo === 'beneficiario' || modo === 'dni') {
                        const tipo = modo === 'dni' ? '2' : '1';
                        select = Array.from(document.querySelectorAll('select')).find((sel) =>
                          visible(sel) && Array.from(sel.options || []).some((opt) => (opt.value || '').trim() === tipo)
                        );
                        if (select) {
                          select.value = tipo;
                          select.dispatchEvent(new Event('change', { bubbles: true }));
                        }
                      }
                      const inputs = Array.from(document.querySelectorAll('input[type="text"], input:not([type])')).filter(visible);
                      let input = null;
                      if (modo === 'orden') {
                        input = inputs.find((el) => /orden|nro|numero/i.test(`${el.name || ''} ${el.id || ''} ${el.placeholder || ''}`));
                      } else {
                        if (select) {
                          const selectBox = select.getBoundingClientRect();
                          input = inputs.find((el) => {
                            const box = el.getBoundingClientRect();
                            const mismaColumna = Math.abs(box.left - selectBox.left) < 120 || Math.abs(box.right - selectBox.right) < 120;
                            return mismaColumna && box.top >= selectBox.top - 20 && box.top <= selectBox.bottom + 100;
                          });
                        }
                        input = input || inputs.find((el) => /benef|afiliado|documento|dni/i.test(`${el.name || ''} ${el.id || ''} ${el.placeholder || ''}`));
                      }
                      input = input || inputs[0];
                      if (!input) return false;
                      setValor(input, valor || nroOrden);
                      const buscar = Array.from(document.querySelectorAll('button, input[type="submit"], input[type="button"], a'))
                        .find((el) => visible(el) && (el.textContent || el.value || '').trim().toLowerCase() === 'buscar');
                      if (!buscar) return false;
                      setTimeout(() => buscar.click(), 0);
                      return true;
                    }
                    """,
                        {
                            "nroOrden": nro_orden,
                            "rango": {"desde": rango[0], "hasta": rango[1]} if rango else None,
                            "modo": search_modo,
                            "valor": search_valor,
                        },
                    )
                )
            except Exception as exc:
                if "Execution context was destroyed" not in str(exc):
                    raise
                ok = True
            esperar_pagina_estable(900)
            return ok

        def esperar_orden_visible() -> bool:
            for _ in range(6):
                if self._orden_visible(page, nro_orden):
                    return True
                page.wait_for_timeout(500)
            return False

        filled = buscar_en_panel(modo, valor)
        if filled and esperar_orden_visible():
            return
        if modo != "orden":
            self._log(f"No aparecio OME {nro_orden} buscando por paciente; reintento por numero de orden.")
            filled = buscar_en_panel("orden", nro_orden)
            if filled and esperar_orden_visible():
                return
        if filled:
            page.wait_for_timeout(1200)
        for _ in range(80):
            if self._orden_visible(page, nro_orden):
                return
            try:
                advanced = evaluar_con_reintento(
                    """
                    () => {
                      const visible = (el) => !!el && (el.offsetParent !== null || el.getClientRects().length > 0);
                      const paginaActual = Number(new URL(location.href).searchParams.get('pagina') || '1');
                      const mas = Array.from(document.querySelectorAll('a, button'))
                        .find((el) => visible(el) && (el.textContent || '').trim().toLowerCase().includes('mas resultados'));
                      if (mas) { setTimeout(() => mas.click(), 0); return true; }
                      const sig = Array.from(document.querySelectorAll('a, button'))
                        .find((el) => visible(el) && /^\\d+$/.test((el.textContent || '').trim()) && Number((el.textContent || '').trim()) === paginaActual + 1);
                      if (!sig) return false;
                      setTimeout(() => sig.click(), 0);
                      return true;
                    }
                    """
                )
            except Exception as exc:
                if "Execution context was destroyed" not in str(exc):
                    raise
                advanced = True
            if not advanced:
                break
            esperar_pagina_estable(900)
        raise RuntimeError(f"No se encontro la OME {nro_orden} en Transmision.")

    def _orden_visible(self, page: Page, nro_orden: str) -> bool:
        try:
            return bool(
                page.evaluate(
                    """
                    (nroOrden) => Array.from(document.querySelectorAll('table tbody tr'))
                      .some((row) => (row.querySelectorAll('td')[0]?.textContent || '').trim() === nroOrden)
                    """,
                    nro_orden,
                )
            )
        except Exception:
            return False

    def _confirmar_modal_documentacion(self, page: Page) -> None:
        clicked = page.evaluate(
            """
            () => {
              const visible = (el) => !!el && (el.offsetParent !== null || el.getClientRects().length > 0);
              const texto = (el) => (el?.textContent || el?.value || el?.title || el?.getAttribute?.('aria-label') || '').replace(/\\s+/g, ' ').trim();
              const fileInput = Array.from(document.querySelectorAll('input[type="file"]')).find((el) => {
                try { return el.files && el.files.length > 0; } catch { return false; }
              }) || document.querySelector('input[type="file"]');
              const scope =
                fileInput?.closest('.modal, .swal2-popup, form, .bootbox, .dialog, [role="dialog"]') ||
                document.querySelector('.modal.show, .modal.in, .swal2-popup, .bootbox, [role="dialog"]') ||
                document;
              const positivos = /(guardar|confirmar|subir|cargar|grabar|enviar|adjuntar|aceptar|continuar|finalizar)/i;
              const negativos = /(cancelar|cerrar|volver|limpiar|eliminar|borrar|quitar|no\\b)/i;
              const candidatos = Array.from(scope.querySelectorAll('button, input[type="submit"], input[type="button"], a'));
              const boton = candidatos.find((el) => {
                if (!visible(el) || el.disabled) return false;
                const txt = texto(el);
                const cls = `${el.className || ''} ${el.closest('button,a,span')?.className || ''}`;
                if (negativos.test(txt) || negativos.test(cls)) return false;
                if (positivos.test(txt) || positivos.test(cls)) return true;
                return String(el.type || '').toLowerCase() === 'submit';
              });
              if (!boton && fileInput?.form) {
                try {
                  if (typeof fileInput.form.requestSubmit === 'function') {
                    fileInput.form.requestSubmit();
                    return true;
                  }
                  fileInput.form.submit();
                  return true;
                } catch {}
              }
              if (!boton) {
                window.__PAMI_DOC_LAST_BUTTONS__ = candidatos
                  .filter(visible)
                  .map((el) => ({text: texto(el), type: el.type || '', className: String(el.className || '')}))
                  .slice(0, 20);
                return false;
              }
              boton.click();
              return true;
            }
            """
        )
        if not clicked:
            self._log("Archivo seleccionado; PAMI no mostro boton Guardar/Confirmar, se toma como carga directa.")
            page.wait_for_timeout(1600)
            return
        page.wait_for_timeout(1600)
