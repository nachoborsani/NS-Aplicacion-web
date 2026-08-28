"""Sincroniza la bandeja del mes de cada cliente: PAMI -> web.

Para cada cliente: baja usuario/clave PAMI de la web, abre PAMI (sin ventana),
exporta la bandeja del mes (mismos filtros que Transmisión, dejando *validada* y
*transmitida* vacías -> trae TODA la bandeja), la parsea y la sube a la web para
que se vea en el panel "Dashboard mes en curso".

Pensado para correr de noche (tarea programada de Windows). NO transmite nada:
solo exporta y sube. Uso manual:  python bandeja_sync.py [2026-07] [slug1,slug2]
"""

from __future__ import annotations

import calendar
import json
import sys
import tempfile
import time
import traceback
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

from ns_web import DEFAULT_BASE_URL, NSWebClient, load_config

_MESES = ["", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio",
          "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]

# Techo de espera para la transmisión de un cliente (el bot puede hacer varios
# barridos). Generoso para el backlog de la 1ra corrida; las diarias terminan
# en segundos. Si vence, seguimos con la descarga igual y reportamos lo que haya.
_TRANSMIT_TIMEOUT_S = 1800
# Corridas a partir de esta hora se consideran "fin del día": transmiten y bajan
# hasta hoy. Antes de esa hora (ej. una corrida manual al mediodía): read-only,
# hasta ayer. El schedule por defecto es una sola corrida a las 20:00.
_HORA_FIN_DIA_H = 19
# Clientes que NO se bajan en el barrido automático (por ahora, por decisión del
# user). Se pueden bajar igual pidiéndolos explícitamente por only_slugs.
EXCLUIDOS_AUTO = {"navarro-mc", "scheffelaar-mc"}


def _current_period() -> str:
    t = date.today()
    return f"{t.year}-{t.month:02d}"


# --------------------------------------------------------------------------- #
# Clasificación del error de un cliente (para avisar distinto)
# --------------------------------------------------------------------------- #
def clasificar_error(msg: str) -> str:
    """Clasifica el error de un sync en: 'credencial' (clave equivocada),
    'sin_clave' (no cargada), 'transitorio' (502/red/timeout) u 'otro'.

    Sirve para avisar distinto: una clave mal hay que ir a corregirla; un 502 se
    reintenta solo. La señal de clave mal viene del login de PAMI
    (`_auto_login_cup` lanza 'No se pudo iniciar sesión ... Revisa usuario y clave')."""
    m = (msg or "").lower()
    if not m:
        return "otro"
    if "sin usuario/clave" in m or "cargados en la web" in m:
        return "sin_clave"
    if ("revisa usuario y clave" in m or "iniciar sesion" in m or "iniciar sesión" in m
            or "usuario y clave" in m or "usuario o clave" in m):
        return "credencial"
    # El fetch de credenciales a la web caído (502) NO es clave mal, es transitorio.
    if m.startswith("credenciales:"):
        return "transitorio"
    if any(tok in m for tok in ("502", "503", "504", "timeout", "timed out",
                                "econnreset", "econnrefused", "getaddrinfo",
                                "connection", "temporarily", "network")):
        return "transitorio"
    return "otro"


def error_legible(msg: str) -> str:
    """Traduce el error técnico a algo entendible para el aviso de Telegram."""
    m = (msg or "").lower()
    if not m:
        return "error desconocido"
    if "execution context was destroyed" in m or "page.evaluate" in m or "navigation" in m:
        return "PAMI se recargó mientras leía"
    if "no encontré el botón exportar" in m or "no encontre el boton exportar" in m or "exportar" in m:
        return "PAMI tardó en mostrar el botón Exportar"
    if "timeout" in m or "timed out" in m:
        return "PAMI tardó demasiado en responder"
    if "502" in m or "503" in m or "504" in m or "bad gateway" in m:
        return "el servidor estaba caído un momento"
    if m.startswith("credenciales:"):
        return "no pude leer la clave desde la web"
    if "econnreset" in m or "econnrefused" in m or "getaddrinfo" in m or "network" in m or "connection" in m:
        return "se cortó la conexión"
    # Si no lo reconozco, devuelvo un recorte corto (sin jerga larga).
    return str(msg or "").strip()[:50]


def armar_aviso_telegram(resultados: list[dict], periodo: str, titulo: str = "Bandejas") -> str:
    """Mensaje de Telegram que separa 'clave equivocada' de 'error transitorio'."""
    ok = [r for r in resultados if r.get("ok")]
    fallo = [r for r in resultados if not r.get("ok")]
    if not fallo:
        return f"✅ <b>{titulo}</b> {periodo}: todo ok — {len(ok)} clientes actualizados."
    cred = [r for r in fallo if clasificar_error(r.get("error")) == "credencial"]
    sin_clave = [r for r in fallo if clasificar_error(r.get("error")) == "sin_clave"]
    otros = [r for r in fallo if clasificar_error(r.get("error")) in ("transitorio", "otro")]
    lineas = [f"⚠️ <b>{titulo}</b> {periodo}: {len(ok)}/{len(resultados)} ok."]
    if cred:
        lineas.append("🔑 <b>Clave equivocada</b> — revisar acceso PAMI: "
                      + ", ".join(r.get("name", "") for r in cred))
    if sin_clave:
        lineas.append("🔒 <b>Sin clave cargada</b>: "
                      + ", ".join(r.get("name", "") for r in sin_clave))
    if otros:
        det = ", ".join(f"{r.get('name', '')} ({error_legible(r.get('error'))})" for r in otros)
        lineas.append("↻ Se reintenta solo: " + det)
    return "\n".join(lineas)


# --------------------------------------------------------------------------- #
# Transmisión automática (1 vez por día por cliente)
# --------------------------------------------------------------------------- #
def _transmit_state_path() -> Path:
    try:
        from app_paths import get_data_dir
        return get_data_dir() / "transmit_state.json"
    except Exception:
        return Path(tempfile.gettempdir()) / "transmit_state.json"


def _load_transmit_state() -> dict:
    try:
        return json.loads(_transmit_state_path().read_text("utf-8"))
    except Exception:
        return {}


def _ya_transmitido_hoy(slug: str) -> bool:
    return _load_transmit_state().get(slug) == date.today().isoformat()


def _marcar_transmitido_hoy(slug: str) -> None:
    st = _load_transmit_state()
    st[slug] = date.today().isoformat()
    try:
        p = _transmit_state_path()
        p.parent.mkdir(parents=True, exist_ok=True)
        p.write_text(json.dumps(st, ensure_ascii=False, indent=2), "utf-8")
    except Exception:
        pass


def _correr_transmision(bot, desde: str, hasta: str, progress=None) -> dict:
    """Corre el bot de transmisión (solo valida=Sí, transmitida=No) y espera a que
    termine. El bot solo transmite OMEs con informe cargado; las que les falta
    informe las saltea. Devuelve {transmitidas, errores, lastError, status}."""
    bot.iniciar_bot({
        "fecha_desde": desde, "fecha_hasta": hasta,
        "validada": "Si", "transmitida": "No",
    })
    estado: dict = {}
    completo = False
    fin = time.monotonic() + _TRANSMIT_TIMEOUT_S
    while time.monotonic() < fin:
        estado = bot.obtener_estado() or {}
        if progress:
            progress(f"transmitiendo… {estado.get('procesados', 0)} enviadas")
        if estado.get("status") in ("DONE", "ERROR"):
            completo = True
            break
        time.sleep(3)
    detalle = estado.get("omitidosDetalle")
    detalle = detalle if isinstance(detalle, list) else []
    return {
        "transmitidas": int(estado.get("procesados") or 0),
        "errores": int(estado.get("errores") or 0),
        "lastError": str(estado.get("lastError") or ""),
        "status": str(estado.get("status") or "TIMEOUT"),
        "completo": completo,  # False = cortó por timeout, quedaron pendientes
        # OMEs que no se pudieron transmitir + la leyenda de PAMI (ej. inactivo).
        "omitidosDetalle": detalle,
    }


def month_range(period: str, hasta_hoy: bool = False) -> tuple[str, str, str]:
    """'2026-07' -> ('01/07/2026', '31/07/2026', 'Julio 2026'). Fechas en formato PAMI.

    Para el MES EN CURSO cortamos en AYER (días cerrados), no en fin de mes: así no
    traemos turnos futuros ya agendados que todavía no pasaron. En la corrida de la
    tarde (fin del día, hasta_hoy=True) sí incluimos HOY, porque los consultorios
    ya cerraron. Los meses pasados se bajan completos.
    """
    year, month = int(period[:4]), int(period[5:7])
    last = calendar.monthrange(year, month)[1]
    hasta_dia = last
    hoy = date.today()
    if (year, month) == (hoy.year, hoy.month):
        hasta_dia = hoy.day if hasta_hoy else max(1, hoy.day - 1)  # hasta hoy / hasta ayer
    return f"01/{month:02d}/{year}", f"{hasta_dia:02d}/{month:02d}/{year}", f"{_MESES[month]} {year}"


def forward_range(period: str, hasta_hoy: bool = False) -> tuple[str, str, str] | None:
    """Rango 'hacia adelante': del día SIGUIENTE al corte del mes en curso hasta fin
    de mes. Ej: si el mes en curso corta el 18/08 (corrida de la tarde), adelante
    va del 19/08 al 31/08. Así el día que pasa de futuro a presente se mete solo en
    el mes en curso y no se solapa. Devuelve None si no quedan días futuros.
    """
    year, month = int(period[:4]), int(period[5:7])
    last = calendar.monthrange(year, month)[1]
    hoy = date.today()
    if (year, month) == (hoy.year, hoy.month):
        corte = hoy.day if hasta_hoy else max(1, hoy.day - 1)
    else:
        corte = last  # meses pasados: no hay futuro
    desde_dia = corte + 1
    if desde_dia > last:
        return None
    return f"{desde_dia:02d}/{month:02d}/{year}", f"{last:02d}/{month:02d}/{year}", f"{_MESES[month]} {year}"


def future_month_ranges(period: str, n: int = 2) -> list[tuple[str, str, str, str]]:
    """Los próximos N meses COMPLETOS después de `period`, cada uno como
    (period, desde, hasta, label). Se bajan APARTE (un filtro por mes) para no
    hacer explotar el export. Ej period=2026-08 → sep y oct 2026 completos."""
    year, month = int(period[:4]), int(period[5:7])
    out = []
    for i in range(1, n + 1):
        m, y = month + i, year
        while m > 12:
            m -= 12
            y += 1
        last = calendar.monthrange(y, m)[1]
        out.append((f"{y}-{m:02d}", f"01/{m:02d}/{y}", f"{last:02d}/{m:02d}/{y}", f"{_MESES[m]} {y}"))
    return out


def parse_bandeja_excel(path: str) -> tuple[list[dict], list[str]]:
    """Lee el Excel exportado y devuelve (filas como dicts, columnas). La primera
    fila no vacía se toma como encabezado."""
    workbook = load_workbook(path, read_only=True, data_only=True)
    ws = workbook.active
    rows_iter = ws.iter_rows(values_only=True)

    header = None
    for raw in rows_iter:
        if raw and any(cell not in (None, "") for cell in raw):
            header = [str(cell).strip() if cell is not None else "" for cell in raw]
            break
    if not header:
        return [], []

    columns = [h for h in header if h]
    data: list[dict] = []
    for raw in rows_iter:
        if not raw or all(cell in (None, "") for cell in raw):
            continue
        row: dict = {}
        for i, name in enumerate(header):
            if not name:
                continue
            value = raw[i] if i < len(raw) else None
            row[name] = "" if value is None else str(value)
        data.append(row)
    return data, columns


def _contar_transmitidas(rows: list[dict]) -> int:
    """Cuántas filas del export están transmitidas (columna TRASMITIDA = 'S')."""
    if not rows:
        return 0
    kt = next((k for k in rows[0] if "TRASMITIDA" in k.upper() or "TRANSMITIDA" in k.upper()), None)
    if not kt:
        return 0
    return sum(1 for r in rows if str(r.get(kt, "")).strip().upper() == "S")


def sync_client(web: NSWebClient, client: dict, period: str, progress=None,
                transmitir: bool = False, forzar_transmision: bool = False) -> dict:
    """Baja la bandeja de un cliente y la sube. Si transmitir=True y todavía no se
    transmitió hoy, corre el bot de transmisión ANTES de exportar (1x/día). Nunca
    lanza: devuelve el resultado."""
    slug = client.get("slug", "")
    name = client.get("name", slug)
    try:
        cred = web.client_pami(slug)
    except Exception as exc:  # noqa: BLE001
        return {"slug": slug, "name": name, "ok": False, "error": f"credenciales: {exc}"}
    if not cred.get("pamiUser") or not cred.get("pamiPassword"):
        return {"slug": slug, "name": name, "ok": False, "error": "sin usuario/clave PAMI cargados en la web"}

    # Cliente EN ANÁLISIS (potencial): se baja la bandeja para analizar pero NUNCA
    # se transmite (no somos su facturador). El flag manda por sobre todo, incluso
    # una corrida forzada. Reemplaza al hardcode de EXCLUIDOS_AUTO.
    if client.get("enAnalisis"):
        transmitir = False

    # La corrida de la tarde (fin del día) transmite y baja HASTA HOY (los
    # consultorios ya cerraron); la del mediodía es read-only y hasta ayer.
    # forzar_transmision (corrida manual): transmite a CUALQUIER hora. El user lo
    # pidió explícito — "cada vez que bajes la bandeja, antes transmití, no importa
    # el horario". El scheduler automático NO fuerza (conserva la ventana de las 19h).
    es_fin_dia = forzar_transmision or datetime.now().hour >= _HORA_FIN_DIA_H
    desde, hasta, label = month_range(period, hasta_hoy=es_fin_dia)
    tmp = Path(tempfile.gettempdir()) / f"bandeja_{slug}_{period}.xlsx"
    transmit_info = None
    exported_futuros = []  # (fperiod, flabel, path) de los meses futuros
    try:
        from pami_transmision import PamiTransmisionController

        bot = PamiTransmisionController()
        try:
            bot.abrir_pami(usuario=cred["pamiUser"], clave=cred["pamiPassword"], headless=True)
            # 1) Transmitir pendientes SOLO en la corrida de la tarde (1x/día). El bot
            #    solo transmite las que tienen informe cargado; el resto queda para
            #    "faltan informes".
            if transmitir and es_fin_dia and not _ya_transmitido_hoy(slug):
                if progress:
                    progress(f"{name}: transmitiendo pendientes…")
                transmit_info = _correr_transmision(bot, desde, hasta, progress=progress)
                _marcar_transmitido_hoy(slug)
            # 2) Exportar la bandeja completa (validada/transmitida vacías).
            if progress:
                progress(f"{name}: bajando bandeja…")
            exported = bot.exportar_excel_panel(str(tmp), {"fecha_desde": desde, "fecha_hasta": hasta})
            # Auto-corrección: si transmitimos N>0 pero el export no trae NINGUNA
            # transmitida, PAMI arrastró el filtro de la transmisión (validada=S,
            # transmitida=N) y exportó solo ese subconjunto (bug detectado en CIMA:
            # 476 filas en vez de 3707, cobro real $0). Re-exportamos una vez.
            if transmit_info and (transmit_info.get("transmitidas") or 0) > 0:
                try:
                    rows_chk, _ = parse_bandeja_excel(exported)
                    if _contar_transmitidas(rows_chk) == 0:
                        if progress:
                            progress(f"{name}: el export salió filtrado por la transmisión, re-exportando…")
                        exported = bot.exportar_excel_panel(str(tmp), {"fecha_desde": desde, "fecha_hasta": hasta})
                except Exception:  # noqa: BLE001 - si el chequeo falla, seguimos con lo que haya
                    pass
            # 3) Bandeja "hacia adelante": turnos del día siguiente al corte hasta
            #    fin de mes (para detectar posibles débitos por adelantado).
            fut = forward_range(period, hasta_hoy=es_fin_dia)
            exported_adelante = None
            fut_range = None
            if fut is not None:
                desde_f, hasta_f, _ = fut
                fut_range = (desde_f, hasta_f)
                if progress:
                    progress(f"{name}: bajando turnos futuros ({desde_f} a {hasta_f})…")
                tmp_f = Path(tempfile.gettempdir()) / f"bandeja_adelante_{slug}_{period}.xlsx"
                exported_adelante = bot.exportar_excel_panel(str(tmp_f), {"fecha_desde": desde_f, "fecha_hasta": hasta_f})
            # 4) Meses FUTUROS completos (sep, oct): cada uno APARTE con su propio
            #    filtro (para no hacer explotar el export en centros grandes).
            for fperiod, fdesde, fhasta, flabel in future_month_ranges(period, n=2):
                try:
                    if progress:
                        progress(f"{name}: bajando {flabel} ({fdesde} a {fhasta})…")
                    tmp_fut = Path(tempfile.gettempdir()) / f"bandeja_fut_{slug}_{fperiod}.xlsx"
                    exp_fut = bot.exportar_excel_panel(str(tmp_fut), {"fecha_desde": fdesde, "fecha_hasta": fhasta})
                    exported_futuros.append((fperiod, flabel, exp_fut))
                except Exception:  # noqa: BLE001 - un mes futuro que falla no corta el resto
                    pass
        finally:
            try:
                bot.cerrar()
            except Exception:  # noqa: BLE001
                pass

        rows, columns = parse_bandeja_excel(exported)
        web.upload_bandeja(slug, period, rows, columns=columns, month_label=label,
                           generated_at=date.today().isoformat())
        # Subir la bandeja de adelante (si había días futuros). Un fallo acá no
        # tumba el sync del mes en curso.
        adelante_count = None
        try:
            if exported_adelante is not None:
                rows_f, cols_f = parse_bandeja_excel(exported_adelante)
                web.upload_bandeja_adelante(slug, period, rows_f, columns=cols_f, month_label=label,
                                            generated_at=date.today().isoformat())
                adelante_count = len(rows_f)
        except Exception:  # noqa: BLE001
            pass
        # Subir los meses FUTUROS (sep, oct), cada uno aparte. Un fallo no tumba
        # el sync del mes en curso.
        futuros_count = 0
        for fperiod, flabel, exp_fut in exported_futuros:
            try:
                rows_fut, cols_fut = parse_bandeja_excel(exp_fut)
                web.upload_bandeja_futura(slug, fperiod, rows_fut, columns=cols_fut,
                                          month_label=flabel, generated_at=date.today().isoformat())
                futuros_count += 1
            except Exception:  # noqa: BLE001
                pass
        res = {"slug": slug, "name": name, "ok": True, "count": len(rows)}
        if futuros_count:
            res["futuros"] = futuros_count
        if adelante_count is not None:
            res["adelante"] = adelante_count
        if transmit_info is not None:
            res["transmit"] = transmit_info
        return res
    except Exception as exc:  # noqa: BLE001
        return {"slug": slug, "name": name, "ok": False, "error": str(exc),
                "trace": traceback.format_exc()}


def sync_all(period: str | None = None, only_slugs: list[str] | None = None,
             base_url: str = "", admin_user: str = "", admin_pass: str = "",
             progress=None, transmitir: bool | None = None,
             forzar_transmision: bool = False) -> list[dict]:
    """Recorre los clientes de la web y sincroniza la bandeja de cada uno.

    Se conecta con una cuenta ADMIN (el endpoint de credenciales es admin-only).
    Por defecto usa la conexión guardada en 'Conexión con NS'. Si transmitir es
    None, lo lee de la config del panel 'Refresco automático'.
    """
    cfg = load_config()
    web = NSWebClient(base_url or cfg.get("base_url") or DEFAULT_BASE_URL)
    web.login(admin_user or cfg.get("username", ""), admin_pass or cfg.get("password", ""))
    period = period or _current_period()
    if forzar_transmision:
        transmitir = True  # forzar implica transmitir sí o sí
    if transmitir is None:
        try:
            import bandeja_schedule
            transmitir = bool(bandeja_schedule.load_config().get("transmitir"))
        except Exception:  # noqa: BLE001
            transmitir = False

    results: list[dict] = []
    for client in web.list_clients():
        slug = client.get("slug")
        if only_slugs and slug not in only_slugs:
            continue
        # Los médicos de cabecera NO bajan bandeja (su flujo es benef/credenciales,
        # no tienen consultas/prácticas para transmitir). Se excluyen por tipo, así
        # cualquier MC actual o futuro queda afuera solo. También se excluyen los
        # slugs de EXCLUIDOS_AUTO. Si se piden por only_slugs, igual se bajan.
        if not only_slugs and (client.get("tipo") == "med_cabecera" or slug in EXCLUIDOS_AUTO):
            if progress:
                progress(f"{client.get('name', slug)}: omitido (no baja bandeja)")
            continue
        res = sync_client(web, client, period, progress=progress, transmitir=transmitir,
                          forzar_transmision=forzar_transmision)
        # Reportamos el resultado (ok/error + transmisión) para el indicador de salud.
        try:
            t = res.get("transmit") or {}
            web.report_bandeja_estado(
                res.get("slug", ""), res.get("ok"), res.get("count"), res.get("error", ""),
                transmitidas=t.get("transmitidas"), transmit_errores=t.get("errores"),
                transmit_error=t.get("lastError"),
                omitidos_detalle=t.get("omitidosDetalle"),
            )
        except Exception:  # noqa: BLE001 - un fallo del reporte no corta el sync
            pass
        results.append(res)
    return results


if __name__ == "__main__":  # pragma: no cover
    # Flags: --transmitir (fuerza transmitir=True), --forzar (transmite a cualquier
    # hora, saltando el candado de las 19h; implica --transmitir). Uso manual:
    #   python bandeja_sync.py 2026-08 caballito-pediatrico --forzar
    # La consola de Windows (cp1252) no sabe imprimir algunos unicode: el "→" del
    # aviso final cortaba con UnicodeEncodeError. UTF-8 con reemplazo: ningún print corta.
    for _s in (sys.stdout, sys.stderr):
        try:
            _s.reconfigure(encoding="utf-8", errors="replace")
        except Exception:
            pass
    _pos = [a for a in sys.argv[1:] if not a.startswith("--")]
    _flags = [a for a in sys.argv[1:] if a.startswith("--")]
    period_arg = _pos[0] if len(_pos) > 0 else None
    slugs_arg = _pos[1].split(",") if len(_pos) > 1 else None
    forzar = "--forzar" in _flags
    transmitir_arg = True if ("--transmitir" in _flags or forzar) else None
    print(f"Sincronizando bandeja {period_arg or _current_period()} "
          + ("(transmitiendo, forzado) …" if forzar else "…"))
    resultados = sync_all(period_arg, only_slugs=slugs_arg, transmitir=transmitir_arg,
                          forzar_transmision=forzar, progress=lambda m: print("  ", m))
    for r in resultados:
        estado = f"OK — {r.get('count')} filas" if r.get("ok") else f"FALLO — {r.get('error')}"
        print(f"  {r['name'][:32]:32} -> {estado}")

    # Aviso por Telegram: que esté todo ok (o qué clientes fallaron), separando
    # 'clave equivocada' (hay que corregirla) de 'error transitorio' (se reintenta).
    try:
        msg = armar_aviso_telegram(resultados, _current_period())
        cfg = load_config()
        w = NSWebClient(cfg.get("base_url") or DEFAULT_BASE_URL)
        w.login(cfg.get("username", ""), cfg.get("password", ""))
        w.avisar(msg)
        print("  → aviso Telegram enviado")
    except Exception as e:  # noqa: BLE001 - el aviso no debe cortar el barrido
        print(f"  → no pude enviar el aviso Telegram: {e!r}")
