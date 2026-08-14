import os
import queue
import threading
import json
import webbrowser
import shutil
import tempfile
from pathlib import Path
from threading import Event
from tkinter import filedialog, messagebox, ttk

import customtkinter as ctk

from ui_kit import apply_button_icon
from openpyxl import load_workbook

from app_paths import get_data_dir, get_log_file
from credencial_scraper import exportar_reporte_credenciales, procesar_lote_credenciales, ruta_reporte_credenciales_por_defecto
from excel_models import guardar_modelo_credencial
from gmail_informes import get_gmail_credentials_path
from google_drive_storage import (
    get_connected_drive_email,
    get_drive_token_path,
    resolve_child_folder_id,
    upload_file_to_drive_folder,
)
from google_sheets_credenciales import (
    OFFICE_FILE_MESSAGE,
    is_office_file_url,
    normalize_spreadsheet_url,
    repair_credencial_sheet_missing_tramites,
    read_credencial_sheet_rows,
    write_credencial_sheet_results,
)
from google_sheets_ome import extract_spreadsheet_id, get_connected_google_email, get_sheets_token_path, list_spreadsheet_sheet_names


CREDENCIAL_HEADERS = {
    "benef": {"benef", "beneficio", "beneficiario", "numero_afiliacion", "afiliacion"},
    "dni": {"dni", "documento"},
    "tramite": {"tramite", "numero_tramite", "nro_tramite", "n_tramite"},
    "sexo": {"sexo", "genero"},
}

CREDENCIAL_TABLE_COLUMNS = [
    ("benef", "BENEF"),
    ("dni", "DNI"),
    ("tramite", "N° TRAMITE"),
    ("sexo", "SEXO"),
    ("estado", "ESTADO"),
    ("archivo_pdf", "ARCHIVO PDF"),
]

DRIVE_PARENT_FOLDER_URL = os.environ.get("PAMI_CREDENCIALES_DRIVE_PARENT_URL", "")
DESTINO_LOCAL = "Local: CREDENCIALES"
DESTINO_DRIVE_SCHEFE = "Drive: CREDENCIALES SCHEFE"
DESTINO_DRIVE_CIMA = "Drive: CREDENCIALES CIMA/DUBE"
DESTINO_OPTIONS = [DESTINO_LOCAL, DESTINO_DRIVE_SCHEFE, DESTINO_DRIVE_CIMA]
DESTINO_DRIVE_FOLDER_NAMES = {
    DESTINO_DRIVE_SCHEFE: "CREDENCIALES SCHEFE",
    DESTINO_DRIVE_CIMA: "CREDENCIALES CIMA/DUBE",
}
SHEET_PROFILE_DUBE = "DUBE"
SHEET_PROFILE_SCHEFE = "SCHEFE"
SHEET_PROFILE_PLAN_SALUD_CIMA = "PLAN SALUD CIMA"
SHEET_PROFILE_OPTIONS = [SHEET_PROFILE_DUBE, SHEET_PROFILE_SCHEFE, SHEET_PROFILE_PLAN_SALUD_CIMA]
PROFILE_DESTINATION_MAP = {
    SHEET_PROFILE_DUBE: DESTINO_DRIVE_CIMA,
    SHEET_PROFILE_SCHEFE: DESTINO_DRIVE_SCHEFE,
    SHEET_PROFILE_PLAN_SALUD_CIMA: DESTINO_DRIVE_CIMA,
}
DESTINATION_PROFILE_MAP = {
    DESTINO_DRIVE_CIMA: SHEET_PROFILE_DUBE,
    DESTINO_DRIVE_SCHEFE: SHEET_PROFILE_SCHEFE,
}
LIMIT_MODE_TOPE = "Tope"
LIMIT_MODE_END_ROW = "Hasta fila"
LIMIT_MODE_OPTIONS = [LIMIT_MODE_TOPE, LIMIT_MODE_END_ROW]


def clear_local_pdf_paths_for_drive_results(resultados: list[dict]) -> None:
    for result in resultados:
        if str(result.get("estado", "")).strip().upper() != "DESCARGADA":
            continue
        drive_url = str(result.get("archivo_drive_url", "") or "").strip()
        result["archivo_pdf"] = drive_url if drive_url else ""


def normalize_limit_mode(value: str | None, *, has_end_row: bool = False) -> str:
    normalized = str(value or "").strip().upper().replace(" ", "_")
    if normalized in {"TOPE", "TOP"}:
        return LIMIT_MODE_TOPE
    if normalized in {"HASTA_FILA", "HASTA FILA", "FILA_FINAL", "FINAL"}:
        return LIMIT_MODE_END_ROW
    return LIMIT_MODE_END_ROW if has_end_row else LIMIT_MODE_TOPE


class CredencialModuleFrame(ctk.CTkFrame):
    def __init__(self, master, on_back, on_restart=None) -> None:
        super().__init__(master, fg_color="transparent")
        self.on_back = on_back
        self.on_restart = on_restart
        self.event_queue: queue.Queue = queue.Queue()
        self.worker_thread: threading.Thread | None = None
        self.processing = False
        self.cancel_requested = Event()
        self.current_results: list[dict] = []
        self.report_path: Path | None = None
        self._syncing_profile_destination = False
        self.data_dir = get_data_dir()
        self.destination_dir = self.data_dir / "CREDENCIALES"
        self.destination_dir.mkdir(parents=True, exist_ok=True)
        self.sheet_settings_file = self.data_dir / "credenciales_sheets_config.json"
        self.sheet_settings = self._load_sheet_settings()
        self.sheet_profiles = self._build_sheet_profiles()
        self.sheet_url_profiles = self._build_sheet_url_profiles()
        self.current_sheet_profile_key = str(self.sheet_settings.get("active_sheet_profile", SHEET_PROFILE_DUBE))
        if self.current_sheet_profile_key not in SHEET_PROFILE_OPTIONS:
            self.current_sheet_profile_key = SHEET_PROFILE_DUBE
        self.sheets_connected = False
        self.drive_connected = False
        self.current_destination_mode = DESTINO_LOCAL
        self.active_cell_value = ""
        self.active_column_id = ""
        self.active_column_title = ""
        self._syncing_sheet_limit_mode = False
        self.last_missing_benef_rows: list[dict] = []
        self.last_missing_tramite_rows: list[dict] = []
        self.last_recovered_tramite_rows: list[dict] = []

        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(0, weight=1)
        self._build_ui()
        self._on_destination_mode_changed(self.destination_mode_var.get())
        self.after(400, self._start_sheets_status_check)
        self.after(500, self._start_drive_status_check)
        self.after(150, self._process_ui_queue)

    def _build_ui(self) -> None:
        content = ctk.CTkScrollableFrame(self, fg_color="#eef3f8")
        content.grid(row=0, column=0, sticky="nsew")
        content.grid_columnconfigure(0, weight=1)

        top = ctk.CTkFrame(content, corner_radius=12, fg_color="#ffffff", border_width=1, border_color="#d8e2ec")
        top.grid(row=0, column=0, padx=8, pady=(6, 4), sticky="ew")
        top.grid_columnconfigure(1, weight=1)
        ctk.CTkButton(top, text="Volver", width=78, height=28, command=self._go_home, fg_color="#9aafc3", hover_color="#7f95aa").grid(
            row=0, column=0, padx=(8, 10), pady=(6, 4), sticky="w"
        )
        ctk.CTkLabel(top, text="Descargar credencial provisoria", font=ctk.CTkFont(size=20, weight="bold"), text_color="#16324f").grid(
            row=0, column=1, padx=(0, 10), pady=(6, 2), sticky="w"
        )
        self.restart_button = ctk.CTkButton(
            top,
            text="Reiniciar app",
            width=150,
            height=28,
            command=self._restart_app,
            fg_color="#66788a",
            hover_color="#536577",
        )
        self.restart_button.grid(row=0, column=2, padx=(8, 6), pady=(6, 2), sticky="e")
        self.load_excel_button = ctk.CTkButton(top, text="Cargar Excel Masivo", width=150, height=28, command=self.load_excel)
        self.load_excel_button.grid(row=0, column=3, padx=(6, 6), pady=(6, 2), sticky="e")
        self.template_button = ctk.CTkButton(top, text="Descargar Modelo Excel Masivo", width=190, height=28, command=self.download_template)
        self.template_button.grid(row=0, column=4, padx=(0, 8), pady=(6, 2), sticky="e")
        ctk.CTkLabel(
            top,
            text="Podes cargar una fila individual, pegar lotes por columnas o subir un Excel modelo.",
            font=ctk.CTkFont(size=11),
            text_color="#51657a",
        ).grid(row=1, column=1, columnspan=4, padx=(0, 8), pady=(0, 6), sticky="w")

        active_sheet_profile = self._get_sheet_profile_values(self.current_sheet_profile_key)
        self.sheet_profile_var = ctk.StringVar(value=self.current_sheet_profile_key)
        self.sheet_internal_name_var = ctk.StringVar(value=str(active_sheet_profile.get("internal_name", "")).strip())
        self.sheet_url_var = ctk.StringVar(value=normalize_spreadsheet_url(str(active_sheet_profile.get("spreadsheet_url", ""))))
        self.sheet_template_display_var = ctk.StringVar(value="")
        self.sheet_name_var = ctk.StringVar(value=str(active_sheet_profile.get("sheet_name", "")))
        self.sheet_start_row_var = ctk.StringVar(value=str(active_sheet_profile.get("start_row", 2)))
        self.sheet_max_rows_var = ctk.StringVar(value=str(active_sheet_profile.get("max_rows", "40")))
        self.sheet_end_row_var = ctk.StringVar(value=str(active_sheet_profile.get("end_row", "")))
        self.sheet_limit_mode_var = ctk.StringVar(
            value=normalize_limit_mode(
                str(active_sheet_profile.get("limit_mode", LIMIT_MODE_TOPE)),
                has_end_row=bool(str(active_sheet_profile.get("end_row", "")).strip()),
            )
        )
        self.sheet_only_with_tramite_var = ctk.BooleanVar(value=bool(active_sheet_profile.get("only_with_tramite", False)))
        self.sheets_status_var = ctk.StringVar(value="Google Sheets no conectado")
        initial_sheet_name = str(active_sheet_profile.get("sheet_name", "")).strip()
        self.sheet_tabs = [initial_sheet_name] if initial_sheet_name else []
        if self.sheet_limit_mode_var.get() not in LIMIT_MODE_OPTIONS:
            self.sheet_limit_mode_var.set(normalize_limit_mode("", has_end_row=bool(self.sheet_end_row_var.get().strip())))

        sheets = ctk.CTkFrame(content, corner_radius=8, fg_color="#ffffff", border_width=1, border_color="#d8e2ec")
        sheets.grid(row=1, column=0, padx=8, pady=(0, 5), sticky="ew")
        sheets.grid_columnconfigure(0, weight=1)
        ctk.CTkLabel(sheets, text="Google Sheets", font=ctk.CTkFont(size=16, weight="bold"), text_color="#16324f").grid(
            row=0, column=0, padx=14, pady=(10, 4), sticky="w"
        )

        row1_box = ctk.CTkFrame(sheets, fg_color="transparent")
        row1_box.grid(row=1, column=0, padx=14, pady=(0, 6), sticky="w")
        ctk.CTkLabel(row1_box, text="Perfil", text_color="#16324f").grid(row=0, column=0, padx=(0, 8), pady=0, sticky="w")
        self.sheet_profile_combo = ctk.CTkComboBox(
            row1_box, values=SHEET_PROFILE_OPTIONS, variable=self.sheet_profile_var,
            width=130, height=30, state="readonly", command=self._on_sheet_profile_changed,
        )
        self.sheet_profile_combo.grid(row=0, column=1, padx=(0, 14), pady=0, sticky="w")
        ctk.CTkLabel(row1_box, text="Hoja", text_color="#16324f").grid(row=0, column=2, padx=(0, 8), pady=0, sticky="w")
        self.sheet_url_entry = ctk.CTkEntry(row1_box, textvariable=self.sheet_template_display_var, width=220, height=30, state="readonly")
        self.sheet_url_entry.grid(row=0, column=3, padx=(0, 6), pady=0, sticky="w")
        self.sheet_url_entry.bind("<Button-1>", lambda _event: self._open_sheet_template_dialog(), add="+")
        self.sheet_url_config_button = ctk.CTkButton(
            row1_box, text="Configurar", width=110, height=30, fg_color="#245b9d", hover_color="#1d4b82", command=self._open_sheet_template_dialog,
        )
        self.sheet_url_config_button.grid(row=0, column=4, padx=(0, 14), pady=0, sticky="w")
        apply_button_icon(self.sheet_url_config_button, "gear.png", "Configurar", size=(18, 18), width=40)
        ctk.CTkLabel(row1_box, text="Pestaña", text_color="#16324f").grid(row=0, column=5, padx=(0, 8), pady=0, sticky="w")
        self.sheet_name_combo = ctk.CTkComboBox(
            row1_box, values=self.sheet_tabs or [""], variable=self.sheet_name_var, width=180, height=30,
            command=lambda _selected: (self._save_sheet_settings(), self._update_sheet_template_display()),
        )
        self.sheet_name_combo.grid(row=0, column=6, padx=(0, 14), pady=0, sticky="w")
        self.sheet_tabs_button = ctk.CTkButton(
            row1_box, text="Cargar pestañas", width=120, height=30, fg_color="#245b9d", hover_color="#1d4b82", command=self._load_sheet_tabs,
        )
        self.sheet_tabs_button.grid(row=0, column=7, padx=0, pady=0, sticky="w")

        row2_box = ctk.CTkFrame(sheets, fg_color="transparent")
        row2_box.grid(row=2, column=0, padx=14, pady=(2, 6), sticky="w")
        ctk.CTkLabel(row2_box, text="Fila inicial", text_color="#16324f").grid(row=0, column=0, padx=(0, 8), pady=0, sticky="w")
        self.sheet_start_row_entry = ctk.CTkEntry(row2_box, textvariable=self.sheet_start_row_var, width=90, height=30)
        self.sheet_start_row_entry.grid(row=0, column=1, padx=(0, 14), pady=0, sticky="w")
        ctk.CTkLabel(row2_box, text="Modo", text_color="#16324f").grid(row=0, column=2, padx=(0, 8), pady=0, sticky="w")
        self.sheet_limit_mode_combo = ctk.CTkComboBox(
            row2_box, values=LIMIT_MODE_OPTIONS, variable=self.sheet_limit_mode_var,
            width=130, height=30, state="readonly", command=self._on_sheet_limit_mode_changed,
        )
        self.sheet_limit_mode_combo.grid(row=0, column=3, padx=(0, 14), pady=0, sticky="w")
        ctk.CTkLabel(row2_box, text="Tope", text_color="#16324f").grid(row=0, column=4, padx=(0, 8), pady=0, sticky="w")
        self.sheet_max_rows_entry = ctk.CTkEntry(row2_box, textvariable=self.sheet_max_rows_var, width=90, height=30)
        self.sheet_max_rows_entry.grid(row=0, column=5, padx=(0, 14), pady=0, sticky="w")
        ctk.CTkLabel(row2_box, text="Hasta fila", text_color="#16324f").grid(row=0, column=6, padx=(0, 8), pady=0, sticky="w")
        self.sheet_end_row_entry = ctk.CTkEntry(row2_box, textvariable=self.sheet_end_row_var, width=90, height=30, placeholder_text="Opcional")
        self.sheet_end_row_entry.grid(row=0, column=7, padx=0, pady=0, sticky="w")
        self.sheets_actions = ctk.CTkFrame(sheets, fg_color="transparent")
        self.sheets_actions.grid(row=3, column=0, columnspan=13, padx=14, pady=(4, 6), sticky="w")
        self.sheet_only_with_tramite_check = ctk.CTkCheckBox(
            self.sheets_actions,
            text="Solo filas con N° trámite",
            variable=self.sheet_only_with_tramite_var,
            command=self._save_sheet_settings,
            width=170,
        )
        self.sheet_only_with_tramite_check.grid(row=0, column=0, padx=(0, 12), pady=0, sticky="w")
        self.sheets_connect_button = ctk.CTkButton(
            self.sheets_actions,
            text="Conectar Sheets",
            width=150,
            height=30,
            command=self._connect_sheets_account,
        )
        self.sheets_connect_button.grid(row=0, column=1, padx=(0, 8), pady=0, sticky="w")
        self.sheets_save_button = ctk.CTkButton(
            self.sheets_actions,
            text="Guardar config",
            width=130,
            height=30,
            fg_color="#245b9d",
            hover_color="#1d4b82",
            command=self._save_current_sheet_profile,
        )
        self.sheets_save_button.grid(row=0, column=2, padx=(0, 8), pady=0, sticky="w")
        self.sheets_scan_button = ctk.CTkButton(
            self.sheets_actions,
            text="Barrido trámite",
            width=150,
            height=30,
            fg_color="#66788a",
            hover_color="#536577",
            command=self.scan_sheet_missing_tramites,
        )
        self.sheets_scan_button.grid(row=0, column=3, padx=(0, 8), pady=0, sticky="w")
        self.sheets_run_button = ctk.CTkButton(
            self.sheets_actions,
            text="Descargar desde Sheets",
            width=190,
            height=30,
            command=self.start_sheet_processing,
        )
        self.sheets_run_button.grid(row=0, column=4, padx=0, pady=0, sticky="w")
        self.sheets_status_label = ctk.CTkLabel(
            sheets,
            textvariable=self.sheets_status_var,
            font=ctk.CTkFont(size=11),
            text_color="#66788a",
        )
        self.sheets_status_label.grid(row=4, column=0, columnspan=13, padx=14, pady=(0, 10), sticky="w")
        self._apply_sheet_limit_mode()
        self._update_sheet_template_display()

        carga_row = ctk.CTkFrame(content, fg_color="transparent")
        carga_row.grid(row=2, column=0, padx=8, pady=(0, 5), sticky="ew")
        carga_row.grid_columnconfigure(2, weight=1)

        manual = ctk.CTkFrame(carga_row, corner_radius=8, fg_color="#ffffff", border_width=1, border_color="#d8e2ec")
        manual.grid(row=0, column=0, padx=(0, 6), pady=0, sticky="nsew")
        manual.grid_columnconfigure(0, weight=0)
        manual.grid_columnconfigure(1, weight=0)
        manual.grid_columnconfigure(2, weight=1)
        ctk.CTkLabel(manual, text="Carga individual", font=ctk.CTkFont(size=16, weight="bold"), text_color="#16324f").grid(
            row=0, column=0, columnspan=3, padx=16, pady=(12, 10), sticky="w"
        )

        self.individual_vars = {
            "benef": ctk.StringVar(),
            "dni": ctk.StringVar(),
            "tramite": ctk.StringVar(),
            "sexo": ctk.StringVar(value="MASC"),
        }
        labels = [("BENEF", "benef"), ("DNI", "dni"), ("N° TRÁMITE", "tramite"), ("SEXO", "sexo")]
        for idx, (title, key) in enumerate(labels):
            ctk.CTkLabel(
                manual, text=title, font=ctk.CTkFont(size=13, weight="bold"), text_color="#16324f", width=110, anchor="w"
            ).grid(row=1 + idx, column=0, padx=(16, 10), pady=(0, 8), sticky="w")
            if key == "sexo":
                widget = ctk.CTkComboBox(manual, values=["MASC", "FEM", "OTRO"], variable=self.individual_vars[key], state="readonly", width=100, height=30)
            else:
                widget = ctk.CTkEntry(manual, textvariable=self.individual_vars[key], height=30, width=260)
            widget.grid(row=1 + idx, column=1, padx=(0, 16), pady=(0, 8), sticky="w")
        ctk.CTkButton(manual, text="Agregar al lote", width=160, height=30, command=self.add_manual_record).grid(
            row=5, column=1, padx=(0, 16), pady=(4, 12), sticky="w"
        )

        batch = ctk.CTkFrame(carga_row, corner_radius=8, fg_color="#ffffff", border_width=1, border_color="#d8e2ec")
        batch.grid(row=0, column=1, padx=0, pady=0, sticky="nw")
        batch.grid_columnconfigure((0, 1, 2, 3), weight=0)
        batch.grid_columnconfigure(4, weight=1)
        ctk.CTkLabel(batch, text="Carga por lote", font=ctk.CTkFont(size=16, weight="bold"), text_color="#16324f").grid(
            row=0, column=0, columnspan=4, padx=16, pady=(12, 6), sticky="w"
        )
        ctk.CTkLabel(
            batch,
            text="Pega una lista por columna. La fila 1 de cada columna forma el paciente 1.",
            font=ctk.CTkFont(size=12),
            text_color="#51657a",
        ).grid(row=1, column=0, columnspan=4, padx=16, pady=(0, 8), sticky="w")

        self.batch_boxes: dict[str, ctk.CTkTextbox] = {}
        for index, (title, key) in enumerate(labels):
            ctk.CTkLabel(batch, text=title, font=ctk.CTkFont(size=13, weight="bold"), text_color="#16324f").grid(
                row=2, column=index, padx=(16 if index == 0 else 0, 8), pady=(0, 4), sticky="w"
            )
            textbox = ctk.CTkTextbox(batch, height=140, width=155, font=ctk.CTkFont(size=13))
            textbox.grid(row=3, column=index, padx=(16 if index == 0 else 0, 8), pady=(0, 12), sticky="w")
            self.batch_boxes[key] = textbox

        controls = ctk.CTkFrame(content, corner_radius=8, fg_color="#ffffff", border_width=1, border_color="#d8e2ec")
        controls.grid(row=4, column=0, padx=8, pady=(0, 5), sticky="ew")
        controls.grid_columnconfigure(3, weight=0)
        controls.grid_columnconfigure(9, weight=1)
        self.destination_mode_var = ctk.StringVar(
            value=str(self.sheet_settings.get("destination_mode", DESTINO_LOCAL)) if str(self.sheet_settings.get("destination_mode", DESTINO_LOCAL)) in DESTINO_OPTIONS else DESTINO_LOCAL
        )
        self.folder_var = ctk.StringVar(value=str(self.sheet_settings.get("local_folder", self.destination_dir)))
        ctk.CTkLabel(controls, text="Destino:", font=ctk.CTkFont(size=13, weight="bold"), text_color="#16324f").grid(
            row=0, column=0, padx=(12, 8), pady=10, sticky="w"
        )
        self.destination_mode_combo = ctk.CTkComboBox(
            controls,
            values=DESTINO_OPTIONS,
            variable=self.destination_mode_var,
            width=220,
            height=30,
            state="readonly",
            command=self._on_destination_mode_changed,
        )
        self.destination_mode_combo.grid(row=0, column=1, padx=(0, 8), pady=10, sticky="w")
        ctk.CTkLabel(controls, text="Carpeta destino:", font=ctk.CTkFont(size=13, weight="bold"), text_color="#16324f").grid(
            row=0, column=2, padx=(8, 8), pady=10, sticky="w"
        )
        self.folder_entry = ctk.CTkEntry(controls, textvariable=self.folder_var, height=30)
        self.folder_entry.grid(row=0, column=3, padx=8, pady=10, sticky="w")
        self.folder_button = ctk.CTkButton(controls, text="Elegir carpeta", width=140, height=30, command=self.choose_destination_folder)
        self.folder_button.grid(row=0, column=4, padx=8, pady=10, sticky="w")
        self.process_button = ctk.CTkButton(
            controls, text="Generar credenciales", width=180, height=34, font=ctk.CTkFont(size=15, weight="bold"), command=self.start_processing
        )
        self.process_button.grid(row=0, column=5, padx=8, pady=10, sticky="w")
        self.stop_button = ctk.CTkButton(
            controls, text="Detener", width=120, height=34, fg_color="#8a5a5a", hover_color="#734949", state="disabled", command=self.request_stop
        )
        self.stop_button.grid(row=0, column=6, padx=8, pady=10, sticky="w")
        self.clear_button = ctk.CTkButton(controls, text="Limpiar", width=110, height=34, fg_color="#9aafc3", hover_color="#7f95aa", command=self.clear_panel)
        self.clear_button.grid(row=0, column=7, padx=8, pady=10, sticky="e")
        self.save_report_button = ctk.CTkButton(
            controls, text="Guardar reporte Excel", width=180, height=34, fg_color="#66788a", hover_color="#536577", command=self.save_report
        )
        self.save_report_button.grid(row=0, column=8, padx=(8, 12), pady=10, sticky="e")

        status = ctk.CTkFrame(content, corner_radius=8, fg_color="#ffffff", border_width=1, border_color="#d8e2ec")
        status.grid(row=5, column=0, padx=8, pady=(0, 8), sticky="nsew")
        status.grid_columnconfigure(0, weight=1)
        status.grid_rowconfigure(3, weight=1)
        self.status_label = ctk.CTkLabel(
            status, text="Preparado para descargar credenciales.", font=ctk.CTkFont(size=14, weight="bold"), text_color="#16324f"
        )
        self.status_label.grid(row=0, column=0, padx=14, pady=(10, 6), sticky="w")
        self.progress = ctk.CTkProgressBar(status, height=14)
        self.progress.grid(row=1, column=0, padx=14, pady=(0, 6), sticky="ew")
        self.progress.set(0)
        self.summary_label = ctk.CTkLabel(status, text="Todavia no hay descargas.", font=ctk.CTkFont(size=12), text_color="#66788a")
        self.summary_label.grid(row=2, column=0, padx=14, pady=(0, 8), sticky="w")

        table_container = ctk.CTkFrame(status, corner_radius=10, fg_color="#ffffff")
        table_container.grid(row=3, column=0, padx=14, pady=(0, 10), sticky="nsew")
        table_container.grid_columnconfigure(0, weight=1)
        table_container.grid_rowconfigure(0, weight=1)
        self.tree = ttk.Treeview(
            table_container,
            columns=[column_id for column_id, _ in CREDENCIAL_TABLE_COLUMNS],
            show="headings",
            height=6,
            style="Report.Treeview",
        )
        self.tree.grid(row=0, column=0, sticky="nsew")
        self.tree.bind("<ButtonRelease-1>", self.on_tree_click)
        self.tree.bind("<Double-1>", self.on_tree_double_click)
        self.tree.bind("<Control-c>", self.on_tree_copy_shortcut)
        self.tree.bind("<MouseWheel>", self.on_tree_mousewheel)
        for column_id, title in CREDENCIAL_TABLE_COLUMNS:
            self.tree.heading(column_id, text=title)
            self.tree.column(column_id, width=340 if column_id == "archivo_pdf" else 130, anchor="w", stretch=True)
        scrollbar_y = ttk.Scrollbar(table_container, orient="vertical", command=self.tree.yview, style="Report.Vertical.TScrollbar")
        scrollbar_y.grid(row=0, column=1, sticky="ns")
        self.tree.configure(yscrollcommand=scrollbar_y.set)

        actions = ctk.CTkFrame(status, corner_radius=10, fg_color="#eef3f8")
        actions.grid(row=4, column=0, padx=14, pady=(0, 12), sticky="ew")
        actions.grid_columnconfigure(2, weight=0)
        actions.grid_columnconfigure(4, weight=1)
        self.open_pdf_button = ctk.CTkButton(actions, text="Abrir PDF seleccionado", height=30, command=self.open_selected_pdf)
        self.open_pdf_button.grid(row=0, column=0, padx=(10, 8), pady=8, sticky="w")
        self.copy_path_button = ctk.CTkButton(actions, text="Copiar ruta PDF", height=30, command=self.copy_selected_pdf_path)
        self.copy_path_button.grid(row=0, column=1, padx=8, pady=8, sticky="w")
        self.open_folder_button = ctk.CTkButton(actions, text="Abrir carpeta", height=30, fg_color="#66788a", hover_color="#536577", command=self.open_destination_folder)
        self.open_folder_button.grid(row=0, column=3, padx=(8, 10), pady=8, sticky="w")

    def _go_home(self) -> None:
        if self.processing:
            messagebox.showwarning("Atencion", "Espera a que termine o detene el proceso antes de volver.")
            return
        self.on_back()

    def add_manual_record(self) -> None:
        values = {key: var.get().strip() for key, var in self.individual_vars.items()}
        if not all(values.values()):
            messagebox.showwarning("Atencion", "Completa BENEF, DNI, tramite y sexo antes de agregar.")
            return
        for key, value in values.items():
            box = self.batch_boxes[key]
            current = box.get("1.0", "end").strip()
            text = f"{current}\n{value}" if current else value
            box.delete("1.0", "end")
            box.insert("1.0", text)
        self.individual_vars["benef"].set("")
        self.individual_vars["dni"].set("")
        self.individual_vars["tramite"].set("")
        self.individual_vars["sexo"].set("MASC")
        self.summary_label.configure(text="Fila individual agregada al lote.")

    def _detect_credencial_column(self, headers: list[str], target: str) -> str:
        clean_headers = [str(item or "").strip() for item in headers if str(item or "").strip()]
        for header in clean_headers:
            if header.lower() in CREDENCIAL_HEADERS[target]:
                return header
        raise ValueError(f"No pude identificar la columna '{target}' en el Excel.")

    def _read_credencial_excel(self, path: Path) -> list[dict]:
        workbook = load_workbook(path, read_only=True, data_only=True)
        rows = list(workbook.active.iter_rows(values_only=True))
        if not rows:
            raise ValueError("El archivo Excel esta vacio.")
        headers = [str(value or "").strip() for value in rows[0]]
        indices = {key: headers.index(self._detect_credencial_column(headers, key)) for key in CREDENCIAL_HEADERS}
        registros = []
        for row in rows[1:]:
            registro = {}
            for key, index in indices.items():
                value = row[index] if index < len(row) else None
                registro[key] = str(value).strip() if value is not None else ""
            if any(registro.values()):
                registros.append(registro)
        return registros

    def _load_records_into_boxes(self, registros: list[dict]) -> None:
        for key, box in self.batch_boxes.items():
            box.delete("1.0", "end")
            box.insert("1.0", "\n".join(str(item.get(key, "")) for item in registros))

    def _build_records_from_boxes(self) -> list[dict]:
        columnas = {}
        for key, box in self.batch_boxes.items():
            contenido = box.get("1.0", "end").strip()
            columnas[key] = [line.strip() for line in contenido.splitlines()] if contenido else []
        cantidad = max((len(values) for values in columnas.values()), default=0)
        if cantidad == 0:
            return []
        largos = {key: len(values) for key, values in columnas.items()}
        if len(set(largos.values())) != 1:
            detalle = ", ".join(f"{key.upper()}: {valor}" for key, valor in largos.items())
            raise ValueError(f"Las columnas no tienen la misma cantidad de filas.\n\n{detalle}")
        registros = []
        for index in range(cantidad):
            registro = {key: columnas[key][index] for key in ("benef", "dni", "tramite", "sexo")}
            if not all(registro.values()):
                raise ValueError(f"La fila {index + 1} tiene datos incompletos.")
            registros.append(registro)
        return registros

    def load_excel(self) -> None:
        file_path = filedialog.askopenfilename(title="Seleccionar Excel", filetypes=[("Excel", "*.xlsx")])
        if not file_path:
            return
        try:
            registros = self._read_credencial_excel(Path(file_path))
        except Exception as exc:
            messagebox.showerror("Error", f"No se pudo leer el archivo.\n\n{exc}")
            return
        self._load_records_into_boxes(registros)
        self.status_label.configure(text=f"Archivo cargado: {Path(file_path).name}")
        self.summary_label.configure(text=f"Se cargaron {len(registros)} fila(s) para credenciales.")

    def download_template(self) -> None:
        ruta = filedialog.asksaveasfilename(
            title="Guardar modelo Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile="modelo_credenciales.xlsx",
        )
        if not ruta:
            return
        try:
            guardar_modelo_credencial(ruta)
            messagebox.showinfo("Modelo guardado", f"Modelo guardado en:\n\n{ruta}")
        except Exception as exc:
            messagebox.showerror("Error", f"No se pudo guardar el modelo.\n\n{exc}")

    def choose_destination_folder(self) -> None:
        if self.destination_mode_var.get() != DESTINO_LOCAL:
            messagebox.showinfo("Destino web", "En modo Drive la carpeta web se elige desde el desplegable de destino.")
            return
        folder = filedialog.askdirectory(title="Seleccionar carpeta destino")
        if folder:
            self.folder_var.set(folder)

    def _restart_app(self) -> None:
        if self.on_restart is None:
            messagebox.showerror("Reiniciar app", "No se encontro el reinicio global de la suite.")
            return
        try:
            self.on_restart()
        except Exception as exc:
            messagebox.showerror("Reiniciar app", f"No se pudo reiniciar la app:\n{exc}")

    def _load_sheet_settings(self) -> dict:
        try:
            if not self.sheet_settings_file.exists():
                return {}
            data = json.loads(self.sheet_settings_file.read_text(encoding="utf-8"))
            if not isinstance(data, dict):
                return {}
            return data
        except Exception:
            return {}

    def _default_sheet_profile_values(self) -> dict:
        return {
            "internal_name": "",
            "spreadsheet_url": "",
            "sheet_name": "",
            "start_row": "2",
            "max_rows": "40",
            "end_row": "",
            "limit_mode": LIMIT_MODE_TOPE,
            "only_with_tramite": False,
        }

    def _build_sheet_profiles(self) -> dict[str, dict]:
        profiles = {key: self._default_sheet_profile_values() for key in SHEET_PROFILE_OPTIONS}
        raw_profiles = self.sheet_settings.get("sheet_profiles")
        if isinstance(raw_profiles, dict):
            for key in SHEET_PROFILE_OPTIONS:
                raw_profile = raw_profiles.get(key, {})
                if not isinstance(raw_profile, dict):
                    continue
                profiles[key] = {
                    "internal_name": str(raw_profile.get("internal_name", "")).strip(),
                    "spreadsheet_url": normalize_spreadsheet_url(str(raw_profile.get("spreadsheet_url", "")).strip()),
                    "sheet_name": str(raw_profile.get("sheet_name", "")).strip(),
                    "start_row": str(raw_profile.get("start_row", "2")).strip() or "2",
                    "max_rows": str(raw_profile.get("max_rows", "40")).strip(),
                    "end_row": str(raw_profile.get("end_row", "")).strip(),
                    "limit_mode": str(raw_profile.get("limit_mode", "")).strip(),
                    "only_with_tramite": bool(raw_profile.get("only_with_tramite", False)),
                }
        else:
            profiles[SHEET_PROFILE_DUBE] = {
                "internal_name": str(self.sheet_settings.get("internal_name", "")).strip(),
                "spreadsheet_url": normalize_spreadsheet_url(str(self.sheet_settings.get("spreadsheet_url", "")).strip()),
                "sheet_name": str(self.sheet_settings.get("sheet_name", "")).strip(),
                "start_row": str(self.sheet_settings.get("start_row", "2")).strip() or "2",
                "max_rows": str(self.sheet_settings.get("max_rows", "40")).strip(),
                "end_row": str(self.sheet_settings.get("end_row", "")).strip(),
                "limit_mode": str(self.sheet_settings.get("limit_mode", "")).strip(),
                "only_with_tramite": bool(self.sheet_settings.get("only_with_tramite", False)),
            }
        schefe_profile = profiles.get(SHEET_PROFILE_SCHEFE, self._default_sheet_profile_values())
        dube_profile = profiles.get(SHEET_PROFILE_DUBE, self._default_sheet_profile_values())
        plan_salud_profile = profiles.get(SHEET_PROFILE_PLAN_SALUD_CIMA, self._default_sheet_profile_values())
        destination_mode = str(self.sheet_settings.get("destination_mode", "")).strip()
        active_profile = str(self.sheet_settings.get("active_sheet_profile", "")).strip().upper()
        if (
            not str(schefe_profile.get("spreadsheet_url", "")).strip()
            and str(dube_profile.get("spreadsheet_url", "")).strip()
            and (
                destination_mode == DESTINO_DRIVE_SCHEFE
                or active_profile == SHEET_PROFILE_SCHEFE
            )
        ):
            profiles[SHEET_PROFILE_SCHEFE] = dict(dube_profile)
        if (
            not str(plan_salud_profile.get("spreadsheet_url", "")).strip()
            and str(dube_profile.get("spreadsheet_url", "")).strip()
            and active_profile == SHEET_PROFILE_PLAN_SALUD_CIMA
        ):
            profiles[SHEET_PROFILE_PLAN_SALUD_CIMA] = dict(dube_profile)
        return profiles

    def _build_sheet_url_profiles(self) -> dict[str, dict[str, dict]]:
        profiles: dict[str, dict[str, dict]] = {key: {} for key in SHEET_PROFILE_OPTIONS}
        raw_profiles = self.sheet_settings.get("sheet_url_profiles")
        if not isinstance(raw_profiles, dict):
            return profiles
        for profile_key in SHEET_PROFILE_OPTIONS:
            raw_entries = raw_profiles.get(profile_key, {})
            if not isinstance(raw_entries, dict):
                continue
            normalized_entries: dict[str, dict] = {}
            for url_key, raw_entry in raw_entries.items():
                if not isinstance(raw_entry, dict):
                    continue
                normalized_url = normalize_spreadsheet_url(str(url_key or "").strip())
                if not normalized_url:
                    continue
                normalized_entries[normalized_url] = {
                    "spreadsheet_url": normalized_url,
                    "internal_name": str(raw_entry.get("internal_name", "")).strip(),
                    "sheet_name": str(raw_entry.get("sheet_name", "")).strip(),
                    "start_row": str(raw_entry.get("start_row", "2")).strip() or "2",
                    "max_rows": str(raw_entry.get("max_rows", "40")).strip(),
                    "end_row": str(raw_entry.get("end_row", "")).strip(),
                    "limit_mode": str(raw_entry.get("limit_mode", "")).strip(),
                    "only_with_tramite": bool(raw_entry.get("only_with_tramite", False)),
                }
            profiles[profile_key] = normalized_entries
        return profiles

    def _get_sheet_profile_values(self, profile_key: str) -> dict:
        raw = self.sheet_profiles.get(profile_key, self._default_sheet_profile_values())
        values = {
            "internal_name": str(raw.get("internal_name", "")).strip(),
            "spreadsheet_url": normalize_spreadsheet_url(str(raw.get("spreadsheet_url", "")).strip()),
            "sheet_name": str(raw.get("sheet_name", "")).strip(),
            "start_row": str(raw.get("start_row", "2")).strip() or "2",
            "max_rows": str(raw.get("max_rows", "40")).strip(),
            "end_row": str(raw.get("end_row", "")).strip(),
            "limit_mode": str(raw.get("limit_mode", "")).strip(),
            "only_with_tramite": bool(raw.get("only_with_tramite", False)),
        }
        values["limit_mode"] = normalize_limit_mode(values["limit_mode"], has_end_row=bool(values["end_row"]))
        url_key = values["spreadsheet_url"]
        saved_for_url = self.sheet_url_profiles.get(profile_key, {}).get(url_key, {}) if url_key else {}
        if saved_for_url:
            values.update(
                {
                    "sheet_name": str(saved_for_url.get("sheet_name", values["sheet_name"])).strip(),
                    "internal_name": str(saved_for_url.get("internal_name", values["internal_name"])).strip(),
                    "start_row": str(saved_for_url.get("start_row", values["start_row"])).strip() or values["start_row"],
                    "max_rows": str(saved_for_url.get("max_rows", values["max_rows"])).strip(),
                    "end_row": str(saved_for_url.get("end_row", values["end_row"])).strip(),
                    "limit_mode": str(saved_for_url.get("limit_mode", values["limit_mode"])).strip() or values["limit_mode"],
                    "only_with_tramite": bool(saved_for_url.get("only_with_tramite", values["only_with_tramite"])),
                }
            )
        values["limit_mode"] = normalize_limit_mode(values["limit_mode"], has_end_row=bool(values["end_row"]))
        return values

    def _current_sheet_form_values(self) -> dict:
        return {
            "internal_name": (self.sheet_internal_name_var.get() or "").strip() if hasattr(self, "sheet_internal_name_var") else "",
            "spreadsheet_url": normalize_spreadsheet_url((self.sheet_url_var.get() or "").strip()),
            "sheet_name": (self.sheet_name_var.get() or "").strip(),
            "start_row": (self.sheet_start_row_var.get() or "").strip() or "2",
            "max_rows": (self.sheet_max_rows_var.get() or "").strip(),
            "end_row": (self.sheet_end_row_var.get() or "").strip(),
            "limit_mode": (self.sheet_limit_mode_var.get() or "").strip() or LIMIT_MODE_TOPE,
            "only_with_tramite": bool(self.sheet_only_with_tramite_var.get()) if hasattr(self, "sheet_only_with_tramite_var") else False,
        }

    def _save_url_specific_sheet_values(self, profile_name: str, values: dict) -> None:
        normalized_url = normalize_spreadsheet_url(str(values.get("spreadsheet_url", "")).strip())
        if not normalized_url:
            return
        if profile_name not in self.sheet_url_profiles:
            self.sheet_url_profiles[profile_name] = {}
        self.sheet_url_profiles[profile_name][normalized_url] = {
            "spreadsheet_url": normalized_url,
            "internal_name": str(values.get("internal_name", "")).strip(),
            "sheet_name": str(values.get("sheet_name", "")).strip(),
            "start_row": str(values.get("start_row", "2")).strip() or "2",
            "max_rows": str(values.get("max_rows", "40")).strip(),
            "end_row": str(values.get("end_row", "")).strip(),
            "limit_mode": normalize_limit_mode(str(values.get("limit_mode", LIMIT_MODE_TOPE)).strip(), has_end_row=bool(str(values.get("end_row", "")).strip())),
            "only_with_tramite": bool(values.get("only_with_tramite", False)),
        }

    def _load_url_specific_sheet_values(self, profile_name: str, url_value: str) -> bool:
        normalized_url = normalize_spreadsheet_url(str(url_value or "").strip())
        if not normalized_url:
            return False
        saved = self.sheet_url_profiles.get(profile_name, {}).get(normalized_url)
        if not saved:
            return False
        current_internal_name = self.sheet_internal_name_var.get() if hasattr(self, "sheet_internal_name_var") else ""
        self.sheet_internal_name_var.set(str(saved.get("internal_name", current_internal_name)).strip())
        self.sheet_url_var.set(normalized_url)
        self.sheet_name_var.set(str(saved.get("sheet_name", "")).strip())
        self.sheet_start_row_var.set(str(saved.get("start_row", "2")).strip() or "2")
        self.sheet_max_rows_var.set(str(saved.get("max_rows", "40")).strip())
        self.sheet_end_row_var.set(str(saved.get("end_row", "")).strip())
        self.sheet_limit_mode_var.set(
            normalize_limit_mode(
                str(saved.get("limit_mode", "")).strip(),
                has_end_row=bool(self.sheet_end_row_var.get().strip()),
            )
        )
        self.sheet_only_with_tramite_var.set(bool(saved.get("only_with_tramite", False)))
        self._apply_sheet_limit_mode()
        return True

    def _sheet_template_name(self, item: dict | None) -> str:
        item = item or {}
        internal_name = str(item.get("internal_name", "") or "").strip()
        if internal_name:
            return internal_name
        sheet_name = str(item.get("sheet_name", "") or "").strip()
        if sheet_name:
            return sheet_name
        url = normalize_spreadsheet_url(str(item.get("spreadsheet_url", "") or ""))
        if url:
            spreadsheet_id = extract_spreadsheet_id(url)
            return f"Planilla {spreadsheet_id[-6:]}" if spreadsheet_id else "Planilla configurada"
        return "Sin hoja"

    def _current_sheet_template_name(self) -> str:
        return self._sheet_template_name(self._current_sheet_form_values())

    def _update_sheet_template_display(self) -> None:
        if hasattr(self, "sheet_template_display_var"):
            self.sheet_template_display_var.set(self._current_sheet_template_name())

    def _apply_sheet_template_config(self, *, display_name: str, spreadsheet_url: str, sheet_name: str) -> None:
        normalized_url = normalize_spreadsheet_url(spreadsheet_url)
        if not normalized_url:
            raise RuntimeError("Ingresa una URL de Google Sheets.")
        if is_office_file_url(spreadsheet_url):
            raise RuntimeError(OFFICE_FILE_MESSAGE)
        if not sheet_name.strip():
            raise RuntimeError("Ingresa el nombre de la pestaña.")

        self.sheet_internal_name_var.set(display_name.strip())
        self.sheet_url_var.set(normalized_url)
        self.sheet_name_var.set(sheet_name.strip())
        self.sheet_tabs = [sheet_name.strip()]
        if hasattr(self, "sheet_name_combo"):
            self.sheet_name_combo.configure(values=self.sheet_tabs or [""])
        self._save_sheet_settings()
        self._update_sheet_template_display()

    def _open_sheet_template_dialog(self) -> None:
        if self.processing:
            messagebox.showwarning("Google Sheets", "Espera a que termine el proceso antes de cambiar la hoja.")
            return

        selected = self._current_sheet_form_values()
        dialog = ctk.CTkToplevel(self)
        dialog.title("Configurar hoja de credenciales")
        dialog.geometry("680x245")
        dialog.transient(self.winfo_toplevel())
        dialog.grab_set()
        dialog.grid_columnconfigure(1, weight=1)

        name_var = ctk.StringVar(value=str(selected.get("internal_name", "") or self._current_sheet_template_name()).strip())
        url_var = ctk.StringVar(value=normalize_spreadsheet_url(str(selected.get("spreadsheet_url", "") or self.sheet_url_var.get())))
        sheet_var = ctk.StringVar(value=str(selected.get("sheet_name", "") or self.sheet_name_var.get()).strip())

        ctk.CTkLabel(dialog, text="Nombre interno", text_color="#16324f").grid(row=0, column=0, padx=(14, 8), pady=(16, 8), sticky="w")
        name_entry = ctk.CTkEntry(dialog, textvariable=name_var, height=30)
        name_entry.grid(row=0, column=1, padx=(0, 14), pady=(16, 8), sticky="ew")

        ctk.CTkLabel(dialog, text="URL", text_color="#16324f").grid(row=1, column=0, padx=(14, 8), pady=8, sticky="w")
        url_entry = ctk.CTkEntry(dialog, textvariable=url_var, height=30)
        url_entry.grid(row=1, column=1, padx=(0, 14), pady=8, sticky="ew")

        ctk.CTkLabel(dialog, text="Pestaña", text_color="#16324f").grid(row=2, column=0, padx=(14, 8), pady=8, sticky="w")
        sheet_entry = ctk.CTkEntry(dialog, textvariable=sheet_var, height=30)
        sheet_entry.grid(row=2, column=1, padx=(0, 14), pady=8, sticky="ew")

        hint = ctk.CTkLabel(
            dialog,
            text="El link queda guardado en la app; en el panel solo se muestra este nombre.",
            text_color="#51657a",
            font=ctk.CTkFont(size=11),
        )
        hint.grid(row=3, column=0, columnspan=2, padx=14, pady=(2, 12), sticky="w")

        buttons = ctk.CTkFrame(dialog, fg_color="transparent")
        buttons.grid(row=4, column=0, columnspan=2, padx=14, pady=(0, 14), sticky="e")

        def save() -> None:
            try:
                self._apply_sheet_template_config(
                    display_name=name_var.get(),
                    spreadsheet_url=url_var.get(),
                    sheet_name=sheet_var.get(),
                )
            except Exception as exc:
                messagebox.showerror("Google Sheets", str(exc))
                return
            dialog.destroy()

        ctk.CTkButton(
            buttons,
            text="Cancelar",
            width=92,
            fg_color="#9aafc3",
            hover_color="#7f95aa",
            command=dialog.destroy,
        ).grid(row=0, column=0, padx=(0, 8), sticky="e")
        ctk.CTkButton(buttons, text="Guardar", width=92, command=save).grid(row=0, column=1, sticky="e")
        name_entry.focus_set()

    def _on_sheet_url_focus_out(self, _event=None):
        self._load_url_specific_sheet_values(
            self.current_sheet_profile_key,
            self.sheet_url_var.get(),
        )

    def _on_sheet_url_enter(self, _event=None):
        self._on_sheet_url_focus_out()
        return "break"

    def _on_sheet_profile_changed(self, selected: str | None = None) -> None:
        new_profile = (selected or self.sheet_profile_var.get() or SHEET_PROFILE_DUBE).strip()
        if new_profile not in SHEET_PROFILE_OPTIONS:
            new_profile = SHEET_PROFILE_DUBE
        current_values = self._current_sheet_form_values()
        self.sheet_profiles[self.current_sheet_profile_key] = current_values
        self._save_url_specific_sheet_values(self.current_sheet_profile_key, current_values)
        self.current_sheet_profile_key = new_profile
        profile_values = self._get_sheet_profile_values(new_profile)
        self.sheet_profile_var.set(new_profile)
        self.sheet_internal_name_var.set(profile_values["internal_name"])
        self.sheet_url_var.set(profile_values["spreadsheet_url"])
        self.sheet_name_var.set(profile_values["sheet_name"])
        current_sheet_name = (profile_values["sheet_name"] or "").strip()
        self.sheet_tabs = [current_sheet_name] if current_sheet_name else []
        self.sheet_name_combo.configure(values=self.sheet_tabs or [""])
        self.sheet_start_row_var.set(profile_values["start_row"])
        self.sheet_max_rows_var.set(profile_values["max_rows"])
        self.sheet_end_row_var.set(profile_values["end_row"])
        self.sheet_limit_mode_var.set(profile_values["limit_mode"])
        self.sheet_only_with_tramite_var.set(bool(profile_values.get("only_with_tramite", False)))
        self._apply_sheet_limit_mode()
        target_destination = PROFILE_DESTINATION_MAP.get(new_profile)
        if (
            not self._syncing_profile_destination
            and target_destination
            and hasattr(self, "destination_mode_var")
            and self.destination_mode_var.get() != target_destination
        ):
            self._syncing_profile_destination = True
            try:
                self.destination_mode_var.set(target_destination)
                self._on_destination_mode_changed(target_destination)
            finally:
                self._syncing_profile_destination = False
        self._save_sheet_settings()
        self._update_sheet_template_display()

    def _save_current_sheet_profile(self) -> None:
        profile_name = self.current_sheet_profile_key or SHEET_PROFILE_DUBE
        current_values = self._current_sheet_form_values()
        self.sheet_profiles[profile_name] = current_values
        self._save_url_specific_sheet_values(profile_name, current_values)
        self._save_sheet_settings()
        self._update_sheet_template_display()
        self.summary_label.configure(text=f"Perfil {profile_name} guardado.")
        messagebox.showinfo("Google Sheets", f"Se guardo la configuracion del perfil {profile_name}.")

    def _save_sheet_settings(self) -> None:
        selected_destination = (self.destination_mode_var.get() or DESTINO_LOCAL).strip() if hasattr(self, "destination_mode_var") else DESTINO_LOCAL
        current_local_folder = ""
        if hasattr(self, "folder_var"):
            if selected_destination == DESTINO_LOCAL:
                current_local_folder = (self.folder_var.get() or str(self.destination_dir)).strip()
            else:
                current_local_folder = str(self.sheet_settings.get("local_folder", self.destination_dir)).strip() or str(self.destination_dir)
        active_profile = self.current_sheet_profile_key if hasattr(self, "current_sheet_profile_key") else SHEET_PROFILE_DUBE
        if active_profile not in SHEET_PROFILE_OPTIONS:
            active_profile = SHEET_PROFILE_DUBE
        if hasattr(self, "sheet_url_var"):
            current_values = self._current_sheet_form_values()
            self.sheet_profiles[active_profile] = current_values
            self._save_url_specific_sheet_values(active_profile, current_values)
        active_profile_values = self._get_sheet_profile_values(active_profile)
        payload = {
            "active_sheet_profile": active_profile,
            "sheet_profiles": self.sheet_profiles,
            "sheet_url_profiles": self.sheet_url_profiles,
            "internal_name": active_profile_values["internal_name"],
            "spreadsheet_url": active_profile_values["spreadsheet_url"],
            "sheet_name": active_profile_values["sheet_name"],
            "start_row": active_profile_values["start_row"],
            "max_rows": active_profile_values["max_rows"],
            "end_row": active_profile_values["end_row"],
            "limit_mode": active_profile_values["limit_mode"],
            "only_with_tramite": bool(active_profile_values.get("only_with_tramite", False)),
            "destination_mode": selected_destination,
            "local_folder": current_local_folder or str(self.destination_dir),
        }
        self.sheet_settings = payload
        self.sheet_settings_file.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

    def _drive_enabled(self) -> bool:
        return self.destination_mode_var.get() in DESTINO_DRIVE_FOLDER_NAMES

    def _on_destination_mode_changed(self, _selected: str | None = None) -> None:
        selected = self.destination_mode_var.get() or DESTINO_LOCAL
        current_profile = getattr(self, "current_sheet_profile_key", "")
        target_profile = DESTINATION_PROFILE_MAP.get(selected)
        if selected == DESTINO_DRIVE_CIMA and current_profile in {SHEET_PROFILE_DUBE, SHEET_PROFILE_PLAN_SALUD_CIMA}:
            target_profile = None
        if (
            target_profile
            and current_profile != target_profile
            and not self._syncing_profile_destination
        ):
            self._syncing_profile_destination = True
            try:
                self._on_sheet_profile_changed(target_profile)
            finally:
                self._syncing_profile_destination = False
        if selected == DESTINO_LOCAL:
            local_folder = str(self.sheet_settings.get("local_folder", "")).strip() or str(self.destination_dir)
            self.folder_var.set(local_folder)
            self.folder_entry.configure(state="normal")
            self.folder_button.configure(state="normal")
        else:
            folder_name = DESTINO_DRIVE_FOLDER_NAMES.get(selected, "")
            self.folder_var.set(f"Drive/{folder_name}" if folder_name else "Drive")
            self.folder_entry.configure(state="disabled")
            self.folder_button.configure(state="disabled")
        if hasattr(self, "sheet_settings_file"):
            self._save_sheet_settings()

    def _start_sheets_status_check(self) -> None:
        token_path = get_sheets_token_path()
        if not token_path.exists():
            self.sheets_connected = False
            self.sheets_status_var.set("Google Sheets no conectado")
            return

        def worker() -> None:
            try:
                email = get_connected_google_email(
                    credentials_path=get_gmail_credentials_path(),
                    token_path=token_path,
                    interactive=False,
                )
                if email:
                    self.event_queue.put(("sheets_connected", email))
                else:
                    self.event_queue.put(("sheets_status", "Token Sheets encontrado, pero hay que reconectar."))
            except Exception:
                self.event_queue.put(("sheets_status", "Token Sheets encontrado, pero hay que reconectar."))

        threading.Thread(target=worker, daemon=True).start()

    def _start_drive_status_check(self) -> None:
        token_path = get_drive_token_path()
        if not token_path.exists():
            self.drive_connected = False
            return

        def worker() -> None:
            try:
                email = get_connected_drive_email(
                    credentials_path=get_gmail_credentials_path(),
                    token_path=token_path,
                    interactive=False,
                )
                if email:
                    self.event_queue.put(("drive_connected", email))
            except Exception:
                pass

        threading.Thread(target=worker, daemon=True).start()

    def _connect_sheets_account(self) -> None:
        self._save_sheet_settings()
        self.status_label.configure(text="Conectando Google Sheets...")
        self.summary_label.configure(text=f"Usando credenciales: {get_gmail_credentials_path()}")

        def worker() -> None:
            try:
                email = get_connected_google_email(
                    credentials_path=get_gmail_credentials_path(),
                    token_path=get_sheets_token_path(),
                    interactive=True,
                )
                sheet_url = normalize_spreadsheet_url((self.sheet_url_var.get() or "").strip())
                if sheet_url:
                    try:
                        tabs = self._sheet_tabs(sheet_url)
                        self.event_queue.put(("sheet_tabs_loaded", tabs))
                    except Exception:
                        pass
                self.event_queue.put(("sheets_connected", email or "cuenta Google"))
            except Exception as exc:
                self.event_queue.put(("sheets_error", str(exc)))

        threading.Thread(target=worker, daemon=True).start()

    def _sheet_spreadsheet_url(self) -> str:
        sheet_url = normalize_spreadsheet_url((self.sheet_url_var.get() or "").strip())
        if not sheet_url:
            raise ValueError("Ingresa la URL de Google Sheets.")
        if is_office_file_url(self.sheet_url_var.get() or ""):
            raise ValueError(OFFICE_FILE_MESSAGE)
        self.sheet_url_var.set(sheet_url)
        return sheet_url

    def _sheet_tabs(self, sheet_url: str) -> list[str]:
        tabs = list_spreadsheet_sheet_names(
            spreadsheet_url_or_id=sheet_url,
            credentials_path=get_gmail_credentials_path(),
            token_path=get_sheets_token_path(),
            interactive=False,
        )
        if not tabs:
            raise ValueError("No se encontraron pestañas en el Google Sheet.")
        return tabs

    def _load_sheet_tabs(self) -> None:
        try:
            sheet_url = self._sheet_spreadsheet_url()
            tabs = self._sheet_tabs(sheet_url)
        except Exception as exc:
            messagebox.showwarning("Google Sheets", self._friendly_google_error_message(str(exc)))
            return
        self.event_queue.put(("sheet_tabs_loaded", tabs))
        self.event_queue.put(("sheets_status", f"Pestañas cargadas: {len(tabs)}"))

    def _sheet_start_row(self) -> int:
        raw_value = (self.sheet_start_row_var.get() or "").strip()
        if not raw_value:
            return 2
        try:
            start_row = int(raw_value)
        except ValueError as exc:
            raise ValueError("La fila inicial de Sheets debe ser un numero entero.") from exc
        if start_row < 2:
            raise ValueError("La fila inicial de Sheets debe ser 2 o mayor.")
        return start_row

    def _on_sheet_limit_mode_changed(self, _selected: str | None = None) -> None:
        self._apply_sheet_limit_mode()
        self._save_sheet_settings()

    def _apply_sheet_limit_mode(self) -> None:
        mode = (self.sheet_limit_mode_var.get() or LIMIT_MODE_TOPE).strip()
        if mode not in LIMIT_MODE_OPTIONS:
            mode = LIMIT_MODE_TOPE
            self.sheet_limit_mode_var.set(mode)
        if mode == LIMIT_MODE_TOPE:
            self.sheet_max_rows_entry.configure(state="normal")
            self.sheet_end_row_entry.configure(state="disabled")
            self.sheet_end_row_var.set("")
        else:
            self.sheet_end_row_entry.configure(state="normal")
            self.sheet_max_rows_entry.configure(state="disabled")
            self.sheet_max_rows_var.set("")

    def _sheet_max_rows(self) -> int | None:
        raw_value = (self.sheet_max_rows_var.get() or "").strip()
        if not raw_value:
            if (self.sheet_end_row_var.get() or "").strip():
                return None
            return 40
        try:
            max_rows = int(raw_value)
        except ValueError as exc:
            raise ValueError("El tope de Sheets debe ser un numero entero.") from exc
        if max_rows < 1:
            raise ValueError("El tope de Sheets debe ser 1 o mayor.")
        return max_rows

    def _sheet_end_row(self, start_row: int) -> int | None:
        raw_value = (self.sheet_end_row_var.get() or "").strip()
        if not raw_value:
            return None
        try:
            end_row = int(raw_value)
        except ValueError as exc:
            raise ValueError("La fila final de Sheets debe ser un numero entero.") from exc
        if end_row < start_row:
            raise ValueError("La fila final de Sheets debe ser mayor o igual a la fila inicial.")
        return end_row

    def _repair_sheet_missing_tramites(self) -> dict[str, object]:
        sheet_url = normalize_spreadsheet_url((self.sheet_url_var.get() or "").strip())
        sheet_name = (self.sheet_name_var.get() or "").strip()
        start_row = self._sheet_start_row()
        max_rows = self._sheet_max_rows()
        end_row = self._sheet_end_row(start_row)
        if not sheet_url:
            raise ValueError("Ingresa la URL de Google Sheets.")
        if is_office_file_url(self.sheet_url_var.get() or ""):
            raise ValueError(OFFICE_FILE_MESSAGE)
        if not sheet_name:
            raise ValueError("Ingresa el nombre de la pestaña.")
        self.sheet_url_var.set(sheet_url)
        result = repair_credencial_sheet_missing_tramites(
            spreadsheet_url_or_id=sheet_url,
            sheet_name=sheet_name,
            start_row=start_row,
            max_rows=max_rows,
            end_row=end_row,
            sheet_profile=self.current_sheet_profile_key,
        )
        self.last_recovered_tramite_rows = list(result.get("matches", []) or [])
        return result

    def scan_sheet_missing_tramites(self) -> None:
        try:
            result = self._repair_sheet_missing_tramites()
        except Exception as exc:
            messagebox.showwarning("Google Sheets", self._friendly_google_error_message(str(exc)))
            return
        updated = int(result.get("updated") or 0)
        reviewed = int(result.get("reviewed") or 0)
        unresolved = int(result.get("unresolved") or 0)
        target_end_row = result.get("target_end_row")
        rango = f"filas {self._sheet_start_row()} a {target_end_row}" if target_end_row else f"desde fila {self._sheet_start_row()}"
        self.sheets_status_var.set(
            f"Barrido tramite listo: recuperadas {updated} | sin resolver {unresolved}"
        )
        self.summary_label.configure(
            text=f"Barrido de tramite en {rango}: revisadas {reviewed}, recuperadas {updated}, sin resolver {unresolved}."
        )
        messagebox.showinfo(
            "Barrido de tramite",
            f"Rango revisado: {rango}\n\n"
            f"Filas sin tramite revisadas: {reviewed}\n"
            f"Tramites recuperados: {updated}\n"
            f"Siguen sin tramite: {unresolved}",
        )

    def _build_sheet_records(self) -> list[dict]:
        sheet_url = normalize_spreadsheet_url((self.sheet_url_var.get() or "").strip())
        sheet_name = (self.sheet_name_var.get() or "").strip()
        start_row = self._sheet_start_row()
        max_rows = self._sheet_max_rows()
        end_row = self._sheet_end_row(start_row)
        if not sheet_url:
            raise ValueError("Ingresa la URL de Google Sheets.")
        if is_office_file_url(self.sheet_url_var.get() or ""):
            raise ValueError(OFFICE_FILE_MESSAGE)
        if not sheet_name:
            raise ValueError("Ingresa el nombre de la pestaña.")

        self.sheet_url_var.set(sheet_url)
        self._save_sheet_settings()
        if bool(self.sheet_only_with_tramite_var.get()):
            self.sheets_status_var.set("Modo activo: se ignoran filas sin N° de tramite.")
        else:
            repaired = self._repair_sheet_missing_tramites()
            recovered_count = int(repaired.get("updated") or 0)
            if recovered_count:
                self.sheets_status_var.set(f"Se recuperaron {recovered_count} tramite(s) antes de descargar.")
        payload = read_credencial_sheet_rows(
            spreadsheet_url_or_id=sheet_url,
            sheet_name=sheet_name,
            start_row=start_row,
            max_rows=max_rows,
            end_row=end_row,
            sheet_profile=self.current_sheet_profile_key,
            only_with_tramite=bool(self.sheet_only_with_tramite_var.get()),
        )
        records = payload.get("records", [])
        self.last_missing_benef_rows = payload.get("missing_benef", [])
        self.last_missing_tramite_rows = payload.get("missing_tramite", [])
        if not records:
            if self.last_missing_benef_rows or self.last_missing_tramite_rows:
                partes: list[str] = []
                if self.last_missing_benef_rows:
                    partes.append(f"{len(self.last_missing_benef_rows)} por falta de BENEF")
                if self.last_missing_tramite_rows:
                    partes.append(f"{len(self.last_missing_tramite_rows)} por falta de N° de tramite")
                raise ValueError(
                    f"No hay filas procesables desde la fila {start_row}. "
                    f"Se omitieron {', '.join(partes)}."
                )
            raise ValueError(f"No hay filas pendientes en la hoja para procesar desde la fila {start_row}.")
        return records

    def start_sheet_processing(self) -> None:
        if self.processing:
            return
        try:
            registros = self._build_sheet_records()
        except Exception as exc:
            messagebox.showwarning("Google Sheets", self._friendly_google_error_message(str(exc)))
            return
        self._start_processing_common(registros, from_sheets=True)

    def start_processing(self) -> None:
        if self.processing:
            return
        try:
            registros = self._build_records_from_boxes()
        except Exception as exc:
            messagebox.showwarning("Atencion", str(exc))
            return
        if not registros:
            messagebox.showwarning("Atencion", "No hay registros para procesar.")
            return
        self._start_processing_common(registros, from_sheets=False)

    def _start_processing_common(self, registros: list[dict], *, from_sheets: bool) -> None:
        destino_label = self.destination_mode_var.get() or DESTINO_LOCAL
        drive_mode = destino_label in DESTINO_DRIVE_FOLDER_NAMES
        local_folder = str(self.sheet_settings.get("local_folder", self.destination_dir)).strip() or str(self.destination_dir)
        if drive_mode:
            destino = Path(tempfile.mkdtemp(prefix="suite_pami_credenciales_"))
        elif destino_label == DESTINO_LOCAL:
            local_folder = self.folder_var.get().strip() or local_folder
            destino = Path(local_folder)
            destino.mkdir(parents=True, exist_ok=True)
            self.folder_var.set(str(destino))
        else:
            destino = Path(local_folder)
            destino.mkdir(parents=True, exist_ok=True)
        self._save_sheet_settings()
        self.processing = True
        self.current_destination_mode = destino_label
        self.cancel_requested.clear()
        self.current_results = []
        self.report_path = None
        self.progress.set(0)
        self._clear_table()
        self._set_controls_enabled(False)
        self.status_label.configure(text="Preparando...")
        origen = "Sheets" if from_sheets else "lote manual"
        omitidas_benef = len(self.last_missing_benef_rows) if from_sheets else 0
        omitidas_tramite = len(self.last_missing_tramite_rows) if from_sheets else 0
        resumen = f"Se van a procesar {len(registros)} credencial(es) desde {origen}. Destino: {destino_label}"
        if omitidas_benef:
            resumen += f" | Omitidas sin BENEF: {omitidas_benef}"
        if omitidas_tramite:
            resumen += f" | Omitidas sin tramite: {omitidas_tramite}"
        self.summary_label.configure(text=resumen)
        self.worker_thread = threading.Thread(
            target=self._run_worker,
            args=(registros, destino, from_sheets, destino_label),
            daemon=True,
        )
        self.worker_thread.start()

    def _run_worker(self, registros: list[dict], destino: Path, from_sheets: bool, destination_mode: str) -> None:
        try:
            resultados = procesar_lote_credenciales(
                registros,
                destino,
                progress_callback=self._progress_callback,
                status_callback=self._status_callback,
                should_cancel=self.cancel_requested.is_set,
            )
            drive_uploaded = 0
            drive_error = ""
            if destination_mode in DESTINO_DRIVE_FOLDER_NAMES:
                try:
                    self.event_queue.put(("status", "Subiendo credenciales a Drive..."))
                    folder_name = DESTINO_DRIVE_FOLDER_NAMES[destination_mode]
                    email = get_connected_drive_email(
                        credentials_path=get_gmail_credentials_path(),
                        token_path=get_drive_token_path(),
                        interactive=True,
                    )
                    if email:
                        self.event_queue.put(("drive_connected", email))
                    folder_id = resolve_child_folder_id(
                        DRIVE_PARENT_FOLDER_URL,
                        folder_name,
                        interactive=False,
                    )
                    for result in resultados:
                        if str(result.get("estado", "")).strip().upper() != "DESCARGADA":
                            continue
                        local_pdf = str(result.get("archivo_pdf", "")).strip()
                        if not local_pdf:
                            continue
                        uploaded = upload_file_to_drive_folder(
                            local_pdf,
                            folder_id,
                            target_name=Path(local_pdf).name,
                            interactive=False,
                        )
                        result["archivo_drive_url"] = uploaded.get("webViewLink", "")
                        result["archivo_drive_id"] = uploaded.get("id", "")
                        result["observaciones"] = (
                            f"{result.get('observaciones', '')} | Drive: {uploaded.get('webViewLink', '')}".strip(" |")
                        )
                        drive_uploaded += 1
                except Exception as exc:
                    drive_error = str(exc)
                    self.event_queue.put(("status", "Credenciales descargadas. Fallo la subida a Drive."))
            if destination_mode in DESTINO_DRIVE_FOLDER_NAMES:
                clear_local_pdf_paths_for_drive_results(resultados)
            if from_sheets:
                self.event_queue.put(("status", "Anotando resultados e hipervinculos en Sheets..."))
                for source_row, result_row in zip(registros, resultados):
                    result_row["sheet_row"] = source_row.get("sheet_row")
                skipped_results = [
                    {
                        "sheet_row": item.get("sheet_row"),
                        "estado": "Falta BENEF",
                        "archivo_pdf": "",
                        "observaciones": str(item.get("observaciones", "")).strip() or "Falta BENEF",
                    }
                    for item in self.last_missing_benef_rows
                    if item.get("sheet_row")
                ]
                skipped_results.extend(
                    [
                    {
                        "sheet_row": item.get("sheet_row"),
                        "estado": "FALTA TRAMITE",
                        "archivo_pdf": "",
                        "observaciones": str(item.get("observaciones", "")).strip() or "Falta Tramite",
                    }
                    for item in self.last_missing_tramite_rows
                    if item.get("sheet_row")
                    ]
                )
                sheet_result_rows = [*resultados, *skipped_results]
                updated_count = write_credencial_sheet_results(
                    spreadsheet_url_or_id=normalize_spreadsheet_url((self.sheet_url_var.get() or "").strip()),
                    sheet_name=(self.sheet_name_var.get() or "").strip(),
                    result_rows=sheet_result_rows,
                    sheet_profile=self.current_sheet_profile_key,
                )
                processed_rows = [
                    int(item.get("sheet_row"))
                    for item in sheet_result_rows
                    if str(item.get("sheet_row", "")).strip().isdigit()
                ]
                if processed_rows:
                    self.event_queue.put(("sheet_advance", {"next_row": max(processed_rows) + 1}))
            else:
                updated_count = 0
            self.event_queue.put(
                (
                    "finished",
                    {
                        "cancelled": self.cancel_requested.is_set(),
                        "resultados": resultados,
                        "from_sheets": from_sheets,
                        "sheet_updates": updated_count,
                        "missing_benef_count": len(self.last_missing_benef_rows) if from_sheets else 0,
                        "missing_tramite_count": len(self.last_missing_tramite_rows) if from_sheets else 0,
                        "drive_uploaded": drive_uploaded,
                        "drive_error": drive_error,
                        "destination_mode": destination_mode,
                    },
                )
            )
        except Exception as exc:
            self.event_queue.put(("fatal_error", str(exc)))
        finally:
            if destination_mode in DESTINO_DRIVE_FOLDER_NAMES:
                shutil.rmtree(destino, ignore_errors=True)

    def _progress_callback(self, current: int, total: int, resultado: dict) -> None:
        self.event_queue.put(("progress", {"current": current, "total": total, "resultado": resultado}))

    def _status_callback(self, text: str) -> None:
        normalized = str(text or "").strip().lower()
        if normalized == "finalizado":
            if self.current_destination_mode in DESTINO_DRIVE_FOLDER_NAMES:
                self.event_queue.put(("status", "Descarga temporal completa. Actualizando Drive/Sheets..."))
            else:
                self.event_queue.put(("status", "Descarga local completa. Actualizando Drive/Sheets..."))
            return
        self.event_queue.put(("status", text))

    def _process_ui_queue(self) -> None:
        had_events = False
        try:
            while True:
                event, payload = self.event_queue.get_nowait()
                had_events = True
                if event == "status":
                    self.status_label.configure(text=payload)
                elif event == "progress":
                    current, total, resultado = payload["current"], payload["total"], payload["resultado"]
                    self.progress.set(current / total if total else 0)
                    self.status_label.configure(text=f"Procesando {current} de {total}")
                    self.summary_label.configure(
                        text=f"Ultimo estado: {resultado['estado']}" + (f" | {Path(resultado['archivo_pdf']).name}" if resultado["archivo_pdf"] else "")
                    )
                    self.current_results.append(resultado)
                    self.tree.insert("", "end", values=[resultado.get(column_id, "") for column_id, _ in CREDENCIAL_TABLE_COLUMNS])
                elif event == "finished":
                    self.processing = False
                    self._set_controls_enabled(True)
                    self._refresh_results_table()
                    if not payload["cancelled"] and self.current_results:
                        self.progress.set(1)
                    exitos = sum(1 for item in self.current_results if item.get("estado") == "DESCARGADA")
                    missing_benef_count = int(payload.get("missing_benef_count") or 0)
                    missing_tramite_count = int(payload.get("missing_tramite_count") or 0)
                    self.status_label.configure(text="Proceso detenido" if payload["cancelled"] else "Finalizado")
                    resumen = f"Descargadas {exitos} de {len(self.current_results)} credenciales."
                    if missing_benef_count:
                        resumen += f" Omitidas sin BENEF: {missing_benef_count}."
                    if missing_tramite_count:
                        resumen += f" Omitidas sin tramite: {missing_tramite_count}."
                    self.summary_label.configure(text=resumen)
                    detalle = "El lote se detuvo cuando termino la descarga en curso." if payload["cancelled"] else "Las credenciales se procesaron y quedaron listadas en pantalla."
                    if payload.get("from_sheets") and payload.get("sheet_updates"):
                        detalle += f"\n\nSheets actualizadas: {payload['sheet_updates']}"
                    if missing_benef_count:
                        detalle += (
                            f"\n\nHubo {missing_benef_count} paciente(s) marcados con "
                            "Falta BENEF."
                        )
                    if missing_tramite_count:
                        detalle += (
                            f"\n\nHubo {missing_tramite_count} paciente(s) que no se pudieron completar "
                            "por falta de N° de tramite."
                        )
                    if payload.get("drive_uploaded"):
                        detalle += f"\nDrive subidos: {payload['drive_uploaded']} ({payload.get('destination_mode', '')})"
                    if payload.get("drive_error"):
                        detalle += (
                            "\n\nLa descarga temporal se completo, pero fallo la subida a Drive."
                            f"\nDetalle Drive: {payload['drive_error']}"
                        )
                    destino_final = (
                        payload.get("destination_mode", "")
                        if payload.get("destination_mode") in DESTINO_DRIVE_FOLDER_NAMES
                        else self.folder_var.get()
                    )
                    messagebox.showinfo(
                        self.status_label.cget("text"),
                        f"{detalle}\n\nDestino:\n{destino_final}\n\nLog local:\n{get_log_file()}",
                    )
                elif event == "sheets_connected":
                    self.sheets_connected = True
                    self.sheets_status_var.set(f"Google Sheets conectado: {payload}")
                    self.summary_label.configure(text=f"Google Sheets conectado: {payload}")
                elif event == "drive_connected":
                    self.drive_connected = True
                    self.summary_label.configure(text=f"Google Drive conectado: {payload}")
                elif event == "sheets_status":
                    self.sheets_connected = False
                    self.sheets_status_var.set(str(payload or "Google Sheets no conectado"))
                elif event == "sheets_error":
                    self.sheets_connected = False
                    self.sheets_status_var.set("Google Sheets no conectado")
                    messagebox.showerror("Google Sheets", self._friendly_google_error_message(str(payload)))
                elif event == "sheet_tabs_loaded":
                    tabs = [str(item or "").strip() for item in (payload or []) if str(item or "").strip()]
                    self.sheet_tabs = tabs
                    self.sheet_name_combo.configure(values=tabs or [""])
                    current = (self.sheet_name_var.get() or "").strip()
                    if tabs and current not in tabs:
                        match = next((item for item in tabs if item.lower() == current.lower()), "")
                        self.sheet_name_var.set(match or tabs[0])
                    self._save_sheet_settings()
                elif event == "sheet_advance":
                    next_row = payload.get("next_row")
                    if next_row:
                        self.sheet_start_row_var.set(str(next_row))
                        self._save_sheet_settings()
                        self.sheets_status_var.set(
                            f"Google Sheets listo. Proxima corrida desde fila {next_row}."
                        )
                elif event == "fatal_error":
                    self.processing = False
                    self._set_controls_enabled(True)
                    self.status_label.configure(text="Error")
                    self.summary_label.configure(text="El proceso se detuvo por un error general.")
                    messagebox.showerror("Error", f"Ocurrio un error general.\n\n{self._friendly_google_error_message(str(payload))}\n\nLog: {get_log_file()}")
        except queue.Empty:
            pass
        finally:
            delay = 120 if (self.processing or had_events) else 350
            self.after(delay, self._process_ui_queue)

    def _friendly_google_error_message(self, message: str) -> str:
        raw = str(message or "")
        lowered = raw.lower()
        if "invalid_grant" in lowered or "expired" in lowered or "revoked" in lowered:
            return (
                "La conexion de Google Sheets vencio o fue revocada.\n\n"
                "Presiona Conectar Sheets para autorizar la cuenta nuevamente y despues reintenta."
            )
        if "insufficient authentication scopes" in lowered or "insufficient permission" in lowered:
            return (
                "Google no dio permisos suficientes para esta accion.\n\n"
                "Presiona Conectar Sheets nuevamente y acepta los permisos solicitados."
            )
        if (
            ("connection aborted" in lowered and "permission denied" in lowered)
            or "permissionerror(13" in lowered
            or "callback local" in lowered
        ):
            return (
                "No se pudo completar la autorizacion de Google Sheets.\n\n"
                "El navegador o Windows corto el permiso de conexion local. "
                "Presiona Conectar Sheets nuevamente, acepta los permisos de Google y espera a que la pagina confirme la conexion."
            )
        if "no hay token google sheets valido" in lowered or "token sheets encontrado" in lowered:
            return (
                "Google Sheets no esta conectado correctamente.\n\n"
                "Presiona Conectar Sheets para autorizar la cuenta."
            )
        return raw

    def save_report(self) -> None:
        if not self.current_results:
            messagebox.showwarning("Atencion", "Todavia no hay resultados para guardar.")
            return
        ruta_default = ruta_reporte_credenciales_por_defecto(self.folder_var.get().strip() or self.destination_dir)
        ruta = filedialog.asksaveasfilename(
            title="Guardar reporte Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile=ruta_default.name,
            initialdir=str(ruta_default.parent),
        )
        if not ruta:
            return
        try:
            self.report_path = exportar_reporte_credenciales(self.current_results, ruta)
            messagebox.showinfo("Reporte guardado", f"Reporte guardado en:\n\n{self.report_path}")
        except Exception as exc:
            messagebox.showerror("Error", f"No se pudo guardar el reporte.\n\n{exc}")

    def open_destination_folder(self) -> None:
        if self._drive_enabled():
            messagebox.showinfo(
                "Destino web",
                f"El destino principal actual es web:\n\n{self.folder_var.get()}\n\n"
                "Los PDFs se descargan en una carpeta temporal solo para subirlos y se eliminan al terminar.",
            )
            return
        destino = Path(self.folder_var.get().strip() or self.destination_dir)
        destino.mkdir(parents=True, exist_ok=True)
        os.startfile(str(destino.resolve()))

    def open_selected_pdf(self) -> None:
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning("Atencion", "Selecciona una fila con PDF para abrir.")
            return
        self._open_result_document(selected[0])

    def copy_selected_pdf_path(self) -> None:
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning("Atencion", "Selecciona una fila para copiar la ruta.")
            return
        values = self.tree.item(selected[0], "values")
        archivo_pdf = values[5] if len(values) > 5 else ""
        if not archivo_pdf:
            messagebox.showwarning("Atencion", "La fila seleccionada no tiene ruta de PDF.")
            return
        self.clipboard_clear()
        self.clipboard_append(archivo_pdf)
        self.update()
        messagebox.showinfo("Copiado", "La ruta del PDF fue copiada al portapapeles.")

    def on_tree_click(self, event) -> None:
        region = self.tree.identify_region(event.x, event.y)
        if region != "cell":
            return

        row_id = self.tree.identify_row(event.y)
        column_token = self.tree.identify_column(event.x)
        if not row_id or not column_token:
            return

        column_index = int(column_token.replace("#", "")) - 1
        if column_index < 0 or column_index >= len(CREDENCIAL_TABLE_COLUMNS):
            return

        values = self.tree.item(row_id, "values")
        self.tree.selection_set(row_id)
        self.active_column_id, self.active_column_title = CREDENCIAL_TABLE_COLUMNS[column_index]
        self.active_cell_value = str(values[column_index]) if column_index < len(values) else ""
        if self.active_column_id == "estado" and self.active_cell_value.strip().upper() == "DESCARGADA":
            self._open_result_document(row_id)
            return
        self.summary_label.configure(
            text=f"Celda seleccionada: {self.active_column_title} = {self.active_cell_value or '(vacio)'}"
        )

    def on_tree_double_click(self, event) -> None:
        region = self.tree.identify_region(event.x, event.y)
        if region == "cell":
            self.on_tree_click(event)
            self._copy_active_cell()
            return

        if region == "heading":
            column_token = self.tree.identify_column(event.x)
            if not column_token:
                return
            column_index = int(column_token.replace("#", "")) - 1
            if column_index < 0 or column_index >= len(CREDENCIAL_TABLE_COLUMNS):
                return
            column_id, column_title = CREDENCIAL_TABLE_COLUMNS[column_index]
            column_text = "\n".join(str(result.get(column_id, "")) for result in self.current_results)
            self.clipboard_clear()
            self.clipboard_append(column_text)
            self.update()
            self.summary_label.configure(text=f"Columna copiada: {column_title}")

    def on_tree_copy_shortcut(self, _event=None):
        if self.active_cell_value:
            self._copy_active_cell()
            return "break"
        return None

    def on_tree_mousewheel(self, event):
        self.tree.yview_scroll(int(-1 * (event.delta / 120)), "units")
        return "break"

    def _copy_active_cell(self) -> None:
        if not self.active_column_title:
            return
        self.clipboard_clear()
        self.clipboard_append(self.active_cell_value)
        self.update()
        self.summary_label.configure(
            text=f"Copiado: {self.active_column_title} = {self.active_cell_value or '(vacio)'}"
        )

    def clear_panel(self) -> None:
        if self.processing:
            return
        for box in self.batch_boxes.values():
            box.delete("1.0", "end")
        self.individual_vars["benef"].set("")
        self.individual_vars["dni"].set("")
        self.individual_vars["tramite"].set("")
        self.individual_vars["sexo"].set("MASC")
        self.current_results = []
        self.report_path = None
        self.progress.set(0)
        self.status_label.configure(text="Preparado para descargar credenciales.")
        self.summary_label.configure(text="Panel limpio. Podes cargar nuevos pacientes.")
        self._clear_table()

    def request_stop(self) -> None:
        if not self.processing:
            return
        self.cancel_requested.set()
        self.stop_button.configure(state="disabled")
        self.status_label.configure(text="Deteniendo...")
        self.summary_label.configure(text="Se va a detener al terminar la descarga en curso.")

    def _set_controls_enabled(self, enabled: bool) -> None:
        state = "normal" if enabled else "disabled"
        self.destination_mode_combo.configure(state="readonly" if enabled else "disabled")
        if enabled:
            self._on_destination_mode_changed(self.destination_mode_var.get())
        else:
            self.folder_entry.configure(state="disabled")
            self.folder_button.configure(state="disabled")
        self.load_excel_button.configure(state=state)
        self.template_button.configure(state=state)
        self.process_button.configure(state=state)
        self.sheet_profile_combo.configure(state="readonly" if enabled else "disabled")
        self.sheet_url_entry.configure(state="readonly" if enabled else "disabled")
        if hasattr(self, "sheet_url_config_button"):
            self.sheet_url_config_button.configure(state=state)
        self.sheet_name_combo.configure(state=state)
        self.sheet_tabs_button.configure(state=state)
        self.sheet_start_row_entry.configure(state=state)
        self.sheet_limit_mode_combo.configure(state="readonly" if enabled else "disabled")
        if hasattr(self, "sheet_only_with_tramite_check"):
            self.sheet_only_with_tramite_check.configure(state=state)
        if enabled:
            self._apply_sheet_limit_mode()
        else:
            self.sheet_max_rows_entry.configure(state="disabled")
            self.sheet_end_row_entry.configure(state="disabled")
        self.sheets_connect_button.configure(state=state)
        self.sheets_save_button.configure(state=state)
        self.sheets_scan_button.configure(state=state)
        self.sheets_run_button.configure(state=state)
        self.clear_button.configure(state=state)
        self.save_report_button.configure(state=state)
        self.stop_button.configure(state=("normal" if not enabled else "disabled"))

    def _clear_table(self) -> None:
        for item in self.tree.get_children():
            self.tree.delete(item)

    def _get_result_for_row_id(self, row_id: str) -> dict | None:
        children = list(self.tree.get_children())
        try:
            index = children.index(row_id)
        except ValueError:
            return None
        if index < 0 or index >= len(self.current_results):
            return None
        return self.current_results[index]

    def _open_result_document(self, row_id: str) -> None:
        result = self._get_result_for_row_id(row_id)
        if not result:
            messagebox.showwarning("Atencion", "No pude identificar el archivo de la fila seleccionada.")
            return

        drive_url = str(result.get("archivo_drive_url", "") or "").strip()
        local_pdf = str(result.get("archivo_pdf", "") or "").strip()

        if drive_url:
            webbrowser.open(drive_url)
            return
        if local_pdf and Path(local_pdf).exists():
            os.startfile(str(Path(local_pdf).resolve()))
            return
        if local_pdf.startswith(("http://", "https://")):
            webbrowser.open(local_pdf)
            return

        messagebox.showwarning("Atencion", "La fila seleccionada no tiene un PDF disponible.")

    def _result_display_value(self, result: dict, column_id: str) -> str:
        if column_id == "archivo_pdf":
            return str(result.get("archivo_drive_url") or result.get("archivo_pdf") or "")
        return str(result.get(column_id, "") or "")

    def _refresh_results_table(self) -> None:
        self._clear_table()
        for result in self.current_results:
            self.tree.insert(
                "",
                "end",
                values=[self._result_display_value(result, column_id) for column_id, _ in CREDENCIAL_TABLE_COLUMNS],
            )

