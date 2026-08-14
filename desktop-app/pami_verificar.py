"""
Motor para verificar OMEs activas en pe.pami.org.ar (Panel de Aceptacion).
Lee el Excel de turnos del cliente seleccionado, busca cada paciente PAMI y extrae sus OMEs futuras.
"""

from __future__ import annotations

import re
import json
import queue
import threading
import time
import unicodedata
from datetime import date, datetime, timedelta
from html.parser import HTMLParser
from pathlib import Path
from typing import Callable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from playwright.sync_api import Page, TimeoutError, sync_playwright

from app_logging import log_message
from app_paths import get_output_dir
from app_settings import load_verificar_omes_config
from pami_scraper import configurar_playwright


PAMI_PANEL_URL = "https://pe.pami.org.ar/controllers/efector.php?registros_por_pagina=50"
PAMI_TRANSMISION_URL = "https://pe.pami.org.ar/controllers/transmision.php?registros_por_pagina=50"
CUP_LOGIN_URL = "https://cup.pami.org.ar/controllers/loginController.php?redirect=https://pe.pami.org.ar"
PAMI_REGEX = re.compile(r"\bp[ao]?[qm]?[a]?m[i]?\b")
ESPECIALIDADES_FILENAME = "medico y especialidad.xlsx"
CLIENTE_CIMA = "cima"
CLIENTE_GJS = "gjs"
AUDITORIA_PROGRESS_VERSION = 3

CIMA_COMBOS_VALIDACION_CONFIG = {
    "lesiones": {
        "principales": {"dermatologia", "otorrinolaringologia", "video_rinofibro"},
        "companeras": {"lesiones"},
        "omitir_companera": {"base:lesiones"},
    },
    "video_otorino_lesiones": {
        "componentes": {"video_rinofibro", "otorrinolaringologia", "lesiones"},
        "omitir": {"717132", "base:otorrinolaringologia", "base:lesiones"},
    },
    "flebo_esclerosante": {
        "componentes": {"tratamiento_esclerosante", "flebologia"},
        "omitir": {"487610", "base:flebologia"},
    },
    "reuma_infiltraciones": {
        "componentes": {"reumatologia", "infiltraciones"},
        "omitir": {"base:reumatologia", "121801", "528801"},
    },
    "hema_frotis": {
        "componentes": {"hematologia", "frotis"},
        "omitir": {"820121", "247130"},
    },
    "mamografia_eco_mamaria": {
        "bases": {"mamografia"},
        "omitir": {"180106", "186001", "base:ecografia_mamaria", "base:ecografia_partes_blandas"},
    },
    "espiro_neumo": {
        "bases": {"espirometria"},
        "omitir_relacionada": {"neumonologia"},
    },
}

CLIENTES_VERIFICACION = {
    CLIENTE_CIMA: {
        "codigo": CLIENTE_CIMA,
        "nombre": "CIMA / Ceintramed",
        "excel_label": "Excel de turnos CIMA",
        "sistema_turnos": "CIMA",
        "especialidades_filename": ESPECIALIDADES_FILENAME,
    },
    CLIENTE_GJS: {
        "codigo": CLIENTE_GJS,
        "nombre": "GJS - Grupo Justo Salud",
        "excel_label": "Listado de consultas GJS",
        "sistema_turnos": "GJS",
        "especialidades_filename": None,
    }
}
PROFESIONAL_ALIASES = {
    "aliaga orlando": "aliaga",
    "aliaga gricel": "aliaga",
    "fattorell mabel mapa": "fattorell mabel mapa",
    "fattorell mabel holter": "fatorell mabel holter",
    "fattorell mabel ecg": "fattorell mabel ecg",
    "diaz b patricia mapa": "diaz b patricia mapa",
    "balderrama maria l martes": "balderrama maria lourdes",
    "balderrama maria": "balderrama maria lourdes",
    "savia gerardo": "savia",
    "castell milton": "castel milton",
    "castell milton ruta": "castel milton",
    "ramirez hernan": "ramirez hernan",
    "rx": "rx",
}
PROFESIONAL_ESPECIALIDADES_MANUALES_CIMA = {
    "carrera jose raul": ["Carrera DBT doble (PAMI)"],
    "diaz b patricia mapa": ["MAPA"],
    "fattorell mabel mapa": ["MAPA"],
    "fattorell mabel ecg": ["ECG"],
    "ramirez hernan": ["Gastroenterologia (PAMI)"],
    "ramirez yanina": ["Nutricion (PAMI)"],
    "ramirez walter": ["Ergometria (PAMI)", "Ecoestres (PAMI)"],
    "vedia esquina esther": ["Endocrinologia (PAMI)"],
    "vedia esther": ["Endocrinologia (PAMI)"],
    "rx": ["RX"],
    "silvestre jimena erika": ["Ecografia (PAMI)"],
    "veliz lisandro": ["Estudios urologicos (PAMI)"],
    "veliz lisandro ignacio": ["Estudios urologicos (PAMI)"],
}
PROFESIONAL_ESPECIALIDADES_MANUALES_POR_SISTEMA = {
    "CIMA": PROFESIONAL_ESPECIALIDADES_MANUALES_CIMA,
    "GJS": {},
}
PROFESIONAL_ESPECIALIDADES_MANUALES = PROFESIONAL_ESPECIALIDADES_MANUALES_CIMA
PERALES_CIMA_BASES_ALTERNATIVAS = {
    "cardiologia",
    "gastroenterologia",
    "diabetologia",
    "dbt_nefro_nutri",
    "nefrologia",
    "nutricion",
}
PERALES_CIMA_REQUISITO_ALTERNATIVO = (
    "570129|820113|base:cardiologia|base:gastroenterologia|"
    "base:diabetologia|base:dbt_nefro_nutri|base:nefrologia|base:nutricion"
)
PERALES_CIMA_DESCRIPCION_ALTERNATIVA = (
    "CONSULTA CON ESPECIALISTA EN CARDIOLOGIA / GASTROENTEROLOGIA / "
    "DIABETOLOGIA / NEFROLOGIA / NUTRICION"
)
BASES_COMPATIBLES_EXCLUSIVAS_POR_SISTEMA = {
    "CIMA": {
        "carrera_dbt",
        "dbt_nefro_nutri",
        "dbt_nefro_nutri_1",
        "dbt_nefro_nutri_2",
        "diabetologia_o_gastro",
        "doppler_o_ecoestres",
        "ecocardio_o_vasos",
    },
    "GJS": set(),
}
PRACTICAS_COMPATIBLES_POR_ESPECIALIDAD = {
    "electroencefalografia": {
        "bases": {"electroencefalografia", "neurologia"},
        "codigos": {"690112", "690119", "690109", "691009", "820129"},
    },
    "hematologia": {
        "bases": {"hematologia"},
        "codigos": {"820121", "247130"},
    },
    "papanicolau": {
        "bases": {"papanicolau"},
        "codigos": {"467119", "467116", "467115"},
    },
    "ginecologia": {
        "bases": {"ginecologia", "papanicolau"},
        "codigos": {"820145", "467119", "467116", "467115"},
    },
    "mamografia": {
        "bases": {"mamografia", "mx"},
        "codigos": {"180105", "180106", "186001"},
    },
    "rx": {
        "bases": {"rx", "radiografia"},
        "codigos": set(),
    },
    "flujometria": {
        "bases": {"flujometria"},
        "codigos": {"507315", "360116"},
    },
    "estudio_urodinamico": {
        "bases": {"estudio_urodinamico"},
        "codigos": {"507313", "360111"},
    },
    "estudios_urologicos": {
        "bases": {"flujometria", "estudio_urodinamico"},
        "codigos": {"507315", "360116", "507313", "360111"},
    },
    "lesiones": {
        "bases": {"lesiones", "dermatologia"},
        "codigos": set(),
    },
    "holter": {
        "bases": {"holter"},
        "codigos": {"570121"},
    },
    "mapa": {
        "bases": {"mapa", "presurometria"},
        "codigos": {"570120"},
    },
    "ecg": {
        "bases": {"ecg"},
        "codigos": {"570129"},
    },
    "ecografia": {
        "bases": {"ecografia"},
        "codigos": set(),
    },
    "ecografia_abdominal": {
        "bases": {"ecografia_abdominal"},
        "codigos": {"180112"},
    },
    "ecografia_renal": {
        "bases": {"ecografia_renal"},
        "codigos": {"180116"},
    },
    "ecografia_vesical": {
        "bases": {"ecografia_vesical"},
        "codigos": set(),
    },
    "ecografia_prostatica": {
        "bases": {"ecografia_prostatica"},
        "codigos": set(),
    },
    "ecografia_testicular": {
        "bases": {"ecografia_testicular"},
        "codigos": set(),
    },
    "ecografia_vias_biliares": {
        "bases": {"ecografia_vias_biliares", "ecografia_abdominal"},
        "codigos": set(),
    },
    "ecografia_endocavitaria": {
        "bases": {"ecografia_endocavitaria"},
        "codigos": {"180128"},
    },
    "ecografia_mamaria": {
        "bases": {"ecografia_mamaria"},
        "codigos": {"180106"},
    },
    "ecografia_partes_blandas": {
        "bases": {"ecografia_partes_blandas"},
        "codigos": {"186001"},
    },
    "ecografia_musculoesqueletica": {
        "bases": {"ecografia_musculoesqueletica"},
        "codigos": {"186001"},
    },
    "ergometria": {
        "bases": {"ergometria"},
        "codigos": set(),
    },
    "ecoestres": {
        "bases": {"ecoestres", "ecostress"},
        "codigos": set(),
    },
    "ecodoppler_cardiaco": {
        "bases": {"ecodoppler_cardiaco"},
        "codigos": {"180301"},
    },
    "ecocardio": {
        "bases": {"ecocardio"},
        "codigos": {"180158"},
    },
    "ecocardio_o_vasos": {
        "bases": {"ecocardio", "doppler"},
        "codigos": {"180158", "180607"},
    },
    "doppler_o_ecoestres": {
        "bases": {"doppler", "ecoestres", "ecodoppler_cardiaco"},
        "codigos": {"180301"},
    },
    "diabetologia_o_gastro": {
        "bases": {"diabetologia", "gastroenterologia"},
        "codigos": set(),
    },
    "dbt_nefro_nutri": {
        "bases": {"diabetologia", "nefrologia", "nutricion"},
        "codigos": {"820171", "820155", "820159"},
    },
    "dbt_nefro_nutri_1": {
        "bases": {"diabetologia", "nefrologia", "nutricion"},
        "codigos": {"820171", "820155", "820159"},
    },
    "dbt_nefro_nutri_2": {
        "bases": {"diabetologia", "nefrologia", "nutricion"},
        "codigos": {"820171", "820155", "820159"},
    },
    "espirometria": {
        "bases": {"espirometria", "neumonologia"},
        "codigos": {"687114"},
    },
    "audiometria": {
        "bases": {"audiometria", "fonoaudiologia"},
        "codigos": {"717150"},
    },
    "logoaudiometria": {
        "bases": {"logoaudiometria", "fonoaudiologia"},
        "codigos": {"717151"},
    },
    "timpanometria": {
        "bases": {"timpanometria", "fonoaudiologia"},
        "codigos": {"717156"},
    },
    "acufenometria": {
        "bases": {"acufenometria", "fonoaudiologia"},
        "codigos": {"717157"},
    },
    "fonoaudiologia": {
        "bases": {"fonoaudiologia"},
        "codigos": {"717150", "717151", "717156", "717157"},
    },
    "fono_estudios": {
        "bases": {"fono_estudios", "audiometria", "logoaudiometria", "timpanometria", "acufenometria"},
        "codigos": {"717150", "717151", "717156", "717157"},
    },
    "impedanciometria": {
        "bases": {"impedanciometria"},
        "codigos": {"717155"},
    },
    "lavaje_oido": {
        "bases": {"lavaje_oido", "otorrinolaringologia"},
        "codigos": {"717111", "717116"},
    },
    "video_rinofibro": {
        "bases": {"video_rinofibro", "otorrinolaringologia"},
        "codigos": {"717132"},
    },
    "urologia": {
        "bases": {"urologia"},
        "codigos": {"820167"},
    },
    "tratamiento_esclerosante": {
        "bases": {"tratamiento_esclerosante", "flebologia"},
        "codigos": {"487610"},
    },
    "infiltraciones": {
        "bases": {"infiltraciones", "reumatologia"},
        "codigos": {"121801", "528801"},
    },
    "control_marcapasos": {
        "bases": {"control_marcapasos"},
        "codigos": {"243003"},
    },
}
PRACTICAS_REQUERIDAS_POR_ESPECIALIDAD = {
    "electroencefalografia": {
        "820129": "CONSULTA CON ESPECIALISTA EN NEUROLOGIA",
    },
    "neurologia": {
        "820129": "CONSULTA CON ESPECIALISTA EN NEUROLOGIA",
    },
    "cardiologia": {
        "570129|820113|base:cardiologia": "CONSULTA CON ESPECIALISTA EN CARDIOLOGIA (INCLUYE ELECTROCARDIOGRAMA)",
    },
    "hematologia": {
        "820121": "CONSULTA CON ESPECIALISTA EN HEMATOLOGIA",
        "247130": "EXAMEN DE FROTIS SANGUINEO",
    },
    "papanicolau": {
        "467119": "CEPILLADO ENDOCERVICAL DE EPITELIO VAGINAL Y CERVICAL PARA DETECCION HPV, CA IN SITU",
        "467116": "COLPOCITOLOGIA CON OBTENCION DE MATERIAL POR CITOLOGIA CERVICOVAGINAL",
        "467115": "COLPOSCOPIA. VULVOSCOPIA. VAGINOSCOPIA",
        "820145": "CONSULTA CON ESPECIALISTA EN GINECOLOGIA (INCLUYE EXAMEN GINECOLOGICO GENERAL Y EXAMEN MAMARIO)",
    },
    "ginecologia": {
        "820145": "CONSULTA CON ESPECIALISTA EN GINECOLOGIA (INCLUYE EXAMEN GINECOLOGICO GENERAL Y EXAMEN MAMARIO)",
    },
    "mamografia": {
        "base:mamografia_bilateral": "MAMOGRAFIA BILATERAL",
        "180106": "ECOGRAFIA MAMARIA BILATERAL",
        "186001": "ECOGRAFIA DE PARTES BLANDAS",
    },
    "mapa": {
        "570120": "PRESUROMETRIA (POR 24HS)",
    },
    "holter": {
        "570121": "HOLTER CARDIACO DE 3 CANALES 24 HS.",
    },
    "ecg": {
        "570129|820113|base:cardiologia": "CONSULTA CON ESPECIALISTA EN CARDIOLOGIA (INCLUYE ELECTROCARDIOGRAMA)",
    },
    "ecografia": {
        "base:ecografia": "ECOGRAFIA",
    },
    "ecografia_abdominal": {
        "base:ecografia_abdominal": "ECOGRAFIA ABDOMINAL COMPLETA",
    },
    "ecografia_renal": {
        "base:ecografia_renal": "ECOGRAFIA RENAL BILATERAL",
    },
    "ecografia_vesical": {
        "base:ecografia_vesical": "ECOGRAFIA VESICAL CON MEDICION DE RESIDUO POST MICCIONAL",
    },
    "ecografia_prostatica": {
        "base:ecografia_prostatica": "ECOGRAFIA PROSTATICA",
    },
    "ecografia_testicular": {
        "base:ecografia_testicular": "ECOGRAFIA TESTICULAR",
    },
    "ecografia_vias_biliares": {
        "base:ecografia_vias_biliares": "ECOGRAFIA DE VIAS BILIARES",
    },
    "ecografia_endocavitaria": {
        "base:ecografia_endocavitaria": "ECOGRAFIA ENDOCAVITARIA GINECOLOGICA",
    },
    "ecografia_mamaria": {
        "base:ecografia_mamaria": "ECOGRAFIA MAMARIA BILATERAL",
    },
    "ecografia_partes_blandas": {
        "base:ecografia_partes_blandas": "ECOGRAFIA DE PARTES BLANDAS",
    },
    "ecografia_musculoesqueletica": {
        "base:ecografia_musculoesqueletica": "ECOGRAFIA MUSCULOESQUELETICA",
    },
    "ergometria": {
        "base:ergometria": "ERGOMETRIA",
    },
    "ecoestres": {
        "base:ecoestres": "ECOCARDIOGRAMA DE STRESS",
    },
    "ecodoppler_cardiaco": {
        "base:ecodoppler_cardiaco": "ECODOPPLER CARDIACO",
    },
    "ecocardio": {
        "base:ecocardio": "ECOCARDIOGRAMA",
    },
    "ecocardio_o_vasos": {
        "base:ecocardio|180607|base:doppler": "ECOCARDIOGRAMA / ECODOPPLER DE VASOS DEL CUELLO",
    },
    "espirometria": {
        "687114": "ESPIROMETRIA (INCLUYE CURVA DE FLUJO VOLUMEN/USO DE BRONCODILATADORES)",
    },
    "doppler": {
        "base:doppler": "ECODOPPLER",
    },
    "doppler_o_ecoestres": {
        "base:doppler|base:ecoestres|base:ecodoppler_cardiaco|180301": "ECODOPPLER / ECOCARDIOGRAMA DE STRESS",
    },
    "diabetologia_o_gastro": {
        "base:diabetologia|base:gastroenterologia": "CONSULTA CON ESPECIALISTA EN DIABETOLOGIA / GASTROENTEROLOGIA",
    },
    "neumonologia": {
        "base:neumonologia": "CONSULTA CON ESPECIALISTA EN NEUMONOLOGIA",
    },
    "reumatologia": {
        "base:reumatologia": "CONSULTA CON ESPECIALISTA EN REUMATOLOGIA",
    },
    "nutricion": {
        "base:nutricion": "CONSULTA CON ESPECIALISTA EN NUTRICION",
    },
    "diabetologia": {
        "base:dbt_nefro_nutri": "CONSULTA CON ESPECIALISTA EN DIABETOLOGIA / NEFROLOGIA / NUTRICION",
    },
    "endocrinologia": {
        "base:endocrinologia": "CONSULTA CON ESPECIALISTA EN ENDOCRINOLOGIA",
    },
    "gastroenterologia": {
        "base:gastroenterologia": "CONSULTA CON ESPECIALISTA EN GASTROENTEROLOGIA",
    },
    "otorrinolaringologia": {
        "base:otorrinolaringologia": "CONSULTA CON ESPECIALISTA EN OTORRINOLARINGOLOGIA",
    },
    "audiometria": {
        "717150": "AUDIOMETRIA",
        "717151": "LOGOAUDIOMETRIA",
        "717156": "TIMPANOMETRIA",
        "717157": "ACUFENOMETRIA",
        "820137": "CONSULTA CON ESPECIALISTA EN FONOAUDIOLOGIA",
    },
    "logoaudiometria": {
        "717151": "LOGOAUDIOMETRIA",
    },
    "timpanometria": {
        "717156": "TIMPANOMETRIA",
    },
    "acufenometria": {
        "717157": "ACUFENOMETRIA",
    },
    "fonoaudiologia": {
        "base:fonoaudiologia": "CONSULTA CON ESPECIALISTA EN FONOAUDIOLOGIA",
    },
    "fono_estudios": {
        "717150": "AUDIOMETRIA",
        "717151": "LOGOAUDIOMETRIA",
        "717156": "TIMPANOMETRIA",
        "717157": "ACUFENOMETRIA",
    },
    "impedanciometria": {
        "717155": "IMPEDANCIOMETRIA",
    },
    "lavaje_oido": {
        "820168": "CONSULTA CON ESPECIALISTA EN OTORRINOLARINGOLOGIA",
        "717111|717116": "LAVAJE DE OIDO / LAVAJE DE SENO PARANASAL",
    },
    "video_rinofibro": {
        "717132": "VIDEO RINOFIBROLARINGOSCOPIA",
    },
    "traumatologia": {
        "base:traumatologia": "CONSULTA CON ESPECIALISTA EN TRAUMATOLOGIA Y ORTOPEDIA",
    },
    "dermatologia": {
        "base:dermatologia": "CONSULTA CON ESPECIALISTA EN DERMATOLOGIA",
    },
    "flebologia": {
        "base:flebologia": "CONSULTA CON ESPECIALISTA EN FLEBOLOGIA",
    },
    "nefrologia": {
        "base:nefrologia": "CONSULTA CON ESPECIALISTA EN NEFROLOGIA",
    },
    "carrera_dbt": {
        "base:dbt_nefro_nutri_1": "PRIMERA CONSULTA DBT / NEFRO / NUTRICION",
        "base:dbt_nefro_nutri_2": "SEGUNDA CONSULTA DBT / NEFRO / NUTRICION",
    },
    "urologia": {
        "820167": "CONSULTA CON ESPECIALISTA EN UROLOGIA",
    },
    "flujometria": {
        "507315|360116|base:flujometria": "FLUJOMETRIA URINARIA COMPUTARIZADA",
    },
    "estudio_urodinamico": {
        "507313|360111|base:estudio_urodinamico": "ESTUDIO URODINAMICO COMPLETO",
    },
    "estudios_urologicos": {
        "507315|360116|507313|360111|base:flujometria|base:estudio_urodinamico": (
            "FLUJOMETRIA URINARIA COMPUTARIZADA / ESTUDIO URODINAMICO COMPLETO"
        ),
    },
    "rx": {
        "base:rx": "RADIOGRAFIA",
    },
    "tratamiento_esclerosante": {
        "487610": "TRATAMIENTO ESCLEROSANTE EN VARICES DE MIEMBROS INFERIORES",
    },
    "infiltraciones": {
        "121801|528801": "INFILTRACIONES MUSCULARES, TENDINOSAS Y ARTICULARES",
    },
    "control_marcapasos": {
        "243003": "CONTROL Y/O REPROGRAMACION DE MARCAPASO",
    },
    "lesiones": {
        "base:lesiones": "LESIONES",
    },
}
PRACTICAS_CONDICIONALES_POR_ESPECIALIDAD = {
    "electroencefalografia": {
        "690112": "HOLTER ELECTROENCEFALOGRAFICO",
        "690119": "ELECTROENCEFALOGRAMA",
        "690109|691009": "ELECTRODIAGNOSTICO",
    },
    "ginecologia": {
        "467119": "CEPILLADO ENDOCERVICAL DE EPITELIO VAGINAL Y CERVICAL PARA DETECCION HPV, CA IN SITU",
        "467116": "COLPOCITOLOGIA CON OBTENCION DE MATERIAL POR CITOLOGIA CERVICOVAGINAL",
        "467115": "COLPOSCOPIA. VULVOSCOPIA. VAGINOSCOPIA",
    },
}

BASES_TURNOS_CON_SUSTITUCION_ESPECIALISTA = {
    "ginecologia",
    "cardiologia",
    "neurologia",
    "traumatologia",
    "diabetologia",
    "dbt_nefro_nutri",
    "diabetologia_o_gastro",
}

BASES_OMES_SUSTITUTAS_ESPECIALISTA = {
    "nutricion",
    "nefrologia",
    "flebologia",
    "cardiologia",
    "gastroenterologia",
    "endocrinologia",
    "otorrinolaringologia",
    "dermatologia",
    "neurologia",
    "traumatologia",
    "diabetologia",
    "ginecologia",
    "neumonologia",
}
PRACTICAS_REQUERIDAS_POR_PROFESIONAL = {}
_ESPECIALIDADES_CACHE: dict[str, dict[str, list[str]]] = {}


def _normalizar_texto(valor) -> str:
    texto = str(valor or "").strip().lower()
    texto = unicodedata.normalize("NFD", texto)
    return "".join(ch for ch in texto if unicodedata.category(ch) != "Mn")


def _clave_texto(valor) -> str:
    texto = _normalizar_texto(valor)
    texto = re.sub(r"\([^)]*\)", " ", texto)
    texto = re.sub(r"[^a-z0-9]+", " ", texto)
    return re.sub(r"\s+", " ", texto).strip()


def _texto_busqueda(valor) -> str:
    texto = _normalizar_texto(valor)
    texto = re.sub(r"[^a-z0-9]+", " ", texto)
    return re.sub(r"\s+", " ", texto).strip()


def _solo_digitos(valor) -> str:
    return re.sub(r"\D", "", str(valor or ""))


def _es_cobertura_pami(valor) -> bool:
    texto = _normalizar_texto(valor)
    return bool(PAMI_REGEX.search(texto)) and not re.fullmatch(r"p", texto)


def get_clientes_verificacion() -> list[dict]:
    return [dict(cliente) for cliente in CLIENTES_VERIFICACION.values()]


def _cliente_config(cliente_codigo: str | None) -> dict:
    codigo = (cliente_codigo or CLIENTE_CIMA).strip().lower()
    return CLIENTES_VERIFICACION.get(codigo, CLIENTES_VERIFICACION[CLIENTE_CIMA])


def _sistema_turnos_key(sistema_turnos: str | None) -> str:
    return str(sistema_turnos or "CIMA").strip().upper() or "CIMA"


def _texto_identificador_excel(valor) -> str:
    if valor is None:
        return ""
    if isinstance(valor, float) and valor.is_integer():
        return str(int(valor))
    return str(valor).strip()


def _es_sistema_cima(sistema_turnos: str | None) -> bool:
    return _sistema_turnos_key(sistema_turnos) == "CIMA"


def _titulo_identificador_reporte(sistema_turnos: str | None) -> str:
    return "BENEF/DNI"


def _identificador_reporte(sistema_turnos: str | None, identificador: str, turno: dict | None) -> str:
    return identificador


def _tramite_reporte(sistema_turnos: str | None, turno: dict | None) -> str:
    if not _es_sistema_cima(sistema_turnos):
        return ""
    return _texto_identificador_excel((turno or {}).get("numero_tramite"))


def _columnas_reporte(sistema_turnos: str | None) -> dict[str, int]:
    if _es_sistema_cima(sistema_turnos):
        return {
            "paciente": 1,
            "identificador": 2,
            "tramite": 3,
            "fecha_turno": 4,
            "profesional": 5,
            "fecha_pami": 6,
            "transmitida": 7,
            "practica": 8,
            "total": 8,
        }
    return {
        "paciente": 1,
        "identificador": 2,
        "tramite": 0,
        "fecha_turno": 3,
        "profesional": 4,
        "fecha_pami": 5,
        "transmitida": 6,
        "practica": 7,
        "total": 7,
    }


def _formatear_fecha(valor) -> str:
    if isinstance(valor, datetime):
        return valor.strftime("%d/%m/%Y")
    if isinstance(valor, date):
        return valor.strftime("%d/%m/%Y")

    texto = str(valor or "").strip()
    if not texto:
        return ""

    formatos = ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d/%m/%y", "%d-%m-%y")
    for formato in formatos:
        try:
            return datetime.strptime(texto.split()[0], formato).strftime("%d/%m/%Y")
        except Exception:
            continue
    return texto.split()[0]


def _fecha_key(valor) -> str:
    fecha = _formatear_fecha(valor)
    for formato in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(fecha, formato).strftime("%Y-%m-%d")
        except Exception:
            continue
    return fecha


def _mismo_mes(fecha_a: str, fecha_b: str) -> bool:
    key_a = _fecha_key(fecha_a)
    key_b = _fecha_key(fecha_b)
    return len(key_a) >= 7 and len(key_b) >= 7 and key_a[:7] == key_b[:7]


def _ome_en_ventana_de_turnos(ome: dict, turnos: list[dict], dias_previos: int = 15) -> bool:
    fecha_ome = _parse_fecha(ome.get("f_agenda", ""))
    fechas_turno = [_parse_fecha(turno.get("fecha")) for turno in turnos]
    fechas_turno = [fecha for fecha in fechas_turno if fecha is not None]
    if fecha_ome is None or not fechas_turno:
        return True
    fecha_min = min(fechas_turno)
    fecha_max = max(fechas_turno)
    return fecha_min - timedelta(days=dias_previos) <= fecha_ome <= fecha_max


def _ome_en_mes_de_algun_turno(ome: dict, turnos: list[dict]) -> bool:
    fecha_ome = _fecha_key(ome.get("f_agenda", ""))
    if not fecha_ome:
        return True
    return any(_mismo_mes(fecha_ome, turno.get("fecha", "")) for turno in turnos)


def _ome_modificable_gjs_para_turno(ome: dict, turno: dict) -> bool:
    if not isinstance(ome, dict) or not ome.get("mismo_efector"):
        return False
    if str(ome.get("estado_accion", "") or "").strip() != "activada_modificable":
        return False
    fecha_turno = _parse_fecha((turno or {}).get("fecha"))
    fecha_vencimiento = _parse_fecha(ome.get("f_vencimiento", ""))
    if fecha_turno is None:
        return True
    return fecha_vencimiento is None or fecha_turno <= fecha_vencimiento


def _parse_fecha(valor) -> date | None:
    fecha = _formatear_fecha(valor)
    if not fecha:
        return None
    for formato in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(fecha, formato).date()
        except Exception:
            continue
    return None


def _formatear_hora(valor) -> str:
    if isinstance(valor, datetime):
        return valor.strftime("%H:%M")
    if hasattr(valor, "strftime"):
        try:
            return valor.strftime("%H:%M")
        except Exception:
            pass
    return str(valor or "").strip()


class _HtmlTableParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.tables: list[list[list[str]]] = []
        self._in_table = False
        self._in_row = False
        self._in_cell = False
        self._depth = 0
        self._rows: list[list[str]] = []
        self._row: list[str] = []
        self._cell: list[str] = []

    def handle_starttag(self, tag: str, attrs) -> None:
        if tag == "table":
            if not self._in_table:
                self._in_table = True
                self._depth = 1
                self._rows = []
            else:
                self._depth += 1
        elif self._in_table and tag == "tr":
            self._in_row = True
            self._row = []
        elif self._in_row and tag in ("td", "th"):
            self._in_cell = True
            self._cell = []
        elif self._in_cell and tag == "br":
            self._cell.append(" ")

    def handle_endtag(self, tag: str) -> None:
        if self._in_cell and tag in ("td", "th"):
            value = " ".join("".join(self._cell).strip().split())
            self._row.append(value)
            self._in_cell = False
        elif self._in_table and tag == "tr":
            if self._row:
                self._rows.append(self._row)
            self._in_row = False
        elif self._in_table and tag == "table":
            self._depth -= 1
            if self._depth == 0:
                self.tables.append(self._rows)
                self._in_table = False

    def handle_data(self, data: str) -> None:
        if self._in_cell:
            self._cell.append(data)


def _leer_tabla_html_xls(ruta_excel: Path) -> list[list[str]]:
    raw = Path(ruta_excel).read_bytes()
    for encoding in ("utf-8", "latin-1"):
        try:
            html = raw.decode(encoding)
            break
        except Exception:
            html = raw.decode("latin-1", errors="replace")
    parser = _HtmlTableParser()
    parser.feed(html)
    return parser.tables[0] if parser.tables else []


def _leer_tabla_xlsx(ruta_excel: Path) -> list[list[str]]:
    wb = load_workbook(ruta_excel, data_only=True, read_only=True)
    ws = wb.active
    rows: list[list[str]] = []
    for row in ws.iter_rows(values_only=True):
        values = ["" if value is None else str(value).strip() for value in row]
        if any(values):
            rows.append(values)
    return rows


def _leer_tabla_excel(ruta_excel: Path) -> list[list[str]]:
    suffix = Path(ruta_excel).suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        return _leer_tabla_xlsx(ruta_excel)
    return _leer_tabla_html_xls(ruta_excel)


def _cell(row: list[str], headers: dict[str, int], name: str) -> str:
    idx = headers.get(name, -1)
    return row[idx] if idx >= 0 and idx < len(row) else ""


def _excluir_turno_gjs(profesional: str, especialidad: str, practica: str, tipo_atencion: str) -> bool:
    texto = _clave_texto(" ".join(str(item or "") for item in (profesional, especialidad, practica, tipo_atencion)))
    return (
        "arias camacho mirko daniel" in texto
        and "medicos de cabecera pami" in texto
        and "consulta" in texto
    )


def _es_turno_clinica_medica_generico(*campos: object) -> bool:
    texto = _clave_texto(" ".join(str(campo or "") for campo in campos))
    return texto in {"clinica medica", "clinica medica pami"}


def leer_pacientes_pami(
    ruta_excel: Path,
    fecha_desde: str | None = None,
    fecha_hasta: str | None = None,
    cliente_codigo: str | None = CLIENTE_CIMA,
) -> list[dict]:
    cliente = _cliente_config(cliente_codigo)
    if cliente["codigo"] == CLIENTE_GJS:
        return leer_pacientes_gjs(ruta_excel, fecha_desde=fecha_desde, fecha_hasta=fecha_hasta)
    if cliente["codigo"] != CLIENTE_CIMA:
        raise NotImplementedError(f"El lector de Excel para {cliente['nombre']} todavia no esta configurado.")
    wb = load_workbook(ruta_excel, data_only=True)
    ws = wb.active
    primera_fila = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
    headers = {_clave_texto(value): idx for idx, value in enumerate(primera_fila)}
    if {"fecha turno", "nombre paciente", "dni paciente"} <= set(headers):
        return _leer_pacientes_cima_por_encabezados(ws, headers, fecha_desde, fecha_hasta)

    agrupados: dict[str, dict] = {}
    desde = _parse_fecha(fecha_desde) if fecha_desde else None
    hasta = _parse_fecha(fecha_hasta) if fecha_hasta else None

    for row in ws.iter_rows(min_row=2, values_only=True):
        cobertura = row[14] if len(row) > 14 else ""
        if not _es_cobertura_pami(cobertura):
            continue

        fecha = row[2] if len(row) > 2 else ""
        fecha_turno = _parse_fecha(fecha)
        if desde and (fecha_turno is None or fecha_turno < desde):
            continue
        if hasta and (fecha_turno is None or fecha_turno > hasta):
            continue

        hora = row[3] if len(row) > 3 else ""
        profesional = row[4] if len(row) > 4 else ""
        if _es_turno_clinica_medica_generico(profesional):
            continue
        nombre = row[5] if len(row) > 5 else ""
        dni = _solo_digitos(row[6] if len(row) > 6 else "")
        beneficio_raw = row[15] if len(row) > 15 else ""
        beneficio_digits = _solo_digitos(beneficio_raw)
        beneficiario = beneficio_digits[:14] if len(beneficio_digits) >= 14 else ""

        if not beneficiario and not dni:
            continue

        key = beneficiario or f"DNI_{dni}"
        if key not in agrupados:
            agrupados[key] = {
                "key": key,
                "nombre": str(nombre or "").strip(),
                "dni": dni,
                "beneficiario": beneficiario,
                "usa_dni": not bool(beneficiario),
                "turnos": [],
            }

        agrupados[key]["turnos"].append(
            {
                "fecha": _formatear_fecha(fecha),
                "fecha_key": _fecha_key(fecha),
                "hora": _formatear_hora(hora),
                "profesional": str(profesional or "").strip(),
                "numero_tramite": "",
            }
        )

    return [paciente for paciente in agrupados.values() if paciente.get("turnos")]


def _leer_pacientes_cima_por_encabezados(
    ws,
    headers: dict[str, int],
    fecha_desde: str | None = None,
    fecha_hasta: str | None = None,
) -> list[dict]:
    agrupados: dict[str, dict] = {}
    desde = _parse_fecha(fecha_desde) if fecha_desde else None
    hasta = _parse_fecha(fecha_hasta) if fecha_hasta else None
    tiene_cobertura = any(key in headers for key in ("cobertura", "obra social"))

    for row in ws.iter_rows(min_row=2, values_only=True):
        cobertura = _cell(row, headers, "cobertura") or _cell(row, headers, "obra social")
        if tiene_cobertura and not _es_cobertura_pami(cobertura):
            continue

        fecha = _cell(row, headers, "fecha turno") or _cell(row, headers, "fecha")
        fecha_turno = _parse_fecha(fecha)
        if desde and (fecha_turno is None or fecha_turno < desde):
            continue
        if hasta and (fecha_turno is None or fecha_turno > hasta):
            continue

        nombre = _cell(row, headers, "nombre paciente")
        dni = _solo_digitos(_cell(row, headers, "dni paciente") or _cell(row, headers, "documento"))
        beneficio_digits = _solo_digitos(
            _cell(row, headers, "numero de beneficiario")
            or _cell(row, headers, "numero afiliado")
            or _cell(row, headers, "nro afiliado")
        )
        beneficiario = beneficio_digits[:14] if len(beneficio_digits) >= 14 else ""

        if not beneficiario and not dni:
            continue

        profesional = (
            _cell(row, headers, "nombre profesional")
            or _cell(row, headers, "profesional")
            or _cell(row, headers, "especialista")
        )
        practica = (
            _cell(row, headers, "especialidad o practica")
            or _cell(row, headers, "practica")
            or _cell(row, headers, "especialidad")
        )
        numero_tramite = _texto_identificador_excel(
            _cell(row, headers, "numero tramite")
            or _cell(row, headers, "numero de tramite")
            or _cell(row, headers, "nro tramite")
            or _cell(row, headers, "n tramite")
            or _cell(row, headers, "tramite")
        )
        if not tiene_cobertura and "plan de salud" in _normalizar_texto(practica):
            continue
        detalle = _cell(row, headers, "detalle estudios realizados")
        detalles_turno = _split_detalles_cima(detalle) or [""]

        key = beneficiario or f"DNI_{dni}"
        if key not in agrupados:
            agrupados[key] = {
                "key": key,
                "nombre": str(nombre or "").strip(),
                "dni": dni,
                "beneficiario": beneficiario,
                "usa_dni": not bool(beneficiario),
                "turnos": [],
            }

        for detalle_turno in detalles_turno:
            profesional_reporte = " - ".join(item for item in (profesional, practica, detalle_turno) if item)
            if _es_turno_clinica_medica_generico(profesional_reporte):
                continue
            agrupados[key]["turnos"].append(
                {
                    "fecha": _formatear_fecha(fecha),
                    "fecha_key": _fecha_key(fecha),
                    "hora": "",
                    "profesional": profesional_reporte,
                    "numero_tramite": numero_tramite,
                }
            )

    return [paciente for paciente in agrupados.values() if paciente.get("turnos")]


def _split_detalles_cima(detalle: str) -> list[str]:
    texto = str(detalle or "").strip()
    if not texto:
        return []
    partes = [parte.strip() for parte in re.split(r"\s*,\s*(?=\d+\s*x\s+)", texto, flags=re.IGNORECASE) if parte.strip()]
    resultado = []
    for parte in partes or [texto]:
        limpio = re.sub(r"^\d+\s*x\s+", "", parte, flags=re.IGNORECASE).strip()
        if limpio:
            resultado.append(limpio)
    return resultado


def leer_pacientes_gjs(ruta_excel: Path, fecha_desde: str | None = None, fecha_hasta: str | None = None) -> list[dict]:
    rows = _leer_tabla_excel(ruta_excel)
    if not rows:
        raise RuntimeError("No se encontro una tabla valida en el archivo GJS.")

    headers = {_clave_texto(value): idx for idx, value in enumerate(rows[0])}
    if {"nro orden", "nro beneficio gp", "apellido y nombre", "practica", "turno"} <= set(headers):
        return _leer_pacientes_gjs_desde_transmision(rows, headers, fecha_desde, fecha_hasta)

    agrupados: dict[str, dict] = {}
    desde = _parse_fecha(fecha_desde) if fecha_desde else None
    hasta = _parse_fecha(fecha_hasta) if fecha_hasta else None
    tiene_cobertura = "obra social" in headers

    for row in rows[1:]:
        cobertura = _cell(row, headers, "obra social")
        if tiene_cobertura and not _es_cobertura_pami(cobertura):
            continue

        fecha = _cell(row, headers, "fecha")
        fecha_turno = _parse_fecha(fecha)
        if desde and (fecha_turno is None or fecha_turno < desde):
            continue
        if hasta and (fecha_turno is None or fecha_turno > hasta):
            continue

        profesional = _cell(row, headers, "profesional")
        especialidad = _cell(row, headers, "especialidad")
        practica = _cell(row, headers, "practica")
        tipo_atencion = _cell(row, headers, "tipo atencion")
        if _excluir_turno_gjs(profesional, especialidad, practica, tipo_atencion):
            continue
        profesional_reporte = " - ".join(
            item for item in (profesional, especialidad, practica if practica and practica.upper() != "CONSULTA" else tipo_atencion) if item
        )
        nombre = _cell(row, headers, "nombre paciente")
        dni = _solo_digitos(_cell(row, headers, "documento"))
        beneficio_digits = _solo_digitos(_cell(row, headers, "numero afiliado"))
        beneficiario = beneficio_digits[:14] if len(beneficio_digits) >= 14 else ""

        if not beneficiario and not dni:
            continue

        key = beneficiario or f"DNI_{dni}"
        if key not in agrupados:
            agrupados[key] = {
                "key": key,
                "nombre": str(nombre or "").strip(),
                "dni": dni,
                "beneficiario": beneficiario,
                "usa_dni": not bool(beneficiario),
                "turnos": [],
            }

        agrupados[key]["turnos"].append(
            {
                "fecha": _formatear_fecha(fecha),
                "fecha_key": _fecha_key(fecha),
                "hora": _formatear_hora(_cell(row, headers, "hora")),
                "profesional": profesional_reporte or str(profesional or "").strip(),
            }
        )

    for paciente in agrupados.values():
        _normalizar_duplicados_cardio_gjs(paciente.get("turnos", []))
        _normalizar_duplicados_fono_gjs(paciente.get("turnos", []))

    return list(agrupados.values())


def _leer_pacientes_gjs_desde_transmision(
    rows: list[list[str]],
    headers: dict[str, int],
    fecha_desde: str | None = None,
    fecha_hasta: str | None = None,
) -> list[dict]:
    agrupados: dict[str, dict] = {}
    desde = _parse_fecha(fecha_desde) if fecha_desde else None
    hasta = _parse_fecha(fecha_hasta) if fecha_hasta else None

    for row in rows[1:]:
        turno_texto = _cell(row, headers, "turno")
        fecha = _extraer_fecha_desde_texto(turno_texto)
        fecha_turno = _parse_fecha(fecha)
        if desde and (fecha_turno is None or fecha_turno < desde):
            continue
        if hasta and (fecha_turno is None or fecha_turno > hasta):
            continue

        beneficiario = _solo_digitos(_cell(row, headers, "nro beneficio gp"))
        if len(beneficiario) > 14:
            beneficiario = beneficiario[:14]
        nombre = str(_cell(row, headers, "apellido y nombre") or "").strip()
        practica = str(_cell(row, headers, "practica") or "").strip()
        if not beneficiario or not practica:
            continue

        key = beneficiario
        if key not in agrupados:
            agrupados[key] = {
                "key": key,
                "nombre": nombre,
                "dni": "",
                "beneficiario": beneficiario,
                "usa_dni": False,
                "turnos": [],
            }

        agrupados[key]["turnos"].append(
            {
                "fecha": _formatear_fecha(fecha),
                "fecha_key": _fecha_key(fecha),
                "hora": _extraer_hora_desde_texto(turno_texto),
                "profesional": practica,
            }
        )

    return list(agrupados.values())


def leer_prestaciones_transmision_export(
    ruta_excel: Path,
    fecha_desde: str | None = None,
    fecha_hasta: str | None = None,
) -> dict[str, list[dict]]:
    rows = _leer_tabla_html_xls(ruta_excel)
    if not rows:
        return {}
    headers = {_clave_texto(value): idx for idx, value in enumerate(rows[0])}
    if {"nro orden", "nro beneficio gp", "apellido y nombre", "practica", "turno"} - set(headers):
        return {}

    desde = _parse_fecha(fecha_desde) if fecha_desde else None
    hasta = _parse_fecha(fecha_hasta) if fecha_hasta else None
    por_paciente: dict[str, list[dict]] = {}
    for row in rows[1:]:
        turno_texto = _cell(row, headers, "turno")
        fecha = _extraer_fecha_desde_texto(turno_texto)
        fecha_turno = _parse_fecha(fecha)
        if desde and (fecha_turno is None or fecha_turno < desde):
            continue
        if hasta and (fecha_turno is None or fecha_turno > hasta):
            continue

        beneficiario = _solo_digitos(_cell(row, headers, "nro beneficio gp"))
        if len(beneficiario) > 14:
            beneficiario = beneficiario[:14]
        practica = str(_cell(row, headers, "practica") or "").strip()
        if not beneficiario or not practica:
            continue

        transmitida_raw = str(_cell(row, headers, "trasmitida") or "").strip()
        f_transmitida = str(_cell(row, headers, "f transmitida") or "").strip()
        validada_raw = str(_cell(row, headers, "validada") or "").strip()
        f_validacion = str(_cell(row, headers, "f validacion") or "").strip()
        transmitida = transmitida_raw.upper().startswith("S") or bool(f_transmitida)
        validada = validada_raw.upper().startswith("S") or bool(f_validacion) or transmitida
        estado_accion = "auditoria_transmitida" if transmitida else "auditoria_validada" if validada else "auditoria_pendiente"
        transmitida_texto = f"SI - {f_transmitida}" if transmitida and f_transmitida else ("SI" if transmitida else "NO")

        por_paciente.setdefault(beneficiario, []).append(
            {
                "n_orden": str(_cell(row, headers, "nro orden") or "").strip(),
                "f_emision": str(_cell(row, headers, "fecha emision") or "").strip(),
                "n_beneficio": beneficiario,
                "nombre_pami": str(_cell(row, headers, "apellido y nombre") or "").strip(),
                "practica": practica,
                "f_agenda": _formatear_fecha(fecha),
                "turno_texto": turno_texto,
                "f_vencimiento": transmitida_texto,
                "transmitida": transmitida,
                "transmitida_texto": transmitida_texto,
                "validada": validada,
                "doc_cargada": transmitida,
                "doc_pendiente": validada and not transmitida,
                "estado_accion": estado_accion,
                "mismo_efector": True,
                "turno_asignado": True,
            }
        )
    return por_paciente


def _extraer_fecha_desde_texto(valor: str) -> str:
    match = re.search(r"\b\d{2}/\d{2}/\d{4}\b", str(valor or ""))
    return match.group(0) if match else str(valor or "").strip()


def _extraer_hora_desde_texto(valor: str) -> str:
    match = re.search(r"\b\d{2}:\d{2}\b", str(valor or ""))
    return match.group(0) if match else ""


def _normalizar_duplicados_fono_gjs(turnos: list[dict]) -> None:
    grupos: dict[tuple[str, str], list[dict]] = {}
    for turno in turnos:
        profesional = str(turno.get("profesional", "") or "")
        partes = [parte.strip() for parte in profesional.split(" - ") if parte.strip()]
        if len(partes) < 3 or _especialidad_base(partes[1]) != "fonoaudiologia":
            continue
        grupo = (turno.get("fecha_key") or _fecha_key(turno.get("fecha")), _clave_texto(" - ".join(partes[:2])))
        grupos.setdefault(grupo, []).append(turno)

    for grupo_turnos in grupos.values():
        tiene_audio = any(_especialidad_base(turno.get("profesional", "")) == "audiometria" for turno in grupo_turnos)
        if not tiene_audio:
            continue
        consulta_asignada = False
        for turno in grupo_turnos:
            if _especialidad_base(turno.get("profesional", "")) != "logoaudiometria":
                continue
            partes = [parte.strip() for parte in str(turno.get("profesional", "") or "").split(" - ") if parte.strip()]
            if len(partes) >= 3:
                partes[-1] = "CONSULTA"
                turno["profesional"] = " - ".join(partes)
                consulta_asignada = True
                break
        if consulta_asignada:
            continue

    consultas_vistas: set[tuple[str, str]] = set()
    filtrados: list[dict] = []
    for turno in turnos:
        profesional = str(turno.get("profesional", "") or "")
        partes = [parte.strip() for parte in profesional.split(" - ") if parte.strip()]
        if (
            len(partes) >= 3
            and _especialidad_base(partes[1]) == "fonoaudiologia"
            and _clave_texto(partes[-1]) == "consulta"
        ):
            grupo = (turno.get("fecha_key") or _fecha_key(turno.get("fecha")), _clave_texto(" - ".join(partes[:2])))
            if grupo in consultas_vistas:
                continue
            consultas_vistas.add(grupo)
        filtrados.append(turno)
    turnos[:] = filtrados


def _normalizar_duplicados_cardio_gjs(turnos: list[dict]) -> None:
    grupos_con_consulta: set[tuple[str, str, str]] = set()
    for turno in turnos:
        profesional = str(turno.get("profesional", "") or "")
        partes = [parte.strip() for parte in profesional.split(" - ") if parte.strip()]
        if len(partes) < 3:
            continue
        if not any(_especialidad_base(parte) == "ecg" for parte in partes[:2]):
            continue
        if _clave_texto(partes[-1]) != "consulta":
            continue
        grupos_con_consulta.add(
            (
                turno.get("fecha_key") or _fecha_key(turno.get("fecha")),
                _formatear_hora(turno.get("hora")),
                _clave_texto(" - ".join(partes[:2])),
            )
        )

    if not grupos_con_consulta:
        return

    filtrados: list[dict] = []
    for turno in turnos:
        profesional = str(turno.get("profesional", "") or "")
        partes = [parte.strip() for parte in profesional.split(" - ") if parte.strip()]
        if len(partes) >= 3 and any(_especialidad_base(parte) == "ecg" for parte in partes[:2]):
            grupo = (
                turno.get("fecha_key") or _fecha_key(turno.get("fecha")),
                _formatear_hora(turno.get("hora")),
                _clave_texto(" - ".join(partes[:2])),
            )
            if grupo in grupos_con_consulta and _especialidad_base(partes[-1]) == "ecg":
                continue
        filtrados.append(turno)
    turnos[:] = filtrados

    vistos: set[tuple[str, str]] = set()
    bases_estudio = {"audiometria", "logoaudiometria", "timpanometria", "acufenometria", "impedanciometria"}
    for turno in turnos:
        profesional = str(turno.get("profesional", "") or "")
        base = _especialidad_base(profesional)
        if base not in bases_estudio:
            continue
        partes = [parte.strip() for parte in profesional.split(" - ") if parte.strip()]
        if len(partes) < 3 or _especialidad_base(partes[1]) != "fonoaudiologia":
            continue
        grupo = (turno.get("fecha_key") or _fecha_key(turno.get("fecha")), _clave_texto(profesional))
        if grupo not in vistos:
            vistos.add(grupo)
            continue
        partes[-1] = "CONSULTA"
        turno["profesional"] = " - ".join(partes)

    consultas_vistas = set()
    filtrados = []
    for turno in turnos:
        profesional = str(turno.get("profesional", "") or "")
        partes = [parte.strip() for parte in profesional.split(" - ") if parte.strip()]
        if (
            len(partes) >= 3
            and _especialidad_base(partes[1]) == "fonoaudiologia"
            and _clave_texto(partes[-1]) == "consulta"
        ):
            grupo = (turno.get("fecha_key") or _fecha_key(turno.get("fecha")), _clave_texto(" - ".join(partes[:2])))
            if grupo in consultas_vistas:
                continue
            consultas_vistas.add(grupo)
        filtrados.append(turno)
    turnos[:] = filtrados


def get_ruta_progreso(ruta_excel: Path, cliente_codigo: str | None = CLIENTE_CIMA) -> Path:
    cliente = _cliente_config(cliente_codigo)
    sufijo_cliente = "" if cliente["codigo"] == CLIENTE_CIMA else f"_{cliente['codigo']}"
    return get_output_dir() / f"{Path(ruta_excel).stem}{sufijo_cliente}_progreso.json"


EXTRAER_OMES_SCRIPT = r"""
async (args) => {
  const fechaHoy = args?.fechaHoy || args || '';
  const fechaDesde = args?.fechaDesde || '';
  const fechaHasta = args?.fechaHasta || '';
  if (!window.bandeja || window.bandeja.length === 0) return [];
  const domInfo = {};
  document.querySelectorAll('table tbody tr').forEach(tr => {
    const calBtn = tr.querySelector('.fa-calendar, .fa-calendar-o, [class*="calendar"]');
    const aceptarBtn = tr.querySelector(
      'i.boton-historial[data-estado="aceptar"], .boton-historial[data-estado="aceptar"], ' +
      '.fa-check'
    );
    const activarBtn = tr.querySelector(
      'i.boton-historial[data-estado="activar"], .boton-historial[data-estado="activar"], ' +
      'i.boton-historial[data-estado="asignar"], .boton-historial[data-estado="asignar"], ' +
      '.fa-users, .fa-user-plus, .fa-user-md'
    );
    const modificarBtn = tr.querySelector('i.boton-historial[data-estado="modificar"], .boton-historial[data-estado="modificar"]');
    const cancelarBtn = tr.querySelector(
      'i.boton-historial[data-estado="cancelar"], .boton-historial[data-estado="cancelar"], ' +
      'i.boton-historial[data-estado="anular"], .boton-historial[data-estado="anular"], ' +
      '.fa-ban, .fa-times, .fa-remove'
    );
    const infoBtn = tr.querySelector(
      '.fa-info, .fa-info-circle, ' +
      'i.boton-historial[data-estado="informar"], .boton-historial[data-estado="informar"]'
    );
    const accionBtn = aceptarBtn || activarBtn || modificarBtn || cancelarBtn || calBtn || infoBtn;
    if (!accionBtn) return;
    const celdas = Array.from(tr.querySelectorAll('td')).map(td => (td.textContent || '').trim());
    const orden = accionBtn.getAttribute('data-orden') || celdas[0] || '';
    if (!orden) return;
    let estadoAccion = 'activada_info';
    let mismoEfector = true;
    if (aceptarBtn || activarBtn || (calBtn && infoBtn && !cancelarBtn)) {
      estadoAccion = 'disponible_activar';
      mismoEfector = true;
    } else if (calBtn && cancelarBtn) {
      estadoAccion = 'activada_modificable';
      mismoEfector = true;
    } else if (calBtn && infoBtn) {
      estadoAccion = 'activada_otro_prestador';
      mismoEfector = false;
    } else if (modificarBtn || calBtn) {
      estadoAccion = 'activada_modificable';
      mismoEfector = true;
    } else if (infoBtn) {
      estadoAccion = 'transmitida';
      mismoEfector = true;
    }
    domInfo[orden] = {
      mismo_efector: mismoEfector,
      estado_accion: estadoAccion
    };
  });
  const hoy = new Date(fechaHoy);
  const parseFecha = (valor) => {
    if (!valor) return null;
    const sf = String(valor).split(' ')[0];
    const p = sf.split('/');
    if (p.length !== 3) return null;
    const ft = new Date(parseInt(p[2]), parseInt(p[1])-1, parseInt(p[0]));
    return isNaN(ft) ? null : ft;
  };
  const desde = parseFecha(fechaDesde);
  const hasta = parseFecha(fechaHasta);
  const dentroRangoInforme = (fecha) => {
    if (!fecha) return false;
    if (desde && fecha < desde) return false;
    if (hasta && fecha > hasta) return false;
    return !!(desde || hasta);
  };
  const omes = [];
  for (const ome of window.bandeja) {
    const domData = domInfo[ome.n_orden];

    if (domData?.estado_accion === 'disponible_activar') {
      const fv = parseFecha(ome.f_vencimiento);
      if (fv && fv < hoy && !dentroRangoInforme(fv)) continue;
      omes.push({
        n_orden: ome.n_orden,
        n_beneficio: ome.n_beneficio,
        c_grado_paren: ome.c_grado_paren,
        practica: ome.d_practica,
        f_vencimiento: ome.f_vencimiento,
        f_agenda: '',
        mismo_efector: true,
        estado_accion: domData.estado_accion || '',
        turno_asignado: false,
        ya_transmitida: false,
        transmitida: false
      });
      continue;
    }

    if (ome.id_estado_efector !== '2' || !domData) continue;
    const infoOrden = domData;
    const resp = await fetch('ajax/efectores_detalle.php', {
      method: 'POST',
      headers: {'Content-Type': 'application/x-www-form-urlencoded'},
      body: `orden=${ome.n_orden}&estado=info&bene=${ome.n_beneficio}&gp=${ome.c_grado_paren}`
    });
    const det = await resp.json();
    const agenda = det.prescripcion?.[0]?.agenda?.[0];
    if (!agenda?.f_agenda) continue;
    const sf = agenda.f_agenda.split(' ')[0];
    const ft = parseFecha(sf);
    const fv = parseFecha(ome.f_vencimiento);
    if (!ft) continue;
    if (ft < hoy && (!fv || fv < hoy) && !dentroRangoInforme(ft)) continue;
    omes.push({
      n_orden: ome.n_orden,
      n_beneficio: ome.n_beneficio,
      c_grado_paren: ome.c_grado_paren,
      practica: ome.d_practica,
      f_vencimiento: ome.f_vencimiento,
      f_agenda: sf,
      mismo_efector: infoOrden.mismo_efector,
      estado_accion: infoOrden.estado_accion || '',
      turno_asignado: infoOrden.mismo_efector && infoOrden.estado_accion !== 'transmitida',
      ya_transmitida: infoOrden.estado_accion === 'transmitida'
    });
  }
  return omes;
}
"""

EXTRAER_TRANSMISION_SCRIPT = r"""
() => {
  const visible = (el) => !!el && (el.offsetParent !== null || el.getClientRects().length > 0);
  const texto = (el) => (el?.textContent || el?.value || '').replace(/\s+/g, ' ').trim();
  const esAzul = (el) => {
    const target = el?.closest('button,a,span') || el;
    const cls = `${target?.className || ''} ${el?.className || ''}`.toLowerCase();
    return cls.includes('btn-primary') || cls.includes('btn-info') || cls.includes('blue');
  };
  const esVerde = (el) => {
    const target = el?.closest('button,a,span') || el;
    const cls = `${target?.className || ''} ${el?.className || ''}`.toLowerCase();
    return cls.includes('btn-success') || cls.includes('green');
  };
  const fechaTurno = (valor) => {
    const match = String(valor || '').match(/\b\d{2}\/\d{2}\/\d{4}\b/);
    return match ? match[0] : '';
  };

  return Array.from(document.querySelectorAll('table tbody tr')).map((tr) => {
    const tds = Array.from(tr.querySelectorAll('td'));
    if (tds.length < 6) return null;
    const celdas = tds.map(texto);
    const acciones = tds[tds.length - 1] || tr;
    const check = acciones.querySelector('.fa-check');
    const doc = acciones.querySelector('.fa-upload, .fa-file, .fa-file-text, .fa-folder, .fa-cloud-upload');
    const transmitidaTexto = celdas[6] || '';
    const transmitida = /^si\b/i.test(transmitidaTexto);
    const validada = check ? esAzul(check) : transmitida;
    const docCargada = doc ? esAzul(doc) : false;
    const docPendiente = doc ? esVerde(doc) : false;
    let estadoAccion = 'auditoria_pendiente';
    if (transmitida) estadoAccion = 'auditoria_transmitida';
    else if (validada) estadoAccion = 'auditoria_validada';

    return {
      n_orden: celdas[0] || '',
      f_emision: celdas[1] || '',
      n_beneficio: celdas[2] || '',
      nombre_pami: celdas[3] || '',
      practica: celdas[4] || '',
      f_agenda: fechaTurno(celdas[5] || ''),
      turno_texto: celdas[5] || '',
      f_vencimiento: transmitidaTexto,
      transmitida: transmitida,
      transmitida_texto: transmitidaTexto,
      validada: validada,
      doc_cargada: docCargada,
      doc_pendiente: docPendiente,
      estado_accion: estadoAccion,
      mismo_efector: true,
      turno_asignado: true
    };
  }).filter(Boolean);
}
"""


class PamiVerificarController:
    def __init__(
        self,
        usuario: str,
        clave: str,
        log_callback: Callable[[str], None],
        resultado_callback: Callable[[str, list], None],
        progreso_callback: Callable[[dict], None] | None = None,
        respuesta_queue: queue.Queue | None = None,
        navegador_visible: bool = True,
    ) -> None:
        self.usuario = (usuario or "").strip()
        self.clave = clave or ""
        self.log_callback = log_callback
        self.resultado_callback = resultado_callback
        self.progreso_callback = progreso_callback
        self.respuesta_queue = respuesta_queue
        self._detener = threading.Event()
        self.thread: threading.Thread | None = None
        self.completed = False
        self.stopped = False
        self.error: str | None = None
        self.pacientes: list[dict] = []
        self.omes_por_key: dict[str, list] = {}
        self.retomar_progreso: bool | None = None
        self.cliente_codigo = CLIENTE_CIMA
        self.navegador_visible = navegador_visible

    def iniciar(
        self,
        ruta_excel: Path,
        fecha_hoy: str,
        fecha_desde: str | None = None,
        fecha_hasta: str | None = None,
        retomar_progreso: bool | None = None,
        cliente_codigo: str | None = CLIENTE_CIMA,
        modo_verificacion: str = "futuro",
    ) -> None:
        if self.thread and self.thread.is_alive():
            self._log("La verificacion ya esta en curso.")
            return

        self.completed = False
        self.stopped = False
        self.error = None
        self.retomar_progreso = retomar_progreso
        self.cliente_codigo = _cliente_config(cliente_codigo)["codigo"]
        modo_verificacion = (modo_verificacion or "futuro").strip().lower()
        self._detener.clear()
        self.thread = threading.Thread(
            target=self._run,
            args=(Path(ruta_excel), fecha_hoy, fecha_desde, fecha_hasta, self.cliente_codigo, modo_verificacion),
            daemon=True,
        )
        self.thread.start()

    def detener(self) -> None:
        self._detener.set()
        self._log("Se solicito detener la verificacion. Se frenara al finalizar el paciente actual.")

    def _run(
        self,
        ruta_excel: Path,
        fecha_hoy: str,
        fecha_desde: str | None,
        fecha_hasta: str | None,
        cliente_codigo: str,
        modo_verificacion: str,
    ) -> None:
        playwright = None
        browser = None
        context = None
        cliente = _cliente_config(cliente_codigo)
        ruta_progreso = get_ruta_progreso(ruta_excel, cliente["codigo"])
        if modo_verificacion == "auditoria":
            ruta_progreso = ruta_progreso.with_name(ruta_progreso.stem + "_auditoria" + ruta_progreso.suffix)
        finalizado_todos = False
        try:
            self._log(f"Cliente de verificacion: {cliente['nombre']}")
            self._log(
                "Modo de verificacion: "
                + ("Auditoria de prestaciones" if modo_verificacion == "auditoria" else "Control futuro")
            )
            self._log(f"Leyendo turnos {cliente['nombre']}: {ruta_excel}")
            self.pacientes = leer_pacientes_pami(
                ruta_excel,
                fecha_desde=fecha_desde,
                fecha_hasta=fecha_hasta,
                cliente_codigo=cliente["codigo"],
            )
            self._log(f"Pacientes PAMI encontrados: {len(self.pacientes)}")
            if fecha_desde or fecha_hasta:
                self._log(f"Filtro de fechas aplicado: desde {fecha_desde or '-'} hasta {fecha_hasta or '-'}")

            if not self.pacientes:
                self.completed = True
                return

            prestaciones_desde_export = (
                leer_prestaciones_transmision_export(ruta_excel, fecha_desde, fecha_hasta)
                if modo_verificacion == "auditoria"
                else {}
            )
            if prestaciones_desde_export:
                self._log(
                    "Archivo de Transmision detectado. La auditoria usara esas prestaciones como fuente directa."
                )
                fecha_panel_auditoria = (_parse_fecha(fecha_desde).isoformat() if _parse_fecha(fecha_desde) else fecha_hoy)
                mapa_especialidades = cargar_especialidades_medicos()
                pacientes_para_panel = [
                    paciente
                    for paciente in self.pacientes
                    if _faltan_requisitos_auditoria(
                        paciente,
                        prestaciones_desde_export.get(paciente.get("key", ""), []),
                        mapa_especialidades,
                        cliente["sistema_turnos"],
                    )
                ]
                if pacientes_para_panel:
                    self._log(
                        f"Se complementaran {len(pacientes_para_panel)} pacientes con el Panel de Aceptacion."
                    )
                    configurar_playwright()
                    playwright = sync_playwright().start()
                    browser = playwright.chromium.launch(
                        headless=not self.navegador_visible,
                        args=["--window-size=1280,900"] if self.navegador_visible else [],
                    )
                    context = browser.new_context(ignore_https_errors=True, viewport={"width": 1280, "height": 900})
                    page = context.new_page()
                    page.set_default_timeout(25000)
                    page.on("console", lambda msg: self._log(f"[Navegador] {msg.text}") if msg.type == "error" else None)
                    self._abrir_panel(page)
                    total_panel = len(pacientes_para_panel)
                    for idx_panel, paciente in enumerate(pacientes_para_panel, start=1):
                        if self._detener.is_set():
                            break
                        key_panel = paciente.get("key", "")
                        nombre_panel = paciente.get("nombre") or key_panel
                        self._log(
                            f"[{idx_panel}/{total_panel}] Complementando disponibles de {nombre_panel} en Panel de Aceptacion."
                        )
                        try:
                            prestaciones_desde_export[key_panel] = self._complementar_auditoria_con_panel(
                                page,
                                paciente,
                                prestaciones_desde_export.get(key_panel, []),
                                fecha_panel_auditoria,
                                mapa_especialidades,
                                cliente["sistema_turnos"],
                            )
                        except Exception as exc:
                            self._log(f"No se pudo complementar {nombre_panel}: {exc}")
                if ruta_progreso.exists():
                    ruta_progreso.unlink()
                    self._log("Archivo de progreso anterior eliminado para evitar resultados viejos.")
                total = len(self.pacientes)
                for idx, paciente in enumerate(self.pacientes, start=1):
                    if self._detener.is_set():
                        self._log("Verificacion detenida por el usuario.")
                        break
                    key = paciente["key"]
                    nombre = paciente.get("nombre") or key
                    self._log(f"[{idx}/{total}] Auditando prestaciones de {nombre} desde el archivo.")
                    omes = prestaciones_desde_export.get(key, [])
                    if not omes:
                        omes = [_ome_sin_resultados_pami()]
                    self.omes_por_key[key] = omes
                    self.resultado_callback(key, omes)
                    self._log(f"[{idx}/{total}] Prestaciones encontradas: {len(omes)}")
                self.stopped = self._detener.is_set()
                self.completed = True
                self._log("Verificacion detenida por el usuario." if self.stopped else "Verificacion finalizada.")
                return

            progreso = self._preparar_progreso(ruta_excel, ruta_progreso)
            completados = progreso.setdefault("completados", {})
            claves_actuales = {paciente["key"] for paciente in self.pacientes}
            for key, omes in list(completados.items()):
                if key in claves_actuales:
                    self.omes_por_key[key] = omes if isinstance(omes, list) else []
                    paciente = next((item for item in self.pacientes if item.get("key") == key), None)
                    if paciente and any(isinstance(ome, dict) and ome.get("verificar_benef") for ome in self.omes_por_key[key]):
                        paciente["verificar_benef"] = True
                        benef_pami = next(
                            (
                                str(ome.get("beneficiario_pami") or ome.get("n_beneficio") or "").strip()
                                for ome in self.omes_por_key[key]
                                if isinstance(ome, dict)
                                and str(ome.get("beneficiario_pami") or ome.get("n_beneficio") or "").strip()
                            ),
                            "",
                        )
                        if benef_pami:
                            paciente["beneficiario_pami"] = benef_pami
                    self.resultado_callback(key, self.omes_por_key[key])

            if claves_actuales and claves_actuales.issubset(set(completados.keys())):
                self.completed = True
                self._log("Todos los pacientes del rango ya estaban procesados en el progreso guardado.")
                if ruta_progreso.exists():
                    ruta_progreso.unlink()
                    self._log("Archivo de progreso eliminado porque el proceso ya estaba completo.")
                return

            configurar_playwright()
            playwright = sync_playwright().start()
            browser = playwright.chromium.launch(
                headless=not self.navegador_visible,
                args=["--window-size=1280,900"] if self.navegador_visible else [],
            )
            context = browser.new_context(ignore_https_errors=True, viewport={"width": 1280, "height": 900})
            page = context.new_page()
            page.set_default_timeout(25000)
            page.on("console", lambda msg: self._log(f"[Navegador] {msg.text}") if msg.type == "error" else None)

            if modo_verificacion == "auditoria":
                self._abrir_panel_transmision(page)
            else:
                self._abrir_panel(page)

            total = len(self.pacientes)
            for idx, paciente in enumerate(self.pacientes, start=1):
                if self._detener.is_set():
                    self._log("Verificacion detenida por el usuario.")
                    break

                key = paciente["key"]
                if key in completados:
                    nombre = paciente.get("nombre") or key
                    self._log(f"[{idx}/{total}] {nombre} ya estaba en progreso guardado. Se omite.")
                    continue

                nombre = paciente.get("nombre") or key
                self._log(
                    f"[{idx}/{total}] "
                    + (
                        f"Auditando prestaciones de {nombre}"
                        if modo_verificacion == "auditoria"
                        else f"Buscando OMEs de {nombre}"
                    )
                )
                try:
                    if modo_verificacion == "auditoria":
                        omes = self._procesar_paciente_transmision(page, paciente, fecha_desde, fecha_hasta)
                        if _faltan_requisitos_auditoria(
                            paciente,
                            omes,
                            cargar_especialidades_medicos(),
                            cliente["sistema_turnos"],
                        ):
                            self._log(
                                "Transmision no cubre todos los turnos. Se verifica tambien Panel de Aceptacion."
                            )
                            self._abrir_panel(page)
                            fecha_panel_auditoria = (
                                _parse_fecha(fecha_desde).isoformat() if _parse_fecha(fecha_desde) else fecha_hoy
                            )
                            omes = self._complementar_auditoria_con_panel(
                                page,
                                paciente,
                                omes,
                                fecha_panel_auditoria,
                                cargar_especialidades_medicos(),
                                cliente["sistema_turnos"],
                            )
                            self._abrir_panel_transmision(page)
                    else:
                        omes = self._procesar_paciente(page, paciente, fecha_hoy, fecha_desde, fecha_hasta)
                except Exception as exc:
                    self._log(f"ERROR al procesar {nombre}: {exc}")
                    omes = []

                self.omes_por_key[key] = omes
                completados[key] = omes
                self._guardar_progreso(ruta_progreso, progreso)
                self.resultado_callback(key, omes)
                self._log(
                    f"[{idx}/{total}] "
                    + (
                        f"Prestaciones encontradas: {len(omes)}"
                        if modo_verificacion == "auditoria"
                        else f"OMEs futuras encontradas: {len(omes)}"
                    )
                )
                page.wait_for_timeout(100)

            self.stopped = self._detener.is_set()
            finalizado_todos = not self.stopped
            self.completed = True
            self._log("Verificacion detenida por el usuario." if self.stopped else "Verificacion finalizada.")
            if finalizado_todos and ruta_progreso.exists():
                ruta_progreso.unlink()
                self._log("Archivo de progreso eliminado porque el proceso finalizo completo.")
        except Exception as exc:
            self.error = str(exc)
            self._log(f"ERROR general en verificacion: {exc}")
        finally:
            try:
                if context:
                    context.close()
                if browser:
                    browser.close()
                if playwright:
                    playwright.stop()
            except Exception:
                pass

    def _preparar_progreso(self, ruta_excel: Path, ruta_progreso: Path) -> dict:
        version = AUDITORIA_PROGRESS_VERSION if ruta_progreso.stem.endswith("_auditoria") else 1
        nuevo = {
            "fecha_inicio": date.today().isoformat(),
            "ruta_excel": str(ruta_excel),
            "version": version,
            "completados": {},
        }
        if not ruta_progreso.exists():
            self._guardar_progreso(ruta_progreso, nuevo)
            return nuevo

        retomar = self.retomar_progreso
        if retomar is None:
            retomar = self._preguntar_retomar_progreso(ruta_progreso)
        if not retomar:
            try:
                ruta_progreso.unlink()
            except Exception:
                pass
            self._log("Progreso anterior descartado. Se empieza desde cero.")
            self._guardar_progreso(ruta_progreso, nuevo)
            return nuevo

        try:
            data = json.loads(ruta_progreso.read_text(encoding="utf-8"))
            if int(data.get("version") or 1) != version:
                self._log("Progreso anterior incompatible con la version actual. Se empieza desde cero.")
                self._guardar_progreso(ruta_progreso, nuevo)
                return nuevo
            completados = data.get("completados", {})
            if not isinstance(completados, dict):
                completados = {}
            data["completados"] = completados
            data["ruta_excel"] = str(ruta_excel)
            self._log(f"Progreso retomado: {len(completados)} pacientes ya procesados.")
            return data
        except Exception as exc:
            self._log(f"No se pudo leer el progreso anterior ({exc}). Se empieza desde cero.")
            self._guardar_progreso(ruta_progreso, nuevo)
            return nuevo

    def _preguntar_retomar_progreso(self, ruta_progreso: Path) -> bool:
        if not self.progreso_callback or not self.respuesta_queue:
            self._log("Se encontro progreso previo y se retomara automaticamente.")
            return True

        self.progreso_callback({"type": "resume_prompt", "ruta": str(ruta_progreso)})
        try:
            respuesta = self.respuesta_queue.get(timeout=3600)
            return str(respuesta).strip().lower() == "retomar"
        except Exception:
            self._log("No hubo respuesta para retomar progreso. Se retomara automaticamente.")
            return True

    def _guardar_progreso(self, ruta_progreso: Path, data: dict) -> None:
        try:
            ruta_progreso.parent.mkdir(parents=True, exist_ok=True)
            ruta_progreso.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
        except Exception as exc:
            self._log(f"No se pudo guardar el progreso: {exc}")

    def _abrir_panel(self, page: Page) -> None:
        self._log("Abriendo panel de ordenes medicas PAMI...")
        self._asegurar_panel_activo(page)
        self._log("Panel PAMI listo.")

    def _abrir_panel_transmision(self, page: Page) -> None:
        self._log("Abriendo panel de transmision PAMI...")
        self._asegurar_transmision_activa(page)
        self._log("Panel de transmision listo.")

    def _asegurar_transmision_activa(self, page: Page, intentos: int = 2) -> None:
        for intento in range(1, intentos + 1):
            page.goto(PAMI_TRANSMISION_URL, wait_until="domcontentloaded")
            page.wait_for_timeout(1000)

            if "cup.pami.org.ar" in (page.url or ""):
                self._log("Sesion PAMI expirada o cerrada. Reingresando automaticamente...")
                self._login(page, force=True)
                page.goto(PAMI_TRANSMISION_URL, wait_until="domcontentloaded")
                page.wait_for_timeout(1000)

            if self._transmision_dom_activa(page):
                return

            self._log(f"No se pudo validar Transmision PAMI en intento {intento}/{intentos}.")

        raise RuntimeError("No se pudo acceder al panel de transmision PAMI.")

    def _transmision_dom_activa(self, page: Page) -> bool:
        try:
            if "cup.pami.org.ar" in (page.url or ""):
                return False
            if "transmision.php" in (page.url or ""):
                return True
            return page.locator("table tbody, button:has-text('Buscar'), input[type='button'][value*='Buscar' i]").count() > 0
        except Exception:
            return False

    def _ir_panel_rapido(self, page: Page) -> None:
        page.goto(PAMI_PANEL_URL, wait_until="domcontentloaded")
        page.wait_for_timeout(700)
        if "cup.pami.org.ar" in (page.url or ""):
            self._log("Sesion PAMI expirada o cerrada. Reingresando automaticamente...")
            self._login(page, force=True)
            page.goto(PAMI_PANEL_URL, wait_until="domcontentloaded")
            page.wait_for_timeout(700)
        if "cup.pami.org.ar" in (page.url or ""):
            raise RuntimeError("No se pudo recuperar la sesion de PAMI.")

    def _asegurar_panel_activo(self, page: Page, intentos: int = 2) -> None:
        for intento in range(1, intentos + 1):
            page.goto(PAMI_PANEL_URL, wait_until="domcontentloaded")
            page.wait_for_timeout(1000)

            if "cup.pami.org.ar" in (page.url or ""):
                self._log("Sesion PAMI expirada o cerrada. Reingresando automaticamente...")
                self._login(page, force=True)

            if self._panel_dom_activo(page):
                return

            self._log(f"No se pudo validar el panel PAMI en intento {intento}/{intentos}.")

        url = page.url or ""
        self._log(f"No se pudo validar el panel PAMI. URL actual: {url}")
        raise RuntimeError("No se pudo acceder al panel PAMI. Revisa usuario y clave.")

    def _login(self, page: Page, force: bool = False) -> None:
        if not self.usuario or not self.clave:
            raise RuntimeError("El portal pidio login y no hay credenciales cargadas.")

        if not force and self._session_is_active(page):
            self._log("La sesion de CUP ya estaba activa.")
            return

        self._log("Iniciando sesion automatica en CUP PAMI para Verificacion...")
        if "cup.pami.org.ar" not in (page.url or ""):
            page.goto(CUP_LOGIN_URL, wait_until="domcontentloaded")
        page.wait_for_timeout(350)

        user_input = page.locator('input[name="usuario"], input[type="text"], #usuario').first
        pass_input = page.locator('input[name="password"], input[type="password"], #password').first

        user_input.wait_for(state="visible")
        user_input.fill("")
        page.wait_for_timeout(120)
        user_input.type(self.usuario, delay=45)
        page.wait_for_timeout(120)

        pass_input.wait_for(state="visible")
        pass_input.fill("")
        page.wait_for_timeout(120)
        pass_input.type(self.clave, delay=45)
        page.wait_for_timeout(250)

        page.locator(
            'button[type="submit"], input[type="submit"], button:has-text("Ingresar"), button:has-text("Entrar")'
        ).first.click()
        page.wait_for_load_state("domcontentloaded")
        page.wait_for_timeout(700)

        if "cup.pami.org.ar" in (page.url or "") and not self._session_is_active(page):
            raise RuntimeError("No se pudo iniciar sesion en CUP PAMI. Revisa usuario y clave.")

        self._log("Sesion iniciada automaticamente en CUP PAMI.")

    def _session_is_active(self, page: Page) -> bool:
        try:
            response = page.goto(PAMI_PANEL_URL, wait_until="domcontentloaded")
            page.wait_for_timeout(1200)
            if "cup.pami.org.ar" in (page.url or ""):
                return False
            if response and response.status >= 400:
                return False
            return self._panel_dom_activo(page)
        except Exception:
            return False

    def _panel_dom_activo(self, page: Page) -> bool:
        try:
            if "cup.pami.org.ar" in (page.url or ""):
                return False
            if "efector.php" in (page.url or ""):
                return True
            return page.locator("table tbody, button:has-text('Buscar'), input[type='button'][value*='Buscar' i]").count() > 0
        except Exception:
            return False

    def _procesar_paciente(
        self,
        page: Page,
        paciente: dict,
        fecha_hoy: str,
        fecha_desde: str | None = None,
        fecha_hasta: str | None = None,
    ) -> list:
        modo = "dni" if paciente.get("usa_dni") else "beneficiario"
        valor = paciente.get("dni") if modo == "dni" else paciente.get("beneficiario")
        omes, cantidad = self._buscar_y_extraer(page, modo, str(valor or ""), fecha_hoy, fecha_desde, fecha_hasta)
        if modo == "dni" and cantidad == 0:
            paciente["no_pami"] = True

        if modo == "beneficiario" and cantidad == 0 and paciente.get("dni"):
            self._log("La busqueda por beneficiario no trajo resultados. Se intenta con DNI para verificar el BENEF.")
            omes, cantidad_dni = self._buscar_y_extraer(
                page,
                "dni",
                str(paciente.get("dni") or ""),
                fecha_hoy,
                fecha_desde,
                fecha_hasta,
            )
            if cantidad_dni > 0:
                paciente["verificar_benef"] = True
                benef_pami = next(
                    (
                        str(ome.get("n_beneficio", "") or "").strip()
                        for ome in omes
                        if isinstance(ome, dict) and str(ome.get("n_beneficio", "") or "").strip()
                    ),
                    "",
                )
                if benef_pami:
                    paciente["beneficiario_pami"] = benef_pami
                for ome in omes:
                    if isinstance(ome, dict):
                        ome["verificar_benef"] = True
                        if benef_pami:
                            ome["beneficiario_pami"] = benef_pami
                if benef_pami:
                    self._log(f"Se encontraron resultados por DNI. BENEF detectado en PAMI: {benef_pami}.")
                else:
                    self._log("Se encontraron resultados por DNI. El reporte marcara Verificar BENEF.")
            else:
                paciente["no_pami"] = True
                self._log("Tampoco se encontraron resultados por DNI.")

        return omes

    def _procesar_paciente_transmision(
        self,
        page: Page,
        paciente: dict,
        fecha_desde: str | None,
        fecha_hasta: str | None,
    ) -> list:
        modo = "dni" if paciente.get("usa_dni") else "beneficiario"
        valor = paciente.get("dni") if modo == "dni" else paciente.get("beneficiario")
        prestaciones, cantidad = self._buscar_y_extraer_transmision(
            page,
            modo,
            str(valor or ""),
            fecha_desde,
            fecha_hasta,
        )
        if modo == "dni" and cantidad == 0:
            paciente["no_pami"] = True

        if modo == "beneficiario" and cantidad == 0 and paciente.get("dni"):
            self._log("Transmision no trajo resultados por beneficiario. Se intenta con DNI para verificar el BENEF.")
            prestaciones, cantidad_dni = self._buscar_y_extraer_transmision(
                page,
                "dni",
                str(paciente.get("dni") or ""),
                fecha_desde,
                fecha_hasta,
            )
            if cantidad_dni > 0:
                paciente["verificar_benef"] = True
                benef_pami = next(
                    (
                        str(item.get("n_beneficio", "") or "").strip()
                        for item in prestaciones
                        if isinstance(item, dict) and str(item.get("n_beneficio", "") or "").strip()
                    ),
                    "",
                )
                if benef_pami:
                    paciente["beneficiario_pami"] = benef_pami
                    for item in prestaciones:
                        if isinstance(item, dict):
                            item["verificar_benef"] = True
                            item["beneficiario_pami"] = benef_pami
            else:
                paciente["no_pami"] = True

        if cantidad == 0 and not prestaciones:
            return [_ome_sin_resultados_pami()]
        return prestaciones

    def _complementar_auditoria_con_panel(
        self,
        page: Page,
        paciente: dict,
        prestaciones: list[dict],
        fecha_panel_auditoria: str,
        mapa_especialidades: dict[str, list[str]],
        sistema_turnos: str,
    ) -> list[dict]:
        omes_panel = self._procesar_paciente(page, paciente, fecha_panel_auditoria)
        self._log(f"Panel de Aceptacion devolvio {len(omes_panel)} OMEs para complementar auditoria.")
        resultado = _fusionar_omes_auditoria(
            prestaciones,
            omes_panel,
            paciente,
            mapa_especialidades,
            sistema_turnos,
        )

        if not _faltan_requisitos_auditoria(paciente, resultado, mapa_especialidades, sistema_turnos):
            return resultado

        if paciente.get("usa_dni") or not paciente.get("beneficiario") or not paciente.get("dni"):
            return resultado

        self._log("Siguen faltando OMEs con beneficiario. Se verifica tambien por DNI por posible BENEF incorrecto.")
        paciente_dni = dict(paciente)
        paciente_dni["usa_dni"] = True
        omes_dni = self._procesar_paciente(page, paciente_dni, fecha_panel_auditoria)
        omes_dni_validas = [
            ome
            for ome in omes_dni
            if isinstance(ome, dict) and ome.get("estado_accion") != "sin_resultados_pami"
        ]
        if not omes_dni_validas:
            paciente["no_pami"] = True
            self._log("La verificacion por DNI no encontro OMEs adicionales.")
            return resultado

        benef_pami = next(
            (
                str(ome.get("n_beneficio", "") or "").strip()
                for ome in omes_dni_validas
                if str(ome.get("n_beneficio", "") or "").strip()
            ),
            "",
        )
        for ome in omes_dni_validas:
            ome["verificar_benef"] = True
            if benef_pami:
                ome["beneficiario_pami"] = benef_pami

        resultado_dni = _fusionar_omes_auditoria(
            resultado,
            omes_dni_validas,
            paciente,
            mapa_especialidades,
            sistema_turnos,
        )
        if len(resultado_dni) <= len(resultado) and _faltan_requisitos_auditoria(
            paciente,
            resultado_dni,
            mapa_especialidades,
            sistema_turnos,
        ):
            self._log("La busqueda por DNI encontro OMEs, pero ninguna cubre los faltantes de auditoria.")
            return resultado

        paciente["verificar_benef"] = True
        if benef_pami:
            paciente["beneficiario_pami"] = benef_pami
            self._log(f"La busqueda por DNI cubrio faltantes. BENEF detectado en PAMI: {benef_pami}.")
        else:
            self._log("La busqueda por DNI cubrio faltantes. El reporte marcara Verificar BENEF.")
        return resultado_dni

    def _buscar_y_extraer_transmision(
        self,
        page: Page,
        modo: str,
        valor: str,
        fecha_desde: str | None,
        fecha_hasta: str | None,
    ) -> tuple[list, int]:
        self._asegurar_transmision_activa(page, intentos=1)
        try:
            self._click_accion_visible(page, "Limpiar", timeout=3000)
            page.wait_for_timeout(250)
        except Exception:
            self._log("No se encontro el boton Limpiar en Transmision; se continua con la busqueda.")

        self._seleccionar_tipo_busqueda(page, "2" if modo == "dni" else "1")
        self._aplicar_fechas_transmision(page, fecha_desde, fecha_hasta)
        self._log(f"Completando busqueda en Transmision por {'DNI' if modo == 'dni' else 'beneficiario'}: {valor}")
        self._llenar_campo_busqueda(page, str(valor or ""), modo)
        self._click_buscar(page)
        page.wait_for_timeout(1200)

        if "cup.pami.org.ar" in (page.url or "") or not self._transmision_dom_activa(page):
            self._log("La sesion se perdio durante la auditoria. Reingresando y reintentando paciente...")
            self._asegurar_transmision_activa(page)
            return self._buscar_y_extraer_transmision(page, modo, valor, fecha_desde, fecha_hasta)

        prestaciones: list[dict] = []
        paginas = 0
        while paginas < 12:
            pagina_items = page.evaluate(EXTRAER_TRANSMISION_SCRIPT)
            if isinstance(pagina_items, list):
                prestaciones.extend(item for item in pagina_items if isinstance(item, dict))
            paginas += 1
            advanced = self._avanzar_pagina_transmision(page)
            if not advanced:
                break
            page.wait_for_timeout(900)

        buscado = _solo_digitos(valor)
        if buscado and modo == "beneficiario":
            prestaciones = [
                item for item in prestaciones
                if not _solo_digitos(item.get("n_beneficio", "")) or _solo_digitos(item.get("n_beneficio", "")) == buscado
            ]
        self._log(f"Registros encontrados en Transmision: {len(prestaciones)}")
        return prestaciones, len(prestaciones)

    def _aplicar_fechas_transmision(self, page: Page, fecha_desde: str | None, fecha_hasta: str | None) -> None:
        page.evaluate(
            """
            ({desde, hasta}) => {
              const setValor = (selector, valor) => {
                const el = document.querySelector(selector);
                if (!el || !valor) return false;
                el.value = valor;
                el.dispatchEvent(new Event('input', { bubbles: true }));
                el.dispatchEvent(new Event('change', { bubbles: true }));
                return true;
              };
              setValor('input[name="f_turno_desde"]', desde || '');
              setValor('input[name="f_turno_hasta"]', hasta || '');
            }
            """,
            {"desde": fecha_desde or "", "hasta": fecha_hasta or ""},
        )

    def _avanzar_pagina_transmision(self, page: Page) -> bool:
        try:
            return bool(
                page.evaluate(
                    """
                    () => {
                      const visible = (el) => !!el && (el.offsetParent !== null || el.getClientRects().length > 0);
                      const paginaActiva = () => {
                        const activa = Array.from(document.querySelectorAll('a, button, li, span'))
                          .find((el) => /^\\d+$/.test((el.textContent || '').trim()) && /active|current|selected|btn-primary/.test(el.className || ''));
                        if (activa) return Number((activa.textContent || '').trim());
                        const url = new URL(window.location.href);
                        return Number(url.searchParams.get('pagina') || '1');
                      };
                      const actual = paginaActiva();
                      const mas = Array.from(document.querySelectorAll('a, button'))
                        .find((el) => visible(el) && (el.textContent || '').trim().toLowerCase().includes('mas resultados'));
                      if (mas) {
                        mas.click();
                        return true;
                      }
                      const siguiente = Array.from(document.querySelectorAll('a, button'))
                        .find((el) => {
                          if (!visible(el) || el.disabled) return false;
                          const txt = (el.textContent || '').trim();
                          if (!/^\\d+$/.test(txt)) return false;
                          if (Number(txt) !== actual + 1) return false;
                          return !/active|current|selected/.test(el.className || '');
                        });
                      if (!siguiente) return false;
                      siguiente.click();
                      return true;
                    }
                    """
                )
            )
        except Exception:
            return False

    def _buscar_y_extraer(
        self,
        page: Page,
        modo: str,
        valor: str,
        fecha_hoy: str,
        fecha_desde: str | None = None,
        fecha_hasta: str | None = None,
        reintentar_sesion: bool = True,
        usar_panel_actual: bool = True,
        limpiar_antes: bool = False,
    ) -> tuple[list, int]:
        self._ir_panel_rapido(page)

        try:
            self._click_accion_visible(page, "Limpiar", timeout=3000)
            page.wait_for_timeout(120)
        except Exception:
            self._log("No se encontro el boton Limpiar; se continua con la busqueda.")

        self._seleccionar_tipo_busqueda(page, "2" if modo == "dni" else "1")
        self._log(f"Completando busqueda por {'DNI' if modo == 'dni' else 'beneficiario'}: {valor}")
        self._llenar_campo_busqueda(page, str(valor or ""), modo)
        self._resetear_bandeja(page)
        self._click_buscar(page)
        self._log("Click en Buscar ejecutado.")

        try:
            page.wait_for_function("window.bandeja !== undefined", timeout=10000)
        except TimeoutError:
            if reintentar_sesion and ("cup.pami.org.ar" in (page.url or "") or not self._panel_dom_activo(page)):
                self._log("La sesion se perdio durante la busqueda. Reingresando y reintentando paciente...")
                self._asegurar_panel_activo(page)
                return self._buscar_y_extraer(
                    page,
                    modo,
                    valor,
                    fecha_hoy,
                    fecha_desde,
                    fecha_hasta,
                    reintentar_sesion=False,
                    usar_panel_actual=False,
                    limpiar_antes=True,
                )
            self._log("No aparecio window.bandeja en 10 segundos; se toma como sin OMEs encontradas.")
            return [], 0

        if "registros_por_pagina=50" not in (page.url or "") and "efector.php" in (page.url or ""):
            self._ir_panel_rapido(page)
            try:
                page.wait_for_function("window.bandeja !== undefined", timeout=10000)
            except TimeoutError:
                self._log("No aparecio window.bandeja despues de recuperar el panel de 50; se continua con el resultado disponible.")

        cantidad = page.evaluate("Array.isArray(window.bandeja) ? window.bandeja.length : 0")
        self._log(f"Registros encontrados en bandeja: {cantidad}")
        if cantidad == 0 and reintentar_sesion and ("cup.pami.org.ar" in (page.url or "") or not self._panel_dom_activo(page)):
            self._log("La busqueda dio 0 y la sesion parece invalida. Reingresando y reintentando paciente...")
            self._asegurar_panel_activo(page)
            return self._buscar_y_extraer(
                page,
                modo,
                valor,
                fecha_hoy,
                fecha_desde,
                fecha_hasta,
                reintentar_sesion=False,
                usar_panel_actual=False,
                limpiar_antes=True,
            )
        resultado_no_valido = self._resultado_no_corresponde_al_paciente(page, modo, valor, cantidad)
        if resultado_no_valido and not limpiar_antes and usar_panel_actual:
            self._log("La busqueda rapida no valido el paciente. Se reintenta limpiando filtros.")
            return self._buscar_y_extraer(
                page,
                modo,
                valor,
                fecha_hoy,
                fecha_desde,
                fecha_hasta,
                reintentar_sesion=reintentar_sesion,
                usar_panel_actual=True,
                limpiar_antes=True,
            )
        if resultado_no_valido and limpiar_antes and usar_panel_actual:
            self._log("La busqueda con Limpiar no valido el paciente. Se recarga el panel y se reintenta.")
            return self._buscar_y_extraer(
                page,
                modo,
                valor,
                fecha_hoy,
                fecha_desde,
                fecha_hasta,
                reintentar_sesion=False,
                usar_panel_actual=False,
                limpiar_antes=True,
            )
        if cantidad == 0 and not limpiar_antes and usar_panel_actual:
            self._log("La busqueda rapida dio 0 resultados. Se reintenta limpiando filtros.")
            return self._buscar_y_extraer(
                page,
                modo,
                valor,
                fecha_hoy,
                fecha_desde,
                fecha_hasta,
                reintentar_sesion=reintentar_sesion,
                usar_panel_actual=True,
                limpiar_antes=True,
            )
        resultado = page.evaluate(
            EXTRAER_OMES_SCRIPT,
            {"fechaHoy": fecha_hoy, "fechaDesde": fecha_desde or "", "fechaHasta": fecha_hasta or ""},
        )
        return (resultado if isinstance(resultado, list) else []), int(cantidad or 0)

    def _resetear_bandeja(self, page: Page) -> None:
        try:
            page.evaluate("window.bandeja = undefined")
        except Exception:
            pass

    def _resultado_no_corresponde_al_paciente(
        self,
        page: Page,
        modo: str,
        valor: str,
        cantidad: int,
    ) -> bool:
        if cantidad <= 0 or modo != "beneficiario":
            return False
        buscado = _solo_digitos(valor)
        if not buscado:
            return False
        try:
            encontrados = page.evaluate(
                """
                () => Array.isArray(window.bandeja)
                  ? window.bandeja.flatMap(item => {
                      const bene = String(item.n_beneficio || '').replace(/\\D/g, '');
                      const gp = String(item.c_grado_paren || '').replace(/\\D/g, '');
                      return [bene, `${bene}${gp}`].filter(Boolean);
                    })
                  : []
                """
            )
        except Exception:
            return False
        return bool(encontrados) and buscado not in set(encontrados)

    def _seleccionar_tipo_busqueda(self, page: Page, value: str) -> None:
        script = """
        (value) => {
          const selects = Array.from(document.querySelectorAll('select')).filter(s => s.offsetParent !== null);
          const select = selects.find(s => Array.from(s.options).some(o => o.value === value));
          if (!select) return false;
          select.value = value;
          select.dispatchEvent(new Event('change', {bubbles: true}));
          return true;
        }
        """
        if not page.evaluate(script, value):
            self._log("No se pudo cambiar el tipo de busqueda; se usa el valor visible por defecto.")

    def _click_buscar(self, page: Page) -> None:
        clicked = page.evaluate(
            """
            () => {
              const visible = (el) => !!el && (el.offsetParent !== null || el.getClientRects().length > 0);
              const candidatos = Array.from(document.querySelectorAll('button, input[type="submit"], input[type="button"], a'));
              const buscar = candidatos.find((el) => {
                const texto = String(el.textContent || el.value || '').trim().toLowerCase();
                return texto === 'buscar' && visible(el);
              });
              if (!buscar) return false;
              buscar.click();
              return true;
            }
            """
        )
        if clicked:
            return

        selectores = [
            'button:has-text("Buscar")',
            'input[type="submit"][value*="Buscar" i]',
            'input[type="button"][value*="Buscar" i]',
            'a:has-text("Buscar")',
        ]
        for selector in selectores:
            try:
                locator = page.locator(selector).first
                if locator.count() and locator.is_visible():
                    locator.click(timeout=5000, force=True)
                    return
            except Exception:
                continue

        raise RuntimeError("No se encontro el boton Buscar del panel.")

    def _click_accion_visible(self, page: Page, texto: str, timeout: int = 5000) -> None:
        selectores = [
            f'button:has-text("{texto}")',
            f'input[type="submit"][value*="{texto}" i]',
            f'input[type="button"][value*="{texto}" i]',
            f'a:has-text("{texto}")',
        ]
        for selector in selectores:
            try:
                locator = page.locator(selector).first
                if locator.count() and locator.is_visible():
                    locator.click(timeout=timeout)
                    return
            except Exception:
                continue
        raise RuntimeError(f"No se encontro el boton {texto}.")

    def _llenar_campo_busqueda(self, page: Page, valor: str, modo: str) -> None:
        filled = page.evaluate(
            """
            ({valor, modo}) => {
              const visible = (el) => !!el && (el.offsetParent !== null || el.getClientRects().length > 0);
              const setValue = (input, value) => {
                if (!input) return false;
                const setter = Object.getOwnPropertyDescriptor(window.HTMLInputElement.prototype, 'value')?.set;
                const assign = (v) => {
                  if (setter) setter.call(input, v);
                  else input.value = v;
                };
                const fire = () => {
                  input.dispatchEvent(new Event('input', {bubbles: true}));
                  input.dispatchEvent(new Event('change', {bubbles: true}));
                  input.dispatchEvent(new KeyboardEvent('keyup', {bubbles: true}));
                  if (window.angular) {
                    try {
                      window.angular.element(input).triggerHandler('input');
                      window.angular.element(input).triggerHandler('change');
                    } catch (_) {}
                  }
                };
                const normalizar = (v) => String(v || '').replace(/\\D/g, '');
                input.focus();
                input.select();
                assign('');
                fire();
                assign(value);
                fire();
                if (normalizar(input.value) !== normalizar(value)) {
                  input.select();
                  assign('');
                  fire();
                  assign(value);
                  fire();
                }
                input.blur();
                return normalizar(input.value) === normalizar(value);
              };

              const clearIfExists = (selector) => {
                const el = Array.from(document.querySelectorAll(selector)).find(visible);
                if (el) setValue(el, '');
              };
              clearIfExists('input[ng-model*="practica"], input[name*="practica"], input[id*="practica"]');

              const selectors = modo === 'dni'
                ? [
                    'input[ng-model*="documento"]',
                    'input[ng-model*="doc"]',
                    'input[name*="documento"]',
                    'input[name*="doc"]',
                    'input[placeholder*="documento" i]'
                  ]
                : [
                    'input[ng-model*="beneficio"]',
                    'input[ng-model*="afiliado"]',
                    'input[name*="beneficio"]',
                    'input[name*="afiliado"]',
                    'input[name="n_afiliado"]',
                    'input[placeholder*="beneficio" i]',
                    'input[placeholder*="afiliado" i]'
                  ];

              for (const selector of selectors) {
                const input = Array.from(document.querySelectorAll(selector)).find(visible);
                if (input) return setValue(input, valor);
              }

              const selects = Array.from(document.querySelectorAll('select')).filter(visible);
              const afiliadoSelect = selects.find((select) => {
                const text = Array.from(select.options).map(o => `${o.value} ${o.textContent || ''}`).join(' ').toLowerCase();
                return text.includes('afiliado') || text.includes('documento') || text.includes('gp');
              });
              if (!afiliadoSelect) return false;

              const selectRect = afiliadoSelect.getBoundingClientRect();
              const inputs = Array.from(document.querySelectorAll('input'))
                .filter(input => visible(input) && !input.disabled && input.type !== 'hidden')
                .filter(input => {
                  const text = [
                    input.name || '',
                    input.id || '',
                    input.getAttribute('ng-model') || '',
                    input.placeholder || ''
                  ].join(' ').toLowerCase();
                  return !text.includes('orden') && !text.includes('practica') && !text.includes('fecha') && !text.includes('emision');
                })
                .map(input => {
                  const r = input.getBoundingClientRect();
                  const horizontal = Math.abs(r.left - selectRect.left);
                  const vertical = Math.abs(r.top - selectRect.bottom);
                  const belowOrSameBlock = r.top >= selectRect.top - 8 && r.top <= selectRect.bottom + 55;
                  return {input, score: horizontal + vertical + (belowOrSameBlock ? 0 : 1000)};
                })
                .sort((a, b) => a.score - b.score);

              return inputs.length ? setValue(inputs[0].input, valor) : false;
            }
            """,
            {"valor": valor, "modo": modo},
        )
        if not filled:
            raise RuntimeError("No se encontro el campo de busqueda del panel.")

    def _log(self, message: str) -> None:
        log_message(message)
        try:
            self.log_callback(message)
        except Exception:
            pass


def _ruta_especialidades_default() -> Path | None:
    candidatos = [
        Path.home() / "Downloads" / ESPECIALIDADES_FILENAME,
        Path.cwd() / ESPECIALIDADES_FILENAME,
    ]
    for ruta in candidatos:
        if ruta.exists():
            return ruta
    return None


def cargar_especialidades_medicos(
    ruta: Path | None = None,
    sistema_turnos: str = "CIMA",
) -> dict[str, list[str]]:
    global _ESPECIALIDADES_CACHE
    sistema_key = _sistema_turnos_key(sistema_turnos)
    if ruta is None and sistema_key in _ESPECIALIDADES_CACHE:
        return _ESPECIALIDADES_CACHE[sistema_key]

    ruta = ruta or (_ruta_especialidades_default() if sistema_key == "CIMA" else None)
    mapa: dict[str, list[str]] = {}
    if not ruta or not ruta.exists():
        manuales = PROFESIONAL_ESPECIALIDADES_MANUALES_POR_SISTEMA.get(sistema_key, {})
        normalizado = {
            _clave_texto(profesional): _filtrar_especialidades_por_sistema(list(especialidades), sistema_key)
            for profesional, especialidades in manuales.items()
        }
        normalizado = {clave: valores for clave, valores in normalizado.items() if valores}
        if ruta is None:
            _ESPECIALIDADES_CACHE[sistema_key] = normalizado
        return normalizado

    def es_especialidad(valor: str) -> bool:
        raw = str(valor or "").lower()
        clave = _clave_texto(valor)
        return (
            "(pami" in raw
            or "(particular" in raw
            or "(plan de salud" in raw
            or "(santa cecilia" in raw
            or clave.startswith("rx ")
        )

    try:
        wb = load_workbook(ruta, read_only=True, data_only=True)
        for sheet_name, col_idx in (("Cima medico y especialidad", 1), ("Catan medico y especialidad", 0)):
            if sheet_name not in wb.sheetnames:
                continue
            current = ""
            for row in wb[sheet_name].iter_rows(values_only=True):
                value = row[col_idx] if len(row) > col_idx else None
                if not value:
                    continue
                text = str(value).strip()
                if not text or text.lower().startswith("profesional"):
                    continue
                if es_especialidad(text) and current:
                    mapa.setdefault(current, []).append(text)
                else:
                    current = text
                    mapa.setdefault(current, [])
    except Exception as exc:
        log_message(f"No se pudo cargar el Excel de especialidades: {exc}")
        mapa = {}

    normalizado: dict[str, list[str]] = {}
    for profesional, especialidades in mapa.items():
        filtradas = _filtrar_especialidades_por_sistema(list(especialidades), sistema_key)
        if filtradas:
            normalizado[_clave_texto(profesional)] = filtradas
    for profesional, especialidades in PROFESIONAL_ESPECIALIDADES_MANUALES_POR_SISTEMA.get(sistema_key, {}).items():
        filtradas = _filtrar_especialidades_por_sistema(list(especialidades), sistema_key)
        if filtradas:
            normalizado[_clave_texto(profesional)] = filtradas

    if ruta is None:
        _ESPECIALIDADES_CACHE[sistema_key] = normalizado
    return normalizado


def _especialidades_profesional(
    profesional: str,
    mapa: dict[str, list[str]],
    sistema_turnos: str = "CIMA",
) -> list[str]:
    sistema_key = _sistema_turnos_key(sistema_turnos)
    manuales = PROFESIONAL_ESPECIALIDADES_MANUALES_POR_SISTEMA.get(sistema_key, {})
    profesional_raw = _normalizar_texto(profesional)
    clave = _clave_texto(profesional)
    partes = [parte.strip() for parte in str(profesional or "").split(" - ") if parte.strip()]
    if len(partes) >= 3:
        practica = _especialidad_base(partes[-1])
        grupo = _clave_texto(" - ".join(partes[:-1]))
        if practica == "holter":
            return ["Holter"]
        if practica == "consulta" and "holter" in grupo and ("cardio" in grupo or "cardiologia" in grupo):
            return ["Cardiologia"]
    if "holter" in profesional_raw:
        return ["Holter"]
    if "mapa" in profesional_raw:
        clave_mapa = f"{clave} mapa"
        if clave_mapa in manuales:
            return manuales[clave_mapa]
        return ["MAPA"]
    if sistema_key == "CIMA" and ("ecocardiograma" in profesional_raw or "ecocardio" in profesional_raw) and "vasos de cuello" in profesional_raw:
        return ["ecocardio_o_vasos"]
    if sistema_key == "CIMA" and "doppler" in profesional_raw and (
        "eco stress" in profesional_raw
        or "eco estres" in profesional_raw
        or "ecoestres" in profesional_raw
        or "ecostress" in profesional_raw
        or (("ecocardio" in profesional_raw or "ecocardiograma" in profesional_raw) and "stress" in profesional_raw)
    ):
        return ["doppler_o_ecoestres"]
    if sistema_key == "CIMA" and "diabetologia" in profesional_raw and "gastro" in profesional_raw:
        return ["diabetologia_o_gastro"]
    if sistema_key == "CIMA" and "aoad" in profesional_raw:
        return ["gastroenterologia"]
    if "ecg" in profesional_raw or "e c g" in profesional_raw:
        clave_ecg = f"{clave} ecg"
        if clave_ecg in manuales:
            return manuales[clave_ecg]
        return ["ECG"]
    clave = PROFESIONAL_ALIASES.get(clave, clave)
    if _turno_cima_ramirez_walter_ambiguo({"profesional": profesional}, sistema_key):
        return []
    if _turno_cima_novelli_dario_ambiguo({"profesional": profesional}, sistema_key):
        return []
    especialidades = mapa.get(clave, [])
    if especialidades:
        return _filtrar_especialidades_por_sistema(especialidades, sistema_key)
    base = _especialidad_base(clave)
    if _base_habilitada_para_sistema(base, sistema_key) and (
        base in PRACTICAS_COMPATIBLES_POR_ESPECIALIDAD or base in PRACTICAS_REQUERIDAS_POR_ESPECIALIDAD
    ):
        return [clave]
    return []


def _especialidad_base(valor: str) -> str:
    texto_con_parentesis = _texto_busqueda(valor)
    if (
        "ecocardio o vasos" in texto_con_parentesis
        or "ecocardiograma o vasos" in texto_con_parentesis
        or "ecocardio o vasos de cuello" in texto_con_parentesis
        or "ecocardiograma o vasos de cuello" in texto_con_parentesis
    ):
        return "ecocardio_o_vasos"
    if "doppler o ecoestres" in texto_con_parentesis or "doppler o eco estres" in texto_con_parentesis:
        return "doppler_o_ecoestres"
    if "diabetologia o gastro" in texto_con_parentesis:
        return "diabetologia_o_gastro"
    texto = _clave_texto(valor)
    if "carrera dbt" in texto or "dbt doble" in texto:
        return "carrera_dbt"
    reemplazos = {
        "gatroenterologia": "gastroenterologia",
        "gastro": "gastroenterologia",
        "cardiologia": "cardiologia",
        "neurologia": "neurologia",
        "neumonologia": "neumonologia",
        "otorrinolaringologia": "otorrinolaringologia",
        "traumatologia": "traumatologia",
        "diabetologia": "diabetologia",
        "endocrinologia": "endocrinologia",
        "dermatologia": "dermatologia",
        "urologia": "urologia",
        "nutricion": "nutricion",
        "hematologia": "hematologia",
        "ginecologia": "ginecologia",
        "reumatologia": "reumatologia",
        "nefrologia": "nefrologia",
        "flebologia": "flebologia",
    }
    if "e e g" in texto or "eeg" in texto or "electroencefal" in texto:
        return "electroencefalografia"
    if "logoaudiometria" in texto or "logoaudimetria" in texto:
        return "logoaudiometria"
    if "audiometria" in texto:
        return "audiometria"
    if "timpanometria" in texto:
        return "timpanometria"
    if "acufenometria" in texto or "acufeno" in texto:
        return "acufenometria"
    if "impedanciometria" in texto or "impedancia" in texto or "inpedancia" in texto:
        return "impedanciometria"
    if "estudios fonoaudiologicos" in texto or "fonoaudiologicos" in texto:
        return "fono_estudios"
    if "fonoaudiologia" in texto:
        return "fonoaudiologia"
    if (
        ("lavaje" in texto and ("oido" in texto or "seno" in texto or "proetz" in texto))
        or ("extraccion" in texto and ("cuerpo extra" in texto or "tapon" in texto or "cerumen" in texto))
    ):
        return "lavaje_oido"
    if "rinofibro" in texto or "rinofibrolaringoscopia" in texto or "rinofibrolangoscopia" in texto:
        return "video_rinofibro"
    if "flujometria" in texto or "flujometria urinaria" in texto:
        return "flujometria"
    if "urodinamico" in texto or "urodinamica" in texto or "urodinamico completo" in texto:
        return "estudio_urodinamico"
    if "estudios urologicos" in texto or "estudio urologico" in texto:
        return "estudios_urologicos"
    if "m a p a" in texto or "mapa" in texto:
        return "mapa"
    if "presurometria" in texto:
        return "mapa"
    if "holter" in texto:
        return "holter"
    if "e c g" in texto or "ecg" in texto or "electrocardiograma" in texto:
        return "ecg"
    if "radiografia" in texto or "radiografias" in texto:
        return "rx"
    if texto == "rx" or texto.startswith("rx ") or "radiologia" in texto:
        return "rx"
    if "lesion" in texto:
        return "lesiones"
    if "esclerosante" in texto or "esclerosis" in texto:
        return "tratamiento_esclerosante"
    if "infiltracion" in texto or "infiltraciones" in texto:
        return "infiltraciones"
    if "marcapasos" in texto or "marcapaso" in texto:
        return "control_marcapasos"
    if "doppler" in texto and ("eco stress" in texto or "eco estres" in texto or "ecoestres" in texto or "ecostress" in texto):
        return "doppler_o_ecoestres"
    if "doppler" in texto and (("ecocardio" in texto or "ecocardiograma" in texto) and "stress" in texto):
        return "doppler_o_ecoestres"
    if "diabetologia o gastro" in texto or ("diabetologia" in texto and "gastro" in texto):
        return "diabetologia_o_gastro"
    if ("ecoestres" in texto or "ecostress" in texto or "eco stress" in texto) or (
        ("ecocardio" in texto or "ecocardiograma" in texto) and "stress" in texto
    ):
        return "ecoestres"
    if (
        "ecocardio o vasos" in texto
        or (("ecocardio" in texto or "ecocardiograma" in texto) and "vasos de cuello" in texto)
    ):
        return "ecocardio_o_vasos"
    if "ecocardio" in texto or "ecocardiograma" in texto:
        return "ecocardio"
    if "ergometria" in texto:
        return "ergometria"
    if (
        "doppler o ecoestres" in texto
        or "doppler o eco estres" in texto
        or ("doppler" in texto and ("eco stress" in texto or "eco estres" in texto or "ecoestres" in texto or "ecostress" in texto))
    ):
        return "doppler_o_ecoestres"
    if ("doppler" in texto or "ecodoppler" in texto) and ("cardiaco" in texto or "cardio" in texto):
        return "ecodoppler_cardiaco"
    if "doppler" in texto or "ecodoppler" in texto:
        return "doppler"
    if "espirometria" in texto:
        return "espirometria"
    if "ecografia" in texto and ("renal" in texto or "rinon" in texto or "riñon" in texto):
        return "ecografia_renal"
    if "ecografia" in texto and ("vesical" in texto or "residuo post miccional" in texto or "rpm" in texto):
        return "ecografia_vesical"
    if "ecografia" in texto and ("prostatica" in texto or "prostata" in texto):
        return "ecografia_prostatica"
    if "ecografia" in texto and "testicular" in texto:
        return "ecografia_testicular"
    if "ecografia" in texto and ("vias biliares" in texto or "via biliar" in texto or "biliar" in texto):
        return "ecografia_vias_biliares"
    if "ecografia" in texto and ("transvaginal" in texto or "endocavitaria" in texto or "endocavitaria" in texto):
        return "ecografia_endocavitaria"
    if "ecografia" in texto and ("abdominal" in texto or "abdomen" in texto or "adbdominal" in texto):
        return "ecografia_abdominal"
    if "ecografia" in texto and "mamaria" in texto:
        return "ecografia_mamaria"
    if "ecografia" in texto and "partes blandas" in texto:
        return "ecografia_partes_blandas"
    if "ecografia" in texto and ("musculoesqueletica" in texto or "musculo esqueletica" in texto or "músculo esquelética" in texto):
        return "ecografia_musculoesqueletica"
    if "ecografia" in texto:
        return "ecografia"
    if "mamografia bilateral" in texto:
        return "mamografia_bilateral"
    if texto == "mx" or texto.startswith("mx ") or "mamografia" in texto or "mamaria" in texto:
        return "mamografia"
    if texto == "pap" or texto.startswith("pap ") or " papanicolau" in f" {texto}" or "papanicolau" in texto:
        return "papanicolau"
    if "clinica medica" in texto:
        return "clinica medica"
    if "consulta con especialista en neurologia" in texto:
        return "neurologia"
    if "otorrino" in texto:
        return "otorrinolaringologia"
    for token, base in reemplazos.items():
        if token in texto:
            return base
    return texto


def _base_habilitada_para_sistema(base: str, sistema_turnos: str | None) -> bool:
    base_normalizada = _especialidad_base(base)
    sistema_actual = _sistema_turnos_key(sistema_turnos)
    for sistema, bases in BASES_COMPATIBLES_EXCLUSIVAS_POR_SISTEMA.items():
        if base_normalizada in bases:
            return sistema_actual == sistema
    return True


def _filtrar_especialidades_por_sistema(especialidades: list[str], sistema_turnos: str | None) -> list[str]:
    return [
        especialidad
        for especialidad in especialidades
        if _base_habilitada_para_sistema(especialidad, sistema_turnos)
    ]


def _codigo_practica(valor: str) -> str:
    match = re.search(r"\b(\d{6})\b", str(valor or ""))
    return match.group(1) if match else ""


def _codigo_cubre_requisito(codigo: str, requisito: str) -> bool:
    opciones = {item.strip() for item in str(requisito or "").split("|") if item.strip()}
    return codigo in opciones


def _ome_cubre_requisito(ome: dict, requisito: str, descripcion: str = "") -> bool:
    practica = str(ome.get("practica", "") or "")
    requisitos = [item.strip() for item in str(requisito or "").split("|") if item.strip()]
    for item in requisitos:
        if item.startswith("base:"):
            base = item.split(":", 1)[1]
            if base == "mamografia_bilateral" and _descripcion_cubre_requisito(practica, "MAMOGRAFIA BILATERAL"):
                return True
            if base != "mamografia_bilateral" and _practica_compatible_con_base(practica, base):
                return True
    if _codigo_cubre_requisito(_codigo_practica(practica), requisito):
        return True
    if descripcion:
        return _descripcion_cubre_requisito(practica, descripcion)
    return False


def _requisito_presente(omes_presentes: list[dict], requisito: str, descripcion: str = "") -> bool:
    return any(_ome_cubre_requisito(ome, requisito, descripcion) for ome in omes_presentes)


def _descripcion_cubre_requisito(practica: str, descripcion: str) -> bool:
    practica_key = _clave_texto(practica)
    descripcion_key = _clave_texto(descripcion)
    practica_busqueda = _texto_busqueda(practica)
    descripcion_busqueda = _texto_busqueda(descripcion)
    if "cardiologia" in descripcion_busqueda and "electrocardiograma" in descripcion_busqueda:
        return (
            _codigo_practica(practica) in {"570129", "820113"}
            or ("cardiologia" in practica_busqueda and "electrocardiograma" in practica_busqueda)
        )
    if "holter electroencefalografico" in descripcion_key:
        return "holter" in practica_key and "electroencefal" in practica_key
    if "electroencefalograma" in descripcion_key:
        return "electroencefalograma" in practica_key
    if "electrodiagnostico" in descripcion_key:
        return "electrodiagnostico" in practica_key
    if "mamografia bilateral" in descripcion_key:
        return "mamografia" in practica_key and "bilateral" in practica_key
    if "ecografia mamaria bilateral" in descripcion_key:
        return "ecografia" in practica_key and "mamaria" in practica_key and "bilateral" in practica_key
    if "ecografia de partes blandas" in descripcion_key:
        return "ecografia" in practica_key and "partes blandas" in practica_key
    if "ecografia renal" in descripcion_key:
        return "ecografia" in practica_key and "renal" in practica_key
    if "ecografia vesical" in descripcion_key:
        return "ecografia" in practica_key and (
            "vesical" in practica_key
            or "vesicoprostatica" in practica_key
            or "residuo post miccional" in practica_key
            or "rpm" in practica_key
        )
    if "ecografia prostatica" in descripcion_key:
        return "ecografia" in practica_key and (
            "prostatica" in practica_key
            or "vesicoprostatica" in practica_key
            or "prostata" in practica_key
        )
    if "ecografia testicular" in descripcion_key:
        return "ecografia" in practica_key and "testicular" in practica_key
    if "ecografia de vias biliares" in descripcion_key or "ecografia vias biliares" in descripcion_key:
        return "ecografia" in practica_key and (
            "vias biliares" in practica_key
            or "via biliar" in practica_key
            or "biliar" in practica_key
            or "abdominal" in practica_key
            or "abdomen" in practica_key
        )
    if "ecografia endocavitaria" in descripcion_key or "ecografia transvaginal" in descripcion_key:
        return "ecografia" in practica_key and (
            "endocavitaria" in practica_key
            or "transvaginal" in practica_key
            or "ginecologica" in practica_key
        )
    if "ecografia abdominal" in descripcion_key or "ecografia adbdominal" in descripcion_key:
        return "ecografia" in practica_key and ("abdominal" in practica_key or "abdomen" in practica_key)
    if "ecografia musculo" in descripcion_key or "ecografia musculoesqueletica" in descripcion_key:
        return "ecografia" in practica_key and ("musculo" in practica_key or "musculoesqueletica" in practica_key)
    if "radiografia" in descripcion_key:
        return "radiografia" in practica_key
    if "presurometria" in descripcion_key:
        return "presurometria" in practica_key
    if "logoaudiometria" in descripcion_key:
        return "logoaudiometria" in practica_key
    if "audiometria" in descripcion_key:
        return "audiometria" in practica_key and "logoaudiometria" not in practica_key
    if "timpanometria" in descripcion_key:
        return "timpanometria" in practica_key
    if "acufenometria" in descripcion_key:
        return "acufenometria" in practica_key
    if "impedanciometria" in descripcion_key:
        return "impedanciometria" in practica_key or "impedancia" in practica_key or "inpedancia" in practica_key
    if "lavaje de oido" in descripcion_key or "lavaje de seno" in descripcion_key:
        return (
            "lavaje" in practica_key
            or "extraccion de tapon" in practica_key
            or "tapon de cerumen" in practica_key
            or "cuerpo extra" in practica_key
        )
    if "electrocardiograma" in descripcion_key:
        return (
            "ecg" in practica_key
            or "e c g" in practica_key
            or "electrocardiograma" in practica_key
            or _codigo_practica(practica) == "570129"
        )
    if descripcion_key == "ecografia":
        return "ecografia" in practica_key
    if descripcion_key == "ergometria":
        return "ergometria" in practica_key
    if descripcion_key == "ecocardiograma de stress":
        return (
            "ecoestres" in practica_key
            or "ecostress" in practica_key
            or "eco stress" in practica_key
            or (("ecocardio" in practica_key or "ecocardiograma" in practica_key) and "stress" in practica_key)
        )
    if descripcion_key == "ecodoppler cardiaco":
        return ("doppler" in practica_key or "ecodoppler" in practica_key) and (
            "cardiaco" in practica_key or "cardio" in practica_key
        )
    if descripcion_key == "ecodoppler":
        return "doppler" in practica_key or "ecodoppler" in practica_key
    if descripcion_key == "radiografia":
        return "radiografia" in practica_key
    if descripcion_key == "lesiones":
        return "lesion" in practica_key
    return _practica_compatible_con_base(practica, _especialidad_base(descripcion))


def _practica_compatible_con_base(practica: str, base_esperada: str) -> bool:
    practica_base = _especialidad_base(practica)
    codigo = _codigo_practica(practica)
    if base_esperada == "ecg":
        practica_key = _texto_busqueda(practica)
        return (
            codigo in {"570129", "820113"}
            or ("cardiologia" in practica_key and "electrocardiograma" in practica_key)
        )
    if base_esperada == "ecografia" and practica_base.startswith("ecografia"):
        return True
    if practica_base == base_esperada:
        return True

    regla = PRACTICAS_COMPATIBLES_POR_ESPECIALIDAD.get(base_esperada)
    if not regla:
        return False

    if codigo == "180114" and base_esperada in {"ecografia_vesical", "ecografia_prostatica"}:
        return True
    if base_esperada in {"audiometria", "logoaudiometria", "timpanometria", "acufenometria"} and codigo == "820137":
        return False
    if base_esperada == "fonoaudiologia" and codigo:
        return codigo == "820137"
    if base_esperada == "cardiologia" and codigo in {"570129", "820113"}:
        return True
    if base_esperada == "rx" and codigo.startswith("340"):
        return True
    if codigo and codigo in regla.get("codigos", set()):
        return True
    return practica_base in regla.get("bases", set())


def _ome_compatible_con_turno(
    ome: dict,
    turno: dict,
    mapa_especialidades: dict[str, list[str]],
    sistema_turnos: str = "CIMA",
) -> bool:
    if ome.get("estado_accion") != "disponible_activar":
        return False
    return _ome_relacionada_con_turno(ome, turno, mapa_especialidades, sistema_turnos)


def _ome_relacionada_con_turno(
    ome: dict,
    turno: dict,
    mapa_especialidades: dict[str, list[str]],
    sistema_turnos: str = "CIMA",
) -> bool:
    if _ome_bloqueada_por_combo_cima(ome, turno, mapa_especialidades, sistema_turnos):
        return False

    sistema_key = _sistema_turnos_key(sistema_turnos)
    practica = str(ome.get("practica", "") or "")
    profesional = str(turno.get("profesional", "") or "")
    codigo_ome = _codigo_practica(practica)
    codigo_turno = _codigo_practica(profesional)
    if codigo_ome and codigo_turno and codigo_ome == codigo_turno:
        return True
    if codigo_turno:
        return False
    esperadas = _especialidades_turno(turno, mapa_especialidades, sistema_key)
    practica_base = _especialidad_base(practica)
    bases_esperadas = {_especialidad_base(esperada) for esperada in esperadas}
    if sistema_key == "CIMA" and "ecg" in bases_esperadas and codigo_ome == "570126":
        return True
    if "fonoaudiologia" in bases_esperadas and practica_base == "fono_estudios":
        return True
    if _turno_cima_ramirez_walter_ambiguo(turno, sistema_key) and practica_base in {
        "doppler",
        "ecodoppler_cardiaco",
        "ecocardio",
        "ecoestres",
    }:
        return True
    if _turno_cima_novelli_dario_ambiguo(turno, sistema_key) and _ome_base_novelli_dario_generico(practica_base):
        return True
    if _turno_diabetologia_perales(turno, sistema_key) and _practica_compatible_con_base(practica, "cardiologia"):
        return True
    if sistema_key == "GJS":
        requisitos = _ajustar_requisitos_por_sistema(
            _requisitos_practicas_turno(turno, mapa_especialidades, sistema_key),
            sistema_key,
        )
        if any(_ome_cubre_requisito(ome, codigo, descripcion) for codigo, descripcion in requisitos.items()):
            return True
    return any(_practica_compatible_con_base(practica, _especialidad_base(esperada)) for esperada in esperadas)


def _ome_relacionada_con_turnos(
    ome: dict,
    turnos: list[dict],
    mapa_especialidades: dict[str, list[str]],
    sistema_turnos: str = "CIMA",
) -> bool:
    return any(_ome_relacionada_con_turno(ome, turno, mapa_especialidades, sistema_turnos) for turno in turnos)


def _bases_esperadas_turno(
    turno: dict,
    mapa_especialidades: dict[str, list[str]],
    sistema_turnos: str = "CIMA",
) -> set[str]:
    return {
        _especialidad_base(esperada)
        for esperada in _especialidades_turno(turno, mapa_especialidades, sistema_turnos)
    }


def _turno_indica_pap(turno: dict) -> bool:
    texto = _clave_texto(turno.get("profesional", ""))
    return bool(re.search(r"\bpap\b", texto)) or "papanicolau" in texto


def _turno_no_reportable(turno: dict, sistema_turnos: str = "CIMA") -> bool:
    if _sistema_turnos_key(sistema_turnos) != "CIMA":
        return False
    texto = _clave_texto(turno.get("profesional", ""))
    if "pap dim" in texto:
        return True
    if _es_turno_clinica_medica_generico(texto):
        return True
    if ("clinica medica" in texto or "clinica médica" in texto) and (
        "dube" in texto or "dubesarsky" in texto or re.search(r"\bmc\b", texto)
    ):
        return True
    return False


def _turno_diabetologia_perales(turno: dict, sistema_turnos: str = "CIMA") -> bool:
    if _sistema_turnos_key(sistema_turnos) != "CIMA":
        return False
    profesional_raw = str(turno.get("profesional", "") or "")
    texto = _clave_texto(profesional_raw)
    if "perales" in texto and "diabetologia" in texto:
        return True
    if " - " in profesional_raw:
        return False
    return texto == "diabetologia"


def _turno_perales_especialidad_alternativa(
    turno: dict,
    mapa_especialidades: dict[str, list[str]],
    sistema_turnos: str = "CIMA",
) -> bool:
    if _sistema_turnos_key(sistema_turnos) != "CIMA":
        return False
    texto = _clave_texto(turno.get("profesional", ""))
    if "perales" not in texto:
        return False
    bases = {
        _especialidad_base(especialidad)
        for especialidad in _especialidades_turno(turno, mapa_especialidades, sistema_turnos)
    }
    return bool(bases & PERALES_CIMA_BASES_ALTERNATIVAS)


def _turno_cima_ramirez_walter_ambiguo(turno: dict, sistema_turnos: str = "CIMA") -> bool:
    if _sistema_turnos_key(sistema_turnos) != "CIMA":
        return False
    texto = _clave_texto((turno or {}).get("profesional", ""))
    texto = PROFESIONAL_ALIASES.get(texto, texto)
    return texto == "ramirez walter"


def _turno_cima_novelli_dario_ambiguo(turno: dict, sistema_turnos: str = "CIMA") -> bool:
    if _sistema_turnos_key(sistema_turnos) != "CIMA":
        return False
    texto = _clave_texto((turno or {}).get("profesional", ""))
    texto = PROFESIONAL_ALIASES.get(texto, texto)
    return texto == "novelli dario"


def _ome_base_novelli_dario_generico(practica_base: str) -> str:
    if str(practica_base or "").startswith("ecografia"):
        return "ecografia"
    if practica_base in {"doppler", "ecodoppler_cardiaco"}:
        return "doppler"
    return ""


def _turno_cima_profesional_ambiguo_sin_requisito(turno: dict, sistema_turnos: str = "CIMA") -> bool:
    return _turno_cima_ramirez_walter_ambiguo(
        turno,
        sistema_turnos,
    ) or _turno_cima_novelli_dario_ambiguo(
        turno,
        sistema_turnos,
    )


def _especialidades_turno(
    turno: dict,
    mapa_especialidades: dict[str, list[str]],
    sistema_turnos: str = "CIMA",
) -> list[str]:
    sistema_key = _sistema_turnos_key(sistema_turnos)
    profesional = str(turno.get("profesional", "") or "")
    especialidades = _especialidades_profesional(profesional, mapa_especialidades, sistema_key)
    especialidades = _filtrar_especialidades_por_sistema(especialidades, sistema_key)
    if _turno_indica_pap(turno):
        return especialidades
    bases = {_especialidad_base(especialidad) for especialidad in especialidades}
    if "ginecologia" in bases and "papanicolau" in bases:
        return [especialidad for especialidad in especialidades if _especialidad_base(especialidad) != "papanicolau"]
    return especialidades


def _requisitos_practicas_turno(
    turno: dict,
    mapa_especialidades: dict[str, list[str]],
    sistema_turnos: str = "CIMA",
) -> dict[str, str]:
    sistema_key = _sistema_turnos_key(sistema_turnos)
    if _turno_perales_especialidad_alternativa(turno, mapa_especialidades, sistema_key):
        return {PERALES_CIMA_REQUISITO_ALTERNATIVO: PERALES_CIMA_DESCRIPCION_ALTERNATIVA}
    requisitos: dict[str, str] = {}
    for base in _bases_esperadas_turno(turno, mapa_especialidades, sistema_key):
        requisitos.update(PRACTICAS_REQUERIDAS_POR_ESPECIALIDAD.get(base, {}))
    profesional = _clave_texto(turno.get("profesional", ""))
    profesional = PROFESIONAL_ALIASES.get(profesional, profesional)
    if sistema_key == "CIMA":
        requisitos.update(PRACTICAS_REQUERIDAS_POR_PROFESIONAL.get(profesional, {}))
    if _turno_diabetologia_perales(turno, sistema_key) and "base:diabetologia" in requisitos:
        descripcion = requisitos.pop("base:diabetologia")
        requisitos["base:diabetologia|base:cardiologia|570129|820113"] = descripcion
    if _turno_diabetologia_perales(turno, sistema_key) and "base:dbt_nefro_nutri" in requisitos:
        descripcion = requisitos.pop("base:dbt_nefro_nutri")
        requisitos["base:dbt_nefro_nutri|base:cardiologia|570129|820113"] = descripcion
    return requisitos


def _ajustar_requisitos_por_sistema(requisitos: dict[str, str], sistema_turnos: str) -> dict[str, str]:
    if _sistema_turnos_key(sistema_turnos) != "GJS" or not requisitos:
        return requisitos

    ajustados: dict[str, str] = {}
    for codigo, descripcion in requisitos.items():
        opciones = {item.strip() for item in str(codigo or "").split("|") if item.strip()}
        if "base:ecodoppler_cardiaco" in opciones:
            opciones.update({"180606", "180607", "180610"})
            codigo = "|".join(item for item in str(codigo or "").split("|") if item.strip())
            codigo = f"{codigo}|180606|180607|180610" if codigo else "180606|180607|180610"
            descripcion = "DOPPLER"
        ajustados[codigo] = descripcion
    return ajustados


def _grupo_profesional_gjs(turno: dict) -> str:
    partes = [parte.strip() for parte in str(turno.get("profesional", "") or "").split(" - ") if parte.strip()]
    return _clave_texto(" - ".join(partes[:2] if len(partes) >= 2 else partes))


def _quitar_requisitos_duplicados_por_bono_gjs(
    requisitos: dict[str, str],
    turno: dict,
    turnos: list[dict],
    mapa_especialidades: dict[str, list[str]],
    sistema_turnos: str,
) -> dict[str, str]:
    sistema_key = _sistema_turnos_key(sistema_turnos)
    if sistema_key != "GJS" or not requisitos:
        return requisitos

    fecha_key = turno.get("fecha_key") or _fecha_key(turno.get("fecha"))
    grupo = _grupo_profesional_gjs(turno)
    bases_otros: set[str] = set()
    for otro in turnos:
        if otro is turno:
            continue
        if (otro.get("fecha_key") or _fecha_key(otro.get("fecha"))) != fecha_key:
            continue
        if _grupo_profesional_gjs(otro) != grupo:
            continue
        bases_otros.update(_bases_esperadas_turno(otro, mapa_especialidades, sistema_key))

    if not bases_otros:
        return requisitos

    requerimientos_por_base = {
        "audiometria": {"717150"},
        "logoaudiometria": {"717151"},
        "timpanometria": {"717156"},
        "acufenometria": {"717157"},
        "impedanciometria": {"717155"},
    }
    bases_actuales = _bases_esperadas_turno(turno, mapa_especialidades, sistema_key)
    if "fonoaudiologia" not in bases_actuales:
        requerimientos_por_base["fonoaudiologia"] = {"820137", "base:fonoaudiologia"}
    codigos_a_quitar = set().union(*(requerimientos_por_base.get(base, set()) for base in bases_otros))
    if not codigos_a_quitar:
        return requisitos

    filtrados: dict[str, str] = {}
    for codigo, descripcion in requisitos.items():
        opciones = {item.strip() for item in str(codigo).split("|") if item.strip()}
        if opciones & codigos_a_quitar:
            continue
        filtrados[codigo] = descripcion
    return filtrados


def _omitir_validacion_combo_cima(
    requisitos: dict[str, str],
    turno: dict,
    turnos: list[dict],
    mapa_especialidades: dict[str, list[str]],
    sistema_turnos: str,
) -> dict[str, str]:
    if _sistema_turnos_key(sistema_turnos) != "CIMA" or not requisitos:
        return requisitos

    fecha_key = turno.get("fecha_key") or _fecha_key(turno.get("fecha"))
    bases_actuales = _bases_turno_para_combo(turno, mapa_especialidades, sistema_turnos)
    if not fecha_key or not bases_actuales:
        return requisitos

    bases_misma_fecha: set[str] = set()
    for otro in turnos:
        if otro is turno:
            continue
        if (otro.get("fecha_key") or _fecha_key(otro.get("fecha"))) != fecha_key:
            continue
        bases_misma_fecha.update(_bases_turno_para_combo(otro, mapa_especialidades, sistema_turnos))

    omitir: set[str] = set()
    for combo, regla in CIMA_COMBOS_VALIDACION_CONFIG.items():
        if _combo_verificar_habilitado(combo):
            continue
        bases_regla = set(regla.get("bases", set()))
        if bases_regla and bases_actuales & bases_regla:
            omitir.update(regla.get("omitir", set()))
            continue

        componentes = set(regla.get("componentes", set()))
        if componentes:
            bases_actuales_combo = bases_actuales & componentes
            bases_otros_combo = bases_misma_fecha & (componentes - bases_actuales_combo)
            opciones_combo = _opciones_requisitos_presentes(requisitos, set(regla.get("omitir", set())))
            if (
                (bases_actuales_combo and bases_otros_combo)
                or len(bases_actuales_combo) >= 2
                or len(opciones_combo) >= 2
            ):
                omitir.update(regla.get("omitir", set()))
            continue

        principales = set(regla.get("principales", set()))
        companeras = set(regla.get("companeras", set()))
        if bases_actuales & companeras and (bases_misma_fecha & principales or bases_actuales & principales):
            omitir.update(regla.get("omitir_companera", set()))

    if not omitir:
        return requisitos

    return {
        codigo: descripcion
        for codigo, descripcion in requisitos.items()
        if not _requisito_tiene_opcion(codigo, omitir)
    }


def _turno_omitido_por_combo_cima(
    turno: dict,
    turnos: list[dict],
    mapa_especialidades: dict[str, list[str]],
    sistema_turnos: str,
    requisitos: dict[str, str] | None = None,
) -> bool:
    if _sistema_turnos_key(sistema_turnos) != "CIMA":
        return False

    fecha_key = turno.get("fecha_key") or _fecha_key(turno.get("fecha"))
    bases_actuales = _bases_turno_para_combo(turno, mapa_especialidades, sistema_turnos)
    if not fecha_key or not bases_actuales:
        return False

    bases_misma_fecha: set[str] = set()
    for otro in turnos:
        if otro is turno:
            continue
        if (otro.get("fecha_key") or _fecha_key(otro.get("fecha"))) != fecha_key:
            continue
        bases_misma_fecha.update(_bases_turno_para_combo(otro, mapa_especialidades, sistema_turnos))

    for combo, regla in CIMA_COMBOS_VALIDACION_CONFIG.items():
        if _combo_verificar_habilitado(combo):
            continue
        componentes = set(regla.get("componentes", set()))
        if componentes:
            bases_actuales_combo = bases_actuales & componentes
            opciones_combo = _opciones_requisitos_presentes(requisitos or {}, set(regla.get("omitir", set())))
            if (
                (bases_actuales_combo and bases_misma_fecha & (componentes - bases_actuales_combo))
                or len(bases_actuales_combo) >= 2
                or len(opciones_combo) >= 2
            ):
                return True
            continue
        principales = set(regla.get("principales", set()))
        companeras = set(regla.get("companeras", set()))
        if bases_actuales & companeras and bases_misma_fecha & principales:
            return True
    return False


def _combo_verificar_habilitado(combo: str) -> bool:
    try:
        return bool(load_verificar_omes_config().get("combos", {}).get(combo, True))
    except Exception:
        return True


def _turno_cima_mamografia_combo_desactivado(
    turno: dict,
    mapa_especialidades: dict[str, list[str]],
    sistema_turnos: str,
) -> bool:
    if _sistema_turnos_key(sistema_turnos) != "CIMA":
        return False
    if _combo_verificar_habilitado("mamografia_eco_mamaria"):
        return False
    return "mamografia" in _bases_turno_para_combo(turno, mapa_especialidades, sistema_turnos)


def _opciones_requisitos_presentes(requisitos: dict[str, str], opciones: set[str]) -> set[str]:
    if not requisitos or not opciones:
        return set()
    presentes: set[str] = set()
    for codigo in requisitos:
        partes = {item.strip() for item in str(codigo or "").split("|") if item.strip()}
        presentes.update(partes & opciones)
    return presentes


def _ome_bloqueada_por_combo_cima(
    ome: dict,
    turno: dict,
    mapa_especialidades: dict[str, list[str]],
    sistema_turnos: str,
) -> bool:
    if _sistema_turnos_key(sistema_turnos) != "CIMA":
        return False
    if _combo_verificar_habilitado("espiro_neumo"):
        return False
    bases = _bases_esperadas_turno(turno, mapa_especialidades, sistema_turnos)
    if "espirometria" not in bases:
        return False
    practica = str((ome or {}).get("practica", "") or "")
    practica_key = _texto_busqueda(practica)
    return "consulta" in practica_key and _especialidad_base(practica) == "neumonologia"


def _bases_turno_para_combo(
    turno: dict,
    mapa_especialidades: dict[str, list[str]],
    sistema_turnos: str = "CIMA",
) -> set[str]:
    bases = {
        _especialidad_base(esperada)
        for esperada in _especialidades_turno(turno, mapa_especialidades, sistema_turnos)
    }
    texto = _clave_texto(turno.get("profesional", ""))
    if texto == "mx" or texto.startswith("mx ") or "mamografia" in texto or "mamaria" in texto:
        bases.add("mamografia")
    if "frotis" in texto:
        bases.add("frotis")
    return bases


def _requisito_tiene_opcion(codigo: str, opciones: set[str]) -> bool:
    partes = {item.strip() for item in str(codigo or "").split("|") if item.strip()}
    return bool(partes & opciones)


def _requisitos_practicas_efectivos(
    turno: dict,
    mapa_especialidades: dict[str, list[str]],
    omes: list[dict],
    fecha_turno_key: str,
    omes_usadas: set[int],
    sistema_turnos: str = "CIMA",
) -> dict[str, str]:
    requisitos = _ajustar_requisitos_por_sistema(
        _requisitos_practicas_turno(turno, mapa_especialidades, sistema_turnos),
        sistema_turnos,
    )
    bases = _bases_esperadas_turno(turno, mapa_especialidades, sistema_turnos)

    for base in bases:
        condicionales = PRACTICAS_CONDICIONALES_POR_ESPECIALIDAD.get(base, {})
        if not condicionales:
            continue

        hay_parte_del_paquete = False
        for idx, ome in enumerate(omes):
            if idx in omes_usadas:
                continue
            if not any(
                _ome_cubre_requisito(ome, requisito, descripcion)
                for requisito, descripcion in condicionales.items()
            ):
                continue
            if _fecha_key(ome.get("f_agenda", "")) == fecha_turno_key or ome.get("estado_accion") == "disponible_activar":
                hay_parte_del_paquete = True
                break

        if hay_parte_del_paquete:
            requisitos.update(condicionales)

    return requisitos


def _paquete_key_turno(
    turno: dict,
    mapa_especialidades: dict[str, list[str]],
    sistema_turnos: str = "CIMA",
) -> tuple:
    fecha_key = turno.get("fecha_key") or _fecha_key(turno.get("fecha"))
    if _turno_perales_especialidad_alternativa(turno, mapa_especialidades, sistema_turnos):
        return (fecha_key, "perales_especialidades_alternativas")
    bases = _bases_esperadas_turno(turno, mapa_especialidades, sistema_turnos)
    if "fono_estudios" in bases:
        return (fecha_key, "fono_estudios")
    if "lavaje_oido" in bases:
        return (fecha_key, "lavaje_oido")
    return (fecha_key, _clave_texto(turno.get("profesional", "")))


def _ome_faltante(codigo: str, descripcion: str) -> dict:
    codigo_visible = str(codigo).replace("base:", "").replace("|", " / ")
    descripcion_limpia = str(descripcion or "").strip()
    prefijo = f"{codigo_visible} - " if codigo_visible and descripcion_limpia.upper() != "DOPPLER" else ""
    return {
        "n_orden": "",
        "practica": f"FALTA OME: {prefijo}{descripcion_limpia}",
        "f_vencimiento": "",
        "f_agenda": "",
        "mismo_efector": True,
        "estado_accion": "faltante",
    }


def _turno_con_practica_faltante(turno: dict, descripcion: str, sistema_turnos: str) -> dict:
    if not turno:
        return turno
    copia = dict(turno)
    profesional = str(copia.get("profesional", "") or "")
    partes = [parte.strip() for parte in profesional.split(" - ") if parte.strip()]
    descripcion_limpia = str(descripcion or "").strip()
    if descripcion_limpia and len(partes) >= 3:
        partes[-1] = descripcion_limpia
        copia["profesional"] = " - ".join(partes)
    elif descripcion_limpia:
        copia["profesional"] = f"{profesional} - {descripcion_limpia}" if profesional else descripcion_limpia
    return copia


def _turno_con_practica_ome(
    turno: dict,
    ome: dict,
    requisitos: dict[str, str],
    sistema_turnos: str,
) -> dict:
    if not turno or not requisitos:
        return turno
    for codigo, descripcion in requisitos.items():
        if _ome_cubre_requisito(ome, codigo, descripcion):
            return _turno_con_practica_faltante(turno, descripcion, sistema_turnos)
    return turno


def _ome_fecha_no_coincide(ome: dict) -> dict:
    copia = dict(ome)
    copia["estado_accion"] = "fecha_no_coincide"
    return copia


def _ome_fecha_auditoria_ok(ome: dict) -> dict:
    copia = dict(ome)
    copia["estado_accion"] = "fecha_auditoria_ok"
    return copia


def _ome_fecha_mismo_mes(ome: dict) -> dict:
    copia = dict(ome)
    copia["estado_accion"] = "fecha_mismo_mes"
    return copia


def _ome_practica_relacionada(ome: dict) -> dict:
    copia = dict(ome)
    copia["estado_accion"] = "practica_relacionada"
    return copia


def _ome_sustitucion_especialista(ome: dict) -> dict:
    copia = dict(ome)
    copia["estado_accion"] = "especialidad_sustituida"
    return copia


def _turno_admite_sustitucion_especialista(
    turno: dict,
    mapa_especialidades: dict[str, list[str]],
    sistema_turnos: str = "CIMA",
) -> bool:
    bases = _bases_esperadas_turno(turno, mapa_especialidades, sistema_turnos)
    return bool(bases & BASES_TURNOS_CON_SUSTITUCION_ESPECIALISTA)


def _ome_puede_sustituir_especialidad(
    ome: dict,
    fecha_turno_key: str,
    sistema_turnos: str = "CIMA",
) -> bool:
    if not isinstance(ome, dict):
        return False
    if _sistema_turnos_key(sistema_turnos) != "CIMA":
        return False
    if not _ome_auditoria_transmitida(ome):
        return False
    if ome.get("estado_accion") in {"disponible_activar", "faltante", "sin_resultados_pami"}:
        return False
    if not ome.get("mismo_efector"):
        return False
    if _fecha_key(ome.get("f_agenda", "")) != fecha_turno_key:
        return False
    practica = str(ome.get("practica", "") or "")
    practica_key = _texto_busqueda(practica)
    if "consulta" not in practica_key:
        return False
    return _especialidad_base(practica) in BASES_OMES_SUSTITUTAS_ESPECIALISTA


def _ome_sustituta_reservada_por_otro_turno(
    ome: dict,
    turno_actual: dict,
    turnos: list[dict],
    mapa_especialidades: dict[str, list[str]],
    sistema_turnos: str = "CIMA",
) -> bool:
    fecha_ome = _fecha_key(ome.get("f_agenda", ""))
    if not fecha_ome:
        return False
    for turno in turnos:
        if turno is turno_actual:
            continue
        if _fecha_key(turno.get("fecha")) != fecha_ome:
            continue
        if _ome_relacionada_con_turno(ome, turno, mapa_especialidades, sistema_turnos):
            return True
        requisitos = _requisitos_practicas_turno(turno, mapa_especialidades, sistema_turnos)
        if any(_ome_cubre_requisito(ome, codigo, descripcion) for codigo, descripcion in requisitos.items()):
            return True
    return False


def _buscar_ome_sustituta_misma_fecha(
    omes: list[dict],
    usadas: set[int],
    turno: dict,
    turnos: list[dict],
    mapa_especialidades: dict[str, list[str]],
    sistema_turnos: str = "CIMA",
) -> tuple[int, dict] | None:
    if not _turno_admite_sustitucion_especialista(turno, mapa_especialidades, sistema_turnos):
        return None
    fecha_turno_key = turno.get("fecha_key") or _fecha_key(turno.get("fecha"))
    for idx, ome in enumerate(omes):
        if idx in usadas:
            continue
        if not _ome_puede_sustituir_especialidad(ome, fecha_turno_key, sistema_turnos):
            continue
        if _ome_sustituta_reservada_por_otro_turno(ome, turno, turnos, mapa_especialidades, sistema_turnos):
            continue
        return idx, ome
    return None


def _ome_sin_resultados_pami() -> dict:
    return {
        "n_orden": "",
        "practica": "SIN OMES EN PAMI PARA ESTE PACIENTE",
        "f_vencimiento": "",
        "f_agenda": "",
        "mismo_efector": True,
        "estado_accion": "sin_resultados_pami",
    }


def _ome_reutilizable_en_reporte(ome: dict) -> bool:
    practica = str((ome or {}).get("practica", "") or "")
    practica_key = _texto_busqueda(practica)
    if (ome or {}).get("estado_accion") == "disponible_activar":
        return True
    if _codigo_practica(practica) in {"570129", "820113", "180106", "186001", "180114"}:
        return True
    return (
        "cardiologia" in practica_key
        and ("consulta" in practica_key or "electrocardiograma" in practica_key)
    )


def _turno_cardio_ecg_gjs(
    turno: dict,
    mapa_especialidades: dict[str, list[str]],
    sistema_turnos: str = "GJS",
) -> bool:
    if _sistema_turnos_key(sistema_turnos) != "GJS":
        return False
    return bool({"cardiologia", "ecg"} & _bases_esperadas_turno(turno, mapa_especialidades, sistema_turnos))


def _ome_cubre_combo_cardio_gjs(
    ome: dict,
    turno: dict,
    turnos: list[dict],
    mapa_especialidades: dict[str, list[str]],
    sistema_turnos: str,
) -> bool:
    sistema_key = _sistema_turnos_key(sistema_turnos)
    if sistema_key != "GJS" or not _ome_reutilizable_en_reporte(ome):
        return False
    if not _turno_cardio_ecg_gjs(turno, mapa_especialidades, sistema_key):
        return False

    fecha_ome = _fecha_key(ome.get("f_agenda", ""))
    if not fecha_ome:
        return False
    return any(
        _turno_cardio_ecg_gjs(otro_turno, mapa_especialidades, sistema_key)
        and _fecha_key(otro_turno.get("fecha")) == fecha_ome
        for otro_turno in turnos
    )


def _ome_reservada_para_turno_exacto(
    ome: dict,
    turno_actual: dict,
    turnos: list[dict],
    mapa_especialidades: dict[str, list[str]],
    codigo: str = "",
    descripcion: str = "",
    sistema_turnos: str = "CIMA",
) -> bool:
    fecha_ome = _fecha_key(ome.get("f_agenda", ""))
    if not fecha_ome:
        return False
    for turno in turnos:
        if turno is turno_actual:
            continue
        if _fecha_key(turno.get("fecha")) != fecha_ome:
            continue
        if codigo and _ome_cubre_requisito(ome, codigo, descripcion):
            return True
        if _ome_relacionada_con_turno(ome, turno, mapa_especialidades, sistema_turnos):
            return True
    return False


def _ome_relacionada_por_codigo_para_revisar(ome: dict, codigo: str, descripcion: str) -> bool:
    practica = str((ome or {}).get("practica", "") or "")
    codigo_ome = _codigo_practica(practica)
    descripcion_key = _clave_texto(descripcion)
    practica_key = _clave_texto(practica)
    return (
        codigo_ome == "186001"
        and "ecografia de partes blandas" in descripcion_key
        and "musculoesqueletica" in practica_key
    )


def generar_reporte(
    pacientes: list[dict],
    omes_por_key: dict[str, list],
    ruta_salida: Path,
    cliente_codigo: str | None = CLIENTE_CIMA,
    modo_verificacion: str = "futuro",
) -> Path:
    if (modo_verificacion or "").strip().lower() == "auditoria":
        return generar_reporte_auditoria(pacientes, omes_por_key, ruta_salida, cliente_codigo=cliente_codigo)

    cliente = _cliente_config(cliente_codigo)
    sistema_turnos = cliente.get("sistema_turnos", "CIMA")
    ruta_salida = Path(ruta_salida)
    if not ruta_salida.suffix:
        ruta_salida = ruta_salida.with_suffix(".xlsx")
    ruta_salida.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws._sistema_turnos = sistema_turnos
    ws.title = "Verificacion OMEs"

    azul_oscuro = PatternFill("solid", fgColor="1F4E79")
    azul = PatternFill("solid", fgColor="2E75B6")
    verde = PatternFill("solid", fgColor="C6EFCE")
    celeste = PatternFill("solid", fgColor="DDEBF7")
    naranja = PatternFill("solid", fgColor="FFE699")
    amarillo = PatternFill("solid", fgColor="FFF2CC")
    rojo = PatternFill("solid", fgColor="FFC7CE")
    gris = PatternFill("solid", fgColor="F2F2F2")
    fill_disponible = rojo if sistema_turnos == "GJS" else celeste
    fill_otro_prestador = rojo if sistema_turnos == "GJS" else naranja
    blanco = Font(color="FFFFFF", bold=True)
    borde = Border(*(Side(style="thin", color="D9E2F3") for _ in range(4)))

    columnas = _columnas_reporte(sistema_turnos)
    ws.merge_cells("A1:A2")
    ws.merge_cells("B1:B2")
    ws["A1"] = "PACIENTE"
    ws["B1"] = _titulo_identificador_reporte(sistema_turnos)
    if _es_sistema_cima(sistema_turnos):
        ws.merge_cells("C1:C2")
        ws.merge_cells("D1:E1")
        ws.merge_cells("F1:H1")
        ws["C1"] = "NRO. TRAMITE"
        ws["D1"] = f"SISTEMA {sistema_turnos}"
        ws["F1"] = "SISTEMA PAMI"
        headers = ["FECHA", "ESPECIALISTA", "FECHA TURNO", "VENCIMIENTO", "ESPECIALIDAD/CODIGO"]
        for col, value in enumerate(headers, start=4):
            ws.cell(row=2, column=col, value=value)
    else:
        ws.merge_cells("C1:D1")
        ws.merge_cells("E1:G1")
        ws["C1"] = f"SISTEMA {sistema_turnos}"
        ws["E1"] = "SISTEMA PAMI"
        headers = ["FECHA", "ESPECIALISTA", "FECHA TURNO", "VENCIMIENTO", "ESPECIALIDAD/CODIGO"]
        for col, value in enumerate(headers, start=3):
            ws.cell(row=2, column=col, value=value)

    for row in (1, 2):
        for col in range(1, columnas["total"] + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = azul_oscuro if row == 1 else azul
            cell.font = blanco
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = borde

    mapa_especialidades = cargar_especialidades_medicos()
    current_row = 3
    for paciente in pacientes:
        if _paciente_excluido_reporte_no_pami(paciente, sistema_turnos):
            continue
        start_row = current_row
        key = paciente.get("key") or paciente.get("beneficiario") or f"DNI_{paciente.get('dni', '')}"
        nombre = paciente.get("nombre") or key
        identificador = _identificador_paciente(paciente)
        turnos = [turno for turno in paciente.get("turnos", []) if not _turno_no_reportable(turno)]
        if not turnos:
            continue
        omes = list(omes_por_key.get(key, []))
        omes_usadas: set[int] = set()
        paquetes_procesados: set[tuple] = set()
        requisitos_cubiertos: set[tuple[str, str]] = set()
        turnos_omitidos_sin_faltante = False

        for turno in turnos:
            fecha_turno_key = turno.get("fecha_key") or _fecha_key(turno.get("fecha"))
            if not omes:
                requisitos_sin_omes = _requisitos_practicas_efectivos(
                    turno,
                    mapa_especialidades,
                    omes,
                    fecha_turno_key,
                    omes_usadas,
                    sistema_turnos,
                )
                requisitos_sin_omes = _quitar_requisitos_duplicados_por_bono_gjs(
                    requisitos_sin_omes,
                    turno,
                    turnos,
                    mapa_especialidades,
                    sistema_turnos,
                )
                requisitos_sin_omes = _omitir_validacion_combo_cima(
                    requisitos_sin_omes,
                    turno,
                    turnos,
                    mapa_especialidades,
                    sistema_turnos,
                )
                if not requisitos_sin_omes:
                    if _turno_cima_mamografia_combo_desactivado(turno, mapa_especialidades, sistema_turnos):
                        turnos_omitidos_sin_faltante = True
                    elif _turno_cima_novelli_dario_ambiguo(turno, sistema_turnos):
                        _append_reporte_row(
                            ws,
                            current_row,
                            nombre,
                            identificador,
                            _turno_con_practica_faltante(turno, "ECOGRAFIA / ECODOPPLER", sistema_turnos),
                            _ome_faltante("", "ECOGRAFIA / ECODOPPLER"),
                            rojo,
                            borde,
                        )
                        current_row += 1
                    elif _turno_cima_profesional_ambiguo_sin_requisito(turno, sistema_turnos):
                        turnos_omitidos_sin_faltante = True
                    continue
                _append_reporte_row(
                    ws,
                    current_row,
                    nombre,
                    identificador,
                    turno,
                    _ome_sin_resultados_pami(),
                    rojo,
                    borde,
                )
                current_row += 1
                continue
            requisitos = _requisitos_practicas_efectivos(
                turno,
                mapa_especialidades,
                omes,
                fecha_turno_key,
                omes_usadas,
                sistema_turnos,
            )
            requisitos = _quitar_requisitos_duplicados_por_bono_gjs(
                requisitos,
                turno,
                turnos,
                mapa_especialidades,
                sistema_turnos,
            )
            requisitos_antes_combo = dict(requisitos)
            requisitos = _omitir_validacion_combo_cima(
                requisitos,
                turno,
                turnos,
                mapa_especialidades,
                sistema_turnos,
            )
            if (requisitos_antes_combo and not requisitos) or _turno_omitido_por_combo_cima(
                turno,
                turnos,
                mapa_especialidades,
                sistema_turnos,
                requisitos_antes_combo,
            ):
                turnos_omitidos_sin_faltante = True
                continue
            if requisitos:
                paquete_key = _paquete_key_turno(turno, mapa_especialidades)
                emitir_faltantes = paquete_key not in paquetes_procesados
                paquetes_procesados.add(paquete_key)

                presentes = [
                    (idx, ome)
                    for idx, ome in enumerate(omes)
                    if (idx not in omes_usadas or _ome_reutilizable_en_reporte(ome))
                    and not _ome_bloqueada_por_combo_cima(ome, turno, mapa_especialidades, sistema_turnos)
                    and any(_ome_cubre_requisito(ome, requisito, descripcion) for requisito, descripcion in requisitos.items())
                    and (
                        _fecha_key(ome.get("f_agenda", "")) == fecha_turno_key
                        or ome.get("estado_accion") == "disponible_activar"
                        or _ome_cubre_combo_cardio_gjs(ome, turno, turnos, mapa_especialidades, sistema_turnos)
                        or (
                            ome.get("mismo_efector")
                            and _fecha_key(ome.get("f_agenda", "")) not in ("", fecha_turno_key)
                            and _mismo_mes(ome.get("f_agenda", ""), fecha_turno_key)
                        )
                    )
                ]
                presentes_filtrados = []
                sistema_es_gjs = _sistema_turnos_key(sistema_turnos) == "GJS"
                for idx, ome in presentes:
                    fecha_ome_key = _fecha_key(ome.get("f_agenda", ""))
                    if ome.get("estado_accion") != "disponible_activar" and fecha_ome_key not in ("", fecha_turno_key):
                        disponible_mismo_requisito = any(
                            otro_idx != idx
                            and otra.get("estado_accion") == "disponible_activar"
                            and any(
                                _ome_cubre_requisito(ome, requisito, descripcion)
                                and _ome_cubre_requisito(otra, requisito, descripcion)
                                for requisito, descripcion in requisitos.items()
                            )
                            for otro_idx, otra in presentes
                        )
                        if disponible_mismo_requisito and not sistema_es_gjs:
                            continue
                    if ome.get("estado_accion") == "disponible_activar":
                        cubierta_por_activa = any(
                            otro_idx != idx
                            and otra.get("estado_accion") != "disponible_activar"
                            and (
                                _fecha_key(otra.get("f_agenda", "")) == fecha_turno_key
                                or (
                                    sistema_es_gjs
                                    and _fecha_key(otra.get("f_agenda", ""))
                                    and _mismo_mes(otra.get("f_agenda", ""), fecha_turno_key)
                                )
                            )
                            and any(
                                _ome_cubre_requisito(ome, requisito, descripcion)
                                and _ome_cubre_requisito(otra, requisito, descripcion)
                                for requisito, descripcion in requisitos.items()
                            )
                            for otro_idx, otra in presentes
                        )
                        if cubierta_por_activa:
                            omes_usadas.add(idx)
                            continue
                    presentes_filtrados.append((idx, ome))
                presentes = presentes_filtrados
                omes_presentes = [ome for _, ome in presentes]
                for idx, ome in presentes:
                    omes_usadas.add(idx)
                    for codigo, descripcion in requisitos.items():
                        if _ome_cubre_requisito(ome, codigo, descripcion):
                            requisitos_cubiertos.add((codigo, descripcion))
                    fecha_ome_key = _fecha_key(ome.get("f_agenda", ""))
                    ome_reporte = (
                        _ome_fecha_mismo_mes(ome)
                        if ome.get("mismo_efector")
                        and fecha_ome_key not in ("", fecha_turno_key)
                        and _mismo_mes(fecha_ome_key, fecha_turno_key)
                        else ome
                    )
                    fill = (
                        fill_disponible
                        if ome.get("estado_accion") == "disponible_activar"
                        else verde
                        if ome.get("mismo_efector")
                        else fill_otro_prestador
                    )
                    _append_reporte_row(
                        ws,
                        current_row,
                        nombre,
                        identificador,
                        _turno_con_practica_ome(turno, ome, requisitos, sistema_turnos),
                        ome_reporte,
                        fill,
                        borde,
                    )
                    current_row += 1
                if emitir_faltantes:
                    for codigo, descripcion in requisitos.items():
                        if (codigo, descripcion) in requisitos_cubiertos:
                            continue
                        if not _requisito_presente(omes_presentes, codigo, descripcion):
                            fuera_de_fecha = next(
                                (
                                    (idx, ome)
                                    for idx, ome in enumerate(omes)
                                    if (idx not in omes_usadas or _ome_reutilizable_en_reporte(ome))
                                    and not _ome_bloqueada_por_combo_cima(ome, turno, mapa_especialidades, sistema_turnos)
                                    and _ome_cubre_requisito(ome, codigo, descripcion)
                                    and _fecha_key(ome.get("f_agenda", "")) not in ("", fecha_turno_key)
                                    and _ome_en_ventana_de_turnos(ome, turnos, dias_previos=15)
                                    and not _ome_reservada_para_turno_exacto(
                                        ome,
                                        turno,
                                        turnos,
                                        mapa_especialidades,
                                        codigo,
                                        descripcion,
                                        sistema_turnos,
                                    )
                                ),
                                None,
                            )
                            if fuera_de_fecha:
                                idx, ome = fuera_de_fecha
                                omes_usadas.add(idx)
                                requisitos_cubiertos.add((codigo, descripcion))
                                fecha_en_mes = bool(
                                    ome.get("mismo_efector")
                                    and _mismo_mes(ome.get("f_agenda", ""), fecha_turno_key)
                                )
                                fill_fuera_fecha = verde if fecha_en_mes else amarillo
                                ome_reporte = _ome_fecha_mismo_mes(ome) if fecha_en_mes else _ome_fecha_no_coincide(ome)
                                _append_reporte_row(
                                    ws,
                                    current_row,
                                    nombre,
                                    identificador,
                                    _turno_con_practica_faltante(turno, descripcion, sistema_turnos),
                                    ome_reporte,
                                    fill_fuera_fecha,
                                    borde,
                                )
                                current_row += 1
                                continue
                            relacionada_para_revisar = next(
                                (
                                    (idx, ome)
                                    for idx, ome in enumerate(omes)
                                    if idx not in omes_usadas
                                    and _ome_relacionada_por_codigo_para_revisar(ome, codigo, descripcion)
                                    and _fecha_key(ome.get("f_agenda", "")) == fecha_turno_key
                                ),
                                None,
                            )
                            if relacionada_para_revisar:
                                idx, ome = relacionada_para_revisar
                                omes_usadas.add(idx)
                                requisitos_cubiertos.add((codigo, descripcion))
                                _append_reporte_row(
                                    ws,
                                    current_row,
                                    nombre,
                                    identificador,
                                    _turno_con_practica_faltante(turno, descripcion, sistema_turnos),
                                    _ome_practica_relacionada(ome),
                                    amarillo,
                                    borde,
                                )
                                current_row += 1
                                continue
                            sustituta_misma_fecha = _buscar_ome_sustituta_misma_fecha(
                                omes,
                                omes_usadas,
                                turno,
                                turnos,
                                mapa_especialidades,
                                sistema_turnos,
                            )
                            if sustituta_misma_fecha:
                                idx, ome = sustituta_misma_fecha
                                omes_usadas.add(idx)
                                requisitos_cubiertos.add((codigo, descripcion))
                                _append_reporte_row(
                                    ws,
                                    current_row,
                                    nombre,
                                    identificador,
                                    _turno_con_practica_faltante(turno, descripcion, sistema_turnos),
                                    _ome_sustitucion_especialista(ome),
                                    verde,
                                    borde,
                                )
                                current_row += 1
                                continue
                            compatible_sin_fecha_exacta = next(
                                (
                                    (idx, ome)
                                    for idx, ome in enumerate(omes)
                                    if (idx not in omes_usadas or _ome_reutilizable_en_reporte(ome))
                                    and not _ome_bloqueada_por_combo_cima(ome, turno, mapa_especialidades, sistema_turnos)
                                    and _ome_cubre_requisito(ome, codigo, descripcion)
                                    and (
                                        ome.get("estado_accion") == "disponible_activar"
                                        or _ome_en_ventana_de_turnos(ome, turnos, dias_previos=15)
                                        or (
                                            _sistema_turnos_key(sistema_turnos) == "GJS"
                                            and _ome_modificable_gjs_para_turno(ome, turno)
                                        )
                                    )
                                    and not _ome_reservada_para_turno_exacto(
                                        ome,
                                        turno,
                                        turnos,
                                        mapa_especialidades,
                                        codigo,
                                        descripcion,
                                        sistema_turnos,
                                    )
                                ),
                                None,
                            )
                            if compatible_sin_fecha_exacta:
                                idx, ome = compatible_sin_fecha_exacta
                                omes_usadas.add(idx)
                                requisitos_cubiertos.add((codigo, descripcion))
                                if ome.get("estado_accion") == "disponible_activar":
                                    ome_reporte = ome
                                    fill_reporte = fill_disponible
                                else:
                                    ome_reporte = _ome_fecha_no_coincide(ome)
                                    fill_reporte = amarillo
                                _append_reporte_row(
                                    ws,
                                    current_row,
                                    nombre,
                                    identificador,
                                    _turno_con_practica_faltante(turno, descripcion, sistema_turnos),
                                    ome_reporte,
                                    fill_reporte,
                                    borde,
                                )
                                current_row += 1
                                continue
                            _append_reporte_row(
                                ws,
                                current_row,
                                nombre,
                                identificador,
                                _turno_con_practica_faltante(turno, descripcion, sistema_turnos),
                                _ome_faltante(codigo, descripcion),
                                rojo,
                                borde,
                            )
                            current_row += 1
                continue

            matches = [
                (idx, ome)
                for idx, ome in enumerate(omes)
                if idx not in omes_usadas
                and _fecha_key(ome.get("f_agenda", "")) == fecha_turno_key
                and _ome_relacionada_con_turno(ome, turno, mapa_especialidades)
            ]
            if not matches:
                compatibles = [
                    (idx, ome)
                    for idx, ome in enumerate(omes)
                    if idx not in omes_usadas and _ome_compatible_con_turno(ome, turno, mapa_especialidades)
                ]
                if compatibles:
                    idx, ome = compatibles[0]
                    omes_usadas.add(idx)
                    _append_reporte_row(ws, current_row, nombre, identificador, turno, ome, fill_disponible, borde)
                    current_row += 1
                else:
                    fuera_de_fecha = [
                        (idx, ome)
                        for idx, ome in enumerate(omes)
                        if idx not in omes_usadas
                        and _ome_relacionada_con_turno(ome, turno, mapa_especialidades)
                        and _fecha_key(ome.get("f_agenda", "")) not in ("", fecha_turno_key)
                        and _ome_en_ventana_de_turnos(ome, turnos, dias_previos=15)
                        and not _ome_reservada_para_turno_exacto(ome, turno, turnos, mapa_especialidades, sistema_turnos=sistema_turnos)
                    ]
                    if fuera_de_fecha:
                        for idx, ome in fuera_de_fecha:
                            omes_usadas.add(idx)
                            fecha_en_mes = bool(
                                ome.get("mismo_efector")
                                and _mismo_mes(ome.get("f_agenda", ""), fecha_turno_key)
                            )
                            fill_fuera_fecha = verde if fecha_en_mes else amarillo
                            ome_reporte = _ome_fecha_mismo_mes(ome) if fecha_en_mes else _ome_fecha_no_coincide(ome)
                            _append_reporte_row(
                                ws,
                                current_row,
                                nombre,
                                identificador,
                                turno,
                                ome_reporte,
                                fill_fuera_fecha,
                                borde,
                            )
                            current_row += 1
                    else:
                        if _turno_cima_novelli_dario_ambiguo(turno, sistema_turnos):
                            _append_reporte_row(
                                ws,
                                current_row,
                                nombre,
                                identificador,
                                _turno_con_practica_faltante(turno, "ECOGRAFIA / ECODOPPLER", sistema_turnos),
                                _ome_faltante("", "ECOGRAFIA / ECODOPPLER"),
                                rojo,
                                borde,
                            )
                            current_row += 1
                            continue
                        if _turno_cima_profesional_ambiguo_sin_requisito(turno, sistema_turnos):
                            turnos_omitidos_sin_faltante = True
                            continue
                        _append_reporte_row(ws, current_row, nombre, identificador, turno, None, rojo, borde)
                        current_row += 1
                continue
            for idx, ome in matches:
                omes_usadas.add(idx)
                fill = fill_disponible if ome.get("estado_accion") == "disponible_activar" else verde if ome.get("mismo_efector") else fill_otro_prestador
                _append_reporte_row(ws, current_row, nombre, identificador, turno, ome, fill, borde)
                current_row += 1

        for idx, ome in enumerate(omes):
            if idx in omes_usadas:
                continue
            turno_relacionado = next(
                (
                    turno
                    for turno in turnos
                    if _ome_relacionada_con_turno(ome, turno, mapa_especialidades)
                ),
                None,
            )
            if not turno_relacionado:
                continue
            fecha_ome_key = _fecha_key(ome.get("f_agenda", ""))
            fecha_turno_key = turno_relacionado.get("fecha_key") or _fecha_key(turno_relacionado.get("fecha"))
            requisitos_relacionados = _requisitos_practicas_turno(turno_relacionado, mapa_especialidades)
            requisitos_relacionados = _quitar_requisitos_duplicados_por_bono_gjs(
                requisitos_relacionados,
                turno_relacionado,
                turnos,
                mapa_especialidades,
                sistema_turnos,
            )
            requisitos_relacionados = _omitir_validacion_combo_cima(
                requisitos_relacionados,
                turno_relacionado,
                turnos,
                mapa_especialidades,
                sistema_turnos,
            )
            if (
                fecha_ome_key
                and fecha_turno_key
                and fecha_ome_key != fecha_turno_key
                and requisitos_relacionados
                and any(
                    (codigo, descripcion) in requisitos_cubiertos
                    and _ome_cubre_requisito(ome, codigo, descripcion)
                    for codigo, descripcion in requisitos_relacionados.items()
                )
            ):
                continue
            if not _ome_en_ventana_de_turnos(ome, turnos, dias_previos=15):
                continue
            fecha_distinta = bool(fecha_ome_key and fecha_turno_key and fecha_ome_key != fecha_turno_key)
            fecha_en_mes = bool(ome.get("mismo_efector") and _mismo_mes(fecha_ome_key, fecha_turno_key))
            if fecha_distinta:
                ome_reporte = _ome_fecha_mismo_mes(ome) if fecha_en_mes else _ome_fecha_no_coincide(ome)
                fill_reporte = verde if fecha_en_mes else amarillo
            else:
                ome_reporte = ome
                fill_reporte = (
                    fill_disponible
                    if ome.get("estado_accion") == "disponible_activar"
                    else verde
                    if ome.get("mismo_efector")
                    else fill_otro_prestador
                )
            _append_reporte_row(
                ws,
                current_row,
                nombre,
                identificador,
                turno_relacionado,
                ome_reporte,
                fill_reporte,
                borde,
            )
            current_row += 1

        if current_row == start_row:
            if turnos_omitidos_sin_faltante:
                continue
            _append_reporte_row(ws, current_row, nombre, identificador, None, None, rojo, borde)
            current_row += 1

        if current_row - start_row > 1:
            ws.merge_cells(start_row=start_row, start_column=1, end_row=current_row - 1, end_column=1)
            ws.cell(row=start_row, column=1).alignment = Alignment(vertical="center", wrap_text=True)
            if not _es_sistema_cima(sistema_turnos):
                ws.merge_cells(start_row=start_row, start_column=2, end_row=current_row - 1, end_column=2)
                ws.cell(row=start_row, column=2).alignment = Alignment(vertical="center", wrap_text=True)

    current_row += 1
    ws.cell(row=current_row, column=1, value="Leyenda").font = Font(bold=True, color="16324F")
    leyendas = [
        ("OME activa para nuestro centro con fecha coincidente", verde),
        ("OME disponible para activar", fill_disponible),
        ("OME activa para otro prestador", fill_otro_prestador),
        (f"OME existente con fecha PAMI distinta al turno {sistema_turnos}", amarillo),
        (f"Turno {sistema_turnos} sin OME en PAMI para esa fecha", rojo),
        (
            f"OME en PAMI sin bono {sistema_turnos} correspondiente"
            if sistema_turnos == "GJS"
            else f"OME en PAMI sin turno {sistema_turnos} correspondiente",
            rojo if sistema_turnos == "GJS" else gris,
        ),
    ]
    for texto, fill in leyendas:
        current_row += 1
        ws.cell(row=current_row, column=1, value=texto)
        ws.cell(row=current_row, column=1).fill = fill

    widths = (
        {"A": 26, "B": 18, "C": 16, "D": 13, "E": 20, "F": 13, "G": 13, "H": 44}
        if _es_sistema_cima(sistema_turnos)
        else {"A": 26, "B": 18, "C": 13, "D": 20, "E": 13, "F": 13, "G": 44}
    )
    for column, width in widths.items():
        ws.column_dimensions[column].width = width

    ws.freeze_panes = "A3"
    _crear_hoja_resumen_reporte(wb, ws, sistema_turnos)
    wb.save(ruta_salida)
    log_message(f"Reporte de verificacion guardado en {ruta_salida}")
    return ruta_salida


def generar_reporte_auditoria(
    pacientes: list[dict],
    omes_por_key: dict[str, list],
    ruta_salida: Path,
    cliente_codigo: str | None = CLIENTE_CIMA,
) -> Path:
    cliente = _cliente_config(cliente_codigo)
    sistema_turnos = cliente.get("sistema_turnos", "CIMA")
    ruta_salida = Path(ruta_salida)
    if not ruta_salida.suffix:
        ruta_salida = ruta_salida.with_suffix(".xlsx")
    ruta_salida.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws._sistema_turnos = sistema_turnos
    ws.title = "Auditoria prestaciones"

    azul_oscuro = PatternFill("solid", fgColor="1F4E79")
    azul = PatternFill("solid", fgColor="2E75B6")
    verde = PatternFill("solid", fgColor="C6EFCE")
    amarillo = PatternFill("solid", fgColor="FFF2CC")
    rojo = PatternFill("solid", fgColor="FFC7CE")
    gris = PatternFill("solid", fgColor="F2F2F2")
    blanco = Font(color="FFFFFF", bold=True)
    borde = Border(*(Side(style="thin", color="D9E2F3") for _ in range(4)))

    columnas = _columnas_reporte(sistema_turnos)
    ws.merge_cells("A1:A2")
    ws.merge_cells("B1:B2")
    ws["A1"] = "PACIENTE"
    ws["B1"] = _titulo_identificador_reporte(sistema_turnos)
    if _es_sistema_cima(sistema_turnos):
        ws.merge_cells("C1:C2")
        ws.merge_cells("D1:E1")
        ws.merge_cells("F1:H1")
        ws["C1"] = "NRO. TRAMITE"
        ws["D1"] = f"SISTEMA {sistema_turnos}"
        ws["F1"] = "PANEL TRANSMISION"
        headers = ["FECHA", "ESPECIALISTA", "FECHA TURNO", "TRANSMITIDA", "PRACTICA / ESTADO"]
        for col, value in enumerate(headers, start=4):
            ws.cell(row=2, column=col, value=value)
    else:
        ws.merge_cells("C1:D1")
        ws.merge_cells("E1:G1")
        ws["C1"] = f"SISTEMA {sistema_turnos}"
        ws["E1"] = "PANEL TRANSMISION"
        headers = ["FECHA", "ESPECIALISTA", "FECHA TURNO", "TRANSMITIDA", "PRACTICA / ESTADO"]
        for col, value in enumerate(headers, start=3):
            ws.cell(row=2, column=col, value=value)

    for row in (1, 2):
        for col in range(1, columnas["total"] + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = azul_oscuro if row == 1 else azul
            cell.font = blanco
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = borde

    mapa_especialidades = cargar_especialidades_medicos()
    current_row = 3
    for paciente in pacientes:
        if _paciente_excluido_reporte_no_pami(paciente, sistema_turnos):
            continue
        start_row = current_row
        key = paciente.get("key") or paciente.get("beneficiario") or f"DNI_{paciente.get('dni', '')}"
        nombre = paciente.get("nombre") or key
        identificador = _identificador_paciente(paciente)
        turnos = [turno for turno in paciente.get("turnos", []) if not _turno_no_reportable(turno)]
        if not turnos:
            continue
        prestaciones = [
            item for item in list(omes_por_key.get(key, []))
            if isinstance(item, dict) and item.get("estado_accion") != "sin_resultados_pami"
        ]
        usadas: set[int] = set()
        requisitos_procesados: set[tuple] = set()
        asignaciones_codigo = _asignaciones_exactas_por_codigo(turnos, prestaciones)

        for turno_idx, turno in enumerate(turnos):
            fecha_turno_key = turno.get("fecha_key") or _fecha_key(turno.get("fecha"))
            idx_asignado = asignaciones_codigo.get(turno_idx)
            if idx_asignado is not None and idx_asignado not in usadas:
                item = prestaciones[idx_asignado]
                usadas.add(idx_asignado)
                _marcar_requisitos_cubiertos(
                    turno,
                    turnos,
                    item,
                    requisitos_procesados,
                    mapa_especialidades,
                    sistema_turnos,
                )
                fill = _fill_auditoria(item, verde, amarillo)
                if _fecha_key(item.get("f_agenda", "")) != fecha_turno_key:
                    item, fill = _item_auditoria_fecha_distinta(item, fecha_turno_key, verde, amarillo)
                _append_reporte_row(
                    ws,
                    current_row,
                    nombre,
                    identificador,
                    turno,
                    item,
                    fill,
                    borde,
                )
                current_row += 1
                current_row = _append_requisitos_auditoria(
                    ws,
                    current_row,
                    nombre,
                    identificador,
                    turno,
                    turnos,
                    prestaciones,
                    usadas,
                    requisitos_procesados,
                    mapa_especialidades,
                    sistema_turnos,
                    verde,
                    amarillo,
                    rojo,
                    borde,
                )
                continue
            if idx_asignado is not None and idx_asignado in usadas and _codigo_practica(str(turno.get("profesional", "") or "")) == "820137":
                item = prestaciones[idx_asignado]
                _append_reporte_row(
                    ws,
                    current_row,
                    nombre,
                    identificador,
                    turno,
                    item,
                    _fill_auditoria(item, verde, amarillo),
                    borde,
                )
                current_row += 1
                current_row = _append_requisitos_auditoria(
                    ws,
                    current_row,
                    nombre,
                    identificador,
                    turno,
                    turnos,
                    prestaciones,
                    usadas,
                    requisitos_procesados,
                    mapa_especialidades,
                    sistema_turnos,
                    verde,
                    amarillo,
                    rojo,
                    borde,
                )
                continue
            exactas = [
                (idx, item)
                for idx, item in enumerate(prestaciones)
                if (idx not in usadas or _ome_reutilizable_en_reporte(item))
                and _fecha_key(item.get("f_agenda", "")) == fecha_turno_key
                and _ome_relacionada_con_turno(item, turno, mapa_especialidades, sistema_turnos)
            ]
            exactas = _filtrar_candidatos_duplicados_auditoria(
                exactas,
                turno,
                turnos,
                mapa_especialidades,
                sistema_turnos,
            )
            requisitos_turno = _requisitos_efectivos_auditoria(turno, turnos, mapa_especialidades, sistema_turnos)
            exactas.sort(
                key=lambda candidato: (
                    _prioridad_ome_turno_auditoria(candidato[1], turno, requisitos_turno, sistema_turnos),
                    candidato[0],
                )
            )
            if exactas:
                for idx, item in exactas:
                    usadas.add(idx)
                    _marcar_requisitos_cubiertos(
                        turno,
                        turnos,
                        item,
                        requisitos_procesados,
                        mapa_especialidades,
                        sistema_turnos,
                    )
                    _append_reporte_row(
                        ws,
                        current_row,
                        nombre,
                        identificador,
                        turno,
                        item,
                        _fill_auditoria(item, verde, amarillo),
                        borde,
                    )
                    current_row += 1
                current_row = _append_requisitos_auditoria(
                    ws,
                    current_row,
                    nombre,
                    identificador,
                    turno,
                    turnos,
                    prestaciones,
                    usadas,
                    requisitos_procesados,
                    mapa_especialidades,
                    sistema_turnos,
                    verde,
                    amarillo,
                    rojo,
                    borde,
                )
                continue

            relacionadas = [
                (idx, item)
                for idx, item in enumerate(prestaciones)
                if (idx not in usadas or _ome_reutilizable_en_reporte(item))
                and _ome_relacionada_con_turno(item, turno, mapa_especialidades, sistema_turnos)
                and _ome_en_mes_de_algun_turno(item, [turno])
            ]
            if relacionadas:
                requisitos_turno = _requisitos_efectivos_auditoria(turno, turnos, mapa_especialidades, sistema_turnos)
                relacionadas.sort(
                    key=lambda candidato: (
                        _prioridad_ome_turno_auditoria(candidato[1], turno, requisitos_turno, sistema_turnos),
                        candidato[0],
                    )
                )
                idx, item = relacionadas[0]
                usadas.add(idx)
                _marcar_requisitos_cubiertos(
                    turno,
                    turnos,
                    item,
                    requisitos_procesados,
                    mapa_especialidades,
                    sistema_turnos,
                )
                _append_reporte_row(
                    ws,
                    current_row,
                    nombre,
                    identificador,
                    turno,
                    *(
                        (item, rojo if sistema_turnos == "GJS" else amarillo)
                        if item.get("estado_accion") == "disponible_activar"
                        else _item_auditoria_fecha_distinta(item, fecha_turno_key, verde, amarillo)
                    ),
                    borde,
                )
                current_row += 1
                current_row = _append_requisitos_auditoria(
                    ws,
                    current_row,
                    nombre,
                    identificador,
                    turno,
                    turnos,
                    prestaciones,
                    usadas,
                    requisitos_procesados,
                    mapa_especialidades,
                    sistema_turnos,
                    verde,
                    amarillo,
                    rojo,
                    borde,
                )
                continue

            fuera_de_mes = [
                (idx, item)
                for idx, item in enumerate(prestaciones)
                if (idx not in usadas or _ome_reutilizable_en_reporte(item))
                and _ome_relacionada_con_turno(item, turno, mapa_especialidades, sistema_turnos)
            ]
            if fuera_de_mes:
                requisitos_turno = _requisitos_efectivos_auditoria(turno, turnos, mapa_especialidades, sistema_turnos)
                fuera_de_mes.sort(
                    key=lambda candidato: (
                        _prioridad_ome_turno_auditoria(candidato[1], turno, requisitos_turno, sistema_turnos),
                        candidato[0],
                    )
                )
                idx, item = fuera_de_mes[0]
                usadas.add(idx)
                _marcar_requisitos_cubiertos(
                    turno,
                    turnos,
                    item,
                    requisitos_procesados,
                    mapa_especialidades,
                    sistema_turnos,
                )
                _append_reporte_row(
                    ws,
                    current_row,
                    nombre,
                    identificador,
                    turno,
                    item if item.get("estado_accion") == "disponible_activar" else _ome_fecha_no_coincide(item),
                    rojo if item.get("estado_accion") == "disponible_activar" and sistema_turnos == "GJS" else amarillo,
                    borde,
                )
                current_row += 1
                current_row = _append_requisitos_auditoria(
                    ws,
                    current_row,
                    nombre,
                    identificador,
                    turno,
                    turnos,
                    prestaciones,
                    usadas,
                    requisitos_procesados,
                    mapa_especialidades,
                    sistema_turnos,
                    verde,
                    amarillo,
                    rojo,
                    borde,
                )
                continue

            fila_antes_faltantes = current_row
            current_row = _append_requisitos_auditoria(
                ws,
                current_row,
                nombre,
                identificador,
                turno,
                turnos,
                prestaciones,
                usadas,
                requisitos_procesados,
                mapa_especialidades,
                sistema_turnos,
                verde,
                amarillo,
                rojo,
                borde,
            )
            if current_row != fila_antes_faltantes:
                continue

            fila_antes_fallback = current_row
            current_row = _append_faltantes_especificos_auditoria(
                ws,
                current_row,
                nombre,
                identificador,
                turno,
                turnos,
                prestaciones,
                usadas,
                mapa_especialidades,
                sistema_turnos,
                rojo,
                borde,
            )
            if current_row != fila_antes_fallback:
                continue
            if _requisitos_efectivos_auditoria(turno, turnos, mapa_especialidades, sistema_turnos):
                continue

            _append_reporte_row(
                ws,
                current_row,
                nombre,
                identificador,
                turno,
                _ome_faltante("", f"SIN PRESTACION EN TRANSMISION PARA ESTE TURNO {sistema_turnos}"),
                rojo,
                borde,
            )
            current_row += 1

        sobrantes_vistos: set[tuple] = set()
        for idx, item in enumerate(prestaciones):
            if idx in usadas:
                continue
            turno_relacionado = next(
                (
                    turno
                    for turno in turnos
                    if _ome_relacionada_con_turno(item, turno, mapa_especialidades, sistema_turnos)
                ),
                None,
            )
            if not turno_relacionado:
                continue
            if _prestacion_ya_cubierta_en_auditoria(
                item,
                [prestacion for j, prestacion in enumerate(prestaciones) if j in usadas],
                turnos,
                mapa_especialidades,
                sistema_turnos,
            ):
                continue
            if not _ome_en_mes_de_algun_turno(item, turnos):
                continue
            firma_sobrante = _firma_prestacion_sobrante_auditoria(item)
            if firma_sobrante in sobrantes_vistos:
                continue
            sobrantes_vistos.add(firma_sobrante)
            _append_reporte_row(
                ws,
                current_row,
                nombre,
                identificador,
                turno_relacionado,
                item,
                rojo if sistema_turnos == "GJS" else gris,
                borde,
            )
            current_row += 1

        if current_row == start_row:
            _append_reporte_row(
                ws,
                current_row,
                nombre,
                identificador,
                None,
                _ome_sin_resultados_pami(),
                rojo,
                borde,
            )
            current_row += 1

        if current_row - start_row > 1:
            ws.merge_cells(start_row=start_row, start_column=1, end_row=current_row - 1, end_column=1)
            ws.cell(row=start_row, column=1).alignment = Alignment(vertical="center", wrap_text=True)
            if not _es_sistema_cima(sistema_turnos):
                ws.merge_cells(start_row=start_row, start_column=2, end_row=current_row - 1, end_column=2)
                ws.cell(row=start_row, column=2).alignment = Alignment(vertical="center", wrap_text=True)

    current_row += 1
    ws.cell(row=current_row, column=1, value="Leyenda").font = Font(bold=True, color="16324F")
    leyendas = [
        ("Prestacion transmitida", verde),
        ("Prestacion encontrada con pendiente de validacion/documentacion/transmision o fecha distinta", amarillo),
        (f"Turno {sistema_turnos} sin prestacion compatible en Transmision", rojo),
        (f"Prestacion en Transmision sin bono/turno {sistema_turnos}", rojo if sistema_turnos == "GJS" else gris),
    ]
    for texto, fill in leyendas:
        current_row += 1
        ws.cell(row=current_row, column=1, value=texto)
        ws.cell(row=current_row, column=1).fill = fill

    widths = (
        {"A": 26, "B": 18, "C": 16, "D": 13, "E": 28, "F": 13, "G": 22, "H": 52}
        if _es_sistema_cima(sistema_turnos)
        else {"A": 26, "B": 18, "C": 13, "D": 28, "E": 13, "F": 22, "G": 52}
    )
    for column, width in widths.items():
        ws.column_dimensions[column].width = width

    ws.freeze_panes = "A3"
    _crear_hoja_resumen_reporte(wb, ws, sistema_turnos)
    wb.save(ruta_salida)
    log_message(f"Reporte de auditoria de prestaciones guardado en {ruta_salida}")
    return ruta_salida


def _asignaciones_exactas_por_codigo(turnos: list[dict], prestaciones: list[dict]) -> dict[int, int]:
    disponibles: dict[tuple[str, str], list[int]] = {}
    comodines: dict[str, list[int]] = {}
    for idx, item in enumerate(prestaciones):
        codigo = _codigo_practica(str(item.get("practica", "") or ""))
        if not codigo:
            continue
        fecha = _fecha_key(item.get("f_agenda", ""))
        disponibles.setdefault((codigo, fecha), []).append(idx)
        comodines.setdefault(codigo, []).append(idx)

    asignadas: dict[int, int] = {}
    usados: set[int] = set()
    for turno_idx, turno in enumerate(turnos):
        codigo = _codigo_practica(str(turno.get("profesional", "") or ""))
        fecha = turno.get("fecha_key") or _fecha_key(turno.get("fecha"))
        if not codigo:
            continue
        candidatos = disponibles.get((codigo, fecha), [])
        if not candidatos:
            candidatos = comodines.get(codigo, [])
        for idx in candidatos:
            if idx not in usados or codigo == "820137":
                asignadas[turno_idx] = idx
                if codigo != "820137":
                    usados.add(idx)
                break
    return asignadas


def _filtrar_candidatos_duplicados_auditoria(
    candidatos: list[tuple[int, dict]],
    turno: dict,
    turnos: list[dict],
    mapa_especialidades: dict[str, list[str]],
    sistema_turnos: str,
) -> list[tuple[int, dict]]:
    if len(candidatos) <= 1:
        return candidatos

    requisitos = _requisitos_efectivos_auditoria(turno, turnos, mapa_especialidades, sistema_turnos)
    if not requisitos:
        return candidatos

    seleccionados: list[tuple[int, dict]] = []
    requisitos_cubiertos: set[tuple[str, str]] = set()
    fecha_turno_key = turno.get("fecha_key") or _fecha_key(turno.get("fecha"))
    candidatos_ordenados = sorted(
        candidatos,
        key=lambda item: (
            _prioridad_ome_turno_auditoria(item[1], turno, requisitos, sistema_turnos),
            item[0],
        ),
    )

    for idx, item in candidatos_ordenados:
        cubre = {
            (codigo, descripcion)
            for codigo, descripcion in requisitos.items()
            if _ome_cubre_requisito(item, codigo, descripcion)
        }
        if not cubre:
            seleccionados.append((idx, item))
            continue
        nuevos = cubre - requisitos_cubiertos
        if not nuevos:
            continue
        requisitos_cubiertos.update(nuevos)
        seleccionados.append((idx, item))

    return sorted(seleccionados, key=lambda item: item[0])


def _prioridad_prestacion_auditoria(ome: dict) -> int:
    if _ome_auditoria_transmitida(ome):
        return 0
    if ome.get("validada"):
        return 1
    if ome.get("estado_accion") == "disponible_activar":
        return 3
    return 2


def _prioridad_candidato_auditoria(ome: dict, fecha_turno_key: str) -> tuple[int, int]:
    fecha_ome = _fecha_key(ome.get("f_agenda", ""))
    if fecha_ome and fecha_ome == fecha_turno_key:
        relacion_fecha = 0
    elif ome.get("estado_accion") == "disponible_activar":
        relacion_fecha = 1
    elif fecha_ome and _mismo_mes(fecha_ome, fecha_turno_key):
        relacion_fecha = 2
    else:
        relacion_fecha = 3
    return relacion_fecha, _prioridad_prestacion_auditoria(ome)


def _categoria_doppler_gjs_texto(valor: str) -> str:
    texto = _texto_busqueda(valor)
    if "vasos" in texto and "cuello" in texto:
        return "vasos_cuello"
    if "carotid" in texto or "cuello" in texto:
        return "vasos_cuello"
    if "venoso" in texto:
        return "venoso"
    if "arterial" in texto or "arterio" in texto:
        return "arterial"
    if "cardiaco" in texto or re.search(r"\bcardio\b", texto):
        return "cardiaco"
    return ""


def _categoria_doppler_gjs_ome(ome: dict) -> str:
    practica = str((ome or {}).get("practica", "") or "")
    codigo = _codigo_practica(practica)
    por_codigo = {
        "180607": "vasos_cuello",
        "180606": "venoso",
        "180610": "arterial",
        "180301": "cardiaco",
    }
    return por_codigo.get(codigo, "") or _categoria_doppler_gjs_texto(practica)


def _categoria_doppler_gjs_turno(turno: dict, descripcion: str) -> str:
    texto = f"{(turno or {}).get('profesional', '')} {descripcion or ''}"
    return _categoria_doppler_gjs_texto(texto)


def _es_requisito_doppler_gjs(codigo: str, descripcion: str) -> bool:
    opciones = {item.strip() for item in str(codigo or "").split("|") if item.strip()}
    if opciones & {"180301", "180606", "180607", "180610", "base:doppler", "base:ecodoppler_cardiaco"}:
        return True
    texto = _texto_busqueda(descripcion)
    return "doppler" in texto or "ecodoppler" in texto


def _distancia_dias_turno(ome: dict, turno: dict) -> int:
    fecha_ome = _parse_fecha((ome or {}).get("f_agenda", ""))
    fecha_turno = _parse_fecha((turno or {}).get("fecha"))
    if fecha_ome is None or fecha_turno is None:
        return 9999
    return abs((fecha_ome - fecha_turno).days)


def _prioridad_doppler_gjs_auditoria(
    ome: dict,
    turno: dict,
    codigo: str,
    descripcion: str,
    sistema_turnos: str,
) -> tuple[int, int]:
    if _sistema_turnos_key(sistema_turnos) != "GJS" or not _es_requisito_doppler_gjs(codigo, descripcion):
        return (0, 0)

    categoria_turno = _categoria_doppler_gjs_turno(turno, descripcion)
    categoria_ome = _categoria_doppler_gjs_ome(ome)
    if categoria_turno:
        if categoria_ome == categoria_turno:
            categoria_rank = 0
        elif not categoria_ome:
            categoria_rank = 1
        else:
            categoria_rank = 2
    else:
        categoria_rank = 0
    return categoria_rank, _distancia_dias_turno(ome, turno)


def _prioridad_candidato_turno_auditoria(
    ome: dict,
    turno: dict,
    codigo: str,
    descripcion: str,
    sistema_turnos: str,
) -> tuple:
    fecha_turno_key = (turno or {}).get("fecha_key") or _fecha_key((turno or {}).get("fecha"))
    return (
        _prioridad_doppler_gjs_auditoria(ome, turno, codigo, descripcion, sistema_turnos),
        _prioridad_candidato_auditoria(ome, fecha_turno_key),
        _distancia_dias_turno(ome, turno),
    )


def _prioridad_ome_turno_auditoria(
    ome: dict,
    turno: dict,
    requisitos: dict[str, str],
    sistema_turnos: str,
) -> tuple:
    fecha_turno_key = (turno or {}).get("fecha_key") or _fecha_key((turno or {}).get("fecha"))
    if not requisitos:
        return (
            (0, _distancia_dias_turno(ome, turno)),
            _prioridad_candidato_auditoria(ome, fecha_turno_key),
            _distancia_dias_turno(ome, turno),
        )
    return min(
        _prioridad_candidato_turno_auditoria(ome, turno, codigo, descripcion, sistema_turnos)
        for codigo, descripcion in requisitos.items()
    )


def _hay_turnos_doppler_duplicados_gjs(
    turno_actual: dict,
    turnos: list[dict],
    mapa_especialidades: dict[str, list[str]],
    sistema_turnos: str,
) -> bool:
    if _sistema_turnos_key(sistema_turnos) != "GJS":
        return False
    fecha_key = (turno_actual or {}).get("fecha_key") or _fecha_key((turno_actual or {}).get("fecha"))
    paquete_key = _paquete_key_turno(turno_actual, mapa_especialidades, sistema_turnos)
    if "ecodoppler_cardiaco" not in _bases_esperadas_turno(turno_actual, mapa_especialidades, sistema_turnos):
        return False
    cantidad = 0
    for turno in turnos:
        if ((turno or {}).get("fecha_key") or _fecha_key((turno or {}).get("fecha"))) != fecha_key:
            continue
        if _paquete_key_turno(turno, mapa_especialidades, sistema_turnos) != paquete_key:
            continue
        if "ecodoppler_cardiaco" not in _bases_esperadas_turno(turno, mapa_especialidades, sistema_turnos):
            continue
        cantidad += 1
    return cantidad > 1


def _requisitos_efectivos_auditoria(
    turno: dict,
    turnos: list[dict],
    mapa_especialidades: dict[str, list[str]],
    sistema_turnos: str,
) -> dict[str, str]:
    requisitos = _requisitos_practicas_turno(turno, mapa_especialidades)
    requisitos = _ajustar_requisitos_por_sistema(requisitos, sistema_turnos)
    requisitos = _ajustar_requisitos_mamografia_auditoria(requisitos, turno, mapa_especialidades, sistema_turnos)
    requisitos = _quitar_requisitos_duplicados_por_bono_gjs(
        requisitos,
        turno,
        turnos,
        mapa_especialidades,
        sistema_turnos,
    )
    return _omitir_validacion_combo_cima(
        requisitos,
        turno,
        turnos,
        mapa_especialidades,
        sistema_turnos,
    )


def _ajustar_requisitos_mamografia_auditoria(
    requisitos: dict[str, str],
    turno: dict,
    mapa_especialidades: dict[str, list[str]],
    sistema_turnos: str,
) -> dict[str, str]:
    if sistema_turnos != "CIMA" or not requisitos:
        return requisitos
    bases = _bases_esperadas_turno(turno, mapa_especialidades)
    if "mamografia" not in bases:
        return requisitos

    # En auditoria hacia atras, el paquete MX no debe generar faltantes de ecos
    # si el Excel no trajo esos turnos/informes explicitamente.
    filtrados = dict(requisitos)
    filtrados.pop("180106", None)
    filtrados.pop("186001", None)
    return filtrados


def _candidatos_requisito_auditoria(
    prestaciones: list[dict],
    turno: dict,
    codigo: str,
    descripcion: str,
) -> list[tuple[int, dict]]:
    fecha_turno_key = turno.get("fecha_key") or _fecha_key(turno.get("fecha"))
    candidatos = []
    for idx, item in enumerate(prestaciones):
        if not isinstance(item, dict) or not _ome_cubre_requisito(item, codigo, descripcion):
            continue
        fecha_item = _fecha_key(item.get("f_agenda", ""))
        if (
            fecha_item == fecha_turno_key
            or item.get("estado_accion") == "disponible_activar"
            or (item.get("mismo_efector") and fecha_item and _mismo_mes(fecha_item, fecha_turno_key))
        ):
            candidatos.append((idx, item))
    return candidatos


def _faltan_requisitos_auditoria(
    paciente: dict,
    prestaciones: list[dict],
    mapa_especialidades: dict[str, list[str]],
    sistema_turnos: str,
) -> bool:
    turnos = paciente.get("turnos", [])
    for turno in turnos:
        requisitos = _requisitos_efectivos_auditoria(turno, turnos, mapa_especialidades, sistema_turnos)
        for codigo, descripcion in requisitos.items():
            if not _candidatos_requisito_auditoria(prestaciones, turno, codigo, descripcion):
                return True
    return False


def _fusionar_omes_auditoria(
    prestaciones: list[dict],
    omes_panel: list[dict],
    paciente: dict,
    mapa_especialidades: dict[str, list[str]],
    sistema_turnos: str,
) -> list[dict]:
    resultado = list(prestaciones or [])
    turnos = paciente.get("turnos", [])
    existentes = {
        (
            str(item.get("n_orden", "") or "").strip(),
            _codigo_practica(str(item.get("practica", "") or "")),
            _fecha_key(item.get("f_agenda", "")),
            str(item.get("estado_accion", "") or ""),
        )
        for item in resultado
        if isinstance(item, dict)
    }
    for ome in omes_panel or []:
        if not isinstance(ome, dict) or ome.get("estado_accion") == "sin_resultados_pami":
            continue
        cubre_requisito = any(
            _ome_cubre_requisito(ome, codigo, descripcion)
            for turno in turnos
            for codigo, descripcion in _requisitos_efectivos_auditoria(
                turno,
                turnos,
                mapa_especialidades,
                sistema_turnos,
            ).items()
        )
        if not cubre_requisito and not _ome_relacionada_con_turnos(ome, turnos, mapa_especialidades, sistema_turnos):
            continue
        if _prestacion_ya_cubierta_en_auditoria(
            ome,
            resultado,
            turnos,
            mapa_especialidades,
            sistema_turnos,
        ):
            continue
        firma = (
            str(ome.get("n_orden", "") or "").strip(),
            _codigo_practica(str(ome.get("practica", "") or "")),
            _fecha_key(ome.get("f_agenda", "")),
            str(ome.get("estado_accion", "") or ""),
        )
        if firma in existentes:
            continue
        resultado.append(_normalizar_ome_panel_para_auditoria(ome))
        existentes.add(firma)
    return resultado


def _prestacion_ya_cubierta_en_auditoria(
    ome: dict,
    prestaciones: list[dict],
    turnos: list[dict],
    mapa_especialidades: dict[str, list[str]],
    sistema_turnos: str,
) -> bool:
    for turno in turnos:
        requisitos = _requisitos_efectivos_auditoria(turno, turnos, mapa_especialidades, sistema_turnos)
        for codigo, descripcion in requisitos.items():
            if not _ome_cubre_requisito(ome, codigo, descripcion):
                continue
            if not _prestacion_aplica_a_turno_auditoria(ome, turno):
                continue
            for existente in prestaciones:
                if not isinstance(existente, dict):
                    continue
                if not _ome_cubre_requisito(existente, codigo, descripcion):
                    continue
                if not _prestacion_aplica_a_turno_auditoria(existente, turno):
                    continue
                if (
                    _hay_turnos_doppler_duplicados_gjs(turno, turnos, mapa_especialidades, sistema_turnos)
                    and _codigo_practica(str(existente.get("practica", "") or ""))
                    != _codigo_practica(str(ome.get("practica", "") or ""))
                ):
                    continue
                if _prioridad_candidato_turno_auditoria(
                    existente,
                    turno,
                    codigo,
                    descripcion,
                    sistema_turnos,
                ) <= _prioridad_candidato_turno_auditoria(
                    ome,
                    turno,
                    codigo,
                    descripcion,
                    sistema_turnos,
                ):
                    return True
    return False


def _requisito_ya_cubierto_por_usadas(
    prestaciones: list[dict],
    usadas: set[int],
    turno: dict,
    codigo: str,
    descripcion: str,
) -> bool:
    for idx in usadas:
        if idx < 0 or idx >= len(prestaciones):
            continue
        item = prestaciones[idx]
        if not isinstance(item, dict):
            continue
        if not _ome_cubre_requisito(item, codigo, descripcion):
            continue
        if _prestacion_aplica_a_turno_auditoria(item, turno):
            return True
    return False


def _prestacion_aplica_a_turno_auditoria(ome: dict, turno: dict) -> bool:
    fecha_turno_key = turno.get("fecha_key") or _fecha_key(turno.get("fecha"))
    fecha_ome = _fecha_key(ome.get("f_agenda", ""))
    return (
        fecha_ome == fecha_turno_key
        or ome.get("estado_accion") == "disponible_activar"
        or (ome.get("mismo_efector") and fecha_ome and _mismo_mes(fecha_ome, fecha_turno_key))
    )


def _normalizar_ome_panel_para_auditoria(ome: dict) -> dict:
    copia = dict(ome)
    if copia.get("estado_accion") in {"transmitida", "disponible_activar"}:
        return copia
    if copia.get("mismo_efector"):
        copia["estado_accion"] = "auditoria_panel_pendiente"
        copia["validada"] = False
        copia["transmitida"] = False
        copia["doc_cargada"] = False
        copia["doc_pendiente"] = False
    return copia


def _append_requisitos_auditoria(
    ws,
    current_row: int,
    nombre: str,
    identificador: str,
    turno: dict,
    turnos: list[dict],
    prestaciones: list[dict],
    usadas: set[int],
    requisitos_procesados: set[tuple],
    mapa_especialidades: dict[str, list[str]],
    sistema_turnos: str,
    verde,
    amarillo,
    rojo,
    borde,
) -> int:
    requisitos = _requisitos_efectivos_auditoria(turno, turnos, mapa_especialidades, sistema_turnos)
    if not requisitos:
        return current_row

    paquete_key = _paquete_key_turno(turno, mapa_especialidades)
    for codigo, descripcion in requisitos.items():
        clave = (paquete_key, codigo, descripcion)
        if clave in requisitos_procesados:
            continue
        requisitos_procesados.add(clave)
        if _requisito_ya_cubierto_por_usadas(prestaciones, usadas, turno, codigo, descripcion):
            continue
        candidatos = _candidatos_requisito_auditoria(prestaciones, turno, codigo, descripcion)
        if not candidatos:
            sustituta_misma_fecha = _buscar_ome_sustituta_misma_fecha(
                prestaciones,
                usadas,
                turno,
                turnos,
                mapa_especialidades,
                sistema_turnos,
            )
            if sustituta_misma_fecha:
                idx, item = sustituta_misma_fecha
                usadas.add(idx)
                _append_reporte_row(
                    ws,
                    current_row,
                    nombre,
                    identificador,
                    _turno_con_practica_faltante(turno, descripcion, sistema_turnos),
                    _ome_sustitucion_especialista(item),
                    _fill_auditoria(item, verde, amarillo),
                    borde,
                )
                current_row += 1
                continue
            candidatos_fuera_fecha = [
                (idx, item)
                for idx, item in enumerate(prestaciones)
                if (idx not in usadas or _ome_reutilizable_en_reporte(item))
                and isinstance(item, dict)
                and _ome_cubre_requisito(item, codigo, descripcion)
                and not _ome_reservada_para_turno_exacto(
                    item,
                    turno,
                    turnos,
                    mapa_especialidades,
                    codigo,
                    descripcion,
                    sistema_turnos,
                )
            ]
            if candidatos_fuera_fecha:
                candidatos_fuera_fecha.sort(
                    key=lambda candidato: (
                        _prioridad_candidato_turno_auditoria(
                            candidato[1],
                            turno,
                            codigo,
                            descripcion,
                            sistema_turnos,
                        ),
                        candidato[0],
                    )
                )
                idx, item = candidatos_fuera_fecha[0]
                usadas.add(idx)
                item_reporte, fill = _item_auditoria_fecha_distinta(
                    item,
                    turno.get("fecha_key") or _fecha_key(turno.get("fecha")),
                    verde,
                    amarillo,
                )
                _append_reporte_row(
                    ws,
                    current_row,
                    nombre,
                    identificador,
                    _turno_con_practica_faltante(turno, descripcion, sistema_turnos),
                    item_reporte,
                    fill,
                    borde,
                )
                current_row += 1
                continue
            _append_reporte_row(
                ws,
                current_row,
                nombre,
                identificador,
                _turno_con_practica_faltante(turno, descripcion, sistema_turnos),
                _ome_faltante(codigo, descripcion),
                rojo,
                borde,
            )
            current_row += 1
            continue

        candidatos_no_usados = [
            (idx, item)
            for idx, item in candidatos
            if idx not in usadas or _ome_reutilizable_en_reporte(item)
        ]
        candidatos_no_usados = [
            (idx, item)
            for idx, item in candidatos_no_usados
            if not _ome_reservada_para_turno_exacto(
                item,
                turno,
                turnos,
                mapa_especialidades,
                codigo,
                descripcion,
                sistema_turnos,
            )
        ]
        if not candidatos_no_usados:
            continue
        fecha_turno_key = turno.get("fecha_key") or _fecha_key(turno.get("fecha"))
        candidatos_no_usados.sort(
            key=lambda candidato: (
                _prioridad_candidato_turno_auditoria(
                    candidato[1],
                    turno,
                    codigo,
                    descripcion,
                    sistema_turnos,
                ),
                candidato[0],
            )
        )
        idx, item = candidatos_no_usados[0]
        usadas.add(idx)
        fecha_item = _fecha_key(item.get("f_agenda", ""))
        item_reporte = item
        fill = _fill_auditoria(item, verde, amarillo)
        if item.get("estado_accion") == "disponible_activar":
            fill = rojo if sistema_turnos == "GJS" else amarillo
        elif fecha_item not in ("", fecha_turno_key):
            item_reporte, fill = _item_auditoria_fecha_distinta(item, fecha_turno_key, verde, amarillo)
        _append_reporte_row(
            ws,
            current_row,
            nombre,
            identificador,
            _turno_con_practica_faltante(turno, descripcion, sistema_turnos),
            item_reporte,
            fill,
            borde,
        )
        current_row += 1
    return current_row


def _append_faltantes_especificos_auditoria(
    ws,
    current_row: int,
    nombre: str,
    identificador: str,
    turno: dict,
    turnos: list[dict],
    prestaciones: list[dict],
    usadas: set[int],
    mapa_especialidades: dict[str, list[str]],
    sistema_turnos: str,
    rojo,
    borde,
) -> int:
    requisitos = _requisitos_efectivos_auditoria(turno, turnos, mapa_especialidades, sistema_turnos)
    if not requisitos:
        return current_row

    for codigo, descripcion in requisitos.items():
        if _requisito_ya_cubierto_por_usadas(prestaciones, usadas, turno, codigo, descripcion):
            continue
        candidatos_disponibles = [
            (idx, item)
            for idx, item in _candidatos_requisito_auditoria(prestaciones, turno, codigo, descripcion)
            if (idx not in usadas or _ome_reutilizable_en_reporte(item))
            and not _ome_reservada_para_turno_exacto(
                item,
                turno,
                turnos,
                mapa_especialidades,
                codigo,
                descripcion,
                sistema_turnos,
            )
        ]
        if candidatos_disponibles:
            continue
        _append_reporte_row(
            ws,
            current_row,
            nombre,
            identificador,
            _turno_con_practica_faltante(turno, descripcion, sistema_turnos),
            _ome_faltante(codigo, descripcion),
            rojo,
            borde,
        )
        current_row += 1

    return current_row


def _marcar_requisitos_cubiertos(
    turno: dict,
    turnos: list[dict],
    ome: dict,
    requisitos_procesados: set[tuple],
    mapa_especialidades: dict[str, list[str]],
    sistema_turnos: str,
) -> None:
    paquete_key = _paquete_key_turno(turno, mapa_especialidades)
    requisitos = _requisitos_efectivos_auditoria(turno, turnos, mapa_especialidades, sistema_turnos)
    for codigo, descripcion in requisitos.items():
        cubre_requisito = _ome_cubre_requisito(ome, codigo, descripcion)
        if (
            not cubre_requisito
            and _sistema_turnos_key(sistema_turnos) == "CIMA"
            and _codigo_practica(str(ome.get("practica", "") or "")) == "570126"
            and "ecg" in _bases_esperadas_turno(
                turno,
                mapa_especialidades,
                sistema_turnos,
            )
        ):
            cubre_requisito = True
        if cubre_requisito:
            requisitos_procesados.add((paquete_key, codigo, descripcion))


def _ome_auditoria_transmitida(ome: dict | None) -> bool:
    if not isinstance(ome, dict):
        return False
    if ome.get("transmitida"):
        return True
    estado_accion = str(ome.get("estado_accion", "") or "").strip()
    if estado_accion in {"transmitida", "auditoria_transmitida", "fecha_auditoria_ok"}:
        return True
    transmitida_texto = str(ome.get("transmitida_texto", "") or "").strip().lower()
    return transmitida_texto.startswith("si")


def _fill_auditoria(ome: dict, verde, amarillo):
    return verde if _ome_auditoria_transmitida(ome) else amarillo


def _firma_prestacion_sobrante_auditoria(ome: dict) -> tuple:
    practica = str((ome or {}).get("practica", "") or "").strip()
    codigo = _codigo_practica(practica)
    practica_key = _texto_busqueda(practica)
    return (
        codigo or practica_key,
        _fecha_key((ome or {}).get("f_agenda", "")),
        _fecha_key((ome or {}).get("f_vencimiento", "")),
        str((ome or {}).get("transmitida_texto", "") or "").strip(),
        bool((ome or {}).get("mismo_efector")),
        str((ome or {}).get("estado_accion", "") or "").strip(),
    )


def _item_auditoria_fecha_distinta(ome: dict, fecha_turno_key: str, verde, amarillo):
    if ome.get("estado_accion") == "disponible_activar":
        return ome, amarillo
    fecha_item = _fecha_key(ome.get("f_agenda", ""))
    if _ome_auditoria_transmitida(ome) and _mismo_mes(fecha_item, fecha_turno_key):
        return _ome_fecha_auditoria_ok(ome), verde
    if not _ome_auditoria_transmitida(ome) and _mismo_mes(fecha_item, fecha_turno_key):
        return _ome_fecha_mismo_mes(ome), amarillo
    return _ome_fecha_no_coincide(ome), amarillo


def _valor_columna_transmitida(ws, ome: dict | None) -> str:
    if getattr(ws, "title", "") != "Auditoria prestaciones":
        return (ome or {}).get("f_vencimiento", "")
    if not _ome_auditoria_transmitida(ome):
        return ""
    transmitida_texto = str((ome or {}).get("transmitida_texto", "") or "").strip()
    if transmitida_texto and transmitida_texto.lower().startswith("si"):
        return transmitida_texto
    valor = str((ome or {}).get("f_vencimiento", "") or "").strip()
    if valor.lower().startswith("si"):
        return valor
    return "SI" if (ome or {}).get("transmitida") else ""


def _append_reporte_row(
    ws,
    row: int,
    paciente: str,
    identificador: str,
    turno: dict | None,
    ome: dict | None,
    fill,
    borde,
) -> None:
    sistema_turnos = getattr(ws, "_sistema_turnos", "CIMA")
    values = [
        paciente,
        _identificador_reporte(sistema_turnos, identificador, turno),
    ]
    if _es_sistema_cima(sistema_turnos):
        values.append(_tramite_reporte(sistema_turnos, turno))
    values.extend(
        [
            (turno or {}).get("fecha", ""),
            (turno or {}).get("profesional", ""),
            (ome or {}).get("f_agenda", ""),
            _valor_columna_transmitida(ws, ome),
            _texto_practica(
                ome,
                sin_turno_cima=turno is None and ome is not None,
                sistema_turnos=sistema_turnos,
            ),
        ]
    )
    for col, value in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col, value=value)
        cell.fill = fill
        cell.border = borde
        cell.alignment = Alignment(vertical="center", wrap_text=True)


def _crear_hoja_resumen_reporte(wb: Workbook, ws_origen, sistema_turnos: str) -> None:
    if "Resumen" in wb.sheetnames:
        del wb["Resumen"]
    indice_resumen = wb.sheetnames.index(ws_origen.title) + 1 if ws_origen.title in wb.sheetnames else None
    ws = wb.create_sheet("Resumen", index=indice_resumen)

    azul_oscuro = PatternFill("solid", fgColor="1F4E79")
    azul = PatternFill("solid", fgColor="2E75B6")
    rojo = PatternFill("solid", fgColor="FFC7CE")
    amarillo = PatternFill("solid", fgColor="FFF2CC")
    blanco = Font(color="FFFFFF", bold=True)
    borde = Border(*(Side(style="thin", color="D9E2F3") for _ in range(4)))

    sistema_key = str(sistema_turnos or "").strip().upper()
    es_gjs = sistema_key == "GJS"
    if es_gjs:
        titulo_pendientes = "Pacientes con documentacion / transmision pendiente"
        titulo_fecha = "Pacientes con fecha fuera del mes"
    else:
        titulo_pendientes = "Pacientes con OMEs disponibles para activar"
        titulo_fecha = f"Pacientes con fecha PAMI fuera del mes {sistema_turnos}"

    categorias = [
        ("faltantes", "Pacientes con faltante de OMEs", rojo),
        ("fecha_fuera_mes", titulo_fecha, amarillo),
    ]
    if es_gjs:
        categorias.insert(1, ("activar", "Pacientes con OMEs disponibles para activar", rojo))
        categorias.insert(2, ("otro_prestador", "Pacientes con OME activa para otro prestador", rojo))
        categorias.insert(3, ("pendientes", titulo_pendientes, amarillo))
    else:
        categorias.insert(1, ("pendientes", titulo_pendientes, amarillo))
    columnas_origen = _columnas_reporte(sistema_turnos)
    es_cima = _es_sistema_cima(sistema_turnos)
    detalle: dict[str, list[tuple[str, str, str, str, str, str]]] = {codigo: [] for codigo, _, _ in categorias}
    vistos: dict[str, set[tuple[str, str, str, str, str, str]]] = {codigo: set() for codigo, _, _ in categorias}

    paciente_actual = ""
    identificador_actual = ""
    tramite_actual = ""
    for row in range(3, ws_origen.max_row + 1):
        paciente = ws_origen.cell(row=row, column=columnas_origen["paciente"]).value
        if _clave_texto(str(paciente or "")) == "leyenda":
            break
        identificador = ws_origen.cell(row=row, column=columnas_origen["identificador"]).value
        tramite = (
            ws_origen.cell(row=row, column=columnas_origen["tramite"]).value
            if columnas_origen["tramite"]
            else ""
        )
        if paciente:
            paciente_actual = str(paciente)
        if identificador:
            identificador_actual = str(identificador)
        if tramite:
            tramite_actual = str(tramite)
        fecha_turno = ws_origen.cell(row=row, column=columnas_origen["fecha_turno"]).value
        profesional = str(ws_origen.cell(row=row, column=columnas_origen["profesional"]).value or "").strip()
        fecha_pami = ws_origen.cell(row=row, column=columnas_origen["fecha_pami"]).value
        practica = str(ws_origen.cell(row=row, column=columnas_origen["practica"]).value or "").strip()
        if not practica:
            continue
        categoria = _categoria_resumen_reporte(practica, sistema_turnos, fecha_turno, fecha_pami)
        if not categoria:
            continue
        firma = (paciente_actual, identificador_actual, tramite_actual, str(fecha_turno or ""), profesional, practica)
        if firma in vistos[categoria]:
            continue
        vistos[categoria].add(firma)
        detalle[categoria].append(firma)

    detalle_columnas = 6 if es_cima else 5
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=detalle_columnas)
    ws["A1"] = "Resumen de observaciones"
    ws["A1"].fill = azul_oscuro
    ws["A1"].font = blanco
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws.append([])
    ws.append(["Categoria", "Pacientes", "Registros"])
    for col in range(1, 4):
        cell = ws.cell(row=3, column=col)
        cell.fill = azul
        cell.font = blanco
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = borde

    row = 4
    for codigo, titulo, fill in categorias:
        pacientes_unicos = {(paciente, identificador) for paciente, identificador, _, _, _, _ in detalle[codigo]}
        ws.cell(row=row, column=1, value=titulo)
        ws.cell(row=row, column=2, value=len(pacientes_unicos))
        ws.cell(row=row, column=3, value=len(detalle[codigo]))
        for col in range(1, 4):
            cell = ws.cell(row=row, column=col)
            cell.fill = fill
            cell.border = borde
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        row += 1

    row += 2
    for codigo, titulo, fill in categorias:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=detalle_columnas)
        cell = ws.cell(row=row, column=1, value=titulo)
        cell.fill = azul_oscuro
        cell.font = blanco
        cell.alignment = Alignment(horizontal="center", vertical="center")
        row += 1
        ws.cell(row=row, column=1, value="Paciente")
        ws.cell(row=row, column=2, value="BENEF/DNI")
        if es_cima:
            ws.cell(row=row, column=3, value="NRO. TRAMITE")
            ws.cell(row=row, column=4, value="Fecha")
            ws.cell(row=row, column=5, value="Turno / sistema")
            ws.cell(row=row, column=6, value="Practica / estado")
        else:
            ws.cell(row=row, column=3, value="Fecha")
            ws.cell(row=row, column=4, value="Turno / sistema")
            ws.cell(row=row, column=5, value="Practica / estado")
        for col in range(1, detalle_columnas + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = azul
            cell.font = blanco
            cell.border = borde
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        row += 1
        if detalle[codigo]:
            for paciente, identificador, tramite, fecha_turno, profesional, practica in detalle[codigo]:
                ws.cell(row=row, column=1, value=paciente)
                ws.cell(row=row, column=2, value=identificador)
                if es_cima:
                    ws.cell(row=row, column=3, value=tramite)
                    ws.cell(row=row, column=4, value=fecha_turno)
                    ws.cell(row=row, column=5, value=profesional)
                    ws.cell(row=row, column=6, value=practica)
                else:
                    ws.cell(row=row, column=3, value=fecha_turno)
                    ws.cell(row=row, column=4, value=profesional)
                    ws.cell(row=row, column=5, value=practica)
                for col in range(1, detalle_columnas + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.fill = fill
                    cell.border = borde
                    cell.alignment = Alignment(vertical="center", wrap_text=True)
                row += 1
        else:
            ws.cell(row=row, column=1, value="Sin registros")
            for col in range(1, detalle_columnas + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = borde
                cell.alignment = Alignment(vertical="center", wrap_text=True)
            row += 1
        row += 2

    widths = (
        {"A": 32, "B": 22, "C": 18, "D": 14, "E": 48, "F": 78}
        if es_cima
        else {"A": 32, "B": 22, "C": 14, "D": 48, "E": 78}
    )
    for column, width in widths.items():
        ws.column_dimensions[column].width = width
    ws.freeze_panes = "A4"


def _categoria_resumen_reporte(
    practica: str,
    sistema_turnos: str,
    fecha_turno=None,
    fecha_pami=None,
) -> str | None:
    texto = _texto_busqueda(practica)
    if not texto:
        return None
    if "fecha pami distinta dentro del mes" in texto:
        return None
    if "falta ome" in texto or "sin prestacion en transmision" in texto or "sin omes en pami" in texto:
        return "faltantes"
    if "fecha pami no coincide" in texto:
        if fecha_turno and fecha_pami and _mismo_mes(str(fecha_turno), str(fecha_pami)):
            return None
        return "fecha_fuera_mes"
    if (
        "disponible para activar" in texto
        or "pendiente de activar" in texto
        or "falta activar" in texto
    ):
        if _sistema_turnos_key(sistema_turnos) == "GJS":
            return "activar"
        return "pendientes"
    if (
        "falta documentacion" in texto
        or "falta validar" in texto
        or "no validada" in texto
        or "no transmitida" in texto
    ):
        return "pendientes"
    if _sistema_turnos_key(sistema_turnos) == "GJS" and (
        "activada otro prestador" in texto
        or "otro prestador" in texto
    ):
        return "otro_prestador"
    return None


def _texto_practica(ome: dict | None, sin_turno_cima: bool = False, sistema_turnos: str = "CIMA") -> str:
    if not ome:
        return ""
    orden = str(ome.get("n_orden", "") or "").strip()
    practica = str(ome.get("practica", "") or "").strip()
    detalle = practica or orden
    notas = []
    estado_accion = str(ome.get("estado_accion", "") or "").strip()
    sistema_key = _sistema_turnos_key(sistema_turnos)
    es_cima = sistema_key == "CIMA"
    if estado_accion == "faltante":
        return detalle
    if estado_accion == "sin_resultados_pami":
        return detalle
    if estado_accion == "fecha_no_coincide":
        notas.append(f"fecha PAMI no coincide con {sistema_turnos}")
        if ome.get("mismo_efector"):
            notas.append("asignada a nuestro centro en otra fecha" if es_cima else "nuestro centro")
        else:
            notas.append("activada por otro prestador" if es_cima else "otro prestador")
    elif estado_accion == "fecha_auditoria_ok":
        notas.append(f"fecha PAMI no coincide con {sistema_turnos}")
        notas.append("nuestro centro" if ome.get("mismo_efector") else "otro prestador")
        if ome.get("transmitida"):
            transmitida_texto = str(ome.get("transmitida_texto", "") or "").strip()
            notas.append(transmitida_texto if transmitida_texto else "ya transmitida")
    elif estado_accion == "fecha_mismo_mes":
        notas.append(f"fecha PAMI distinta, dentro del mes {sistema_turnos}")
        if ome.get("mismo_efector"):
            notas.append("asignada a nuestro centro en otra fecha" if es_cima else "nuestro centro")
        else:
            notas.append("activada por otro prestador" if es_cima else "otro prestador")
    elif estado_accion == "practica_relacionada":
        notas.append("practica relacionada por codigo, revisar detalle")
    elif estado_accion == "especialidad_sustituida":
        notas.append("OME transmitida de otra especialidad, asumida por coincidencia exacta de fecha")
    elif estado_accion == "disponible_activar":
        notas.append("pendiente de activar" if es_cima else "disponible para activar")
        if sistema_key == "GJS":
            notas.append("Falta activar")
    elif estado_accion == "transmitida":
        notas.append("ya transmitida")
    elif estado_accion == "auditoria_panel_pendiente":
        notas.append("turno asignado")
        notas.append("falta validar/transmitir")
    elif estado_accion in {"auditoria_transmitida", "auditoria_validada", "auditoria_pendiente"}:
        if ome.get("validada"):
            notas.append("validada")
        else:
            notas.append("no validada")
        mostrar_documentacion = _mostrar_documentacion_en_reporte(ome, sistema_turnos)
        if mostrar_documentacion and ome.get("doc_cargada"):
            notas.append("documentacion cargada")
        elif mostrar_documentacion and ome.get("doc_pendiente"):
            notas.append("falta documentacion para transmitir")
        if ome.get("transmitida"):
            transmitida_texto = str(ome.get("transmitida_texto", "") or "").strip()
            notas.append(transmitida_texto if transmitida_texto else "transmitida")
        else:
            notas.append("no transmitida")
    elif not ome.get("mismo_efector"):
        notas.append("Activada otro prestador" if sistema_key == "GJS" else "activada por otro prestador")
    elif sin_turno_cima:
        notas.append("nuestro centro")
    if ome.get("turno_asignado"):
        notas.append("turno asignado")
    if sin_turno_cima:
        notas.append(f"Falta bono {sistema_turnos}" if sistema_turnos == "GJS" else f"sin turno {sistema_turnos}")
    if notas:
        return f"{detalle} ({', '.join(notas)})" if detalle else f"({', '.join(notas)})"
    return detalle


def _mostrar_documentacion_en_reporte(ome: dict | None, sistema_turnos: str) -> bool:
    if sistema_turnos != "GJS":
        return True
    practica = str((ome or {}).get("practica", "") or "")
    texto = _texto_busqueda(practica)
    codigo = _codigo_practica(practica)
    if codigo in {"570129", "820113"} or ("cardiologia" in texto and "electrocardiograma" in texto):
        return True
    es_consulta = codigo.startswith("820") or "consulta con especialista" in texto
    return not es_consulta


def _paciente_excluido_reporte_no_pami(paciente: dict, sistema_turnos: str) -> bool:
    return bool((paciente or {}).get("no_pami"))


def _identificador_paciente(paciente: dict) -> str:
    beneficiario = str(paciente.get("beneficiario", "") or "").strip()
    beneficiario_pami = str(paciente.get("beneficiario_pami", "") or "").strip()
    dni = str(paciente.get("dni", "") or "").strip()
    no_pami = bool(paciente.get("no_pami"))
    if beneficiario:
        suffix = ""
        if paciente.get("verificar_benef"):
            suffix = f" (Verificar BENEF; PAMI: {beneficiario_pami})" if beneficiario_pami else " (Verificar BENEF)"
        elif no_pami:
            suffix = " (NO PAMI*)"
        return f"{beneficiario}{suffix}"
    if dni:
        suffix = " (NO PAMI*)" if no_pami else ""
        return f"DNI {dni}{suffix}"
    key = str(paciente.get("key", "") or "").strip()
    identificador = key.replace("DNI_", "DNI ", 1)
    return f"{identificador} (NO PAMI*)" if no_pami and identificador else identificador


def build_default_report_path(cliente_codigo: str | None = CLIENTE_CIMA) -> Path:
    cliente = _cliente_config(cliente_codigo)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    sufijo_cliente = "" if cliente["codigo"] == CLIENTE_CIMA else f"_{cliente['codigo']}"
    return get_output_dir() / f"verificacion_omes{sufijo_cliente}_{stamp}.xlsx"
