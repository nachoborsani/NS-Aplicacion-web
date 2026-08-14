from pathlib import Path

from openpyxl import Workbook

from pami_scraper import FUENTE_CARTILLA, MODO_DNI


def guardar_modelo_padron(ruta_excel: str | Path, modo_busqueda: str, fuente_consulta: str = "") -> Path:
    ruta = Path(ruta_excel)
    ruta.parent.mkdir(parents=True, exist_ok=True)

    libro = Workbook()
    hoja = libro.active
    hoja.title = "Padron"

    if fuente_consulta == FUENTE_CARTILLA:
        hoja.append(["beneficio", "dni"])
        hoja.append(["15000000000000", "12345678"])
        hoja.column_dimensions["A"].width = 22
        hoja.column_dimensions["B"].width = 14
    else:
        columna = "dni" if modo_busqueda == MODO_DNI else "beneficio"
        ejemplo = "12345678" if modo_busqueda == MODO_DNI else "15000000000000"
        hoja.append([columna])
        hoja.append([ejemplo])
        hoja.column_dimensions["A"].width = 24

    libro.save(ruta)
    return ruta


def guardar_modelo_credencial(ruta_excel: str | Path) -> Path:
    ruta = Path(ruta_excel)
    ruta.parent.mkdir(parents=True, exist_ok=True)

    libro = Workbook()
    hoja = libro.active
    hoja.title = "Credenciales"
    hoja.append(["benef", "dni", "tramite", "sexo"])
    hoja.append(["15000000000000", "12345678", "00123456789", "MASC"])
    hoja.column_dimensions["A"].width = 22
    hoja.column_dimensions["B"].width = 14
    hoja.column_dimensions["C"].width = 18
    hoja.column_dimensions["D"].width = 12
    libro.save(ruta)
    return ruta


def guardar_modelo_ome(ruta_excel: str | Path) -> Path:
    ruta = Path(ruta_excel)
    ruta.parent.mkdir(parents=True, exist_ok=True)

    libro = Workbook()
    hoja = libro.active
    hoja.title = "OME"
    hoja.append(["modo", "afiliado", "diagnostico", "practica"])
    hoja.append(["BENEF", "15000000000000", "Z000", "427109"])
    hoja.append(["BENEF", "15000000000001", "Z000", "427122"])
    hoja.append(["DNI", "12345678", "Z000", "427109"])
    hoja.column_dimensions["A"].width = 12
    hoja.column_dimensions["B"].width = 22
    hoja.column_dimensions["C"].width = 16
    hoja.column_dimensions["D"].width = 16
    libro.save(ruta)
    return ruta
