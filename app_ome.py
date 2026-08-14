import csv
import json
import os
import queue
import subprocess
import tempfile
import threading
import unicodedata
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from tkinter import filedialog, messagebox, ttk
from uuid import uuid4

import customtkinter as ctk
from openpyxl import load_workbook

import tkinter as tk

from app_credentials import sync_profile_records, upsert_shared_credentials_from_records
from app_paths import get_data_dir, get_log_file, get_output_dir, get_resource_path
from excel_models import guardar_modelo_ome
from gmail_informes import get_gmail_credentials_path


class _Tooltip:
    """Tooltip simple para botones-icono (CustomTkinter no trae uno)."""

    def __init__(self, widget, text: str) -> None:
        self.widget = widget
        self.text = text
        self.tip = None
        widget.bind("<Enter>", self._show, add="+")
        widget.bind("<Leave>", self._hide, add="+")
        widget.bind("<Destroy>", self._hide, add="+")

    def _show(self, _event=None) -> None:
        if self.tip is not None or not self.text:
            return
        try:
            x = self.widget.winfo_rootx() + 12
            y = self.widget.winfo_rooty() + self.widget.winfo_height() + 6
            self.tip = tk.Toplevel(self.widget)
            self.tip.wm_overrideredirect(True)
            self.tip.wm_geometry(f"+{x}+{y}")
            tk.Label(
                self.tip, text=self.text, background="#16324f", foreground="#ffffff",
                font=("Segoe UI", 9), padx=8, pady=3, bd=0,
            ).pack()
        except Exception:
            self.tip = None

    def _hide(self, _event=None) -> None:
        if self.tip is not None:
            try:
                self.tip.destroy()
            except Exception:
                pass
            self.tip = None


def attach_tooltip(widget, text: str) -> None:
    _Tooltip(widget, text)
from google_sheets_ome import (
    OFFICE_FILE_MESSAGE,
    extract_spreadsheet_id,
    get_connected_google_email,
    get_sheets_token_path,
    is_office_file_url,
    list_spreadsheet_sheet_names,
    normalize_spreadsheet_url,
    read_ome_sheet_rows,
    write_ome_sheet_results,
)
from pami_plan_salud_resolver import (
    explain_unresolved_plan_salud_practice,
    resolve_plan_salud_practices,
)
from pami_ome import PamiOmeController
from pami_ome_generator import (
    build_default_output_path,
    export_results_csv_to_excel,
    run_batch_sync,
)


MED_CABECERA_PRACTICA_OPTIONS = ["427109", "427120", "427121", "427122"]
MANUAL_PRACTICE_OPTION = "MANUAL"
DEFAULT_DIAGNOSTICO_OPTIONS = [
    "Z000 - Examen medico general",
    "Z001 - Examen de ojos y vision",
    "Z002 - Examen dental",
    "Z009 - Examen medico no especificado",
    "I10 - Hipertension esencial primaria",
    "I119 - Enfermedad cardiaca hipertensiva sin insuficiencia cardiaca",
    "E119 - Diabetes mellitus tipo 2 sin complicaciones",
    "E149 - Diabetes mellitus no especificada sin complicaciones",
    "I259 - Enfermedad isquemica cronica del corazon no especificada",
    "I489 - Fibrilacion auricular no especificada",
    "I509 - Insuficiencia cardiaca no especificada",
    "J449 - EPOC no especificada",
    "J459 - Asma no especificada",
    "M819 - Osteoporosis no especificada",
    "M179 - Artrosis de rodilla no especificada",
    "M549 - Dorsalgia no especificada",
    "F329 - Episodio depresivo no especificado",
    "F419 - Trastorno de ansiedad no especificado",
    "G309 - Alzheimer no especificado",
    "G209 - Parkinson no especificado",
    "E039 - Hipotiroidismo no especificado",
    "E059 - Tirotoxicosis no especificada",
]

DEFAULT_NOMENCLADOR_FILENAME = "nomenclador_pami.xlsx"
DEFAULT_NOMENCLADOR_ALT_FILENAME = "nomenclador_pami.xls"
LEGACY_NOMENCLADOR_FILENAME = "73-Nomenclador Pami valorizado 04-2026.xlsx"
DEFAULT_NOMENCLADOR_PATH = get_resource_path(DEFAULT_NOMENCLADOR_FILENAME)
DEFAULT_NOMENCLADOR_ALT_PATH = get_resource_path(DEFAULT_NOMENCLADOR_ALT_FILENAME)
LEGACY_NOMENCLADOR_PATH = get_resource_path(LEGACY_NOMENCLADOR_FILENAME)
LEGACY_EXTERNAL_NOMENCLADOR_PATH = Path(r"E:\NomencladoresPAMI\73-Nomenclador Pami valorizado 04-2026.xlsx")
LEGACY_NOMENCLADOR_ALT_PATH = Path(r"E:\NomencladoresPAMI\nomenclador_pami.xls")
DESKTOP_NOMENCLADOR_ALT_PATH = Path.home() / "Desktop" / DEFAULT_NOMENCLADOR_ALT_FILENAME
HIDDEN_MODULE_KEYWORDS = (
    "internacion",
    "alta complejidad",
    "unidad coronaria",
    "terapia intensiva",
    "hospital de dia",
    "prestacion sanatorial",
    "modulo global",
    "hemodialisis",
    "dialisis",
    "trasplante",
    "internación",
)

SPECIALIST_PRACTICE_CODE_HINTS = {
    "gastroenterologia": "820139",
    "gatroenterologia": "820139",
    "urologia": "820167",
    "urolologia": "820167",
    "diabetologia": "820171",
    "dermatologia": "820116",
    "reumatologia": "820163",
    "otorrinolaringologia": "820168",
    "traumatologia": "820165",
    "traumatologia y ortopedia": "820165",
    "flebologia": "820143",
    "hematologia": "820121",
    "m a p a": "570120",
    "mapa": "570120",
    "presurometria": "570120",
    "ecografia partes blandas": "186001",
    "ecografia de partes blandas": "186001",
    "partes blandas": "186001",
}


def _default_nomenclador_candidates() -> list[Path]:
    candidates = [
        DEFAULT_NOMENCLADOR_PATH,
        DEFAULT_NOMENCLADOR_ALT_PATH,
        DESKTOP_NOMENCLADOR_ALT_PATH,
        LEGACY_NOMENCLADOR_PATH,
        LEGACY_NOMENCLADOR_ALT_PATH,
        LEGACY_EXTERNAL_NOMENCLADOR_PATH,
    ]
    unique_candidates: list[Path] = []
    seen: set[str] = set()
    for path in candidates:
        resolved = str(path)
        if resolved in seen:
            continue
        seen.add(resolved)
        unique_candidates.append(path)
    return unique_candidates


@dataclass(frozen=True)
class OmeModuleConfig:
    module_key: str
    module_title: str
    module_subtitle: str
    profile_filename: str
    template_filename: str
    practice_options: tuple[str, ...]
    diagnosis_options: tuple[str, ...]
    default_practice_code: str = ""
    show_open_button: bool = True
    show_visible_batch_button: bool = True
    full_batch_button_text: str = "Ejecutar lote OME"
    show_stop_button: bool = False
    show_clear_panel_button: bool = False
    show_completion_alert: bool = False
    show_mass_load_panel: bool = True
    link_sheet_profiles_to_pami: bool = True


@dataclass(frozen=True)
class PracticeCatalogItem:
    module_id: str
    module_name: str
    code: str
    description: str

    @property
    def display(self) -> str:
        return f"{self.code} - {self.description}"


MED_CABECERA_CONFIG = OmeModuleConfig(
    module_key="ome_med_cabecera",
    module_title="Generar OME Med Cabecera",
    module_subtitle="Carga masiva por listas y edicion por paciente. El resultado final visible se guarda en Excel.",
    profile_filename="usuarios_ome_med_cabecera.json",
    template_filename="modelo_ome_med_cabecera.xlsx",
    practice_options=tuple(MED_CABECERA_PRACTICA_OPTIONS),
    diagnosis_options=tuple(DEFAULT_DIAGNOSTICO_OPTIONS),
    default_practice_code="427122",
    show_open_button=False,
    show_visible_batch_button=False,
    full_batch_button_text="Ejecutar BOT",
    show_stop_button=True,
    show_clear_panel_button=True,
    show_completion_alert=False,
    show_mass_load_panel=False,
)

ESPECIALISTA_CONFIG = OmeModuleConfig(
    module_key="ome_especialista",
    module_title="Generar OME Especialista",
    module_subtitle="Carga masiva por listas y edicion por paciente. El resultado final visible se guarda en Excel.",
    profile_filename="usuarios_ome_especialista.json",
    template_filename="modelo_ome_especialista.xlsx",
    practice_options=tuple(MED_CABECERA_PRACTICA_OPTIONS),
    diagnosis_options=tuple(DEFAULT_DIAGNOSTICO_OPTIONS),
    default_practice_code="427122",
    show_open_button=False,
    show_visible_batch_button=False,
    full_batch_button_text="Ejecutar BOT",
    show_stop_button=True,
    show_clear_panel_button=True,
    show_completion_alert=True,
    link_sheet_profiles_to_pami=False,
)


class PracticeCatalogDialog(ctk.CTkToplevel):
    def __init__(
        self,
        master,
        *,
        module_title: str,
        catalog_path: Path,
        items: list[PracticeCatalogItem],
        active_modules: set[str],
        active_practices: set[str],
    ) -> None:
        super().__init__(master)
        self.title(f"Codigos disponibles - {module_title}")
        self.geometry("1180x720")
        self.minsize(980, 620)
        self.transient(master)
        self.grab_set()

        self.catalog_path = catalog_path
        self.items = items
        self.active_modules = set(active_modules)
        self.active_practices = set(active_practices)
        self.result: dict | None = None

        self.module_items: dict[str, list[PracticeCatalogItem]] = {}
        for item in self.items:
            self.module_items.setdefault(item.module_id, []).append(item)
        self.module_order = list(self.module_items.keys())
        self.selected_module_id = self.module_order[0] if self.module_order else None
        self.module_search_var = ctk.StringVar(value="")
        self.search_var = ctk.StringVar(value="")
        self.marked_practices: set[str] = set()

        style = ttk.Style(self)
        style.configure(
            "Catalog.Treeview",
            rowheight=24,
            font=("Segoe UI", 10),
            background="#ffffff",
            foreground="#16324f",
            fieldbackground="#ffffff",
            borderwidth=0,
        )
        style.configure(
            "Catalog.Treeview.Heading",
            font=("Segoe UI", 10, "bold"),
            background="#edf3f9",
            foreground="#16324f",
            relief="flat",
        )
        style.map("Catalog.Treeview", background=[("selected", "#245b9d")], foreground=[("selected", "#ffffff")])

        self._build_ui()
        self._refresh_modules_tree()
        self._refresh_practices_tree()

    @staticmethod
    def _normalize_search_text(value: str) -> str:
        normalized = unicodedata.normalize("NFKD", value or "")
        ascii_text = normalized.encode("ascii", "ignore").decode("ascii")
        return " ".join(ascii_text.lower().split())

    def _build_ui(self) -> None:
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=1)

        top = ctk.CTkFrame(self, fg_color="#eef3f8", corner_radius=12)
        top.grid(row=0, column=0, padx=14, pady=(14, 8), sticky="ew")
        top.grid_columnconfigure(0, weight=1)
        ctk.CTkLabel(
            top,
            text="Activa modulos y códigos del nomenclador. Doble click para alternar un item.",
            font=ctk.CTkFont(size=13, weight="bold"),
            text_color="#16324f",
        ).grid(row=0, column=0, padx=12, pady=(10, 4), sticky="w")
        ctk.CTkLabel(
            top,
            text=str(self.catalog_path),
            font=ctk.CTkFont(size=11),
            text_color="#51657a",
        ).grid(row=1, column=0, padx=12, pady=(0, 10), sticky="w")

        body = ctk.CTkFrame(self, fg_color="transparent")
        body.grid(row=1, column=0, padx=14, pady=(0, 8), sticky="nsew")
        body.grid_columnconfigure(0, weight=1)
        body.grid_columnconfigure(1, weight=2)
        body.grid_rowconfigure(0, weight=1)

        modules_box = ctk.CTkFrame(body, fg_color="#f4f8fb", corner_radius=12)
        modules_box.grid(row=0, column=0, padx=(0, 8), pady=0, sticky="nsew")
        modules_box.grid_columnconfigure(0, weight=1)
        modules_box.grid_rowconfigure(2, weight=1)
        ctk.CTkLabel(
            modules_box,
            text="Modulos visibles",
            font=ctk.CTkFont(size=14, weight="bold"),
            text_color="#16324f",
        ).grid(row=0, column=0, padx=12, pady=(10, 6), sticky="w")

        module_search_row = ctk.CTkFrame(modules_box, fg_color="transparent")
        module_search_row.grid(row=1, column=0, columnspan=2, padx=12, pady=(0, 8), sticky="ew")
        module_search_row.grid_columnconfigure(1, weight=1)
        ctk.CTkLabel(
            module_search_row,
            text="Buscar",
            font=ctk.CTkFont(size=12, weight="bold"),
            text_color="#16324f",
        ).grid(row=0, column=0, padx=(0, 8), pady=0, sticky="w")
        module_search_entry = ctk.CTkEntry(
            module_search_row,
            textvariable=self.module_search_var,
            placeholder_text="Modulo",
        )
        module_search_entry.grid(row=0, column=1, padx=0, pady=0, sticky="ew")
        module_search_entry.bind("<KeyRelease>", lambda _event: self._refresh_modules_tree(), add="+")

        self.modules_tree = ttk.Treeview(
            modules_box,
            columns=("activo", "modulo", "seleccion"),
            show="headings",
            height=16,
            style="Catalog.Treeview",
        )
        self.modules_tree.grid(row=2, column=0, padx=(12, 0), pady=(0, 8), sticky="nsew")
        self.modules_tree.heading("activo", text="Activo")
        self.modules_tree.heading("modulo", text="Modulo")
        self.modules_tree.heading("seleccion", text="Códigos")
        self.modules_tree.column("activo", width=70, anchor="center")
        self.modules_tree.column("modulo", width=250, anchor="w")
        self.modules_tree.column("seleccion", width=90, anchor="center")
        modules_scroll = ttk.Scrollbar(modules_box, orient="vertical", command=self.modules_tree.yview)
        modules_scroll.grid(row=2, column=1, padx=(0, 12), pady=(0, 8), sticky="ns")
        self.modules_tree.configure(yscrollcommand=modules_scroll.set)
        self.modules_tree.bind("<<TreeviewSelect>>", self._on_module_selected, add="+")
        self.modules_tree.bind("<Double-1>", self._toggle_selected_module, add="+")

        module_buttons = ctk.CTkFrame(modules_box, fg_color="transparent")
        module_buttons.grid(row=3, column=0, columnspan=2, padx=12, pady=(0, 10), sticky="ew")
        for col in range(3):
            module_buttons.grid_columnconfigure(col, weight=1, uniform="module_actions")
        ctk.CTkButton(module_buttons, text="Alternar modulo", command=self._toggle_selected_module).grid(
            row=0, column=0, padx=(0, 6), pady=0, sticky="ew"
        )
        ctk.CTkButton(module_buttons, text="Activar todos", command=self._activate_all_modules).grid(
            row=0, column=1, padx=6, pady=0, sticky="ew"
        )
        ctk.CTkButton(module_buttons, text="Desactivar todos", command=self._deactivate_all_modules).grid(
            row=0, column=2, padx=(6, 0), pady=0, sticky="ew"
        )

        practices_box = ctk.CTkFrame(body, fg_color="#f4f8fb", corner_radius=12)
        practices_box.grid(row=0, column=1, padx=(8, 0), pady=0, sticky="nsew")
        practices_box.grid_columnconfigure(0, weight=1)
        practices_box.grid_rowconfigure(2, weight=1)
        ctk.CTkLabel(
            practices_box,
            text="Códigos del modulo seleccionado",
            font=ctk.CTkFont(size=14, weight="bold"),
            text_color="#16324f",
        ).grid(row=0, column=0, padx=12, pady=(10, 6), sticky="w")

        search_row = ctk.CTkFrame(practices_box, fg_color="transparent")
        search_row.grid(row=1, column=0, padx=12, pady=(0, 8), sticky="ew")
        search_row.grid_columnconfigure(1, weight=1)
        ctk.CTkLabel(search_row, text="Buscar", font=ctk.CTkFont(size=12, weight="bold"), text_color="#16324f").grid(
            row=0, column=0, padx=(0, 8), pady=0, sticky="w"
        )
        search_entry = ctk.CTkEntry(search_row, textvariable=self.search_var, placeholder_text="Código o descripción")
        search_entry.grid(row=0, column=1, padx=0, pady=0, sticky="ew")
        search_entry.bind("<KeyRelease>", lambda _event: self._refresh_practices_tree(), add="+")

        self.practices_tree = ttk.Treeview(
            practices_box,
            columns=("marcado", "activo", "codigo", "descripcion"),
            show="headings",
            height=16,
            selectmode="extended",
            style="Catalog.Treeview",
        )
        self.practices_tree.grid(row=2, column=0, padx=(12, 0), pady=(0, 8), sticky="nsew")
        self.practices_tree.heading("marcado", text="Sel")
        self.practices_tree.heading("activo", text="Activo")
        self.practices_tree.heading("codigo", text="Código")
        self.practices_tree.heading("descripcion", text="Descripción")
        self.practices_tree.column("marcado", width=52, anchor="center")
        self.practices_tree.column("activo", width=70, anchor="center")
        self.practices_tree.column("codigo", width=90, anchor="center")
        self.practices_tree.column("descripcion", width=420, anchor="w")
        practices_scroll = ttk.Scrollbar(practices_box, orient="vertical", command=self.practices_tree.yview)
        practices_scroll.grid(row=2, column=1, padx=(0, 12), pady=(0, 8), sticky="ns")
        self.practices_tree.configure(yscrollcommand=practices_scroll.set)
        self.practices_tree.bind("<Button-1>", self._handle_practice_tree_click, add="+")
        self.practices_tree.bind("<Double-1>", self._toggle_selected_practice, add="+")

        practice_buttons = ctk.CTkFrame(practices_box, fg_color="transparent")
        practice_buttons.grid(row=3, column=0, columnspan=2, padx=12, pady=(0, 10), sticky="ew")
        for col in range(5):
            practice_buttons.grid_columnconfigure(col, weight=1, uniform="practice_actions")
        ctk.CTkButton(practice_buttons, text="Marcar selección", command=self._mark_selected_practices).grid(
            row=0, column=0, padx=(0, 6), pady=0, sticky="ew"
        )
        ctk.CTkButton(practice_buttons, text="Desmarcar selección", command=self._unmark_selected_practices).grid(
            row=0, column=1, padx=6, pady=0, sticky="ew"
        )
        ctk.CTkButton(practice_buttons, text="Activar / desactivar código", command=self._toggle_selected_practice).grid(
            row=0, column=2, padx=6, pady=0, sticky="ew"
        )
        ctk.CTkButton(practice_buttons, text="Activar marcados", command=self._activate_marked_practices).grid(
            row=0, column=3, padx=6, pady=0, sticky="ew"
        )
        ctk.CTkButton(practice_buttons, text="Desactivar marcados", command=self._deactivate_marked_practices).grid(
            row=0, column=4, padx=(6, 0), pady=0, sticky="ew"
        )

        bottom = ctk.CTkFrame(self, fg_color="transparent")
        bottom.grid(row=2, column=0, padx=14, pady=(0, 14), sticky="ew")
        bottom.grid_columnconfigure(0, weight=1)
        bottom.grid_columnconfigure(1, weight=0)
        bottom.grid_columnconfigure(2, weight=0)
        ctk.CTkButton(bottom, text="Cancelar", fg_color="#9aafc3", hover_color="#7f95aa", command=self.destroy).grid(
            row=0, column=1, padx=(0, 8), pady=0, sticky="e"
        )
        ctk.CTkButton(bottom, text="Guardar", command=self._save_and_close).grid(
            row=0, column=2, padx=0, pady=0, sticky="e"
        )

    def _refresh_modules_tree(self) -> None:
        current = self.selected_module_id
        search = self._normalize_search_text(self.module_search_var.get())
        for item_id in self.modules_tree.get_children():
            self.modules_tree.delete(item_id)
        filtered_module_ids: list[str] = []
        for module_id in self.module_order:
            module_items = self.module_items[module_id]
            module_name = module_items[0].module_name
            if search and search not in self._normalize_search_text(module_name):
                continue
            active_count = sum(1 for item in module_items if item.code in self.active_practices)
            total_count = len(module_items)
            values = (
                "Si" if module_id in self.active_modules else "No",
                module_name,
                f"{active_count}/{total_count}",
            )
            self.modules_tree.insert("", "end", iid=module_id, values=values)
            filtered_module_ids.append(module_id)
        if current and current in self.modules_tree.get_children():
            self.modules_tree.selection_set(current)
            self.modules_tree.focus(current)
        elif filtered_module_ids:
            self.selected_module_id = filtered_module_ids[0]
            self.modules_tree.selection_set(self.selected_module_id)
            self.modules_tree.focus(self.selected_module_id)
        else:
            self.selected_module_id = None
        self._refresh_practices_tree()

    def _refresh_practices_tree(self) -> None:
        for item_id in self.practices_tree.get_children():
            self.practices_tree.delete(item_id)
        module_id = self.selected_module_id
        if not module_id:
            return
        search = self.search_var.get().strip().lower()
        for item in self.module_items.get(module_id, []):
            haystack = f"{item.code} {item.description}".lower()
            if search and search not in haystack:
                continue
            self.practices_tree.insert(
                "",
                "end",
                iid=item.code,
                values=(
                    "â˜‘" if item.code in self.marked_practices else "â˜",
                    "Si" if item.code in self.active_practices else "No",
                    item.code,
                    item.description,
                ),
            )

    def _on_module_selected(self, _event=None) -> None:
        selected = self.modules_tree.selection()
        if selected:
            self.selected_module_id = selected[0]
            self._refresh_practices_tree()

    def _handle_practice_tree_click(self, event) -> str | None:
        item_id = self.practices_tree.identify_row(event.y)
        column_id = self.practices_tree.identify_column(event.x)
        if not item_id:
            return None
        if column_id == "#1":
            if item_id in self.marked_practices:
                self.marked_practices.discard(item_id)
            else:
                self.marked_practices.add(item_id)
            self._refresh_practices_tree()
            self.practices_tree.selection_set(item_id)
            return "break"
        return None

    def _toggle_selected_module(self, _event=None) -> None:
        module_id = self.selected_module_id or next(iter(self.modules_tree.selection()), None)
        if not module_id:
            return
        module_items = self.module_items.get(module_id, [])
        if module_id in self.active_modules:
            self.active_modules.discard(module_id)
            for item in module_items:
                self.active_practices.discard(item.code)
        else:
            self.active_modules.add(module_id)
            for item in module_items:
                self.active_practices.add(item.code)
        self._refresh_modules_tree()
        self._refresh_practices_tree()

    def _activate_all_modules(self) -> None:
        self.active_modules = set(self.module_order)
        self.active_practices = {item.code for item in self.items}
        self._refresh_modules_tree()
        self._refresh_practices_tree()

    def _deactivate_all_modules(self) -> None:
        self.active_modules.clear()
        self.active_practices.clear()
        self._refresh_modules_tree()
        self._refresh_practices_tree()

    def _toggle_selected_practice(self, _event=None) -> None:
        selected = self.practices_tree.selection()
        if not selected:
            return
        module_id = self.selected_module_id
        if not module_id:
            return
        for code in selected:
            if code in self.active_practices:
                self.active_practices.discard(code)
            else:
                self.active_practices.add(code)
                self.active_modules.add(module_id)
        if not any(item.code in self.active_practices for item in self.module_items.get(module_id, [])):
            self.active_modules.discard(module_id)
        self._refresh_modules_tree()
        self._refresh_practices_tree()
        if selected:
            self.practices_tree.selection_set(selected)

    def _mark_selected_practices(self) -> None:
        selected = self.practices_tree.selection()
        if not selected:
            return
        self.marked_practices.update(selected)
        self._refresh_practices_tree()
        self.practices_tree.selection_set(selected)

    def _unmark_selected_practices(self) -> None:
        selected = self.practices_tree.selection()
        if not selected:
            return
        for code in selected:
            self.marked_practices.discard(code)
        self._refresh_practices_tree()
        self.practices_tree.selection_set(selected)

    def _activate_marked_practices(self) -> None:
        module_id = self.selected_module_id
        if not module_id or not self.marked_practices:
            return
        visible_codes = {self.practices_tree.item(item_id, "values")[2] for item_id in self.practices_tree.get_children()}
        target_codes = self.marked_practices.intersection(visible_codes)
        if not target_codes:
            return
        self.active_modules.add(module_id)
        self.active_practices.update(target_codes)
        self._refresh_modules_tree()
        self._refresh_practices_tree()

    def _deactivate_marked_practices(self) -> None:
        module_id = self.selected_module_id
        if not module_id or not self.marked_practices:
            return
        visible_codes = {self.practices_tree.item(item_id, "values")[2] for item_id in self.practices_tree.get_children()}
        target_codes = self.marked_practices.intersection(visible_codes)
        if not target_codes:
            return
        for code in target_codes:
            self.active_practices.discard(code)
        if not any(item.code in self.active_practices for item in self.module_items.get(module_id, [])):
            self.active_modules.discard(module_id)
        self._refresh_modules_tree()
        self._refresh_practices_tree()

    def _activate_visible_practices(self) -> None:
        module_id = self.selected_module_id
        if not module_id:
            return
        visible_codes = [self.practices_tree.item(item_id, "values")[1] for item_id in self.practices_tree.get_children()]
        if not visible_codes:
            return
        self.active_modules.add(module_id)
        self.active_practices.update(visible_codes)
        self._refresh_modules_tree()
        self._refresh_practices_tree()

    def _deactivate_visible_practices(self) -> None:
        module_id = self.selected_module_id
        if not module_id:
            return
        for item_id in self.practices_tree.get_children():
            code = self.practices_tree.item(item_id, "values")[1]
            self.active_practices.discard(code)
        if not any(item.code in self.active_practices for item in self.module_items.get(module_id, [])):
            self.active_modules.discard(module_id)
        self._refresh_modules_tree()
        self._refresh_practices_tree()

    def _save_and_close(self) -> None:
        self.result = {
            "catalog_path": str(self.catalog_path),
            "active_modules": sorted(self.active_modules),
            "active_practices": sorted(self.active_practices),
        }
        self.destroy()


class StopBatchDialog(ctk.CTkToplevel):
    def __init__(self, master) -> None:
        super().__init__(master)
        self.title("Detener lote")
        self.geometry("460x190")
        self.minsize(460, 190)
        self.maxsize(460, 190)
        self.transient(master)
        self.grab_set()
        self.result = False

        frame = ctk.CTkFrame(self, corner_radius=12, fg_color="#eef3f8")
        frame.pack(fill="both", expand=True, padx=12, pady=12)

        ctk.CTkLabel(
            frame,
            text="El lote sigue en ejecución.",
            font=ctk.CTkFont(size=16, weight="bold"),
            text_color="#16324f",
        ).pack(anchor="w", padx=16, pady=(16, 8))

        ctk.CTkLabel(
            frame,
            text="Podés continuar o cortar ahora e informar lo procesado hasta este momento.",
            font=ctk.CTkFont(size=13),
            text_color="#66788a",
            justify="left",
            wraplength=400,
        ).pack(anchor="w", padx=16, pady=(0, 18))

        buttons = ctk.CTkFrame(frame, fg_color="transparent")
        buttons.pack(fill="x", padx=16, pady=(0, 16))
        buttons.grid_columnconfigure(0, weight=1)
        buttons.grid_columnconfigure(1, weight=0)
        buttons.grid_columnconfigure(2, weight=0)

        ctk.CTkButton(
            buttons,
            text="Continuar",
            width=118,
            fg_color="#6d7f90",
            hover_color="#536577",
            command=self._continue_running,
        ).grid(row=0, column=1, padx=(0, 8), pady=0, sticky="e")

        ctk.CTkButton(
            buttons,
            text="Informar hasta ahora",
            width=160,
            fg_color="#bd6b2a",
            hover_color="#9d571f",
            command=self._stop_and_report,
        ).grid(row=0, column=2, padx=0, pady=0, sticky="e")

        self.protocol("WM_DELETE_WINDOW", self._continue_running)

    def _continue_running(self) -> None:
        self.result = False
        self.destroy()

    def _stop_and_report(self) -> None:
        self.result = True
        self.destroy()


class PamiOmeModuleFrame(ctk.CTkFrame):
    def __init__(self, master, on_back=None, *, config: OmeModuleConfig = MED_CABECERA_CONFIG) -> None:
        super().__init__(master, fg_color="transparent")
        self.on_back = on_back
        self.config_data = config

        self.event_queue: queue.Queue = queue.Queue()
        self.controller_queue: queue.Queue = queue.Queue()
        self.action_running = False
        self.stop_requested = False
        self.password_visible = False
        self._bulk_scroll_syncing = False
        self._syncing_sheet_profile = False
        self._syncing_pami_profile = False
        self.data_dir = get_data_dir()
        self.module_key = self.config_data.module_key
        self.module_title = self.config_data.module_title
        self.module_subtitle = self.config_data.module_subtitle
        self.practice_options = list(self.config_data.practice_options)
        self.diagnosis_options = list(self.config_data.diagnosis_options)
        self.profiles_file = Path(self.data_dir) / self.config_data.profile_filename
        self.practice_catalog_file = Path(self.data_dir) / f"{self.module_key}_nomenclador_practicas.json"
        self.sheet_settings_file = Path(self.data_dir) / f"{self.module_key}_sheets_config.json"
        self.saved_profiles = self._load_saved_profiles()
        self.sheet_settings = self._load_sheet_settings()
        self.sheet_profiles = self._extract_sheet_profiles(self.sheet_settings)
        self.selected_sheet_profile_id = str(self.sheet_settings.get("selected_profile_id", "")).strip()
        self.sheet_profile_lookup: dict[str, dict] = {}
        self.row_widgets: list[dict] = []
        self.current_result_rows: list[dict] = []
        self.selected_result_item: str | None = None
        self.selected_result_column: str | None = None
        self.last_results_csv_path: Path | None = None
        self.practice_catalog_items: list[PracticeCatalogItem] = []
        self.practice_catalog_path: Path | None = None
        self.active_catalog_modules: set[str] = set()
        self.active_catalog_practices: set[str] = set()
        self.filtered_practice_options: list[str] = []
        self.last_bulk_source_textbox: ctk.CTkTextbox | None = None
        self.last_bulk_line: str | None = None
        self.sheets_connected = False

        self.controller: PamiOmeController | None = None
        self.controller_thread: threading.Thread | None = None

        self._build_ui()
        self._initialize_practice_catalog()
        self.after(400, self._start_sheets_status_check)
        self.after(150, self._process_ui_queue)

    def _load_button_icon(self, filename: str, size: tuple[int, int]):
        try:
            from PIL import Image

            return ctk.CTkImage(Image.open(get_resource_path(f"assets/icons/{filename}")), size=size)
        except Exception:
            return None

    def _build_ui(self) -> None:
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=1)
        self._icon_save = self._load_button_icon("save.png", (18, 18))
        self._icon_trash = self._load_button_icon("trash.png", (16, 16))
        self._icon_eye = self._load_button_icon("eye.png", (18, 18))
        self._icon_eye_off = self._load_button_icon("eye_off.png", (18, 18))

        self.top_bar = ctk.CTkFrame(self, corner_radius=12, fg_color="#ffffff", border_width=1, border_color="#d8e2ec")
        self.top_bar.grid(row=0, column=0, padx=8, pady=(6, 4), sticky="ew")
        self.top_bar.grid_columnconfigure(1, weight=1)
        self.top_bar.grid_columnconfigure(2, weight=0)

        if self.on_back:
            self.back_button = ctk.CTkButton(
                self.top_bar,
                text="Volver",
                width=78,
                command=self._go_home,
                fg_color="#66788a",
                hover_color="#536577",
            )
            self.back_button.grid(row=0, column=0, rowspan=2, padx=(8, 10), pady=6, sticky="w")
        else:
            self.back_button = None

        self.title_label = ctk.CTkLabel(
            self.top_bar,
            text=self.module_title,
            font=ctk.CTkFont(size=20, weight="bold"),
            text_color="#16324f",
        )
        self.title_label.grid(row=0, column=1, padx=10, pady=(6, 1), sticky="w")

        self.subtitle_label = ctk.CTkLabel(
            self.top_bar,
            text=self.module_subtitle,
            font=ctk.CTkFont(size=11),
            text_color="#51657a",
        )
        self.subtitle_label.grid(row=1, column=1, padx=10, pady=(0, 6), sticky="w")

        # Tiles compactos de resumen: siempre visibles en el encabezado.
        self.summary_tiles = ctk.CTkFrame(self.top_bar, fg_color="transparent")
        self.summary_tiles.grid(row=0, column=2, rowspan=2, padx=(8, 6), pady=6, sticky="e")

        def _build_summary_tile(col: int, label: str, accent: str):
            tile = ctk.CTkFrame(
                self.summary_tiles, fg_color="#ffffff", corner_radius=8,
                border_width=1, border_color="#d8e2ec",
            )
            tile.grid(row=0, column=col, padx=(0 if col == 0 else 6, 0), sticky="w")
            ctk.CTkFrame(tile, fg_color=accent, width=3, height=1, corner_radius=0).grid(
                row=0, column=0, rowspan=2, sticky="ns"
            )
            value = ctk.CTkLabel(
                tile, text="0", font=ctk.CTkFont(size=15, weight="bold"), text_color="#16324f",
            )
            value.grid(row=0, column=1, padx=(7, 11), pady=(3, 0), sticky="w")
            ctk.CTkLabel(
                tile, text=label, font=ctk.CTkFont(size=9, weight="bold"), text_color="#6d7f90",
            ).grid(row=1, column=1, padx=(7, 11), pady=(0, 3), sticky="w")
            return value

        self.tile_grilla_value = _build_summary_tile(0, "EN GRILLA", "#66788a")
        self.tile_gen_value = _build_summary_tile(1, "GENERADAS", "#1f7a46")
        self.tile_err_value = _build_summary_tile(2, "ERRORES", "#8a5a5a")

        self.restart_button = ctk.CTkButton(
            self.top_bar,
            text="Reiniciar app",
            width=132,
            height=28,
            command=self._restart_app,
            fg_color="#66788a",
            hover_color="#536577",
            font=ctk.CTkFont(size=11, weight="bold"),
        )
        self.restart_button.grid(row=0, column=3, rowspan=2, padx=(8, 10), pady=6, sticky="e")

        content = ctk.CTkScrollableFrame(self, corner_radius=8, fg_color="#eef3f8")
        self.content_frame = content
        content.grid(row=1, column=0, padx=10, pady=(0, 10), sticky="nsew")
        content.grid_columnconfigure(0, weight=1, minsize=860)

        pasos = ctk.CTkFrame(content, corner_radius=12, fg_color="#ffffff", border_width=1, border_color="#d8e2ec")
        self.pasos_frame = pasos
        pasos.grid(row=0, column=0, padx=8, pady=(0, 4), sticky="ew")
        pasos.grid_columnconfigure(0, weight=0)
        pasos.grid_columnconfigure(1, weight=1)
        pasos.grid_columnconfigure(2, weight=0)

        ctk.CTkLabel(
            pasos,
            text="Modo de carga",
            font=ctk.CTkFont(size=14, weight="bold"),
            text_color="#16324f",
        ).grid(row=0, column=0, padx=(12, 10), pady=6, sticky="w")

        self.pasos_steps_label = ctk.CTkLabel(
            pasos,
            text="1 Guardar perfil  |  2 Elegir BENEF o DNI  |  3 Pegar listas  |  4 Cargar grilla  |  5 Ejecutar y revisar",
            justify="left",
            font=ctk.CTkFont(size=11),
            text_color="#536577",
        )
        self.pasos_steps_label.grid(row=0, column=1, padx=(0, 12), pady=6, sticky="w")
        # El label de texto se conserva (el layout responsive lo referencia) pero se
        # oculta y se reemplaza visualmente por un stepper de chips.
        self.pasos_steps_label.grid_remove()
        self.pasos_stepper = ctk.CTkFrame(pasos, fg_color="transparent")
        self.pasos_stepper.grid(row=0, column=1, padx=(0, 12), pady=5, sticky="w")
        for _step_index, _step_text in enumerate(
            ("Guardar perfil", "Elegir BENEF/DNI", "Pegar listas", "Cargar grilla", "Ejecutar y revisar")
        ):
            _chip = ctk.CTkFrame(
                self.pasos_stepper, fg_color="#f4f8fb", corner_radius=8,
                border_width=1, border_color="#d8e2ec",
            )
            _chip.grid(row=0, column=_step_index, padx=(0 if _step_index == 0 else 6, 0), sticky="w")
            ctk.CTkLabel(
                _chip, text=str(_step_index + 1), font=ctk.CTkFont(size=10, weight="bold"),
                fg_color="#245b9d", text_color="#ffffff", corner_radius=9, width=18, height=18,
            ).grid(row=0, column=0, padx=(6, 6), pady=4)
            ctk.CTkLabel(
                _chip, text=_step_text, font=ctk.CTkFont(size=10, weight="bold"), text_color="#51657a",
            ).grid(row=0, column=1, padx=(0, 8), pady=4)

        self.pasos_actions = ctk.CTkFrame(pasos, fg_color="transparent")
        self.pasos_actions.grid(row=0, column=2, padx=(8, 12), pady=5, sticky="e")

        ctk.CTkLabel(
            self.pasos_actions,
            text="Carga masiva",
            font=ctk.CTkFont(size=12, weight="bold"),
            text_color="#16324f",
        ).grid(row=0, column=0, padx=(0, 10), pady=0, sticky="w")

        self.load_rows_button = ctk.CTkButton(
            self.pasos_actions,
            text="Cargar grilla",
            width=124,
            height=30,
            command=self._load_rows_from_bulk,
        )
        self.load_rows_button.grid(row=0, column=1, padx=(0, 8), pady=0, sticky="e")

        self.template_button = ctk.CTkButton(
            self.pasos_actions,
            text="↓",
            command=self._download_template,
            width=36,
            height=30,
            font=ctk.CTkFont(size=15, weight="bold"),
            fg_color="#66788a",
            hover_color="#536577",
        )
        self.template_button.grid(row=0, column=2, padx=0, pady=0, sticky="e")
        attach_tooltip(self.template_button, "Descargar modelo")
        if not self.config_data.show_mass_load_panel:
            pasos.grid_remove()

        self.sheet_url_var = ctk.StringVar(value=normalize_spreadsheet_url(str(self.sheet_settings.get("spreadsheet_url", ""))))
        self.sheet_internal_name_var = ctk.StringVar(value=str(self.sheet_settings.get("display_name", "")).strip())
        self.sheet_template_display_var = ctk.StringVar(value="")
        self.sheet_name_var = ctk.StringVar(value=str(self.sheet_settings.get("sheet_name", "Mc Dube")))
        self.sheet_start_row_var = ctk.StringVar(value=str(self.sheet_settings.get("start_row", 2)))
        self.sheet_max_rows_var = ctk.StringVar(value=str(self.sheet_settings.get("max_rows", 40)))
        self.sheet_limit_mode_var = ctk.StringVar(
            value=self._sheet_limit_mode_label(self.sheet_settings.get("limit_mode", "cantidad"))
        )
        self.sheet_profile_var = ctk.StringVar(value="")
        self.sheet_tabs = self._settings_sheet_tabs(self.sheet_settings, self.sheet_name_var.get())
        self.sheet_complete_benef_var = ctk.BooleanVar(value=False)
        self.sheet_complete_dni_var = ctk.BooleanVar(value=False)
        self.sheet_check_credential_var = ctk.BooleanVar(value=bool(self.sheet_settings.get("check_credential", False)))
        self.headless_var = ctk.BooleanVar(value=False)
        self._sheet_precheck_result_rows: list[dict] = []
        self.sheets_status_var = ctk.StringVar(value="Google Sheets no conectado")

        self.sheets_frame = ctk.CTkFrame(content, corner_radius=12, fg_color="#ffffff", border_width=1, border_color="#d8e2ec")
        self.frame_sheets = self.sheets_frame
        self.sheets_frame.grid(row=1, column=0, padx=10, pady=(0, 20), sticky="ew")
        for col in range(10):
            self.sheets_frame.grid_columnconfigure(col, weight=0, minsize=0)
        self.sheets_frame.grid_columnconfigure(1, weight=0, minsize=170)
        self.sheets_frame.grid_columnconfigure(3, weight=0, minsize=150)
        self.sheets_frame.grid_columnconfigure(5, weight=0, minsize=150)
        self.sheets_frame.grid_columnconfigure(9, weight=1, minsize=150)

        ctk.CTkLabel(
            self.sheets_frame,
            text="Google Sheets",
            font=ctk.CTkFont(size=14, weight="bold"),
            text_color="#16324f",
        ).grid(row=0, column=0, padx=(12, 8), pady=(10, 4), sticky="w")

        self.sheet_profile_label = ctk.CTkLabel(self.sheets_frame, text="Plantilla guardada", text_color="#16324f")
        self.sheet_profile_label.grid(row=0, column=1, padx=(8, 8), pady=(10, 4), sticky="e")
        self.sheet_profile_combo = ctk.CTkComboBox(
            self.sheets_frame,
            values=self._sheet_profile_options() or [""],
            variable=self.sheet_profile_var,
            command=self._on_sheet_profile_selected,
            width=360,
            state="readonly",
        )
        self.sheet_profile_combo.grid(row=0, column=2, padx=(0, 8), pady=(10, 4), sticky="w")

        # Acciones de plantilla agrupadas en un contenedor compacto (evita los
        # huecos que dejaban los botones sueltos en columnas anchas de la grilla).
        self.sheet_profile_actions = ctk.CTkFrame(self.sheets_frame, fg_color="transparent")
        self.sheet_profile_actions.grid(row=0, column=3, columnspan=3, padx=(0, 8), pady=(10, 4), sticky="w")

        self.new_sheet_profile_button = ctk.CTkButton(
            self.sheet_profile_actions,
            text="+",
            command=self._new_sheet_profile,
            width=36,
            font=ctk.CTkFont(size=16, weight="bold"),
            fg_color="#245b9d",
            hover_color="#1d4b82",
        )
        self.new_sheet_profile_button.grid(row=0, column=0, padx=(0, 6), pady=0, sticky="w")
        attach_tooltip(self.new_sheet_profile_button, "Nueva plantilla")

        self.save_sheet_profile_button = ctk.CTkButton(
            self.sheet_profile_actions,
            text="Guardar",
            command=self._save_current_sheet_profile,
            width=84,
            fg_color="#66788a",
            hover_color="#536577",
        )
        self.save_sheet_profile_button.grid(row=0, column=1, padx=(0, 6), pady=0, sticky="w")
        attach_tooltip(self.save_sheet_profile_button, "Guardar plantilla")
        if self._icon_save is not None:
            self.save_sheet_profile_button.configure(image=self._icon_save, text="", width=40)

        self.delete_sheet_profile_button = ctk.CTkButton(
            self.sheet_profile_actions,
            text="✕",
            command=self._delete_current_sheet_profile,
            width=36,
            font=ctk.CTkFont(size=15, weight="bold"),
            fg_color="#8a5a5a",
            hover_color="#734949",
        )
        self.delete_sheet_profile_button.grid(row=0, column=2, padx=0, pady=0, sticky="w")
        attach_tooltip(self.delete_sheet_profile_button, "Borrar plantilla")
        if self._icon_trash is not None:
            self.delete_sheet_profile_button.configure(image=self._icon_trash, text="")

        self.sheet_url_label = ctk.CTkLabel(self.sheets_frame, text="Plantilla", text_color="#16324f")
        self.sheet_url_label.grid(row=1, column=0, padx=(12, 8), pady=(0, 8), sticky="w")
        self.sheet_url_entry = ctk.CTkEntry(
            self.sheets_frame,
            textvariable=self.sheet_template_display_var,
            state="readonly",
        )
        self.sheet_url_entry.grid(row=1, column=1, columnspan=3, padx=8, pady=(0, 8), sticky="w")
        self.sheet_url_entry.bind("<Button-1>", lambda _event: self._open_sheet_template_dialog(), add="+")

        self.sheet_url_config_button = ctk.CTkButton(
            self.sheets_frame,
            text="...",
            command=self._open_sheet_template_dialog,
            width=32,
            height=28,
            fg_color="#51657a",
            hover_color="#536577",
        )
        self.sheet_url_config_button.grid(row=1, column=4, padx=(0, 8), pady=(0, 8), sticky="w")

        self.sheet_name_label = ctk.CTkLabel(self.sheets_frame, text="Pestaña", text_color="#16324f")
        self.sheet_name_label.grid(
            row=1, column=5, padx=(8, 8), pady=(0, 8), sticky="w"
        )
        self.sheet_name_combo = ctk.CTkComboBox(
            self.sheets_frame,
            values=self.sheet_tabs or [""],
            variable=self.sheet_name_var,
            width=160,
            state="readonly",
        )
        self.sheet_name_combo.grid(row=1, column=6, padx=(0, 8), pady=(0, 8), sticky="w")

        self.load_sheet_tabs_button = ctk.CTkButton(
            self.sheets_frame,
            text="Cargar pestañas",
            command=lambda: self._run_action(self._load_sheet_tabs),
            width=128,
            fg_color="#66788a",
            hover_color="#536577",
        )
        self.load_sheet_tabs_button.grid(row=0, column=6, padx=(0, 8), pady=(10, 4), sticky="w")

        self.sheet_start_row_label = ctk.CTkLabel(self.sheets_frame, text="Fila inicial", text_color="#16324f")
        self.sheet_start_row_label.grid(
            row=1, column=7, padx=(8, 8), pady=(0, 8), sticky="w"
        )
        self.sheet_start_row_entry = ctk.CTkEntry(
            self.sheets_frame,
            textvariable=self.sheet_start_row_var,
            width=90,
        )
        self.sheet_start_row_entry.grid(row=1, column=8, padx=(0, 8), pady=(0, 8), sticky="w")

        self.sheet_limit_mode_combo = ctk.CTkComboBox(
            self.sheets_frame,
            values=["Cantidad", "Fila final"],
            variable=self.sheet_limit_mode_var,
            command=lambda _value: self._save_sheet_settings(),
            width=116,
            state="readonly",
        )
        self.sheet_limit_mode_combo.grid(row=2, column=6, padx=(8, 8), pady=(0, 10), sticky="w")
        self.sheet_max_rows_entry = ctk.CTkEntry(
            self.sheets_frame,
            textvariable=self.sheet_max_rows_var,
            width=80,
        )
        self.sheet_max_rows_entry.grid(row=2, column=7, padx=(0, 8), pady=(0, 10), sticky="w")

        self.sheets_run_actions = ctk.CTkFrame(self.sheets_frame, fg_color="transparent")
        self.sheets_run_actions.grid(row=2, column=8, columnspan=2, padx=(0, 8), pady=(0, 10), sticky="w")

        self.sheets_connect_button = ctk.CTkButton(
            self.sheets_run_actions,
            text="Conectar Sheets",
            command=lambda: self._run_action(self._connect_sheets_account),
            width=138,
        )
        self.sheets_connect_button.grid(row=0, column=0, padx=(0, 8), pady=0, sticky="w")

        self.sheets_run_button = ctk.CTkButton(
            self.sheets_run_actions,
            text="Ejecutar desde Sheets",
            command=lambda: self._run_action(
                self._run_sheet_batch_with_profile,
                pending_status="Leyendo Google Sheets y preparando lote. No toques nada...",
                pending_summary=f"Preparando lote desde Sheets... Log en: {get_log_file()}",
            ),
            width=170,
            fg_color="#1f7a46",
            hover_color="#176238",
        )
        self.sheets_run_button.grid(row=0, column=1, padx=0, pady=0, sticky="w")

        # Checkboxes agrupados con separacion pareja (evita huecos raros entre ellos).
        self.sheet_checks = ctk.CTkFrame(self.sheets_frame, fg_color="transparent")
        self.sheet_checks.grid(row=2, column=1, columnspan=6, padx=(8, 8), pady=(0, 10), sticky="w")

        self.sheet_complete_benef_checkbox = ctk.CTkCheckBox(
            self.sheet_checks,
            text="Completar BENEF",
            variable=self.sheet_complete_benef_var,
            text_color="#16324f",
            checkbox_width=18,
            checkbox_height=18,
        )
        self.sheet_complete_benef_checkbox.grid(row=0, column=0, padx=(0, 24), pady=0, sticky="w")

        self.sheet_complete_dni_checkbox = ctk.CTkCheckBox(
            self.sheet_checks,
            text="Completar DNI",
            variable=self.sheet_complete_dni_var,
            text_color="#16324f",
            checkbox_width=18,
            checkbox_height=18,
        )
        self.sheet_complete_dni_checkbox.grid(row=0, column=1, padx=(0, 24), pady=0, sticky="w")

        self.sheet_check_credential_checkbox = ctk.CTkCheckBox(
            self.sheet_checks,
            text="Validar credencial",
            variable=self.sheet_check_credential_var,
            text_color="#16324f",
            checkbox_width=18,
            checkbox_height=18,
        )
        self.sheet_check_credential_checkbox.grid(row=0, column=2, padx=(0, 24), pady=0, sticky="w")

        self.sheet_headless_checkbox = ctk.CTkCheckBox(
            self.sheet_checks,
            text="No ver navegador",
            variable=self.headless_var,
            text_color="#16324f",
            checkbox_width=18,
            checkbox_height=18,
        )
        self.sheet_headless_checkbox.grid(row=0, column=3, padx=0, pady=0, sticky="w")

        self.sheets_status_label = ctk.CTkLabel(
            self.sheets_frame,
            textvariable=self.sheets_status_var,
            text_color="#51657a",
        )
        self.sheets_status_label.grid(row=3, column=1, columnspan=9, padx=(8, 12), pady=(0, 10), sticky="w")
        self._restore_sheet_profile_selection()

        self.profiles_frame = ctk.CTkFrame(content, corner_radius=12, fg_color="#ffffff", border_width=1, border_color="#d8e2ec")
        self.frame_perfil = self.profiles_frame
        self.profiles_frame.grid(row=2, column=0, padx=10, pady=(0, 20), sticky="ew")
        self.profiles_frame.grid_columnconfigure(1, weight=0, minsize=300)
        self.profiles_frame.grid_columnconfigure(2, weight=0, minsize=120)
        self.profiles_frame.grid_columnconfigure(3, weight=0, minsize=140)
        self.profiles_frame.grid_columnconfigure(4, weight=1, minsize=150)

        profile_options = self._profile_options()
        self.profile_var = ctk.StringVar(value=profile_options[0] if profile_options else "")
        self.profile_name_var = ctk.StringVar(value="")
        self.profile_user_var = ctk.StringVar(value="")
        self.profile_password_var = ctk.StringVar(value="")

        ctk.CTkLabel(
            self.profiles_frame,
            text="Perfil guardado",
            font=ctk.CTkFont(size=12, weight="bold"),
            text_color="#16324f",
        ).grid(row=0, column=0, padx=(12, 8), pady=(6, 4), sticky="w")

        self.profile_combo = ctk.CTkComboBox(
            self.profiles_frame,
            values=profile_options or [""],
            variable=self.profile_var,
            width=520,
            height=28,
            command=self._on_profile_selected,
        )
        self.profile_combo.grid(row=0, column=1, padx=8, pady=(5, 4), sticky="w")

        self.profile_actions = ctk.CTkFrame(self.profiles_frame, fg_color="transparent")
        self.profile_actions.grid(row=0, column=2, columnspan=2, padx=8, pady=(5, 4), sticky="w")

        self.new_profile_button = ctk.CTkButton(
            self.profile_actions,
            text="+",
            command=self._new_profile,
            width=36,
            height=28,
            font=ctk.CTkFont(size=16, weight="bold"),
            fg_color="#245b9d",
            hover_color="#1d4b82",
        )
        self.new_profile_button.grid(row=0, column=0, padx=(0, 6), pady=0, sticky="w")
        attach_tooltip(self.new_profile_button, "Nuevo perfil")

        self.delete_profile_button = ctk.CTkButton(
            self.profile_actions,
            text="✕",
            command=self._delete_current_profile,
            width=36,
            height=28,
            font=ctk.CTkFont(size=15, weight="bold"),
            fg_color="#8a5a5a",
            hover_color="#734949",
        )
        self.delete_profile_button.grid(row=0, column=1, padx=0, pady=0, sticky="w")
        attach_tooltip(self.delete_profile_button, "Borrar perfil")
        if self._icon_trash is not None:
            self.delete_profile_button.configure(image=self._icon_trash, text="")

        ctk.CTkLabel(
            self.profiles_frame,
            text="Médico / cliente",
            font=ctk.CTkFont(size=12, weight="bold"),
            text_color="#16324f",
        ).grid(row=1, column=0, padx=(12, 8), pady=(0, 4), sticky="w")

        self.client_entry = ctk.CTkEntry(
            self.profiles_frame,
            textvariable=self.profile_name_var,
            placeholder_text="Nombre del médico o centro",
            width=520,
            height=28,
        )
        self.client_entry.grid(row=1, column=1, padx=8, pady=(0, 4), sticky="w")

        self.save_profile_button = ctk.CTkButton(
            self.profiles_frame,
            text="Guardar perfil",
            command=self._save_current_profile,
            width=112,
            height=28,
            fg_color="#66788a",
            hover_color="#536577",
        )
        self.save_profile_button.grid(row=1, column=2, padx=8, pady=(0, 4), sticky="w")
        attach_tooltip(self.save_profile_button, "Guardar perfil")
        if self._icon_save is not None:
            self.save_profile_button.configure(image=self._icon_save, text="", width=40)

        ctk.CTkLabel(
            self.profiles_frame,
            text="Usuario",
            font=ctk.CTkFont(size=12, weight="bold"),
            text_color="#16324f",
        ).grid(row=2, column=0, padx=(12, 8), pady=(0, 6), sticky="w")

        self.credentials_row = ctk.CTkFrame(self.profiles_frame, fg_color="transparent")
        self.credentials_row.grid(row=2, column=1, columnspan=4, padx=8, pady=(0, 6), sticky="ew")
        self.credentials_row.grid_columnconfigure(0, weight=0)
        self.credentials_row.grid_columnconfigure(2, weight=0)
        self.credentials_row.grid_columnconfigure(1, weight=0)
        self.credentials_row.grid_columnconfigure(3, weight=0)
        self.credentials_row.grid_columnconfigure(4, weight=1)

        self.user_entry = ctk.CTkEntry(
            self.credentials_row,
            textvariable=self.profile_user_var,
            placeholder_text="Usuario médico",
            width=220,
            height=28,
        )
        self.user_entry.grid(row=0, column=0, padx=(0, 12), pady=0, sticky="w")

        ctk.CTkLabel(
            self.credentials_row,
            text="Clave",
            font=ctk.CTkFont(size=12, weight="bold"),
            text_color="#16324f",
        ).grid(row=0, column=1, padx=(0, 8), pady=0, sticky="w")

        self.password_entry = ctk.CTkEntry(
            self.credentials_row,
            textvariable=self.profile_password_var,
            placeholder_text="Clave PAMI",
            show="*",
            width=220,
            height=28,
        )
        self.password_entry.grid(row=0, column=2, padx=(0, 8), pady=0, sticky="w")

        self.toggle_password_button = ctk.CTkButton(
            self.credentials_row,
            text="Ver",
            command=self._toggle_password_visibility,
            width=70,
            height=28,
            fg_color="#51657a",
            hover_color="#536577",
        )
        self.toggle_password_button.grid(row=0, column=3, padx=(0, 0), pady=0, sticky="w")
        attach_tooltip(self.toggle_password_button, "Ver / ocultar clave")
        if self._icon_eye is not None:
            self.toggle_password_button.configure(image=self._icon_eye, text="", width=40)

        self.controls = ctk.CTkFrame(content, corner_radius=12, fg_color="#ffffff", border_width=1, border_color="#d8e2ec")
        self.frame_botones_ejecucion = self.controls
        self.controls.grid(row=3, column=0, padx=10, pady=(0, 20), sticky="ew")
        for col in range(6):
            self.controls.grid_columnconfigure(col, weight=0, minsize=150)
        self.controls.grid_columnconfigure(6, weight=1, minsize=150)

        self.open_button: ctk.CTkButton | None = None
        if self.config_data.show_open_button:
            self.open_button = ctk.CTkButton(self.controls, text="Abrir CUP", command=self._handle_open_pami, width=118, height=30)
            self.open_button.grid(row=0, column=0, padx=(12, 8), pady=10, sticky="w")

        self.close_button = ctk.CTkButton(
            self.controls,
            text="Cerrar navegador",
            command=self._handle_close_browser,
            width=146,
            height=30,
            fg_color="#8a5a5a",
            hover_color="#734949",
        )
        self.close_button.grid(row=0, column=1 if self.config_data.show_open_button else 0, padx=8 if self.config_data.show_open_button else (12, 8), pady=10, sticky="w")

        self.run_batch_button: ctk.CTkButton | None = None
        if self.config_data.show_visible_batch_button:
            self.run_batch_button = ctk.CTkButton(
                self.controls,
                text="Iniciar bot",
                command=lambda: self._run_action(self._run_batch_with_profile),
                width=140,
                height=30,
                fg_color="#1f7a46",
                hover_color="#176238",
                font=ctk.CTkFont(size=12, weight="bold"),
            )
            self.run_batch_button.grid(row=0, column=2 if self.config_data.show_open_button else 1, padx=8, pady=10, sticky="w")

        self.full_batch_button = ctk.CTkButton(
            self.controls,
            text=self.config_data.full_batch_button_text,
            command=lambda: self._run_action(
                self._run_full_batch_with_profile,
                pending_status="Iniciando proceso. Abriendo navegador y preparando la sesion. No toques nada...",
                pending_summary=f"Preparando lote... Log en: {get_log_file()}",
            ),
            width=170,
            height=30,
            fg_color="#1f7a46",
            hover_color="#176238",
            font=ctk.CTkFont(size=12, weight="bold"),
        )
        full_batch_column = 3
        if not self.config_data.show_open_button and not self.config_data.show_visible_batch_button:
            full_batch_column = 1
        elif not self.config_data.show_visible_batch_button:
            full_batch_column = 2
        self.full_batch_button.grid(row=0, column=full_batch_column, padx=8, pady=10, sticky="w")

        self.stop_button: ctk.CTkButton | None = None
        if self.config_data.show_stop_button:
            self.stop_button = ctk.CTkButton(
                self.controls,
                text="Detener",
                command=self._request_stop,
                width=118,
                height=30,
                fg_color="#bd6b2a",
                hover_color="#9d571f",
                font=ctk.CTkFont(size=12, weight="bold"),
                state="disabled",
            )
            self.stop_button.grid(row=0, column=full_batch_column + 1, padx=(8, 12), pady=10, sticky="w")

        self.clear_panel_button: ctk.CTkButton | None = None
        if self.config_data.show_clear_panel_button:
            self.clear_panel_button = ctk.CTkButton(
                self.controls,
                text="Limpiar panel",
                command=self._clear_rows,
                width=128,
                height=30,
                fg_color="#66788a",
                hover_color="#536577",
            )
            self.clear_panel_button.grid(row=0, column=full_batch_column + 2, padx=(0, 12), pady=10, sticky="w")

        self.manual_frame = ctk.CTkFrame(content, corner_radius=12, fg_color="#ffffff", border_width=1, border_color="#d8e2ec")
        self.frame_controles = self.manual_frame
        self.manual_frame.grid(row=4, column=0, padx=10, pady=(0, 20), sticky="nsew")
        self.manual_frame.configure(height=1)
        self.manual_frame.grid_propagate(True)
        self.manual_frame.grid_columnconfigure((0, 1, 2), weight=1)
        self.manual_frame.grid_rowconfigure(2, weight=0, minsize=0)
        self.manual_frame.grid_rowconfigure(4, weight=0, minsize=0)
        self.manual_frame.grid_rowconfigure(5, weight=0)

        self.search_mode_var = ctk.StringVar(value="AUTO")
        self.practice_var = ctk.StringVar(value=self._preferred_practice_option(self.practice_options))
        self.diagnosis_preset_var = ctk.StringVar(value=self.diagnosis_options[0] if self.diagnosis_options else "")
        self.output_var = ctk.StringVar()
        self.selected_row_var = ctk.IntVar(value=-1)

        self.top_row = ctk.CTkFrame(self.manual_frame, fg_color="transparent")
        self.top_row.grid(row=0, column=0, columnspan=3, padx=10, pady=(8, 5), sticky="ew")
        self.top_row.grid_columnconfigure(3, weight=0)
        self.top_row.grid_columnconfigure(8, weight=0)
        self.top_row.grid_columnconfigure(9, weight=0)

        self.tipo_group = ctk.CTkFrame(self.top_row, fg_color="transparent")
        self.tipo_group.grid(row=0, column=0, padx=(0, 14), pady=0, sticky="w")

        self.search_mode_label = ctk.CTkLabel(
            self.tipo_group,
            text="Tipo",
            font=ctk.CTkFont(size=12, weight="bold"),
            text_color="#16324f",
        )
        self.search_mode_label.grid(row=0, column=0, padx=(0, 8), pady=0, sticky="w")
        self.search_mode_combo = ctk.CTkComboBox(
            self.tipo_group, values=["AUTO", "BENEF", "DNI"], variable=self.search_mode_var, width=120, state="readonly"
        )
        self.search_mode_combo.grid(row=0, column=1, padx=0, pady=0, sticky="w")

        self.practice_label = ctk.CTkLabel(
            self.top_row,
            text="Práctica general",
            font=ctk.CTkFont(size=12, weight="bold"),
            text_color="#16324f",
        )
        self.practice_label.grid(
            row=0, column=2, padx=(0, 8), pady=0, sticky="w"
        )
        self.practice_combo = ctk.CTkComboBox(
            self.top_row, values=self.practice_options, variable=self.practice_var, width=520, state="normal"
        )
        self.practice_combo.grid(row=0, column=3, padx=(0, 14), pady=0, sticky="w")
        if hasattr(self.practice_combo, "_entry"):
            self.practice_combo._entry.bind("<KeyRelease>", self._on_practice_combo_keyrelease, add="+")
            self.practice_combo._entry.bind("<FocusOut>", self._on_practice_combo_focus_out, add="+")

        self.practice_catalog_button = ctk.CTkButton(
            self.top_row,
            text="Códigos...",
            width=108,
            command=self._open_practice_catalog_manager,
            fg_color="#66788a",
            hover_color="#536577",
        )
        self.practice_catalog_button.grid(row=0, column=4, padx=(0, 10), pady=0, sticky="w")

        self.apply_practice_button = ctk.CTkButton(
            self.top_row,
            text="Práctica a fila",
            width=114,
            command=self._apply_general_practice_to_row,
            fg_color="#66788a",
            hover_color="#536577",
        )
        self.apply_practice_button.grid(row=0, column=5, padx=(0, 8), pady=0, sticky="w")

        self.apply_practice_all_button = ctk.CTkButton(
            self.top_row,
            text="Práctica a todos",
            width=126,
            command=self._apply_general_practice_to_all,
            fg_color="#66788a",
            hover_color="#536577",
        )
        self.apply_practice_all_button.grid(row=0, column=6, padx=(0, 14), pady=0, sticky="w")

        self.diagnosis_label = ctk.CTkLabel(
            self.top_row,
            text="Diagnóstico rápido",
            font=ctk.CTkFont(size=12, weight="bold"),
            text_color="#16324f",
        )
        self.diagnosis_label.grid(row=1, column=0, padx=(0, 8), pady=(8, 0), sticky="w")
        self.diagnosis_preset_combo = ctk.CTkComboBox(
            self.top_row,
            values=self.diagnosis_options,
            variable=self.diagnosis_preset_var,
            width=280,
        )
        self.diagnosis_preset_combo.grid(row=1, column=1, columnspan=2, padx=(0, 10), pady=(8, 0), sticky="w")

        self.diag_actions = ctk.CTkFrame(self.top_row, fg_color="transparent")
        self.diag_actions.grid(row=1, column=3, columnspan=2, padx=(0, 8), pady=(8, 0), sticky="w")

        self.apply_diag_button = ctk.CTkButton(
            self.diag_actions,
            text="Aplicar a fila",
            width=114,
            command=self._apply_selected_diagnosis_to_row,
        )
        self.apply_diag_button.grid(row=0, column=0, padx=(0, 8), pady=0, sticky="w")

        self.apply_diag_all_button = ctk.CTkButton(
            self.diag_actions,
            text="Dx a todos",
            width=102,
            command=self._apply_selected_diagnosis_to_all,
        )
        self.apply_diag_all_button.grid(row=0, column=1, padx=0, pady=0, sticky="w")

        self.headless_checkbox = ctk.CTkCheckBox(self.top_row, text="No ver navegador", variable=self.headless_var)
        self.headless_checkbox.grid(row=1, column=5, padx=(8, 0), pady=(8, 0), sticky="w")

        self.quick_help_label = ctk.CTkLabel(
            self.top_row,
            text=(
                "Uso rapido: pega BENEF/DNI, diagnosticos y practicas por columna. "
                "Si dejas AUTO, cada fila se interpreta sola por longitud. "
                "Si estas escribiendo en una linea, los botones aplican en esa misma fila."
            ),
            font=ctk.CTkFont(size=11),
            text_color="#51657a",
        )
        self.quick_help_label.grid(row=2, column=0, columnspan=10, padx=(0, 0), pady=(6, 0), sticky="w")

        quick_load_frame = ctk.CTkFrame(self.manual_frame, corner_radius=8, fg_color="#f4f8fb", border_width=1, border_color="#d8e2ec")
        self.frame_grilla = quick_load_frame
        quick_load_frame.grid(row=1, column=0, columnspan=3, padx=10, pady=(0, 5), sticky="nsew")
        quick_load_frame.grid_columnconfigure(1, weight=2)
        quick_load_frame.grid_columnconfigure(2, weight=2)
        quick_load_frame.grid_columnconfigure(3, weight=2)
        quick_load_frame.grid_rowconfigure(1, weight=1)

        ctk.CTkLabel(
            quick_load_frame,
            text="Fila",
            font=ctk.CTkFont(size=12, weight="bold"),
            text_color="#51657a",
        ).grid(row=0, column=0, padx=(8, 6), pady=(6, 3), sticky="w")
        ctk.CTkLabel(
            quick_load_frame,
            text="BENEF / DNI",
            font=ctk.CTkFont(size=12, weight="bold"),
            text_color="#16324f",
        ).grid(row=0, column=1, padx=6, pady=(6, 3), sticky="w")
        ctk.CTkLabel(
            quick_load_frame,
            text="Diagnóstico",
            font=ctk.CTkFont(size=12, weight="bold"),
            text_color="#16324f",
        ).grid(row=0, column=2, padx=6, pady=(6, 3), sticky="w")
        ctk.CTkLabel(
            quick_load_frame,
            text="Práctica",
            font=ctk.CTkFont(size=12, weight="bold"),
            text_color="#16324f",
        ).grid(row=0, column=3, padx=(6, 8), pady=(6, 3), sticky="w")

        self.bulk_line_numbers = ctk.CTkTextbox(
            quick_load_frame,
            width=34,
            height=58,
            font=ctk.CTkFont(size=11),
            fg_color="#eef3f8",
            text_color="#51657a",
            activate_scrollbars=False,
            border_width=0,
        )
        self.bulk_line_numbers.grid(row=1, column=0, padx=(8, 6), pady=(0, 7), sticky="ns")
        self.bulk_line_numbers.configure(state="disabled")

        self.identifiers_text = ctk.CTkTextbox(quick_load_frame, height=58, font=ctk.CTkFont(size=12), corner_radius=8, border_width=1, border_color="#d8e2ec")
        self.identifiers_text.grid(row=1, column=1, padx=6, pady=(0, 7), sticky="nsew")
        self.diagnostics_text = ctk.CTkTextbox(quick_load_frame, height=58, font=ctk.CTkFont(size=12), corner_radius=8, border_width=1, border_color="#d8e2ec")
        self.diagnostics_text.grid(row=1, column=2, padx=6, pady=(0, 7), sticky="nsew")
        self.practices_text = ctk.CTkTextbox(quick_load_frame, height=58, font=ctk.CTkFont(size=12), corner_radius=8, border_width=1, border_color="#d8e2ec")
        self.practices_text.grid(row=1, column=3, padx=(6, 8), pady=(0, 7), sticky="nsew")

        for textbox in (self.identifiers_text, self.diagnostics_text, self.practices_text):
            try:
                textbox._y_scrollbar.configure(width=4)
            except Exception:
                pass
            textbox.configure(yscrollcommand=lambda first, last, box=textbox: self._on_bulk_text_scroll(box, first, last))
            textbox.bind("<KeyRelease>", lambda _event: self._refresh_bulk_line_numbers(), add="+")
            textbox.bind("<FocusIn>", lambda _event: self._refresh_bulk_line_numbers(), add="+")
            textbox.bind("<ButtonRelease-1>", lambda _event: self._refresh_bulk_line_numbers(), add="+")
            textbox.bind("<KeyRelease>", lambda _event, box=textbox: self._remember_bulk_line(box), add="+")
            textbox.bind("<FocusIn>", lambda _event, box=textbox: self._remember_bulk_line(box), add="+")
            textbox.bind("<ButtonRelease-1>", lambda _event, box=textbox: self._remember_bulk_line(box), add="+")

        self._refresh_bulk_line_numbers()

        bulk_buttons = ctk.CTkFrame(self.manual_frame, fg_color="transparent")
        bulk_buttons.grid(row=3, column=0, columnspan=3, padx=10, pady=(0, 5), sticky="ew")
        bulk_buttons.grid_columnconfigure(1, weight=1)

        ctk.CTkLabel(
            bulk_buttons,
            text=(
                "Cada fila visual representa una OME. Si pegas 1 diagnostico o 1 practica, "
                "se replica a toda la lista. Si la fila ya tiene valor, no se sobreescribe."
            ),
            font=ctk.CTkFont(size=11),
            text_color="#51657a",
        ).grid(row=0, column=0, padx=(0, 12), pady=0, sticky="w")

        self.clear_rows_button = ctk.CTkButton(
            bulk_buttons,
            text="Limpiar grilla",
            width=120,
            command=self._clear_rows,
            fg_color="#9aafc3",
            hover_color="#66788a",
        )
        self.clear_rows_button.grid(row=0, column=2, padx=0, pady=0, sticky="e")

        grid_container = ctk.CTkFrame(self.manual_frame, corner_radius=8, fg_color="#ffffff", border_width=1, border_color="#d8e2ec")
        self.grid_container = grid_container
        grid_container.grid(row=4, column=0, columnspan=3, padx=10, pady=(0, 5), sticky="nsew")
        grid_container.grid_columnconfigure(0, weight=1)
        grid_container.grid_rowconfigure(1, weight=1)

        self.header_row = ctk.CTkFrame(grid_container, fg_color="#eef3f8", corner_radius=6)
        self.header_row.grid(row=0, column=0, padx=5, pady=(5, 3), sticky="ew")
        for idx, text in enumerate(("Sel", "Fila", "BENEF / DNI", "Diagnostico", "Practica", "Resultado", "Nro OME")):
            if idx == 0:
                width = 50
            elif idx == 1:
                width = 70
            elif idx == 2:
                width = 180
            elif idx == 3:
                width = 220
            elif idx == 4:
                width = 100
            elif idx == 5:
                width = 110
            else:
                width = 120
            ctk.CTkLabel(
                self.header_row,
                text=text,
                width=width,
                anchor="w",
                font=ctk.CTkFont(size=12, weight="bold"),
                text_color="#16324f",
            ).grid(row=0, column=idx, padx=6, pady=5, sticky="w")

        self.rows_scroll = ctk.CTkScrollableFrame(grid_container, fg_color="#ffffff")
        self.rows_scroll.grid(row=1, column=0, padx=5, pady=(0, 5), sticky="nsew")
        self.rows_scroll.grid_columnconfigure(0, weight=1)

        self.results_table_container = ctk.CTkFrame(grid_container, fg_color="#ffffff")
        self.results_table_container.grid(row=1, column=0, padx=5, pady=(0, 5), sticky="nsew")
        self.results_table_container.grid_columnconfigure(0, weight=1)
        self.results_table_container.grid_rowconfigure(0, weight=1)
        self.results_table = ttk.Treeview(
            self.results_table_container,
            columns=("afiliado", "diagnostico", "practica", "nro_ome"),
            show="headings",
            height=4,
            selectmode="none",
        )
        self.results_table.grid(row=0, column=0, sticky="nsew")
        self.results_table.heading("afiliado", text="BENEF / DNI")
        self.results_table.heading("diagnostico", text="DIAGNOSTICO")
        self.results_table.heading("practica", text="PRACTICA")
        self.results_table.heading("nro_ome", text="NRO OME")
        self.results_table.column("afiliado", width=180, anchor="center")
        self.results_table.column("diagnostico", width=140, anchor="center")
        self.results_table.column("practica", width=90, anchor="center")
        self.results_table.column("nro_ome", width=150, anchor="center")
        self.results_table_scroll = ttk.Scrollbar(self.results_table_container, orient="vertical", command=self.results_table.yview)
        self.results_table_scroll.grid(row=0, column=1, sticky="ns")
        self.results_table.configure(yscrollcommand=self.results_table_scroll.set)
        style = ttk.Style(self)
        try:
            style.theme_use("clam")
        except Exception:
            pass
        style.configure(
            "OmeResults.Treeview",
            rowheight=25,
            font=("Segoe UI", 10),
            background="#ffffff",
            fieldbackground="#ffffff",
            foreground="#16324f",
            borderwidth=0,
            relief="flat",
        )
        style.configure(
            "OmeResults.Treeview.Heading",
            font=("Segoe UI", 10, "bold"),
            background="#eef3f8",
            foreground="#16324f",
            borderwidth=0,
            relief="flat",
        )
        style.map(
            "OmeResults.Treeview",
            background=[("selected", "#d7eaf8")],
            foreground=[("selected", "#16324f")],
        )
        self.results_table.configure(style="OmeResults.Treeview")
        self.results_table.bind("<Button-1>", self._handle_results_table_click, add="+")
        self.results_table_container.grid_remove()
        self.grid_container.grid_remove()

        self.report_row = ctk.CTkFrame(self.manual_frame, fg_color="transparent")
        self.report_row.grid(row=5, column=0, columnspan=3, padx=10, pady=(0, 8), sticky="ew")
        self.report_row.grid_columnconfigure(2, weight=1)

        self.copy_row_button = ctk.CTkButton(
            self.report_row,
            text="Copiar fila",
            command=self.copy_selected_row,
            width=100,
            height=32,
            fg_color="#66788a",
            hover_color="#536577",
        )
        self.copy_row_button.grid(row=0, column=0, padx=(0, 8), pady=0, sticky="w")

        self.copy_ome_button = ctk.CTkButton(
            self.report_row,
            text="Copiar OMEs",
            command=self.copy_ome_column,
            width=110,
            height=32,
            fg_color="#245b9d",
            hover_color="#1d4b82",
        )
        self.copy_ome_button.grid(row=0, column=1, padx=(0, 12), pady=0, sticky="w")

        self.output_entry = ctk.CTkEntry(
            self.report_row,
            textvariable=self.output_var,
            placeholder_text="El Excel se genera solo cuando lo descargas",
        )
        self.output_entry.grid(row=0, column=2, padx=(0, 10), pady=0, sticky="ew")

        self.pick_output_button = ctk.CTkButton(
            self.report_row,
            text="Descargar reporte Excel",
            command=self._download_report_excel,
            width=160,
            height=32,
        )
        self.pick_output_button.grid(row=0, column=3, padx=(0, 10), pady=0, sticky="e")

        self.open_report_button = ctk.CTkButton(
            self.report_row,
            text="Abrir reporte",
            command=self._open_report_file,
            width=120,
            height=32,
            fg_color="#66788a",
            hover_color="#536577",
        )
        self.open_report_button.grid(row=0, column=4, padx=0, pady=0, sticky="e")
        self.report_row.grid_remove()

        status_frame = ctk.CTkFrame(content, corner_radius=8, fg_color="#16324f", border_width=1, border_color="#0f2436")
        self.status_frame = status_frame
        self.frame_log = status_frame
        status_frame.grid(row=5, column=0, padx=10, pady=(0, 10), sticky="ew")
        status_frame.configure(height=74)
        status_frame.grid_columnconfigure(0, weight=1)
        status_frame.grid_rowconfigure(2, weight=0)

        self.status_label = ctk.CTkLabel(
            status_frame,
            text="Preparado para abrir CUP o ejecutar un lote.",
            font=ctk.CTkFont(size=14, weight="bold"),
            text_color="#ffffff",
        )
        self.status_label.grid(row=0, column=0, padx=12, pady=(8, 4), sticky="w")

        self.summary_label = ctk.CTkLabel(
            status_frame,
            text=f"Log en: {get_log_file()}",
            font=ctk.CTkFont(size=12),
            text_color="#e8eef5",
        )
        self.summary_label.grid(row=1, column=0, padx=12, pady=(0, 6), sticky="w")

        self._load_initial_profile_into_form()
        self._apply_professional_skin()
        self.bind("<Configure>", self._handle_responsive_layout, add="+")
        self.after(120, self._apply_responsive_layout)

    def _apply_professional_skin(self) -> None:
        input_widgets = (
            self.sheet_profile_combo,
            self.sheet_url_entry,
            self.sheet_name_combo,
            self.sheet_start_row_entry,
            self.sheet_limit_mode_combo,
            self.sheet_max_rows_entry,
            self.profile_combo,
            self.client_entry,
            self.user_entry,
            self.password_entry,
            self.search_mode_combo,
            self.practice_combo,
            self.diagnosis_preset_combo,
            self.output_entry,
        )
        for widget in input_widgets:
            try:
                widget.configure(
                    fg_color="#ffffff",
                    border_color="#aebbc8",
                    text_color="#16324f",
                )
            except Exception:
                pass
            try:
                widget.configure(
                    button_color="#66788a",
                    button_hover_color="#536577",
                    dropdown_fg_color="#ffffff",
                    dropdown_hover_color="#eef3f8",
                    dropdown_text_color="#16324f",
                )
            except Exception:
                pass

        for textbox in (self.bulk_line_numbers, self.identifiers_text, self.diagnostics_text, self.practices_text):
            try:
                textbox.configure(
                    fg_color="#ffffff",
                    border_color="#d8e2ec",
                    text_color="#16324f",
                )
            except Exception:
                pass
        try:
            self.bulk_line_numbers.configure(fg_color="#eef3f8", text_color="#51657a")
        except Exception:
            pass

        for checkbox in (
            self.sheet_complete_benef_checkbox,
            self.sheet_complete_dni_checkbox,
            self.sheet_check_credential_checkbox,
            self.sheet_headless_checkbox,
            self.headless_checkbox,
        ):
            try:
                checkbox.configure(
                    text_color="#16324f",
                    border_color="#51657a",
                    fg_color="#245b9d",
                    hover_color="#1d4b82",
                    checkmark_color="#ffffff",
                )
            except Exception:
                pass

    def _handle_responsive_layout(self, event=None) -> None:
        # Ancho nuevo del frame: del evento <Configure> si viene, sino consultado.
        new_width = getattr(event, "width", None) if event is not None else None
        if not new_width:
            try:
                new_width = self.winfo_width()
            except Exception:
                new_width = 0
        # Si el ancho no cambio (ej. solo se movio la ventana o cambio el alto),
        # no re-maquetar: evita re-layouts al pedo y mejora fluidez.
        if new_width == getattr(self, "_last_layout_width", -1):
            return
        self._last_layout_width = new_width
        # Debounce: colapsar la ráfaga de <Configure> en un solo re-layout.
        after_id = getattr(self, "_relayout_after_id", None)
        if after_id is not None:
            try:
                self.after_cancel(after_id)
            except Exception:
                pass
        self._relayout_after_id = self.after(90, self._run_responsive_layout)

    def _run_responsive_layout(self) -> None:
        self._relayout_after_id = None
        self._apply_responsive_layout()

    def _apply_responsive_layout(self) -> None:
        raw_width = max(self.winfo_width(), self.winfo_toplevel().winfo_width(), 900)
        try:
            dpi_scale = max(self.winfo_fpixels("1i") / 72, 1)
        except Exception:
            dpi_scale = 1
        layout_width = max(raw_width / dpi_scale, 900)
        medium_width = layout_width < 1500 or raw_width < 2600
        compact = layout_width < 1220
        very_compact = layout_width < 1040
        ultra_compact = layout_width < 900
        # Flags de layout partido: estaban usados mas abajo pero NUNCA definidos,
        # lo que tiraba NameError en _apply_responsive_layout para ventanas
        # medianas/grandes (dejaba el layout a medio aplicar). Se definen aca.
        fixed_split_layout = False
        diagnosis_split = True
        min_content_width = 860
        full_width = int(max(layout_width - 32, min_content_width))
        self.content_frame.grid_columnconfigure(0, weight=1, minsize=min_content_width)
        self.content_frame.grid_columnconfigure(1, weight=0, minsize=0)
        self.content_frame.grid_columnconfigure(2, weight=0, minsize=0)
        card_column = 0

        for section in (
            getattr(self, "pasos_frame", None),
            self.sheets_frame,
            self.profiles_frame,
            self.controls,
            self.manual_frame,
            getattr(self, "status_frame", None),
        ):
            if section is None:
                continue
            try:
                section.configure(width=full_width)
                section.grid_configure(column=card_column, sticky="ew")
            except Exception:
                pass

        for col in range(5):
            self.top_bar.grid_columnconfigure(col, weight=0, minsize=0)
        self.top_bar.grid_columnconfigure(1, weight=1)
        back_column = 0
        title_column = 1
        restart_column = 3

        if self.back_button is not None:
            self.back_button.grid_configure(row=0, column=back_column, rowspan=2 if not compact else 1, sticky="w")
        self.title_label.grid_configure(row=0, column=title_column, sticky="w")
        self.subtitle_label.grid_configure(row=1, column=title_column, sticky="w", pady=(0, 8))
        self.restart_button.grid_configure(row=0, column=restart_column, rowspan=2, sticky="e")
        if getattr(self, "summary_tiles", None) is not None:
            if ultra_compact:
                self.summary_tiles.grid_remove()
            else:
                self.summary_tiles.grid()
                self.summary_tiles.grid_configure(row=0, column=2, rowspan=2, padx=(8, 6), sticky="e")
        self.title_label.configure(font=ctk.CTkFont(size=19 if ultra_compact else 21 if very_compact else 23, weight="bold"))
        self.subtitle_label.configure(font=ctk.CTkFont(size=11 if ultra_compact else 12))
        self.restart_button.configure(width=116 if ultra_compact else 124 if very_compact else 132)

        if very_compact:
            self.pasos_actions.grid_configure(row=1, column=0, columnspan=3, padx=12, pady=(0, 6), sticky="w")
            self.load_rows_button.grid_configure(row=0, column=1, padx=(0, 8), pady=0, sticky="w")
            self.template_button.grid_configure(row=0, column=2, padx=0, pady=0, sticky="w")
        else:
            self.pasos_actions.grid_configure(row=0, column=2, columnspan=1, padx=(8, 12), pady=5, sticky="e")
            self.load_rows_button.grid_configure(row=0, column=1, padx=(0, 8), pady=0, sticky="e")
            self.template_button.grid_configure(row=0, column=2, padx=0, pady=0, sticky="e")
        self.pasos_steps_label.configure(font=ctk.CTkFont(size=10 if ultra_compact else 11))

        def reset_columns(frame: ctk.CTkFrame, total: int, fill_col: int | None = None, minsize: int = 0) -> None:
            for column in range(14):
                try:
                    frame.grid_columnconfigure(column, weight=0, minsize=0)
                except Exception:
                    pass
            for column in range(total):
                try:
                    frame.grid_columnconfigure(column, weight=0, minsize=minsize)
                except Exception:
                    pass
            if fill_col is not None:
                frame.grid_columnconfigure(fill_col, weight=1, minsize=max(minsize, 150))

        def set_min_columns(frame: ctk.CTkFrame, columns: dict[int, int]) -> None:
            for column, minsize in columns.items():
                try:
                    frame.grid_columnconfigure(column, minsize=minsize)
                except Exception:
                    pass

        def place(widget, row: int, column: int, *, span: int = 1, padx=(0, 8), pady=(0, 6), sticky="w") -> None:
            widget.grid_configure(row=row, column=column, columnspan=span, padx=padx, pady=pady, sticky=sticky)

        def field_width(small: int, medium: int, large: int) -> int:
            if ultra_compact:
                return small
            if very_compact or medium_width:
                return medium
            return large

        side_pad = (10, 8)
        control_pad = (0, 10)

        # Bloque 1: Google Sheets y plantillas. El frame ocupa todo el ancho,
        # pero los controles administrativos mantienen anchos de formulario.
        reset_columns(self.sheets_frame, 10, 9)
        set_min_columns(self.sheets_frame, {1: 170, 3: 150, 5: 150, 7: 90, 9: 150})
        self.sheet_profile_combo.configure(width=field_width(220, 300, 360))
        self.sheet_url_entry.configure(width=field_width(220, 300, 360))
        self.sheet_name_combo.configure(width=field_width(130, 150, 170))
        self.sheet_start_row_entry.configure(width=88)
        self.sheet_max_rows_entry.configure(width=88)
        self.sheet_limit_mode_combo.configure(width=112)
        self.sheets_connect_button.configure(width=132)
        self.sheets_run_button.configure(width=150, text="Ejecutar Sheets")
        self.load_sheet_tabs_button.configure(width=128)

        place(self.sheet_profile_label, 0, 0, padx=side_pad, pady=(8, 5), sticky="w")
        place(self.sheet_profile_combo, 0, 1, padx=control_pad, pady=(8, 5), sticky="w")
        place(self.sheet_name_label, 0, 2, padx=(10, 8), pady=(8, 5), sticky="w")
        place(self.sheet_name_combo, 0, 3, pady=(8, 5), sticky="w")
        place(self.sheet_start_row_label, 0, 4, padx=(10, 8), pady=(8, 5), sticky="w")
        place(self.sheet_start_row_entry, 0, 5, pady=(8, 5), sticky="w")
        place(self.sheet_limit_mode_combo, 0, 6, padx=(10, 8), pady=(8, 5), sticky="w")
        place(self.sheet_max_rows_entry, 0, 7, pady=(8, 5), sticky="w")

        place(self.sheet_url_label, 1, 0, padx=side_pad, pady=(0, 5), sticky="w")
        place(self.sheet_url_entry, 1, 1, padx=control_pad, pady=(0, 5), sticky="w")
        place(self.sheet_url_config_button, 1, 2, padx=(0, 10), pady=(0, 5), sticky="w")
        place(self.sheet_profile_actions, 1, 3, span=3, pady=(0, 5), sticky="w")
        place(self.load_sheet_tabs_button, 1, 6, pady=(0, 5), sticky="w")

        place(self.sheet_checks, 2, 1, span=6, pady=(0, 5), sticky="w")
        place(self.sheets_run_actions, 3, 1, span=3, pady=(0, 10), sticky="w")
        place(self.sheets_status_label, 3, 4, span=6, padx=(10, 10), pady=(0, 10), sticky="w")

        if medium_width and not compact:
            self.sheet_profile_combo.configure(width=340)
            self.sheet_url_entry.configure(width=340)
            self.sheet_name_combo.configure(width=130)
            place(self.sheet_profile_label, 0, 0, padx=side_pad, pady=(8, 5), sticky="w")
            place(self.sheet_profile_combo, 0, 1, padx=control_pad, pady=(8, 5), sticky="w")
            place(self.sheet_name_label, 0, 2, padx=(10, 8), pady=(8, 5), sticky="w")
            place(self.sheet_name_combo, 0, 3, pady=(8, 5), sticky="w")
            place(self.sheet_start_row_label, 0, 4, padx=(10, 8), pady=(8, 5), sticky="w")
            place(self.sheet_start_row_entry, 0, 5, pady=(8, 5), sticky="w")
            place(self.sheet_limit_mode_combo, 0, 6, padx=(10, 8), pady=(8, 5), sticky="w")
            place(self.sheet_max_rows_entry, 0, 7, pady=(8, 5), sticky="w")

            place(self.sheet_url_label, 1, 0, padx=side_pad, pady=(0, 5), sticky="w")
            place(self.sheet_url_entry, 1, 1, padx=control_pad, pady=(0, 5), sticky="w")
            place(self.sheet_url_config_button, 1, 2, padx=(0, 10), pady=(0, 5), sticky="w")
            place(self.new_sheet_profile_button, 1, 3, pady=(0, 5), sticky="w")
            place(self.save_sheet_profile_button, 1, 4, pady=(0, 5), sticky="w")
            place(self.delete_sheet_profile_button, 1, 5, pady=(0, 5), sticky="w")
            place(self.load_sheet_tabs_button, 1, 6, pady=(0, 5), sticky="w")

            place(self.sheet_checks, 2, 1, span=6, pady=(0, 10), sticky="w")
            place(self.sheets_connect_button, 3, 1, pady=(0, 10), sticky="w")
            place(self.sheets_run_button, 3, 2, span=2, pady=(0, 10), sticky="w")
            place(self.sheets_status_label, 3, 4, span=5, padx=(10, 10), pady=(0, 10), sticky="w")
        elif compact:
            set_min_columns(self.sheets_frame, {1: 170, 2: 150, 3: 150, 5: 150})
            self.sheet_profile_combo.configure(width=field_width(240, 330, 360))
            self.sheet_url_entry.configure(width=field_width(240, 330, 360))
            self.sheet_name_combo.configure(width=140)
            place(self.sheet_profile_label, 0, 0, padx=side_pad, pady=(8, 5), sticky="w")
            place(self.sheet_profile_combo, 0, 1, pady=(8, 5), sticky="w")
            place(self.sheet_name_label, 0, 2, padx=(10, 8), pady=(8, 5), sticky="w")
            place(self.sheet_name_combo, 0, 3, pady=(8, 5), sticky="w")
            place(self.sheet_start_row_label, 1, 0, padx=side_pad, pady=(0, 5), sticky="w")
            place(self.sheet_start_row_entry, 1, 1, pady=(0, 5), sticky="w")
            place(self.sheet_limit_mode_combo, 1, 2, padx=(10, 8), pady=(0, 5), sticky="w")
            place(self.sheet_max_rows_entry, 1, 3, pady=(0, 5), sticky="w")
            place(self.sheet_url_label, 2, 0, padx=side_pad, pady=(0, 5), sticky="w")
            place(self.sheet_url_entry, 2, 1, pady=(0, 5), sticky="w")
            place(self.sheet_url_config_button, 2, 2, padx=(10, 8), pady=(0, 5), sticky="w")
            place(self.load_sheet_tabs_button, 2, 3, pady=(0, 5), sticky="w")
            place(self.sheet_profile_actions, 3, 1, span=3, pady=(0, 5), sticky="w")
            place(self.sheet_checks, 4, 1, span=6, pady=(0, 5), sticky="w")
            place(self.sheets_run_actions, 6, 1, span=3, pady=(0, 10), sticky="w")
            place(self.sheets_status_label, 6, 3, span=3, padx=(10, 10), pady=(0, 10), sticky="w")

        # Bloque 2: Perfil, medico y credenciales. Varias columnas, campos
        # controlados; no se estiran hasta el borde.
        reset_columns(self.profiles_frame, 8, 7)
        reset_columns(self.credentials_row, 6, 5)
        set_min_columns(self.profiles_frame, {1: 300, 2: 120, 3: 140, 7: 150})
        set_min_columns(self.credentials_row, {0: 180, 2: 180, 3: 70, 5: 150})
        self.profile_combo.configure(width=field_width(260, 330, 400))
        self.client_entry.configure(width=field_width(260, 330, 400))
        self.user_entry.configure(width=field_width(150, 180, 210))
        self.password_entry.configure(width=field_width(150, 180, 210))

        place(self.profile_combo, 0, 1, padx=control_pad, pady=(8, 5), sticky="w")
        place(self.profile_actions, 0, 2, span=2, pady=(8, 5), sticky="w")
        place(self.client_entry, 1, 1, padx=control_pad, pady=(0, 5), sticky="w")
        place(self.save_profile_button, 1, 2, pady=(0, 5), sticky="w")
        self.credentials_row.grid_configure(row=2, column=1, columnspan=6, padx=0, pady=(0, 8), sticky="w")
        self.user_entry.grid_configure(sticky="w")
        self.password_entry.grid_configure(sticky="w")

        # Bloque 3a: acciones principales.
        reset_columns(self.controls, 7, 6, minsize=150)
        visible_controls = []
        if self.open_button is not None:
            visible_controls.append(self.open_button)
        visible_controls.append(self.close_button)
        if self.run_batch_button is not None:
            visible_controls.append(self.run_batch_button)
        visible_controls.append(self.full_batch_button)
        if self.stop_button is not None:
            visible_controls.append(self.stop_button)
        if self.clear_panel_button is not None:
            visible_controls.append(self.clear_panel_button)
        buttons_per_row = 3 if ultra_compact else 5
        for index, button in enumerate(visible_controls):
            place(
                button,
                index // buttons_per_row,
                index % buttons_per_row,
                padx=(12, 8) if index % buttons_per_row == 0 else (0, 8),
                pady=(8, 4) if index < buttons_per_row else (0, 8),
            )

        # Bloque 3b: tipo, practica y diagnostico. Layout ERP: grupos cortos
        # en varias columnas, grilla de carga debajo a ancho total.
        reset_columns(self.top_row, 10, 9)
        set_min_columns(self.top_row, {2: 260, 4: 108, 5: 114, 6: 126, 8: 240, 9: 114, 10: 102})
        self.practice_combo.configure(width=field_width(260, 340, 400))
        self.diagnosis_preset_combo.configure(width=field_width(260, 340, 400))
        place(self.tipo_group, 0, 0, pady=(0, 5), sticky="w")
        place(self.practice_label, 0, 1, padx=(6, 8), pady=(0, 5), sticky="w")
        place(self.practice_combo, 0, 2, span=2, pady=(0, 5), sticky="w")
        place(self.practice_catalog_button, 0, 4, pady=(0, 5), sticky="w")
        place(self.apply_practice_button, 0, 5, pady=(0, 5), sticky="w")
        place(self.apply_practice_all_button, 0, 6, pady=(0, 5), sticky="w")

        place(self.diagnosis_label, 1, 0, padx=(0, 8), pady=(0, 5), sticky="w")
        place(self.diagnosis_preset_combo, 1, 1, span=3, pady=(0, 5), sticky="w")
        place(self.diag_actions, 1, 4, span=2, pady=(0, 5), sticky="w")
        place(self.headless_checkbox, 1, 6, padx=(12, 0), pady=(0, 5), sticky="w")
        place(self.quick_help_label, 2, 0, span=10, padx=(0, 0), pady=(2, 0), sticky="w")

        if compact:
            self.practice_combo.configure(width=field_width(260, 340, 400))
            self.diagnosis_preset_combo.configure(width=field_width(260, 340, 400))
            place(self.tipo_group, 0, 0, pady=(0, 5), sticky="w")
            place(self.practice_label, 1, 0, padx=(0, 8), pady=(0, 5), sticky="w")
            place(self.practice_combo, 1, 1, span=3, pady=(0, 5), sticky="w")
            place(self.practice_catalog_button, 1, 4, pady=(0, 5), sticky="w")
            place(self.apply_practice_button, 2, 1, pady=(0, 5), sticky="w")
            place(self.apply_practice_all_button, 2, 2, pady=(0, 5), sticky="w")
            place(self.diagnosis_label, 3, 0, padx=(0, 8), pady=(0, 5), sticky="w")
            place(self.diagnosis_preset_combo, 3, 1, span=3, pady=(0, 5), sticky="w")
            place(self.diag_actions, 3, 4, span=2, pady=(0, 5), sticky="w")
            place(self.headless_checkbox, 4, 2, padx=(12, 0), pady=(0, 5), sticky="w")
            place(self.quick_help_label, 5, 0, span=10, padx=(0, 0), pady=(2, 0), sticky="w")

        if self.report_row.winfo_ismapped():
            reset_columns(self.report_row, 5, 2)
            place(self.copy_row_button, 0, 0, padx=(0, 8), pady=(0, 0))
            place(self.copy_ome_button, 0, 1, padx=(0, 12), pady=(0, 0))
            self.output_entry.grid_configure(row=0, column=2, padx=(0, 10), pady=0, sticky="ew")
            place(self.pick_output_button, 0, 3, padx=(0, 10), pady=(0, 0), sticky="e")
            place(self.open_report_button, 0, 4, padx=(0, 0), pady=(0, 0), sticky="e")

        return

    def _normalize_code(self, value: str) -> str:
        text = (value or "").strip()
        if not text:
            return ""
        digits = "".join(ch if ch.isdigit() else " " for ch in text)
        digit_tokens = [token for token in digits.split() if token]
        if digit_tokens:
            longest = max(digit_tokens, key=len)
            if len(longest) >= 6:
                return longest[:6].strip().upper()
        if " - " in text:
            return text.split(" - ", 1)[0].strip().upper()
        if " â€” " in text:
            return text.split(" â€” ", 1)[0].strip().upper()
        for separator in ("/", "\\", ",", ";"):
            if separator in text:
                return text.split(separator, 1)[0].strip().upper()
        return text.split()[0].strip().upper()

    def _resolve_practice_code(self, value: str) -> str:
        text = (value or "").strip()
        if not text:
            return ""
        if text == MANUAL_PRACTICE_OPTION:
            return text

        normalized = self._normalize_code(text)
        if normalized.isdigit():
            return normalized

        normalized_search = self._normalize_search_text(text)
        for item in self.practice_catalog_items:
            if (
                normalized_search == self._normalize_search_text(item.display)
                or normalized_search == self._normalize_search_text(item.description)
                or normalized_search in self._normalize_search_text(item.display)
                or normalized_search in self._normalize_search_text(item.description)
            ):
                return item.code

        for option in self.practice_options:
            option_text = (option or "").strip()
            if not option_text or option_text == MANUAL_PRACTICE_OPTION:
                continue
            if normalized_search == self._normalize_search_text(option_text):
                option_code = self._normalize_code(option_text)
                if option_code:
                    return option_code
            if normalized_search in self._normalize_search_text(option_text):
                option_code = self._normalize_code(option_text)
                if option_code:
                    return option_code

        return normalized

    def _normalize_search_text(self, value: str) -> str:
        text = unicodedata.normalize("NFKD", value or "")
        text = "".join(ch for ch in text if not unicodedata.combining(ch))
        lowered = text.lower()
        for token in ("(", ")", "/", "\\", ",", ";", "-", "_", ".", ":"):
            lowered = lowered.replace(token, " ")
        return " ".join(lowered.split())

    def _looks_like_cie10_value(self, value: str) -> bool:
        code = self._normalize_code(value)
        if not code or len(code) < 3:
            return False
        if not code[0].isalpha():
            return False
        return code[1:].isdigit()

    def _resolve_specialist_practice_from_text(self, raw_text: str, *, allow_inactive: bool = False) -> str:
        if self.config_data.module_key != "ome_especialista":
            return ""
        original_hint = self._normalize_search_text(raw_text)
        hint = original_hint
        if not hint:
            return ""
        candidate_items = self.practice_catalog_items if allow_inactive else [
            item
            for item in self.practice_catalog_items
            if item.code in self.active_catalog_practices
            and (not self.active_catalog_modules or item.module_id in self.active_catalog_modules)
        ]
        for alias, code in SPECIALIST_PRACTICE_CODE_HINTS.items():
            if alias in original_hint:
                direct_match = next((item for item in candidate_items if item.code == code), None)
                if direct_match:
                    return direct_match.display
                if allow_inactive:
                    fallback_match = next((item for item in self.practice_catalog_items if item.code == code), None)
                    if fallback_match:
                        return fallback_match.display
        for noise in (
            "plan de salud",
            "consulta",
            "medico",
            "especialista",
            "control",
            "primera vez",
            "1 x",
        ):
            hint = hint.replace(noise, " ")
        hint = " ".join(hint.split())
        if not hint:
            return ""

        candidates = candidate_items
        if not candidates:
            candidates = [
                PracticeCatalogItem(module_id="", module_name="", code=self._normalize_code(option), description=option)
                for option in self.practice_options
                if option != MANUAL_PRACTICE_OPTION
            ]

        preferred: list[str] = []
        fallback: list[str] = []
        for item in candidates:
            haystack = self._normalize_search_text(f"{item.module_name} {item.description}")
            if hint and hint in haystack:
                display = item.display if item.description and " - " not in item.description else f"{item.code} - {item.description}".strip(" -")
                if "consulta con especialista" in self._normalize_search_text(item.description):
                    preferred.append(display)
                else:
                    fallback.append(display)
        if preferred:
            return preferred[0]
        if fallback:
            return fallback[0]
        return ""

    def _is_hidden_module_name(self, module_name: str) -> bool:
        normalized = (module_name or "").strip().lower()
        return any(keyword in normalized for keyword in HIDDEN_MODULE_KEYWORDS)

    def _default_active_catalog_practices(self, items: list[PracticeCatalogItem] | None = None) -> set[str]:
        if self.config_data.module_key == "ome_especialista" and items:
            specialist_codes = {
                item.code
                for item in items
                if (
                    "consulta con especialista" in self._normalize_search_text(item.description)
                    or "consulta con medico especialista" in self._normalize_search_text(item.description)
                )
            }
            if specialist_codes:
                return specialist_codes
        return {self._normalize_code(value) for value in self.config_data.practice_options}

    def _preferred_practice_option(self, options: list[str] | tuple[str, ...]) -> str:
        normalized_options = [value.strip() for value in options if value and value.strip()]
        if not normalized_options:
            return ""
        preferred_code = self._normalize_code(self.config_data.default_practice_code)
        if preferred_code:
            for option in normalized_options:
                if self._normalize_code(option).startswith(preferred_code):
                    return option
        return normalized_options[0]

    def _resolve_catalog_path_from_state(self, state: dict) -> Path:
        if state.get("catalog_path"):
            saved_path = Path(state.get("catalog_path", "")).expanduser()
            if saved_path.exists():
                return saved_path
        return next((path for path in _default_nomenclador_candidates() if path.exists()), DEFAULT_NOMENCLADOR_PATH)

    def _load_practice_catalog_state(self) -> dict:
        if not self.practice_catalog_file.exists():
            return {}
        try:
            return json.loads(self.practice_catalog_file.read_text(encoding="utf-8"))
        except Exception:
            return {}

    def _save_practice_catalog_state(self) -> None:
        payload = {
            "catalog_path": str(self.practice_catalog_path) if self.practice_catalog_path else "",
            "active_modules": sorted(self.active_catalog_modules),
            "active_practices": sorted(self.active_catalog_practices),
        }
        self.practice_catalog_file.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

    def _prepare_nomenclador_workbook_path(self, catalog_path: Path) -> Path:
        suffix = catalog_path.suffix.lower()
        if suffix != ".xls":
            return catalog_path
        cache_dir = Path(self.data_dir) / "catalog_cache"
        cache_dir.mkdir(parents=True, exist_ok=True)
        safe_stem = "".join(ch if ch.isalnum() else "_" for ch in catalog_path.stem).strip("_") or "nomenclador"
        stamp = int(catalog_path.stat().st_mtime)
        cache_path = cache_dir / f"{safe_stem}_{stamp}.xlsx"
        if cache_path.exists():
            return cache_path

        temp_dir = Path(tempfile.gettempdir())
        temp_path = temp_dir / f"{safe_stem}_{stamp}.xlsx"
        temp_path.parent.mkdir(parents=True, exist_ok=True)
        source = str(catalog_path)
        target = str(temp_path)
        powershell_script = f"""
$ErrorActionPreference = 'Stop'
$source = '{source.replace("'", "''")}'
$target = '{target.replace("'", "''")}'
$excel = New-Object -ComObject Excel.Application
$excel.Visible = $false
$excel.DisplayAlerts = $false
try {{
  $wb = $excel.Workbooks.Open($source)
  $wb.SaveAs($target, 51)
  $wb.Close($false) | Out-Null
}} finally {{
  if ($wb) {{
    try {{ $wb.Close($false) | Out-Null }} catch {{}}
  }}
  $excel.Quit()
  [System.Runtime.Interopservices.Marshal]::ReleaseComObject($excel) | Out-Null
  [GC]::Collect()
  [GC]::WaitForPendingFinalizers()
}}
"""
        try:
            subprocess.run(
                ["powershell", "-NoProfile", "-ExecutionPolicy", "Bypass", "-Command", powershell_script],
                check=True,
                capture_output=True,
                text=True,
                timeout=60,
            )
        except subprocess.CalledProcessError as exc:
            detail = (exc.stderr or exc.stdout or str(exc)).strip()
            raise RuntimeError(f"No se pudo convertir el nomenclador .xls: {detail}") from exc
        except subprocess.TimeoutExpired as exc:
            raise RuntimeError("La conversion del nomenclador .xls excedio el tiempo de espera.") from exc
        if not temp_path.exists():
            raise RuntimeError("No se pudo generar la copia .xlsx del nomenclador .xls.")
        temp_path.replace(cache_path)
        return cache_path

    def _read_nomenclador_catalog(self, catalog_path: Path) -> list[PracticeCatalogItem]:
        workbook_path = self._prepare_nomenclador_workbook_path(catalog_path)
        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
        if "Nomenclador" not in workbook.sheetnames:
            raise RuntimeError("El archivo no tiene una hoja 'Nomenclador'.")
        ws = workbook["Nomenclador"]
        items: list[PracticeCatalogItem] = []
        for row in ws.iter_rows(min_row=11, values_only=True):
            if not row or len(row) < 4:
                continue
            module_code = str(row[0] or "").strip()
            module_name = " ".join(str(row[1] or "").split())
            practice_code = str(row[2] or "").strip()
            description_candidates = [
                " ".join(str(row[index] or "").split())
                for index in (3, 4)
                if index < len(row)
            ]
            normalized_practice_code = self._normalize_code(practice_code)
            description = next(
                (
                    value
                    for value in description_candidates
                    if value and self._normalize_code(value) != normalized_practice_code
                ),
                "",
            )
            if not module_code or not module_name or not practice_code or not description:
                continue
            if self._is_hidden_module_name(module_name):
                continue
            items.append(
                PracticeCatalogItem(
                    module_id=f"{module_code}::{module_name}",
                    module_name=module_name,
                    code=self._normalize_code(practice_code),
                    description=description,
                )
            )
        items.sort(key=lambda item: (item.module_name.lower(), item.code))
        return items

    def _apply_practice_options(self, displays: list[str]) -> None:
        normalized = [value.strip() for value in displays if value.strip()]
        if not normalized:
            normalized = list(self.config_data.practice_options)
        if self.config_data.module_key == "ome_especialista" and MANUAL_PRACTICE_OPTION not in normalized:
            normalized = [MANUAL_PRACTICE_OPTION, *normalized]
        self.practice_options = normalized
        self.filtered_practice_options = list(self.practice_options)
        current_value = self.practice_var.get().strip()
        if current_value not in self.practice_options:
            self.practice_var.set(self._preferred_practice_option(self.practice_options))
        self.practice_combo.configure(values=self.filtered_practice_options)
        for item in self.row_widgets:
            if item.get("readonly"):
                continue
            widget = item.get("practice_widget")
            if isinstance(widget, ctk.CTkComboBox):
                widget.configure(values=self.practice_options)

    def _on_practice_combo_keyrelease(self, _event=None) -> None:
        query = self.practice_var.get().strip().lower()
        if not query:
            self.filtered_practice_options = list(self.practice_options)
        else:
            self.filtered_practice_options = [
                option for option in self.practice_options if query in option.lower()
            ]
        if not self.filtered_practice_options:
            self.filtered_practice_options = list(self.practice_options)
        self.practice_combo.configure(values=self.filtered_practice_options)

    def _on_practice_combo_focus_out(self, _event=None) -> None:
        current_value = self.practice_var.get().strip()
        if not current_value:
            if self.practice_options:
                self.practice_var.set(self._preferred_practice_option(self.practice_options))
        elif current_value not in self.practice_options:
            matches = [option for option in self.practice_options if current_value.lower() in option.lower()]
            if len(matches) == 1:
                self.practice_var.set(matches[0])
        self.filtered_practice_options = list(self.practice_options)
        self.practice_combo.configure(values=self.filtered_practice_options)

    def _initialize_practice_catalog(self) -> None:
        state = self._load_practice_catalog_state()
        candidate_path = self._resolve_catalog_path_from_state(state)
        if not candidate_path.exists():
            self._apply_practice_options(list(self.config_data.practice_options))
            return
        try:
            items = self._read_nomenclador_catalog(candidate_path)
        except Exception as exc:
            self._push_log(f"Nomenclador no cargado: {exc}")
            self._apply_practice_options(list(self.config_data.practice_options))
            return

        self.practice_catalog_path = candidate_path
        self.practice_catalog_items = items
        known_module_ids = {item.module_id for item in items}
        known_codes = {item.code for item in items}
        self.active_catalog_modules = {module_id for module_id in state.get("active_modules", []) if module_id in known_module_ids}
        self.active_catalog_practices = {
            self._normalize_code(code)
            for code in state.get("active_practices", [])
            if self._normalize_code(code) in known_codes
        }
        if not self.active_catalog_practices:
            default_codes = self._default_active_catalog_practices(items)
            self.active_catalog_practices = {code for code in default_codes if code in known_codes}
            self.active_catalog_modules = {
                item.module_id for item in items if item.code in self.active_catalog_practices
            }
        elif not self.active_catalog_modules:
            self.active_catalog_modules = {
                item.module_id for item in items if item.code in self.active_catalog_practices
            }
        self._refresh_practice_options_from_catalog(save=False)

    def _refresh_practice_options_from_catalog(self, *, save: bool = True) -> None:
        displays = [
            item.display
            for item in self.practice_catalog_items
            if item.code in self.active_catalog_practices and (not self.active_catalog_modules or item.module_id in self.active_catalog_modules)
        ]
        if self.practice_catalog_items and not displays:
            default_codes = self._default_active_catalog_practices(self.practice_catalog_items)
            known_codes = {item.code for item in self.practice_catalog_items}
            self.active_catalog_practices = {code for code in default_codes if code in known_codes}
            self.active_catalog_modules = {
                item.module_id for item in self.practice_catalog_items if item.code in self.active_catalog_practices
            }
            displays = [
                item.display
                for item in self.practice_catalog_items
                if item.code in self.active_catalog_practices and (not self.active_catalog_modules or item.module_id in self.active_catalog_modules)
            ]
        self._apply_practice_options(displays)
        if save and self.practice_catalog_path:
            self._save_practice_catalog_state()

    def _open_practice_catalog_manager(self) -> None:
        state = self._load_practice_catalog_state()
        catalog_path = self.practice_catalog_path or self._resolve_catalog_path_from_state(state)
        if not catalog_path.exists():
            initial_dir = next(
                (path.parent for path in _default_nomenclador_candidates() if path.parent.exists()),
                get_output_dir(),
            )
            selected = filedialog.askopenfilename(
                title="Seleccionar nomenclador PAMI",
                initialdir=str(initial_dir),
                filetypes=[("Excel", "*.xlsx *.xlsm *.xls")],
            )
            if not selected:
                return
            catalog_path = Path(selected).expanduser().resolve()
        try:
            items = self._read_nomenclador_catalog(catalog_path)
        except Exception as exc:
            messagebox.showerror("Nomenclador", str(exc))
            return

        active_modules = self.active_catalog_modules or set(state.get("active_modules", []))
        active_practices = self.active_catalog_practices or {self._normalize_code(code) for code in state.get("active_practices", [])}
        if not active_practices:
            active_practices = self._default_active_catalog_practices(items)
            active_modules = {item.module_id for item in items if item.code in active_practices}

        dialog = PracticeCatalogDialog(
            self,
            module_title=self.module_title,
            catalog_path=catalog_path,
            items=items,
            active_modules=active_modules,
            active_practices=active_practices,
        )
        self.wait_window(dialog)
        if not dialog.result:
            return
        self.practice_catalog_path = Path(dialog.result["catalog_path"])
        self.practice_catalog_items = items
        self.active_catalog_modules = set(dialog.result.get("active_modules", []))
        self.active_catalog_practices = {self._normalize_code(code) for code in dialog.result.get("active_practices", [])}
        self._refresh_practice_options_from_catalog(save=True)
        self._push_log(
            f"Nomenclador aplicado: {len(self.active_catalog_practices)} codigo(s) activos en {self.module_title.lower()}."
        )

    def _records_from_bulk_boxes(self) -> list[dict]:
        identifiers = [line.strip() for line in self.identifiers_text.get("1.0", "end").splitlines() if line.strip()]
        if not identifiers:
            raise RuntimeError("Pega al menos un BENEF o DNI para cargar filas.")

        diagnostics = [line.strip() for line in self.diagnostics_text.get("1.0", "end").splitlines() if line.strip()]
        if not diagnostics:
            raise RuntimeError("Pega al menos un diagnostico.")

        if len(diagnostics) == 1:
            diagnostics = diagnostics * len(identifiers)
        elif len(diagnostics) != len(identifiers):
            raise RuntimeError("La cantidad de diagnosticos debe ser 1 o coincidir con la cantidad de filas.")

        practices = [line.strip() for line in self.practices_text.get("1.0", "end").splitlines() if line.strip()]
        if not practices:
            base = self.practice_var.get().strip()
            if not base:
                raise RuntimeError("Selecciona una practica general o pega practicas por fila.")
            practices = [base] * len(identifiers)
        elif len(practices) == 1:
            practices = practices * len(identifiers)
        elif len(practices) != len(identifiers):
            raise RuntimeError("La cantidad de practicas debe ser 1 o coincidir con la cantidad de filas.")

        preset_diag = self._normalize_code(self.diagnosis_preset_var.get())
        selected_general_practice = self._resolve_practice_code(self.practice_var.get())
        records: list[dict] = []
        for identifier, diagnostico, practica in zip(identifiers, diagnostics, practices):
            resolved_diag = self._normalize_code(diagnostico)
            resolved_practice = self._resolve_practice_code(practica)
            manual_mode = resolved_practice == MANUAL_PRACTICE_OPTION or selected_general_practice == MANUAL_PRACTICE_OPTION

            if self.config_data.module_key == "ome_especialista" and not self._looks_like_cie10_value(diagnostico):
                inferred_practice = self._resolve_specialist_practice_from_text(diagnostico, allow_inactive=manual_mode)
                if inferred_practice:
                    resolved_practice = self._resolve_practice_code(inferred_practice)
                    if preset_diag:
                        resolved_diag = preset_diag
            if self.config_data.module_key == "ome_especialista" and resolved_practice == MANUAL_PRACTICE_OPTION:
                inferred_from_practice = self._resolve_specialist_practice_from_text(practica, allow_inactive=True)
                if inferred_from_practice:
                    resolved_practice = self._resolve_practice_code(inferred_from_practice)

            records.append(
                {
                    "modo": self._resolve_record_mode(identifier),
                    "afiliado": identifier,
                    "diagnostico": resolved_diag,
                    "practica": resolved_practice,
                }
            )
        return records

    def _load_rows_from_bulk(self) -> None:
        records = self._records_from_bulk_boxes()
        self._render_rows(records)
        self._push_log(f"Grilla cargada con {len(records)} fila(s).")

    def _clear_row_widgets(self) -> None:
        for item in self.row_widgets:
            item["frame"].destroy()
        self.row_widgets = []
        self.selected_row_var.set(-1)

    def _clear_result_table(self) -> None:
        self.current_result_rows = []
        self.selected_result_item = None
        self.selected_result_column = None
        for item_id in self.results_table.get_children():
            self.results_table.delete(item_id)

    def _show_edit_grid(self) -> None:
        self.grid_container.grid()
        self.report_row.grid()
        self.results_table_container.grid_remove()
        self.header_row.grid()
        self.rows_scroll.grid()

    def _show_result_table(self) -> None:
        self.grid_container.grid()
        self.report_row.grid()
        self.rows_scroll.grid_remove()
        self.header_row.grid_remove()
        self.results_table_container.grid()

    def _clear_rows(self) -> None:
        self._clear_row_widgets()
        self._clear_result_table()
        self.grid_container.grid_remove()
        self.report_row.grid_remove()
        for textbox in (self.identifiers_text, self.diagnostics_text, self.practices_text):
            textbox.delete("1.0", "end")
        self.output_var.set("")
        self._refresh_bulk_line_numbers()
        self.status_label.configure(text="Preparado para abrir CUP o ejecutar un lote.")
        self.summary_label.configure(text=f"Log en: {get_log_file()}")
        self._update_summary_tiles()
        self._push_log("Grilla y carga rapida limpiadas.")

    def _render_rows(self, records: list[dict]) -> None:
        self._clear_row_widgets()
        self._clear_result_table()
        self._show_edit_grid()
        for index, record in enumerate(records):
            readonly_row = bool(record.get("_readonly"))
            frame = ctk.CTkFrame(self.rows_scroll, fg_color="#ffffff")
            frame.grid(row=index, column=0, padx=0, pady=4, sticky="ew")
            for col in range(7):
                frame.grid_columnconfigure(col, weight=1 if col >= 2 else 0)

            ctk.CTkRadioButton(frame, text="", variable=self.selected_row_var, value=index, width=30).grid(
                row=0, column=0, padx=(6, 8), pady=6, sticky="w"
            )
            ctk.CTkLabel(frame, text=str(index + 1), width=40, text_color="#16324f").grid(
                row=0, column=1, padx=(0, 8), pady=6, sticky="w"
            )

            ident_var = ctk.StringVar(value=record["afiliado"])
            ident_entry = ctk.CTkEntry(frame, textvariable=ident_var, width=220)
            ident_entry.grid(row=0, column=2, padx=(0, 8), pady=6, sticky="ew")
            if readonly_row:
                ident_entry.configure(state="readonly")

            diag_var = ctk.StringVar(value=record["diagnostico"])
            if readonly_row:
                diag_widget = ctk.CTkEntry(frame, textvariable=diag_var, width=320)
                diag_widget.grid(row=0, column=3, padx=(0, 8), pady=6, sticky="ew")
                diag_widget.configure(state="readonly")
            else:
                diag_widget = ctk.CTkComboBox(frame, values=self.diagnosis_options, variable=diag_var, width=320)
                diag_widget.grid(row=0, column=3, padx=(0, 8), pady=6, sticky="ew")

            practice_var = ctk.StringVar(value=record["practica"])
            if readonly_row:
                practice_widget = ctk.CTkEntry(frame, textvariable=practice_var, width=140)
                practice_widget.grid(row=0, column=4, padx=(0, 6), pady=6, sticky="ew")
                practice_widget.configure(state="readonly")
            else:
                practice_widget = ctk.CTkComboBox(frame, values=self.practice_options, variable=practice_var, width=140)
                practice_widget.grid(row=0, column=4, padx=(0, 6), pady=6, sticky="ew")

            result_var = ctk.StringVar(value=record.get("resultado", ""))
            result_entry = ctk.CTkEntry(frame, textvariable=result_var, width=140)
            result_entry.grid(row=0, column=5, padx=(0, 6), pady=6, sticky="ew")
            result_entry.configure(state="readonly")

            ome_var = ctk.StringVar(value=record.get("nro_ome", ""))
            ome_entry = ctk.CTkEntry(frame, textvariable=ome_var, width=140)
            ome_entry.grid(row=0, column=6, padx=(0, 6), pady=6, sticky="ew")
            ome_entry.configure(state="readonly")

            self._bind_row_selection(frame, index)
            self._bind_row_selection(ident_entry, index)
            self._bind_row_selection(diag_widget, index)
            self._bind_row_selection(practice_widget, index)
            self._bind_row_selection(result_entry, index)
            self._bind_row_selection(ome_entry, index)

            self.row_widgets.append(
                {
                    "frame": frame,
                    "identifier_var": ident_var,
                    "diagnosis_var": diag_var,
                    "practice_var": practice_var,
                    "identifier_entry": ident_entry,
                    "diagnosis_widget": diag_widget,
                    "practice_widget": practice_widget,
                    "result_var": result_var,
                    "ome_var": ome_var,
                    "result_entry": result_entry,
                    "ome_entry": ome_entry,
                    "readonly": readonly_row,
                }
            )

        if self.row_widgets:
            self.selected_row_var.set(0)
        self._update_summary_tiles()

    def _apply_selected_diagnosis_to_row(self) -> None:
        line_status = self._apply_value_to_active_bulk_line(self.diagnostics_text, self.diagnosis_preset_var.get())
        if line_status == "applied":
            self._push_log("Diagnostico aplicado en la linea activa.")
            return
        if line_status == "blocked":
            messagebox.showinfo("Diagnostico", "La linea activa ya tiene un diagnostico cargado.")
            self._push_log("Diagnostico no aplicado: la linea activa ya tenía valor.")
            return
        index = self.selected_row_var.get()
        if index < 0 or index >= len(self.row_widgets):
            messagebox.showwarning("Fila", "Selecciona una fila para aplicar el diagnostico.")
            return

        current_value = self._normalize_code(self.row_widgets[index]["diagnosis_var"].get())
        if current_value:
            messagebox.showinfo("Diagnostico", f"La fila {index + 1} ya tiene un diagnostico cargado.")
            self._push_log(f"Diagnostico no aplicado: la fila {index + 1} ya tenía valor.")
            return

        self.row_widgets[index]["diagnosis_var"].set(self.diagnosis_preset_var.get())
        self._push_log(f"Diagnostico aplicado en la fila {index + 1}.")

    def _apply_selected_diagnosis_to_all(self) -> None:
        value = self.diagnosis_preset_var.get().strip()
        if not value:
            raise RuntimeError("Selecciona un diagnostico rapido antes de aplicar.")
        if not self.row_widgets:
            applied, skipped = self._fill_bulk_textbox_to_all_rows(self.diagnostics_text, value)
            self._push_log(f"Diagnostico aplicado a todas las lineas vacias: {applied} | omitidas: {skipped}.")
            if applied == 0:
                messagebox.showinfo("Diagnostico", "No habia lineas vacias para completar.")
            return
        applied = 0
        skipped = 0
        for item in self.row_widgets:
            current_value = self._normalize_code(item["diagnosis_var"].get())
            if current_value:
                skipped += 1
                continue
            item["diagnosis_var"].set(value)
            applied += 1
        self._push_log(f"Diagnostico aplicado a todas las filas vacias: {applied} | omitidas: {skipped}.")
        if applied == 0:
            messagebox.showinfo("Diagnostico", "No habia filas vacias para completar.")

    def _apply_general_practice_to_row(self) -> None:
        line_status = self._apply_value_to_active_bulk_line(self.practices_text, self.practice_var.get())
        if line_status == "applied":
            self._push_log("Practica aplicada en la linea activa.")
            return
        if line_status == "blocked":
            messagebox.showinfo("Practica", "La linea activa ya tiene una practica cargada.")
            self._push_log("Practica no aplicada: la linea activa ya tenía valor.")
            return
        index = self.selected_row_var.get()
        if index < 0 or index >= len(self.row_widgets):
            messagebox.showwarning("Fila", "Selecciona una fila para aplicar la practica.")
            return

        current_value = self._normalize_code(self.row_widgets[index]["practice_var"].get())
        if current_value:
            messagebox.showinfo("Practica", f"La fila {index + 1} ya tiene una practica cargada.")
            self._push_log(f"Practica no aplicada: la fila {index + 1} ya tenía valor.")
            return

        self.row_widgets[index]["practice_var"].set(self.practice_var.get())
        self._push_log(f"Practica aplicada en la fila {index + 1}.")

    def _apply_general_practice_to_all(self) -> None:
        value = self.practice_var.get().strip()
        if not value:
            raise RuntimeError("Selecciona una practica general antes de aplicar.")
        if not self.row_widgets:
            applied, skipped = self._fill_bulk_textbox_to_all_rows(self.practices_text, value)
            self._push_log(f"Practica aplicada a todas las lineas vacias: {applied} | omitidas: {skipped}.")
            if applied == 0:
                messagebox.showinfo("Practica", "No habia lineas vacias para completar.")
            return
        applied = 0
        skipped = 0
        for item in self.row_widgets:
            current_value = self._normalize_code(item["practice_var"].get())
            if current_value:
                skipped += 1
                continue
            item["practice_var"].set(value)
            applied += 1
        self._push_log(f"Practica aplicada a todas las filas vacias: {applied} | omitidas: {skipped}.")
        if applied == 0:
            messagebox.showinfo("Practica", "No habia filas vacias para completar.")

    def _bind_row_selection(self, widget, index: int) -> None:
        widget.bind("<FocusIn>", lambda _event, row=index: self.selected_row_var.set(row), add="+")
        widget.bind("<Button-1>", lambda _event, row=index: self.selected_row_var.set(row), add="+")

    def _handle_results_table_click(self, event) -> str:
        item_id = self.results_table.identify_row(event.y)
        column_id = self.results_table.identify_column(event.x)
        self.selected_result_item = item_id or None
        self.selected_result_column = column_id or None
        if item_id:
            self.results_table.focus(item_id)
            values = self.results_table.item(item_id, "values")
            row_number = self.results_table.index(item_id) + 1
            column_name = {
                "#1": "BENEF / DNI",
                "#2": "DIAGNOSTICO",
                "#3": "PRACTICA",
                "#4": "NRO OME",
            }.get(column_id, "FILA")
            cell_text = ""
            if column_id.startswith("#"):
                column_index = int(column_id[1:]) - 1
                if 0 <= column_index < len(values):
                    cell_text = str(values[column_index]).strip()
            self.status_label.configure(
                text=f"Reporte: fila {row_number} | columna {column_name}{f' | {cell_text}' if cell_text else ''}"
            )
        return "break"

    def _widget_belongs_to_textbox(self, widget, textbox: ctk.CTkTextbox) -> bool:
        current = widget
        inner_text = getattr(textbox, "_textbox", None)
        while current is not None:
            if current == textbox or current == inner_text:
                return True
            current = getattr(current, "master", None)
        return False

    def _get_active_bulk_line(self) -> tuple[ctk.CTkTextbox | None, str | None]:
        widget = self.focus_get()
        if widget is not None:
            for textbox in (self.identifiers_text, self.diagnostics_text, self.practices_text):
                if self._widget_belongs_to_textbox(widget, textbox):
                    line = textbox.index("insert").split(".", 1)[0]
                    self.last_bulk_source_textbox = textbox
                    self.last_bulk_line = line
                    return textbox, line

        return self.last_bulk_source_textbox, self.last_bulk_line

    def _remember_bulk_line(self, textbox: ctk.CTkTextbox) -> None:
        try:
            self.last_bulk_source_textbox = textbox
            self.last_bulk_line = textbox.index("insert").split(".", 1)[0]
        except Exception:
            pass

    def _get_non_empty_line_count(self, textbox: ctk.CTkTextbox) -> int:
        return len([line for line in textbox.get("1.0", "end").splitlines() if line.strip()])

    def _fill_bulk_textbox_to_all_rows(self, target_textbox: ctk.CTkTextbox, value: str) -> tuple[int, int]:
        identifiers = [line.strip() for line in self.identifiers_text.get("1.0", "end").splitlines() if line.strip()]
        if not identifiers:
            raise RuntimeError("Pega al menos un BENEF o DNI antes de aplicar a todas las filas.")

        target_lines = target_textbox.get("1.0", "end").splitlines()
        row_count = len(identifiers)
        while len(target_lines) < row_count:
            target_lines.append("")

        applied = 0
        skipped = 0
        for index in range(row_count):
            current_value = target_lines[index].strip() if index < len(target_lines) else ""
            if current_value:
                skipped += 1
                continue
            target_lines[index] = value
            applied += 1

        target_textbox.delete("1.0", "end")
        target_textbox.insert("1.0", "\n".join(target_lines[:row_count]))
        self.last_bulk_source_textbox = target_textbox
        self.last_bulk_line = str(row_count if row_count > 0 else 1)
        self._refresh_bulk_line_numbers()
        return applied, skipped

    def _set_line_numbers(self, target: ctk.CTkTextbox, count: int) -> None:
        count = max(1, count)
        text = "\n".join(str(index) for index in range(1, count + 1))
        target.configure(state="normal")
        target.delete("1.0", "end")
        target.insert("1.0", text)
        target.configure(state="disabled")

    def _refresh_bulk_line_numbers(self) -> None:
        identifier_count = self._get_non_empty_line_count(self.identifiers_text)
        diagnostics_count = self._get_non_empty_line_count(self.diagnostics_text)
        practices_count = self._get_non_empty_line_count(self.practices_text)
        base_count = max(identifier_count, diagnostics_count, practices_count, 1)
        self._set_line_numbers(self.bulk_line_numbers, base_count)
        self._align_bulk_scroll_with(self.identifiers_text)

    def _bulk_scroll_targets(self) -> tuple[ctk.CTkTextbox, ...]:
        return (
            self.bulk_line_numbers,
            self.identifiers_text,
            self.diagnostics_text,
            self.practices_text,
        )

    def _align_bulk_scroll_with(self, source: ctk.CTkTextbox) -> None:
        try:
            first, _last = source._textbox.yview()
        except Exception:
            return
        self._sync_bulk_text_scroll(source, str(first), "1.0")

    def _on_bulk_text_scroll(self, source: ctk.CTkTextbox, first: str, last: str) -> None:
        try:
            source._y_scrollbar.set(first, last)
        except Exception:
            pass
        self._sync_bulk_text_scroll(source, first, last)

    def _sync_bulk_text_scroll(self, source: ctk.CTkTextbox, first: str, _last: str) -> None:
        if self._bulk_scroll_syncing:
            return
        try:
            fraction = float(first)
        except (TypeError, ValueError):
            return
        self._bulk_scroll_syncing = True
        try:
            for target in self._bulk_scroll_targets():
                if target is source:
                    continue
                target._textbox.yview_moveto(fraction)
        finally:
            self._bulk_scroll_syncing = False

    def _apply_value_to_active_bulk_line(self, target_textbox: ctk.CTkTextbox, value: str) -> str:
        source_textbox, line = self._get_active_bulk_line()
        if source_textbox is None or line is None:
            return "no_target"

        line_number = int(line)
        start = f"{line}.0"
        end = f"{line}.end"
        current_line = target_textbox.get(start, end).strip()

        if current_line:
            return "blocked"
        else:
            target_textbox.insert(start, value)

        max_target_lines = 0
        if target_textbox in (self.diagnostics_text, self.practices_text):
            max_target_lines = self._get_non_empty_line_count(self.identifiers_text)

        target_textbox.focus_set()
        last_line = int(target_textbox.index("end-1c").split(".", 1)[0])
        next_line = line_number + 1

        if max_target_lines and next_line > max_target_lines:
            target_textbox.mark_set("insert", f"{line_number}.end")
            target_textbox.see(f"{line_number}.0")
            self.last_bulk_source_textbox = target_textbox
            self.last_bulk_line = str(line_number)
            self._refresh_bulk_line_numbers()
            return "applied"

        if next_line > last_line:
            target_textbox.insert("end", "\n")
        target_textbox.mark_set("insert", f"{next_line}.0")
        target_textbox.see(f"{next_line}.0")
        self.last_bulk_source_textbox = target_textbox
        self.last_bulk_line = str(next_line)
        self._refresh_bulk_line_numbers()
        return "applied"

    def _records_from_rows(self) -> list[dict]:
        records: list[dict] = []
        for item in self.row_widgets:
            afiliado = item["identifier_var"].get().strip()
            diagnostico = self._normalize_code(item["diagnosis_var"].get())
            practica = self._resolve_practice_code(item["practice_var"].get())
            if not afiliado:
                continue
            if not diagnostico or not practica:
                raise RuntimeError("Todas las filas de la grilla deben tener diagnostico y practica.")
            records.append(
                {
                    "modo": self._resolve_record_mode(afiliado),
                    "afiliado": afiliado,
                    "diagnostico": diagnostico,
                    "practica": practica,
                }
            )
        return records

    def _resolve_record_mode(self, raw_value: str) -> str:
        selected = self.search_mode_var.get().strip().upper() or "AUTO"
        cleaned = (raw_value or "").strip()
        digits = "".join(ch for ch in cleaned if ch.isdigit())
        if selected in {"BENEF", "DNI"}:
            return selected
        if digits and digits == cleaned:
            if len(digits) <= 8:
                return "DNI"
            if len(digits) >= 12:
                return "BENEF"
        return "BENEF"

    def _read_results_csv(self, path: Path) -> list[dict]:
        if not path.exists():
            return []

        rows: list[dict] = []
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            for row in reader:
                rows.append(
                    {
                        "modo": (row.get("modo") or "").strip().upper(),
                        "afiliado": (row.get("afiliado") or "").strip(),
                        "beneficio": (row.get("beneficio") or "").strip(),
                        "dni": (row.get("dni") or "").strip(),
                        "nombre": (row.get("nombre") or "").strip(),
                        "diagnostico": self._normalize_code(row.get("diagnostico") or ""),
                        "practica": self._normalize_code(row.get("practica") or ""),
                        "resultado": (row.get("resultado") or "").strip(),
                        "nro_ome": (row.get("nro_ome") or "").strip(),
                    }
                )
        return rows

    def _render_result_rows(self, rows: list[dict]) -> None:
        self._clear_row_widgets()
        self._clear_result_table()
        if not rows:
            self.grid_container.grid_remove()
            self.report_row.grid_remove()
            self._update_summary_tiles()
            return
        self._show_result_table()
        self.current_result_rows = [
            {
                "modo": (row.get("modo") or self.search_mode_var.get().strip().upper() or "BENEF").strip().upper(),
                "afiliado": (row.get("afiliado") or "").strip(),
                "beneficio": (row.get("beneficio") or "").strip(),
                "dni": (row.get("dni") or "").strip(),
                "diagnostico": self._normalize_code(row.get("diagnostico") or ""),
                "practica": self._normalize_code(row.get("practica") or ""),
                "resultado": (row.get("resultado") or "").strip(),
                "nro_ome": (row.get("nro_ome") or "").strip(),
            }
            for row in rows
        ]
        for row in self.current_result_rows:
            self.results_table.insert(
                "",
                "end",
                values=(
                    row["afiliado"],
                    row["diagnostico"],
                    row["practica"],
                    self._format_visible_nro_ome(row),
                ),
            )
        self._update_summary_tiles()

    def _collect_ome_numbers(self, rows: list[dict]) -> list[str]:
        return [(row.get("nro_ome") or "").strip() for row in rows if (row.get("nro_ome") or "").strip()]

    def _format_visible_result(self, result: str) -> str:
        normalized = (result or "").strip().upper()
        if normalized.startswith("ERROR"):
            return "ERROR"
        if normalized == "NO_DNI":
            return "NO DNI"
        if normalized == "NO_PAMI":
            return "NO PAMI"
        if normalized == "DOBLE_DNI":
            return "DOBLE DNI"
        if normalized == "BENEF_COMPLETADO":
            return "BENEF COMPLETADO"
        if normalized == "DNI_COMPLETADO":
            return "DNI COMPLETADO"
        if normalized == "YA_TIENE_OME":
            return "YA TIENE OME"
        if normalized == "LIMITE_ANUAL":
            return "LIMITE ANUAL"
        if normalized == "VERIFICAR_CREDENCIAL":
            return "Verificar credencial"
        return normalized

    def _format_visible_nro_ome(self, row: dict) -> str:
        visible_result = self._format_visible_result(row.get("resultado") or "")
        nro_ome = (row.get("nro_ome") or "").strip()
        if nro_ome:
            return nro_ome
        if visible_result == "YA TIENE OME":
            return visible_result
        if visible_result in {"GENERADA", "YA TIENE OME", "LIMITE", "LIMITE ANUAL", "BAJA", "ERROR", "NO PAMI", "DOBLE DNI", "BENEF COMPLETADO", "Verificar credencial"}:
            return visible_result
        return ""

    def copy_selected_row(self) -> None:
        if self.current_result_rows:
            item_id = self.selected_result_item
            if not item_id:
                messagebox.showwarning("Atencion", "Selecciona una fila para copiar.")
                return
            values = self.results_table.item(item_id, "values")
            row_text = "\t".join(str(value) for value in values)
        else:
            index = self.selected_row_var.get()
            if index < 0 or index >= len(self.row_widgets):
                messagebox.showwarning("Atencion", "Selecciona una fila para copiar.")
                return

            item = self.row_widgets[index]
            row_text = "\t".join(
                [
                    item["identifier_var"].get().strip(),
                    self._normalize_code(item["diagnosis_var"].get()),
                    self._normalize_code(item["practice_var"].get()),
                    self._format_visible_nro_ome(
                        {
                            "resultado": item["result_var"].get().strip(),
                            "nro_ome": item["ome_var"].get().strip(),
                        }
                    ),
                ]
            )
        self.clipboard_clear()
        self.clipboard_append(row_text)
        self.update()
        messagebox.showinfo("Copiado", "La fila seleccionada fue copiada al portapapeles.")

    def copy_ome_column(self) -> None:
        if self.current_result_rows:
            ome_numbers = [
                self._format_visible_nro_ome(row)
                for row in self.current_result_rows
                if self._format_visible_nro_ome(row)
            ]
        else:
            ome_numbers = [
                self._format_visible_nro_ome(
                    {
                        "resultado": item["result_var"].get().strip(),
                        "nro_ome": item["ome_var"].get().strip(),
                    }
                )
                for item in self.row_widgets
                if self._format_visible_nro_ome(
                    {
                        "resultado": item["result_var"].get().strip(),
                        "nro_ome": item["ome_var"].get().strip(),
                    }
                )
            ]
        if not ome_numbers:
            messagebox.showwarning("Atencion", "Todavia no hay numeros de OME para copiar.")
            return

        self.clipboard_clear()
        self.clipboard_append("\n".join(ome_numbers))
        self.update()
        messagebox.showinfo("Copiado", "La columna de OMEs fue copiada al portapapeles.")

    def _build_manual_input_file(self) -> Path:
        if not self.row_widgets:
            self._load_rows_from_bulk()
        records = self._records_from_rows()
        if not records:
            raise RuntimeError("No hay filas cargadas para ejecutar.")

        manual_input_path = get_output_dir() / f"{self.module_key}_lote_manual.csv"
        with manual_input_path.open("w", encoding="utf-8", newline="") as handle:
            handle.write("modo,afiliado,diagnostico,practica\n")
            for record in records:
                handle.write(
                    f"{record['modo']},{record['afiliado']},{record['diagnostico']},{record['practica']}\n"
                )

        modos = sorted({record.get("modo", "") for record in records if record.get("modo")})
        self._push_log(f"Entrada manual preparada: {len(records)} fila(s) | modos={', '.join(modos) or '-'}")
        return manual_input_path

    def _run_action(self, action, pending_status: str | None = None, pending_summary: str | None = None) -> None:
        if self.action_running:
            return
        self.stop_requested = False
        self._ensure_controller()
        self.controller.clear_stop_request()
        self.action_running = True
        self._set_controls_enabled(False)
        if pending_status:
            self.status_label.configure(text=pending_status)
        if pending_summary:
            self.summary_label.configure(text=pending_summary)
        self.controller_queue.put(action)

    def _ensure_controller(self) -> None:
        if self.controller is None:
            self.controller = PamiOmeController(
                log_callback=self._push_log,
                status_callback=self._push_status,
            )
        if self.controller_thread is None or not self.controller_thread.is_alive():
            self.controller_thread = threading.Thread(target=self._controller_loop, daemon=True)
            self.controller_thread.start()

    def _controller_loop(self) -> None:
        while True:
            action = self.controller_queue.get()
            if action is None:
                break
            try:
                action()
                self.event_queue.put(("action_finished", None))
            except Exception as exc:
                raw_error = str(exc)
                if self._is_playwright_driver_closed_error(raw_error) and self.controller is not None:
                    try:
                        self.controller.descartar_sesion_rota()
                    except Exception:
                        pass
                self.event_queue.put(("action_error", self._friendly_action_error(raw_error)))

    def _is_playwright_driver_closed_error(self, message: str) -> bool:
        normalized = (message or "").lower()
        return (
            "connection closed while reading from the driver" in normalized
            or "target page, context or browser has been closed" in normalized
            or "playwright" in normalized and "driver" in normalized and "closed" in normalized
        )

    def _friendly_action_error(self, message: str) -> str:
        if self._is_playwright_driver_closed_error(message):
            return (
                "Se corto la conexion con el navegador de CUP. "
                "La sesion quedo descartada; volve a ejecutar la accion para abrir un navegador limpio."
            )
        return message

    def _handle_open_pami(self) -> None:
        self._run_action(self._open_pami_with_profile)

    def _handle_close_browser(self) -> None:
        self._run_action(lambda: self._get_controller().cerrar_navegador())

    def _request_stop(self) -> None:
        if not self.action_running or self.stop_requested:
            return
        dialog = StopBatchDialog(self)
        self.wait_window(dialog)
        if not dialog.result:
            self._push_log("Detencion cancelada. El lote sigue en ejecucion.")
            return
        self.stop_requested = True
        if self.controller is not None:
            self.controller.request_stop()
        if self.stop_button is not None:
            self.stop_button.configure(state="disabled")
        self.status_label.configure(text="Detencion solicitada. Se detiene al terminar la fila actual.")
        self._push_log("Detencion solicitada. El lote se cortara al terminar la fila en curso.")

    def _apply_credentials_to_login(self) -> None:
        profile = self._current_profile_from_form()
        if not profile["usuario"]:
            raise RuntimeError("Ingresa un usuario para autocompletar.")
        self._upsert_profile(profile)
        self._get_controller().autocompletar_credenciales(profile["usuario"], profile["clave"])

    def _open_pami_with_profile(self) -> None:
        profile = self._current_profile_from_form()
        if profile["usuario"]:
            self._upsert_profile(profile)
        self._get_controller().abrir_pami(usuario=profile["usuario"] or None, clave=profile["clave"] or None)

    def _run_batch_with_profile(self) -> None:
        profile = self._current_profile_from_form()
        if not profile["usuario"] or not profile["clave"]:
            raise RuntimeError("Guarda o completa un perfil con usuario y clave antes de ejecutar el lote.")

        records = self._records_from_rows() if self.row_widgets else self._records_from_bulk_boxes()
        if not records:
            raise RuntimeError("No hay filas cargadas para ejecutar.")

        manual_input_path = get_output_dir() / f"{self.module_key}_lote_manual.csv"
        temp_csv_path = self._reserve_temp_output_path(get_output_dir() / f"{self.module_key}_lote_manual_resultados.csv")
        self.last_results_csv_path = None
        self.event_queue.put(("batch_started", {"output_path": str(temp_csv_path)}))

        result_rows = self._get_controller().ejecutar_lote_en_pagina_actual(
            records,
            progress_callback=self._push_batch_progress,
        )
        self._write_results_csv(temp_csv_path, result_rows)
        self.last_results_csv_path = temp_csv_path
        summary = self._summarize_result_rows(result_rows)
        summary["output_path"] = str(temp_csv_path)
        summary["rows"] = result_rows
        self.event_queue.put(("batch_summary", summary))

    def _run_full_batch_with_profile(self) -> None:
        profile = self._current_profile_from_form()
        if not profile["usuario"] or not profile["clave"]:
            raise RuntimeError("Guarda o completa un perfil con usuario y clave antes de ejecutar el lote.")

        records = self._records_from_rows() if self.row_widgets else self._records_from_bulk_boxes()
        if not records:
            raise RuntimeError("No hay filas cargadas para ejecutar.")

        input_path = get_output_dir() / f"{self.module_key}_lote_manual.csv"
        with input_path.open("w", encoding="utf-8", newline="") as handle:
            handle.write("modo,afiliado,diagnostico,practica\n")
            for record in records:
                handle.write(f"{record['modo']},{record['afiliado']},{record['diagnostico']},{record['practica']}\n")

        temp_csv_path = self._reserve_temp_output_path(build_default_output_path(input_path))
        self.last_results_csv_path = None
        self.event_queue.put(("batch_started", {"output_path": str(temp_csv_path)}))

        if not bool(self.headless_var.get()) and self.controller is not None and self.controller.sesion_activa():
            self._push_log("Sesion OME ya abierta: se reutiliza el navegador actual para el lote.")
            result_rows = self._get_controller().ejecutar_lote_en_pagina_actual(
                records,
                progress_callback=self._push_batch_progress,
            )
            self._write_results_csv(temp_csv_path, result_rows)
            self.last_results_csv_path = temp_csv_path
            summary = self._summarize_result_rows(result_rows)
            summary["output_path"] = str(temp_csv_path)
            summary["rows"] = result_rows
            self.event_queue.put(("batch_summary", summary))
            return

        result = run_batch_sync(
            input_path=input_path,
            output_path=temp_csv_path,
            user=profile["usuario"],
            password=profile["clave"],
            headless=bool(self.headless_var.get()),
            log_callback=self._push_log,
            progress_callback=self._push_batch_progress,
            stop_requested=lambda: self.stop_requested,
        )
        self.last_results_csv_path = temp_csv_path
        result["rows"] = self._read_results_csv(temp_csv_path)
        result["output_path"] = str(temp_csv_path)
        self.event_queue.put(("batch_summary", result))

    def _push_log(self, message: str) -> None:
        self.event_queue.put(("log", message))

    def _push_status(self, message: str) -> None:
        self.event_queue.put(("status", message))

    def _push_batch_progress(self, payload: dict) -> None:
        self.event_queue.put(("batch_progress", payload))

    def _get_controller(self) -> PamiOmeController:
        self._ensure_controller()
        assert self.controller is not None
        return self.controller

    def _process_ui_queue(self) -> None:
        try:
            while True:
                event, payload = self.event_queue.get_nowait()

                if event == "log":
                    pass
                elif event == "status":
                    self.status_label.configure(text=payload)
                elif event == "batch_started":
                    self._render_result_rows([])
                    self.output_var.set("")
                    self.status_label.configure(text="Lote OME en ejecucion...")
                    self.summary_label.configure(text=f"CSV temporal: {payload.get('output_path')} | Log en: {get_log_file()}")
                elif event == "batch_progress":
                    total = payload.get("total", 0) or 0
                    current = payload.get("current", 0) or 0
                    resultado = payload.get("resultado", "-")
                    afiliado = payload.get("afiliado", "-")
                    modo = payload.get("modo", "")
                    nro_ome = payload.get("nro_ome", "")
                    ome_suffix = f" | OME {nro_ome}" if nro_ome else ""
                    self.status_label.configure(text=f"Lote OME: {resultado} en {modo} {afiliado}{ome_suffix}")
                    progress_text = f"{current}/{total}" if total else str(current)
                    csv_text = str(self.last_results_csv_path) if self.last_results_csv_path else "-"
                    self.summary_label.configure(text=f"Progreso: {progress_text} | CSV temporal: {csv_text}")
                elif event == "batch_summary":
                    result_rows = payload.get("rows", []) or []
                    self._render_result_rows(result_rows)
                    sheet_error = str(payload.get("sheet_error", "") or "").strip()
                    status_text = "Lote OME finalizado."
                    if sheet_error:
                        status_text = "Lote OME finalizado. No se pudo escribir en Google Sheets."
                    self.status_label.configure(text=status_text)
                    ome_numbers = self._collect_ome_numbers(result_rows)
                    ome_text = ", ".join(ome_numbers[:8])
                    if len(ome_numbers) > 8:
                        ome_text += f" ... (+{len(ome_numbers) - 8})"
                    sheet_updates = payload.get("sheet_updates")
                    sheet_suffix = f" | Sheets actualizadas: {sheet_updates}" if sheet_updates is not None else ""
                    if sheet_error:
                        sheet_suffix += " | Sheets: sin permiso de escritura"
                    self.summary_label.configure(
                        text=(
                            f"OK: {payload.get('ok', 0)} | "
                            f"Existentes: {payload.get('generadas', 0)} | "
                            f"Limite: {payload.get('limite', 0)} | "
                            f"Bajas: {payload.get('bajas', 0)} | "
                            f"No DNI: {payload.get('no_dni', 0)} | "
                            f"Doble DNI: {payload.get('doble_dni', 0)} | "
                            f"Benef completados: {payload.get('benef_completados', 0)} | "
                            f"DNI completados: {payload.get('dni_completados', 0)} | "
                            f"Errores: {payload.get('errores', 0)} | "
                            f"OMEs: {ome_text or '-'} | "
                            f"CSV temporal: {payload.get('output_path')} | Excel: descargar a demanda"
                            f"{sheet_suffix}"
                        )
                    )
                    if self.config_data.show_completion_alert:
                        completion_message = (
                            f"OK: {payload.get('ok', 0)}\n"
                            f"Existentes: {payload.get('generadas', 0)}\n"
                            f"Limite: {payload.get('limite', 0)}\n"
                            f"Bajas: {payload.get('bajas', 0)}\n"
                            f"No DNI: {payload.get('no_dni', 0)}\n"
                            f"Doble DNI: {payload.get('doble_dni', 0)}\n"
                            f"Benef completados: {payload.get('benef_completados', 0)}\n"
                            f"DNI completados: {payload.get('dni_completados', 0)}\n"
                            f"Errores: {payload.get('errores', 0)}"
                        )
                        if sheet_updates is not None:
                            completion_message += f"\nSheets actualizadas: {sheet_updates}"
                        if sheet_error:
                            completion_message += (
                                "\n\nNo se pudo anotar en Google Sheets. "
                                "El resultado quedo guardado en el CSV temporal para cargarlo manualmente."
                            )
                            messagebox.showwarning("Lote finalizado con aviso", completion_message)
                        else:
                            messagebox.showinfo(
                                "Lote finalizado",
                                completion_message,
                            )
                elif event == "sheets_connected":
                    self.sheets_connected = True
                    self.sheets_status_var.set(f"Google Sheets conectado: {payload}")
                elif event == "sheets_status":
                    self.sheets_connected = False
                    self.sheets_status_var.set(str(payload or "Google Sheets no conectado"))
                elif event == "sheet_tabs_loaded":
                    tabs = [str(item or "").strip() for item in (payload or []) if str(item or "").strip()]
                    self.sheet_tabs = tabs
                    self.sheet_name_combo.configure(values=tabs or [""])
                    current = (self.sheet_name_var.get() or "").strip()
                    if tabs and current not in tabs:
                        match = next((item for item in tabs if item.lower() == current.lower()), "")
                        self.sheet_name_var.set(match or tabs[0])
                    self._update_sheet_template_display()
                    self._save_sheet_settings()
                elif event == "sheet_advance":
                    self._apply_sheet_start_row_advance(payload or {})
                elif event == "action_finished":
                    self.action_running = False
                    self.stop_requested = False
                    self._set_controls_enabled(True)
                elif event == "action_error":
                    self.action_running = False
                    self.stop_requested = False
                    self._set_controls_enabled(True)
                    self.status_label.configure(text="Ocurrio un error.")
                    self._push_log(f"ERROR: {payload}")
                    if str(payload or "") == OFFICE_FILE_MESSAGE:
                        self.sheets_status_var.set(OFFICE_FILE_MESSAGE)
                        messagebox.showwarning("Google Sheets", payload)
                    else:
                        messagebox.showerror("Error", payload)
        except queue.Empty:
            pass

        self.after(150, self._process_ui_queue)

    def _set_controls_enabled(self, enabled: bool) -> None:
        state = "normal" if enabled else "disabled"
        for button in (
            self.open_button,
            self.close_button,
            self.run_batch_button,
            self.full_batch_button,
            self.save_profile_button,
            self.new_profile_button,
            self.delete_profile_button,
            self.toggle_password_button,
            self.copy_row_button,
            self.copy_ome_button,
            self.pick_output_button,
            self.open_report_button,
            self.template_button,
            self.load_rows_button,
            self.load_sheet_tabs_button,
            self.sheets_connect_button,
            self.sheets_run_button,
            self.new_sheet_profile_button,
            self.save_sheet_profile_button,
            self.delete_sheet_profile_button,
            self.sheet_complete_benef_checkbox,
            self.sheet_complete_dni_checkbox,
            self.sheet_check_credential_checkbox,
            self.sheet_headless_checkbox,
            self.clear_rows_button,
            self.apply_diag_button,
            self.apply_diag_all_button,
            self.apply_practice_button,
            self.apply_practice_all_button,
            self.practice_catalog_button,
        ):
            if button is not None:
                button.configure(state=state)
        if hasattr(self, "sheet_profile_combo") and self.sheet_profile_combo is not None:
            self.sheet_profile_combo.configure(state="readonly" if enabled else "disabled")
        if hasattr(self, "sheet_name_combo") and self.sheet_name_combo is not None:
            self.sheet_name_combo.configure(state="readonly" if enabled else "disabled")
        if hasattr(self, "sheet_limit_mode_combo") and self.sheet_limit_mode_combo is not None:
            self.sheet_limit_mode_combo.configure(state="readonly" if enabled else "disabled")
        if self.stop_button is not None:
            stop_state = "disabled"
            if not enabled and self.config_data.show_stop_button and not self.stop_requested:
                stop_state = "normal"
            self.stop_button.configure(state=stop_state)
        self.profile_combo.configure(state=state)
        self.search_mode_combo.configure(state=state)
        self.practice_combo.configure(state=state)
        self.diagnosis_preset_combo.configure(state=state)
        self.client_entry.configure(state=state)
        self.user_entry.configure(state=state)
        self.password_entry.configure(state=state)
        self.sheet_url_entry.configure(state="readonly" if enabled else "disabled")
        if hasattr(self, "sheet_url_config_button") and self.sheet_url_config_button is not None:
            self.sheet_url_config_button.configure(state=state)
        self.sheet_start_row_entry.configure(state=state)
        self.sheet_max_rows_entry.configure(state=state)
        self.output_entry.configure(state=state)
        self.identifiers_text.configure(state=state)
        self.diagnostics_text.configure(state=state)
        self.practices_text.configure(state=state)
        self.headless_checkbox.configure(state=state)
        for item in self.row_widgets:
            item["frame"].winfo_children()[0].configure(state=state)

    def _write_results_csv(self, output_path: Path, rows: list[dict]) -> None:
        output_path.parent.mkdir(parents=True, exist_ok=True)
        fieldnames = ["sheet_row", "modo", "afiliado", "beneficio", "dni", "nombre", "diagnostico", "practica", "resultado", "nro_ome"]
        with output_path.open("w", encoding="utf-8", newline="") as handle:
            writer = csv.DictWriter(
                handle,
                fieldnames=fieldnames,
                extrasaction="ignore",
            )
            writer.writeheader()
            for row in rows:
                writer.writerow(row)

    def _reserve_temp_output_path(self, preferred_path: Path) -> Path:
        preferred_path.parent.mkdir(parents=True, exist_ok=True)
        if not preferred_path.exists():
            return preferred_path
        try:
            preferred_path.unlink()
            self._push_log(f"Salida temporal reiniciada: {preferred_path}")
            return preferred_path
        except PermissionError:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            alternative = preferred_path.with_name(f"{preferred_path.stem}_{timestamp}{preferred_path.suffix}")
            self._push_log(
                f"Salida temporal ocupada por otro proceso. Se usa un archivo nuevo: {alternative}"
            )
            return alternative

    def _update_summary_tiles(self) -> None:
        if not hasattr(self, "tile_grilla_value"):
            return
        try:
            results = getattr(self, "current_result_rows", None) or []
            if results:
                summary = self._summarize_result_rows(results)
                en_grilla = len(results)
                generadas = summary.get("generadas", 0) + summary.get("ok", 0)
                errores = summary.get("errores", 0)
            else:
                en_grilla = len(getattr(self, "row_widgets", []) or [])
                generadas = 0
                errores = 0
            self.tile_grilla_value.configure(text=str(en_grilla))
            self.tile_gen_value.configure(text=str(generadas))
            self.tile_err_value.configure(text=str(errores))
        except Exception:
            pass

    def _summarize_result_rows(self, rows: list[dict]) -> dict[str, int]:
        summary = {
            "ok": 0,
            "generadas": 0,
            "limite": 0,
            "bajas": 0,
            "no_dni": 0,
            "doble_dni": 0,
            "benef_completados": 0,
            "dni_completados": 0,
            "errores": 0,
        }
        for row in rows:
            resultado = (row.get("resultado") or "").upper()
            if resultado == "OK":
                summary["ok"] += 1
            elif resultado in {"GENERADA", "YA_TIENE_OME"}:
                summary["generadas"] += 1
            elif resultado in {"LIMITE", "LIMITE_ANUAL"}:
                summary["limite"] += 1
            elif resultado == "BAJA":
                summary["bajas"] += 1
            elif resultado == "NO_DNI":
                summary["no_dni"] += 1
            elif resultado == "DOBLE_DNI":
                summary["doble_dni"] += 1
            elif resultado == "BENEF_COMPLETADO":
                summary["benef_completados"] += 1
            elif resultado == "DNI_COMPLETADO":
                summary["dni_completados"] += 1
            elif resultado.startswith("ERROR"):
                summary["errores"] += 1
            else:
                summary["errores"] += 1
        return summary

    def _go_home(self) -> None:
        if self.action_running:
            messagebox.showwarning("Atencion", "Espera a que termine la accion actual antes de volver.")
            return
        if self.on_back:
            self.on_back()

    def _restart_app(self) -> None:
        if self.action_running:
            messagebox.showwarning("Atencion", "Espera a que termine la accion actual antes de reiniciar.")
            return
        root = self.winfo_toplevel()
        restart = getattr(root, "_restart_app", None)
        if callable(restart):
            restart(self.module_key)
        else:
            messagebox.showerror("Reiniciar app", "No se encontro el reinicio global de la suite.")

    def on_close(self) -> None:
        try:
            if self.controller is not None:
                self.controller_queue.put(self.controller.cerrar)
            if self.controller_thread is not None and self.controller_thread.is_alive():
                self.controller_queue.put(None)
        except Exception:
            pass
        finally:
            self.destroy()

    def _download_report_excel(self) -> None:
        csv_path = self.last_results_csv_path
        if csv_path is None or not csv_path.exists():
            messagebox.showwarning("Atencion", "Todavia no hay un resultado listo para exportar a Excel.")
            return

        suggested_name = f"{csv_path.stem}.xlsx"
        path = filedialog.asksaveasfilename(
            title="Guardar reporte Excel",
            defaultextension=".xlsx",
            initialfile=suggested_name,
            initialdir=str(get_output_dir()),
            filetypes=[("Excel", "*.xlsx")],
        )
        if not path:
            return

        excel_path = Path(path).expanduser().resolve()
        export_results_csv_to_excel(csv_path, excel_path)
        self.output_var.set(str(excel_path))
        self._push_log(f"Reporte Excel exportado en {excel_path}")
        messagebox.showinfo("Reporte listo", f"Se guardo el reporte en:\n{excel_path}")

    def _open_report_file(self) -> None:
        report_path = Path((self.output_var.get() or "").strip())
        if not report_path or not str(report_path):
            messagebox.showwarning("Atencion", "Todavia no hay un reporte Excel descargado para abrir.")
            return
        if not report_path.exists():
            messagebox.showwarning("Atencion", f"No encuentro el archivo:\n{report_path}")
            return
        try:
            os.startfile(str(report_path))
        except Exception as exc:
            raise RuntimeError(f"No se pudo abrir el reporte: {exc}") from exc

    def _download_template(self) -> None:
        default_path = get_output_dir() / self.config_data.template_filename
        path = filedialog.asksaveasfilename(
            title="Guardar modelo OME",
            defaultextension=".xlsx",
            initialfile=default_path.name,
            initialdir=str(default_path.parent),
            filetypes=[("Excel", "*.xlsx")],
        )
        if not path:
            return
        ruta = guardar_modelo_ome(path)
        self._push_log(f"Modelo OME guardado en {ruta}")
        messagebox.showinfo("Modelo listo", f"Se guardo el modelo en:\n{ruta}")

    def _load_sheet_settings(self) -> dict:
        try:
            if not self.sheet_settings_file.exists():
                return {}
            data = json.loads(self.sheet_settings_file.read_text(encoding="utf-8"))
            if isinstance(data, dict):
                return data
        except Exception:
            pass
        return {}

    def _merge_sheet_tabs(self, tabs: list[str] | None, current_sheet: str = "") -> list[str]:
        merged: list[str] = []
        for item in tabs or []:
            text = str(item or "").strip()
            if text and text not in merged:
                merged.append(text)
        current = str(current_sheet or "").strip()
        if current and current not in merged:
            merged.insert(0, current)
        return merged

    def _settings_sheet_tabs(self, settings: dict | None, current_sheet: str = "") -> list[str]:
        raw_tabs = settings.get("sheet_tabs") if isinstance(settings, dict) else None
        tabs: list[str] = []
        if isinstance(raw_tabs, list):
            tabs = [str(item or "").strip() for item in raw_tabs if str(item or "").strip()]
        fallback_sheet = ""
        if isinstance(settings, dict):
            fallback_sheet = str(settings.get("sheet_name", "") or "").strip()
        return self._merge_sheet_tabs(tabs, current_sheet or fallback_sheet)

    def _sheet_template_merge_signature(self, item: dict) -> tuple[str, str, str]:
        display_name = str(item.get("display_name") or item.get("internal_name") or item.get("label") or "").strip().lower()
        pami_key = ""
        if self.config_data.link_sheet_profiles_to_pami:
            pami_key = str(item.get("pami_usuario") or item.get("pami_nombre") or "").strip().lower()
        return (
            normalize_spreadsheet_url(str(item.get("spreadsheet_url", ""))).strip().lower(),
            display_name,
            pami_key,
        )

    def _extract_sheet_profiles(self, data: dict) -> list[dict]:
        profiles: list[dict] = []
        profile_by_template: dict[tuple[str, str, str], dict] = {}
        selected_profile_id = str(data.get("selected_profile_id", "")).strip() if isinstance(data, dict) else ""

        def add_profile(profile: dict) -> None:
            if not profile["spreadsheet_url"]:
                return
            signature = self._sheet_template_merge_signature(profile)
            existing = profile_by_template.get(signature)
            if existing:
                merged_tabs = self._merge_sheet_tabs(existing.get("sheet_tabs", []), profile.get("sheet_name", ""))
                for tab in profile.get("sheet_tabs", []) or []:
                    merged_tabs = self._merge_sheet_tabs(merged_tabs, tab)
                if selected_profile_id and profile.get("profile_id") == selected_profile_id:
                    existing.update(profile)
                existing["sheet_tabs"] = merged_tabs
                return
            profile_by_template[signature] = profile
            profiles.append(profile)

        raw_profiles = data.get("profiles") if isinstance(data, dict) else None
        if isinstance(raw_profiles, list):
            for item in raw_profiles:
                if not isinstance(item, dict):
                    continue
                profile = {
                    "profile_id": str(item.get("profile_id", "")).strip() or str(uuid4()),
                    "display_name": str(
                        item.get("display_name") or item.get("internal_name") or item.get("label") or ""
                    ).strip(),
                    "spreadsheet_url": normalize_spreadsheet_url(str(item.get("spreadsheet_url", ""))),
                    "sheet_name": str(item.get("sheet_name", "")).strip(),
                    "start_row": str(item.get("start_row", "2")).strip() or "2",
                    "max_rows": str(item.get("max_rows", "40")).strip() or "40",
                    "limit_mode": str(item.get("limit_mode", "cantidad")).strip() or "cantidad",
                    "check_credential": bool(item.get("check_credential", False)),
                    "pami_usuario": str(item.get("pami_usuario", "")).strip(),
                    "pami_nombre": str(item.get("pami_nombre", "")).strip(),
                    "sheet_tabs": self._settings_sheet_tabs(item, str(item.get("sheet_name", "")).strip()),
                }
                add_profile(profile)

        legacy_url = normalize_spreadsheet_url(str(data.get("spreadsheet_url", ""))) if isinstance(data, dict) else ""
        if legacy_url:
            legacy_profile = {
                "profile_id": str(uuid4()),
                "display_name": str(
                    data.get("display_name") or data.get("internal_name") or data.get("label") or ""
                ).strip(),
                "spreadsheet_url": legacy_url,
                "sheet_name": str(data.get("sheet_name", "")).strip(),
                "start_row": str(data.get("start_row", "2")).strip() or "2",
                "max_rows": str(data.get("max_rows", "40")).strip() or "40",
                "limit_mode": str(data.get("limit_mode", "cantidad")).strip() or "cantidad",
                "check_credential": bool(data.get("check_credential", False)),
                "pami_usuario": str(data.get("pami_usuario", "")).strip(),
                "pami_nombre": str(data.get("pami_nombre", "")).strip(),
                "sheet_tabs": self._settings_sheet_tabs(data, str(data.get("sheet_name", "")).strip()),
            }
            add_profile(legacy_profile)

        return profiles[:25]

    def _sheet_profile_signature(self, item: dict) -> tuple[str, str, str, str, str]:
        return (
            str(item.get("spreadsheet_url", "")).strip().lower(),
            str(item.get("sheet_name", "")).strip().lower(),
            str(item.get("start_row", "2")).strip() or "2",
            str(item.get("max_rows", "40")).strip() or "40",
            self._sheet_limit_mode_value(item.get("limit_mode", "cantidad")),
        )

    def _sheet_profile_key(self, item: dict) -> str:
        return (
            str(item.get("profile_id", "")).strip()
            or "|".join(self._sheet_profile_signature(item))
        )

    def _sheet_profile_options(self) -> list[str]:
        self.sheet_profile_lookup = {}
        options: list[str] = []
        for item in self.sheet_profiles:
            display = self._format_sheet_profile_entry(item)
            if display in self.sheet_profile_lookup:
                display = f"{display} | {self._sheet_profile_key(item)[-6:]}"
            self.sheet_profile_lookup[display] = item
            options.append(display)
        return options

    def _format_sheet_profile_entry(self, item: dict) -> str:
        display_name = self._sheet_template_name(item)
        detail = display_name
        pami_nombre = str(item.get("pami_nombre", "")).strip()
        pami_usuario = str(item.get("pami_usuario", "")).strip()
        pami_label = pami_nombre or pami_usuario
        if self.config_data.link_sheet_profiles_to_pami and pami_label and pami_label.lower() not in display_name.lower():
            detail += f" | {pami_label}"
        return detail

    def _sheet_template_name(self, item: dict | None) -> str:
        item = item or {}
        display_name = str(item.get("display_name") or item.get("internal_name") or item.get("label") or "").strip()
        if display_name:
            return display_name
        sheet_name = str(item.get("sheet_name", "") or "").strip()
        pami_name = str(item.get("pami_nombre") or item.get("pami_usuario") or "").strip()
        if sheet_name and pami_name:
            return f"{sheet_name} - {pami_name}"
        if pami_name:
            return pami_name
        if sheet_name:
            return sheet_name
        url = normalize_spreadsheet_url(str(item.get("spreadsheet_url", "") or ""))
        if url:
            spreadsheet_id = extract_spreadsheet_id(url)
            return f"Plantilla {spreadsheet_id[-6:]}" if spreadsheet_id else "Plantilla configurada"
        return "Sin plantilla"

    def _current_sheet_template_name(self) -> str:
        selected = self._selected_sheet_profile()
        if selected:
            return self._sheet_template_name(selected)
        display = self.sheet_template_display_var.get().strip() if hasattr(self, "sheet_template_display_var") else ""
        if display and display != "Sin plantilla":
            return display
        return self._sheet_template_name(self._current_sheet_profile_from_form())

    def _update_sheet_template_display(self) -> None:
        if hasattr(self, "sheet_template_display_var"):
            self.sheet_template_display_var.set(self._current_sheet_template_name())

    def _apply_sheet_template_config(self, *, display_name: str, spreadsheet_url: str, sheet_name: str) -> None:
        normalized_url = normalize_spreadsheet_url(spreadsheet_url)
        if not normalized_url:
            raise RuntimeError("Ingresa una URL de Google Sheets.")
        if is_office_file_url(spreadsheet_url):
            raise RuntimeError(OFFICE_FILE_MESSAGE)
        if not sheet_name.strip():
            raise RuntimeError("Ingresa el nombre de la pestana.")

        self.sheet_internal_name_var.set(display_name.strip())
        self.sheet_url_var.set(normalized_url)
        self.sheet_name_var.set(sheet_name.strip())
        self.sheet_tabs = self._merge_sheet_tabs(self.sheet_tabs, sheet_name.strip())
        if hasattr(self, "sheet_name_combo"):
            self.sheet_name_combo.configure(values=self.sheet_tabs or [""])
        self._save_current_sheet_profile()
        self._update_sheet_template_display()

    def _open_sheet_template_dialog(self) -> None:
        if self.action_running:
            messagebox.showwarning("Plantilla", "Espera a que termine la accion actual antes de cambiar la plantilla.")
            return

        selected = self._selected_sheet_profile() or self._current_sheet_profile_from_form()
        dialog = ctk.CTkToplevel(self)
        dialog.title(f"Configurar plantilla - {self.module_title}")
        dialog.geometry("680x245")
        dialog.transient(self.winfo_toplevel())
        dialog.grab_set()
        dialog.grid_columnconfigure(1, weight=1)

        name_var = ctk.StringVar(value=str(selected.get("display_name", "") or self._current_sheet_template_name()).strip())
        url_var = ctk.StringVar(value=normalize_spreadsheet_url(str(selected.get("spreadsheet_url", "") or self.sheet_url_var.get())))
        sheet_var = ctk.StringVar(value=str(selected.get("sheet_name", "") or self.sheet_name_var.get()).strip())

        ctk.CTkLabel(dialog, text="Nombre interno", text_color="#16324f").grid(row=0, column=0, padx=(14, 8), pady=(16, 8), sticky="w")
        name_entry = ctk.CTkEntry(dialog, textvariable=name_var, height=30)
        name_entry.grid(row=0, column=1, padx=(0, 14), pady=(16, 8), sticky="ew")

        ctk.CTkLabel(dialog, text="URL", text_color="#16324f").grid(row=1, column=0, padx=(14, 8), pady=8, sticky="w")
        url_entry = ctk.CTkEntry(dialog, textvariable=url_var, height=30)
        url_entry.grid(row=1, column=1, padx=(0, 14), pady=8, sticky="ew")

        ctk.CTkLabel(dialog, text="Pestaña", text_color="#16324f").grid(row=2, column=0, padx=(14, 8), pady=8, sticky="w")
        sheet_row = ctk.CTkFrame(dialog, fg_color="transparent")
        sheet_row.grid(row=2, column=1, padx=(0, 14), pady=8, sticky="ew")
        sheet_row.grid_columnconfigure(0, weight=1)
        known_tabs = self._merge_sheet_tabs(selected.get("sheet_tabs", self.sheet_tabs), sheet_var.get())
        sheet_entry = ctk.CTkComboBox(
            sheet_row,
            values=known_tabs or [""],
            variable=sheet_var,
            height=30,
            state="normal",
        )
        sheet_entry.grid(row=0, column=0, padx=(0, 8), pady=0, sticky="ew")

        def load_tabs() -> None:
            normalized_url = normalize_spreadsheet_url(url_var.get())
            if not normalized_url:
                messagebox.showwarning("Plantilla", "Ingresa una URL de Google Sheets para cargar pestañas.")
                return
            if is_office_file_url(url_var.get()):
                messagebox.showwarning("Plantilla", OFFICE_FILE_MESSAGE)
                return
            try:
                tabs = self._sheet_tabs(normalized_url)
            except Exception as exc:
                messagebox.showerror("Plantilla", f"No se pudieron cargar las pestañas:\n{exc}")
                return
            self.sheet_tabs = tabs
            sheet_entry.configure(values=tabs or [""])
            if tabs:
                current_value = sheet_var.get().strip()
                if current_value not in tabs:
                    match = next((item for item in tabs if item.lower() == current_value.lower()), "")
                    sheet_var.set(match or tabs[0])
            if hasattr(self, "sheet_name_combo"):
                self.sheet_name_combo.configure(values=tabs or [""])
            messagebox.showinfo("Plantilla", f"Pestañas cargadas: {len(tabs)}")

        ctk.CTkButton(
            sheet_row,
            text="Cargar pestañas",
            width=132,
            height=30,
            fg_color="#66788a",
            hover_color="#536577",
            command=load_tabs,
        ).grid(row=0, column=1, padx=0, pady=0, sticky="e")

        hint = ctk.CTkLabel(
            dialog,
            text="El link queda guardado en la app; en el panel solo se muestra este nombre.",
            text_color="#51657a",
            font=ctk.CTkFont(size=11),
        )
        hint.grid(row=3, column=0, columnspan=2, padx=14, pady=(2, 12), sticky="w")

        buttons = ctk.CTkFrame(dialog, fg_color="transparent")
        buttons.grid(row=4, column=0, columnspan=2, padx=14, pady=(0, 14), sticky="e")

        def save() -> None:
            try:
                self._apply_sheet_template_config(
                    display_name=name_var.get(),
                    spreadsheet_url=url_var.get(),
                    sheet_name=sheet_var.get(),
                )
            except Exception as exc:
                messagebox.showerror("Plantilla", str(exc))
                return
            dialog.destroy()

        ctk.CTkButton(
            buttons,
            text="Cancelar",
            width=92,
            fg_color="#9aafc3",
            hover_color="#7f95aa",
            command=dialog.destroy,
        ).grid(row=0, column=0, padx=(0, 8), sticky="e")
        ctk.CTkButton(buttons, text="Guardar", width=92, command=save).grid(row=0, column=1, sticky="e")
        name_entry.focus_set()

    def _current_sheet_profile_from_form(self) -> dict:
        pami_nombre = ""
        pami_usuario = ""
        if self.config_data.link_sheet_profiles_to_pami and hasattr(self, "profile_name_var"):
            pami_nombre = (self.profile_name_var.get() or "").strip()
        if self.config_data.link_sheet_profiles_to_pami and hasattr(self, "profile_user_var"):
            pami_usuario = (self.profile_user_var.get() or "").strip()
        return {
            "profile_id": self.selected_sheet_profile_id or str(uuid4()),
            "display_name": (self.sheet_internal_name_var.get() or "").strip()
            if hasattr(self, "sheet_internal_name_var")
            else "",
            "spreadsheet_url": normalize_spreadsheet_url((self.sheet_url_var.get() or "").strip()),
            "sheet_name": (self.sheet_name_var.get() or "").strip(),
            "sheet_tabs": self._merge_sheet_tabs(self.sheet_tabs, (self.sheet_name_var.get() or "").strip()),
            "start_row": (self.sheet_start_row_var.get() or "").strip() or "2",
            "max_rows": (self.sheet_max_rows_var.get() or "").strip() or "40",
            "limit_mode": self._sheet_limit_mode_value(),
            "check_credential": bool(self.sheet_check_credential_var.get()) if hasattr(self, "sheet_check_credential_var") else False,
            "pami_usuario": pami_usuario,
            "pami_nombre": pami_nombre,
        }

    def _restore_sheet_profile_selection(self) -> None:
        options = self._sheet_profile_options()
        if hasattr(self, "sheet_profile_combo"):
            self.sheet_profile_combo.configure(values=options or [""])
        current_key = (
            str(self.selected_sheet_profile_id or "").strip()
            or self._sheet_profile_key(self._current_sheet_profile_from_form())
        )
        for display, item in self.sheet_profile_lookup.items():
            if self._sheet_profile_key(item) == current_key:
                self.selected_sheet_profile_id = self._sheet_profile_key(item)
                self.sheet_profile_var.set(display)
                self._update_sheet_template_display()
                return
        self.sheet_profile_var.set("")
        self._update_sheet_template_display()

    def _on_sheet_profile_selected(self, selected: str) -> None:
        profile = self.sheet_profile_lookup.get(selected)
        if not profile:
            return
        self._apply_sheet_profile(profile, select_pami=self.config_data.link_sheet_profiles_to_pami, save=True)

    def _apply_sheet_profile(self, profile: dict, *, select_pami: bool = True, save: bool = False) -> None:
        if not profile:
            return
        self._syncing_sheet_profile = True
        try:
            self.selected_sheet_profile_id = self._sheet_profile_key(profile)
            if hasattr(self, "sheet_internal_name_var"):
                self.sheet_internal_name_var.set(profile.get("display_name", ""))
            self.sheet_url_var.set(profile.get("spreadsheet_url", ""))
            self.sheet_name_var.set(profile.get("sheet_name", ""))
            sheet_name = (self.sheet_name_var.get() or "").strip()
            self.sheet_tabs = self._merge_sheet_tabs(profile.get("sheet_tabs", self.sheet_tabs), sheet_name)
            if hasattr(self, "sheet_name_combo"):
                self.sheet_name_combo.configure(values=self.sheet_tabs or [""])
            self.sheet_start_row_var.set(profile.get("start_row", "2"))
            self.sheet_max_rows_var.set(profile.get("max_rows", "40"))
            if hasattr(self, "sheet_limit_mode_var"):
                self.sheet_limit_mode_var.set(self._sheet_limit_mode_label(profile.get("limit_mode", "cantidad")))
            if hasattr(self, "sheet_check_credential_var"):
                self.sheet_check_credential_var.set(bool(profile.get("check_credential", self.sheet_check_credential_var.get())))
            if select_pami:
                self._select_pami_profile_for_sheet(profile)
            self._update_sheet_template_display()
            self._restore_sheet_profile_selection()
            if save:
                self._save_sheet_settings()
        finally:
            self._syncing_sheet_profile = False

    def _new_sheet_profile(self) -> None:
        self.selected_sheet_profile_id = ""
        self.sheet_profile_var.set("")
        if hasattr(self, "sheet_internal_name_var"):
            self.sheet_internal_name_var.set("")
        self.sheet_url_var.set("")
        self.sheet_name_var.set("Mc Dube")
        self.sheet_tabs = self._merge_sheet_tabs([], "Mc Dube")
        if hasattr(self, "sheet_name_combo"):
            self.sheet_name_combo.configure(values=self.sheet_tabs)
        if hasattr(self, "sheet_check_credential_var"):
            self.sheet_check_credential_var.set(False)
        self.sheet_start_row_var.set("2")
        self.sheet_max_rows_var.set("40")
        if hasattr(self, "sheet_limit_mode_var"):
            self.sheet_limit_mode_var.set("Cantidad")
        self._update_sheet_template_display()
        self._save_sheet_settings()

    def _delete_current_sheet_profile(self) -> None:
        selected = (self.sheet_profile_var.get() or "").strip()
        profile = self.sheet_profile_lookup.get(selected)

        if not profile:
            current_key = self._sheet_profile_key(self._current_sheet_profile_from_form())
            for item in self.sheet_profiles:
                if self._sheet_profile_key(item) == current_key:
                    profile = item
                    break

        if not profile:
            messagebox.showwarning("Google Sheets", "No hay una plantilla guardada seleccionada para borrar.")
            return

        if not messagebox.askyesno(
            "Google Sheets",
            f"Vas a borrar la plantilla guardada '{self._sheet_template_name(profile)}'.",
        ):
            return

        target_key = self._sheet_profile_key(profile)
        self.sheet_profiles = [item for item in self.sheet_profiles if self._sheet_profile_key(item) != target_key]
        if self.selected_sheet_profile_id == target_key:
            self.selected_sheet_profile_id = ""

        self.sheet_profile_var.set("")
        if hasattr(self, "sheet_internal_name_var"):
            self.sheet_internal_name_var.set("")
        self.sheet_url_var.set("")
        self.sheet_name_var.set("Mc Dube")
        self.sheet_tabs = self._merge_sheet_tabs([], "Mc Dube")
        if hasattr(self, "sheet_name_combo"):
            self.sheet_name_combo.configure(values=self.sheet_tabs)
        if hasattr(self, "sheet_check_credential_var"):
            self.sheet_check_credential_var.set(False)
        self.sheet_start_row_var.set("2")
        self.sheet_max_rows_var.set("40")
        if hasattr(self, "sheet_limit_mode_var"):
            self.sheet_limit_mode_var.set("Cantidad")
        self._update_sheet_template_display()
        self._save_sheet_settings()
        self._restore_sheet_profile_selection()
        self._push_log(f"Plantilla borrada: {self._sheet_template_name(profile)} | {profile.get('spreadsheet_url', '')}")

    def _save_current_sheet_profile(self) -> None:
        profile = self._current_sheet_profile_from_form()
        if not profile["spreadsheet_url"]:
            messagebox.showwarning("Google Sheets", "Ingresa la URL de la hoja para guardarla.")
            return
        if is_office_file_url(self.sheet_url_var.get() or ""):
            messagebox.showwarning("Google Sheets", OFFICE_FILE_MESSAGE)
            return
        if not profile["sheet_name"]:
            messagebox.showwarning("Google Sheets", "Ingresa el nombre de la pestana para guardarla.")
            return

        target_key = self._sheet_profile_key(profile)
        self.selected_sheet_profile_id = target_key
        updated = False
        for idx, item in enumerate(self.sheet_profiles):
            if self._sheet_profile_key(item) == target_key:
                self.sheet_profiles[idx] = profile
                updated = True
                break
        if not updated:
            self.sheet_profiles.insert(0, profile)

        self.sheet_profiles = self.sheet_profiles[:25]
        self._save_sheet_settings()
        self._restore_sheet_profile_selection()
        selected = self._format_sheet_profile_entry(profile)
        self.sheet_profile_var.set(selected)
        self._update_sheet_template_display()
        self._push_log(f"Plantilla guardada: {self._sheet_template_name(profile)} | {profile['spreadsheet_url']}")

    def _save_sheet_settings(self) -> None:
        if is_office_file_url(self.sheet_url_var.get() or ""):
            return
        payload = self._current_sheet_profile_from_form()
        selected_profile = self._selected_sheet_profile()
        if selected_profile and self._sheet_profile_key(selected_profile) == self._sheet_profile_key(payload):
            if (
                self.config_data.link_sheet_profiles_to_pami
                and not payload.get("pami_usuario")
                and selected_profile.get("pami_usuario")
            ):
                payload["pami_usuario"] = selected_profile.get("pami_usuario", "")
                payload["pami_nombre"] = selected_profile.get("pami_nombre", "")
        payload["selected_profile_id"] = str(self.selected_sheet_profile_id or "").strip()
        if hasattr(self, "profile_user_var"):
            payload["selected_pami_usuario"] = (self.profile_user_var.get() or "").strip()
        if hasattr(self, "profile_name_var"):
            payload["selected_pami_nombre"] = (self.profile_name_var.get() or "").strip()
        payload["profiles"] = self.sheet_profiles
        self.sheet_settings_file.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

    def _apply_sheet_start_row_advance(self, payload: dict) -> None:
        next_row = payload.get("next_row")
        if not next_row:
            return

        try:
            next_row_int = int(next_row)
        except (TypeError, ValueError):
            return

        spreadsheet_url = str(payload.get("spreadsheet_url", "")).strip()
        sheet_name = str(payload.get("sheet_name", "")).strip()
        target_profile_id = str(payload.get("sheet_profile_id", "")).strip()
        current_profile = self._current_sheet_profile_from_form()
        current_key = self._sheet_profile_key(current_profile)
        target_key = target_profile_id or "|".join(
            (
                spreadsheet_url.lower(),
                sheet_name.lower(),
                str(payload.get("start_row", "")).strip(),
                str(payload.get("max_rows", "")).strip(),
                self._sheet_limit_mode_value(payload.get("limit_mode", "cantidad")),
            )
        )

        if target_key == current_key:
            self.sheet_start_row_var.set(str(next_row_int))
            self.selected_sheet_profile_id = current_key

        updated = False
        for item in self.sheet_profiles:
            if self._sheet_profile_key(item) == target_key:
                item["start_row"] = str(next_row_int)
                updated = True

        if updated or target_key == current_key:
            self._save_sheet_settings()
            self._restore_sheet_profile_selection()
            self._update_sheet_template_display()
            self._push_log(f"Sheets: proxima fila sugerida {next_row_int} para {sheet_name or 'hoja actual'}.")

    def _start_sheets_status_check(self) -> None:
        token_path = get_sheets_token_path()
        if not token_path.exists():
            self.sheets_connected = False
            self.sheets_status_var.set("Google Sheets no conectado")
            return

        def worker() -> None:
            try:
                email = get_connected_google_email(
                    credentials_path=get_gmail_credentials_path(),
                    token_path=token_path,
                    interactive=False,
                )
                if email:
                    self.event_queue.put(("sheets_connected", email))
                else:
                    self.event_queue.put(("sheets_status", "Token Sheets encontrado, pero hay que reconectar."))
            except Exception:
                self.event_queue.put(("sheets_status", "Token Sheets encontrado, pero hay que reconectar."))

        threading.Thread(target=worker, daemon=True).start()

    def _connect_sheets_account(self) -> None:
        self._save_sheet_settings()
        self._push_log(f"Conectando Google Sheets con credenciales: {get_gmail_credentials_path()}")
        email = get_connected_google_email(
            credentials_path=get_gmail_credentials_path(),
            token_path=get_sheets_token_path(),
            interactive=True,
        )
        sheet_url = normalize_spreadsheet_url((self.sheet_url_var.get() or "").strip())
        if sheet_url and not is_office_file_url(self.sheet_url_var.get() or ""):
            try:
                self.event_queue.put(("sheet_tabs_loaded", self._sheet_tabs(sheet_url)))
            except Exception:
                pass
        self.event_queue.put(("sheets_connected", email or "cuenta Google"))
        self._push_log(f"Google Sheets conectado: {email or 'cuenta Google'}")

    def _sheet_spreadsheet_url(self) -> str:
        sheet_url = normalize_spreadsheet_url((self.sheet_url_var.get() or "").strip())
        if not sheet_url:
            raise RuntimeError("Ingresa la URL de Google Sheets.")
        if is_office_file_url(self.sheet_url_var.get() or ""):
            raise RuntimeError(OFFICE_FILE_MESSAGE)
        self.sheet_url_var.set(sheet_url)
        return sheet_url

    def _sheet_tabs(self, sheet_url: str) -> list[str]:
        tabs = list_spreadsheet_sheet_names(
            spreadsheet_url_or_id=sheet_url,
            credentials_path=get_gmail_credentials_path(),
            token_path=get_sheets_token_path(),
            interactive=False,
        )
        if not tabs:
            raise RuntimeError("No se encontraron pestanas en el Google Sheet.")
        return tabs

    def _load_sheet_tabs(self) -> None:
        sheet_url = self._sheet_spreadsheet_url()
        tabs = self._sheet_tabs(sheet_url)
        self.event_queue.put(("sheet_tabs_loaded", tabs))
        self.event_queue.put(("sheets_status", f"Pestanas cargadas: {len(tabs)}"))

    def _sheet_fallback_diagnostico(self) -> str:
        return self._normalize_code(self.diagnosis_preset_var.get())

    def _sheet_default_practice(self) -> str:
        return self._resolve_practice_code(self.practice_var.get())

    def _sheet_limit_mode_value(self, raw_value: str | None = None) -> str:
        value = raw_value
        if value is None and hasattr(self, "sheet_limit_mode_var"):
            value = self.sheet_limit_mode_var.get()
        normalized = str(value or "cantidad").strip().lower()
        if normalized in {"fila_final", "fila final", "hasta fila", "final"}:
            return "fila_final"
        return "cantidad"

    def _sheet_limit_mode_label(self, raw_value: str | None = None) -> str:
        return "Fila final" if self._sheet_limit_mode_value(raw_value) == "fila_final" else "Cantidad"

    def _sheet_start_row(self) -> int:
        raw_value = (self.sheet_start_row_var.get() or "").strip()
        if not raw_value:
            return 2
        try:
            start_row = int(raw_value)
        except ValueError as exc:
            raise RuntimeError("La fila inicial de Sheets debe ser un numero entero.") from exc
        if start_row < 1:
            raise RuntimeError("La fila inicial de Sheets debe ser 1 o mayor.")
        return start_row

    def _sheet_max_rows(self, start_row: int | None = None) -> int:
        raw_value = (self.sheet_max_rows_var.get() or "").strip()
        if not raw_value:
            return 40
        try:
            value = int(raw_value)
        except ValueError as exc:
            raise RuntimeError("El limite de Sheets debe ser un numero entero.") from exc
        if self._sheet_limit_mode_value() == "fila_final":
            start_row = start_row if start_row is not None else self._sheet_start_row()
            if value < start_row:
                raise RuntimeError("La fila final de Sheets debe ser igual o mayor que la fila inicial.")
            return value - start_row + 1
        if value < 1:
            raise RuntimeError("La cantidad de filas de Sheets debe ser 1 o mayor.")
        return value

    def _build_sheet_records(self) -> list[dict]:
        sheet_url = normalize_spreadsheet_url((self.sheet_url_var.get() or "").strip())
        sheet_name = (self.sheet_name_var.get() or "").strip()
        start_row = self._sheet_start_row()
        max_rows = self._sheet_max_rows(start_row)
        complete_benef_only = bool(self.sheet_complete_benef_var.get())
        complete_dni_only = bool(self.sheet_complete_dni_var.get())
        if complete_benef_only and complete_dni_only:
            raise RuntimeError("Elegir Completar BENEF o Completar DNI, no los dos a la vez.")
        only_complete_data = complete_benef_only or complete_dni_only
        check_credential = bool(self.sheet_check_credential_var.get()) and not only_complete_data
        self._sheet_precheck_result_rows = []
        if not sheet_url:
            raise RuntimeError("Ingresa la URL de Google Sheets.")
        if is_office_file_url(self.sheet_url_var.get() or ""):
            raise RuntimeError(OFFICE_FILE_MESSAGE)
        if not sheet_name:
            raise RuntimeError("Ingresa el nombre de la pestana.")

        self.sheet_url_var.set(sheet_url)
        self._save_sheet_settings()
        sheet_read_kwargs = {
            "spreadsheet_url_or_id": sheet_url,
            "sheet_name": sheet_name,
            "fallback_diagnostico": self._sheet_fallback_diagnostico(),
            "default_practica": self._sheet_default_practice(),
            "start_row": start_row,
            "max_rows": max_rows,
            "complete_benef_only": complete_benef_only,
            "complete_dni_only": complete_dni_only,
        }
        if check_credential:
            sheet_read_kwargs["require_credential_ok"] = True
        records = read_ome_sheet_rows(**sheet_read_kwargs)
        if not records:
            raise RuntimeError(f"No hay filas pendientes en la hoja para procesar desde la fila {start_row}.")

        prepared: list[dict] = []
        skipped_practices: list[str] = []
        for row in records:
            if row.get("credential_blocked"):
                self._sheet_precheck_result_rows.append(
                    {
                        "sheet_row": row.get("sheet_row"),
                        "modo": row.get("modo") or ("BENEF" if row.get("beneficio") else "DNI"),
                        "afiliado": (row.get("afiliado") or "").strip(),
                        "beneficio": (row.get("beneficio") or "").strip(),
                        "dni": (row.get("dni") or "").strip(),
                        "nombre": (row.get("nombre") or "").strip(),
                        "diagnostico": self._normalize_code(row.get("diagnostico") or "") or self._sheet_fallback_diagnostico(),
                        "practica": (row.get("practica_origen") or row.get("practica") or "").strip(),
                        "resultado": "VERIFICAR_CREDENCIAL",
                        "nro_ome": "",
                    }
                )
                continue
            if complete_benef_only and not (row.get("dni") or "").strip():
                continue
            if complete_dni_only and not (row.get("beneficio") or row.get("afiliado") or "").strip():
                continue
            diagnostico = self._normalize_code(row.get("diagnostico") or "")
            raw_practica = (row.get("practica_origen") or row.get("practica") or "").strip()
            if complete_dni_only:
                practicas = [self._sheet_default_practice() or "427122"]
            elif self.config_data.module_key == "ome_especialista" and not complete_benef_only:
                practicas = resolve_plan_salud_practices(raw_practica) if raw_practica else []
                if not practicas and raw_practica:
                    skipped_practices.append(
                        f"fila {row.get('sheet_row')}: {raw_practica} "
                        f"({explain_unresolved_plan_salud_practice(raw_practica)})"
                    )
                    continue
            else:
                practica = self._resolve_practice_code(raw_practica)
                practicas = [practica] if practica else []
            if not diagnostico:
                diagnostico = self._sheet_fallback_diagnostico()
            if not practicas:
                default_practice = self._sheet_default_practice()
                practicas = [default_practice] if default_practice else []
            if not only_complete_data and (not diagnostico or not practicas):
                raise RuntimeError("Hay filas en la hoja sin diagnostico o sin practica resoluble.")

            for practica in practicas:
                modo = row.get("modo") or ("BENEF" if row.get("beneficio") else "DNI")
                afiliado = (row.get("afiliado") or "").strip()
                beneficio = (row.get("beneficio") or "").strip()
                dni = (row.get("dni") or "").strip()
                if complete_dni_only:
                    modo = "BENEF"
                    afiliado = beneficio or afiliado
                    beneficio = beneficio or afiliado
                    dni = ""
                prepared.append(
                    {
                        "sheet_row": row.get("sheet_row"),
                        "modo": modo,
                        "afiliado": afiliado,
                        "beneficio": beneficio,
                        "dni": dni,
                        "nombre": (row.get("nombre") or "").strip(),
                        "diagnostico": diagnostico,
                        "practica": practica,
                        "practica_origen": raw_practica,
                        "completar_benef": "1" if complete_benef_only else "",
                        "completar_dni": "1" if complete_dni_only else "",
                    }
                )
        if skipped_practices:
            sample = "; ".join(skipped_practices[:5])
            extra = "" if len(skipped_practices) <= 5 else f" y {len(skipped_practices) - 5} mas"
            self._push_log(f"Sheets: se omitieron practicas sin codigo seguro: {sample}{extra}.")
        if complete_benef_only and not prepared:
            raise RuntimeError(f"No hay filas con DNI para completar/verificar BENEF desde la fila {start_row}.")
        if complete_dni_only and not prepared:
            raise RuntimeError(f"No hay filas con BENEF y DNI faltante desde la fila {start_row}.")
        if not complete_benef_only and not prepared and not self._sheet_precheck_result_rows:
            raise RuntimeError(f"No hay filas con practica resoluble para procesar desde la fila {start_row}.")
        return prepared

    def _run_sheet_batch_with_profile(self) -> None:
        records = self._build_sheet_records()
        precheck_rows = list(getattr(self, "_sheet_precheck_result_rows", []))
        profile = self._current_profile_from_form()
        if records and (not profile["usuario"] or not profile["clave"]):
            raise RuntimeError("Guarda o completa un perfil con usuario y clave antes de ejecutar el lote.")

        temp_csv_path = self._reserve_temp_output_path(get_output_dir() / f"{self.module_key}_sheets_resultados.csv")
        self.last_results_csv_path = None
        self.event_queue.put(("batch_started", {"output_path": str(temp_csv_path)}))

        if not records:
            result_rows = []
            self._push_log(f"Sheets: {len(precheck_rows)} fila(s) marcadas como Verificar credencial sin abrir PAMI.")
        elif not bool(self.headless_var.get()) and self.controller is not None and self.controller.sesion_activa():
            self._push_log("Sesion OME ya abierta: se reutiliza el navegador actual para el lote desde Sheets.")
            result_rows = self._get_controller().ejecutar_lote_en_pagina_actual(
                records,
                progress_callback=self._push_batch_progress,
            )
        else:
            input_path = get_output_dir() / f"{self.module_key}_sheets_input.csv"
            with input_path.open("w", encoding="utf-8", newline="") as handle:
                writer = csv.DictWriter(
                    handle,
                    fieldnames=[
                        "sheet_row",
                        "modo",
                        "afiliado",
                        "beneficio",
                        "dni",
                        "nombre",
                        "diagnostico",
                        "practica",
                        "completar_benef",
                        "completar_dni",
                    ],
                )
                writer.writeheader()
                for record in records:
                    writer.writerow(
                        {
                            "sheet_row": record.get("sheet_row", ""),
                            "modo": record.get("modo", ""),
                            "afiliado": record.get("afiliado", ""),
                            "beneficio": record.get("beneficio", ""),
                            "dni": record.get("dni", ""),
                            "nombre": record.get("nombre", ""),
                            "diagnostico": record.get("diagnostico", ""),
                            "practica": record.get("practica", ""),
                            "completar_benef": record.get("completar_benef", ""),
                            "completar_dni": record.get("completar_dni", ""),
                        }
                    )
            result = run_batch_sync(
                input_path=input_path,
                output_path=temp_csv_path,
                user=profile["usuario"],
                password=profile["clave"],
                headless=bool(self.headless_var.get()),
                log_callback=self._push_log,
                progress_callback=self._push_batch_progress,
                stop_requested=lambda: self.stop_requested,
            )
            result_rows = result.get("rows") or self._read_results_csv(temp_csv_path)

        enriched_rows: list[dict] = list(precheck_rows)
        for source_row, result_row in zip(records, result_rows):
            merged = dict(result_row)
            merged["sheet_row"] = source_row.get("sheet_row")
            for field in ("modo", "afiliado", "beneficio", "dni", "nombre", "diagnostico", "practica"):
                if not str(merged.get(field, "") or "").strip():
                    merged[field] = source_row.get(field, "")
            enriched_rows.append(merged)

        self._write_results_csv(temp_csv_path, enriched_rows)
        self.last_results_csv_path = temp_csv_path
        updated_count = 0
        sheet_write_error = ""
        complete_benef_only_run = bool(records) and all(
            str(record.get("completar_benef", "") or "").strip() for record in records
        )
        complete_dni_only_run = bool(records) and all(
            str(record.get("completar_dni", "") or "").strip() for record in records
        )
        sheet_result_rows = enriched_rows
        if complete_benef_only_run:
            sheet_result_rows = [row for row in enriched_rows if row.get("resultado") == "BENEF_COMPLETADO"]
        elif complete_dni_only_run:
            sheet_result_rows = [row for row in enriched_rows if row.get("resultado") == "DNI_COMPLETADO"]
        try:
            updated_count = write_ome_sheet_results(
                spreadsheet_url_or_id=normalize_spreadsheet_url((self.sheet_url_var.get() or "").strip()),
                sheet_name=(self.sheet_name_var.get() or "").strip(),
                result_rows=sheet_result_rows,
            )
        except Exception as exc:
            sheet_write_error = str(exc)
            self._push_log(
                "El lote termino, pero Google Sheets rechazo la escritura de resultados. "
                "Los numeros generados quedan en el CSV temporal para carga manual. "
                f"Detalle: {sheet_write_error}"
            )
        processed_sheet_rows = [
            int(row.get("sheet_row"))
            for row in enriched_rows
            if str(row.get("sheet_row", "")).strip().isdigit()
        ]
        if processed_sheet_rows and not sheet_write_error and not (complete_benef_only_run or complete_dni_only_run):
            self.event_queue.put(
                (
                    "sheet_advance",
                    {
                        "sheet_profile_id": str(self.selected_sheet_profile_id or "").strip(),
                        "spreadsheet_url": normalize_spreadsheet_url((self.sheet_url_var.get() or "").strip()),
                        "sheet_name": (self.sheet_name_var.get() or "").strip(),
                        "start_row": (self.sheet_start_row_var.get() or "").strip() or "2",
                        "max_rows": (self.sheet_max_rows_var.get() or "").strip() or "40",
                        "limit_mode": self._sheet_limit_mode_value(),
                        "next_row": max(processed_sheet_rows) + 1,
                    },
                )
            )
        elif processed_sheet_rows and not sheet_write_error and complete_benef_only_run:
            self._push_log("Sheets: BENEF completado; se mantiene la fila inicial para generar OMEs despues.")
        elif processed_sheet_rows and not sheet_write_error and complete_dni_only_run:
            self._push_log("Sheets: DNI completado; se mantiene la fila inicial para generar OMEs despues.")
        summary = self._summarize_result_rows(enriched_rows)
        summary["output_path"] = str(temp_csv_path)
        summary["rows"] = enriched_rows
        summary["sheet_updates"] = updated_count
        summary["sheet_error"] = sheet_write_error
        self.event_queue.put(("batch_summary", summary))

    def _load_saved_profiles(self) -> list[dict]:
        try:
            if not self.profiles_file.exists():
                return []
            data = json.loads(self.profiles_file.read_text(encoding="utf-8"))
            profiles = data.get("usuarios", [])
            normalizados = []
            for item in profiles:
                if not isinstance(item, dict):
                    continue
                usuario = str(item.get("usuario", "")).strip()
                nombre = str(item.get("nombre", item.get("cliente", ""))).strip()
                clave = str(item.get("clave", "")).strip()
                if usuario:
                    normalizados.append(
                        {
                            "usuario": usuario,
                            "nombre": nombre,
                            "clave": clave,
                        }
                    )
            return sync_profile_records(normalizados)
        except Exception:
            return []

    def _save_saved_profiles(self) -> None:
        payload = {"usuarios": self.saved_profiles[:20]}
        self.profiles_file.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
        upsert_shared_credentials_from_records(self.saved_profiles[:20])

    def _upsert_profile(self, profile: dict) -> None:
        usuario = profile["usuario"].strip()
        if not usuario:
            return
        self.saved_profiles = [p for p in self.saved_profiles if p["usuario"].lower() != usuario.lower()]
        self.saved_profiles.insert(
            0,
            {
                "usuario": usuario,
                "nombre": profile["nombre"].strip(),
                "clave": profile["clave"],
            },
        )
        self.saved_profiles = self.saved_profiles[:20]
        self._save_saved_profiles()
        self.profile_combo.configure(values=self._profile_options() or [""])
        self.profile_var.set(self._format_profile_entry(self.saved_profiles[0]))

    def _save_current_profile(self) -> None:
        profile = self._current_profile_from_form()
        if not profile["usuario"]:
            messagebox.showwarning("Perfil", "Ingresa un usuario para guardar el perfil.")
            return
        self._upsert_profile(profile)
        self._push_log(f"Perfil guardado: {profile['usuario']} ({profile['nombre'] or 'sin nombre'})")

    def _delete_current_profile(self) -> None:
        profile = self._current_profile_from_form()
        usuario = profile["usuario"].strip()
        if not usuario:
            messagebox.showwarning("Perfil", "No hay un perfil seleccionado para borrar.")
            return
        if not messagebox.askyesno("Borrar perfil", f"Quieres borrar el perfil {usuario}?"):
            return

        remaining = [p for p in self.saved_profiles if p["usuario"].lower() != usuario.lower()]
        if len(remaining) == len(self.saved_profiles):
            messagebox.showwarning("Perfil", "Ese perfil no estaba guardado.")
            return

        self.saved_profiles = remaining
        self._save_saved_profiles()
        opciones = self._profile_options() or [""]
        self.profile_combo.configure(values=opciones)
        if self.saved_profiles:
            primero = self._format_profile_entry(self.saved_profiles[0])
            self.profile_var.set(primero)
            self._on_profile_selected(primero)
        else:
            self._new_profile()
        self._push_log(f"Perfil borrado: {usuario}")

    def _profile_options(self) -> list[str]:
        return [self._format_profile_entry(item) for item in self.saved_profiles]

    def _format_profile_entry(self, item: dict) -> str:
        usuario = item.get("usuario", "").strip()
        nombre = item.get("nombre", "").strip()
        if usuario and nombre:
            return f"{usuario} - {nombre}"
        return usuario

    def _extract_user_code(self, value: str) -> str:
        return value.split(" - ", 1)[0].strip()

    def _on_profile_selected(self, selected: str) -> None:
        profile = self._find_profile_by_display(selected)
        if not profile:
            return
        self._syncing_pami_profile = True
        self.profile_name_var.set(profile.get("nombre", ""))
        self.profile_user_var.set(profile.get("usuario", ""))
        self.profile_password_var.set(profile.get("clave", ""))
        self._syncing_pami_profile = False
        if self.config_data.link_sheet_profiles_to_pami and not self._syncing_sheet_profile:
            self._select_sheet_profile_for_pami(profile.get("usuario", ""))
        self._save_sheet_settings()

    def _select_profile_by_user(self, usuario: str) -> bool:
        usuario = (usuario or "").strip()
        if not usuario:
            return False
        for profile in self.saved_profiles:
            if profile.get("usuario", "").strip().lower() == usuario.lower():
                display = self._format_profile_entry(profile)
                self.profile_var.set(display)
                self._on_profile_selected(display)
                return True
        return False

    def _selected_sheet_profile(self) -> dict | None:
        selected = (self.sheet_profile_var.get() or "").strip()
        profile = self.sheet_profile_lookup.get(selected)
        if profile:
            return profile
        current_key = str(self.selected_sheet_profile_id or "").strip()
        if not current_key:
            return None
        for item in self.sheet_profiles:
            if self._sheet_profile_key(item) == current_key:
                return item
        return None

    def _select_pami_profile_for_sheet(self, sheet_profile: dict | None) -> None:
        if not self.config_data.link_sheet_profiles_to_pami:
            return
        if not sheet_profile or not hasattr(self, "profile_user_var"):
            return
        usuario = str(sheet_profile.get("pami_usuario", "")).strip()
        if not usuario:
            return
        if self._select_profile_by_user(usuario):
            self._push_log(
                f"Perfil PAMI sugerido por hoja: {usuario} ({sheet_profile.get('pami_nombre', '') or 'sin nombre'})"
            )

    def _select_sheet_profile_for_pami(self, usuario: str) -> bool:
        if not self.config_data.link_sheet_profiles_to_pami:
            return False
        usuario = (usuario or "").strip()
        if not usuario or not hasattr(self, "sheet_url_var"):
            return False
        current = self._selected_sheet_profile()
        if current and str(current.get("pami_usuario", "")).strip().lower() == usuario.lower():
            return True
        for profile in self.sheet_profiles:
            if str(profile.get("pami_usuario", "")).strip().lower() == usuario.lower():
                self._apply_sheet_profile(profile, select_pami=False, save=True)
                self._push_log(
                    f"Hoja predeterminada sugerida por usuario: {profile.get('sheet_name', '')} | {usuario}"
                )
                return True
        return False

    def _find_profile_by_display(self, selected: str) -> dict | None:
        usuario = self._extract_user_code(selected)
        for profile in self.saved_profiles:
            if profile["usuario"].lower() == usuario.lower():
                return profile
        return None

    def _current_profile_from_form(self) -> dict:
        return {
            "nombre": self.profile_name_var.get().strip(),
            "usuario": self.profile_user_var.get().strip(),
            "clave": self.profile_password_var.get(),
        }

    def _new_profile(self) -> None:
        self.profile_var.set("")
        self.profile_name_var.set("")
        self.profile_user_var.set("")
        self.profile_password_var.set("")
        self._set_password_visibility(False)

    def _load_initial_profile_into_form(self) -> None:
        if self.config_data.link_sheet_profiles_to_pami and self._selected_sheet_profile() and self._select_profile_by_user(
            str(self._selected_sheet_profile().get("pami_usuario", ""))
        ):
            return
        selected_usuario = str(self.sheet_settings.get("selected_pami_usuario", "")).strip()
        if selected_usuario and self._select_profile_by_user(selected_usuario):
            return
        if self.saved_profiles:
            self._on_profile_selected(self._format_profile_entry(self.saved_profiles[0]))
        else:
            self._new_profile()

    def _toggle_password_visibility(self) -> None:
        self._set_password_visibility(not self.password_visible)

    def _set_password_visibility(self, visible: bool) -> None:
        self.password_visible = visible
        self.password_entry.configure(show="" if visible else "*")
        if getattr(self, "_icon_eye", None) is not None:
            self.toggle_password_button.configure(image=self._icon_eye_off if visible else self._icon_eye, text="")
        else:
            self.toggle_password_button.configure(text="Ocultar" if visible else "Ver")


class PamiOmeMedCabeceraModuleFrame(PamiOmeModuleFrame):
    def __init__(self, master, on_back=None) -> None:
        super().__init__(master, on_back, config=MED_CABECERA_CONFIG)


class PamiOmeEspecialistaModuleFrame(PamiOmeModuleFrame):
    def __init__(self, master, on_back=None) -> None:
        super().__init__(master, on_back, config=ESPECIALISTA_CONFIG)


class PamiOmeApp(ctk.CTk):
    def __init__(self) -> None:
        super().__init__()

        ctk.set_appearance_mode("light")
        ctk.set_default_color_theme("blue")

        self.title(MED_CABECERA_CONFIG.module_title)
        self.minsize(1120, 760)
        self._set_initial_window_size()
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(0, weight=1)

        self.module_frame = PamiOmeMedCabeceraModuleFrame(self)
        self.module_frame.grid(row=0, column=0, sticky="nsew")
        self.protocol("WM_DELETE_WINDOW", self._on_close)

    def _on_close(self) -> None:
        self.module_frame.on_close()
        self.destroy()

    def _set_initial_window_size(self) -> None:
        screen_w = max(self.winfo_screenwidth(), 1280)
        screen_h = max(self.winfo_screenheight(), 800)
        width = min(max(int(screen_w * 0.94), 1120), max(screen_w - 40, 1120))
        height = min(max(int(screen_h * 0.9), 760), max(screen_h - 80, 760))
        pos_x = max((screen_w - width) // 2, 0)
        pos_y = max((screen_h - height) // 2, 0)
        self.geometry(f"{width}x{height}+{pos_x}+{pos_y}")


if __name__ == "__main__":
    app = PamiOmeApp()
    app.mainloop()
