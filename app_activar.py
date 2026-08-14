from __future__ import annotations

import json
import os
import queue
import re
import threading
import unicodedata
from calendar import Calendar, month_name
from datetime import datetime, timedelta
from pathlib import Path
import tkinter as tk
from tkinter import filedialog, messagebox

import customtkinter as ctk

from ui_kit import apply_button_icon, attach_tooltip, button_icon
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl import load_workbook

from app_credentials import sync_profile_payload, upsert_shared_credentials_from_payload
from app_ome import DEFAULT_NOMENCLADOR_PATH, PracticeCatalogDialog, PracticeCatalogItem, _default_nomenclador_candidates
from app_logging import log_message
from app_paths import get_data_dir, get_log_file, get_output_dir
from google_sheets_ome import (
    OFFICE_FILE_MESSAGE,
    build_sheets_service,
    extract_spreadsheet_id,
    get_connected_google_email,
    get_sheets_token_path,
    is_office_file_url,
    normalize_spreadsheet_url,
)
from pami_plan_salud_resolver import resolve_plan_salud_practice, resolve_plan_salud_practices
from pami_activar import PamiActivarController, ResultadoLote


BOCAS_CONOCIDAS = {
    "Cualquiera disponible": "",
}

MODALIDADES = {
    "Presencial": "P",
    "Telemedicina": "T",
}

PRACTICAS_ACTIVAR = {
    "427109 - CONSULTA CLINICA PRESENCIAL": "427109",
    "427120 - CONSULTA AFILIADO SIN ASIGNACION": "427120",
    "427121 - CONSULTA AFILIADO EN TRANSITO": "427121",
    "427122 - CONSULTA AFILIADO EXTRACAPITA": "427122",
    "Cualquiera disponible": "",
}

DEFAULT_PRACTICA_ACTIVAR = "427122 - CONSULTA AFILIADO EXTRACAPITA"


class DatePickerDialog(ctk.CTkToplevel):
    def __init__(self, master, initial_value: str = "") -> None:
        super().__init__(master)
        self.title("Seleccionar fecha")
        self.geometry("320x360")
        self.resizable(False, False)
        self.transient(master.winfo_toplevel())
        self.grab_set()

        self.result = None
        self._calendar = Calendar(firstweekday=0)
        self._selected_date = self._parse_initial(initial_value)
        self._current_year = self._selected_date.year
        self._current_month = self._selected_date.month

        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=1)

        header = ctk.CTkFrame(self, fg_color="transparent")
        header.grid(row=0, column=0, padx=14, pady=(14, 8), sticky="ew")
        header.grid_columnconfigure(1, weight=1)

        ctk.CTkButton(header, text="<", width=34, command=self._prev_month).grid(row=0, column=0, padx=(0, 8), sticky="w")
        self.month_label = ctk.CTkLabel(header, text="", font=ctk.CTkFont(size=16, weight="bold"), text_color="#16324f")
        self.month_label.grid(row=0, column=1, sticky="ew")
        ctk.CTkButton(header, text=">", width=34, command=self._next_month).grid(row=0, column=2, padx=(8, 0), sticky="e")

        self.days_frame = ctk.CTkFrame(self, corner_radius=12, fg_color="#eef3f8")
        self.days_frame.grid(row=1, column=0, padx=14, pady=(0, 10), sticky="nsew")

        footer = ctk.CTkFrame(self, fg_color="transparent")
        footer.grid(row=2, column=0, padx=14, pady=(0, 14), sticky="ew")
        footer.grid_columnconfigure(0, weight=1)

        ctk.CTkButton(footer, text="Limpiar", width=80, fg_color="#9aafc3", hover_color="#7f95aa", command=self._clear).grid(row=0, column=0, padx=(0, 8), sticky="w")
        ctk.CTkButton(footer, text="Hoy", width=80, fg_color="#66788a", hover_color="#536577", command=self._today).grid(row=0, column=1, padx=8, sticky="w")
        ctk.CTkButton(footer, text="Cancelar", width=90, fg_color="#9aafc3", hover_color="#7f95aa", command=self._cancel).grid(row=0, column=2, padx=8, sticky="e")

        self._render_days()
        self.protocol("WM_DELETE_WINDOW", self._cancel)

    def _parse_initial(self, value: str) -> datetime:
        try:
            return datetime.strptime(value.strip(), "%d/%m/%Y")
        except Exception:
            return datetime.now()

    def _render_days(self) -> None:
        for child in self.days_frame.winfo_children():
            child.destroy()

        self.month_label.configure(text=f"{month_name[self._current_month]} {self._current_year}")

        week_names = ["Lu", "Ma", "Mi", "Ju", "Vi", "Sa", "Do"]
        for idx, name in enumerate(week_names):
            ctk.CTkLabel(
                self.days_frame,
                text=name,
                font=ctk.CTkFont(size=12, weight="bold"),
                text_color="#51657a",
                width=36,
            ).grid(row=0, column=idx, padx=4, pady=(10, 4))

        month_weeks = self._calendar.monthdayscalendar(self._current_year, self._current_month)
        for row_index, week in enumerate(month_weeks, start=1):
            for col_index, day in enumerate(week):
                if day == 0:
                    ctk.CTkLabel(self.days_frame, text="", width=36).grid(row=row_index, column=col_index, padx=4, pady=4)
                    continue

                is_selected = (
                    self._selected_date.day == day
                    and self._selected_date.month == self._current_month
                    and self._selected_date.year == self._current_year
                )
                ctk.CTkButton(
                    self.days_frame,
                    text=str(day),
                    width=36,
                    height=32,
                    fg_color="#245b9d" if is_selected else "#ffffff",
                    hover_color="#1d4b82" if is_selected else "#dbe7f2",
                    text_color="#ffffff" if is_selected else "#16324f",
                    command=lambda d=day: self._select_day(d),
                ).grid(row=row_index, column=col_index, padx=4, pady=4)

    def _select_day(self, day: int) -> None:
        self._selected_date = datetime(self._current_year, self._current_month, day)
        self.result = self._selected_date.strftime("%d/%m/%Y")
        self.destroy()

    def _prev_month(self) -> None:
        if self._current_month == 1:
            self._current_month = 12
            self._current_year -= 1
        else:
            self._current_month -= 1
        self._render_days()

    def _next_month(self) -> None:
        if self._current_month == 12:
            self._current_month = 1
            self._current_year += 1
        else:
            self._current_month += 1
        self._render_days()

    def _today(self) -> None:
        today = datetime.now()
        self._current_year = today.year
        self._current_month = today.month
        self._selected_date = today
        self.result = today.strftime("%d/%m/%Y")
        self.destroy()

    def _clear(self) -> None:
        self.result = ""
        self.destroy()

    def _cancel(self) -> None:
        self.result = None
        self.destroy()


class PamiActivarModuleFrame(ctk.CTkFrame):
    """
    Modulo de activacion (y modificacion) de turnos de OMEs
    en el Panel de Aceptacion de PAMI.
    """

    def __init__(self, master, on_back=None) -> None:
        super().__init__(master, fg_color="transparent")
        self.on_back = on_back

        self.event_queue: queue.Queue = queue.Queue()
        self.controller_queue: queue.Queue = queue.Queue()
        self.action_running = False
        self.password_visible = False
        self.data_dir = get_data_dir()
        self.profiles_file = Path(self.data_dir) / "usuarios_activar.json"
        self.saved_profiles = self._load_saved_profiles()
        self.practice_catalog_file = Path(self.data_dir) / "activar_nomenclador_practicas.json"
        self.sheet_settings_file = Path(self.data_dir) / "activar_sheets_config.json"
        self.sheet_settings = self._load_sheet_settings()
        self._syncing_sheet_profile = False
        self._syncing_pami_profile = False
        self.sheet_profiles = self._extract_sheet_profiles(self.sheet_settings)
        self.selected_sheet_profile_id = str(self.sheet_settings.get("selected_profile_id", "")).strip()
        self.sheet_profile_lookup: dict[str, dict] = {}
        initial_sheet_name = str(self.sheet_settings.get("sheet_name", "Mc Dube")).strip()
        self.sheet_name_options = [initial_sheet_name] if initial_sheet_name else [""]
        self.practice_catalog_items: list[PracticeCatalogItem] = []
        self.practice_catalog_path: Path | None = None
        self.active_catalog_modules: set[str] = set()
        self.active_catalog_practices: set[str] = set()
        self.practice_options = list(PRACTICAS_ACTIVAR.keys())
        self.filtered_practice_options = list(self.practice_options)
        self.sheets_connected = False
        self.browser_visible_var = ctk.BooleanVar(value=bool(self.sheet_settings.get("browser_visible", True)))

        self.controller = PamiActivarController(
            log_callback=self._push_log,
            status_callback=self._push_status,
        )
        self.last_report_result: ResultadoLote | None = None
        self.last_report_export_path: str | None = None
        self.controller_thread: threading.Thread | None = None

        self._build_ui()
        self._initialize_practice_catalog()
        self.after(400, self._start_sheets_status_check)
        self.after(150, self._process_ui_queue)

    def _build_ui(self) -> None:
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=1)

        top_bar = ctk.CTkFrame(self, corner_radius=12, fg_color="#ffffff", border_width=1, border_color="#d8e2ec")
        top_bar.grid(row=0, column=0, padx=8, pady=(6, 4), sticky="ew")
        top_bar.grid_columnconfigure(1, weight=1)
        top_bar.grid_columnconfigure(2, weight=0)

        if self.on_back:
            ctk.CTkButton(
                top_bar,
                text="Volver",
                width=78,
                height=30,
                command=self._go_home,
                fg_color="#9aafc3",
                hover_color="#7f95aa",
            ).grid(row=0, column=0, rowspan=2, padx=(8, 10), pady=6, sticky="w")

        ctk.CTkLabel(
            top_bar,
            text="Activar OME PAMI",
            font=ctk.CTkFont(size=22, weight="bold"),
            text_color="#16324f",
        ).grid(row=0, column=1, padx=10, pady=(6, 1), sticky="w")

        ctk.CTkLabel(
            top_bar,
            text="Insertar DNI o BENEF, completar OME o DNI y ejecutar el BOT con reporte exportable.",
            font=ctk.CTkFont(size=12),
            text_color="#51657a",
        ).grid(row=1, column=1, padx=10, pady=(0, 6), sticky="w")

        self.restart_app_button = ctk.CTkButton(
            top_bar,
            text="Reiniciar app",
            width=132,
            height=30,
            command=self._restart_application,
            fg_color="#66788a",
            hover_color="#536577",
        )
        self.restart_app_button.grid(row=0, column=2, rowspan=2, padx=(8, 10), pady=6, sticky="e")

        content = ctk.CTkScrollableFrame(self, corner_radius=8, fg_color="#ffffff")
        content.grid(row=1, column=0, padx=8, pady=(0, 8), sticky="nsew")
        content.grid_columnconfigure(0, weight=1)

        pasos = ctk.CTkFrame(content, corner_radius=12, fg_color="#ffffff", border_width=1, border_color="#d8e2ec")
        pasos.grid(row=0, column=0, padx=8, pady=(0, 4), sticky="ew")
        pasos.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(
            pasos,
            text="Flujo rápido",
            font=ctk.CTkFont(size=16, weight="bold"),
            text_color="#16324f",
        ).grid(row=0, column=0, padx=12, pady=(8, 4), sticky="w")

        ctk.CTkLabel(
            pasos,
            text="Insertar DNI/BENEF + OME o práctica --> Ejecutar BOT --> Exportar resultado",
            font=ctk.CTkFont(size=12),
            text_color="#43576b",
        ).grid(row=1, column=0, padx=12, pady=(0, 8), sticky="w")

        profiles_frame = ctk.CTkFrame(content, corner_radius=12, fg_color="#ffffff", border_width=1, border_color="#d8e2ec")
        profiles_frame.grid(row=1, column=0, padx=8, pady=(0, 4), sticky="ew")
        profiles_frame.grid_columnconfigure(0, weight=1)

        options = self._profile_options()
        self.profile_var = ctk.StringVar(value=options[0] if options else "")
        self.profile_name_var = ctk.StringVar()
        self.profile_user_var = ctk.StringVar()
        self.profile_password_var = ctk.StringVar()

        profile_form = ctk.CTkFrame(profiles_frame, fg_color="transparent")
        profile_form.grid(row=0, column=0, padx=(10, 8), pady=8, sticky="w")
        profile_form.grid_columnconfigure(7, weight=1)

        ctk.CTkLabel(
            profile_form,
            text="Perfil guardado",
            font=ctk.CTkFont(size=13, weight="bold"),
            text_color="#16324f",
        ).grid(row=0, column=0, padx=(0, 8), pady=(0, 8), sticky="w")

        self.profile_combo = ctk.CTkComboBox(
            profile_form,
            values=options or [""],
            variable=self.profile_var,
            width=340,
            command=self._on_profile_selected,
        )
        self.profile_combo.grid(row=0, column=1, padx=(0, 12), pady=(0, 8), sticky="w")

        self.new_profile_button = ctk.CTkButton(
            profile_form,
            text="+",
            command=self._new_profile,
            width=40,
            height=28,
            font=ctk.CTkFont(size=16, weight="bold"),
            fg_color="#245b9d",
            hover_color="#1d4b82",
        )
        self.new_profile_button.grid(row=0, column=2, padx=(0, 6), pady=(0, 8), sticky="w")
        attach_tooltip(self.new_profile_button, "Nuevo perfil")

        self.delete_profile_button = ctk.CTkButton(
            profile_form,
            text="Borrar perfil",
            command=self._delete_current_profile,
            width=40,
            height=28,
            fg_color="#8a5a5a",
            hover_color="#734949",
        )
        self.delete_profile_button.grid(row=0, column=3, padx=(0, 8), pady=(0, 8), sticky="w")
        apply_button_icon(self.delete_profile_button, "trash.png", "Borrar perfil", size=(16, 16), width=40)

        self.save_profile_button = ctk.CTkButton(
            profile_form,
            text="Guardar perfil",
            command=self._save_current_profile,
            width=40,
            height=28,
            fg_color="#66788a",
            hover_color="#536577",
        )
        self.save_profile_button.grid(row=0, column=4, padx=(0, 16), pady=(0, 8), sticky="w")
        apply_button_icon(self.save_profile_button, "save.png", "Guardar perfil", size=(18, 18), width=40)

        ctk.CTkLabel(
            profile_form,
            text="Médico / cliente",
            font=ctk.CTkFont(size=13, weight="bold"),
            text_color="#16324f",
        ).grid(row=0, column=5, padx=(0, 8), pady=(0, 8), sticky="w")

        self.client_entry = ctk.CTkEntry(
            profile_form,
            textvariable=self.profile_name_var,
            placeholder_text="Nombre del médico o centro",
            width=300,
        )
        self.client_entry.grid(row=0, column=6, padx=(0, 0), pady=(0, 8), sticky="w")

        credentials_row = ctk.CTkFrame(profile_form, fg_color="transparent")
        credentials_row.grid(row=1, column=0, columnspan=8, sticky="w")

        ctk.CTkLabel(
            credentials_row,
            text="Usuario",
            font=ctk.CTkFont(size=13, weight="bold"),
            text_color="#16324f",
        ).grid(row=0, column=0, padx=(0, 8), sticky="w")

        self.user_entry = ctk.CTkEntry(
            credentials_row,
            textvariable=self.profile_user_var,
            placeholder_text="Usuario PAMI",
            width=170,
        )
        self.user_entry.grid(row=0, column=1, padx=(0, 14), sticky="w")

        ctk.CTkLabel(
            credentials_row,
            text="Clave",
            font=ctk.CTkFont(size=13, weight="bold"),
            text_color="#16324f",
        ).grid(row=0, column=2, padx=(0, 8), sticky="w")

        self.password_entry = ctk.CTkEntry(
            credentials_row,
            textvariable=self.profile_password_var,
            placeholder_text="Clave PAMI",
            show="*",
            width=170,
        )
        self.password_entry.grid(row=0, column=3, padx=(0, 10), sticky="w")

        self.toggle_password_button = ctk.CTkButton(
            credentials_row,
            text="Ver",
            command=self._toggle_password,
            width=42,
            height=30,
            fg_color="#66788a",
            hover_color="#536577",
        )
        self.toggle_password_button.grid(row=0, column=4, padx=(0, 12), sticky="w")
        apply_button_icon(self.toggle_password_button, "eye.png", "Ver / ocultar clave", size=(18, 18), width=42)

        self.browser_visible_checkbox = ctk.CTkCheckBox(
            credentials_row,
            text="Ver navegador",
            variable=self.browser_visible_var,
            text_color="#16324f",
            command=self._on_browser_visibility_changed,
        )
        self.browser_visible_checkbox.grid(row=0, column=5, padx=(0, 0), sticky="w")

        self.sheet_url_var = ctk.StringVar(value=normalize_spreadsheet_url(str(self.sheet_settings.get("spreadsheet_url", ""))))
        self.sheet_url_display_var = ctk.StringVar(value="")
        self.sheet_name_var = ctk.StringVar(value=str(self.sheet_settings.get("sheet_name", "Mc Dube")))
        self.sheet_start_row_var = ctk.StringVar(value=str(self.sheet_settings.get("start_row", 2)))
        self.sheet_max_rows_var = ctk.StringVar(value=str(self.sheet_settings.get("max_rows", 40)))
        self.sheet_profile_var = ctk.StringVar(value="")
        self.sheets_status_var = ctk.StringVar(value="Google Sheets no conectado")
        self.sheet_url_editable = False

        sheets_frame = ctk.CTkFrame(content, corner_radius=12, fg_color="#ffffff", border_width=1, border_color="#d8e2ec")
        sheets_frame.grid(row=2, column=0, padx=8, pady=(0, 4), sticky="ew")

        sheet_header = ctk.CTkFrame(sheets_frame, fg_color="transparent")
        sheet_header.grid(row=0, column=0, padx=10, pady=(8, 4), sticky="w")

        ctk.CTkLabel(
            sheet_header,
            text="Google Sheets",
            font=ctk.CTkFont(size=16, weight="bold"),
            text_color="#16324f",
        ).grid(row=0, column=0, padx=(0, 12), pady=0, sticky="w")

        ctk.CTkLabel(sheet_header, text="Hoja guardada", text_color="#16324f").grid(
            row=0, column=1, padx=(0, 8), pady=0, sticky="w"
        )
        self.sheet_profile_combo = ctk.CTkComboBox(
            sheet_header,
            values=self._sheet_profile_options() or [""],
            variable=self.sheet_profile_var,
            width=360,
            command=self._on_sheet_profile_selected,
            state="readonly",
        )
        self.sheet_profile_combo.grid(row=0, column=2, padx=(0, 8), pady=0, sticky="w")

        sheet_profile_actions = ctk.CTkFrame(sheet_header, fg_color="transparent")
        sheet_profile_actions.grid(row=0, column=3, padx=(0, 0), pady=0, sticky="w")

        self.new_sheet_profile_button = ctk.CTkButton(
            sheet_profile_actions,
            text="+",
            width=40,
            height=28,
            font=ctk.CTkFont(size=16, weight="bold"),
            command=self._new_sheet_profile,
            fg_color="#245b9d",
            hover_color="#1d4b82",
        )
        self.new_sheet_profile_button.grid(row=0, column=0, padx=(0, 6), pady=0, sticky="w")
        attach_tooltip(self.new_sheet_profile_button, "Nueva hoja")

        self.save_sheet_profile_button = ctk.CTkButton(
            sheet_profile_actions,
            text="Guardar hoja",
            width=40,
            height=28,
            command=self._save_current_sheet_profile,
            fg_color="#66788a",
            hover_color="#536577",
        )
        self.save_sheet_profile_button.grid(row=0, column=1, padx=(0, 6), pady=0, sticky="w")
        apply_button_icon(self.save_sheet_profile_button, "save.png", "Guardar hoja", size=(18, 18), width=40)

        self.delete_sheet_profile_button = ctk.CTkButton(
            sheet_profile_actions,
            text="Borrar hoja",
            width=40,
            height=28,
            command=self._delete_current_sheet_profile,
            fg_color="#8a5a5a",
            hover_color="#734949",
        )
        self.delete_sheet_profile_button.grid(row=0, column=2, padx=(0, 0), pady=0, sticky="w")
        apply_button_icon(self.delete_sheet_profile_button, "trash.png", "Borrar hoja", size=(16, 16), width=40)

        sheet_url_row = ctk.CTkFrame(sheets_frame, fg_color="transparent")
        sheet_url_row.grid(row=1, column=0, padx=10, pady=(0, 4), sticky="w")

        ctk.CTkLabel(sheet_url_row, text="Plantilla", text_color="#16324f").grid(row=0, column=0, padx=(0, 8), pady=0, sticky="w")
        self.sheet_url_entry = ctk.CTkEntry(
            sheet_url_row,
            textvariable=self.sheet_url_display_var,
            placeholder_text="https://docs.google.com/spreadsheets/d/...",
            width=720,
        )
        self.sheet_url_entry.grid(row=0, column=1, padx=(0, 8), pady=0, sticky="w")

        self.sheet_url_edit_button = ctk.CTkButton(
            sheet_url_row,
            text="Editar URL",
            width=104,
            height=30,
            command=self._toggle_sheet_url_edit,
            fg_color="#6d7f90",
            hover_color="#536577",
        )
        self.sheet_url_edit_button.grid(row=0, column=2, padx=(0, 0), pady=0, sticky="e")

        inputs_row = ctk.CTkFrame(sheets_frame, fg_color="transparent")
        inputs_row.grid(row=2, column=0, padx=10, pady=(0, 4), sticky="w")

        ctk.CTkLabel(inputs_row, text="Pestaña", text_color="#16324f").grid(
            row=0, column=0, padx=(0, 8), pady=0, sticky="w"
        )
        self.sheet_name_entry = ctk.CTkComboBox(
            inputs_row,
            values=self._sheet_name_values(),
            variable=self.sheet_name_var,
            width=220,
            state="normal",
        )
        self.sheet_name_entry.grid(row=0, column=1, padx=(0, 14), pady=0, sticky="w")

        ctk.CTkLabel(inputs_row, text="Fila inicial", text_color="#16324f").grid(
            row=0, column=2, padx=(0, 8), pady=0, sticky="w"
        )
        self.sheet_start_row_entry = ctk.CTkEntry(inputs_row, textvariable=self.sheet_start_row_var, width=76)
        self.sheet_start_row_entry.grid(row=0, column=3, padx=(0, 12), pady=0, sticky="w")

        ctk.CTkLabel(inputs_row, text="Tope", text_color="#16324f").grid(
            row=0, column=4, padx=(0, 8), pady=0, sticky="w"
        )
        self.sheet_max_rows_entry = ctk.CTkEntry(inputs_row, textvariable=self.sheet_max_rows_var, width=76)
        self.sheet_max_rows_entry.grid(row=0, column=5, padx=(0, 12), pady=0, sticky="w")

        self.sheet_tabs_button = ctk.CTkButton(
            inputs_row,
            text="Leer páginas",
            command=lambda: self._run_action(self._load_sheet_tabs_for_current_url),
            width=108,
            height=28,
            fg_color="#6d7f90",
            hover_color="#536577",
        )
        self.sheet_tabs_button.grid(row=0, column=6, padx=(0, 8), pady=0, sticky="w")

        self.sheets_connect_button = ctk.CTkButton(
            inputs_row,
            text="Conectar Sheets",
            command=lambda: self._run_action(self._connect_sheets_account),
            width=128,
            height=28,
        )
        self.sheets_connect_button.grid(row=0, column=7, padx=(0, 8), pady=0, sticky="w")

        self.sheets_run_button = ctk.CTkButton(
            inputs_row,
            text="Ejecutar desde Sheets",
            command=lambda: self._run_action(self._run_sheet_batch_with_profile),
            width=166,
            height=28,
            fg_color="#1f7a46",
            hover_color="#176238",
        )
        self.sheets_run_button.grid(row=0, column=8, padx=(0, 8), pady=0, sticky="w")

        self.sheet_lookup_button = ctk.CTkButton(
            inputs_row,
            text="Buscar OME / DNI",
            command=lambda: self._run_action(self._run_sheet_ome_lookup_with_profile),
            width=156,
            height=28,
            fg_color="#bd6b2a",
            hover_color="#9d571f",
        )
        self.sheet_lookup_button.grid(row=0, column=9, padx=(0, 0), pady=0, sticky="w")

        sheets_footer = ctk.CTkFrame(sheets_frame, fg_color="transparent")
        sheets_footer.grid(row=3, column=0, padx=10, pady=(0, 8), sticky="w")

        self.sheet_lookup_hint_label = ctk.CTkLabel(
            sheets_footer,
            text="Buscar OME / DNI: completa OME faltante o DNI cuando la OME ya existe.",
            font=ctk.CTkFont(size=12),
            text_color="#7a6338",
        )
        self.sheet_lookup_hint_label.grid(row=0, column=0, padx=(0, 16), pady=0, sticky="w")

        self.sheets_status_label = ctk.CTkLabel(
            sheets_footer,
            textvariable=self.sheets_status_var,
            font=ctk.CTkFont(size=12),
            text_color="#66788a",
        )
        self.sheets_status_label.grid(row=0, column=1, padx=(0, 0), pady=0, sticky="w")
        self._set_sheet_url_edit_mode(False)
        self._restore_sheet_profile_selection()

        agenda_frame = ctk.CTkFrame(content, corner_radius=12, fg_color="#ffffff", border_width=1, border_color="#d8e2ec")
        agenda_frame.grid(row=3, column=0, padx=8, pady=(0, 4), sticky="ew")
        agenda_frame.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(
            agenda_frame,
            text="Agenda del lote",
            font=ctk.CTkFont(size=16, weight="bold"),
            text_color="#16324f",
        ).grid(row=0, column=0, padx=10, pady=(8, 6), sticky="w")

        agenda_row_1 = ctk.CTkFrame(agenda_frame, fg_color="transparent")
        agenda_row_1.grid(row=1, column=0, padx=10, pady=(0, 6), sticky="ew")
        agenda_row_1.grid_columnconfigure(8, weight=1)

        self.fecha_var = ctk.StringVar(value=(datetime.today() + timedelta(days=1)).strftime("%d/%m/%Y"))
        ctk.CTkLabel(
            agenda_row_1,
            text="Fecha",
            font=ctk.CTkFont(size=13, weight="bold"),
            text_color="#16324f",
        ).grid(row=0, column=0, padx=(0, 8), pady=0, sticky="w")

        fecha_row = ctk.CTkFrame(agenda_row_1, fg_color="transparent")
        fecha_row.grid(row=0, column=1, padx=(0, 18), pady=0, sticky="w")

        self.fecha_entry = ctk.CTkEntry(
            fecha_row,
            textvariable=self.fecha_var,
            placeholder_text="Seleccionar fecha",
            width=126,
            state="readonly",
        )
        self.fecha_entry.grid(row=0, column=0, padx=(0, 6), pady=0, sticky="w")
        self.fecha_entry.bind("<Button-1>", lambda _event: self._pick_date(), add="+")

        self.fecha_button = ctk.CTkButton(
            fecha_row,
            text="...",
            width=34,
            height=30,
            command=self._pick_date,
        )
        self.fecha_button.grid(row=0, column=1, padx=0, pady=0, sticky="w")

        ctk.CTkLabel(
            agenda_row_1,
            text="Hora inicio",
            font=ctk.CTkFont(size=13, weight="bold"),
            text_color="#16324f",
        ).grid(row=0, column=2, padx=(0, 8), pady=0, sticky="w")

        hora_row = ctk.CTkFrame(agenda_row_1, fg_color="transparent")
        hora_row.grid(row=0, column=3, padx=(0, 18), pady=0, sticky="w")

        self.hora_var = ctk.StringVar(value="08")
        self.minuto_var = ctk.StringVar(value="00")
        self.modalidad_var = ctk.StringVar(value="Presencial")

        horas = [str(h).zfill(2) for h in range(8, 20)]
        self.hora_combo = ctk.CTkComboBox(hora_row, values=horas, variable=self.hora_var, width=76, state="readonly")
        self.hora_combo.grid(row=0, column=0, padx=(0, 6))

        ctk.CTkLabel(hora_row, text=":", font=ctk.CTkFont(size=18, weight="bold"), text_color="#16324f").grid(
            row=0, column=1, padx=2
        )

        minutos = ["00", "10", "15", "20", "30", "40", "45", "50"]
        self.minuto_combo = ctk.CTkComboBox(hora_row, values=minutos, variable=self.minuto_var, width=76, state="readonly")
        self.minuto_combo.grid(row=0, column=2, padx=(6, 0))

        self.modalidad_combo = ctk.CTkComboBox(
            hora_row,
            values=list(MODALIDADES.keys()),
            variable=self.modalidad_var,
            width=128,
            state="readonly",
        )
        self.modalidad_combo.grid(row=0, column=3, padx=(10, 0))

        ctk.CTkLabel(
            agenda_row_1,
            text="Intervalo (min)",
            font=ctk.CTkFont(size=13, weight="bold"),
            text_color="#16324f",
        ).grid(row=0, column=4, padx=(0, 8), pady=0, sticky="w")

        self.intervalo_var = ctk.StringVar(value="10")
        self.intervalo_combo = ctk.CTkComboBox(
            agenda_row_1,
            values=["5", "10", "15", "20", "30"],
            variable=self.intervalo_var,
            width=102,
            state="readonly",
        )
        self.intervalo_combo.grid(row=0, column=5, padx=(0, 0), pady=0, sticky="w")

        agenda_row_2 = ctk.CTkFrame(agenda_frame, fg_color="transparent")
        agenda_row_2.grid(row=2, column=0, padx=10, pady=(0, 8), sticky="ew")
        agenda_row_2.grid_columnconfigure(5, weight=1)

        ctk.CTkLabel(
            agenda_row_2,
            text="Práctica",
            font=ctk.CTkFont(size=13, weight="bold"),
            text_color="#16324f",
        ).grid(row=0, column=0, padx=(0, 8), pady=0, sticky="w")

        self.practica_var = ctk.StringVar(value=DEFAULT_PRACTICA_ACTIVAR)
        self.practica_combo = ctk.CTkComboBox(
            agenda_row_2,
            values=self.practice_options,
            variable=self.practica_var,
            width=300,
            state="normal",
        )
        self.practica_combo.grid(row=0, column=1, padx=(0, 8), pady=0, sticky="w")
        self.practica_combo.bind("<KeyRelease>", self._on_practice_combo_keyrelease)
        self.practica_combo.bind("<FocusOut>", self._on_practice_combo_focus_out)

        self.practice_catalog_button = ctk.CTkButton(
            agenda_row_2,
            text="Códigos...",
            width=116,
            height=30,
            command=self._open_practice_catalog_manager,
            fg_color="#66788a",
            hover_color="#536577",
        )
        self.practice_catalog_button.grid(row=0, column=2, padx=(0, 14), pady=0, sticky="w")

        ctk.CTkLabel(
            agenda_row_2,
            text="Boca de atención",
            font=ctk.CTkFont(size=13, weight="bold"),
            text_color="#16324f",
        ).grid(row=0, column=3, padx=(0, 8), pady=0, sticky="w")

        self.boca_preset_var = ctk.StringVar(value="Cualquiera disponible")
        self.boca_combo = ctk.CTkComboBox(
            agenda_row_2,
            values=list(BOCAS_CONOCIDAS.keys()),
            variable=self.boca_preset_var,
            width=330,
        )
        self.boca_combo.grid(row=0, column=4, padx=(0, 0), pady=0, sticky="w")

        bulk_frame = ctk.CTkFrame(content, corner_radius=12, fg_color="#ffffff", border_width=1, border_color="#d8e2ec")
        bulk_frame.grid(row=4, column=0, padx=8, pady=(0, 4), sticky="ew")
        bulk_frame.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(
            bulk_frame,
            text="Inserta los pacientes",
            font=ctk.CTkFont(size=16, weight="bold"),
            text_color="#16324f",
        ).grid(row=0, column=0, padx=10, pady=(8, 2), sticky="w")

        ctk.CTkLabel(
            bulk_frame,
            text="Pega BENEF/DNI y OME por fila. Si completas ambas columnas, la busqueda sera mas precisa.",
            font=ctk.CTkFont(size=12),
            text_color="#51657a",
        ).grid(row=1, column=0, padx=10, pady=(0, 2), sticky="w")
        ctk.CTkLabel(
            bulk_frame,
            text="Selecciona una fila y pega hacia abajo. 'GENERADA' o vacio se toma como celda en blanco.",
            font=ctk.CTkFont(size=12),
            text_color="#51657a",
        ).grid(row=2, column=0, padx=10, pady=(0, 6), sticky="w")

        sheet_actions = ctk.CTkFrame(bulk_frame, fg_color="transparent")
        sheet_actions.grid(row=3, column=0, padx=10, pady=(0, 6), sticky="ew")
        sheet_actions.grid_columnconfigure(3, weight=1)

        self.paste_benef_button = ctk.CTkButton(
            sheet_actions,
            text="Pegar benef/DNI",
            width=132,
            height=30,
            fg_color="#66788a",
            hover_color="#536577",
            command=lambda: self._paste_identifier_clipboard("benef"),
        )
        self.paste_benef_button.grid(row=0, column=0, padx=(0, 8), sticky="w")

        self.paste_ome_button = ctk.CTkButton(
            sheet_actions,
            text="Pegar OMEs",
            width=116,
            height=30,
            fg_color="#66788a",
            hover_color="#536577",
            command=lambda: self._paste_identifier_clipboard("orden"),
        )
        self.paste_ome_button.grid(row=0, column=1, padx=8, sticky="w")

        self.clear_sheet_button = ctk.CTkButton(
            sheet_actions,
            text="Limpiar",
            width=96,
            height=30,
            fg_color="#8a5a5a",
            hover_color="#734949",
            command=self._clear_identifier_grid,
        )
        self.clear_sheet_button.grid(row=0, column=2, padx=8, sticky="w")

        ctk.CTkLabel(
            sheet_actions,
            text="Cada fila representa una activacion. Si no eliges una boca puntual, se intentara usar una opción valida disponible.",
            font=ctk.CTkFont(size=12),
            text_color="#7a5c27",
        ).grid(row=0, column=3, padx=(14, 0), sticky="w")

        table_frame = ctk.CTkFrame(bulk_frame, corner_radius=10, fg_color="#ffffff")
        table_frame.grid(row=4, column=0, padx=10, pady=(0, 8), sticky="ew")
        table_frame.grid_columnconfigure(0, weight=1)
        table_frame.grid_rowconfigure(1, weight=1)

        header_frame = ctk.CTkFrame(table_frame, fg_color="#ffffff")
        header_frame.grid(row=0, column=0, sticky="ew", padx=6, pady=(6, 0))
        header_frame.grid_columnconfigure(0, weight=1)
        header_frame.grid_columnconfigure(1, weight=1)

        ctk.CTkLabel(
            header_frame,
            text="BENEF. / DNI",
            font=ctk.CTkFont(size=13, weight="bold"),
            text_color="#16324f",
        ).grid(row=0, column=0, sticky="ew", padx=(0, 6), pady=(0, 6))
        ctk.CTkLabel(
            header_frame,
            text="NRO. OME / ORDEN",
            font=ctk.CTkFont(size=13, weight="bold"),
            text_color="#16324f",
        ).grid(row=0, column=1, sticky="ew", padx=(6, 0), pady=(0, 6))

        self.sheet_divider = tk.Frame(header_frame, width=1, bg="#c7d3df", height=24)
        self.sheet_divider.place(relx=0.5, rely=0.0, relheight=1.0)

        self.identifier_grid = ctk.CTkScrollableFrame(table_frame, height=120, corner_radius=0, fg_color="#ffffff")
        self.identifier_grid.grid(row=1, column=0, sticky="ew", padx=6, pady=(0, 6))
        self.identifier_grid.grid_columnconfigure(0, weight=1)
        self.identifier_grid.grid_columnconfigure(1, weight=1)
        self.identifier_grid.bind("<MouseWheel>", self._on_identifier_mousewheel)
        self.identifier_grid.bind("<Button-4>", self._on_identifier_mousewheel)
        self.identifier_grid.bind("<Button-5>", self._on_identifier_mousewheel)

        self.identificador_rows = []
        self.identificador_edit_enabled = True
        self.active_identifier_column = "benef"
        self.active_identifier_row = 0
        self._ensure_identifier_rows(6)
        if self.identificador_rows:
            self.identificador_rows[0][0].focus_set()

        controls_frame = ctk.CTkFrame(content, corner_radius=12, fg_color="#ffffff", border_width=1, border_color="#d8e2ec")
        controls_frame.grid(row=5, column=0, padx=8, pady=(0, 4), sticky="ew")
        controls_frame.grid_columnconfigure(4, weight=1)

        self.close_button = ctk.CTkButton(
            controls_frame,
            text="Cerrar navegador",
            command=lambda: self._run_action(self.controller.cerrar_navegador),
            width=146,
            height=32,
            fg_color="#8a5a5a",
            hover_color="#734949",
        )
        self.close_button.grid(row=0, column=0, padx=(10, 8), pady=8, sticky="w")

        self.run_button = ctk.CTkButton(
            controls_frame,
            text="Ejecutar BOT",
            command=lambda: self._run_action(self._run_lote_with_profile),
            width=166,
            height=32,
            fg_color="#1f7a46",
            hover_color="#176238",
            font=ctk.CTkFont(size=14, weight="bold"),
        )
        self.run_button.grid(row=0, column=1, padx=8, pady=8, sticky="w")

        self.stop_button = ctk.CTkButton(
            controls_frame,
            text="Detener",
            command=self._stop_bot,
            width=114,
            height=32,
            fg_color="#bd6b2a",
            hover_color="#9d571f",
        )
        self.stop_button.grid(row=0, column=2, padx=8, pady=8, sticky="w")

        self.reset_button = ctk.CTkButton(
            controls_frame,
            text="Reiniciar",
            command=self._reset_module_state,
            width=114,
            height=32,
            fg_color="#66788a",
            hover_color="#536577",
        )
        self.reset_button.grid(row=0, column=3, padx=8, pady=8, sticky="w")

        status_frame = ctk.CTkFrame(content, corner_radius=12, fg_color="#ffffff", border_width=1, border_color="#d8e2ec")
        status_frame.grid(row=6, column=0, padx=8, pady=(0, 8), sticky="ew")
        status_frame.grid_columnconfigure(0, weight=1)
        status_frame.grid_columnconfigure(1, weight=1)

        status_header = ctk.CTkFrame(status_frame, fg_color="transparent")
        status_header.grid(row=0, column=0, columnspan=2, padx=10, pady=(8, 6), sticky="ew")
        status_header.grid_columnconfigure(0, weight=1)

        self.status_label = ctk.CTkLabel(
            status_header,
            text="Preparado para ejecutar el BOT.",
            font=ctk.CTkFont(size=14, weight="bold"),
            text_color="#16324f",
        )
        self.status_label.grid(row=0, column=0, padx=(0, 0), pady=(0, 2), sticky="w")

        self.summary_label = ctk.CTkLabel(
            status_header,
            text=f"Log en: {get_log_file()}",
            font=ctk.CTkFont(size=12),
            text_color="#66788a",
        )
        self.summary_label.grid(row=1, column=0, padx=(0, 0), pady=0, sticky="w")

        report_frame = ctk.CTkFrame(status_frame, corner_radius=10, fg_color="#eef3f8")
        report_frame.grid(row=1, column=0, padx=(10, 6), pady=(0, 10), sticky="nsew")
        report_frame.grid_columnconfigure(0, weight=1)

        report_header = ctk.CTkFrame(report_frame, fg_color="transparent")
        report_header.grid(row=0, column=0, padx=10, pady=(8, 4), sticky="ew")
        report_header.grid_columnconfigure(0, weight=1)

        self.report_label = ctk.CTkLabel(
            report_header,
            text="Reporte del lote",
            font=ctk.CTkFont(size=13, weight="bold"),
            text_color="#16324f",
        )
        self.report_label.grid(row=0, column=0, padx=(0, 0), pady=0, sticky="w")

        self.export_report_button = ctk.CTkButton(
            report_header,
            text="Exportar Excel",
            width=132,
            height=30,
            fg_color="#1f7a46",
            hover_color="#176238",
            command=self._export_report_excel,
        )
        self.export_report_button.grid(row=0, column=1, padx=(8, 8), pady=0, sticky="e")

        self.open_report_button = ctk.CTkButton(
            report_header,
            text="Abrir Excel",
            width=116,
            height=30,
            fg_color="#66788a",
            hover_color="#536577",
            command=self._open_exported_report,
            state="disabled",
        )
        self.open_report_button.grid(row=0, column=2, padx=(0, 0), pady=0, sticky="e")

        self.report_text = ctk.CTkTextbox(report_frame, height=104, font=ctk.CTkFont(family="Consolas", size=12))
        self.report_text.grid(row=1, column=0, padx=10, pady=(0, 10), sticky="ew")
        self.report_text.insert("1.0", self._format_report_rows([]))
        self.report_text.configure(state="disabled")

        log_frame = ctk.CTkFrame(status_frame, corner_radius=10, fg_color="#eef3f8")
        log_frame.grid(row=1, column=1, padx=(6, 10), pady=(0, 10), sticky="nsew")
        log_frame.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(
            log_frame,
            text="Log del proceso",
            font=ctk.CTkFont(size=13, weight="bold"),
            text_color="#16324f",
        ).grid(row=0, column=0, padx=10, pady=(8, 4), sticky="w")

        self.log_text = ctk.CTkTextbox(log_frame, height=104, font=ctk.CTkFont(size=13))
        self.log_text.grid(row=1, column=0, padx=10, pady=(0, 10), sticky="ew")
        self.log_text.insert("1.0", "Modulo de activacion de OMEs listo.\n")
        self.log_text.configure(state="disabled")

        self._load_initial_profile_into_form()
        self._render_estado({"url": "", "titulo": "", "en_panel": False, "timestamp": ""})

    # ------------------------------------------------------------------
    # Acciones
    # ------------------------------------------------------------------

    def _open_panel_with_profile(self) -> None:
        profile = self._current_profile_from_form()
        if profile["usuario"]:
            self._upsert_profile(profile)
        self.controller.abrir_panel(
            usuario=profile["usuario"] or None,
            clave=profile["clave"] or None,
            headless=not bool(self.browser_visible_var.get()),
        )

    def _stop_bot(self) -> None:
        if not self.action_running:
            return
        self.controller.solicitar_detencion()
        self.status_label.configure(text="Lote en proceso. Detencion solicitada.")

    def _show_estado(self) -> None:
        estado = self.controller.obtener_estado()
        self.event_queue.put(("estado_detalle", estado))

    def _restart_application(self) -> None:
        if self.action_running:
            return
        root = self.winfo_toplevel()
        restart = getattr(root, "_restart_app", None)
        if callable(restart):
            restart("activar")

    def _reset_module_state(self) -> None:
        if self.action_running:
            return
        self._load_initial_profile_into_form()
        self.fecha_var.set((datetime.today() + timedelta(days=1)).strftime("%d/%m/%Y"))
        self.hora_var.set("08")
        self.minuto_var.set("00")
        self.modalidad_var.set("Presencial")
        self.intervalo_var.set("10")
        self.practica_var.set(DEFAULT_PRACTICA_ACTIVAR)
        self.boca_preset_var.set("Cualquiera disponible")
        self._clear_identifier_grid()
        self.last_report_result = None
        self.last_report_export_path = None
        self._set_report_text(self._format_report_rows([]))
        self.log_text.configure(state="normal")
        self.log_text.delete("1.0", "end")
        self.log_text.insert("1.0", "Modulo de activacion de OMEs listo.\n")
        self.log_text.configure(state="disabled")
        self.status_label.configure(text="Preparado para ejecutar el BOT.")
        self.summary_label.configure(text=f"Log en: {get_log_file()}")

    def _export_report_excel(self) -> None:
        if not self.last_report_result or not self.last_report_result.detalle:
            messagebox.showwarning("Sin reporte", "Todavia no hay resultados para exportar.")
            return

        default_name = f"reporte_activar_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        destino = filedialog.asksaveasfilename(
            title="Guardar reporte de activacion",
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel", "*.xlsx")],
        )
        if not destino:
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte"

        headers = [
            "BENEF / DNI",
            "NRO OME",
            "NOMBRE APELLIDO",
            "CODIGO PRACTICA",
            "HORARIO ACTIVACION",
            "MENSAJE",
        ]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)

        for detalle in self.last_report_result.detalle:
            horario = f"{detalle.fecha} {detalle.hora}:{detalle.minuto}"
            ws.append([
                self._display_beneficio(detalle),
                getattr(detalle, "n_orden_encontrada", "") or getattr(detalle, "n_orden_solicitada", "") or "",
                getattr(detalle, "nombre_afiliado", "") or "",
                self._extract_codigo_practica(detalle),
                horario,
                detalle.mensaje or "",
            ])

        widths = {
            "A": 18,
            "B": 18,
            "C": 34,
            "D": 18,
            "E": 22,
            "F": 48,
        }
        for column, width in widths.items():
            ws.column_dimensions[column].width = width

        wb.save(destino)
        self.last_report_export_path = destino
        self.status_label.configure(text="Reporte exportado a Excel.")
        self.summary_label.configure(text=f"Reporte guardado en: {destino}")
        self.open_report_button.configure(state="normal")
        self._append_log(f"Reporte Excel exportado: {destino}")

    def _open_exported_report(self) -> None:
        if not self.last_report_export_path:
            messagebox.showwarning("Sin archivo", "Todavia no exportaste ningun reporte.")
            return
        if not Path(self.last_report_export_path).exists():
            self.last_report_export_path = None
            self.open_report_button.configure(state="disabled")
            messagebox.showwarning("Archivo no encontrado", "El ultimo reporte exportado ya no existe en esa ubicacion.")
            return

        os.startfile(self.last_report_export_path)
        self._append_log(f"Reporte Excel abierto: {self.last_report_export_path}")

    def _apply_credentials_to_login(self) -> None:
        profile = self._current_profile_from_form()
        if not profile["usuario"]:
            raise RuntimeError("Ingresa un usuario para autocompletar.")
        self._upsert_profile(profile)
        self.controller._try_autofill_credenciales(
            self.controller._get_page(),
            profile["usuario"],
            profile["clave"],
            timeout_ms=12000,
        )

    def _normalize_code(self, value: str) -> str:
        text = (value or "").strip()
        if not text:
            return ""
        digits = "".join(ch if ch.isdigit() else " " for ch in text)
        digit_tokens = [token for token in digits.split() if token]
        if digit_tokens:
            longest = max(digit_tokens, key=len)
            if len(longest) >= 6:
                return longest[:6].strip().upper()
        if " - " in text:
            return text.split(" - ", 1)[0].strip().upper()
        for separator in ("/", "\\", ",", ";"):
            if separator in text:
                return text.split(separator, 1)[0].strip().upper()
        return text.split()[0].strip().upper()

    def _normalize_search_text(self, value: str) -> str:
        text = unicodedata.normalize("NFKD", value or "")
        text = "".join(ch for ch in text if not unicodedata.combining(ch))
        lowered = text.lower()
        for token in ("(", ")", "/", "\\", ",", ";", "-", "_", ".", ":"):
            lowered = lowered.replace(token, " ")
        return " ".join(lowered.split())

    def _resolve_practice_code(self, value: str) -> str:
        text = (value or "").strip()
        if not text:
            return ""
        if text == "Cualquiera disponible":
            return ""

        normalized = self._normalize_code(text)
        if normalized.isdigit():
            return normalized

        normalized_search = self._normalize_search_text(text)
        for option_text in self.practice_options:
            if not option_text or option_text == "Cualquiera disponible":
                continue
            option_search = self._normalize_search_text(option_text)
            if (
                normalized_search == option_search
                or normalized_search in option_search
                or option_search in normalized_search
            ):
                option_code = self._normalize_code(option_text)
                if option_code:
                    return option_code

        return normalized

    def _on_practice_combo_keyrelease(self, _event=None) -> None:
        current_value = self.practica_var.get().strip().lower()
        options = list(self.practice_options)
        if not current_value:
            filtered = options
        else:
            filtered = [option for option in options if current_value in option.lower()]
        self.practica_combo.configure(values=filtered or options)

    def _on_practice_combo_focus_out(self, _event=None) -> None:
        current_value = self.practica_var.get().strip()
        options = list(self.practice_options)
        if not current_value:
            self.practica_var.set(DEFAULT_PRACTICA_ACTIVAR if DEFAULT_PRACTICA_ACTIVAR in options else "Cualquiera disponible")
        elif current_value in options:
            pass
        else:
            matches = [option for option in options if current_value.lower() in option.lower()]
            if len(matches) == 1:
                self.practica_var.set(matches[0])
        self.practica_combo.configure(values=options)

    def _apply_practice_options(self, displays: list[str]) -> None:
        normalized = [value.strip() for value in displays if value.strip()]
        if not normalized:
            normalized = list(PRACTICAS_ACTIVAR.keys())
        if "Cualquiera disponible" not in normalized:
            normalized.append("Cualquiera disponible")
        self.practice_options = normalized
        self.filtered_practice_options = list(normalized)
        current_value = self.practica_var.get().strip()
        if current_value not in self.practice_options:
            self.practica_var.set(
                DEFAULT_PRACTICA_ACTIVAR
                if DEFAULT_PRACTICA_ACTIVAR in self.practice_options
                else self.practice_options[0]
                if self.practice_options
                else "Cualquiera disponible"
            )
        self.practica_combo.configure(values=self.filtered_practice_options)

    def _default_active_catalog_practices(self, items: list[PracticeCatalogItem] | None = None) -> set[str]:
        if items:
            return {item.code for item in items}
        return {self._normalize_code(value) for value in PRACTICAS_ACTIVAR if self._normalize_code(value)}

    def _load_practice_catalog_state(self) -> dict:
        if not self.practice_catalog_file.exists():
            return {}
        try:
            return json.loads(self.practice_catalog_file.read_text(encoding="utf-8"))
        except Exception:
            return {}

    def _save_practice_catalog_state(self) -> None:
        payload = {
            "catalog_path": str(self.practice_catalog_path) if self.practice_catalog_path else "",
            "active_modules": sorted(self.active_catalog_modules),
            "active_practices": sorted(self.active_catalog_practices),
        }
        self.practice_catalog_file.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

    def _resolve_catalog_path_from_state(self, state: dict) -> Path:
        if state.get("catalog_path"):
            saved_path = Path(state.get("catalog_path", "")).expanduser()
            if saved_path.exists():
                return saved_path
        return next((path for path in _default_nomenclador_candidates() if path.exists()), DEFAULT_NOMENCLADOR_PATH)

    def _read_nomenclador_catalog(self, catalog_path: Path) -> list[PracticeCatalogItem]:
        workbook = load_workbook(catalog_path, read_only=True, data_only=True)
        if "Nomenclador" not in workbook.sheetnames:
            raise RuntimeError("El archivo no tiene una hoja 'Nomenclador'.")
        ws = workbook["Nomenclador"]
        items: list[PracticeCatalogItem] = []
        for row in ws.iter_rows(min_row=11, values_only=True):
            if not row or len(row) < 4:
                continue
            module_code = str(row[0] or "").strip()
            module_name = " ".join(str(row[1] or "").split())
            practice_code = str(row[2] or "").strip()
            description_candidates = [
                " ".join(str(row[index] or "").split())
                for index in (3, 4)
                if index < len(row)
            ]
            normalized_practice_code = self._normalize_code(practice_code)
            description = next(
                (
                    value
                    for value in description_candidates
                    if value and self._normalize_code(value) != normalized_practice_code
                ),
                "",
            )
            if not module_code or not module_name or not practice_code or not description:
                continue
            items.append(
                PracticeCatalogItem(
                    module_id=f"{module_code}::{module_name}",
                    module_name=module_name,
                    code=self._normalize_code(practice_code),
                    description=description,
                )
            )
        items.sort(key=lambda item: (item.module_name.lower(), item.code))
        return items

    def _refresh_practice_options_from_catalog(self, *, save: bool = True) -> None:
        displays = [
            item.display
            for item in self.practice_catalog_items
            if item.code in self.active_catalog_practices and (not self.active_catalog_modules or item.module_id in self.active_catalog_modules)
        ]
        if self.practice_catalog_items and not displays:
            self.active_catalog_practices = self._default_active_catalog_practices(self.practice_catalog_items)
            self.active_catalog_modules = {
                item.module_id for item in self.practice_catalog_items if item.code in self.active_catalog_practices
            }
            displays = [
                item.display
                for item in self.practice_catalog_items
                if item.code in self.active_catalog_practices and (not self.active_catalog_modules or item.module_id in self.active_catalog_modules)
            ]
        self._apply_practice_options(displays)
        if save and self.practice_catalog_path:
            self._save_practice_catalog_state()

    def _initialize_practice_catalog(self) -> None:
        state = self._load_practice_catalog_state()
        candidate_path = self._resolve_catalog_path_from_state(state)
        if not candidate_path.exists():
            self._apply_practice_options(list(PRACTICAS_ACTIVAR.keys()))
            return
        try:
            items = self._read_nomenclador_catalog(candidate_path)
        except Exception as exc:
            self._push_log(f"Nomenclador no cargado: {exc}")
            self._apply_practice_options(list(PRACTICAS_ACTIVAR.keys()))
            return

        self.practice_catalog_path = candidate_path
        self.practice_catalog_items = items
        known_module_ids = {item.module_id for item in items}
        known_codes = {item.code for item in items}
        self.active_catalog_modules = {module_id for module_id in state.get("active_modules", []) if module_id in known_module_ids}
        self.active_catalog_practices = {
            self._normalize_code(code)
            for code in state.get("active_practices", [])
            if self._normalize_code(code) in known_codes
        }
        if not self.active_catalog_practices:
            self.active_catalog_practices = self._default_active_catalog_practices(items)
            self.active_catalog_modules = {
                item.module_id for item in items if item.code in self.active_catalog_practices
            }
        elif not self.active_catalog_modules:
            self.active_catalog_modules = {
                item.module_id for item in items if item.code in self.active_catalog_practices
            }
        self._refresh_practice_options_from_catalog(save=False)

    def _open_practice_catalog_manager(self) -> None:
        state = self._load_practice_catalog_state()
        catalog_path = self.practice_catalog_path or self._resolve_catalog_path_from_state(state)
        if not catalog_path.exists():
            initial_dir = next(
                (path.parent for path in _default_nomenclador_candidates() if path.parent.exists()),
                get_output_dir(),
            )
            selected = filedialog.askopenfilename(
                title="Seleccionar nomenclador PAMI",
                initialdir=str(initial_dir),
                filetypes=[("Excel", "*.xlsx *.xlsm")],
            )
            if not selected:
                return
            catalog_path = Path(selected).expanduser().resolve()
        try:
            items = self._read_nomenclador_catalog(catalog_path)
        except Exception as exc:
            messagebox.showerror("Nomenclador", str(exc))
            return

        active_modules = self.active_catalog_modules or set(state.get("active_modules", []))
        active_practices = self.active_catalog_practices or {self._normalize_code(code) for code in state.get("active_practices", [])}
        if not active_practices:
            active_practices = self._default_active_catalog_practices(items)
            active_modules = {item.module_id for item in items if item.code in active_practices}

        dialog = PracticeCatalogDialog(
            self,
            module_title="Activar OME",
            catalog_path=catalog_path,
            items=items,
            active_modules=active_modules,
            active_practices=active_practices,
        )
        self.wait_window(dialog)
        if not dialog.result:
            return
        self.practice_catalog_path = Path(dialog.result["catalog_path"])
        self.practice_catalog_items = items
        self.active_catalog_modules = set(dialog.result.get("active_modules", []))
        self.active_catalog_practices = {self._normalize_code(code) for code in dialog.result.get("active_practices", [])}
        self._refresh_practice_options_from_catalog(save=True)
        self._push_log(f"Nomenclador aplicado: {len(self.active_catalog_practices)} codigo(s) activos en activar OME.")

    def _load_sheet_settings(self) -> dict:
        try:
            if not self.sheet_settings_file.exists():
                return {}
            data = json.loads(self.sheet_settings_file.read_text(encoding="utf-8"))
            if isinstance(data, dict):
                return data
        except Exception:
            pass
        return {}

    def _extract_sheet_profiles(self, data: dict) -> list[dict]:
        profiles: list[dict] = []
        raw_profiles = data.get("profiles") if isinstance(data, dict) else None
        if isinstance(raw_profiles, list):
            for item in raw_profiles:
                if not isinstance(item, dict):
                    continue
                profile = {
                    "profile_id": str(item.get("profile_id", "")).strip(),
                    "spreadsheet_url": normalize_spreadsheet_url(str(item.get("spreadsheet_url", ""))),
                    "sheet_name": str(item.get("sheet_name", "")).strip(),
                    "start_row": str(item.get("start_row", "2")).strip() or "2",
                    "max_rows": str(item.get("max_rows", "40")).strip() or "40",
                    "pami_usuario": str(item.get("pami_usuario", "")).strip(),
                    "pami_nombre": str(item.get("pami_nombre", "")).strip(),
                    "browser_visible": bool(item.get("browser_visible", True)),
                }
                if profile["spreadsheet_url"]:
                    profiles.append(profile)

        legacy_url = normalize_spreadsheet_url(str(data.get("spreadsheet_url", ""))) if isinstance(data, dict) else ""
        if legacy_url:
            legacy_profile = {
                "profile_id": "",
                "spreadsheet_url": legacy_url,
                "sheet_name": str(data.get("sheet_name", "")).strip(),
                "start_row": str(data.get("start_row", "2")).strip() or "2",
                "max_rows": str(data.get("max_rows", "40")).strip() or "40",
                "pami_usuario": str(data.get("pami_usuario", "")).strip(),
                "pami_nombre": str(data.get("pami_nombre", "")).strip(),
                "browser_visible": bool(data.get("browser_visible", True)),
            }
            legacy_signature = self._sheet_profile_signature(legacy_profile)
            if not any(self._sheet_profile_signature(item) == legacy_signature for item in profiles):
                profiles.insert(0, legacy_profile)
        return profiles[:25]

    def _sheet_profile_signature(self, item: dict) -> tuple[str, str, str, str]:
        return (
            str(item.get("spreadsheet_url", "")).strip().lower(),
            str(item.get("sheet_name", "")).strip().lower(),
            str(item.get("start_row", "2")).strip() or "2",
            str(item.get("max_rows", "40")).strip() or "40",
        )

    def _sheet_profile_key(self, item: dict) -> str:
        return str(item.get("profile_id", "")).strip() or "|".join(self._sheet_profile_signature(item))

    def _sheet_profile_options(self) -> list[str]:
        self.sheet_profile_lookup = {}
        options: list[str] = []
        for item in self.sheet_profiles:
            display = self._format_sheet_profile_entry(item)
            self.sheet_profile_lookup[display] = item
            options.append(display)
        return options

    def _format_sheet_profile_entry(self, item: dict) -> str:
        sheet_name = str(item.get("sheet_name", "")).strip() or "Sin pestana"
        start_row = str(item.get("start_row", "2")).strip() or "2"
        max_rows = str(item.get("max_rows", "40")).strip() or "40"
        detail = f"{sheet_name} | fila {start_row} | tope {max_rows}"
        pami_nombre = str(item.get("pami_nombre", "")).strip()
        pami_usuario = str(item.get("pami_usuario", "")).strip()
        if pami_nombre or pami_usuario:
            detail += f" | {pami_nombre or pami_usuario}"
        return detail

    def _sheet_name_values(self) -> list[str]:
        values = [str(value or "").strip() for value in self.sheet_name_options if str(value or "").strip()]
        current = (self.sheet_name_var.get() or "").strip() if hasattr(self, "sheet_name_var") else ""
        if current and current not in values:
            values.insert(0, current)
        return values or ([current] if current else [""])

    def _refresh_sheet_name_values(self, values: list[str] | None = None, *, selected: str | None = None) -> None:
        if values is not None:
            cleaned = [str(value or "").strip() for value in values if str(value or "").strip()]
            self.sheet_name_options = cleaned or self.sheet_name_options
        target = (selected or self.sheet_name_var.get() or "").strip()
        options = self._sheet_name_values()
        if hasattr(self, "sheet_name_entry") and self.sheet_name_entry is not None:
            self.sheet_name_entry.configure(values=options)
        if target:
            self.sheet_name_var.set(target)
        elif options:
            self.sheet_name_var.set(options[0])

    def _current_sheet_profile_from_form(self) -> dict:
        return {
            "profile_id": self.selected_sheet_profile_id or "",
            "spreadsheet_url": normalize_spreadsheet_url((self.sheet_url_var.get() or "").strip()),
            "sheet_name": (self.sheet_name_var.get() or "").strip(),
            "start_row": (self.sheet_start_row_var.get() or "").strip() or "2",
            "max_rows": (self.sheet_max_rows_var.get() or "").strip() or "40",
            "pami_usuario": (self.profile_user_var.get() or "").strip(),
            "pami_nombre": (self.profile_name_var.get() or "").strip(),
            "browser_visible": bool(self.browser_visible_var.get()),
        }

    def _restore_sheet_profile_selection(self) -> None:
        options = self._sheet_profile_options()
        if hasattr(self, "sheet_profile_combo") and self.sheet_profile_combo is not None:
            self.sheet_profile_combo.configure(values=options or [""])
        current_key = str(self.selected_sheet_profile_id or "").strip() or self._sheet_profile_key(self._current_sheet_profile_from_form())
        for display, item in self.sheet_profile_lookup.items():
            if self._sheet_profile_key(item) == current_key:
                self.selected_sheet_profile_id = self._sheet_profile_key(item)
                self.sheet_profile_var.set(display)
                self._refresh_sheet_url_display()
                return
        self.sheet_profile_var.set("")
        self._refresh_sheet_url_display()

    def _sheet_url_display_text(self) -> str:
        selected = str(self.sheet_profile_var.get() or "").strip()
        if selected:
            return selected
        current_sheet = (self.sheet_name_var.get() or "").strip()
        current_start = (self.sheet_start_row_var.get() or "").strip() or "2"
        current_max = (self.sheet_max_rows_var.get() or "").strip() or "40"
        current_pami = (self.profile_name_var.get() or self.profile_user_var.get() or "").strip()
        if current_sheet:
            detail = f"{current_sheet} | fila {current_start} | tope {current_max}"
            if current_pami:
                detail += f" | {current_pami}"
            return detail
        if normalize_spreadsheet_url((self.sheet_url_var.get() or "").strip()):
            return "Hoja sin guardar"
        return "Sin hoja seleccionada"

    def _refresh_sheet_url_display(self) -> None:
        if hasattr(self, "sheet_url_display_var"):
            self.sheet_url_display_var.set(self._sheet_url_display_text())

    def _set_sheet_url_edit_mode(self, editable: bool, *, focus: bool = False) -> None:
        self.sheet_url_editable = bool(editable)
        entry_state = "normal" if self.sheet_url_editable else "readonly"
        button_text = "Listo" if self.sheet_url_editable else "Editar URL"
        button_color = "#245b9d" if self.sheet_url_editable else "#6d7f90"
        button_hover = "#1d4b82" if self.sheet_url_editable else "#536577"
        if hasattr(self, "sheet_url_entry") and self.sheet_url_entry is not None:
            if self.sheet_url_editable:
                self.sheet_url_entry.configure(textvariable=self.sheet_url_var, state=entry_state)
            else:
                self._refresh_sheet_url_display()
                self.sheet_url_entry.configure(textvariable=self.sheet_url_display_var, state=entry_state)
            if focus and self.sheet_url_editable:
                self.sheet_url_entry.focus_set()
                self.sheet_url_entry.icursor("end")
        if hasattr(self, "sheet_url_edit_button") and self.sheet_url_edit_button is not None:
            self.sheet_url_edit_button.configure(text=button_text, fg_color=button_color, hover_color=button_hover)

    def _toggle_sheet_url_edit(self) -> None:
        self._set_sheet_url_edit_mode(not bool(getattr(self, "sheet_url_editable", False)), focus=True)

    def _apply_sheet_profile(self, profile: dict, *, select_pami: bool = True, save: bool = False) -> None:
        if not profile:
            return
        self._syncing_sheet_profile = True
        try:
            self.selected_sheet_profile_id = self._sheet_profile_key(profile)
            self.sheet_url_var.set(profile.get("spreadsheet_url", ""))
            self._refresh_sheet_name_values([profile.get("sheet_name", "")], selected=profile.get("sheet_name", ""))
            self.sheet_start_row_var.set(profile.get("start_row", "2"))
            self.sheet_max_rows_var.set(profile.get("max_rows", "40"))
            self.browser_visible_var.set(bool(profile.get("browser_visible", True)))
            self._set_sheet_url_edit_mode(False)
            if select_pami:
                self._select_pami_profile_for_sheet(profile)
            self._restore_sheet_profile_selection()
            if save:
                self._save_sheet_settings()
        finally:
            self._syncing_sheet_profile = False

    def _on_sheet_profile_selected(self, selected: str) -> None:
        profile = self.sheet_profile_lookup.get(selected)
        if not profile:
            return
        self._apply_sheet_profile(profile, select_pami=True, save=True)

    def _new_sheet_profile(self) -> None:
        self.selected_sheet_profile_id = ""
        self.sheet_profile_var.set("")
        self.sheet_url_var.set("")
        self._refresh_sheet_name_values(["Mc Dube"], selected="Mc Dube")
        self.sheet_start_row_var.set("2")
        self.sheet_max_rows_var.set("40")
        self._set_sheet_url_edit_mode(True, focus=True)
        self._save_sheet_settings()

    def _delete_current_sheet_profile(self) -> None:
        selected = (self.sheet_profile_var.get() or "").strip()
        profile = self.sheet_profile_lookup.get(selected)
        if not profile:
            current_key = self._sheet_profile_key(self._current_sheet_profile_from_form())
            for item in self.sheet_profiles:
                if self._sheet_profile_key(item) == current_key:
                    profile = item
                    break
        if not profile:
            messagebox.showwarning("Google Sheets", "No hay una hoja guardada seleccionada para borrar.")
            return
        if not messagebox.askyesno("Google Sheets", f"Vas a borrar la hoja guardada '{profile.get('sheet_name', 'Sin pestana')}'."):
            return
        target_key = self._sheet_profile_key(profile)
        self.sheet_profiles = [item for item in self.sheet_profiles if self._sheet_profile_key(item) != target_key]
        if self.selected_sheet_profile_id == target_key:
            self.selected_sheet_profile_id = ""
        self.sheet_profile_var.set("")
        self.sheet_url_var.set("")
        self._refresh_sheet_name_values(["Mc Dube"], selected="Mc Dube")
        self.sheet_start_row_var.set("2")
        self.sheet_max_rows_var.set("40")
        self._set_sheet_url_edit_mode(True, focus=True)
        self._save_sheet_settings()
        self._restore_sheet_profile_selection()
        self._push_log(f"Hoja borrada: {profile.get('sheet_name', '')} | {profile.get('spreadsheet_url', '')}")

    def _save_current_sheet_profile(self) -> None:
        profile = self._current_sheet_profile_from_form()
        if not profile["spreadsheet_url"]:
            messagebox.showwarning("Google Sheets", "Ingresa la URL de la hoja para guardarla.")
            return
        if is_office_file_url(self.sheet_url_var.get() or ""):
            messagebox.showwarning("Google Sheets", OFFICE_FILE_MESSAGE)
            return
        if not profile["sheet_name"]:
            messagebox.showwarning("Google Sheets", "Ingresa el nombre de la pestana para guardarla.")
            return
        target_key = self._sheet_profile_key(profile)
        self.selected_sheet_profile_id = target_key
        updated = False
        for idx, item in enumerate(self.sheet_profiles):
            if self._sheet_profile_key(item) == target_key:
                self.sheet_profiles[idx] = profile
                updated = True
                break
        if not updated:
            self.sheet_profiles.insert(0, profile)
        self.sheet_profiles = self.sheet_profiles[:25]
        self._save_sheet_settings()
        self._restore_sheet_profile_selection()
        self.sheet_profile_var.set(self._format_sheet_profile_entry(profile))
        self._set_sheet_url_edit_mode(False)
        self._push_log(f"Hoja guardada: {profile['sheet_name']} | {profile['spreadsheet_url']}")

    def _save_sheet_settings(self) -> None:
        payload = self._current_sheet_profile_from_form() if hasattr(self, "sheet_url_var") else {
            "spreadsheet_url": "",
            "sheet_name": "",
            "start_row": "2",
            "max_rows": "40",
            "pami_usuario": "",
            "pami_nombre": "",
        }
        payload["browser_visible"] = bool(self.browser_visible_var.get()) if hasattr(self, "browser_visible_var") else True
        current_key = self._sheet_profile_key(payload)
        for idx, item in enumerate(self.sheet_profiles):
            if self._sheet_profile_key(item) == current_key:
                self.sheet_profiles[idx] = dict(payload)
                break
        payload["selected_profile_id"] = str(self.selected_sheet_profile_id or "").strip()
        payload["profiles"] = self.sheet_profiles
        try:
            self.sheet_settings_file.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
        except Exception as exc:
            self._push_log(f"Sheets config: no se pudo guardar configuracion ({exc}).")

    def _start_sheets_status_check(self) -> None:
        token_path = get_sheets_token_path()
        if not token_path.exists():
            self.sheets_connected = False
            self.sheets_status_var.set("Google Sheets no conectado")
            return

        def worker():
            try:
                email = get_connected_google_email(token_path=token_path, interactive=False)
                if email:
                    self.event_queue.put(("sheets_connected", email))
                else:
                    self.event_queue.put(("sheets_status", "Token Sheets encontrado, pero hay que reconectar."))
            except Exception:
                self.event_queue.put(("sheets_status", "Token Sheets encontrado, pero hay que reconectar."))

        threading.Thread(target=worker, daemon=True).start()

    def _connect_sheets_account(self) -> None:
        self._save_sheet_settings()
        email = get_connected_google_email(token_path=get_sheets_token_path(), interactive=True)
        self.event_queue.put(("sheets_connected", email or "cuenta Google"))
        self._push_log(f"Google Sheets conectado: {email or 'cuenta Google'}")

    def _load_sheet_tabs_for_current_url(self) -> None:
        sheet_url = normalize_spreadsheet_url((self.sheet_url_var.get() or "").strip())
        if not sheet_url:
            raise RuntimeError("Ingresa la URL de Google Sheets antes de leer paginas.")
        if is_office_file_url(self.sheet_url_var.get() or ""):
            raise RuntimeError(OFFICE_FILE_MESSAGE)

        spreadsheet_id = extract_spreadsheet_id(sheet_url)
        service = build_sheets_service(interactive=False)
        response = self._execute_sheets_request(
            service.spreadsheets().get(
                spreadsheetId=spreadsheet_id,
                fields="sheets.properties.title",
            )
        )
        titles = [
            str(item.get("properties", {}).get("title", "")).strip()
            for item in response.get("sheets", [])
            if str(item.get("properties", {}).get("title", "")).strip()
        ]
        if not titles:
            raise RuntimeError("No se encontraron paginas en esa hoja.")

        current = (self.sheet_name_var.get() or "").strip()
        target = current if current in titles else titles[0]
        self._refresh_sheet_name_values(titles, selected=target)
        self._save_sheet_settings()
        self._push_log(f"Paginas cargadas desde Sheets: {', '.join(titles[:8])}{'...' if len(titles) > 8 else ''}")

    def _sheet_start_row(self) -> int:
        raw_value = (self.sheet_start_row_var.get() or "").strip()
        if not raw_value:
            return 2
        try:
            start_row = int(raw_value)
        except ValueError as exc:
            raise RuntimeError("La fila inicial de Sheets debe ser un numero entero.") from exc
        if start_row < 2:
            raise RuntimeError("La fila inicial de Sheets debe ser 2 o mayor.")
        return start_row

    def _sheet_max_rows(self) -> int:
        raw_value = (self.sheet_max_rows_var.get() or "").strip()
        if not raw_value:
            return 40
        try:
            max_rows = int(raw_value)
        except ValueError as exc:
            raise RuntimeError("El tope de Sheets debe ser un numero entero.") from exc
        if max_rows < 1:
            raise RuntimeError("El tope de Sheets debe ser 1 o mayor.")
        return max_rows

    def _execute_sheets_request(self, request):
        try:
            return request.execute()
        except Exception as exc:
            raise RuntimeError(f"No se pudo operar con Google Sheets: {exc}") from exc

    def _normalize_sheet_header(self, value: str) -> str:
        text = unicodedata.normalize("NFKD", str(value or ""))
        text = "".join(char for char in text if not unicodedata.combining(char))
        text = re.sub(r"[^a-zA-Z0-9]+", " ", text).strip().lower()
        return re.sub(r"\s+", " ", text)

    def _find_header_index(self, headers: list[str], *aliases: str) -> int | None:
        normalized_headers = [self._normalize_sheet_header(header) for header in headers]
        normalized_aliases = [self._normalize_sheet_header(alias) for alias in aliases if alias]
        for alias in normalized_aliases:
            for index, header in enumerate(normalized_headers):
                if not header:
                    continue
                if header == alias or alias in header or header in alias:
                    return index
        return None

    def _column_letter(self, index: int) -> str:
        if index < 0:
            raise ValueError("El indice de columna no puede ser negativo.")
        result = ""
        current = index + 1
        while current:
            current, remainder = divmod(current - 1, 26)
            result = chr(65 + remainder) + result
        return result

    def _looks_like_headerless_plan_salud_row(self, values: list[str]) -> bool:
        if len(values) < 6:
            return False
        first_cell = str(values[0] or "").strip()
        if not re.match(r"^\d{1,2}/\d{1,2}/\d{2,4}$", first_cell):
            return False
        dni_value = re.sub(r"\D+", "", str(values[2] or ""))
        practice_value = str(values[5] or "").strip()
        return len(dni_value) >= 6 and bool(practice_value)

    def _detect_activation_sheet_layout(self, headers: list[str]) -> dict[str, int | None]:
        if self._looks_like_headerless_plan_salud_row(headers):
            return {
                "beneficio_col": 3,
                "dni_col": 2,
                "practica_col": 5,
                "ome_col": 6,
                "nombre_col": 1,
            }
        return {
            "beneficio_col": self._find_header_index(headers, "benef", "beneficio", "num benef", "nro benef", "numero benef"),
            "dni_col": self._find_header_index(headers, "dni", "nro dni", "numero dni", "documento"),
            "practica_col": self._find_header_index(headers, "practica", "practica general", "especialidad o practica", "especialidad practica", "especialidad", "prestacion", "f"),
            "ome_col": self._find_header_index(headers, "ome", "nro ome", "numero ome", "orden"),
            "nombre_col": self._find_header_index(headers, "nombre y apellido", "apellido y nombre", "nombre", "nombre paciente"),
        }

    def _sheet_cell_value(self, row: list[str], index: int | None) -> str:
        if index is None or index < 0 or index >= len(row):
            return ""
        return str(row[index] or "").strip()

    def _sheet_marker_requires_lookup(self, value: str) -> bool:
        normalized = (value or "").strip().upper()
        if not normalized:
            return True
        if normalized in {"GENERADA", "YA TIENE", "YA TIENE OME", "ERROR", "NO ENCONTRADO", "NO ENCONTRADA"}:
            return True
        return not bool(re.fullmatch(r"\d{8,}", re.sub(r"\D+", "", normalized)))

    def _is_numeric_ome_value(self, value: str) -> bool:
        digits = re.sub(r"\D+", "", str(value or ""))
        return bool(re.fullmatch(r"\d{8,}", digits))

    def _split_sheet_ome_slots(self, value: str) -> list[str]:
        text = str(value or "").strip()
        if not text:
            return []
        if "//" not in text:
            return [text]
        return [part.strip() for part in re.split(r"\s*//\s*", text)]

    def _build_sheet_ome_slots(self, ome_raw: str, candidate_practices: list[str]) -> list[dict]:
        raw_slots = self._split_sheet_ome_slots(ome_raw)
        minimum_slots = 1 if (self._sheet_marker_requires_lookup(ome_raw) or candidate_practices) else 0
        slot_count = max(len(raw_slots), len(candidate_practices), minimum_slots)
        if slot_count <= 0:
            return []

        padded_slots = list(raw_slots) + [""] * max(slot_count - len(raw_slots), 0)
        prepared: list[dict] = []
        for index in range(slot_count):
            raw_value = padded_slots[index].strip()
            practice_code = candidate_practices[index] if index < len(candidate_practices) else ""
            prepared.append(
                {
                    "slot_index": index,
                    "raw_value": raw_value,
                    "practice_code": practice_code,
                    "is_numeric": self._is_numeric_ome_value(raw_value),
                    "requires_lookup": (not self._is_numeric_ome_value(raw_value)) and self._sheet_marker_requires_lookup(raw_value),
                }
            )
        return prepared

    def _sheet_practice_candidates(self, raw_practice: str) -> list[str]:
        text = (raw_practice or "").strip()
        if not text:
            return []
        candidates = resolve_plan_salud_practices(text)
        if not candidates:
            single = resolve_plan_salud_practice(text)
            if single:
                candidates = [single]
        if not candidates:
            single = self._resolve_practice_code(text)
            if single:
                candidates = [single]
        deduped: list[str] = []
        seen: set[str] = set()
        for code in candidates:
            normalized = self._normalize_code(code)
            if not normalized or normalized in seen:
                continue
            deduped.append(normalized)
            seen.add(normalized)
        return deduped

    def _build_sheet_records(self) -> list[dict]:
        sheet_url = normalize_spreadsheet_url((self.sheet_url_var.get() or "").strip())
        sheet_name = (self.sheet_name_var.get() or "").strip()
        start_row = self._sheet_start_row()
        max_rows = self._sheet_max_rows()
        if not sheet_url:
            raise RuntimeError("Ingresa la URL de Google Sheets.")
        if is_office_file_url(self.sheet_url_var.get() or ""):
            raise RuntimeError(OFFICE_FILE_MESSAGE)
        if not sheet_name:
            raise RuntimeError("Ingresa el nombre de la pestaña.")

        self.sheet_url_var.set(sheet_url)
        self._save_sheet_settings()

        spreadsheet_id = extract_spreadsheet_id(sheet_url)
        service = build_sheets_service(interactive=False)

        headers_response = self._execute_sheets_request(
            service.spreadsheets().values().get(spreadsheetId=spreadsheet_id, range=f"'{sheet_name}'!A1:Z1")
        )
        header_values = headers_response.get("values", [])
        headers = [str(value or "").strip() for value in (header_values[0] if header_values else [])]
        layout = self._detect_activation_sheet_layout(headers)
        if layout["ome_col"] is None:
            raise RuntimeError(f"No se encontro la columna OME en la hoja '{sheet_name}'.")
        if layout["beneficio_col"] is None and layout["dni_col"] is None:
            raise RuntimeError(f"No se encontro una columna de BENEF o DNI en la hoja '{sheet_name}'.")

        response = self._execute_sheets_request(
            service.spreadsheets().values().get(spreadsheetId=spreadsheet_id, range=f"'{sheet_name}'!A{start_row}:Z")
        )
        values = response.get("values", [])

        start_dt = datetime.strptime(f"{self.fecha_var.get().strip()} {self.hora_var.get().strip().zfill(2)}:{self.minuto_var.get().strip().zfill(2)}", "%d/%m/%Y %H:%M")
        intervalo = int(self.intervalo_var.get().strip() or "10")
        modalidad = MODALIDADES.get(self.modalidad_var.get().strip(), "P")
        boca = BOCAS_CONOCIDAS.get(self.boca_preset_var.get().strip(), "")

        prepared: list[dict] = []
        current_dt = start_dt
        default_practice = self._resolve_practice_code(self.practica_var.get().strip())
        for offset, row in enumerate(values, start=start_row):
            beneficio = self._sheet_cell_value(row, layout["beneficio_col"])
            dni = self._sheet_cell_value(row, layout["dni_col"])
            practica_raw = self._sheet_cell_value(row, layout.get("practica_col"))
            ome_raw = self._sheet_cell_value(row, layout["ome_col"])
            ome_digits = self._normalize_identifier_cell(ome_raw)
            identificador = beneficio or dni
            if not identificador:
                continue

            lookup_needed = self._sheet_marker_requires_lookup(ome_raw)
            candidate_practices = self._sheet_practice_candidates(practica_raw)
            if not candidate_practices and default_practice:
                candidate_practices = [default_practice]
            if not ome_digits and not lookup_needed:
                continue
            if not ome_digits and not candidate_practices:
                continue

            prepared.append(
                {
                    "sheet_row": offset,
                    "input_beneficio": beneficio,
                    "input_dni": dni,
                    "n_afiliado": identificador,
                    "n_orden": ome_digits,
                    "fecha": current_dt.strftime("%d/%m/%Y"),
                    "hora": current_dt.strftime("%H"),
                    "minuto": current_dt.strftime("%M"),
                    "modalidad": modalidad,
                    "practica": "" if ome_digits else candidate_practices[0],
                    "boca": boca,
                    "generated_search": not bool(ome_digits),
                    "allow_legacy_4271_fallback": (not ome_digits) and not bool(practica_raw.strip()),
                    "candidate_practices": candidate_practices,
                    "practica_origen": practica_raw,
                    "ome_marker": (ome_raw or "").strip(),
                }
            )
            current_dt += timedelta(minutes=intervalo)
            if max_rows and len(prepared) >= max_rows:
                break

        if not prepared:
            raise RuntimeError(f"No hay filas pendientes en la hoja para procesar desde la fila {start_row}.")
        return prepared

    def _write_sheet_results(self, result_rows: list[dict]) -> int:
        sheet_url = normalize_spreadsheet_url((self.sheet_url_var.get() or "").strip())
        sheet_name = (self.sheet_name_var.get() or "").strip()
        spreadsheet_id = extract_spreadsheet_id(sheet_url)
        service = build_sheets_service(interactive=False)

        headers_response = self._execute_sheets_request(
            service.spreadsheets().values().get(spreadsheetId=spreadsheet_id, range=f"'{sheet_name}'!A1:Z1")
        )
        header_values = headers_response.get("values", [])
        headers = [str(value or "").strip() for value in (header_values[0] if header_values else [])]
        layout = self._detect_activation_sheet_layout(headers)
        beneficio_letter = self._column_letter(layout["beneficio_col"]) if layout["beneficio_col"] is not None else None
        activacion_letter = "M"

        def is_success(row: dict) -> bool:
            estado = str(row.get("estado_final", "")).strip().upper()
            escenario = str(row.get("escenario", "")).strip().upper()
            return escenario in {"A", "B"} and estado not in {"ERROR", "PENDIENTE", "VERIFICACION_FALLIDA", "SIN_RESULTADOS", "SIN_ICONO", "NO_ACTIVABLE"}

        data = []
        touched_rows: set[int] = set()
        for row in result_rows:
            sheet_row = row.get("sheet_row")
            if not sheet_row:
                continue
            sheet_row_int = int(sheet_row)
            beneficio_input = str(row.get("input_beneficio", "") or "").strip()
            beneficio_found = str(row.get("beneficio_encontrado", "") or "").strip()
            if not beneficio_input and beneficio_found and beneficio_letter:
                data.append({"range": f"'{sheet_name}'!{beneficio_letter}{sheet_row_int}", "values": [[beneficio_found]]})
                touched_rows.add(sheet_row_int)
            if is_success(row):
                data.append({"range": f"'{sheet_name}'!{activacion_letter}{sheet_row_int}", "values": [[f"{row.get('fecha', '')} {row.get('hora', '')}:{row.get('minuto', '')}"]]})
                touched_rows.add(sheet_row_int)

        if not data:
            return 0

        self._execute_sheets_request(
            service.spreadsheets().values().batchUpdate(
                spreadsheetId=spreadsheet_id,
                body={"valueInputOption": "USER_ENTERED", "data": data},
            )
        )
        return len(touched_rows)

    def _apply_next_schedule_defaults(self, detalles) -> None:
        if not detalles:
            return
        intervalo = int(self.intervalo_var.get().strip() or "10")
        exitos = []
        for detalle in detalles:
            estado = str(getattr(detalle, "estado_final", "") or "").strip().upper()
            escenario = str(getattr(detalle, "escenario", "") or "").strip().upper()
            if escenario in {"A", "B"} and estado not in {"ERROR", "PENDIENTE", "VERIFICACION_FALLIDA", "SIN_RESULTADOS", "SIN_ICONO", "NO_ACTIVABLE"}:
                try:
                    exitos.append(datetime.strptime(f"{detalle.fecha} {detalle.hora}:{detalle.minuto}", "%d/%m/%Y %H:%M"))
                except Exception:
                    continue
        if not exitos:
            return
        siguiente = max(exitos) + timedelta(minutes=intervalo)
        self.fecha_var.set(siguiente.strftime("%d/%m/%Y"))
        self.hora_var.set(siguiente.strftime("%H"))
        self.minuto_var.set(siguiente.strftime("%M"))

    def _run_sheet_batch_with_profile(self) -> None:
        profile = self._current_profile_from_form()
        if not profile["usuario"] or not profile["clave"]:
            raise RuntimeError("Guarda o completa un perfil con usuario y clave antes de ejecutar el lote.")

        try:
            self.controller.obtener_estado()
        except Exception:
            self._open_panel_with_profile()

        records = self._build_sheet_records()
        self.event_queue.put(("lote_started", {"total": len(records)}))

        lote = [
            {
                "n_afiliado": record["n_afiliado"],
                "n_orden": record["n_orden"],
                "fecha": record["fecha"],
                "hora": record["hora"],
                "minuto": record["minuto"],
                "modalidad": record["modalidad"],
                "practica": record["practica"],
                "boca": record["boca"],
            }
            for record in records
        ]
        resultado = self.controller.activar_lote(lote=lote, progress_callback=self._push_progress)

        retry_indices = []
        retry_lote = []
        for idx, (record, detalle) in enumerate(zip(records, resultado.detalle)):
            if record.get("allow_legacy_4271_fallback") and str(detalle.escenario).upper() == "SIN_RESULTADOS":
                retry_indices.append(idx)
                retry_lote.append(
                    {
                        "n_afiliado": record["n_afiliado"],
                        "n_orden": "",
                        "fecha": record["fecha"],
                        "hora": record["hora"],
                        "minuto": record["minuto"],
                        "modalidad": record["modalidad"],
                        "practica": "427109",
                        "boca": record["boca"],
                    }
                )

        if retry_lote:
            retry_resultado = self.controller.activar_lote(lote=retry_lote, progress_callback=self._push_progress)
            for pos, detalle_retry in zip(retry_indices, retry_resultado.detalle):
                resultado.detalle[pos] = detalle_retry

        resultado.ok = 0
        resultado.errores = 0
        resultado.sin_icono = 0
        for detalle in resultado.detalle:
            if detalle.escenario == "SIN_ICONO" or detalle.estado_final == "SIN_ICONO":
                resultado.sin_icono += 1
            elif detalle.escenario == "ERROR" or detalle.estado_final in {"ERROR", "PENDIENTE", "VERIFICACION_FALLIDA"}:
                resultado.errores += 1
            else:
                resultado.ok += 1

        merged_rows = []
        for record, detalle in zip(records, resultado.detalle):
            merged_rows.append(
                {
                    "sheet_row": record.get("sheet_row"),
                    "input_beneficio": record.get("input_beneficio", ""),
                    "input_dni": record.get("input_dni", ""),
                    "beneficio_encontrado": getattr(detalle, "beneficio_encontrado", "") or "",
                    "fecha": detalle.fecha,
                    "hora": detalle.hora,
                    "minuto": detalle.minuto,
                    "estado_final": detalle.estado_final,
                    "escenario": detalle.escenario,
                }
            )

        updated_count = self._write_sheet_results(merged_rows)
        processed_rows = [int(row.get("sheet_row")) for row in merged_rows if str(row.get("sheet_row", "")).isdigit()]
        if processed_rows:
            self.sheet_start_row_var.set(str(max(processed_rows) + 1))
            self._save_sheet_settings()

        self._apply_next_schedule_defaults(resultado.detalle)
        self._push_log(f"Sheets actualizadas: {updated_count}")
        self.event_queue.put(("lote_summary", resultado))

    def _build_sheet_lookup_records(self) -> list[dict]:
        sheet_url = normalize_spreadsheet_url((self.sheet_url_var.get() or "").strip())
        sheet_name = (self.sheet_name_var.get() or "").strip()
        start_row = self._sheet_start_row()
        max_rows = self._sheet_max_rows()
        if not sheet_url:
            raise RuntimeError("Ingresa la URL de Google Sheets.")
        if is_office_file_url(self.sheet_url_var.get() or ""):
            raise RuntimeError(OFFICE_FILE_MESSAGE)
        if not sheet_name:
            raise RuntimeError("Ingresa el nombre de la pestana.")

        self.sheet_url_var.set(sheet_url)
        self._save_sheet_settings()

        spreadsheet_id = extract_spreadsheet_id(sheet_url)
        service = build_sheets_service(interactive=False)
        headers_response = self._execute_sheets_request(
            service.spreadsheets().values().get(spreadsheetId=spreadsheet_id, range=f"'{sheet_name}'!A1:Z1")
        )
        header_values = headers_response.get("values", [])
        headers = [str(value or "").strip() for value in (header_values[0] if header_values else [])]
        layout = self._detect_activation_sheet_layout(headers)
        if layout["ome_col"] is None:
            raise RuntimeError(f"No se encontro la columna OME en la hoja '{sheet_name}'.")
        if layout["beneficio_col"] is None and layout["dni_col"] is None:
            raise RuntimeError(f"No se encontro una columna de BENEF o DNI en la hoja '{sheet_name}'.")

        response = self._execute_sheets_request(
            service.spreadsheets().values().get(spreadsheetId=spreadsheet_id, range=f"'{sheet_name}'!A{start_row}:Z")
        )
        values = response.get("values", [])

        default_practice = self._resolve_practice_code(self.practica_var.get().strip())
        records: list[dict] = []
        processed_source_rows = 0
        for offset, row in enumerate(values, start=start_row):
            beneficio = self._sheet_cell_value(row, layout["beneficio_col"])
            dni = self._sheet_cell_value(row, layout["dni_col"])
            identificador = beneficio or dni
            if not identificador:
                continue

            ome_raw = self._sheet_cell_value(row, layout["ome_col"])

            practica_raw = self._sheet_cell_value(row, layout.get("practica_col"))
            nombre = self._sheet_cell_value(row, layout.get("nombre_col"))
            candidate_practices = self._sheet_practice_candidates(practica_raw)
            if not candidate_practices and default_practice:
                candidate_practices = [default_practice]

            slots = self._build_sheet_ome_slots(ome_raw, candidate_practices)
            numeric_ome_value = ""
            has_lookup_work = False
            for slot in slots:
                if slot.get("is_numeric") and not numeric_ome_value:
                    numeric_ome_value = self._normalize_identifier_cell(str(slot.get("raw_value", "") or ""))
                if not slot.get("requires_lookup"):
                    continue
                practice_code = str(slot.get("practice_code", "") or "").strip()
                if not practice_code:
                    continue
                has_lookup_work = True
                records.append(
                    {
                        "sheet_row": offset,
                        "input_beneficio": beneficio,
                        "input_dni": dni,
                        "input_ome": ome_raw,
                        "n_afiliado": identificador,
                        "n_orden": "",
                        "nombre": nombre,
                        "candidate_practices": [practice_code],
                        "practica_origen": practica_raw,
                        "slot_index": int(slot.get("slot_index", 0) or 0),
                        "slot_count": len(slots),
                        "existing_slots": [str(item.get("raw_value", "") or "").strip() for item in slots],
                        "write_ome": True,
                        "buscar_dni": not bool(dni),
                    }
                )
            if not has_lookup_work and numeric_ome_value and not dni:
                records.append(
                    {
                        "sheet_row": offset,
                        "input_beneficio": beneficio,
                        "input_dni": dni,
                        "input_ome": ome_raw,
                        "n_afiliado": identificador,
                        "n_orden": numeric_ome_value,
                        "nombre": nombre,
                        "candidate_practices": [""],
                        "practica_origen": practica_raw,
                        "slot_index": 0,
                        "slot_count": len(slots) or 1,
                        "existing_slots": [str(item.get("raw_value", "") or "").strip() for item in slots] or [ome_raw],
                        "write_ome": False,
                        "buscar_dni": True,
                    }
                )
                has_lookup_work = True
            if has_lookup_work:
                processed_source_rows += 1
            if max_rows and processed_source_rows >= max_rows:
                break

        if not records:
            raise RuntimeError(f"No hay filas pendientes para completar OME o DNI desde la fila {start_row}.")
        return records

    def _write_sheet_lookup_results(self, result_rows: list[dict]) -> int:
        sheet_url = normalize_spreadsheet_url((self.sheet_url_var.get() or "").strip())
        sheet_name = (self.sheet_name_var.get() or "").strip()
        spreadsheet_id = extract_spreadsheet_id(sheet_url)
        service = build_sheets_service(interactive=False)

        headers_response = self._execute_sheets_request(
            service.spreadsheets().values().get(spreadsheetId=spreadsheet_id, range=f"'{sheet_name}'!A1:Z1")
        )
        header_values = headers_response.get("values", [])
        headers = [str(value or "").strip() for value in (header_values[0] if header_values else [])]
        layout = self._detect_activation_sheet_layout(headers)
        if layout["ome_col"] is None:
            raise RuntimeError(f"No se encontro la columna OME en la hoja '{sheet_name}'.")

        ome_letter = self._column_letter(layout["ome_col"])
        beneficio_letter = self._column_letter(layout["beneficio_col"]) if layout["beneficio_col"] is not None else None
        dni_letter = self._column_letter(layout["dni_col"]) if layout["dni_col"] is not None else None

        grouped: dict[int, dict] = {}
        for row in result_rows:
            sheet_row = row.get("sheet_row")
            if not sheet_row:
                continue
            sheet_row_int = int(sheet_row)
            existing_slots = [str(value or "").strip() for value in (row.get("existing_slots") or [])]
            slot_count = max(int(row.get("slot_count", 0) or 0), len(existing_slots), 1)
            while len(existing_slots) < slot_count:
                existing_slots.append("")
            group = grouped.setdefault(
                sheet_row_int,
                {
                    "slots": list(existing_slots),
                    "input_beneficio": str(row.get("input_beneficio", "") or "").strip(),
                    "input_dni": str(row.get("input_dni", "") or "").strip(),
                    "beneficio_found": "",
                    "dni_found": "",
                    "ome_changed": False,
                },
            )
            while len(group["slots"]) < slot_count:
                group["slots"].append("")
            slot_index = int(row.get("slot_index", 0) or 0)
            if bool(row.get("write_ome", True)):
                located_ome = str(row.get("n_orden_encontrada", "") or "").strip()
                group["slots"][slot_index] = located_ome or "NO ENCONTRADA"
                group["ome_changed"] = True
            beneficio_found = str(row.get("beneficio_encontrado", "") or "").strip()
            if beneficio_found:
                group["beneficio_found"] = beneficio_found
            dni_found = str(row.get("dni_encontrado", "") or "").strip()
            if dni_found:
                group["dni_found"] = dni_found

        data = []
        touched_rows: set[int] = set()
        for sheet_row_int, payload in grouped.items():
            row_changed = False
            if payload.get("ome_changed"):
                visible_value = " // ".join(str(value or "").strip() for value in payload["slots"])
                data.append({"range": f"'{sheet_name}'!{ome_letter}{sheet_row_int}", "values": [[visible_value]]})
                row_changed = True

            beneficio_input = str(payload.get("input_beneficio", "") or "").strip()
            beneficio_found = str(payload.get("beneficio_found", "") or "").strip()
            if not beneficio_input and beneficio_found and beneficio_letter:
                data.append({"range": f"'{sheet_name}'!{beneficio_letter}{sheet_row_int}", "values": [[beneficio_found]]})
                row_changed = True

            dni_input = str(payload.get("input_dni", "") or "").strip()
            dni_found = str(payload.get("dni_found", "") or "").strip()
            if not dni_input and dni_found and dni_letter:
                data.append({"range": f"'{sheet_name}'!{dni_letter}{sheet_row_int}", "values": [[dni_found]]})
                row_changed = True

            if row_changed:
                touched_rows.add(sheet_row_int)

        if not data:
            return 0

        self._execute_sheets_request(
            service.spreadsheets().values().batchUpdate(
                spreadsheetId=spreadsheet_id,
                body={"valueInputOption": "USER_ENTERED", "data": data},
            )
        )
        return len(touched_rows)

    def _run_sheet_ome_lookup_with_profile(self) -> None:
        profile = self._current_profile_from_form()
        if not profile["usuario"] or not profile["clave"]:
            raise RuntimeError("Guarda o completa un perfil con usuario y clave antes de buscar OMEs.")

        try:
            self.controller.obtener_estado()
        except Exception:
            self._open_panel_with_profile()

        records = self._build_sheet_lookup_records()
        self.event_queue.put(("lote_started", {"total": len(records)}))

        lookup_lote = []
        for record in records:
            candidate_practices = list(record.get("candidate_practices") or [])
            if not candidate_practices:
                candidate_practices = [""]
            lookup_lote.append(
                {
                    "n_afiliado": record["n_afiliado"],
                    "n_orden": record.get("n_orden", ""),
                    "candidate_practices": candidate_practices,
                    "buscar_dni": bool(record.get("buscar_dni")),
                }
            )

        resultado = self.controller.buscar_ome_lote(lote=lookup_lote, progress_callback=self._push_progress)

        merged_rows = []
        for record, detalle in zip(records, resultado.detalle):
            merged_rows.append(
                {
                    "sheet_row": record.get("sheet_row"),
                    "input_beneficio": record.get("input_beneficio", ""),
                    "input_dni": record.get("input_dni", ""),
                    "input_ome": record.get("input_ome", ""),
                    "beneficio_encontrado": getattr(detalle, "beneficio_encontrado", "") or "",
                    "dni_encontrado": getattr(detalle, "dni_encontrado", "") or "",
                    "n_orden_encontrada": getattr(detalle, "n_orden_encontrada", "") or "",
                    "slot_index": record.get("slot_index", 0),
                    "slot_count": record.get("slot_count", 1),
                    "existing_slots": record.get("existing_slots", []),
                    "write_ome": bool(record.get("write_ome", True)),
                    "buscar_dni": bool(record.get("buscar_dni")),
                }
            )

        updated_count = self._write_sheet_lookup_results(merged_rows)
        processed_rows = [int(row.get("sheet_row")) for row in merged_rows if str(row.get("sheet_row", "")).isdigit()]
        if processed_rows:
            self.sheet_start_row_var.set(str(max(processed_rows) + 1))
            self._save_sheet_settings()

        found_count = sum(
            1
            for row in merged_rows
            if (bool(row.get("write_ome", True)) and str(row.get("n_orden_encontrada", "") or "").strip())
            or ((not bool(row.get("write_ome", True))) and str(row.get("dni_encontrado", "") or "").strip())
        )
        dni_count = sum(1 for row in merged_rows if str(row.get("dni_encontrado", "") or "").strip())
        self._push_log(f"OMEs/DNI completados en Sheets: {updated_count}")
        self.event_queue.put(("lote_summary", resultado))
        self.event_queue.put(
            (
                "sheet_lookup_finished",
                {
                    "total_rows": len({int(row.get("sheet_row")) for row in merged_rows if str(row.get("sheet_row", "")).isdigit()}),
                    "total_searches": len(merged_rows),
                    "found": found_count,
                    "dni_found": dni_count,
                    "not_found": max(len(merged_rows) - found_count, 0),
                    "updated": updated_count,
                },
            )
        )

    def _build_lote(self) -> list[dict]:
        filas_identificadores = []
        for bene_entry, orden_entry in self.identificador_rows:
            bene = self._normalize_identifier_cell(bene_entry.get())
            orden = self._normalize_identifier_cell(orden_entry.get())
            if bene or orden:
                filas_identificadores.append((bene, orden))

        if not filas_identificadores:
            raise RuntimeError("Completa al menos una fila con beneficio, DNI, numero de orden o ambos.")

        fecha = self.fecha_var.get().strip()
        if not fecha:
            raise RuntimeError("Indica la fecha del turno.")

        hora_inicio = int(self.hora_var.get().strip() or "8")
        minuto_inicio = int(self.minuto_var.get().strip() or "0")
        intervalo = int(self.intervalo_var.get().strip() or "10")
        modalidad = MODALIDADES.get(self.modalidad_var.get().strip(), "P")
        practica = self._resolve_practice_code(self.practica_var.get().strip())
        boca = BOCAS_CONOCIDAS.get(self.boca_preset_var.get().strip(), "")

        dt = datetime.strptime(f"{hora_inicio:02d}:{minuto_inicio:02d}", "%H:%M")
        lote = []
        for index, (bene, orden) in enumerate(filas_identificadores, start=1):
            practica_fila = "" if orden else practica

            if not orden and not practica_fila:
                raise RuntimeError(
                    f"La fila {index} no tiene N° OME y tampoco una practica seleccionada. "
                    "Sin OME y sin practica el modulo no ejecuta para evitar activar al azar."
                )

            lote.append(
                {
                    "n_afiliado": bene,
                    "n_orden": orden,
                    "fecha": fecha,
                    "hora": dt.strftime("%H"),
                    "minuto": dt.strftime("%M"),
                    "modalidad": modalidad,
                    "practica": practica_fila,
                    "boca": boca,
                }
            )
            dt += timedelta(minutes=intervalo)

        if not lote:
            raise RuntimeError("No hay filas validas: cada linea debe tener beneficio, DNI, nro. de orden o ambos.")

        return lote

    def _preview_lote(self) -> None:
        try:
            lote = self._build_lote()
        except Exception as exc:
            messagebox.showwarning("Error en lote", str(exc))
            return

        lineas = [
            f"benef={e['n_afiliado'] or '-'} | orden={e['n_orden'] or '-'} -> {e['fecha']} {e['hora']}:{e['minuto']} mod={e['modalidad']} pract={e['practica'] or 'auto'} boca={e['boca'] or 'auto'}"
            for e in lote
        ]
        messagebox.showinfo("Vista previa del lote", "\n".join(lineas[:40]))

    def _run_lote_with_profile(self) -> None:
        profile = self._current_profile_from_form()
        if not profile["usuario"] or not profile["clave"]:
            raise RuntimeError("Guarda o completa un perfil con usuario y clave antes de ejecutar el lote.")

        try:
            self.controller.obtener_estado()
        except Exception:
            self._open_panel_with_profile()

        lote = self._build_lote()
        self.event_queue.put(("lote_started", {"total": len(lote)}))

        resultado = self.controller.activar_lote(
            lote=lote,
            progress_callback=self._push_progress,
        )
        self.event_queue.put(("lote_summary", resultado))

    # ------------------------------------------------------------------
    # Threading
    # ------------------------------------------------------------------

    def _run_action(self, action) -> None:
        if self.action_running:
            return
        self._ensure_controller_thread()
        self.action_running = True
        self._set_controls_enabled(False)
        self.controller_queue.put(action)

    def _ensure_controller_thread(self) -> None:
        if self.controller_thread is None or not self.controller_thread.is_alive():
            self.controller_thread = threading.Thread(target=self._controller_loop, daemon=True)
            self.controller_thread.start()

    def _controller_loop(self) -> None:
        while True:
            action = self.controller_queue.get()
            if action is None:
                break
            try:
                action()
                self.event_queue.put(("action_finished", None))
            except Exception as exc:
                self.event_queue.put(("action_error", str(exc)))

    def _push_log(self, message: str) -> None:
        self.event_queue.put(("log", message))

    def _push_status(self, message: str) -> None:
        self.event_queue.put(("status", message))

    def _push_progress(self, payload: dict) -> None:
        self.event_queue.put(("progress", payload))

    def _process_ui_queue(self) -> None:
        had_events = False
        try:
            while True:
                event, payload = self.event_queue.get_nowait()
                had_events = True

                try:
                    if event == "log":
                        self._append_log(payload)
                    elif event == "status":
                        self.status_label.configure(text=payload)
                    elif event == "sheets_connected":
                        self.sheets_connected = True
                        self.sheets_status_var.set(f"Google Sheets conectado: {payload}")
                    elif event == "sheets_status":
                        self.sheets_connected = False
                        self.sheets_status_var.set(str(payload or "Google Sheets no conectado"))
                    elif event == "estado_detalle":
                        self._render_estado(payload)
                    elif event == "lote_started":
                        total = payload.get("total", 0)
                        self.last_report_result = None
                        self.last_report_export_path = None
                        self.status_label.configure(text="Lote en proceso.")
                        self.summary_label.configure(text=f"Log en: {get_log_file()}")
                        self._set_report_text(self._format_report_rows([]))
                        self.open_report_button.configure(state="disabled")
                    elif event == "progress":
                        self.status_label.configure(text="Lote en proceso.")
                        self.summary_label.configure(text=f"Log en: {get_log_file()}")
                    elif event == "lote_summary":
                        resultado: ResultadoLote = payload
                        self.last_report_result = resultado
                        resumen = (
                            f"OK: {resultado.ok} | Errores: {resultado.errores} | "
                            f"Sin icono: {resultado.sin_icono} | Log en: {get_log_file()}"
                        )
                        self.status_label.configure(text="Lote finalizado.")
                        self.summary_label.configure(text=f"Reporte listo. Log en: {get_log_file()}")
                        self._set_report_text(self._format_report_rows(resultado.detalle))
                        self._append_log(resumen)
                        for detalle in resultado.detalle:
                            linea = (
                                f"  benef={detalle.n_afiliado or '-'} | orden={getattr(detalle, 'n_orden_solicitada', '') or '-'} "
                                f"-> {detalle.estado_final} (esc={detalle.escenario}) {detalle.fecha} {detalle.hora}:{detalle.minuto}"
                            )
                            orden_real = getattr(detalle, "n_orden_encontrada", "") or "-"
                            practica_real = getattr(detalle, "practica_encontrada", "") or getattr(detalle, "practica", "") or "-"
                            linea += f" | orden real={orden_real} | practica real={practica_real}"
                            if detalle.mensaje:
                                linea += f" | {detalle.mensaje}"
                            self._append_log(linea)
                        self.action_running = False
                        self._set_controls_enabled(True)
                    elif event == "sheet_lookup_finished":
                        total_rows = int(payload.get("total_rows", 0) or 0)
                        total_searches = int(payload.get("total_searches", 0) or 0)
                        found = int(payload.get("found", 0) or 0)
                        dni_found = int(payload.get("dni_found", 0) or 0)
                        not_found = int(payload.get("not_found", 0) or 0)
                        updated = int(payload.get("updated", 0) or 0)
                        messagebox.showinfo(
                            "Buscar OME / DNI",
                            (
                                "Busqueda finalizada.\n\n"
                                f"Filas revisadas: {total_rows}\n"
                                f"Busquedas realizadas: {total_searches}\n"
                                f"Resultados primarios encontrados: {found}\n"
                                f"DNIs completados: {dni_found}\n"
                                f"No encontradas: {not_found}\n"
                                f"Filas escritas en Sheets: {updated}"
                            ),
                        )
                    elif event == "action_finished":
                        self.action_running = False
                        self._set_controls_enabled(True)
                    elif event == "action_error":
                        self.action_running = False
                        self._set_controls_enabled(True)
                        self.status_label.configure(text="Ocurrio un error.")
                        self._append_log(f"ERROR: {payload}")
                        messagebox.showerror("Error", str(payload))
                except Exception as exc:
                    self.action_running = False
                    self._set_controls_enabled(True)
                    log_message(f"Error procesando evento de activacion '{event}': {exc}")
                    try:
                        self._append_log(f"ERROR UI ({event}): {exc}")
                    except Exception:
                        pass
                    try:
                        self.status_label.configure(text="Ocurrio un error de interfaz.")
                    except Exception:
                        pass
        except queue.Empty:
            pass

        delay = 120 if (self.action_running or had_events) else 350
        self.after(delay, self._process_ui_queue)

    def _append_log(self, text: str) -> None:
        self.log_text.configure(state="normal")
        self.log_text.insert("end", text + "\n")
        self.log_text.see("end")
        self.log_text.configure(state="disabled")

    def _ensure_identifier_rows(self, total: int) -> None:
        actual = len(self.identificador_rows)
        for _ in range(actual, total):
            row_index = len(self.identificador_rows)
            bene_entry = tk.Entry(
                self.identifier_grid,
                relief="solid",
                bd=1,
                font=("Consolas", 11),
                highlightthickness=1,
                highlightbackground="#d9e2ec",
                highlightcolor="#245b9d",
            )
            bene_entry.grid(row=row_index, column=0, padx=(0, 6), pady=2, sticky="ew")
            bene_entry.bind("<FocusIn>", lambda _event, r=row_index: self._set_active_identifier_cell(r, "benef"))
            bene_entry.bind("<Control-v>", self._handle_identifier_ctrl_v)
            bene_entry.bind("<Control-V>", self._handle_identifier_ctrl_v)
            bene_entry.bind("<MouseWheel>", self._on_identifier_mousewheel)
            bene_entry.bind("<Button-4>", self._on_identifier_mousewheel)
            bene_entry.bind("<Button-5>", self._on_identifier_mousewheel)
            bene_entry.bind("<KeyRelease>", lambda _event, r=row_index: self._maybe_expand_identifier_rows(r))

            orden_entry = tk.Entry(
                self.identifier_grid,
                relief="solid",
                bd=1,
                font=("Consolas", 11),
                highlightthickness=1,
                highlightbackground="#d9e2ec",
                highlightcolor="#245b9d",
            )
            orden_entry.grid(row=row_index, column=1, padx=(6, 0), pady=2, sticky="ew")
            orden_entry.bind("<FocusIn>", lambda _event, r=row_index: self._set_active_identifier_cell(r, "orden"))
            orden_entry.bind("<Control-v>", self._handle_identifier_ctrl_v)
            orden_entry.bind("<Control-V>", self._handle_identifier_ctrl_v)
            orden_entry.bind("<MouseWheel>", self._on_identifier_mousewheel)
            orden_entry.bind("<Button-4>", self._on_identifier_mousewheel)
            orden_entry.bind("<Button-5>", self._on_identifier_mousewheel)
            orden_entry.bind("<KeyRelease>", lambda _event, r=row_index: self._maybe_expand_identifier_rows(r))

            self.identificador_rows.append((bene_entry, orden_entry))

    def _selected_identifier_index(self) -> int:
        return self.active_identifier_row

    def _set_active_identifier_cell(self, row_index: int, column_name: str) -> None:
        self.active_identifier_row = row_index
        self.active_identifier_column = column_name
        self._maybe_expand_identifier_rows(row_index)

    def _maybe_expand_identifier_rows(self, row_index: int) -> None:
        total = len(self.identificador_rows)
        if total - row_index <= 2:
            self._ensure_identifier_rows(total + 8)

    def _normalize_identifier_cell(self, value: str) -> str:
        text = (value or "").strip()
        if not text:
            return ""
        if text.upper() in {"GENERADA", "VACIA", "VACIO", "-"}:
            return ""
        return text

    def _parse_identifier_line(self, line: str) -> tuple[str, str]:
        text = line.strip()
        if not text:
            return "", ""
        if "\t" in text:
            partes = [p.strip() for p in text.split("\t", 1)]
            return self._normalize_identifier_cell(partes[0]), self._normalize_identifier_cell(partes[1] if len(partes) > 1 else "")
        if ";" in text:
            partes = [p.strip() for p in text.split(";", 1)]
            return self._normalize_identifier_cell(partes[0]), self._normalize_identifier_cell(partes[1] if len(partes) > 1 else "")
        if "," in text:
            partes = [p.strip() for p in text.split(",", 1)]
            return self._normalize_identifier_cell(partes[0]), self._normalize_identifier_cell(partes[1] if len(partes) > 1 else "")
        return self._normalize_identifier_cell(text), ""

    def _paste_identifier_clipboard(self, mode: str) -> None:
        try:
            raw = self.clipboard_get()
        except tk.TclError:
            messagebox.showwarning("Portapapeles vacio", "No hay texto para pegar.")
            return

        lineas = [line.strip() for line in raw.splitlines() if line.strip()]
        if not lineas:
            messagebox.showwarning("Portapapeles vacio", "No hay lineas validas para pegar.")
            return

        inicio = self._selected_identifier_index()
        self._ensure_identifier_rows(inicio + len(lineas) + 20)

        for offset, line in enumerate(lineas):
            bene_entry, orden_entry = self.identificador_rows[inicio + offset]
            benef = self._normalize_identifier_cell(bene_entry.get())
            orden = self._normalize_identifier_cell(orden_entry.get())

            if mode == "tabla":
                benef, orden = self._parse_identifier_line(line)
            elif mode == "benef":
                benef = self._normalize_identifier_cell(line)
            elif mode == "orden":
                orden = self._normalize_identifier_cell(line)

            bene_entry.delete(0, "end")
            bene_entry.insert(0, benef)
            orden_entry.delete(0, "end")
            orden_entry.insert(0, orden)

        target_entry = self.identificador_rows[inicio][0 if self.active_identifier_column == "benef" else 1]
        target_entry.focus_set()

    def _handle_identifier_ctrl_v(self, _event) -> str:
        modo = "benef" if self.active_identifier_column == "benef" else "orden"
        self._paste_identifier_clipboard(modo)
        return "break"

    def _clear_identifier_grid(self) -> None:
        for bene_entry, orden_entry in self.identificador_rows:
            bene_entry.delete(0, "end")
            orden_entry.delete(0, "end")
        self.active_identifier_row = 0
        self.active_identifier_column = "benef"
        if self.identificador_rows:
            self.identificador_rows[0][0].focus_set()

    def _on_identifier_mousewheel(self, event) -> str:
        canvas = getattr(self.identifier_grid, "_parent_canvas", None)
        if canvas is None:
            return "break"

        if getattr(event, "num", None) == 4:
            canvas.yview_scroll(-1, "units")
        elif getattr(event, "num", None) == 5:
            canvas.yview_scroll(1, "units")
        else:
            delta = int(getattr(event, "delta", 0))
            if delta != 0:
                canvas.yview_scroll(int(-delta / 120), "units")
        return "break"

    def _set_report_text(self, text: str) -> None:
        self.report_text.configure(state="normal")
        self.report_text.delete("1.0", "end")
        self.report_text.insert("1.0", text)
        self.report_text.configure(state="disabled")

    def _extract_codigo_practica(self, detalle) -> str:
        def extract_from_text(text: str) -> str:
            match = re.search(r"\b(\d{6})\b", text or "")
            return match.group(1) if match else ""

        codigo = (getattr(detalle, "codigo_practica", "") or "").strip()
        if codigo.isdigit():
            return codigo

        practica_encontrada = (getattr(detalle, "practica_encontrada", "") or "").strip()
        if practica_encontrada:
            candidato = extract_from_text(practica_encontrada)
            if candidato:
                return candidato

        practica_solicitada = (getattr(detalle, "practica", "") or "").strip()
        if practica_solicitada:
            candidato = extract_from_text(practica_solicitada)
            if candidato:
                return candidato

        return "-"

    def _display_beneficio(self, detalle) -> str:
        beneficio = (getattr(detalle, "beneficio_encontrado", "") or "").strip()
        if beneficio:
            return beneficio
        return (getattr(detalle, "n_afiliado", "") or "").strip() or "-"

    def _format_report_rows(self, detalles) -> str:
        headers = [
            ("BENEF / DNI", 16),
            ("NRO OME", 14),
            ("NOMBRE APELLIDO", 30),
            ("CODIGO PRACTICA", 18),
            ("HORARIO ACTIVACION", 20),
        ]

        def fit(value: str, width: int) -> str:
            text = (value or "-").strip()
            if len(text) > width:
                return text[: width - 1] + "…"
            return text.ljust(width)

        lineas = [" | ".join(fit(label, width) for label, width in headers)]
        lineas.append("-" * len(lineas[0]))

        if not detalles:
            lineas.append("Sin resultados todavia.")
            return "\n".join(lineas)

        for detalle in detalles:
            horario = f"{detalle.fecha} {detalle.hora}:{detalle.minuto}"
            fila = [
                fit(self._display_beneficio(detalle), 16),
                fit(getattr(detalle, "n_orden_encontrada", "") or getattr(detalle, "n_orden_solicitada", "") or "-", 14),
                fit(getattr(detalle, "nombre_afiliado", "") or "-", 30),
                fit(self._extract_codigo_practica(detalle), 18),
                fit(horario, 20),
            ]
            lineas.append(" | ".join(fila))
        return "\n".join(lineas)

    def _render_estado(self, estado: dict) -> None:
        url = str(estado.get("url", "") or "")
        titulo = str(estado.get("titulo", "") or "")
        en_panel = bool(estado.get("en_panel"))
        timestamp = str(estado.get("timestamp", "") or "")

        navegador = "Abierto" if url else "Cerrado"
        panel = "Si" if en_panel else "No"
        if url:
            self.summary_label.configure(
                text=f"Navegador: {navegador} | Panel: {panel} | Pantalla: {titulo or url} | {timestamp or 'Sin marca de tiempo'}"
            )
        else:
            self.summary_label.configure(text=f"Log en: {get_log_file()}")

    def _set_controls_enabled(self, enabled: bool) -> None:
        state = "normal" if enabled else "disabled"
        for button in (
            self.restart_app_button,
            self.close_button,
            self.run_button,
            self.stop_button,
            self.reset_button,
            self.paste_benef_button,
            self.paste_ome_button,
            self.clear_sheet_button,
            self.export_report_button,
            self.practice_catalog_button,
            self.new_profile_button,
            self.delete_profile_button,
            self.save_profile_button,
            self.toggle_password_button,
            self.fecha_button,
            self.sheets_connect_button,
            self.sheets_run_button,
            self.sheet_lookup_button,
            self.sheet_tabs_button,
            self.new_sheet_profile_button,
            self.save_sheet_profile_button,
            self.delete_sheet_profile_button,
        ):
            button.configure(state=state)

        if enabled and self.last_report_export_path and Path(self.last_report_export_path).exists():
            self.open_report_button.configure(state="normal")
        else:
            self.open_report_button.configure(state="disabled")

        self.profile_combo.configure(state=state)
        self.client_entry.configure(state=state)
        self.user_entry.configure(state=state)
        self.password_entry.configure(state=state)
        self.browser_visible_checkbox.configure(state=state)
        self.sheet_profile_combo.configure(state="readonly" if enabled else "disabled")
        self.sheet_url_entry.configure(state=("normal" if self.sheet_url_editable else "readonly") if enabled else "disabled")
        self.sheet_url_edit_button.configure(state=state)
        self.sheet_name_entry.configure(state=state)
        self.sheet_start_row_entry.configure(state=state)
        self.sheet_max_rows_entry.configure(state=state)
        self.fecha_entry.configure(state="readonly")
        self.hora_combo.configure(state=state)
        self.minuto_combo.configure(state=state)
        self.modalidad_combo.configure(state=state)
        self.intervalo_combo.configure(state=state)
        self.practica_combo.configure(state=state)
        self.boca_combo.configure(state=state)
        self.identificador_edit_enabled = enabled
        for bene_entry, orden_entry in self.identificador_rows:
            bene_entry.configure(state=state)
            orden_entry.configure(state=state)
        self.stop_button.configure(state="normal" if self.action_running else "disabled")

    # ------------------------------------------------------------------
    # Perfil
    # ------------------------------------------------------------------

    def _profile_options(self) -> list[str]:
        return [p.get("nombre", "") for p in self.saved_profiles.get("usuarios", [])]

    def _load_saved_profiles(self) -> dict:
        try:
            if self.profiles_file.exists():
                data = json.loads(self.profiles_file.read_text(encoding="utf-8"))
                if isinstance(data, dict):
                    data.setdefault("usuarios", [])
                    data.setdefault("selected_profile_name", "")
                    return sync_profile_payload(data)
        except Exception:
            pass
        return {"usuarios": [], "selected_profile_name": ""}

    def _save_profiles_to_disk(self) -> None:
        try:
            self.profiles_file.write_text(
                json.dumps(self.saved_profiles, indent=2, ensure_ascii=False),
                encoding="utf-8",
            )
            upsert_shared_credentials_from_payload(self.saved_profiles)
        except Exception as exc:
            log_message(f"Error guardando perfiles de activacion: {exc}")

    def _current_profile_from_form(self) -> dict:
        return {
            "nombre": self.profile_name_var.get().strip(),
            "usuario": self.profile_user_var.get().strip(),
            "clave": self.profile_password_var.get(),
        }

    def _upsert_profile(self, profile: dict) -> None:
        if not profile.get("nombre"):
            return
        usuarios = self.saved_profiles.setdefault("usuarios", [])
        for saved in usuarios:
            if saved.get("nombre") == profile["nombre"]:
                saved.update(profile)
                self.saved_profiles["selected_profile_name"] = profile["nombre"]
                self._save_profiles_to_disk()
                self._refresh_profile_combo(selected=profile["nombre"])
                return
        usuarios.append(profile)
        self.saved_profiles["selected_profile_name"] = profile["nombre"]
        self._save_profiles_to_disk()
        self._refresh_profile_combo(selected=profile["nombre"])

    def _save_current_profile(self) -> None:
        profile = self._current_profile_from_form()
        if not profile["nombre"]:
            messagebox.showwarning("Perfil", "Ingresa un nombre de medico o centro antes de guardar.")
            return
        self._upsert_profile(profile)
        self._save_sheet_settings()
        messagebox.showinfo("Perfil", f"Perfil '{profile['nombre']}' guardado.")

    def _new_profile(self) -> None:
        self.profile_var.set("")
        self.profile_name_var.set("")
        self.profile_user_var.set("")
        self.profile_password_var.set("")
        self.saved_profiles["selected_profile_name"] = ""
        self._save_profiles_to_disk()
        if not self._syncing_sheet_profile:
            self._save_sheet_settings()

    def _delete_current_profile(self) -> None:
        nombre = self.profile_var.get().strip()
        if not nombre:
            return
        usuarios = self.saved_profiles.get("usuarios", [])
        self.saved_profiles["usuarios"] = [p for p in usuarios if p.get("nombre") != nombre]
        if self.saved_profiles.get("selected_profile_name") == nombre:
            self.saved_profiles["selected_profile_name"] = ""
        self._save_profiles_to_disk()
        self._refresh_profile_combo()
        self._new_profile()

    def _on_profile_selected(self, choice: str) -> None:
        self._syncing_pami_profile = True
        for profile in self.saved_profiles.get("usuarios", []):
            if profile.get("nombre") == choice:
                self.profile_var.set(choice)
                self.profile_name_var.set(profile.get("nombre", ""))
                self.profile_user_var.set(profile.get("usuario", ""))
                self.profile_password_var.set(profile.get("clave", ""))
                self.saved_profiles["selected_profile_name"] = choice
                break
        self._syncing_pami_profile = False
        if not self._syncing_sheet_profile:
            self._select_sheet_profile_for_pami(self.profile_user_var.get().strip())
        self._save_profiles_to_disk()
        self._save_sheet_settings()

    def _load_initial_profile_into_form(self) -> None:
        usuarios = self.saved_profiles.get("usuarios", [])
        if usuarios:
            selected = str(self.saved_profiles.get("selected_profile_name", "")).strip()
            if selected and any(profile.get("nombre") == selected for profile in usuarios):
                self._on_profile_selected(selected)
            else:
                self._on_profile_selected(usuarios[0].get("nombre", ""))
        else:
            self._new_profile()

    def _refresh_profile_combo(self, selected: str | None = None) -> None:
        options = self._profile_options()
        self.profile_combo.configure(values=options or [""])
        target = (selected or str(self.saved_profiles.get("selected_profile_name", ""))).strip()
        if target and target in options:
            self.profile_var.set(target)
        else:
            self.profile_var.set(options[0] if options else "")

    def _select_profile_by_user(self, usuario: str) -> bool:
        usuario = (usuario or "").strip().lower()
        if not usuario:
            return False
        for profile in self.saved_profiles.get("usuarios", []):
            if str(profile.get("usuario", "")).strip().lower() == usuario:
                nombre = str(profile.get("nombre", "")).strip()
                if nombre:
                    self.profile_var.set(nombre)
                    self._on_profile_selected(nombre)
                    return True
        return False

    def _selected_sheet_profile(self) -> dict | None:
        selected = (self.sheet_profile_var.get() or "").strip()
        profile = self.sheet_profile_lookup.get(selected)
        if profile:
            return profile
        current_key = str(self.selected_sheet_profile_id or "").strip()
        if not current_key:
            return None
        for item in self.sheet_profiles:
            if self._sheet_profile_key(item) == current_key:
                return item
        return None

    def _select_pami_profile_for_sheet(self, sheet_profile: dict | None) -> None:
        if not sheet_profile:
            return
        usuario = str(sheet_profile.get("pami_usuario", "")).strip()
        if usuario:
            self._select_profile_by_user(usuario)

    def _select_sheet_profile_for_pami(self, usuario: str) -> bool:
        usuario = (usuario or "").strip().lower()
        if not usuario:
            return False
        current = self._selected_sheet_profile()
        if current and str(current.get("pami_usuario", "")).strip().lower() == usuario:
            return False
        for profile in self.sheet_profiles:
            if str(profile.get("pami_usuario", "")).strip().lower() == usuario:
                self._apply_sheet_profile(profile, select_pami=False, save=True)
                return True
        return False

    # ------------------------------------------------------------------
    # Helpers de UI
    # ------------------------------------------------------------------

    def _pick_date(self) -> None:
        dialog = DatePickerDialog(self, self.fecha_var.get().strip())
        self.wait_window(dialog)
        if dialog.result is not None:
            self.fecha_var.set(dialog.result)

    def _toggle_password(self) -> None:
        self.password_visible = not self.password_visible
        self.password_entry.configure(show="" if self.password_visible else "*")
        _eye = button_icon("eye_off.png" if self.password_visible else "eye.png", (18, 18))
        if _eye is not None:
            self.toggle_password_button.configure(image=_eye, text="")

    def _on_browser_visibility_changed(self) -> None:
        self._save_sheet_settings()

    def _go_home(self) -> None:
        if self.action_running:
            messagebox.showwarning("Atencion", "Espera a que termine la accion actual antes de volver.")
            return
        if self.on_back:
            self.on_back()

    def on_close(self) -> None:
        try:
            if self.controller_thread is not None and self.controller_thread.is_alive():
                self.controller_queue.put(self.controller.cerrar_navegador)
                self.controller_queue.put(None)
        except Exception:
            pass
        finally:
            self.destroy()
